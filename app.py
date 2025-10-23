import openpyxl
from pathlib import Path
import os
import shutil
import subprocess
import sys
import threading
from datetime import datetime, timedelta
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Dict, Tuple, List
import time
import re
import glob
import zipfile

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from threading import Lock
from collections import deque
import openpyxl

app = Flask(__name__)

# Suppress Waitress task queue warnings
import logging
logging.getLogger('waitress.queue').setLevel(logging.ERROR)

def kill_excel_instances():
    """Kill any running Excel instances to prevent COM conflicts."""
    try:
        print("Checking for running Excel instances...")
        result = subprocess.run(
            ['tasklist', '/FI', 'IMAGENAME eq EXCEL.EXE'],
            capture_output=True,
            text=True,
            timeout=5
        )

        if 'EXCEL.EXE' in result.stdout:
            print("Found running Excel instances. Killing them...")
            subprocess.run(
                ['taskkill', '/F', '/IM', 'EXCEL.EXE'],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                timeout=5
            )
            import time
            time.sleep(2)
            print("Excel instances killed successfully")
        else:
            print("No Excel instances found")
    except Exception as e:
        print(f"Error killing Excel instances: {e}")

def _append_run_history(
    frequency: str,
    picked_date: str,
    scripts: list[str],
    checklist: list[dict],
    output_dir: Path | None
) -> None:
    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    
    if HISTORY_XLSX.exists():
        wb = openpyxl.load_workbook(HISTORY_XLSX)
    else:
        wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "History"

    # Add header if file is new
    if ws.max_row == 1 and ws.cell(1, 1).value is None:
        ws.append(["timestamp", "frequency", "picked_date", "reports", "statuses", "output_folder"])

    from datetime import datetime as _dt
    ts = _dt.now().strftime("%Y-%m-%d %H:%M:%S")

    reports = ", ".join(Path(s).stem for s in scripts)
    statuses = ", ".join(f"{it['name']}={it['status']}" for it in checklist)
    out_name = output_dir.name if output_dir else ""

    ws.append([ts, frequency, picked_date, reports, statuses, out_name])

    # Keep last 8 records + header
    while ws.max_row > 9:
        ws.delete_rows(2)

    wb.save(HISTORY_XLSX)
    wb.close()

# ==================== GLOBAL CONFIGURATION ====================
BASE_DIR = Path(__file__).resolve().parent
OUTPUTS_DIR = BASE_DIR / "outputs"
WORKING_DIR = BASE_DIR / "working"
REPORT_AUTOMATIONS_DIR = BASE_DIR / "report_automations"
LOGS_DIR = BASE_DIR / "logs"
MASTER_INPUTS_DIR = BASE_DIR / "Master Inputs"
OLD_OUTPUTS_DIR = BASE_DIR / "old_outputs"
HISTORY_XLSX = LOGS_DIR / "run_history.xlsx"
MASTER_DATA_XLSX = BASE_DIR / "Master_Data.xlsx"

# Status tracking
_status_lock = Lock()
_status: dict[str, str | bool] = {
    "running": False,
    "stage": "idle",
    "message": "",
}

_log_lock = Lock()
_log_buffer: deque[str] = deque(maxlen=2000)
_proc_lock = Lock()
_stop_requested: bool = False
_running_processes: dict[str, subprocess.Popen] = {}

# Create timestamped log file for each run
_timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S')
_run_log_dir = LOGS_DIR / f"run_{_timestamp_str}"
_run_log_dir.mkdir(parents=True, exist_ok=True)
RUN_LOG_FILE = _run_log_dir / "web_run.log"
_status_feed: deque[str] = deque(maxlen=200)
_report_checklist: list[dict] = []
_report_messages: dict[str, deque[str]] = {}
_last_output_dirs: dict[str, Path] = {}  # Track output dirs per report
_download_files: list[Path] = []  # Track files to download
_sofp_download_file: Path | None = None  # Track SOFP file for download after save

# ==================== REPORT DEPENDENCY STRUCTURE ====================
REPORT_STRUCTURE = {
    # LEVEL 1 Reports
    "NBD-MF-01-SOFP-SOCI": {
        "level": 1,
        "set": 1,
        "step": 1,
        "name": "Statement of Financial position",
        "script": "NBD_MF_01_SOFP_SOCI.py",
        "folder": "NBD_MF_01_SOFP_SOCI",
        "dependencies": [],
        "estimated_minutes": 5
    },
    "NBD-WF-18-DM": {
        "level": 1,
        "set": 2,
        "step": 1,
        "name": "Deposits Movement in Previous Week",
        "script": "NBD_WF_18_DM.py",
        "folder": "NBD-WF-18-DM",
        "dependencies": [],
        "estimated_minutes": 30
    },
    "NBD-WF-15-LA": {
        "level": 1,
        "set": 2,
        "step": 2,
        "name": "Liquid Assets",
        "script": "NBD_MF_15_LA.py",
        "folder": "NBD_MF_15_LA",
        "dependencies": [],
        "estimated_minutes": 3
    },
    "NBD-WF-18-ECS": {
        "level": 1,
        "set": 2,
        "step": 3,
        "name": "Expected Cashflow Summary",
        "script": "NBD_WF_18_ECS.py",
        "folder": "NBD_WF_18_ECS",
        "dependencies": [],
        "estimated_minutes": 4
    },

    # LEVEL 2 Reports
    "NBD-MF-04-LA": {
        "level": 2,
        "set": 3,
        "step": 1,
        "name": "Liquid Assets",
        "script": "NBD_MF_04_LA.py",
        "folder": "NBD_MF_04_LA",
        "dependencies": [],
        "estimated_minutes": 8
    },
    "NBD-MF-06-ID": {
        "level": 2,
        "set": 4,
        "step": 1,
        "name": "Interest Deposit",
        "script": "NBD_MF_06_ID.py",
        "folder": "NBD_MF_06_ID",
        "dependencies": [],
        "estimated_minutes": 10
    },
    "NBD-MF-10-GA-11-IS-SET8": {
        "level": 2,
        "set": 5,
        "step": 1,
        "name": "GA & IS Pipeline",
        "script": "NBD_MF_10_GA_NBD_MF_11_IS.py",
        "folder": "NBD-MF-10-GA & NBD-MF-11-IS",
        "dependencies": [],
        "estimated_minutes": 25
    },
    "NBD-MF-23-IA": {
        "level": 2,
        "set": 6,
        "step": 1,
        "name": "Product wise Classification of Loans",
        "script": "NBD_MF_23_IA.py",
        "folder": "NBD_MF_23_IA",
        "dependencies": [],
        "estimated_minutes": 15
    },
    "NBD-MF-23-C1-C2": {
        "level": 2,
        "set": 6,
        "step": 2,
        "name": "Product wise Classification of Loans (Amounts)",
        "script": "NBD_MF_23_C1C2.py",
        "folder": "NBD_MF_23_C1C2",
        "dependencies": [],
        "estimated_minutes": 10
    },
    "NBD-MF-20-C1": {
        "level": 2,
        "set": 7,
        "step": 1,
        "name": "Form 1 : Computation of Capital Ratios",
        "script": "NBD_MF_20_C1.py",
        "folder": "NBD_MF_20_C1_C6",
        "dependencies": [],
        "estimated_minutes": 10
    },
    "NBD-MF-20-C2": {
        "level": 2,
        "set": 7,
        "step": 2,
        "name": "Form 2 : Computation of Total Capital",
        "script": "NBD_MF_20_C2.py",
        "folder": "NBD_MF_20_C1_C6",
        "dependencies": [],
        "estimated_minutes": 12
    },
    "NBD-MF-20-C3": {
        "level": 2,
        "set": 7,
        "step": 3,
        "name": "Form 3: Computation of Risk Weighted Amount for Credit Risk",
        "script": "NBD_MF_20_C3.py",
        "folder": "NBD_MF_20_C1_C6",
        "dependencies": [],
        "estimated_minutes": 15
    },
    "NBD-MF-20-C4": {
        "level": 2,
        "set": 7,
        "step": 4,
        "name": "Form 4: Credit equivalent of Off-Balance Sheet Item",
        "script": "NBD_MF_20_C4.py",
        "folder": "NBD_MF_20_C1_C6",
        "dependencies": [],
        "estimated_minutes": 10
    },
    "NBD-MF-20-C5": {
        "level": 2,
        "set": 7,
        "step": 5,
        "name": "Form 5: Exposures recognized under Credit Risk Mitigation CRM",
        "script": "NBD_MF_20_C5.py",
        "folder": "NBD_MF_20_C1_C6",
        "dependencies": [],
        "estimated_minutes": 12
    },
    "NBD-MF-20-C6": {
        "level": 2,
        "set": 7,
        "step": 6,
        "name": "Form 6: Computation of Risk Weighted Amount for Operational Risk",
        "script": "NBD_MF_20_C6.py",
        "folder": "NBD_MF_20_C1_C6",
        "dependencies": [],
        "estimated_minutes": 15
    },

    # LEVEL 2 Reports - Set 5 (GA & IS combined pipeline)
    "NBD-MF-10-GA-11-IS-SET8": {
        "level": 2,
        "set": 5,
        "step": 1,
        "name": "GA & IS Pipeline",
        "script": "NBD_MF_10_GA_NBD_MF_11_IS.py",
        "dependencies": [],
        "folder": "NBD-MF-10-GA & NBD-MF-11-IS",
        "estimated_minutes": 25
    },

    # LEVEL 2 Reports - Set 8 (C4 standalone)
    "NBD-MF-23-C4": {
        "level": 2,
        "set": 8,
        "step": 1,
        "name": "Loan CF Analysis",
        "script": "NBD-MF-23-C4.py",
        "dependencies": [],
        "folder": "NBD_MF_23_C4",
        "estimated_minutes": 30
    },

    # LEVEL 3 Reports
    "NBD-QF-23-C9": {
        "level": 3,
        "set": 8,
        "step": 1,
        "name": "Sector Wise Credit Exposures",
        "script": "NBD_QF_23_C9.py",
        "folder": "NBD_QF_23_C9",
        "dependencies": [],
        "estimated_minutes": 8
    },
    "NBD-QF-23-C10": {
        "level": 3,
        "set": 8,
        "step": 2,
        "name": "District Wise Credit Exposures",
        "script": "NBD_QF_23_C10.py",
        "folder": "NBD_QF_23_C10",
        "dependencies": [],
        "estimated_minutes": 8
    },
    "NBD-QF-23-C6": {
        "level": 3,
        "set": 8,
        "step": 3,
        "name": "Large Exposures (Top 50 Borrowers)",
        "script": "NBD_QF_23_C6.py",
        "folder": "NBD_QF_23_C6",
        "dependencies": [],
        "estimated_minutes": 6
    },
    "NBD-QF-23-C7": {
        "level": 3,
        "set": 9,
        "step": 1,
        "name": "Detailed breakdown of the stage-wise movement of loans",
        "script": "NBD_QF_23_C7.py",
        "folder": "NBD_QF_23_C7",
        "dependencies": [],
        "estimated_minutes": 10
    },
    "NBD-QF-23-C8": {
        "level": 3,
        "set": 10,
        "step": 1,
        "name": "Classification of Loans based on Internal Credit Risk Rating",
        "script": "NBD_QF_23_C8.py",
        "folder": "NBD_QF_23_C8",
        "dependencies": [],
        "estimated_minutes": 12
    },
    "NBD-QF-23-C4": {
        "level": 3,
        "set": 11,
        "step": 1,
        "name": "Product wise Probability of Default and Loss Given Default of Loans",
        "script": "NBD_QF_23_C4.py",
        "folder": "NBD_MF_23_C4",
        "dependencies": [],
        "estimated_minutes": 15
    },
    "NBD-QF-23-C3": {
        "level": 3,
        "set": 12,
        "step": 1,
        "name": "Measurement of Credit Risk in Loans Based on Regulatory Classification",
        "script": "NBD_QF_23_C3.py",
        "folder": "NBD_QF_23_C3",
        "dependencies": [],
        "estimated_minutes": 12
    },
    "NBD-QF-16-SF": {
        "level": 3,
        "set": 13,
        "step": 1,
        "name": "Sustainable Finance Activities (SFAs)",
        "script": "NBD_QF_16_SF.py",
        "folder": "NBD_QF_16_SF",
        "dependencies": [],
        "estimated_minutes": 10
    },
    "NBD-QF-16-FI": {
        "level": 3,
        "set": 14,
        "step": 1,
        "name": "National Financial Inclusion Council Survey",
        "script": "NBD_QF_16_FI.py",
        "folder": "NBD_QF_16_FI",
        "dependencies": [],
        "estimated_minutes": 8
    },
}

# ==================== HELPER FUNCTIONS ====================

def _emit_log(message: str, *, important: bool = False) -> None:
    """Emit log message."""
    line = f"[{datetime.now().strftime('%H:%M:%S')}] {message}"
    with _log_lock:
        _log_buffer.append(line)
    try:
        with RUN_LOG_FILE.open("a", encoding="utf-8") as f:
            f.write(line + "\n")
    except Exception:
        pass


def _set_status(**kwargs) -> None:
    with _status_lock:
        _status.update(kwargs)


def _reset_status() -> None:
    _set_status(running=False, stage="idle", message="")


def _initialize_new_run() -> None:
    """Initialize a new run with fresh timestamp and log directory."""
    global _timestamp_str, _run_log_dir, RUN_LOG_FILE
    _timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    _run_log_dir = LOGS_DIR / f"run_{_timestamp_str}"
    _run_log_dir.mkdir(parents=True, exist_ok=True)
    RUN_LOG_FILE = _run_log_dir / "web_run.log"
    _emit_log(f"Starting new run: {_timestamp_str}", important=True)


def _init_report_checklist(report_ids: list[str]) -> None:
    """Initialize report checklist for the UI."""
    global _report_checklist, _report_messages
    _report_checklist = []
    _report_messages = {}
    for report_id in report_ids:
        if report_id in REPORT_STRUCTURE:
            name = REPORT_STRUCTURE[report_id]["name"]
            est_min = REPORT_STRUCTURE[report_id].get("estimated_minutes", 10)
            _report_checklist.append({
                "name": report_id,
                "display_name": name,
                "status": "pending",
                "estimated_minutes": est_min,
                "start_time": None,
                "elapsed_seconds": 0
            })
            _report_messages[report_id] = deque(maxlen=200)


def _set_report_status(report_id: str, status: str) -> None:
    """Set status for a report by ID."""
    for item in _report_checklist:
        if item["name"] == report_id:
            item["status"] = status
            if status == "running":
                item["start_time"] = time.time()
            elif status in ["completed", "failed", "stopped"]:
                if item["start_time"]:
                    item["elapsed_seconds"] = int(time.time() - item["start_time"])
            break


def _update_report_elapsed_time(report_id: str) -> None:
    """Update elapsed time for running report."""
    for item in _report_checklist:
        if item["name"] == report_id and item["status"] == "running" and item["start_time"]:
            item["elapsed_seconds"] = int(time.time() - item["start_time"])
            break


def parse_date_folder_name(folder_name: str) -> datetime:
    """Parse date from folder name format: 30/09/2025 or 31/10/2025(1)"""
    base_name = re.sub(r'\(\d+\)$', '', folder_name)
    try:
        return datetime.strptime(base_name, "%d/%m/%Y")
    except ValueError:
        try:
            return datetime.strptime(base_name, "%d-%m-%Y")
        except ValueError:
            try:
                return datetime.strptime(base_name, "%m/%d/%Y")
            except ValueError:
                raise ValueError(f"Cannot parse date from folder name: {folder_name}")


def find_latest_date_folder(report_folder: Path) -> Path | None:
    """Find the latest versioned date folder in a report folder."""
    if not report_folder.exists():
        return None
    
    candidates: list[tuple[datetime, int, Path]] = []
    
    for child in report_folder.iterdir():
        if not child.is_dir():
            continue
        
        folder_name = child.name
        try:
            date_obj = parse_date_folder_name(folder_name)
            version_match = re.search(r'\((\d+)\)$', folder_name)
            version = int(version_match.group(1)) if version_match else 0
            candidates.append((date_obj, version, child))
        except ValueError:
            continue
    
    if not candidates:
        return None
    
    candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
    return candidates[0][2]


def find_latest_completed_date(source_folder: Path) -> tuple[datetime | None, Path | None]:
    """
    Find the latest completed date from completed_dates.txt in source folder.
    Returns tuple of (datetime, date_folder_path)
    """
    completed_dates_file = source_folder / "completed_dates.txt"
    
    if not completed_dates_file.exists():
        _emit_log(f"No completed_dates.txt found in {source_folder}", important=True)
        return None, None
    
    try:
        with open(completed_dates_file, 'r') as f:
            lines = [line.strip() for line in f.readlines() if line.strip()]
        
        if not lines:
            _emit_log(f"completed_dates.txt is empty in {source_folder}", important=True)
            return None, None
        
        # Get the last date from file
        last_date_str = lines[-1]
        _emit_log(f"Found last completed date: {last_date_str}", important=True)
        
        # Parse the date (supports DD-MM-YYYY and DD/MM/YYYY)
        try:
            last_date = datetime.strptime(last_date_str, "%d-%m-%Y")
        except ValueError:
            last_date = datetime.strptime(last_date_str, "%d/%m/%Y")
        
        # Find the folder with this date (including variations)
        date_folder = find_date_folder_with_variations(source_folder, last_date_str)
        
        if date_folder:
            _emit_log(f"Found date folder: {date_folder}", important=True)
            return last_date, date_folder
        else:
            _emit_log(f"Date folder not found for: {last_date_str}", important=True)
            return last_date, None
            
    except Exception as e:
        _emit_log(f"Error reading completed_dates.txt: {e}", important=True)
        return None, None


def find_date_folder_with_variations(parent_folder: Path, date_str: str) -> Path | None:
    """
    Find the latest variation of a date folder (e.g., 30/09/2025, 30/09/2025(1), etc.)
    """
    if not parent_folder.exists():
        return None
    
    candidates: list[tuple[int, Path]] = []
    
    date_variants = {date_str}
    # Accept both DD/MM/YYYY and DD-MM-YYYY
    try:
        _dt = datetime.strptime(date_str, "%d/%m/%Y")
        date_variants.add(_dt.strftime("%d-%m-%Y"))
    except Exception:
        try:
            _dt = datetime.strptime(date_str, "%d-%m-%Y")
            date_variants.add(_dt.strftime("%d/%m/%Y"))
        except Exception:
            pass
    
    for child in parent_folder.iterdir():
        if not child.is_dir():
            continue
        
        folder_name = child.name
        
        # Check if folder name starts with the date
        if any(folder_name.startswith(v) for v in date_variants):
            # Extract version number
            version_match = re.search(r'\((\d+)\)$', folder_name)
            version = int(version_match.group(1)) if version_match else 0
            candidates.append((version, child))
    
    if not candidates:
        return None
    
    # Sort by version descending and return the latest
    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1]


def format_date_folder_name(date_str: str) -> str:
    """Convert date from HTML format (YYYY-MM-DD) to folder format (DD-MM-YYYY)."""
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    return dt.strftime("%d-%m-%Y")


def copy_ia_to_additional_outputs(ia_file_path: Path, gui_date_str: str) -> None:
    """
    Copy the IA 'Prod. wise Class. of Loans' file to three additional output folders:
    1. outputs/NBD_MF_20_C1_C6/{latest-date-folder}/
    2. outputs/NBD_QF_23_C8/{latest-date-folder}/
    3. outputs/NBD-QF-23-C3-C10-SF-FI/{latest-date-folder}/Input/

    Before copying, deletes old 'Prod. wise Class. of Loans*.xlsb' files.
    Uses completed_dates.txt to find the latest date folder.
    """
    import shutil

    # Define the three target output folders
    target_configs = [
        {
            "name": "C1-C6",
            "base_folder": OUTPUTS_DIR / "NBD_MF_20_C1_C6",
            "subfolder": None  # Copy directly to date folder
        },
        {
            "name": "C8",
            "base_folder": OUTPUTS_DIR / "NBD_QF_23_C8",
            "subfolder": None  # Copy directly to date folder
        },
        {
            "name": "C3-C10",
            "base_folder": OUTPUTS_DIR / "NBD-QF-23-C3-C10-SF-FI",
            "subfolder": "Input"  # Copy to Input subfolder
        }
    ]

    for config in target_configs:
        try:
            base_folder = config["base_folder"]

            # Check if base folder exists
            if not base_folder.exists():
                _emit_log(f"{config['name']}: Base folder does not exist: {base_folder}", important=True)
                continue

            # Find latest completed date folder
            _, date_folder = find_latest_completed_date(base_folder)

            if not date_folder or not date_folder.exists():
                _emit_log(f"{config['name']}: Could not find latest date folder in {base_folder}", important=True)
                continue

            # Determine the actual target directory
            if config["subfolder"]:
                target_dir = date_folder / config["subfolder"]
                target_dir.mkdir(parents=True, exist_ok=True)
            else:
                target_dir = date_folder

            # Remove old "Prod. wise Class. of Loans*.xlsb" files
            old_files = list(target_dir.glob("Prod. wise Class. of Loans*.xlsb"))
            for old_file in old_files:
                try:
                    old_file.unlink()
                    _emit_log(f"{config['name']}: Removed old file: {old_file.name}", important=True)
                except Exception as e:
                    _emit_log(f"{config['name']}: Failed to remove old file {old_file.name}: {e}", important=True)

            # Copy the new file
            dest_file = target_dir / ia_file_path.name
            shutil.copy2(ia_file_path, dest_file)
            _emit_log(f"{config['name']}: Copied to {dest_file}", important=True)

        except Exception as e:
            _emit_log(f"{config['name']}: Error copying file: {e}", important=True)


def _append_report_message(report_id: str, message: str) -> None:
    """Append a single log line to the per-report buffer for UI display."""
    buf = _report_messages.get(report_id)
    if buf is None:
        _report_messages[report_id] = deque(maxlen=200)
        buf = _report_messages[report_id]
    ts = datetime.now().strftime('%H:%M:%S')
    buf.append(f"[{ts}] {message.strip()}")


def save_master_data(sofp_data: dict | None, ia_data: dict | None, c1c6_data: dict | None = None) -> tuple[bool, str]:
    """Save master data to Master_Data.xlsx"""
    try:
        wb = openpyxl.load_workbook(MASTER_DATA_XLSX)
        
        # Generate current timestamp
        current_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Save SOFP-SOCI data only if provided and non-empty
        if sofp_data:
            has_sofp_values = any((sofp_data.get("bank_loans"), sofp_data.get("securitization_loans"), sofp_data.get("foreign_funding")))
            if has_sofp_values and "NBD-MF-01-SOFP-SOCI" in wb.sheetnames:
                ws = wb["NBD-MF-01-SOFP-SOCI"]
                next_row = ws.max_row + 1
                ws[f"A{next_row}"] = current_timestamp
                ws[f"B{next_row}"] = sofp_data.get("bank_loans", "")
                ws[f"C{next_row}"] = sofp_data.get("securitization_loans", "")
                ws[f"D{next_row}"] = sofp_data.get("foreign_funding", "")
        
        # Save IA data only if provided and non-empty
        if ia_data:
            has_ia_values = bool(ia_data.get("annual_interest_rate"))
            if has_ia_values and "NBD-MF-23-IA-C1&C2-C5-IP" in wb.sheetnames:
                ws = wb["NBD-MF-23-IA-C1&C2-C5-IP"]
                next_row = ws.max_row + 1
                ws[f"A{next_row}"] = current_timestamp
                ws[f"B{next_row}"] = ia_data.get("annual_interest_rate", "")

        # Save C1-C6 data (Guarantees)
        if c1c6_data is not None:
            has_c1c6_values = bool(c1c6_data.get("guarantees"))
            if has_c1c6_values and "NBD-MF-20-C1-C6" in wb.sheetnames:
                ws = wb["NBD-MF-20-C1-C6"]
                next_row = ws.max_row + 1
                ws[f"A{next_row}"] = current_timestamp
                ws[f"B{next_row}"] = c1c6_data.get("guarantees", "")
        
        wb.save(MASTER_DATA_XLSX)
        wb.close()
        return True, "Master data saved successfully"
    except Exception as e:
        return False, f"Failed to save master data: {e}"


def get_latest_master_data() -> tuple[dict, dict, dict]:
    """Get latest master data from Excel sheets for pre-filling form"""
    sofp_data = {"timestamp": "", "bank_loans": "", "securitization_loans": "", "foreign_funding": ""}
    ia_data = {"timestamp": "", "annual_interest_rate": ""}
    c1c6_data = {"timestamp": "", "guarantees": ""}
    
    try:
        wb = openpyxl.load_workbook(MASTER_DATA_XLSX)
        
        # Get latest SOFP-SOCI data
        if "NBD-MF-01-SOFP-SOCI" in wb.sheetnames:
            ws = wb["NBD-MF-01-SOFP-SOCI"]
            if ws.max_row > 1:  # Has data beyond header
                last_row = ws.max_row
                sofp_data = {
                    "timestamp": str(ws[f"A{last_row}"].value or ""),
                    "bank_loans": str(ws[f"B{last_row}"].value or ""),
                    "securitization_loans": str(ws[f"C{last_row}"].value or ""),
                    "foreign_funding": str(ws[f"D{last_row}"].value or "")
                }
        
        # Get latest IA data
        if "NBD-MF-23-IA-C1&C2-C5-IP" in wb.sheetnames:
            ws = wb["NBD-MF-23-IA-C1&C2-C5-IP"]
            if ws.max_row > 1:  # Has data beyond header
                last_row = ws.max_row
                ia_data = {
                    "timestamp": str(ws[f"A{last_row}"].value or ""),
                    "annual_interest_rate": str(ws[f"B{last_row}"].value or "")
                }
        
        # Get latest C1-C6 data
        if "NBD-MF-20-C1-C6" in wb.sheetnames:
            ws = wb["NBD-MF-20-C1-C6"]
            if ws.max_row > 1:
                last_row = ws.max_row
                c1c6_data = {
                    "timestamp": str(ws[f"A{last_row}"].value or ""),
                    "guarantees": str(ws[f"B{last_row}"].value or "")
                }

        wb.close()
    except Exception as e:
        # Return empty data if file doesn't exist or can't be read
        pass
    
    return sofp_data, ia_data, c1c6_data


def get_latest_uploaded_files() -> dict:
    """Get latest uploaded files from Master Inputs directories"""
    file_info = {}

    categories = [
        "ALCL Management Accounts",
        "Investment Schedule",
        "Loan Schedule",
        "Supporting Schedules",
        "Borrowing report",
        "Disbursement with Budget",
        "Information Request from Credit",
        "Net Portfolio",
        "YARD STOCK AS AT",
        "Cadre",
        "Unutilized Amount",
        "Rec Target",
        "Daily Bank Balances",
        "M2M",
        "FD Base as at"
    ]
    
    for category in categories:
        category_dir = MASTER_INPUTS_DIR / category
        if category_dir.exists():
            # Find the most recent file (.xlsx for most categories, .xlsm for Borrowing report)
            if category == "Borrowing report":
                files = list(category_dir.glob("*.xlsm"))
            else:
                files = list(category_dir.glob("*.xlsx"))
            
            if files:
                latest_file = max(files, key=lambda x: x.stat().st_mtime)
                file_info[category] = latest_file.name
            else:
                file_info[category] = "No files uploaded"
        else:
            file_info[category] = "No files uploaded"
    
    return file_info


def copy_uploaded_file_to_sofp_outputs(category: str, file_path: str, month: str) -> None:
    """Copy a single uploaded file to SOFP outputs folder immediately"""
    try:
        # Find the latest SOFP outputs date folder
        sofp_outputs_dir = OUTPUTS_DIR / "NBD_MF_01_SOFP_SOCI"
        if not sofp_outputs_dir.exists():
            _emit_log("SOFP outputs directory not found", important=True)
            return
            
        # Find latest completed date
        last_date, date_folder = find_latest_completed_date(sofp_outputs_dir)
        if not date_folder or not date_folder.exists():
            _emit_log("No completed date folder found for SOFP", important=True)
            return
            
        # Use the original filename (already renamed when stored)
        new_name = Path(file_path).name
            
        # Remove existing files with same prefix
        removed = 0
        _emit_log(f"Scanning {date_folder} for existing files to remove for category: {category}", important=True)
        
        for f in date_folder.iterdir():
            if f.is_file() and f.suffix.lower() == ".xlsx":
                _emit_log(f"Found file: {f.name}", important=True)
                should_remove = False
                
                if category == "ALCL Management Accounts" and f.name.startswith("ALCL Management Accounts"):
                    should_remove = True
                    _emit_log(f"Marking for removal (ALCL): {f.name}", important=True)
                elif category == "Investment Schedule" and f.name.startswith("Investment Schedule"):
                    should_remove = True
                    _emit_log(f"Marking for removal (Investment): {f.name}", important=True)
                elif category == "Loan Schedule" and f.name.startswith("Loan Schedule"):
                    should_remove = True
                    _emit_log(f"Marking for removal (Loan): {f.name}", important=True)
                elif category == "Supporting Schedules" and f.name.startswith("Supporting Schedules"):
                    should_remove = True
                    _emit_log(f"Marking for removal (Supporting): {f.name}", important=True)
                
                if should_remove:
                    try:
                        f.unlink()
                        removed += 1
                        _emit_log(f"Successfully removed existing file: {f.name}", important=True)
                    except Exception as e:
                        _emit_log(f"ERROR: Failed to remove existing file {f.name}: {e}", important=True)
        
        if removed:
            _emit_log(f"Removed {removed} existing {category} file(s) in {date_folder}", important=True)
        
        # Copy the file
        dest_path = date_folder / new_name
        shutil.copy2(file_path, dest_path)
        _emit_log(f"Copied {category} file to {dest_path}", important=True)
        
    except Exception as e:
        _emit_log(f"Failed to copy {category} file to SOFP outputs: {e}", important=True)


def validate_sofp_file_for_ia(report_ids_with_dates: dict) -> tuple[bool, str]:
    """Validate that SOFP file exists for the selected month before running IA"""
    try:
        # Get the selected date from the first report (they should all be the same)
        if not report_ids_with_dates:
            return False, "No report data found"
            
        first_report_id = list(report_ids_with_dates.keys())[0]
        selected_date_str = report_ids_with_dates[first_report_id][0]  # Get the date string
        
        # Parse the selected date to get month and year
        from datetime import datetime
        try:
            # Try different date formats
            try:
                selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d")  # YYYY-MM-DD format
            except ValueError:
                try:
                    selected_date = datetime.strptime(selected_date_str, "%d/%m/%Y")  # DD/MM/YYYY format
                except ValueError:
                    try:
                        selected_date = datetime.strptime(selected_date_str, "%m/%d/%Y")  # MM/DD/YYYY format
                    except ValueError:
                        return False, f"Invalid date format: {selected_date_str}"
            
            month_name = selected_date.strftime("%B")  # Full month name like "October"
            year = selected_date.year
        except Exception as e:
            return False, f"Error parsing date: {e}"
        
        # Check if SOFP file exists in IA outputs using selected month and year
        ia_outputs_dir = OUTPUTS_DIR / "NBD_MF_23_IA"
        if not ia_outputs_dir.exists():
            return False, f"IA outputs directory not found"
            
        # Find latest completed date folder
        last_date, date_folder = find_latest_completed_date(ia_outputs_dir)
        if not date_folder or not date_folder.exists():
            return False, f"No completed IA date folder found"
        
        # Look for the SOFP file with the expected name using selected month and year
        expected_filename = f"NBD-MF-01-SOFP & SOCI AFL Monthly FS {month_name} {year}.xlsx"
        sofp_file_path = date_folder / expected_filename

        if not sofp_file_path.exists():
            # Try to find any SOFP file for debugging
            import glob
            sofp_pattern = str(date_folder / "NBD-MF-01-SOFP & SOCI AFL Monthly FS*.xlsx")
            found_files = glob.glob(sofp_pattern)
            if found_files:
                _emit_log(f"SOFP file exists but name doesn't match. Expected: '{expected_filename}', Found: {[Path(f).name for f in found_files]}", important=True)
                return False, f"'{expected_filename}' file not available. Please run NBD-MF-01-SOFP-SOCI - Statement of Financial position first for {month_name} {year}."
            else:
                return False, f"'{expected_filename}' file not available. Please run NBD-MF-01-SOFP-SOCI - Statement of Financial position first for {month_name} {year}."

        _emit_log(f"SOFP file validation passed for IA: {expected_filename} found", important=True)
        return True, "SOFP file validation passed"
        
    except Exception as e:
        return False, f"Error validating SOFP file for IA: {e}"


def validate_ia_file_for_c1c2(report_ids_with_dates: dict) -> tuple[bool, str]:
    """Validate that IA file exists for the selected month before running C1-C2"""
    try:
        # Get the selected date from the first report (they should all be the same)
        if not report_ids_with_dates:
            return False, "No report data found"
            
        first_report_id = list(report_ids_with_dates.keys())[0]
        selected_date_str = report_ids_with_dates[first_report_id][0]  # Get the date string
        
        # Parse the selected date to get month and year
        from datetime import datetime
        try:
            # Try different date formats
            try:
                selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d")  # YYYY-MM-DD format
            except ValueError:
                try:
                    selected_date = datetime.strptime(selected_date_str, "%d/%m/%Y")  # DD/MM/YYYY format
                except ValueError:
                    try:
                        selected_date = datetime.strptime(selected_date_str, "%m/%d/%Y")  # MM/DD/YYYY format
                    except ValueError:
                        return False, f"Invalid date format: {selected_date_str}"
            
            month_name = selected_date.strftime("%B")  # Full month name like "October"
            year = selected_date.year
        except Exception as e:
            return False, f"Error parsing date: {e}"
        
        # Check if IA file exists in IA outputs using selected month and year
        ia_outputs_dir = OUTPUTS_DIR / "NBD_MF_23_IA"
        if not ia_outputs_dir.exists():
            return False, f"IA outputs directory not found"
            
        # Find latest completed date folder
        last_date, date_folder = find_latest_completed_date(ia_outputs_dir)
        if not date_folder or not date_folder.exists():
            return False, f"No completed IA date folder found"
        
        # Look for the IA file with the expected name using selected month and year
        expected_filename = f"Prod. wise Class. of Loans - {month_name} {year}.xlsb"
        ia_file_path = date_folder / expected_filename
        
        if not ia_file_path.exists():
            return False, f"Prod. wise Class. of Loans - {month_name} {year}.xlsb file not available, run NBD-MF-23-IA - Product wise Classification of Loans first for the {month_name} {year}."
        
        _emit_log(f"IA file validation passed for C1-C2: {expected_filename} found", important=True)
        return True, "IA file validation passed"
        
    except Exception as e:
        return False, f"Error validating IA file for C1-C2: {e}"


def validate_set7_required_files(report_ids_with_dates: dict) -> tuple[bool, str]:
    """Validate if required files exist for Set 7 reports (C2-C6)"""
    try:
        # Get the first report's date to determine month/year
        first_report_id = next(iter(report_ids_with_dates.keys()))
        date_data = report_ids_with_dates[first_report_id]
        
        # Handle both tuple format (date_str, month, year) and string format
        if isinstance(date_data, tuple):
            date_str = date_data[0]  # Extract date string from tuple
        else:
            date_str = date_data  # Use string directly
        
        # Parse the date to get month and year
        try:
            # Try YYYY-MM-DD format first
            parsed_date = datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError:
            try:
                # Try DD/MM/YYYY format
                parsed_date = datetime.strptime(date_str, "%d/%m/%Y")
            except ValueError:
                # Try MM/DD/YYYY format
                parsed_date = datetime.strptime(date_str, "%m/%d/%Y")
        
        month_name = parsed_date.strftime("%B")
        year = parsed_date.year
        
        # Check if C1-C6 outputs directory exists
        c1c6_outputs_dir = OUTPUTS_DIR / "NBD_MF_20_C1_C6"
        if not c1c6_outputs_dir.exists():
            return False, "C1-C6 outputs directory not found. Please run C1-C6 reports first."
            
        # Find latest completed date
        last_date, date_folder = find_latest_completed_date(c1c6_outputs_dir)
        if not date_folder or not date_folder.exists():
            return False, f"No completed C1-C6 date folder found. Please run C1-C6 reports first for {month_name} {year}."
            
        # Check for the IA file (Prod. wise Class. of Loans)
        ia_filename = f"Prod. wise Class. of Loans - {month_name} {year}.xlsb"
        ia_file_path = date_folder / ia_filename
        
        if not ia_file_path.exists():
            return False, f"'{ia_filename}' file not available. Please run NBD-MF-23-IA - Product wise Classification of Loans first for {month_name} {year}."
            
        # Check for the SOFP file
        sofp_filename = f"NBD-MF-01-SOFP & SOCI AFL Monthly FS {month_name} {year}.xlsx"
        sofp_file_path = date_folder / sofp_filename
        
        if not sofp_file_path.exists():
            return False, f"'{sofp_filename}' file not available. Please run NBD-MF-01-SOFP-SOCI - Statement of Financial position first for {month_name} {year}."
            
        return True, f"Set 7 validation passed: '{ia_filename}' and '{sofp_filename}' found"
        
    except Exception as e:
        return False, f"Error validating Set 7 required files: {str(e)}"


def validate_prod_wise_file_for_c8(report_ids_with_dates: dict) -> tuple[bool, str]:
    """Validate that Prod. wise Class. of Loans file exists for the selected month before running C8"""
    try:
        # Get the selected date from the first report
        if not report_ids_with_dates:
            return False, "No report data found"

        first_report_id = list(report_ids_with_dates.keys())[0]
        selected_date_str = report_ids_with_dates[first_report_id][0]  # Get the date string

        # Parse the selected date to get month and year
        try:
            # Try different date formats
            try:
                selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d")  # YYYY-MM-DD format
            except ValueError:
                try:
                    selected_date = datetime.strptime(selected_date_str, "%d/%m/%Y")  # DD/MM/YYYY format
                except ValueError:
                    try:
                        selected_date = datetime.strptime(selected_date_str, "%m/%d/%Y")  # MM/DD/YYYY format
                    except ValueError:
                        return False, f"Invalid date format: {selected_date_str}"

            month_name = selected_date.strftime("%B")  # Full month name like "October"
            year = selected_date.year
        except Exception as e:
            return False, f"Error parsing date: {e}"

        # Check if Prod. wise Class. of Loans file exists in C8 outputs using selected month and year
        c8_outputs_dir = OUTPUTS_DIR / "NBD_QF_23_C8"
        if not c8_outputs_dir.exists():
            return False, f"C8 outputs directory not found"

        # Find latest completed date folder
        last_date, date_folder = find_latest_completed_date(c8_outputs_dir)
        if not date_folder or not date_folder.exists():
            return False, f"No completed C8 date folder found"

        # Look for the Prod. wise Class. of Loans file with the expected name
        expected_filename = f"Prod. wise Class. of Loans - {month_name} {year}.xlsb"
        prod_wise_file_path = date_folder / expected_filename

        if not prod_wise_file_path.exists():
            # Try to find any Prod. wise file for debugging
            import glob
            prod_wise_pattern = str(date_folder / "Prod. wise Class. of Loans*.xlsb")
            found_files = glob.glob(prod_wise_pattern)
            if found_files:
                _emit_log(f"Prod. wise file exists but name doesn't match. Expected: '{expected_filename}', Found: {[Path(f).name for f in found_files]}", important=True)
                return False, f"'{expected_filename}' file not available. Please run NBD-MF-23-IA - Product wise Classification of Loans first for {month_name} {year}."
            else:
                return False, f"'{expected_filename}' file not available. Please run NBD-MF-23-IA - Product wise Classification of Loans first for {month_name} {year}."

        _emit_log(f"Prod. wise file validation passed for C8: {expected_filename} found", important=True)
        return True, "Prod. wise file validation passed"

    except Exception as e:
        return False, f"Error validating Prod. wise file for C8: {e}"


def validate_sofp_file_for_ga_is(report_ids_with_dates: dict) -> tuple[bool, str]:
    """Validate that SOFP file exists for the selected month before running GA IS"""
    try:
        # Get the selected date from the first report (they should all be the same)
        if not report_ids_with_dates:
            return False, "No report data found"
            
        first_report_id = list(report_ids_with_dates.keys())[0]
        selected_date_str = report_ids_with_dates[first_report_id][0]  # Get the date string
        
        # Parse the selected date to get month and year
        from datetime import datetime
        try:
            # Try different date formats
            try:
                selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d")  # YYYY-MM-DD format
            except ValueError:
                try:
                    selected_date = datetime.strptime(selected_date_str, "%d/%m/%Y")  # DD/MM/YYYY format
                except ValueError:
                    try:
                        selected_date = datetime.strptime(selected_date_str, "%m/%d/%Y")  # MM/DD/YYYY format
                    except ValueError:
                        return False, f"Invalid date format: {selected_date_str}"
            
            month_name = selected_date.strftime("%B")  # July, August, etc.
            year = selected_date.year
        except Exception as e:
            return False, f"Error parsing date: {e}"
        
        # Check if SOFP file exists in outputs using selected month and year
        sofp_outputs_dir = OUTPUTS_DIR / "NBD-MF-10-GA & NBD-MF-11-IS"
        if not sofp_outputs_dir.exists():
            return False, f"SOFP outputs directory not found"
            
        # Find latest completed date folder
        last_date, date_folder = find_latest_completed_date(sofp_outputs_dir)
        if not date_folder or not date_folder.exists():
            return False, f"No completed SOFP date folder found"
        
        # Look for the SOFP file with the expected name using selected month and year
        expected_filename = f"NBD-MF-01-SOFP & SOCI AFL Monthly FS {month_name} {year}.xlsx"
        sofp_file_path = date_folder / expected_filename

        if not sofp_file_path.exists():
            # Try to find any SOFP file for debugging
            import glob
            sofp_pattern = str(date_folder / "NBD-MF-01-SOFP & SOCI AFL Monthly FS*.xlsx")
            found_files = glob.glob(sofp_pattern)
            if found_files:
                _emit_log(f"SOFP file exists but name doesn't match. Expected: '{expected_filename}', Found: {[Path(f).name for f in found_files]}", important=True)
                return False, f"'{expected_filename}' file not available. Please run NBD-MF-01-SOFP-SOCI - Statement of Financial position first for {month_name} {year}."
            else:
                return False, f"'{expected_filename}' file not available. Please run NBD-MF-01-SOFP-SOCI - Statement of Financial position first for {month_name} {year}."

        _emit_log(f"SOFP file validation passed: {expected_filename} found", important=True)
        return True, "SOFP file validation passed"
        
    except Exception as e:
        return False, f"Error validating SOFP file: {e}"


def cleanup_duplicate_ia_files() -> None:
    """Clean up duplicate files in IA outputs folder, keeping only the latest ones"""
    try:
        # Find the latest IA outputs date folder
        ia_outputs_dir = OUTPUTS_DIR / "NBD_MF_23_IA"
        if not ia_outputs_dir.exists():
            return
            
        # Find latest completed date
        last_date, date_folder = find_latest_completed_date(ia_outputs_dir)
        if not date_folder or not date_folder.exists():
            return
            
        _emit_log(f"Cleaning up duplicate files in {date_folder}", important=True)
        
        # Use the new general cleanup function
        removed_count = cleanup_duplicate_files_in_folder(date_folder)
        if removed_count > 0:
            _emit_log(f"Removed {removed_count} duplicate files from IA folder", important=True)
        else:
            _emit_log(f"No duplicate files found in IA folder", important=True)
                
    except Exception as e:
        _emit_log(f"Error cleaning up duplicate IA files: {e}", important=True)


def get_full_month_name(month_abbr: str) -> str:
    """Convert month abbreviation to full month name"""
    month_mapping = {
        'Jan': 'January', 'Feb': 'February', 'Mar': 'March', 'Apr': 'April',
        'May': 'May', 'Jun': 'June', 'Jul': 'July', 'Aug': 'August',
        'Sep': 'September', 'Oct': 'October', 'Nov': 'November', 'Dec': 'December',
        'January': 'January', 'February': 'February', 'March': 'March', 'April': 'April',
        'May': 'May', 'June': 'June', 'July': 'July', 'August': 'August',
        'September': 'September', 'October': 'October', 'November': 'November', 'December': 'December'
    }
    return month_mapping.get(month_abbr, month_abbr)


def copy_uploaded_files_to_ia_outputs(uploaded_files: dict, months: dict, years: dict) -> None:
    """Copy uploaded files to IA outputs folder after renaming"""
    try:
        # Find the latest IA outputs date folder
        ia_outputs_dir = OUTPUTS_DIR / "NBD_MF_23_IA"
        if not ia_outputs_dir.exists():
            _emit_log("IA outputs directory not found", important=True)
            return
            
        # Find latest completed date
        last_date, date_folder = find_latest_completed_date(ia_outputs_dir)
        if not date_folder or not date_folder.exists():
            _emit_log("No completed date folder found for IA", important=True)
            return
            
        # Process each uploaded file
        for category, file_path in uploaded_files.items():
            if not file_path or not Path(file_path).exists():
                continue
                
            # Get the month and year for this category
            month_field = f"{category.lower().replace(' ', '_')}_month"
            year_field = f"{category.lower().replace(' ', '_')}_year"
            month = months.get(month_field, "")
            year = years.get(year_field, "")
            
            # For file removal, we don't need month/year - just remove files with matching prefix
            # But we still need month/year for the actual file processing
            if not month or not year:
                _emit_log(f"No month/year specified for {category}, skipping file processing", important=True)
                # Don't continue - we still want to remove old files even if month/year are missing
                
            # Use the original filename (already renamed when stored)
            new_name = Path(file_path).name
                
            # Only copy the file if month and year are available
            if month and year:
                dest_path = date_folder / new_name
                shutil.copy2(file_path, dest_path)
                _emit_log(f"Copied {category} file to {dest_path}", important=True)
                
                # AFTER copying, remove ALL OLD files with same prefix (excluding the one we just copied)
                removed = 0
                _emit_log(f"Scanning {date_folder} for OLD files to remove for category: {category}", important=True)

                # Determine the prefix to match based on category
                prefix_to_remove = None
                if category == "Disbursement with Budget":
                    prefix_to_remove = "Disbursement with Budget"
                elif category == "Information Request from Credit":
                    prefix_to_remove = "Information Request from Credit"
                elif category == "Net Portfolio":
                    prefix_to_remove = "Net Portfolio"

                if prefix_to_remove:
                    for f in date_folder.iterdir():
                        if f.is_file() and f.suffix.lower() == ".xlsx":
                            _emit_log(f"Found xlsx file: {f.name}", important=True)
                            if f.name.startswith(prefix_to_remove) and f.name != new_name:
                                try:
                                    _emit_log(f"Removing OLD file: {f.name} (keeping latest: {new_name})", important=True)
                                    f.unlink()
                                    removed += 1
                                except Exception as e:
                                    _emit_log(f"ERROR: Failed to remove OLD file {f.name}: {e}", important=True)
                            elif f.name.startswith(prefix_to_remove) and f.name == new_name:
                                _emit_log(f"Keeping latest file: {f.name}", important=True)
                            else:
                                _emit_log(f"File does not match prefix '{prefix_to_remove}': {f.name}", important=True)

                    if removed:
                        _emit_log(f"Removed {removed} OLD {category} file(s) in {date_folder}, kept latest: {new_name}", important=True)
                    else:
                        _emit_log(f"No OLD {category} files found to remove, latest file: {new_name}", important=True)
            else:
                _emit_log(f"Skipped copying {category} file due to missing month/year", important=True)
            
    except Exception as e:
        _emit_log(f"Failed to copy uploaded files to IA outputs: {e}", important=True)


def cleanup_duplicate_c1c2_files() -> None:
    """Clean up duplicate files in C1-C2 outputs folder, keeping only the latest ones"""
    try:
        # Find the latest IA outputs date folder (C1-C2 uses IA outputs)
        ia_outputs_dir = OUTPUTS_DIR / "NBD_MF_23_IA"
        if not ia_outputs_dir.exists():
            return
            
        # Find latest completed date
        last_date, date_folder = find_latest_completed_date(ia_outputs_dir)
        if not date_folder or not date_folder.exists():
            return
            
        _emit_log(f"Cleaning up duplicate C1-C2 files in {date_folder}", important=True)
        
        # Use the new general cleanup function
        removed_count = cleanup_duplicate_files_in_folder(date_folder)
        if removed_count > 0:
            _emit_log(f"Removed {removed_count} duplicate files from C1-C2 folder", important=True)
        else:
            _emit_log(f"No duplicate files found in C1-C2 folder", important=True)
                
    except Exception as e:
        _emit_log(f"Error cleaning up duplicate C1-C2 files: {e}", important=True)


def copy_uploaded_files_to_c1c2_outputs(uploaded_files: dict, dates: dict, months: dict, years: dict) -> None:
    """Copy uploaded files to C1-C2 outputs folder after renaming"""
    try:
        # Find the latest IA outputs date folder (C1-C2 uses IA outputs)
        ia_outputs_dir = OUTPUTS_DIR / "NBD_MF_23_IA"
        if not ia_outputs_dir.exists():
            _emit_log("IA outputs directory not found", important=True)
            return
            
        # Find latest completed date
        last_date, date_folder = find_latest_completed_date(ia_outputs_dir)
        if not date_folder or not date_folder.exists():
            _emit_log("No completed date folder found for IA", important=True)
            return
            
        # Process each uploaded file
        for category, file_path in uploaded_files.items():
            if not file_path or not Path(file_path).exists():
                continue
                
            # Get the date/month/year for this category
            if category == "Cadre":
                date_field = f"{category.lower().replace(' ', '_')}_date"
                date_value = dates.get(date_field, "")
                if not date_value:
                    _emit_log(f"No date specified for {category}, skipping", important=True)
                    continue
            else:
                month_field = f"{category.lower().replace(' ', '_')}_month"
                year_field = f"{category.lower().replace(' ', '_')}_year"
                month = months.get(month_field, "")
                year = years.get(year_field, "")
                if not month or not year:
                    _emit_log(f"No month/year specified for {category}, skipping", important=True)
                    continue
                
            # Use the original filename (already renamed when stored)
            new_name = Path(file_path).name
                
            # Remove ALL existing files with same prefix (regardless of date/month/year)
            removed = 0
            _emit_log(f"Scanning {date_folder} for existing files to remove for category: {category}", important=True)

            # Determine the prefix to match based on category
            prefix_to_remove = None
            if category == "Cadre":
                prefix_to_remove = "Cadre"
            elif category == "Unutilized Amount":
                prefix_to_remove = "Unutilized Amount"

            if prefix_to_remove:
                for f in date_folder.iterdir():
                    if f.is_file() and f.suffix.lower() == ".xlsx":
                        _emit_log(f"Found xlsx file: {f.name}", important=True)
                        if f.name.startswith(prefix_to_remove):
                            try:
                                _emit_log(f"Removing old file: {f.name}", important=True)
                                f.unlink()
                                removed += 1
                            except Exception as e:
                                _emit_log(f"ERROR: Failed to remove existing file {f.name}: {e}", important=True)
                        else:
                            _emit_log(f"File does not match prefix '{prefix_to_remove}': {f.name}", important=True)

                if removed:
                    _emit_log(f"Removed {removed} existing {category} file(s) in {date_folder}", important=True)
                else:
                    _emit_log(f"No old {category} files found to remove", important=True)
            
            # Copy the file
            dest_path = date_folder / new_name
            shutil.copy2(file_path, dest_path)
            _emit_log(f"Copied {category} file to {dest_path}", important=True)
            
    except Exception as e:
        _emit_log(f"Failed to copy uploaded files to C1-C2 outputs: {e}", important=True)


def copy_uploaded_files_to_ga_is_outputs(uploaded_files: dict, dates: dict) -> None:
    """Copy uploaded files to GA IS outputs folder after renaming"""
    try:
        # Find the latest GA IS outputs date folder
        ga_is_outputs_dir = OUTPUTS_DIR / "NBD-MF-10-GA & NBD-MF-11-IS"
        if not ga_is_outputs_dir.exists():
            _emit_log("GA IS outputs directory not found", important=True)
            return
            
        # Find latest completed date
        last_date, date_folder = find_latest_completed_date(ga_is_outputs_dir)
        if not date_folder or not date_folder.exists():
            _emit_log("No completed date folder found for GA IS", important=True)
            return
            
        # Process each uploaded file
        for category, file_path in uploaded_files.items():
            if not file_path or not Path(file_path).exists():
                continue
                
            # Use the original filename (already renamed when stored)
            new_name = Path(file_path).name
                
            # Remove existing files with same prefix
            removed = 0
            _emit_log(f"Scanning {date_folder} for existing files to remove for category: {category}", important=True)
            
            for f in date_folder.iterdir():
                if f.is_file() and f.suffix.lower() in [".xlsm", ".xlsx"]:
                    _emit_log(f"Found file: {f.name}", important=True)
                    should_remove = False
                    
                    if category == "Borrowing report" and f.name.startswith("Borrowing report"):
                        should_remove = True
                        _emit_log(f"Marking for removal (Borrowing report): {f.name}", important=True)
                    
                    if should_remove:
                        try:
                            f.unlink()
                            removed += 1
                            _emit_log(f"Successfully removed existing file: {f.name}", important=True)
                        except Exception as e:
                            _emit_log(f"ERROR: Failed to remove existing file {f.name}: {e}", important=True)
            
            if removed:
                _emit_log(f"Removed {removed} existing {category} file(s) in {date_folder}", important=True)
            
            # Copy the file
            dest_path = date_folder / new_name
            shutil.copy2(file_path, dest_path)
            _emit_log(f"Copied {category} file to {dest_path}", important=True)
            
    except Exception as e:
        _emit_log(f"Failed to copy uploaded files to GA IS outputs: {e}", important=True)


def copy_uploaded_files_to_la_outputs(uploaded_files: dict, months: dict, dates: dict) -> None:
    """Copy uploaded files to LA outputs folder after renaming."""
    try:
        # Get the LA outputs directory
        la_outputs_dir = OUTPUTS_DIR / "NBD_MF_15_LA"
        if not la_outputs_dir.exists():
            _emit_log(f"LA outputs directory not found: {la_outputs_dir}", important=True)
            return
        
        # Find the latest completed date folder
        _, date_folder = find_latest_completed_date(la_outputs_dir)
        if not date_folder or not date_folder.exists():
            _emit_log(f"No completed date folder found for LA in {la_outputs_dir}", important=True)
            return
        
        _emit_log(f"Found LA date folder: {date_folder}", important=True)
        
        # Process each uploaded file
        for category, file_path in uploaded_files.items():
            if not file_path or not Path(file_path).exists():
                continue
                
            # Use the original filename (already renamed when stored)
            new_name = Path(file_path).name
            
            # LA only handles specific files - skip others
            if category not in ["Daily Bank Balances", "M2M", "FD Base as at"]:
                _emit_log(f"Skipping {category} - not an LA file", important=True)
                continue

            # LA files go to Input subfolder
            target_dir = date_folder / "Input"
            target_dir.mkdir(parents=True, exist_ok=True)

            # Remove existing files with same prefix (with force removal for locked files)
            removed = 0
            _emit_log(f"Scanning {target_dir} for existing files to remove for category: {category}", important=True)

            for f in target_dir.iterdir():
                if f.is_file() and f.suffix.lower() in [".xlsx", ".xlsm"]:
                    _emit_log(f"Found file: {f.name}", important=True)
                    should_remove = False

                    if category == "Daily Bank Balances" and f.name.startswith("Daily Bank Balances"):
                        should_remove = True
                    elif category == "M2M" and f.name.startswith("M2M"):
                        should_remove = True
                    elif category == "FD Base as at" and f.name.startswith("FD Base as at"):
                        should_remove = True

                    if should_remove:
                        # Try to remove with retries and force kill Excel if needed
                        max_attempts = 5
                        for attempt in range(1, max_attempts + 1):
                            try:
                                f.unlink()
                                removed += 1
                                _emit_log(f"Successfully removed existing file: {f.name}", important=True)
                                break
                            except PermissionError as e:
                                if attempt < max_attempts:
                                    _emit_log(f"File locked, retrying in 2 seconds... (attempt {attempt}/{max_attempts})", important=True)
                                    time.sleep(2)

                                    # On 3rd attempt, try killing Excel processes
                                    if attempt == 3:
                                        _emit_log(f"Attempting to kill Excel processes to unlock file: {f.name}", important=True)
                                        try:
                                            subprocess.run(["taskkill", "/F", "/IM", "EXCEL.EXE"],
                                                         capture_output=True, timeout=5)
                                            time.sleep(1)
                                        except Exception:
                                            pass
                                else:
                                    # Last attempt - try changing permissions and force delete
                                    try:
                                        import stat
                                        os.chmod(str(f), stat.S_IWRITE)
                                        time.sleep(1)
                                        f.unlink()
                                        removed += 1
                                        _emit_log(f"Force removed file after permission change: {f.name}", important=True)
                                    except Exception as final_error:
                                        _emit_log(f"ERROR: Failed to remove old {category} file {f.name} after {max_attempts} attempts: {final_error}", important=True)
                            except Exception as e:
                                _emit_log(f"ERROR: Failed to remove existing file {f.name}: {e}", important=True)
                                break
            
            if removed:
                _emit_log(f"Removed {removed} existing {category} file(s) in {target_dir}", important=True)
            
            # Copy the file
            dest_path = target_dir / new_name
            shutil.copy2(file_path, dest_path)
            _emit_log(f"Copied {category} file to {dest_path}", important=True)
            
    except Exception as e:
        _emit_log(f"Failed to copy uploaded files to LA outputs: {e}", important=True)


def copy_uploaded_files_to_sofp_outputs(uploaded_files: dict, months: dict) -> None:
    """Copy uploaded files to SOFP outputs folder after renaming."""
    try:
        # Get the SOFP outputs directory
        sofp_outputs_dir = OUTPUTS_DIR / "NBD_MF_01_SOFP_SOCI"
        if not sofp_outputs_dir.exists():
            _emit_log(f"SOFP outputs directory not found: {sofp_outputs_dir}", important=True)
            return
        
        # Find the latest completed date folder
        _, date_folder = find_latest_completed_date(sofp_outputs_dir)
        if not date_folder or not date_folder.exists():
            _emit_log(f"No completed date folder found for SOFP in {sofp_outputs_dir}", important=True)
            return
        
        _emit_log(f"Found SOFP date folder: {date_folder}", important=True)
        
        # Process each uploaded file
        for category, file_path in uploaded_files.items():
            if not file_path or not Path(file_path).exists():
                continue
                
            # Use the original filename (already renamed when stored)
            new_name = Path(file_path).name
            
            # Remove existing files with same prefix
            removed = 0
            _emit_log(f"Scanning {date_folder} for existing files to remove for category: {category}", important=True)
            
            for f in date_folder.iterdir():
                if f.is_file() and f.suffix.lower() in [".xlsx", ".xlsm"]:
                    _emit_log(f"Found file: {f.name}", important=True)
                    should_remove = False
                    
                    if category == "ALCL Management Accounts" and f.name.startswith("ALCL Management Accounts"):
                        should_remove = True
                    elif category == "Investment Schedule" and f.name.startswith("Investment Schedule"):
                        should_remove = True
                    elif category == "Loan Schedule" and f.name.startswith("Loan Schedule"):
                        should_remove = True
                    elif category == "Supporting Schedules" and f.name.startswith("Supporting Schedules"):
                        should_remove = True
                    
                    if should_remove:
                        try:
                            f.unlink()
                            removed += 1
                            _emit_log(f"Successfully removed existing file: {f.name}", important=True)
                        except Exception as e:
                            _emit_log(f"ERROR: Failed to remove existing file {f.name}: {e}", important=True)
            
            if removed:
                _emit_log(f"Removed {removed} existing {category} file(s) in {date_folder}", important=True)
            
            # Copy the file
            dest_path = date_folder / new_name
            shutil.copy2(file_path, dest_path)
            _emit_log(f"Copied {category} file to {dest_path}", important=True)
                    
    except Exception as e:
        _emit_log(f"Failed to copy uploaded files to SOFP outputs: {e}", important=True)


def copy_working_to_outputs_with_versioning(report_ids_with_dates: dict) -> None:
    """Copy working folders to outputs with proper versioning after completion."""
    for report_id in report_ids_with_dates:
        if report_id not in ("NBD-MF-01-SOFP-SOCI", "NBD-MF-23-IA"):
            continue
            
        date_str, _, _ = report_ids_with_dates[report_id]
        folder_name = REPORT_STRUCTURE[report_id]["folder"]
        
        # Get working folder
        working_report_dir = WORKING_DIR / folder_name
        if not working_report_dir.exists():
            _emit_log(f"Working folder not found: {working_report_dir}", important=True)
            continue
            
        # Find the date folder in working
        date_folders = [p for p in working_report_dir.iterdir() if p.is_dir()]
        if len(date_folders) != 1:
            _emit_log(f"Expected one date folder in {working_report_dir}, found {len(date_folders)}", important=True)
            continue
            
        source_folder = date_folders[0]
        target_date_name = format_date_folder_name(date_str)
        
        # Create output report directory
        output_report_dir = OUTPUTS_DIR / folder_name
        output_report_dir.mkdir(parents=True, exist_ok=True)
        
        # Check if target date folder exists and find next version
        target_path = output_report_dir / target_date_name
        if target_path.exists():
            version = 1
            while (output_report_dir / f"{target_date_name}({version})").exists():
                version += 1
            target_path = output_report_dir / f"{target_date_name}({version})"
        
        try:
            shutil.copytree(source_folder, target_path)
            _emit_log(f"Copied {source_folder} to {target_path}", important=True)
        except Exception as e:
            _emit_log(f"Failed to copy {source_folder} to {target_path}: {e}", important=True)


def auto_download_files() -> None:
    """Automatically prepare and download files after successful completion."""
    global _download_files

    if not _download_files:
        _emit_log("No files in _download_files to prepare for download", important=True)
        return
    
    # Filter out non-existent files
    existing_files = [f for f in _download_files if f.exists() and f.is_file()]
    
    if not existing_files:
        _emit_log("No existing files found in _download_files", important=True)
        return
    
    # If only one file, keep it as-is for direct download
    if len(existing_files) == 1:
        _download_files = existing_files
        _emit_log(f"Single file ready for download: {existing_files[0].name}", important=True)
        return
    
    # Multiple files - create zip
    temp_zip_dir = BASE_DIR / "temp_zips"
    temp_zip_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    zip_path = temp_zip_dir / f"reports_{timestamp}.zip"

    try:
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            files_added = 0
            
            for file_path in existing_files:
                arcname = f"output_files/{file_path.name}"
                zipf.write(file_path, arcname=arcname)
                files_added += 1
                _emit_log(f"Added to zip: {arcname}", important=False)

            _emit_log(f"Created zip with {files_added} files", important=True)

        if files_added > 0:
            # Store the zip path for potential download
            _download_files = [zip_path]
            _emit_log(f"Files prepared for download: {zip_path}", important=True)
        else:
            _emit_log("No files found to download", important=True)

    except Exception as e:
        _emit_log(f"Failed to prepare files for download: {e}", important=True)


def copy_folder_from_outputs_to_working(source_report_id: str, dest_report_id: str) -> tuple[bool, str, Path | None]:
    """
    Copy the latest completed date folder from outputs/source to working/dest.
    For NBD-MF-23-IA, source is NBD_MF_23_C4.
    """
    if source_report_id not in REPORT_STRUCTURE or dest_report_id not in REPORT_STRUCTURE:
        return False, f"Unknown report: {source_report_id} or {dest_report_id}", None
    
    source_folder_name = REPORT_STRUCTURE[source_report_id]["folder"]
    dest_folder_name = REPORT_STRUCTURE[dest_report_id]["folder"]
    
    source_dir = OUTPUTS_DIR / source_folder_name
    dest_report_dir = WORKING_DIR / dest_folder_name
    
    if not source_dir.exists():
        return False, f"Source folder not found: {source_dir}", None
    
    # Find latest completed date
    last_date, date_folder = find_latest_completed_date(source_dir)
    
    if not date_folder or not date_folder.exists():
        return False, f"No completed date folder found in: {source_dir}", None
    
    # STEP 1: Clean up duplicate files in the source folder BEFORE copying
    _emit_log(f"*** CLEANING DUPLICATES IN SOURCE FOLDER *** {date_folder}", important=True)
    removed_count = cleanup_duplicate_files_in_folder(date_folder)
    if removed_count > 0:
        _emit_log(f"Removed {removed_count} duplicate files from source folder", important=True)
    
    # Create working report directory
    dest_report_dir.mkdir(parents=True, exist_ok=True)
    
    # Remove existing date folders in working
    for child in dest_report_dir.iterdir():
        if child.is_dir():
            shutil.rmtree(child, ignore_errors=True)
    
    # Copy latest folder to working
    dest_path = dest_report_dir / date_folder.name
    try:
        shutil.copytree(date_folder, dest_path)
        _emit_log(f"Copied {date_folder} to {dest_path}", important=True)
        return True, f"Copied {date_folder.name}", dest_path
    except Exception as e:
        return False, f"Failed to copy: {e}", None


def cleanup_duplicate_files_in_folder(folder_path: Path) -> int:
    """
    Clean up duplicate files in a folder by keeping only the latest file.
    Returns the number of files removed.
    """
    if not folder_path.exists():
        return 0
    
    _emit_log(f"*** CLEANING UP DUPLICATE FILES *** in {folder_path}", important=True)
    
    # Group files by base name (without extension)
    file_groups = {}
    removed_count = 0
    
    for file_path in folder_path.iterdir():
        if file_path.is_file():
            # Get base name without extension
            base_name = file_path.stem
            extension = file_path.suffix
            
            if base_name not in file_groups:
                file_groups[base_name] = []
            file_groups[base_name].append(file_path)
    
    # Process each group of files with the same base name
    for base_name, files in file_groups.items():
        if len(files) > 1:
            _emit_log(f"Found {len(files)} duplicate files for '{base_name}':", important=True)
            
            # Sort by modification time (newest first)
            files.sort(key=lambda f: f.stat().st_mtime, reverse=True)
            
            # Keep the newest file, remove the rest
            latest_file = files[0]
            old_files = files[1:]
            
            from datetime import datetime
            _emit_log(f"  KEEPING (latest): {latest_file.name} (modified: {datetime.fromtimestamp(latest_file.stat().st_mtime)})", important=True)
            
            for old_file in old_files:
                try:
                    _emit_log(f"  REMOVING (older): {old_file.name} (modified: {datetime.fromtimestamp(old_file.stat().st_mtime)})", important=True)
                    
                    # Check if file is locked first
                    if is_file_locked(old_file):
                        _emit_log(f"  FILE IS LOCKED - RENAMING: {old_file.name}", important=True)
                        if rename_locked_file(old_file):
                            removed_count += 1
                            _emit_log(f"  SUCCESSFULLY RENAMED LOCKED FILE: {old_file.name}", important=True)
                        else:
                            _emit_log(f"  FAILED TO RENAME LOCKED FILE: {old_file.name}", important=True)
                    else:
                        # File is not locked, remove it normally
                        try:
                            old_file.unlink()
                            removed_count += 1
                            _emit_log(f"  SUCCESSFULLY REMOVED: {old_file.name}", important=True)
                        except Exception as e:
                            _emit_log(f"  FAILED TO REMOVE: {old_file.name} - {e}", important=True)
                except Exception as e:
                    _emit_log(f"  FAILED TO REMOVE: {old_file.name} - {e}", important=True)
        else:
            _emit_log(f"No duplicates for '{base_name}' - keeping: {files[0].name}", important=True)
    
    _emit_log(f"*** DUPLICATE CLEANUP COMPLETE *** Removed {removed_count} duplicate files", important=True)
    return removed_count


def is_file_locked(filepath: Path) -> bool:
    """Check if a file is locked by trying to open it exclusively"""
    if not filepath.exists():
        return False
    
    try:
        # Try to open the file in exclusive mode
        with open(filepath, 'r+b') as f:
            # Try to acquire an exclusive lock (Windows specific)
            import msvcrt
            msvcrt.locking(f.fileno(), msvcrt.LK_NBLCK, 1)
            msvcrt.locking(f.fileno(), msvcrt.LK_UNLCK, 1)
        _emit_log(f"File is NOT locked: {filepath.name}")
        return False
    except (IOError, OSError, PermissionError) as e:
        _emit_log(f"File IS locked: {filepath.name} - {e}")
        return True
    except ImportError:
        # If not on Windows, try a simpler approach
        try:
            with open(filepath, 'r+b'):
                pass
            _emit_log(f"File is NOT locked (non-Windows): {filepath.name}")
            return False
        except (IOError, OSError, PermissionError) as e:
            _emit_log(f"File IS locked (non-Windows): {filepath.name} - {e}")
            return True


def force_remove_file(filepath: Path, max_retries: int = 3, retry_delay: float = 1.0) -> bool:
    """
    Attempt to forcefully remove a file with retries.
    Returns True if successful, False otherwise.
    """
    for attempt in range(max_retries):
        try:
            if filepath.exists():
                # Try normal deletion first
                filepath.unlink()
                _emit_log(f"Successfully removed file: {filepath.name}")
                return True
        except PermissionError:
            if attempt < max_retries - 1:
                _emit_log(f"File locked, attempt {attempt + 1}/{max_retries}. Waiting {retry_delay}s...")
                time.sleep(retry_delay)
                
                # Try to close any file handles (Windows specific)
                try:
                    import subprocess
                    # Use handle.exe if available to close file handles
                    result = subprocess.run(
                        ['handle.exe', str(filepath), '-c', '-nobanner'],
                        capture_output=True,
                        text=True,
                        timeout=5
                    )
                except:
                    pass
            else:
                _emit_log(f"Failed to remove locked file after {max_retries} attempts: {filepath.name}")
                return False
    
    return not filepath.exists()


def rename_locked_file(filepath: Path) -> bool:
    """
    If a file can't be deleted, try renaming it with a timestamp suffix.
    This allows the new file to be copied without creating a duplicate name.
    """
    try:
        if not filepath.exists():
            _emit_log(f"File does not exist, nothing to rename: {filepath.name}")
            return True
            
        # Generate a unique backup name with timestamp
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        stem = filepath.stem
        suffix = filepath.suffix
        backup_name = f"{stem}_OLD_{timestamp}{suffix}"
        backup_path = filepath.parent / backup_name
        
        _emit_log(f"Attempting to rename locked file: {filepath.name} -> {backup_name}")
        
        # Try to rename the file
        filepath.rename(backup_path)
        _emit_log(f"SUCCESSFULLY renamed locked file to: {backup_name}", important=True)
        return True
    except Exception as e:
        _emit_log(f"FAILED to rename locked file {filepath.name}: {e}", important=True)
        return False


def get_files_with_prefix(directory: Path, prefix: str, extensions: Tuple[str, ...] = ('.xlsx', '.xlsm', '.xlsb')) -> List[Path]:
    """Get all files in directory that start with the given prefix and have valid extensions"""
    matching_files = []
    if directory.exists():
        for file_path in directory.iterdir():
            if file_path.is_file():
                file_name = file_path.name
                if file_name.startswith(prefix) and file_name.lower().endswith(extensions):
                    matching_files.append(file_path)
    return matching_files


def copy_master_data_file_with_cleanup_improved(
    file_category: str, 
    source_file_path: Path, 
    target_dir: Path, 
    target_filename: str,
    force_overwrite: bool = True,
    handle_locked_files: str = "rename"  # Options: "rename", "skip", "retry"
) -> Tuple[bool, str]:
    """
    Improved version of copy_master_data_file_with_cleanup that handles locked files better.
    
    Parameters:
    -----------
    file_category : str
        The category of master data file being copied
    source_file_path : Path
        Source file to copy
    target_dir : Path
        Target directory to copy to
    target_filename : str
        Name for the target file
    force_overwrite : bool
        If True, will attempt to overwrite locked files
    handle_locked_files : str
        Strategy for handling locked files:
        - "rename": Rename locked files with timestamp suffix
        - "skip": Skip copying if any old file is locked
        - "retry": Retry deletion with delays
    
    Returns:
    --------
    Tuple[bool, str]: (Success status, Message describing the result)
    """
    try:
        _emit_log(f"COPYING MASTER DATA FILE WITH IMPROVED CLEANUP: {file_category}", important=True)
        _emit_log(f"Source: {source_file_path}")
        _emit_log(f"Target: {target_dir / target_filename}")
        _emit_log(f"Strategy for locked files: {handle_locked_files}")
        
        # Validate source file
        if not source_file_path.exists():
            error_msg = f"Source file does not exist: {source_file_path}"
            _emit_log(error_msg, important=True)
            return False, error_msg
        
        # Create target directory if it doesn't exist
        target_dir.mkdir(parents=True, exist_ok=True)
        
        # Define prefix mapping (same as original)
        prefix_mapping = {
            "Disbursement with Budget": "Disbursement with Budget",
            "Information Request from Credit": "Information Request from Credit",
            "Net Portfolio": "Net Portfolio",
            "YARD STOCK AS AT": "YARD STOCK AS AT",
            "ALCL Management Accounts": "ALCL Management Accounts",
            "Investment Schedule": "Investment Schedule",
            "Loan Schedule": "Loan Schedule",
            "Supporting Schedules": "Supporting Schedules",
            "Borrowing report": "Borrowing report",
            "Cadre": "Cadre",
            "Unutilized Amount": "Unutilized Amount",
            "Daily Bank Balances": "Daily Bank Balances",
            "M2M": "M2M",
            "FD Base as at": "FD Base as at",
            "Rec Target": "Rec Target"
        }
        
        prefix = prefix_mapping.get(file_category, file_category)
        
        # STEP 1: Find and handle old files with the same prefix
        _emit_log(f"STEP 1: CLEANING UP OLD FILES", important=True)
        old_files = get_files_with_prefix(target_dir, prefix)
        _emit_log(f"Found {len(old_files)} old file(s) with prefix '{prefix}' in {target_dir}")
        
        for old_file in old_files:
            _emit_log(f"  - {old_file.name}")
        
        removed_count = 0
        renamed_count = 0
        locked_files = []
        
        for old_file in old_files:
            _emit_log(f"Processing old file: {old_file.name}")
            
            # Check if file is locked
            if is_file_locked(old_file):
                locked_files.append(old_file)
                _emit_log(f"File is locked: {old_file.name}")
                
                # Handle locked file based on strategy
                if handle_locked_files == "rename":
                    if rename_locked_file(old_file):
                        renamed_count += 1
                    else:
                        # If rename fails and we're trying to copy a file with the same name, this is a problem
                        if old_file.name == target_filename:
                            error_msg = f"Cannot overwrite locked file with same name: {old_file.name}"
                            _emit_log(error_msg, important=True)
                            if not force_overwrite:
                                return False, error_msg
                                
                elif handle_locked_files == "retry":
                    if force_remove_file(old_file, max_retries=3, retry_delay=1.0):
                        removed_count += 1
                    elif old_file.name == target_filename:
                        error_msg = f"Cannot remove locked file after retries: {old_file.name}"
                        _emit_log(error_msg, important=True)
                        if not force_overwrite:
                            return False, error_msg
                            
                elif handle_locked_files == "skip":
                    error_msg = f"Skipping copy due to locked file: {old_file.name}"
                    _emit_log(error_msg, important=True)
                    return False, error_msg
            else:
                # File is not locked, remove it normally
                try:
                    old_file.unlink()
                    removed_count += 1
                    _emit_log(f"Removed old file: {old_file.name}")
                except Exception as e:
                    _emit_log(f"Failed to remove file {old_file.name}: {e}")
        
        _emit_log(f"Cleanup complete: Removed {removed_count} file(s), Renamed {renamed_count} file(s)")
        
        # STEP 2: Check if target file already exists and is locked
        target_path = target_dir / target_filename
        if target_path.exists() and is_file_locked(target_path):
            _emit_log(f"Target file already exists and is locked: {target_filename}", important=True)
            
            if handle_locked_files == "rename":
                if not rename_locked_file(target_path):
                    if not force_overwrite:
                        return False, f"Cannot rename existing locked target file: {target_filename}"
            elif handle_locked_files == "retry":
                if not force_remove_file(target_path):
                    if not force_overwrite:
                        return False, f"Cannot remove existing locked target file: {target_filename}"
            else:
                return False, f"Target file is locked: {target_filename}"
        
        # STEP 3: Copy the new file
        _emit_log(f"STEP 2: COPYING NEW FILE", important=True)
        
        try:
            # Use shutil.copy2 to preserve metadata
            shutil.copy2(source_file_path, target_path)
            _emit_log(f"Successfully copied {file_category} to {target_path}")
            
            # Verify the copy
            if target_path.exists():
                source_size = source_file_path.stat().st_size
                target_size = target_path.stat().st_size
                
                if source_size == target_size:
                    success_msg = f"Successfully copied and verified {file_category} ({source_size} bytes)"
                    _emit_log(success_msg, important=True)
                    return True, success_msg
                else:
                    warning_msg = f"File copied but sizes don't match. Source: {source_size}, Target: {target_size}"
                    _emit_log(warning_msg, important=True)
                    return True, warning_msg
            else:
                error_msg = f"File not found after copy operation"
                _emit_log(error_msg, important=True)
                return False, error_msg
                
        except Exception as e:
            error_msg = f"Failed to copy file: {e}"
            _emit_log(error_msg, important=True)
            return False, error_msg
            
    except Exception as e:
        error_msg = f"Unexpected error in copy_master_data_file_with_cleanup_improved: {e}"
        _emit_log(error_msg, important=True)
        return False, error_msg


def copy_folder_to_working(report_id: str) -> tuple[bool, str, Path | None]:
    """Copy latest completed date folder from outputs/<folder> to working/<folder>.

    For this app, the authoritative latest date comes from completed_dates.txt in the
    corresponding outputs/<folder>. We find the last line, locate the matching date
    folder (including version suffix), clear existing subfolders in working/<folder>,
    then copy that folder in full.
    """
    if report_id not in REPORT_STRUCTURE:
        return False, f"Unknown report: {report_id}", None

    folder_name = REPORT_STRUCTURE[report_id]["folder"]
    # Special-case GA & IS SET8: outputs folder uses ampersand name; working uses underscore name
    if report_id == "NBD-MF-10-GA-11-IS-SET8":
        source_report_dir = OUTPUTS_DIR / "NBD-MF-10-GA & NBD-MF-11-IS"
        dest_report_dir = WORKING_DIR / "NBD_MF_10_GA_NBD_MF_11_IS"
    else:
        source_report_dir = OUTPUTS_DIR / folder_name
        dest_report_dir = WORKING_DIR / folder_name
    
    if not source_report_dir.exists():
        return False, f"Report folder not found: {source_report_dir}", None
    
    # Find latest completed date from completed_dates.txt
    last_date, date_folder = find_latest_completed_date(source_report_dir)
    if last_date is None:
        return False, f"No completed_dates.txt or last date in: {source_report_dir}", None
    if not date_folder or not date_folder.exists():
        return False, f"Date folder not found for: {last_date.strftime('%d-%m-%Y')} under {source_report_dir}", None
    
    # STEP 1: Clean up duplicate files in the source folder BEFORE copying
    _emit_log(f"*** CLEANING DUPLICATES IN SOURCE FOLDER *** {date_folder}", important=True)
    removed_count = cleanup_duplicate_files_in_folder(date_folder)
    if removed_count > 0:
        _emit_log(f"Removed {removed_count} duplicate files from source folder", important=True)
    
    dest_report_dir.mkdir(parents=True, exist_ok=True)

    # Preserve CBSL Provision Comparison files ONLY for IA (not for GA/IS or other reports)
    # These files are created by GA/IS and should persist in IA folder across C1-C2 runs
    # But GA/IS process should manage its own CBSL files independently
    cbsl_files_to_preserve = []
    if report_id == "NBD-MF-23-IA":
        for child in list(dest_report_dir.iterdir()):
            if child.is_dir():
                try:
                    for f in child.rglob("CBSL Provision Comparison*.xlsb"):
                        # Save file to temp location
                        temp_path = dest_report_dir / f.name
                        shutil.copy2(f, temp_path)
                        cbsl_files_to_preserve.append(temp_path)
                        _emit_log(f"Preserving CBSL file for IA: {f.name}", important=True)
                except Exception as e:
                    _emit_log(f"Warning: Failed to preserve CBSL files: {e}", important=True)

    # Remove existing subfolders before copying
    for child in list(dest_report_dir.iterdir()):
        if child.is_dir():
            try:
                _emit_log(f"Removing existing folder: {child}", important=True)
                shutil.rmtree(child)
                _emit_log(f"Successfully removed: {child.name}", important=True)
            except PermissionError as e:
                _emit_log(f"Permission error removing {child}: {e}. Trying to handle locked files...", important=True)
                # Try to handle locked files by using onerror callback
                def handle_remove_error(func, path, exc_info):
                    """Error handler for shutil.rmtree - tries to unlock and remove files"""
                    import stat
                    import time
                    try:
                        # Try to change permissions and remove
                        os.chmod(path, stat.S_IWRITE)
                        time.sleep(0.1)  # Brief pause
                        func(path)
                    except Exception:
                        # If still fails, skip this file
                        _emit_log(f"Cannot remove locked file: {path}", important=True)

                try:
                    shutil.rmtree(child, onerror=handle_remove_error)
                    _emit_log(f"Successfully removed {child.name} with error handling", important=True)
                except Exception as e2:
                    _emit_log(f"Still cannot remove {child.name}: {e2}. Attempting force delete...", important=True)
                    # Last resort - try Windows robocopy to mirror an empty directory
                    import subprocess
                    import tempfile
                    try:
                        # Create a temporary empty directory
                        with tempfile.TemporaryDirectory() as empty_dir:
                            # Use robocopy to mirror empty dir over the target (effectively deleting it)
                            subprocess.run(
                                ["robocopy", empty_dir, str(child), "/MIR", "/R:0", "/W:0"],
                                capture_output=True,
                                timeout=30
                            )
                            # robocopy returns 0-7 for success, > 7 for failure
                            # Now try to remove the directory
                            time.sleep(0.5)
                            if child.exists():
                                shutil.rmtree(child, ignore_errors=True)

                        if not child.exists():
                            _emit_log(f"Force deleted {child.name} using robocopy", important=True)
                        else:
                            _emit_log(f"WARNING: Could not fully remove {child.name}, but continuing...", important=True)
                    except Exception as e3:
                        _emit_log(f"Force delete also failed: {e3}. Continuing anyway...", important=True)
            except Exception as e:
                _emit_log(f"Unexpected error removing {child}: {e}", important=True)
                return False, f"Cannot remove existing folder {child.name}: {e}", None

    dest_path = dest_report_dir / date_folder.name

    # Final check - if destination exists, try one more time to remove it
    if dest_path.exists():
        _emit_log(f"Destination still exists after cleanup, attempting final removal: {dest_path}", important=True)
        try:
            import time
            time.sleep(1)  # Wait for any file handles to close
            shutil.rmtree(dest_path, ignore_errors=True)
            if dest_path.exists():
                _emit_log(f"WARNING: Could not remove destination, will try dirs_exist_ok=True", important=True)
        except Exception as e:
            _emit_log(f"Final removal attempt failed: {e}", important=True)

    try:
        # Try to copy with dirs_exist_ok=True to handle any remaining files
        if dest_path.exists():
            shutil.copytree(date_folder, dest_path, dirs_exist_ok=True)
            _emit_log(f"Copied {date_folder} to {dest_path} (with dirs_exist_ok=True)", important=True)
        else:
            shutil.copytree(date_folder, dest_path)
            _emit_log(f"Copied {date_folder} to {dest_path}", important=True)

        # Restore preserved CBSL files to the new date folder
        for cbsl_file in cbsl_files_to_preserve:
            try:
                restored_path = dest_path / cbsl_file.name
                shutil.move(str(cbsl_file), str(restored_path))
                _emit_log(f"Restored CBSL file: {cbsl_file.name}", important=True)
            except Exception as e:
                _emit_log(f"Warning: Failed to restore CBSL file {cbsl_file.name}: {e}", important=True)
                # Clean up temp file if move failed
                try:
                    cbsl_file.unlink(missing_ok=True)
                except Exception:
                    pass

        return True, f"Copied {date_folder.name}", dest_path
    except Exception as e:
        # Clean up preserved files if copy failed
        for cbsl_file in cbsl_files_to_preserve:
            try:
                cbsl_file.unlink(missing_ok=True)
            except Exception:
                pass
        return False, f"Failed to copy: {e}", None


def run_report_script(report_id: str, date_str: str, report_month: str, report_year: str) -> tuple[bool, str]:
    """Run a single report script."""
    if report_id not in REPORT_STRUCTURE:
        return False, f"Unknown report: {report_id}"

    script_name = REPORT_STRUCTURE[report_id]["script"]
    folder_name = REPORT_STRUCTURE[report_id]["folder"]
    script_path = REPORT_AUTOMATIONS_DIR / script_name

    if not script_path.exists():
        msg = f"Script not found: {script_path}"
        _emit_log(msg, important=True)
        return False, msg

    # Determine working directory. For C1/C2, reuse IA's date folder.
    if report_id in ("NBD-MF-23-C1", "NBD-MF-23-C2", "NBD-MF-23-C1-C2"):
        ia_folder_name = REPORT_STRUCTURE["NBD-MF-23-IA"]["folder"]
        ia_working_dir = WORKING_DIR / ia_folder_name

        # For C1-C2 only mode: Use IA outputs directly if working folder doesn't exist
        if not ia_working_dir.exists():
            _emit_log(f"Working directory not found for {report_id}, checking IA outputs", important=True)
            ia_outputs_dir = OUTPUTS_DIR / "NBD_MF_23_IA"
            if not ia_outputs_dir.exists():
                return False, f"Neither working nor outputs directory found for IA"
            # Find latest completed date folder in outputs
            from pathlib import Path
            date_folders = [p for p in ia_outputs_dir.iterdir() if p.is_dir() and not p.name.startswith('.')]
            if not date_folders:
                return False, f"No date folders found in IA outputs"
            # Get the most recent folder
            latest_folder = max(date_folders, key=lambda x: x.stat().st_mtime)
            working_dir = latest_folder
            _emit_log(f"Using IA outputs folder: {working_dir}", important=True)
        else:
            ia_date_folders = [p for p in ia_working_dir.iterdir() if p.is_dir()]
            if len(ia_date_folders) != 1:
                return False, f"Expected exactly one date folder in {ia_working_dir}, found {len(ia_date_folders)}"
            working_dir = ia_date_folders[0]
    elif report_id == "NBD-MF-10-GA-11-IS-SET8":
        # Special case for GA & IS SET8: working folder uses underscore name
        working_report_dir = WORKING_DIR / "NBD_MF_10_GA_NBD_MF_11_IS"
        if not working_report_dir.exists():
            return False, f"Working directory not found: {working_report_dir}"
        date_folders = [p for p in working_report_dir.iterdir() if p.is_dir()]
        if len(date_folders) != 1:
            return False, f"Expected exactly one date folder in {working_report_dir}, found {len(date_folders)}"
        working_dir = date_folders[0]
    else:
        working_report_dir = WORKING_DIR / folder_name
        if not working_report_dir.exists():
            return False, f"Working directory not found: {working_report_dir}"
        date_folders = [p for p in working_report_dir.iterdir() if p.is_dir()]
        if len(date_folders) != 1:
            # Provide detailed error message showing what folders were found
            folder_names = [f.name for f in date_folders] if date_folders else []
            return False, f"Expected exactly one date folder in {working_report_dir}, found {len(date_folders)}: {folder_names}"
        working_dir = date_folders[0]

    _emit_log(f"Starting {report_id}: {REPORT_STRUCTURE[report_id]['name']}", important=True)
    _set_report_status(report_id, "running")

    # Build command according to script-specific CLI expectations
    ui_date = None
    # Convert UI date to the formats scripts expect
    try:
        ui_dt = datetime.strptime(date_str, "%Y-%m-%d")
    except Exception:
        try:
            ui_dt = datetime.strptime(date_str, "%d-%m-%Y")
        except Exception:
            ui_dt = None
    ui_date_mmddyyyy = ui_dt.strftime("%m/%d/%Y") if ui_dt else date_str

    if script_name == "NBD_MF_01_SOFP_SOCI.py":
        cmd = [
            sys.executable,
            str(script_path),
            "--date", ui_date_mmddyyyy,
        ]
    elif script_name == "NBD_MF_23_IA.py":
        cmd = [
            sys.executable,
            str(script_path),
            ui_date_mmddyyyy,
        ]
    elif script_name == "NBD_MF_23_C1C2.py":
        cmd = [
            sys.executable,
            str(script_path),
        ]
    elif script_name == "NBD_MF_20_C3.py":
        # C3 now works like C4 - auto-discovers files from working directory
        # Pass the parent NBD_MF_20_C1_C6 folder, not the dated folder
        base_working_dir = working_dir.parent if working_dir.name != "NBD_MF_20_C1_C6" else working_dir
        cmd = [
            sys.executable,
            str(script_path),
            "--working-dir", str(base_working_dir),
        ]
        if report_month and report_year:
            cmd.extend(["--month", report_month, "--year", report_year])
    elif script_name == "NBD_MF_10_GA_NBD_MF_11_IS.py":
        # GA & IS script expects --working-dir parameter
        cmd = [
            sys.executable,
            str(script_path),
            "--working-dir", str(working_dir),
        ]
    elif script_name == "NBD_QF_23_C8.py":
        # C8 script expects date in MM/DD/YYYY format
        cmd = [
            sys.executable,
            str(script_path),
            ui_date_mmddyyyy,
        ]
    elif script_name == "NBD_MF_15_LA.py":
        # LA script expects date in MM/DD/YYYY format
        cmd = [
            sys.executable,
            str(script_path),
            ui_date_mmddyyyy,
        ]
    else:
        cmd = [
            sys.executable,
            str(script_path),
            "--month", report_month,
            "--year", report_year,
        ]

    log_file = _run_log_dir / f"{report_id}.log"

    try:
        _emit_log(f"Running: {' '.join(cmd)} (cwd={working_dir})", important=True)

        env = os.environ.copy()
        env["PYTHONIOENCODING"] = "utf-8"

        with subprocess.Popen(
            cmd,
            cwd=str(working_dir),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding='utf-8',
            errors='replace',
            bufsize=1,
            env=env,
        ) as proc:
            with _proc_lock:
                _running_processes[report_id] = proc

            with open(log_file, 'w', encoding='utf-8') as lf:
                for line in proc.stdout:
                    if _stop_requested:
                        proc.terminate()
                        _set_report_status(report_id, "stopped")
                        return False, "Stopped by user"

                    raw = line.rstrip()
                    lf.write(raw + '\n')
                    lf.flush()
                    _emit_log(raw, important=False)
                    _append_report_message(report_id, raw)

            ret = proc.wait()

            with _proc_lock:
                if report_id in _running_processes:
                    del _running_processes[report_id]

            if ret != 0:
                _emit_log(f"{report_id} failed with exit code {ret}", important=True)
                _set_report_status(report_id, "failed")
                return False, f"Failed with exit code {ret}"

            _emit_log(f"{report_id} completed successfully", important=True)
            _set_report_status(report_id, "completed")
            return True, "Completed"

    except Exception as exc:
        _emit_log(f"Failed to run {report_id}: {exc}", important=True)
        _set_report_status(report_id, "failed")
        return False, str(exc)


def find_output_files(report_id: str, file_pattern: str) -> list[Path]:
    """
    Find output files in working directory matching the pattern.
    Returns list of matching file paths.
    """
    if report_id not in REPORT_STRUCTURE:
        return []
    
    folder_name = REPORT_STRUCTURE[report_id]["folder"]
    working_report_dir = WORKING_DIR / folder_name
    
    if not working_report_dir.exists():
        return []
    
    date_folders = [p for p in working_report_dir.iterdir() if p.is_dir()]
    if len(date_folders) != 1:
        return []
    
    date_folder = date_folders[0]
    
    # Find files matching pattern
    matches = glob.glob(str(date_folder / file_pattern))
    return [Path(p) for p in matches]


def _copy_sofp_fs_to_outputs_targets() -> None:
    """Copy latest SOFP Monthly FS file from working SOFP date folder to
    latest versioned date folders of outputs/NBD_MF_23_IA and outputs/NBD_MF_20_C1_C6.
    If a file with same name exists at destination, remove it before copying.
    """
    try:
        # Locate source file in working/SOFP date folder
        sofp_dir = WORKING_DIR / REPORT_STRUCTURE["NBD-MF-01-SOFP-SOCI"]["folder"]
        sofp_dates = [p for p in sofp_dir.iterdir() if p.is_dir()]
        if len(sofp_dates) != 1:
            _emit_log("SOFP working folder not in expected single-date structure; skipping extra copies", important=True)
            return
        sofp_date_dir = sofp_dates[0]
        matches = glob.glob(str(sofp_date_dir / "NBD-MF-01-SOFP & SOCI AFL Monthly FS*.xlsx"))
        if not matches:
            _emit_log("No SOFP Monthly FS file found to copy", important=True)
            return
        src_file = Path(max(matches, key=lambda x: Path(x).stat().st_mtime))

        # Helper to copy into a target outputs/<folder> latest date version
        def copy_into_outputs_folder(folder_key: str) -> None:
            out_dir = OUTPUTS_DIR / REPORT_STRUCTURE[folder_key]["folder"]
            last_date, date_folder = find_latest_completed_date(out_dir)
            if not last_date or not date_folder or not date_folder.exists():
                _emit_log(f"Skip copy to {out_dir} (no completed date folder found)", important=True)
                return
            dest_path = date_folder / src_file.name
            try:
                # Remove any existing .xlsx file with the same prefix in the target folder
                prefix = "NBD-MF-01-SOFP & SOCI AFL Monthly FS"
                removed = 0
                for f in date_folder.iterdir():
                    if f.is_file() and f.suffix.lower() == ".xlsx" and f.name.startswith(prefix):
                        try:
                            f.unlink()
                            removed += 1
                        except Exception as _e:
                            _emit_log(f"Warning: failed removing existing file {f}: {_e}", important=True)
                if removed:
                    _emit_log(f"Removed {removed} existing SOFP Monthly FS .xlsx file(s) in {date_folder}", important=True)
                shutil.copy2(src_file, dest_path)
                _emit_log(f"Copied {src_file.name} to {dest_path}", important=True)
            except Exception as e:
                _emit_log(f"Failed to copy {src_file.name} to {dest_path}: {e}", important=True)

        # Copy to IA outputs latest date folder
        copy_into_outputs_folder("NBD-MF-23-IA")
        # Copy to C1_C6 outputs latest date folder
        copy_into_outputs_folder("NBD-MF-20-C1")
        # Copy to GA & IS outputs latest date folder
        try:
            copy_into_outputs_folder("NBD-MF-10-GA-11-IS-SET8")
        except Exception as _e:
            _emit_log(f"Skip copy to GA & IS outputs: {_e}", important=True)

    except Exception as e:
        _emit_log(f"Extra copy of SOFP FS to outputs targets failed: {e}", important=True)


def save_folder_to_outputs(report_id: str, selected_date: str) -> tuple[bool, str, Path | None]:
    """
    Copy report folder from working back to outputs with selected date.
    Returns (success, message, output_path)
    """
    if report_id not in REPORT_STRUCTURE:
        return False, f"Unknown report: {report_id}", None

    folder_name = REPORT_STRUCTURE[report_id]["folder"]
    # Special-case GA & IS SET8: working folder uses underscore name; outputs uses ampersand name
    if report_id == "NBD-MF-10-GA-11-IS-SET8":
        working_report_dir = WORKING_DIR / "NBD_MF_10_GA_NBD_MF_11_IS"
        output_report_dir = OUTPUTS_DIR / "NBD-MF-10-GA & NBD-MF-11-IS"
    else:
        working_report_dir = WORKING_DIR / folder_name
        output_report_dir = OUTPUTS_DIR / folder_name
    
    # Find the date folder in working
    date_folders = [p for p in working_report_dir.iterdir() if p.is_dir()]
    if len(date_folders) != 1:
        return False, f"Expected one date folder in {working_report_dir}", None
    
    source_folder = date_folders[0]
    target_date_name = format_date_folder_name(selected_date)
    
    # Create output report directory
    output_report_dir.mkdir(parents=True, exist_ok=True)
    
    # Check if target date folder exists
    target_path = output_report_dir / target_date_name
    if target_path.exists():
        # Find next version number
        version = 1
        while (output_report_dir / f"{target_date_name}({version})").exists():
            version += 1
        target_path = output_report_dir / f"{target_date_name}({version})"
    
    try:
        shutil.copytree(source_folder, target_path)
        _emit_log(f"Saved {report_id} to {target_path}", important=True)
        return True, f"Saved to {target_path.name}", target_path
    except Exception as e:
        return False, f"Failed to save: {e}", None


def get_report_dependencies(report_id: str) -> list[str]:
    """Get all dependencies for a report (recursive)."""
    if report_id not in REPORT_STRUCTURE:
        return []

    deps = []
    for dep in REPORT_STRUCTURE[report_id]["dependencies"]:
        deps.append(dep)
        deps.extend(get_report_dependencies(dep))

    seen = set()
    result = []
    for item in deps:
        if item not in seen:
            seen.add(item)
            result.append(item)
    return result


def get_execution_order(report_ids: list[str]) -> list[str]:
    """Get execution order including all dependencies."""
    all_reports = set()

    # IA and C1C2 always run together - no standalone C1C2 anymore
    for report_id in report_ids:
        all_reports.add(report_id)
        all_reports.update(get_report_dependencies(report_id))

    sorted_reports = sorted(
        all_reports,
        key=lambda x: (REPORT_STRUCTURE[x]["level"], REPORT_STRUCTURE[x]["set"], REPORT_STRUCTURE[x]["step"])
    )

    return sorted_reports


def run_selected_reports(report_ids_with_dates: dict) -> tuple[bool, str]:
    """Run selected reports in correct order with modular approach."""
    global _download_files, _sofp_download_file
    _download_files = []
    _sofp_download_file = None

    try:
        report_ids = list(report_ids_with_dates.keys())
        execution_order = get_execution_order(report_ids)
    except Exception as e:
        _emit_log(f"Error in execution order: {e}", important=True)
        import traceback
        _emit_log(traceback.format_exc(), important=True)
        return False, f"Error in execution order: {e}"
    # Ensure SOFP appears first when IA is selected WITHOUT C1-C2
    # When both IA and C1-C2 are selected, don't include SOFP
    if "NBD-MF-23-IA" in report_ids and "NBD-MF-23-C1-C2" not in report_ids:
        if "NBD-MF-01-SOFP-SOCI" in execution_order:
            execution_order = ["NBD-MF-01-SOFP-SOCI"] + [r for r in execution_order if r != "NBD-MF-01-SOFP-SOCI"]

    _emit_log(f"Execution order: {', '.join(execution_order)}", important=True)
    only_sofp_flow = set(report_ids) == {"NBD-MF-01-SOFP-SOCI"}

    # Build checklist: Only show reports that will actually run
    checklist_ids = list(execution_order)

    # Determine if we should show SOFP in the checklist
    # SOFP is shown only when IA is selected WITHOUT C1-C2 (but now they always run together)
    # Since IA and C1C2 always run together, we don't show SOFP in the checklist
    # The validation will ensure SOFP file exists before running

    _init_report_checklist(checklist_ids)
    
    # Detect if this is the C1–C6 set flow
    c1c6_ids = {"NBD-MF-20-C1","NBD-MF-20-C2","NBD-MF-20-C3","NBD-MF-20-C4","NBD-MF-20-C5","NBD-MF-20-C6"}
    is_c1c6_flow = any(r in c1c6_ids for r in execution_order)
    is_ga_is_set8 = any(r == "NBD-MF-10-GA-11-IS-SET8" for r in execution_order)
    is_c8_flow = any(r == "NBD-QF-23-C8" for r in execution_order)
    is_c4_flow = any(r == "NBD-MF-23-C4" for r in execution_order)
    is_dm_flow = any(r == "NBD-WF-18-DM" for r in execution_order)
    is_la_flow = any(r == "NBD-WF-15-LA" for r in execution_order)

    # Detect IA flow (IA and C1C2 always run together)
    is_ia_flow = "NBD-MF-23-IA" in report_ids or "NBD-MF-23-C1-C2" in report_ids
    
    # Copy all required reports to working
    _set_status(stage="copying", message="Copying reports to working directory")
    try:
        if is_c1c6_flow:
            # Copy only NBD_MF_20_C1_C6 based on completed_dates.txt
            _emit_log("Preparing working folder for C1–C6 set", important=True)
            ok, msg, _ = copy_folder_to_working("NBD-MF-20-C1")
            if not ok:
                return False, f"Failed to copy NBD-MF-20-C1_C6: {msg}"
        elif is_ga_is_set8:
            _emit_log("Preparing working folder for GA & IS Set 8", important=True)
            ok, msg, _ = copy_folder_to_working("NBD-MF-10-GA-11-IS-SET8")
            if not ok:
                return False, f"Failed to copy GA & IS: {msg}"
        elif is_c8_flow:
            _emit_log("Preparing working folder for C8", important=True)
            ok, msg, _ = copy_folder_to_working("NBD-QF-23-C8")
            if not ok:
                return False, f"Failed to copy C8: {msg}"
        elif is_c4_flow:
            _emit_log("Preparing working folder for C4", important=True)
            ok, msg, _ = copy_folder_to_working("NBD-MF-23-C4")
            if not ok:
                return False, f"Failed to copy C4: {msg}"
        elif is_dm_flow:
            _emit_log("Preparing working folder for DM", important=True)
            # Copy the latest date folder from outputs to working (remove existing folder first)
            try:
                dm_outputs_dir = OUTPUTS_DIR / "NBD-WF-18-DM"
                if dm_outputs_dir.exists():
                    last_date, date_folder = find_latest_completed_date(dm_outputs_dir)
                    if date_folder and date_folder.exists():
                        dm_working_dir = WORKING_DIR / "NBD-WF-18-DM"
                        
                        # Remove existing working folder if it exists
                        if dm_working_dir.exists():
                            import shutil
                            shutil.rmtree(dm_working_dir)
                            _emit_log("Removed existing DM working folder", important=True)
                        
                        # Create new working directory
                        dm_working_dir.mkdir(parents=True, exist_ok=True)
                        
                        # Copy the latest date folder to working
                        target_working_folder = dm_working_dir / date_folder.name
                        shutil.copytree(date_folder, target_working_folder)
                        _emit_log(f"Copied latest DM date folder from outputs to working: {date_folder.name}", important=True)
                    else:
                        _emit_log("No completed DM date folder found in outputs, proceeding with normal copy", important=True)
                        # Do the normal copy if no completed folder found
                        ok, msg, _ = copy_folder_to_working("NBD-WF-18-DM")
                        if not ok:
                            return False, f"Failed to copy DM: {msg}"
                else:
                    _emit_log("DM outputs directory not found, proceeding with normal copy", important=True)
                    # Do the normal copy if outputs directory doesn't exist
                    ok, msg, _ = copy_folder_to_working("NBD-WF-18-DM")
                    if not ok:
                        return False, f"Failed to copy DM: {msg}"
            except Exception as e:
                _emit_log(f"Failed to copy latest DM date folder from outputs: {e}, proceeding with normal copy", important=True)
                # Do the normal copy if copying fails
                ok, msg, _ = copy_folder_to_working("NBD-WF-18-DM")
                if not ok:
                    return False, f"Failed to copy DM: {msg}"
        elif is_la_flow:
            _emit_log("Preparing working folder for LA", important=True)
            ok, msg, _ = copy_folder_to_working("NBD-WF-15-LA")
            if not ok:
                return False, f"Failed to copy LA: {msg}"
        elif is_ia_flow:
            _emit_log("Preparing working folder for IA flow (IA and C1C2 run together)", important=True)
            # Copy IA folder first
            ok, msg, _ = copy_folder_to_working("NBD-MF-23-IA")
            if not ok:
                return False, f"Failed to copy IA: {msg}"
            # C1-C2 will use the same IA working folder
    except Exception as e:
        _emit_log(f"Error during copy phase: {e}", important=True)
        import traceback
        _emit_log(traceback.format_exc(), important=True)
        return False, f"Copy phase error: {e}"
    else:
        allowed_copy = {"NBD-MF-01-SOFP-SOCI"}
        for report_id in execution_order:
            # Copy only SOFP-SOCI and IA; skip others
            if report_id not in allowed_copy:
                _emit_log(f"Skipping copy for {report_id} (copy only SOFP-SOCI and IA)", important=True)
                continue
            success, msg, _ = copy_folder_to_working(report_id)
            if not success:
                return False, f"Failed to copy {report_id}: {msg}"
    
    # Track which scripts have been run to avoid duplicates
    executed_scripts = {}
    all_success = True
    failed_reports = []
    
    # Run reports sequentially
    if is_c1c6_flow:
        # Specific enforced order for C1–C6: C6 → C5 → C4 → C3 → C2
        enforced_order = ["NBD-MF-20-C6","NBD-MF-20-C5","NBD-MF-20-C4","NBD-MF-20-C3","NBD-MF-20-C2"]
        sequence = [rid for rid in enforced_order if rid in execution_order]
    else:
        sequence = execution_order

    for report_id in sequence:
        if _stop_requested:
            _emit_log("Stop requested, halting execution", important=True)
            break
            
        # Get date for this report
        if report_id in report_ids_with_dates:
            date_str, report_month, report_year = report_ids_with_dates[report_id]
        else:
            date_str, report_month, report_year = list(report_ids_with_dates.values())[0]
        
        script_name = REPORT_STRUCTURE[report_id]["script"]

        # Determine if this report should run based on the flow
        # Normal flows (IA and C1C2 always run together)
        run_allowed = (report_id in ("NBD-MF-01-SOFP-SOCI", "NBD-MF-23-IA", "NBD-MF-23-C1-C2")) or (is_c1c6_flow and report_id in c1c6_ids) or (is_ga_is_set8 and report_id == "NBD-MF-10-GA-11-IS-SET8") or (is_c8_flow and report_id == "NBD-QF-23-C8") or (is_c4_flow and report_id == "NBD-MF-23-C4") or (is_dm_flow and report_id == "NBD-WF-18-DM") or (is_la_flow and report_id == "NBD-WF-15-LA") or (is_ia_flow and report_id in ("NBD-MF-23-IA", "NBD-MF-23-C1-C2"))
        
        # Check if this script was already executed
        if script_name in executed_scripts:
            prev_report_id = executed_scripts[script_name]
            _emit_log(f"Skipping {report_id} - script {script_name} already executed by {prev_report_id}", important=True)
            _set_report_status(report_id, "completed")
            continue
        
        # Skip running non-required reports while keeping them visible
        if not run_allowed:
            _emit_log(f"Skipping run for {report_id} (only SOFP-SOCI → IA → C1C2 run in this set)", important=True)
            _set_report_status(report_id, "skipped")
            continue
        
        # Emit clear sequence status line for the UI
        try:
            seq_name = REPORT_STRUCTURE[report_id]["folder"]
        except Exception:
            seq_name = report_id
        _set_status(stage="running", message=f"running {seq_name}")

        # For IA, copy its latest working folder only after SOFP completes
        if not is_c1c6_flow and report_id == "NBD-MF-23-IA":
            _emit_log("Preparing IA working folder after SOFP completion", important=True)
            ok_copy, copy_msg, _copy_path = copy_folder_to_working("NBD-MF-23-IA")
            if not ok_copy:
                _set_report_status(report_id, "failed")
                return False, f"Failed to copy NBD-MF-23-IA: {copy_msg}"
            # Remove old SOFP monthly FS file in IA date folder, if exists
            try:
                ia_working_dir = WORKING_DIR / REPORT_STRUCTURE["NBD-MF-23-IA"]["folder"]
                date_folders = [p for p in ia_working_dir.iterdir() if p.is_dir()]
                if len(date_folders) == 1:
                    ia_date_dir = date_folders[0]
                    pattern = ia_date_dir / "NBD-MF-01-SOFP & SOCI AFL Monthly FS*.xlsx"
                    match_paths = [Path(p) for p in glob.glob(str(pattern))]
                    if len(match_paths) > 1:
                        # Keep the latest by modified time; remove older ones
                        match_paths.sort(key=lambda p: p.stat().st_mtime, reverse=True)
                        keep = match_paths[0]
                        to_remove = match_paths[1:]
                        removed = 0
                        for fp in to_remove:
                            try:
                                fp.unlink(missing_ok=True)
                                removed += 1
                            except Exception:
                                pass
                        if removed > 0:
                            _emit_log(f"Kept latest '{keep.name}', removed {removed} older SOFP Monthly FS file(s) in {ia_date_dir}", important=True)
                else:
                    _emit_log("IA working folder structure unexpected; skip old file cleanup", important=True)
            except Exception as _e:
                _emit_log(f"IA old file cleanup skipped: {_e}", important=True)
        
        # Run the report script
        success, msg = run_report_script(report_id, date_str, report_month, report_year)
        
        if success:
            executed_scripts[script_name] = report_id
            _set_status(stage="running", message=f"completed {seq_name}")
            # After SOFP completes: rename output file with GUI month/year (don't queue for download yet)
            if not is_c1c6_flow and report_id == "NBD-MF-01-SOFP-SOCI":
                try:
                    # Determine GUI month/year from provided date
                    any_date = report_ids_with_dates.get(report_id) or list(report_ids_with_dates.values())[0]
                    _gui_date_str, gui_month, gui_year = any_date
                    sofp_work_dir = WORKING_DIR / REPORT_STRUCTURE["NBD-MF-01-SOFP-SOCI"]["folder"]
                    date_folders = [p for p in sofp_work_dir.iterdir() if p.is_dir()]
                    if len(date_folders) == 1:
                        date_folder = date_folders[0]
                        matches = glob.glob(str(date_folder / "NBD-MF-01-SOFP & SOCI AFL Monthly FS*.xlsx"))
                        if matches:
                            src = Path(max(matches, key=lambda x: Path(x).stat().st_mtime))
                            base = "NBD-MF-01-SOFP & SOCI AFL Monthly FS"
                            new_name = f"{base} {gui_month} {gui_year}.xlsx"
                            dst = src.with_name(new_name)
                            if src.name != new_name:
                                try:
                                    src.rename(dst)
                                    src = dst
                                    _emit_log(f"Renamed SOFP Monthly FS to {dst.name}", important=True)
                                except Exception as e:
                                    _emit_log(f"SOFP rename failed (continuing with original): {e}", important=True)
                            # Store the renamed file info for later download queueing after save
                            _sofp_download_file = src
                        else:
                            _emit_log("SOFP Monthly FS .xlsx not found for renaming", important=True)
                    else:
                        _emit_log("Unexpected SOFP working folder structure; skipping rename", important=True)
                except Exception as e:
                    _emit_log(f"SOFP post-processing failed: {e}", important=True)
            
            # Special handling for NBD-MF-23-C1 (download output file)
            if report_id == "NBD-MF-23-C1":
                _emit_log("Looking for output files to download...", important=True)
                output_files = find_output_files("NBD-MF-23-IA", "Prod. wise Class. of Loans*.xlsb")
                if output_files:
                    _download_files.extend(output_files)
                    _emit_log(f"Added {len(output_files)} file(s) for download", important=True)
                else:
                    _emit_log("Warning: No output files found matching pattern", important=True)
            # After SOFP completes in IA flow, copy its Monthly FS file to outputs/IA and outputs/C1_C6 latest date folders
            if not is_c1c6_flow and report_id == "NBD-MF-01-SOFP-SOCI":
                try:
                    _copy_sofp_fs_to_outputs_targets()
                except Exception as e:
                    _emit_log(f"Warning: extra copy of SOFP FS failed: {e}", important=True)
            # Small delay before starting C1C2 after IA completes
            if not is_c1c6_flow and report_id == "NBD-MF-23-IA":
                try:
                    time.sleep(3)
                except Exception:
                    pass
        else:
            all_success = False
            failed_reports.append(f"{report_id}: {msg}")
    
    # Post-processing: rename/download/archive
    if is_c1c6_flow:
        # Rename the C1–C6 Excel file in working to GUI month/year
        try:
            c1c6_dir = WORKING_DIR / REPORT_STRUCTURE["NBD-MF-20-C1"]["folder"]
            date_folders = [p for p in c1c6_dir.iterdir() if p.is_dir()]
            if len(date_folders) == 1:
                date_folder = date_folders[0]
                # Find target file
                matches = glob.glob(str(date_folder / "NBD-MF-20-C1 to C6*.xlsx"))
                if matches:
                    src = Path(max(matches, key=lambda x: Path(x).stat().st_mtime))
                    # Build new name from GUI month/year (use one of selected dates)
                    any_date = list(report_ids_with_dates.values())[0]
                    _, gui_month, gui_year = any_date
                    new_name = f"NBD-MF-20-C1 to C6 {gui_month} {gui_year}.xlsx"
                    dst = src.with_name(new_name)
                    if src.name != new_name:
                        try:
                            src.rename(dst)
                            _emit_log(f"Renamed file to {dst.name}", important=True)
                            src = dst
                        except Exception as e:
                            _emit_log(f"Rename failed (continuing with original): {e}", important=True)
                    # Queue for download
                    _download_files.append(src)
                else:
                    _emit_log("No C1–C6 output file found to rename/download", important=True)
            else:
                _emit_log("Unexpected working folder structure for C1–C6; skipping rename/download", important=True)
        except Exception as e:
            _emit_log(f"C1–C6 post-processing failed: {e}", important=True)

        # Save working folder back to outputs with versioning using GUI date
        _set_status(stage="saving", message="Saving C1–C6 to outputs")
        date_str, _, _ = list(report_ids_with_dates.values())[0]
        ok, msg, output_path = save_folder_to_outputs("NBD-MF-20-C1", date_str)
        if ok and output_path:
            _emit_log(f"Saved C1–C6: {msg}", important=True)
            _last_output_dirs["NBD-MF-20-C1"] = output_path
    elif is_ga_is_set8:
        # Post-processing for GA & IS: rename files, queue downloads, archive working folder
        try:
            # Determine GUI date
            any_date = list(report_ids_with_dates.values())[0]
            gui_date_str, gui_month, gui_year = any_date
            # working date folder - use underscore name for working directory
            ga_work_dir = WORKING_DIR / "NBD_MF_10_GA_NBD_MF_11_IS"
            date_folders = [p for p in ga_work_dir.iterdir() if p.is_dir()]
            if len(date_folders) == 1:
                date_folder = date_folders[0]
                # Rename GA/IS file
                matches = glob.glob(str(date_folder / "NBD-MF-10-GA & NBD-MF-11-IS*.xlsx"))
                if matches:
                    src = Path(max(matches, key=lambda x: Path(x).stat().st_mtime))
                    new_name = f"NBD-MF-10-GA & NBD-MF-11-IS {gui_month} {gui_year}.xlsx"
                    dst = src.with_name(new_name)
                    if src.name != new_name:
                        try:
                            src.rename(dst)
                            src = dst
                            _emit_log(f"Renamed GA/IS file to {dst.name}", important=True)
                        except Exception as e:
                            _emit_log(f"GA/IS rename failed (continuing): {e}", important=True)
                    _download_files.append(src)
                    _emit_log(f"Added GA/IS file to download queue: {src.name}", important=True)
                else:
                    _emit_log("GA/IS .xlsx file not found for renaming", important=True)

                # Rename CBSL converted file to GUI month/year with .xlsb extension
                cbsl_matches = glob.glob(str(date_folder / "CBSL Provision Comparison*converted.xlsx"))
                cbsl_path_for_delete = None
                if cbsl_matches:
                    cbsl_src = Path(max(cbsl_matches, key=lambda x: Path(x).stat().st_mtime))
                    new_cbsl = cbsl_src.with_name(f"CBSL Provision Comparison - {gui_month} {gui_year}.xlsb")
                    try:
                        cbsl_src.rename(new_cbsl)
                        _emit_log(f"Renamed CBSL converted to {new_cbsl.name}", important=True)
                        _download_files.append(new_cbsl)
                        _emit_log(f"Added CBSL file to download queue: {new_cbsl.name}", important=True)
                        cbsl_path_for_delete = new_cbsl
                    except Exception as e:
                        _emit_log(f"CBSL rename failed: {e}", important=True)

                # Rename working folder to picked date format DD-MM-YYYY
                try:
                    dt = datetime.strptime(gui_date_str, "%Y-%m-%d")
                except Exception:
                    try:
                        dt = datetime.strptime(gui_date_str, "%m/%d/%Y")
                    except Exception:
                        dt = None
                if dt is not None:
                    picked_name = dt.strftime("%d-%m-%Y")
                    if date_folder.name != picked_name:
                        target = date_folder.parent / picked_name
                        if target.exists():
                            shutil.rmtree(target)
                        try:
                            date_folder.rename(target)
                            date_folder = target
                            _emit_log(f"Renamed working folder to {picked_name}", important=True)
                        except Exception as e:
                            _emit_log(f"Failed to rename working folder: {e}", important=True)

                # Save to outputs with versioning using selected date
                ok, msg, output_path = save_folder_to_outputs("NBD-MF-10-GA-11-IS-SET8", gui_date_str)
                if ok and output_path:
                    _emit_log(f"Saved GA & IS: {msg}", important=True)
                    _last_output_dirs["NBD-MF-10-GA-11-IS-SET8"] = output_path
                    
                    # Update file paths in _download_files to point to the new output location
                    updated_download_files = []
                    for file_path in _download_files:
                        if file_path.exists():
                            # File still exists at original location, keep it
                            updated_download_files.append(file_path)
                        else:
                            # File was moved, try to find it in the new output location
                            new_path = output_path / file_path.name
                            if new_path.exists():
                                updated_download_files.append(new_path)
                                _emit_log(f"Updated file path: {file_path.name} -> {new_path}", important=True)
                            else:
                                _emit_log(f"Could not find moved file: {file_path.name}", important=True)
                    
                    _download_files = updated_download_files
                    _emit_log(f"Updated download files after move: {len(_download_files)} files", important=True)

                # Prepare downloads
                _set_status(stage="saving", message="Preparing files for download")
                try:
                    auto_download_files()
                except Exception as e:
                    _emit_log(f"Warning: GA&IS download prep failed: {e}", important=True)

                # Note: CBSL Provision Comparison file is kept for download
                _emit_log("CBSL Provision Comparison file kept for download", important=True)
            else:
                _emit_log("Unexpected working folder structure for GA & IS; skipping post-processing", important=True)
        except Exception as e:
            _emit_log(f"GA & IS post-processing failed: {e}", important=True)
    elif is_c8_flow:
        # Post-processing for C8: rename file, queue download, archive working folder
        try:
            # Determine GUI date
            any_date = list(report_ids_with_dates.values())[0]
            gui_date_str, gui_month, gui_year = any_date
            # working date folder
            c8_work_dir = WORKING_DIR / "NBD_QF_23_C8"
            date_folders = [p for p in c8_work_dir.iterdir() if p.is_dir()]
            if len(date_folders) == 1:
                date_folder = date_folders[0]
                # Rename C8 file: C8-Client Rating-Jun-2025.xlsx
                matches = glob.glob(str(date_folder / "C8-Client Rating*.xlsx"))
                if matches:
                    src = Path(max(matches, key=lambda x: Path(x).stat().st_mtime))
                    new_name = f"C8-Client Rating-{gui_month}-{gui_year}.xlsx"
                    dst = src.with_name(new_name)
                    if src.name != new_name:
                        try:
                            src.rename(dst)
                            src = dst
                            _emit_log(f"Renamed C8 file to {dst.name}", important=True)
                        except Exception as e:
                            _emit_log(f"C8 rename failed (continuing): {e}", important=True)
                    _download_files.append(src)
                else:
                    _emit_log("C8 .xlsx file not found for renaming", important=True)

                # Rename working folder to picked date format DD-MM-YYYY
                try:
                    dt = datetime.strptime(gui_date_str, "%Y-%m-%d")
                except Exception:
                    try:
                        dt = datetime.strptime(gui_date_str, "%m/%d/%Y")
                    except Exception:
                        dt = None
                if dt is not None:
                    picked_name = dt.strftime("%d-%m-%Y")
                    if date_folder.name != picked_name:
                        target = date_folder.parent / picked_name
                        if target.exists():
                            shutil.rmtree(target)
                        try:
                            date_folder.rename(target)
                            date_folder = target
                            _emit_log(f"Renamed working folder to {picked_name}", important=True)
                        except Exception as e:
                            _emit_log(f"Failed to rename working folder: {e}", important=True)

                # Save to outputs with versioning using selected date
                ok, msg, output_path = save_folder_to_outputs("NBD-QF-23-C8", gui_date_str)
                if ok and output_path:
                    _emit_log(f"Saved C8: {msg}", important=True)
                    _last_output_dirs["NBD-QF-23-C8"] = output_path

                    # Update file paths in _download_files to point to the new output location
                    updated_download_files = []
                    for file_path in _download_files:
                        if file_path.exists():
                            # File still exists at original location, keep it
                            updated_download_files.append(file_path)
                        else:
                            # File was moved, try to find it in the new output location
                            new_path = output_path / file_path.name
                            if new_path.exists():
                                updated_download_files.append(new_path)
                                _emit_log(f"Updated file path: {file_path.name} -> {new_path}", important=True)
                            else:
                                _emit_log(f"Could not find moved file: {file_path.name}", important=True)

                    _download_files = updated_download_files
                    _emit_log(f"Updated download files after move: {len(_download_files)} files", important=True)

                # Prepare downloads
                _set_status(stage="saving", message="Preparing files for download")
                try:
                    auto_download_files()
                except Exception as e:
                    _emit_log(f"Warning: C8 download prep failed: {e}", important=True)
            else:
                _emit_log("Unexpected working folder structure for C8; skipping post-processing", important=True)
        except Exception as e:
            _emit_log(f"C8 post-processing failed: {e}", important=True)
    elif is_c4_flow:
        # Post-processing for C4: rename file, queue download, archive working folder
        try:
            # Determine GUI date
            any_date = list(report_ids_with_dates.values())[0]
            gui_date_str, gui_month, gui_year = any_date
            # working date folder
            c4_work_dir = WORKING_DIR / "NBD_MF_23_C4"
            date_folders = [p for p in c4_work_dir.iterdir() if p.is_dir()]
            if len(date_folders) == 1:
                date_folder = date_folders[0]
                # Rename C4 file: Loan CF Analysis NBD-MF-23-C4-Jun-2025.xlsb
                matches = glob.glob(str(date_folder / "Loan CF Analysis*.xlsb"))
                if matches:
                    src = Path(max(matches, key=lambda x: Path(x).stat().st_mtime))
                    new_name = f"Loan CF Analysis NBD-MF-23-C4-{gui_month}-{gui_year}.xlsb"
                    dst = src.with_name(new_name)
                    if src.name != new_name:
                        try:
                            src.rename(dst)
                            src = dst
                            _emit_log(f"Renamed C4 file to {dst.name}", important=True)
                        except Exception as e:
                            _emit_log(f"C4 rename failed (continuing): {e}", important=True)
                    _download_files.append(src)
                else:
                    _emit_log("C4 .xlsb file not found for renaming", important=True)

                # Rename working folder to picked date format DD-MM-YYYY
                try:
                    dt = datetime.strptime(gui_date_str, "%Y-%m-%d")
                except Exception:
                    try:
                        dt = datetime.strptime(gui_date_str, "%m/%d/%Y")
                    except Exception:
                        dt = None
                if dt is not None:
                    picked_name = dt.strftime("%d-%m-%Y")
                    if date_folder.name != picked_name:
                        target = date_folder.parent / picked_name
                        if target.exists():
                            shutil.rmtree(target)
                        try:
                            date_folder.rename(target)
                            date_folder = target
                            _emit_log(f"Renamed working folder to {picked_name}", important=True)
                        except Exception as e:
                            _emit_log(f"Failed to rename working folder: {e}", important=True)

                # Save to outputs with versioning using selected date
                ok, msg, output_path = save_folder_to_outputs("NBD-MF-23-C4", gui_date_str)
                if ok and output_path:
                    _emit_log(f"Saved C4: {msg}", important=True)
                    _last_output_dirs["NBD-MF-23-C4"] = output_path

                # Prepare downloads
                _set_status(stage="saving", message="Preparing files for download")
                try:
                    auto_download_files()
                except Exception as e:
                    _emit_log(f"Warning: C4 download prep failed: {e}", important=True)
            else:
                _emit_log("Unexpected working folder structure for C4; skipping post-processing", important=True)
        except Exception as e:
            _emit_log(f"C4 post-processing failed: {e}", important=True)
    elif is_dm_flow:
        # Post-processing for DM: rename file with week number, queue download, archive working folder
        try:
            # Determine GUI date
            any_date = list(report_ids_with_dates.values())[0]
            gui_date_str, gui_month, gui_year = any_date

            # Calculate week number from the date
            try:
                dt = datetime.strptime(gui_date_str, "%Y-%m-%d")
            except Exception:
                try:
                    dt = datetime.strptime(gui_date_str, "%m/%d/%Y")
                except Exception:
                    dt = datetime.now()

            # Get ISO week number
            week_number = dt.isocalendar()[1]

            # working date folder
            dm_work_dir = WORKING_DIR / "NBD-WF-18-DM"
            date_folders = [p for p in dm_work_dir.iterdir() if p.is_dir()]
            if len(date_folders) == 1:
                date_folder = date_folders[0]
                _emit_log(f"Looking for DM files in: {date_folder}", important=True)
                
                # List all files in the directory for debugging
                all_files = list(date_folder.glob("*"))
                _emit_log(f"Files found in directory: {[f.name for f in all_files]}", important=True)
                
                # Find DM file: NBD-WF-18-DM Deposit Liability Week*.xlsx
                matches = glob.glob(str(date_folder / "NBD-WF-18-DM Deposit Liability Week*.xlsx")) + \
                         glob.glob(str(date_folder / "NBD-WF-18-DM*.xlsx")) + \
                         glob.glob(str(date_folder / "NBD*WF*18*DM*.xlsx"))
                _emit_log(f"Pattern matches found: {matches}", important=True)
                if matches:
                    src = Path(max(matches, key=lambda x: Path(x).stat().st_mtime))
                    new_name = f"NBD-WF-18-DM Deposit Liability Week {week_number}.xlsx"
                    dst = src.with_name(new_name)
                    if src.name != new_name:
                        try:
                            src.rename(dst)
                            src = dst
                            _emit_log(f"Renamed DM file to {dst.name}", important=True)
                        except Exception as e:
                            _emit_log(f"DM rename failed (continuing): {e}", important=True)
                    _download_files.append(src)
                    _emit_log(f"Added main DM file to download queue: {src.name}", important=True)
                else:
                    _emit_log("DM .xlsx file not found for renaming", important=True)
                    # Try to find any Excel file that might be the DM file
                    excel_files = list(date_folder.glob("*.xlsx"))
                    if excel_files:
                        _emit_log(f"Found Excel files but none matched pattern: {[f.name for f in excel_files]}", important=True)
                        # Use the most recent Excel file as fallback
                        src = max(excel_files, key=lambda x: x.stat().st_mtime)
                        _download_files.append(src)
                        _emit_log(f"Added fallback Excel file to download queue: {src.name}", important=True)

                # Add exceptions file if it exists
                exceptions_file = date_folder / "NBD_WF_18_DM_Detailed_Exceptions.txt"
                if exceptions_file.exists():
                    _download_files.append(exceptions_file)
                    _emit_log(f"Added exceptions file to download: {exceptions_file.name}", important=True)
                else:
                    _emit_log("No exceptions file found for DM", important=True)
                
                _emit_log(f"Total files in download queue: {len(_download_files)}", important=True)

                # Rename working folder to picked date format DD-MM-YYYY
                try:
                    dt = datetime.strptime(gui_date_str, "%Y-%m-%d")
                except Exception:
                    try:
                        dt = datetime.strptime(gui_date_str, "%m/%d/%Y")
                    except Exception:
                        dt = None
                if dt is not None:
                    picked_name = dt.strftime("%d-%m-%Y")
                    if date_folder.name != picked_name:
                        target = date_folder.parent / picked_name
                        if target.exists():
                            shutil.rmtree(target)
                        try:
                            date_folder.rename(target)
                            date_folder = target
                            _emit_log(f"Renamed working folder to {picked_name}", important=True)
                        except Exception as e:
                            _emit_log(f"Failed to rename working folder: {e}", important=True)

                # Save to outputs with versioning using selected date
                ok, msg, output_path = save_folder_to_outputs("NBD-WF-18-DM", gui_date_str)
                if ok and output_path:
                    _emit_log(f"Saved DM: {msg}", important=True)
                    _last_output_dirs["NBD-WF-18-DM"] = output_path
                    
                    # Update file paths in _download_files to point to the new output location
                    updated_download_files = []
                    for file_path in _download_files:
                        if file_path.exists():
                            # File still exists at original location, keep it
                            updated_download_files.append(file_path)
                        else:
                            # File was moved, try to find it in the new output location
                            new_path = output_path / file_path.name
                            if new_path.exists():
                                updated_download_files.append(new_path)
                                _emit_log(f"Updated file path: {file_path.name} -> {new_path}", important=True)
                            else:
                                _emit_log(f"Could not find moved file: {file_path.name}", important=True)
                    
                    _download_files = updated_download_files
                    _emit_log(f"Updated download files after move: {len(_download_files)} files", important=True)

                # Prepare downloads
                _set_status(stage="saving", message="Preparing files for download")
                try:
                    auto_download_files()
                except Exception as e:
                    _emit_log(f"Warning: DM download prep failed: {e}", important=True)
            else:
                _emit_log("Unexpected working folder structure for DM; skipping post-processing", important=True)
        except Exception as e:
            _emit_log(f"DM post-processing failed: {e}", important=True)
    elif is_la_flow:
        # Post-processing for LA: rename AFL file with week number, queue download, archive working folder
        try:
            # Determine GUI date
            any_date = list(report_ids_with_dates.values())[0]
            gui_date_str, gui_month, gui_year = any_date

            # Calculate week number from the date
            try:
                dt = datetime.strptime(gui_date_str, "%Y-%m-%d")
            except Exception:
                try:
                    dt = datetime.strptime(gui_date_str, "%m/%d/%Y")
                except Exception:
                    dt = datetime.now()

            # Get ISO week number
            week_number = dt.isocalendar()[1]

            # working date folder
            la_work_dir = WORKING_DIR / "NBD_MF_15_LA"
            date_folders = [p for p in la_work_dir.iterdir() if p.is_dir()]
            if len(date_folders) == 1:
                date_folder = date_folders[0]
                _emit_log(f"Looking for LA files in: {date_folder}", important=True)

                # List all files in the directory for debugging
                all_files = list(date_folder.glob("*"))
                _emit_log(f"Files found in directory: {[f.name for f in all_files]}", important=True)

                # Find AFL file: AFL Liquidity - Week_*.xlsx
                matches = list(date_folder.glob("AFL Liquidity - Week_*.xlsx"))
                _emit_log(f"AFL files found: {[f.name for f in matches]}", important=True)
                if matches:
                    src = Path(max(matches, key=lambda x: Path(x).stat().st_mtime))
                    new_name = f"AFL Liquidity - Week_{week_number}.xlsx"
                    dst = src.with_name(new_name)
                    if src.name != new_name:
                        try:
                            src.rename(dst)
                            src = dst
                            _emit_log(f"Renamed AFL file to {dst.name}", important=True)
                        except Exception as e:
                            _emit_log(f"AFL rename failed (continuing): {e}", important=True)
                    _download_files.append(src)
                    _emit_log(f"Added AFL file to download queue: {src.name}", important=True)
                else:
                    _emit_log("AFL .xlsx file not found", important=True)

                _emit_log(f"Total files in download queue: {len(_download_files)}", important=True)

                # Rename working folder to picked date format DD-MM-YYYY
                try:
                    dt = datetime.strptime(gui_date_str, "%Y-%m-%d")
                except Exception:
                    try:
                        dt = datetime.strptime(gui_date_str, "%m/%d/%Y")
                    except Exception:
                        dt = None
                if dt is not None:
                    picked_name = dt.strftime("%d-%m-%Y")
                    if date_folder.name != picked_name:
                        target = date_folder.parent / picked_name
                        if target.exists():
                            shutil.rmtree(target)
                        try:
                            date_folder.rename(target)
                            date_folder = target
                            _emit_log(f"Renamed working folder to {picked_name}", important=True)
                        except Exception as e:
                            _emit_log(f"Failed to rename working folder: {e}", important=True)

                # Save to outputs with versioning using selected date
                ok, msg, output_path = save_folder_to_outputs("NBD-WF-15-LA", gui_date_str)
                if ok and output_path:
                    _emit_log(f"Saved LA: {msg}", important=True)
                    _last_output_dirs["NBD-WF-15-LA"] = output_path

                    # Update file paths in _download_files to point to the new output location
                    updated_download_files = []
                    for file_path in _download_files:
                        if file_path.exists():
                            # File still exists at original location, keep it
                            updated_download_files.append(file_path)
                        else:
                            # File was moved, try to find it in the new output location
                            new_path = output_path / file_path.name
                            if new_path.exists():
                                updated_download_files.append(new_path)
                                _emit_log(f"Updated file path: {file_path.name} -> {new_path}", important=True)
                            else:
                                _emit_log(f"Could not find moved file: {file_path.name}", important=True)

                    _download_files = updated_download_files
                    _emit_log(f"Updated download files after move: {len(_download_files)} files", important=True)

                # Prepare downloads
                _set_status(stage="saving", message="Preparing files for download")
                try:
                    auto_download_files()
                except Exception as e:
                    _emit_log(f"Warning: LA download prep failed: {e}", important=True)
            else:
                _emit_log("Unexpected working folder structure for LA; skipping post-processing", important=True)
        except Exception as e:
            _emit_log(f"LA post-processing failed: {e}", important=True)
    elif is_ia_flow:
        # Post-processing for IA flow: handle both IA and C1-C2 reports
        try:
            # Determine GUI date
            any_date = list(report_ids_with_dates.values())[0]
            gui_date_str, gui_month, gui_year = any_date
            
            # Handle IA report
            # IA and C1C2 always run together, so IA will save to outputs and C1C2 will also save
            # Both will save their respective outputs

            if "NBD-MF-23-IA" in report_ids_with_dates:
                ia_work_dir = WORKING_DIR / "NBD_MF_23_IA"
                date_folders = [p for p in ia_work_dir.iterdir() if p.is_dir()]
                if len(date_folders) == 1:
                    date_folder = date_folders[0]

                    # Find and rename IA output file
                    ia_matches = glob.glob(str(date_folder / "Prod. wise Class. of Loans*.xlsb"))
                    if ia_matches:
                        src = Path(max(ia_matches, key=lambda x: Path(x).stat().st_mtime))
                        new_name = f"Prod. wise Class. of Loans - {gui_month} {gui_year}.xlsb"
                        dst = src.with_name(new_name)
                        if src.name != new_name:
                            try:
                                src.rename(dst)
                                src = dst
                                _emit_log(f"Renamed IA file to {dst.name}", important=True)
                            except Exception as e:
                                _emit_log(f"IA rename failed (continuing): {e}", important=True)

                    # Rename working folder to picked date format DD-MM-YYYY
                    try:
                        dt = datetime.strptime(gui_date_str, "%Y-%m-%d")
                    except Exception:
                        try:
                            dt = datetime.strptime(gui_date_str, "%m/%d/%Y")
                        except Exception:
                            dt = None
                    if dt is not None:
                        picked_name = dt.strftime("%d-%m-%Y")
                        if date_folder.name != picked_name:
                            target = date_folder.parent / picked_name
                            if target.exists():
                                shutil.rmtree(target)
                            try:
                                date_folder.rename(target)
                                date_folder = target
                                _emit_log(f"Renamed working folder to {picked_name}", important=True)
                            except Exception as e:
                                _emit_log(f"Failed to rename working folder: {e}", important=True)

                    # IA files stay in working folder for now
                    # They will be saved to outputs after C1-C2 completes
                    _emit_log("IA completed - files ready in working folder", important=True)

            # Handle C1-C2 report (if selected)
            if "NBD-MF-23-C1-C2" in report_ids_with_dates:
                # C1-C2 uses the same working folder as IA
                ia_work_dir = WORKING_DIR / "NBD_MF_23_IA"
                date_folders = [p for p in ia_work_dir.iterdir() if p.is_dir()]
                if len(date_folders) == 1:
                    date_folder = date_folders[0]

                    # Find the updated C1-C2 file (already renamed by IA section)
                    c1c2_file_path = date_folder / f"Prod. wise Class. of Loans - {gui_month} {gui_year}.xlsb"
                    if c1c2_file_path.exists():
                        # Add to download queue
                        _download_files.append(c1c2_file_path)
                        _emit_log(f"Added C1-C2 file to download queue: {c1c2_file_path.name}", important=True)

                        # Copy updated C1-C2 file from working folder to C1-C6, C8, and C3-C10/Input
                        _emit_log(f"Copying updated C1-C2 file to additional output folders...", important=True)
                        copy_ia_to_additional_outputs(c1c2_file_path, gui_date_str)
                    else:
                        _emit_log(f"Warning: C1-C2 file not found in working folder", important=True)

                    # Find and add EXCEPTIONS.xlsx if it exists
                    exceptions_file = date_folder / "EXCEPTIONS.xlsx"
                    if exceptions_file.exists():
                        _download_files.append(exceptions_file)
                        _emit_log(f"Added EXCEPTIONS.xlsx to download queue", important=True)

                    # Find and add Minimum_Rate_Exceptions.xlsx if it exists
                    min_rate_exceptions = date_folder / "Minimum_Rate_Exceptions.xlsx"
                    if min_rate_exceptions.exists():
                        _download_files.append(min_rate_exceptions)
                        _emit_log(f"Added Minimum_Rate_Exceptions.xlsx to download queue", important=True)

                    # Save working folder to IA outputs (not C1-C2 outputs)
                    ok, msg, output_path = save_folder_to_outputs("NBD-MF-23-IA", gui_date_str)
                    if ok and output_path:
                        _emit_log(f"Saved IA/C1-C2 working folder to IA outputs: {msg}", important=True)
                        _last_output_dirs["NBD-MF-23-IA"] = output_path
                        _last_output_dirs["NBD-MF-23-C1-C2"] = output_path
                    else:
                        _emit_log(f"Warning: Could not save working folder to outputs: {msg}", important=True)

                    # C1C2 uses Ctrl+S method - files stay in working location for download
                    _emit_log("C1C2 completed - files ready for download from working location", important=True)

            # Prepare downloads
            _set_status(stage="saving", message="Preparing files for download")
            try:
                auto_download_files()
            except Exception as e:
                _emit_log(f"Warning: IA flow download prep failed: {e}", important=True)
                
        except Exception as e:
            _emit_log(f"IA flow post-processing failed: {e}", important=True)
    else:
        # Save completed reports back to outputs with selected dates (IA/SOFP flow)
        _set_status(stage="saving", message="Saving reports to outputs")
        for report_id in report_ids:
            # Skip saving for C1/C2 as they don't have their own working folder
            if report_id in ("NBD-MF-23-C1", "NBD-MF-23-C2"):
                continue
            if _stop_requested:
                break
            
            date_str, _, _ = report_ids_with_dates[report_id]
            success, msg, output_path = save_folder_to_outputs(report_id, date_str)
            if success and output_path:
                _emit_log(f"Saved {report_id}: {msg}", important=True)
                _last_output_dirs[report_id] = output_path
                # For SOFP-only flow, queue the file from working directory for download
                if only_sofp_flow and report_id == "NBD-MF-01-SOFP-SOCI" and _sofp_download_file:
                    # Download from working directory where the file was renamed
                    if _sofp_download_file.exists():
                        _download_files = [_sofp_download_file]
                        _emit_log(f"Queued SOFP file from working directory for download: {_sofp_download_file}", important=True)
    
    if _stop_requested:
        return False, "Process stopped by user"
    
    if all_success:
        # Outputs have already been saved explicitly per-flow; avoid duplicating version folders
        
        # Prepare downloads: for SOFP-only, return the single renamed .xlsx; otherwise create zip
        if not only_sofp_flow:
            _set_status(stage="saving", message="Preparing files for download")
            try:
                auto_download_files()
            except Exception as e:
                _emit_log(f"Warning: Failed to prepare files for download: {e}", important=True)
        
        return True, "All reports completed successfully"
    else:
        return False, f"Some reports failed: {'; '.join(failed_reports)}"


# Master data file mapping - defines which files are used by which reports
MASTER_DATA_FILE_MAPPING = {
    "ALCL Management Accounts": {
        "reports": ["NBD-MF-01-SOFP-SOCI", "NBD-WF-15-LA"],
        "filename_pattern": "ALCL Management Accounts {month}_{year}.xlsx",
        "target_subfolder": None  # Goes to main date folder
    },
    "Investment Schedule": {
        "reports": ["NBD-MF-01-SOFP-SOCI", "NBD-MF-10-GA-11-IS-SET8", "NBD-WF-15-LA"],
        "filename_pattern": "Investment Schedule - {month} {year}.xlsx",
        "target_subfolder": None,  # Default: main date folder
        "report_specific_subfolders": {
            "NBD-WF-15-LA": "Input"  # LA needs Input subfolder
        }
    },
    "Loan Schedule": {
        "reports": ["NBD-MF-01-SOFP-SOCI", "NBD-WF-15-LA"],
        "filename_pattern": "Loan Schedule - {month} {year}.xlsx",
        "target_subfolder": None,  # Default: main date folder
        "report_specific_subfolders": {
            "NBD-WF-15-LA": "Input"  # LA needs Input subfolder
        }
    },
    "Supporting Schedules": {
        "reports": ["NBD-MF-01-SOFP-SOCI"],
        "filename_pattern": "Supporting Schedules - {month} {year}.xlsx",
        "target_subfolder": None  # Goes to main date folder
    },
    "Borrowing report": {
        "reports": ["NBD-MF-10-GA-11-IS-SET8"],
        "filename_pattern": "Borrowing report {date}.xlsm",
        "target_subfolder": None  # Goes to main date folder
    },
    "Disbursement with Budget": {
        "reports": ["NBD-MF-23-IA"],
        "filename_pattern": "Disbursement with Budget - {month} {year}.xlsx",
        "target_subfolder": None  # Goes to main date folder
    },
    "Information Request from Credit": {
        "reports": ["NBD-MF-23-IA"],
        "filename_pattern": "Information Request from Credit-{month} {year}.xlsx",
        "target_subfolder": None  # Goes to main date folder
    },
    "Net Portfolio": {
        "reports": ["NBD-MF-23-IA"],
        "filename_pattern": "Net Portfolio-{month}-{year}.xlsx",
        "target_subfolder": None  # Goes to main date folder
    },
    "YARD STOCK AS AT": {
        "reports": ["NBD-MF-23-IA"],
        "filename_pattern": "YARD STOCK AS AT {date} FINAL.xlsx",
        "target_subfolder": None  # Goes to main date folder
    },
    "Cadre": {
        "reports": ["NBD-MF-23-C1-C2"],
        "filename_pattern": "Cadre{date}.xlsx",
        "target_subfolder": None  # Goes to main date folder
    },
    "Unutilized Amount": {
        "reports": ["NBD-MF-23-C1-C2"],
        "filename_pattern": "Unutilized Amount {month} {year}.xlsx",
        "target_subfolder": None  # Goes to main date folder
    },
    "Daily Bank Balances": {
        "reports": ["NBD-WF-15-LA"],
        "filename_pattern": "Daily Bank Balances - {month} {year}.xlsx",
        "target_subfolder": "Input"  # Goes to Input subfolder
    },
    "M2M": {
        "reports": ["NBD-WF-15-LA"],
        "filename_pattern": "M2M {mmdd}.xlsx",
        "target_subfolder": "Input"  # Goes to Input subfolder
    },
    "FD Base as at": {
        "reports": ["NBD-WF-15-LA"],
        "filename_pattern": "FD Base as at {date}.xlsx",
        "target_subfolder": "Input"  # Goes to Input subfolder
    },
    "Rec Target": {
        "reports": ["NBD-MF-23-C4"],
        "filename_pattern": "Rec Target {month} {year}.xlsx",
        "target_subfolder": None  # Goes to main date folder
    }
}

# Global counter for debugging
_centralized_call_count = 0
_processed_files = set()  # Track processed files to prevent duplicates

def update_master_data_across_reports(file_category: str, source_file_path: Path, month: str = None, year: str = None, date: str = None) -> None:
    global _centralized_call_count, _processed_files
    
    # Create a unique key for this file processing
    file_key = f"{file_category}_{source_file_path.name}_{month}_{year}_{date}"
    
    # Check if this file has already been processed
    if file_key in _processed_files:
        _emit_log(f"*** SKIPPING DUPLICATE PROCESSING *** {file_category} from {source_file_path} (already processed)", important=True)
        return
    
    # Mark this file as being processed
    _processed_files.add(file_key)
    
    # Additional check: if the source file doesn't exist, skip processing
    if not source_file_path.exists():
        _emit_log(f"*** SKIPPING PROCESSING - SOURCE FILE NOT FOUND *** {file_category} from {source_file_path}", important=True)
        return
    
    _centralized_call_count += 1
    _emit_log(f"*** CENTRALIZED FUNCTION CALLED #{_centralized_call_count} *** {file_category} from {source_file_path}", important=True)
    """
    Update a master data file across all reports that use it.
    
    Args:
        file_category: The category name from MASTER_DATA_FILE_MAPPING
        source_file_path: Path to the source file in Master Inputs
        month: Month name (for files that use month/year)
        year: Year (for files that use month/year)
        date: Date string (for files that use date)
    """
    if file_category not in MASTER_DATA_FILE_MAPPING:
        _emit_log(f"Unknown file category: {file_category}", important=True)
        return
    
    file_config = MASTER_DATA_FILE_MAPPING[file_category]
    reports = file_config["reports"]
    filename_pattern = file_config["filename_pattern"]
    target_subfolder = file_config["target_subfolder"]
    
    _emit_log(f"=== CENTRALIZED UPDATE START === {file_category} across {len(reports)} reports: {reports}", important=True)
    _emit_log(f"Source file: {source_file_path}", important=True)
    _emit_log(f"Parameters - month: {month}, year: {year}, date: {date}", important=True)
    
    # Generate the target filename
    if month and year:
        target_filename = filename_pattern.format(month=month, year=year)
    elif date:
        if file_category == "Borrowing report":
            # Convert YYYY-MM-DD to DD-MM-YYYY for Borrowing report
            date_obj = datetime.strptime(date, "%Y-%m-%d")
            formatted_date = date_obj.strftime("%d-%m-%Y")
            target_filename = filename_pattern.format(date=formatted_date)
        elif file_category == "M2M":
            # Convert YYYY-MM-DD to MMDD for M2M
            date_obj = datetime.strptime(date, "%Y-%m-%d")
            mmdd = date_obj.strftime("%m%d")
            target_filename = filename_pattern.format(mmdd=mmdd)
        elif file_category == "FD Base as at":
            # Convert YYYY-MM-DD to DD.MM.YYYY for FD Base as at
            date_obj = datetime.strptime(date, "%Y-%m-%d")
            formatted_date = date_obj.strftime("%d.%m.%Y")
            target_filename = filename_pattern.format(date=formatted_date)
        elif file_category == "Cadre":
            # Convert YYYY-MM-DD to DD MM YYYY for Cadre
            date_obj = datetime.strptime(date, "%Y-%m-%d")
            formatted_date = date_obj.strftime("%d %m %Y")
            target_filename = filename_pattern.format(date=formatted_date)
        elif file_category == "YARD STOCK AS AT":
            # Convert YYYY-MM-DD to DD-MM-YYYY for YARD STOCK AS AT
            date_obj = datetime.strptime(date, "%Y-%m-%d")
            formatted_date = date_obj.strftime("%d-%m-%Y")
            target_filename = filename_pattern.format(date=formatted_date)
        else:
            target_filename = filename_pattern.format(date=date)
    else:
        _emit_log(f"Missing required parameters for {file_category}", important=True)
        return
    
    # Copy to each report's outputs folder
    _emit_log(f"Processing {file_category} for {len(reports)} reports: {reports}", important=True)
    for report_id in reports:
        try:
            _emit_log(f"*** PROCESSING {file_category} FOR REPORT: {report_id} ***", important=True)
            # Get the outputs folder for this report
            if report_id == "NBD-MF-10-GA-11-IS-SET8":
                # Special case for GA IS
                report_outputs_dir = OUTPUTS_DIR / "NBD-MF-10-GA & NBD-MF-11-IS"
            else:
                report_outputs_dir = OUTPUTS_DIR / REPORT_STRUCTURE[report_id]["folder"]
            
            _emit_log(f"Processing {file_category} for {report_id} -> {report_outputs_dir}", important=True)
            
            if not report_outputs_dir.exists():
                _emit_log(f"Outputs directory not found for {report_id}: {report_outputs_dir}, creating it", important=True)
                report_outputs_dir.mkdir(parents=True, exist_ok=True)
            
            # Find the latest completed date folder
            _, date_folder = find_latest_completed_date(report_outputs_dir)
            if not date_folder or not date_folder.exists():
                _emit_log(f"No completed date folder found for {report_id} in {report_outputs_dir}, creating placeholder", important=True)
                # Create a placeholder date folder using today's date
                today = datetime.now()
                date_folder_name = today.strftime("%d-%m-%Y")
                date_folder = report_outputs_dir / date_folder_name
                date_folder.mkdir(parents=True, exist_ok=True)
                _emit_log(f"Created placeholder date folder: {date_folder}", important=True)
            
            _emit_log(f"Found date folder for {report_id}: {date_folder}", important=True)
            
            # Determine target path
            # Check for report-specific subfolder first, then fall back to default
            report_specific_subfolders = file_config.get("report_specific_subfolders", {})
            if report_id in report_specific_subfolders:
                target_subfolder = report_specific_subfolders[report_id]
            else:
                target_subfolder = file_config.get("target_subfolder")
            
            if target_subfolder:
                target_dir = date_folder / target_subfolder
                target_dir.mkdir(parents=True, exist_ok=True)
                _emit_log(f"Using subfolder for {report_id}: {target_subfolder}", important=True)
            else:
                target_dir = date_folder
                _emit_log(f"Using main date folder for {report_id}", important=True)
            
            # Copy the new file FIRST using win32
            target_path = target_dir / target_filename
            _emit_log(f"Copying {file_category} from {source_file_path} to {target_path}", important=True)
            _emit_log(f"Source file exists: {source_file_path.exists()}", important=True)
            _emit_log(f"Target directory exists: {target_dir.exists()}", important=True)
            _emit_log(f"Target path: {target_path}", important=True)
            # Skip the old win32 copy logic - we'll use the new approach below
            
            # Use the new master data file copying approach
            success, message = copy_master_data_file_with_cleanup_improved(
                file_category, 
                source_file_path, 
                target_dir, 
                target_filename,
                force_overwrite=True,
                handle_locked_files="rename"
            )
            if not success:
                _emit_log(f"Failed to copy {file_category} to {report_id}: {message}", important=True)
                return
            else:
                _emit_log(f"Successfully copied {file_category} to {report_id}: {message}", important=True)
            
            # Simple approach: No complex file removal needed
            # The shutil.copy2 will overwrite existing files automatically
            _emit_log(f"*** SIMPLE APPROACH COMPLETE *** {file_category} -> {report_id}", important=True)
            
            # Simple validation: just check if the file exists
            if target_path.exists():
                _emit_log(f"*** SUCCESS: FILE COPIED AND VERIFIED *** {file_category} -> {report_id}: {target_path.name}", important=True)
            else:
                _emit_log(f"*** ERROR: FILE NOT FOUND AFTER COPY *** {file_category} -> {report_id}", important=True)
            
        except Exception as e:
            _emit_log(f"Failed to update {file_category} for {report_id}: {e}", important=True)
    
    _emit_log(f"=== CENTRALIZED UPDATE COMPLETE === {file_category}", important=True)


def test_file_removal_logic(target_dir: Path, prefix: str) -> None:
    """Test function to verify file removal logic works correctly."""
    _emit_log(f"*** TESTING FILE REMOVAL LOGIC ***", important=True)
    _emit_log(f"Target directory: {target_dir}", important=True)
    _emit_log(f"Prefix: '{prefix}'", important=True)
    
    if not target_dir.exists():
        _emit_log(f"ERROR: Target directory does not exist: {target_dir}", important=True)
        return
    
    # Get all files in the directory
    all_files = [f for f in target_dir.iterdir() if f.is_file()]
    _emit_log(f"Found {len(all_files)} files in target directory", important=True)
    
    for f in all_files:
        file_name = f.name
        file_suffix = f.suffix.lower()
        starts_with_prefix = file_name.startswith(prefix)
        has_valid_suffix = file_suffix in [".xlsx", ".xlsm"]
        
        _emit_log(f"File: '{file_name}' (prefix: '{prefix}', starts_with: {starts_with_prefix}, suffix: '{file_suffix}', valid_suffix: {has_valid_suffix})", important=True)
        
        if starts_with_prefix and has_valid_suffix:
            _emit_log(f"WOULD REMOVE: {file_name}", important=True)
        else:
            _emit_log(f"WOULD KEEP: {file_name}", important=True)

def ensure_directories_exist() -> None:
    """Ensure all necessary directories exist."""
    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    WORKING_DIR.mkdir(parents=True, exist_ok=True)
    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    MASTER_INPUTS_DIR.mkdir(parents=True, exist_ok=True)
    
    # Create Master Inputs subdirectories
    (MASTER_INPUTS_DIR / "ALCL Management Accounts").mkdir(parents=True, exist_ok=True)
    (MASTER_INPUTS_DIR / "Investment Schedule").mkdir(parents=True, exist_ok=True)
    (MASTER_INPUTS_DIR / "Loan Schedule").mkdir(parents=True, exist_ok=True)
    (MASTER_INPUTS_DIR / "Supporting Schedules").mkdir(parents=True, exist_ok=True)
    (MASTER_INPUTS_DIR / "Borrowing report").mkdir(parents=True, exist_ok=True)
    (MASTER_INPUTS_DIR / "Disbursement with Budget").mkdir(parents=True, exist_ok=True)
    (MASTER_INPUTS_DIR / "Information Request from Credit").mkdir(parents=True, exist_ok=True)
    (MASTER_INPUTS_DIR / "Net Portfolio").mkdir(parents=True, exist_ok=True)
    (MASTER_INPUTS_DIR / "YARD STOCK AS AT").mkdir(parents=True, exist_ok=True)
    (MASTER_INPUTS_DIR / "Cadre").mkdir(parents=True, exist_ok=True)
    (MASTER_INPUTS_DIR / "Unutilized Amount").mkdir(parents=True, exist_ok=True)
    (MASTER_INPUTS_DIR / "Daily Bank Balances").mkdir(parents=True, exist_ok=True)
    (MASTER_INPUTS_DIR / "M2M").mkdir(parents=True, exist_ok=True)
    (MASTER_INPUTS_DIR / "FD Base as at").mkdir(parents=True, exist_ok=True)
    
    # Ensure Master_Data.xlsx exists with required sheets
    if not MASTER_DATA_XLSX.exists():
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        wb.create_sheet("NBD-MF-01-SOFP-SOCI")
        wb.create_sheet("NBD-MF-23-IA-C1&C2-C5-IP")
        wb.create_sheet("NBD-MF-20-C1-C6")
        wb.save(MASTER_DATA_XLSX)
        wb.close()


# ==================== FLASK ROUTES ====================

app.secret_key = os.environ.get("FLASK_SECRET_KEY", "change-me-in-prod")


@app.route("/", methods=["GET"])
def index():
    from flask import session

    ensure_directories_exist()
    today_iso = datetime.today().strftime("%Y-%m-%d")

    # Clear session if requested (after user clicks "Thanks" in modal)
    if request.args.get('clear') == '1':
        session.pop('download_completed', None)
        session.pop('download_filename', None)

    # Group reports by level and set
    reports_by_level = {}
    sets_info = {}

    for report_id, info in REPORT_STRUCTURE.items():
        level = info["level"]
        if level not in reports_by_level:
            reports_by_level[level] = {}

        set_num = info["set"]
        if set_num not in reports_by_level[level]:
            reports_by_level[level][set_num] = []
            sets_info[set_num] = {"estimated_minutes": 0}

        # Filter out specific reports from Level 2 Set 7 UI (handled by C1–C6 flow)
        if level == 2 and set_num == 7 and report_id in ("NBD-MF-20-C1", "NBD-MF-01-SOFP-SOCI"):
            continue

        # Filter out old GA & IS reports from Set 5 UI (replaced by NBD-MF-10-GA-11-IS-SET8)
        if level == 2 and set_num == 5 and report_id in ("NBD-MF-10-GA", "NBD-MF-11-IS"):
            continue

        # Filter out SOFP-SOCI from Set 6 UI (not required to be listed)
        if level == 2 and set_num == 6 and report_id == "NBD-MF-01-SOFP-SOCI":
            continue

        report_entry = {
            "id": report_id,
            "name": info["name"],
            "step": info["step"],
            "estimated_minutes": info.get("estimated_minutes", 10)
        }
        reports_by_level[level][set_num].append(report_entry)
        sets_info[set_num]["estimated_minutes"] += report_entry["estimated_minutes"]

    # Sort reports within each set by step
    for level in reports_by_level:
        for set_num in reports_by_level[level]:
            reports_by_level[level][set_num].sort(key=lambda x: x["step"])


    return render_template("index_new.html",
                         today_iso=today_iso,
                         reports_by_level=reports_by_level,
                         sets_info=sets_info,
                         report_structure=REPORT_STRUCTURE)


@app.route("/run-reports", methods=["POST"])
def run_reports():
    """Run selected reports with individual dates per set."""
    ensure_directories_exist()

    form_data = request.form
    report_ids_with_dates = {}

    for key in form_data:
        if key.startswith("date_"):
            set_num = key.replace("date_", "")
            date_str = form_data[key]

            set_reports = form_data.getlist(f"reports_{set_num}")

            if set_reports and date_str:
                dt = datetime.strptime(date_str, "%Y-%m-%d")
                report_month = dt.strftime("%B")
                report_year = str(dt.year)

                for report_id in set_reports:
                    report_ids_with_dates[report_id] = (date_str, report_month, report_year)

    if not report_ids_with_dates:
        flash("Please select at least one report with a date.", "error")
        return redirect(url_for("index"))

    # Check if this set needs master data input
    c1c6_ids = {"NBD-MF-20-C1","NBD-MF-20-C2","NBD-MF-20-C3","NBD-MF-20-C4","NBD-MF-20-C5","NBD-MF-20-C6"}
    needs_master_data = any(report_id in ("NBD-MF-01-SOFP-SOCI", "NBD-MF-23-IA", "NBD-MF-23-C1-C2", "NBD-MF-10-GA-11-IS-SET8", "NBD-MF-23-C4", "NBD-WF-15-LA") or report_id in c1c6_ids for report_id in report_ids_with_dates)
    _emit_log(f"*** MASTER DATA CHECK *** needs_master_data: {needs_master_data}, report_ids: {list(report_ids_with_dates.keys())}", important=True)
    
    if needs_master_data:
        _emit_log(f"*** REDIRECTING TO MASTER DATA FORM ***", important=True)
        # Store report data in session and redirect to master data form
        from flask import session
        session['report_ids_with_dates'] = report_ids_with_dates
        return redirect(url_for("master_data_form"))
    
    # Proceed directly if no master data needed
    return start_automation_with_data(report_ids_with_dates, {}, {})


@app.route("/master-data-form", methods=["GET"])
def master_data_form():
    """Display master data input form with latest data pre-filled."""
    _emit_log(f"*** MASTER DATA FORM ROUTE REACHED ***", important=True)
    from flask import session
    sofp_data, ia_data, c1c6_data = get_latest_master_data()
    uploaded_files = get_latest_uploaded_files()
    report_ids_with_dates = session.get('report_ids_with_dates', {})
    selected_ids = set(report_ids_with_dates.keys())
    c1c6_ids = {"NBD-MF-20-C1","NBD-MF-20-C2","NBD-MF-20-C3","NBD-MF-20-C4","NBD-MF-20-C5","NBD-MF-20-C6"}
    show_sofp = any(x in ("NBD-MF-01-SOFP-SOCI",) for x in selected_ids)
    show_ia = any(x in ("NBD-MF-23-IA",) for x in selected_ids)
    show_c1c2 = any(x in ("NBD-MF-23-C1-C2",) for x in selected_ids)
    show_c1c6 = any(x in c1c6_ids for x in selected_ids)
    show_ga_is = any(x in ("NBD-MF-10-GA-11-IS-SET8",) for x in selected_ids)
    show_c4 = any(x in ("NBD-MF-23-C4",) for x in selected_ids)
    show_la = any(x in ("NBD-WF-15-LA",) for x in selected_ids)

    # Get form data from request parameters (for validation error preservation)
    form_data = request.args.to_dict()

    return render_template("master_data_form.html",
                         sofp_data=sofp_data,
                         ia_data=ia_data,
                         c1c6_data=c1c6_data,
                         show_sofp=show_sofp,
                         show_ia=show_ia,
                         show_c1c2=show_c1c2,
                         show_c1c6=show_c1c6,
                         show_ga_is=show_ga_is,
                         show_c4=show_c4,
                         show_la=show_la,
                         uploaded_files=uploaded_files,
                         form_data=form_data)


@app.route("/submit-master-data", methods=["POST"])
def submit_master_data():
    """Process master data form submission."""
    _emit_log(f"*** SUBMIT MASTER DATA ROUTE REACHED ***", important=True)
    from flask import session
    
    # Clear processed files set for this submission
    global _processed_files
    _processed_files.clear()
    _emit_log(f"*** CLEARED PROCESSED FILES SET ***", important=True)

    # Clear download files from any previous run to prevent unwanted downloads on validation failure
    global _download_files, _sofp_download_file
    _download_files = []
    _sofp_download_file = None

    # Get master data from form (exclude timestamp as it's auto-generated)
    sofp_data = None
    ia_data = None
    c1c6_data = None

    # Determine which sections were shown
    show_sofp = request.form.get("_show_sofp") == "1"
    show_ia = request.form.get("_show_ia") == "1"
    show_c1c2 = request.form.get("_show_c1c2") == "1"
    show_c1c6 = request.form.get("_show_c1c6") == "1"
    show_ga_is = request.form.get("_show_ga_is") == "1"
    show_c4 = request.form.get("_show_c4") == "1"

    if show_sofp:
        sofp_data = {
            "bank_loans": request.form.get("bank_loans", ""),
            "securitization_loans": request.form.get("securitization_loans", ""),
            "foreign_funding": request.form.get("foreign_funding", "")
        }
    if show_ia:
        ia_data = {
            "annual_interest_rate": request.form.get("annual_interest_rate", "")
        }
    if show_c1c6:
        c1c6_data = {
            "guarantees": request.form.get("c1c6_guarantees", "")
        }
    
    # Handle file uploads for SOFP
    _emit_log("*** SOFP FILE UPLOAD SECTION REACHED ***", important=True)
    uploaded_files = {}
    months = {}
    if show_sofp:
        categories = [
            "ALCL Management Accounts",
            "Investment Schedule", 
            "Loan Schedule",
            "Supporting Schedules"
        ]
        
        for category in categories:
            field_name = category.lower().replace(' ', '_')
            month_field = f"{field_name}_month"

            # Get month selection
            month = request.form.get(month_field, "").strip()

            # Handle file upload
            if f"{field_name}_file" in request.files:
                file = request.files[f"{field_name}_file"]
                if file and file.filename:
                    # Month is REQUIRED when file is uploaded
                    if not month:
                        flash(f"Month is required when uploading {category} file.", "error")
                        return redirect(url_for("master_data_form", **request.form.to_dict()))

                    # Save month for later use
                    months[month_field] = month

                    # Save file to Master Inputs directory
                    category_dir = MASTER_INPUTS_DIR / category
                    category_dir.mkdir(parents=True, exist_ok=True)

                    # Generate filename with month and year (exact format as specified)
                    current_year = datetime.now().year
                    if category == "ALCL Management Accounts":
                        filename = f"ALCL Management Accounts {month}_{current_year}.xlsx"
                    elif category == "Investment Schedule":
                        filename = f"Investment Schedule - {month} {current_year}.xlsx"
                    elif category == "Loan Schedule":
                        filename = f"Loan Schedule {month} {current_year}.xlsx"
                    elif category == "Supporting Schedules":
                        filename = f"Supporting Schedules - {month} {current_year}.xlsx"
                    else:
                        filename = file.filename

                    file_path = category_dir / filename
                    file.save(str(file_path))
                    uploaded_files[category] = str(file_path)
                    _emit_log(f"Uploaded {category} file: {filename}", important=True)

                    # Use centralized master data update system
                    _emit_log(f"*** CALLING CENTRALIZED SYSTEM FOR {category} FROM SOFP ***", important=True)
                    update_master_data_across_reports(category, file_path, month=month, year=current_year)
            elif month:
                # Month provided but no file - still save the month
                months[month_field] = month
    
    # Handle file uploads for GA IS
    ga_is_uploaded_files = {}
    ga_is_dates = {}
    if show_ga_is:
        # Handle Borrowing report upload
        if "borrowing_report_file" in request.files:
            file = request.files["borrowing_report_file"]
            if file and file.filename:
                # Get date selection
                borrowing_date = request.form.get("borrowing_date", "").strip()
                if borrowing_date:
                    ga_is_dates["borrowing_date"] = borrowing_date
                    
                    # Save file to Master Inputs directory
                    category_dir = MASTER_INPUTS_DIR / "Borrowing report"
                    category_dir.mkdir(parents=True, exist_ok=True)
                    
                    # Generate filename with date (format: Borrowing report 31-08-2025.xlsm)
                    # Convert date from YYYY-MM-DD to DD-MM-YYYY format
                    date_obj = datetime.strptime(borrowing_date, "%Y-%m-%d")
                    formatted_date = date_obj.strftime("%d-%m-%Y")
                    filename = f"Borrowing report {formatted_date}.xlsm"
                    
                    file_path = category_dir / filename
                    file.save(str(file_path))
                    ga_is_uploaded_files["Borrowing report"] = str(file_path)
                    _emit_log(f"Uploaded Borrowing report file: {filename}", important=True)

                    # Use centralized master data update system
                    update_master_data_across_reports("Borrowing report", file_path, date=borrowing_date)

    # Handle file uploads for C4 (Rec Target)
    c4_uploaded_files = {}
    c4_months = {}
    c4_years = {}
    if show_c4:
        # Handle Rec Target upload
        if "rectarget_file" in request.files:
            file = request.files["rectarget_file"]
            if file and file.filename:
                # Get month and year selection
                month = request.form.get("rectarget_month", "").strip()
                year = request.form.get("rectarget_year", "").strip()

                # Month and Year are REQUIRED when file is uploaded
                if not month or not year:
                    flash(f"Month and Year are required when uploading Rec Target file.", "error")
                    return redirect(url_for("master_data_form", **request.form.to_dict()))

                # Validate year - must be exactly 4 digits, no symbols or characters
                if not year.isdigit() or len(year) != 4:
                    flash(f"Invalid year for Rec Target: {year}. Please enter a valid 4-digit year (e.g., 2025).", "error")
                    return redirect(url_for("master_data_form", **request.form.to_dict()))
                year_int = int(year)
                if year_int < 2000 or year_int > 2100:
                    flash(f"Invalid year for Rec Target: {year}. Please enter a year between 2000 and 2100.", "error")
                    return redirect(url_for("master_data_form", **request.form.to_dict()))

                # Save month/year for later use
                c4_months["rectarget_month"] = month
                c4_years["rectarget_year"] = year

                # Save file to Master Inputs directory
                category_dir = MASTER_INPUTS_DIR / "Rec Target"
                category_dir.mkdir(parents=True, exist_ok=True)

                # Generate filename: Rec Target - July 2025.xlsx
                filename = f"Rec Target - {month} {year}.xlsx"

                file_path = category_dir / filename
                file.save(str(file_path))
                c4_uploaded_files["Rec Target"] = str(file_path)
                _emit_log(f"Uploaded Rec Target file: {filename}", important=True)

                # Use centralized master data update system
                update_master_data_across_reports("Rec Target", file_path, month=month, year=year)

                # Copy to latest C4 outputs folder
                try:
                    c4_outputs_dir = OUTPUTS_DIR / "NBD_MF_23_C4"
                    _emit_log(f"C4 outputs directory: {c4_outputs_dir}", important=True)

                    if c4_outputs_dir.exists():
                        _emit_log(f"C4 outputs directory exists, finding latest completed date...", important=True)
                        # Find latest completed date in C4
                        _, c4_date_folder = find_latest_completed_date(c4_outputs_dir)

                        if c4_date_folder and c4_date_folder.exists():
                            _emit_log(f"Found C4 latest date folder: {c4_date_folder}", important=True)

                            # Remove old Rec Target files (files starting with "Rec Target" and ending with ".xlsx")
                            removed_count = 0
                            for f in c4_date_folder.iterdir():
                                if f.is_file() and f.suffix.lower() == ".xlsx":
                                    if f.name.startswith("Rec Target"):
                                        try:
                                            _emit_log(f"Removing old Rec Target file: {f.name}", important=True)
                                            f.unlink()
                                            removed_count += 1
                                        except Exception as e:
                                            _emit_log(f"Failed to remove old Rec Target file {f.name}: {e}", important=True)

                            if removed_count > 0:
                                _emit_log(f"Removed {removed_count} old Rec Target file(s) from C4 outputs", important=True)
                            else:
                                _emit_log(f"No old Rec Target files found to remove", important=True)

                            # Copy the new file
                            dest_path = c4_date_folder / filename
                            shutil.copy2(file_path, dest_path)
                            _emit_log(f"Copied Rec Target file to C4 outputs: {dest_path}", important=True)
                        else:
                            _emit_log("No completed date folder found for C4", important=True)
                    else:
                        _emit_log(f"C4 outputs directory does not exist: {c4_outputs_dir}", important=True)
                except Exception as e:
                    _emit_log(f"Failed to copy Rec Target file to C4 outputs: {e}", important=True)
                    import traceback
                    _emit_log(f"Traceback: {traceback.format_exc()}", important=True)

    # Handle file uploads for LA
    la_uploaded_files = {}
    la_months = {}
    la_years = {}
    la_dates = {}
    show_la = request.form.get("_show_la") == "1"
    if show_la:
        # Handle Daily Bank Balances upload
        if "daily_bank_balances_file" in request.files:
            file = request.files["daily_bank_balances_file"]
            if file and file.filename:
                # Get month and year
                month = request.form.get("daily_bank_balances_month", "").strip()
                year = request.form.get("daily_bank_balances_year", "").strip()

                # Month and Year are REQUIRED when file is uploaded
                if not month or not year:
                    flash(f"Month and Year are required when uploading Daily Bank Balances file.", "error")
                    return redirect(url_for("master_data_form", **request.form.to_dict()))

                # Validate year
                if not year.isdigit() or len(year) != 4:
                    flash(f"Invalid year for Daily Bank Balances: {year}. Please enter a valid 4-digit year (e.g., 2025).", "error")
                    return redirect(url_for("master_data_form", **request.form.to_dict()))
                year_int = int(year)
                if year_int < 2000 or year_int > 2100:
                    flash(f"Invalid year for Daily Bank Balances: {year}. Please enter a year between 2000 and 2100.", "error")
                    return redirect(url_for("master_data_form", **request.form.to_dict()))

                la_months["daily_bank_balances_month"] = month
                la_years["daily_bank_balances_year"] = year

                # Save file to Master Inputs
                category_dir = MASTER_INPUTS_DIR / "Daily Bank Balances"
                category_dir.mkdir(parents=True, exist_ok=True)

                # Generate filename: Daily Bank Balances - Jul 2025.xlsx
                filename = f"Daily Bank Balances - {month} {year}.xlsx"

                file_path = category_dir / filename
                file.save(str(file_path))
                la_uploaded_files["Daily Bank Balances"] = str(file_path)
                _emit_log(f"Uploaded Daily Bank Balances file: {filename}", important=True)

                # Use centralized master data update system
                update_master_data_across_reports("Daily Bank Balances", file_path, month=month, year=year)

        # Handle M2M upload
        if "m2m_file" in request.files:
            file = request.files["m2m_file"]
            if file and file.filename:
                # Get date
                m2m_date = request.form.get("m2m_date", "").strip()

                # Date is REQUIRED when file is uploaded
                if not m2m_date:
                    flash(f"Date is required when uploading M2M file.", "error")
                    return redirect(url_for("master_data_form", **request.form.to_dict()))

                la_dates["m2m_date"] = m2m_date

                # Save file to Master Inputs
                category_dir = MASTER_INPUTS_DIR / "M2M"
                category_dir.mkdir(parents=True, exist_ok=True)

                # Generate filename: M2M 0727.xlsx (MMDD format)
                date_obj = datetime.strptime(m2m_date, "%Y-%m-%d")
                mmdd = date_obj.strftime("%m%d")
                filename = f"M2M {mmdd}.xlsx"

                file_path = category_dir / filename
                file.save(str(file_path))
                la_uploaded_files["M2M"] = str(file_path)
                _emit_log(f"Uploaded M2M file: {filename}", important=True)

                # Use centralized master data update system
                update_master_data_across_reports("M2M", file_path, date=m2m_date)

        # Handle FD Base as at upload
        if "fdbastat_file" in request.files:
            file = request.files["fdbastat_file"]
            if file and file.filename:
                # Get date
                fdbastat_date = request.form.get("fdbastat_date", "").strip()

                # Date is REQUIRED when file is uploaded
                if not fdbastat_date:
                    flash(f"Date is required when uploading FD Base as at file.", "error")
                    return redirect(url_for("master_data_form", **request.form.to_dict()))

                la_dates["fdbastat_date"] = fdbastat_date

                # Save file to Master Inputs
                category_dir = MASTER_INPUTS_DIR / "FD Base as at"
                category_dir.mkdir(parents=True, exist_ok=True)

                # Generate filename: FD Base as at 31.07.2025.xlsx (DD.MM.YYYY format)
                date_obj = datetime.strptime(fdbastat_date, "%Y-%m-%d")
                formatted_date = date_obj.strftime("%d.%m.%Y")
                filename = f"FD Base as at {formatted_date}.xlsx"

                file_path = category_dir / filename
                file.save(str(file_path))
                la_uploaded_files["FD Base as at"] = str(file_path)
                _emit_log(f"Uploaded FD Base as at file: {filename}", important=True)

                # Use centralized master data update system
                update_master_data_across_reports("FD Base as at", file_path, date=fdbastat_date)

        # Handle Loan Schedule upload
        if "loanW_schedule_file" in request.files:
            file = request.files["loanW_schedule_file"]
            if file and file.filename:
                # Get month and year
                month = request.form.get("loanW_schedule_month", "").strip()
                year = request.form.get("loanW_year", "").strip()

                # Month and Year are REQUIRED when file is uploaded
                if not month or not year:
                    flash(f"Month and Year are required when uploading Loan Schedule file.", "error")
                    return redirect(url_for("master_data_form", **request.form.to_dict()))

                # Validate year
                if not year.isdigit() or len(year) != 4:
                    flash(f"Invalid year for Loan Schedule: {year}. Please enter a valid 4-digit year (e.g., 2025).", "error")
                    return redirect(url_for("master_data_form", **request.form.to_dict()))
                year_int = int(year)
                if year_int < 2000 or year_int > 2100:
                    flash(f"Invalid year for Loan Schedule: {year}. Please enter a year between 2000 and 2100.", "error")
                    return redirect(url_for("master_data_form", **request.form.to_dict()))

                la_months["loanW_schedule_month"] = month
                la_years["loanW_year"] = year

                # Save file to Master Inputs
                category_dir = MASTER_INPUTS_DIR / "Loan Schedule"
                category_dir.mkdir(parents=True, exist_ok=True)

                # Generate filename: Loan Schedule - July 2025.xlsx
                filename = f"Loan Schedule - {month} {year}.xlsx"

                file_path = category_dir / filename
                file.save(str(file_path))
                la_uploaded_files["Loan Schedule"] = str(file_path)
                _emit_log(f"Uploaded Loan Schedule file: {filename}", important=True)

                # Use centralized master data update system
                update_master_data_across_reports("Loan Schedule", file_path, month=month, year=year)

        # Handle Investment Schedule upload
        if "investmentW_file" in request.files:
            file = request.files["investmentW_file"]
            if file and file.filename:
                # Get month and year
                month = request.form.get("investmentW_month", "").strip()
                year = request.form.get("investmentW_year", "").strip()

                # Month and Year are REQUIRED when file is uploaded
                if not month or not year:
                    flash(f"Month and Year are required when uploading Investment Schedule file.", "error")
                    return redirect(url_for("master_data_form", **request.form.to_dict()))

                # Validate year
                if not year.isdigit() or len(year) != 4:
                    flash(f"Invalid year for Investment Schedule: {year}. Please enter a valid 4-digit year (e.g., 2025).", "error")
                    return redirect(url_for("master_data_form", **request.form.to_dict()))
                year_int = int(year)
                if year_int < 2000 or year_int > 2100:
                    flash(f"Invalid year for Investment Schedule: {year}. Please enter a year between 2000 and 2100.", "error")
                    return redirect(url_for("master_data_form", **request.form.to_dict()))

                la_months["investmentW_month"] = month
                la_years["investmentW_year"] = year

                # Save file to Master Inputs
                category_dir = MASTER_INPUTS_DIR / "Investment Schedule"
                category_dir.mkdir(parents=True, exist_ok=True)

                # Generate filename: Investment Schedule - July 2025.xlsx
                filename = f"Investment Schedule - {month} {year}.xlsx"

                file_path = category_dir / filename
                file.save(str(file_path))
                la_uploaded_files["Investment Schedule"] = str(file_path)
                _emit_log(f"Uploaded Investment Schedule file: {filename}", important=True)

                # Use centralized master data update system
                _emit_log(f"*** CALLING CENTRALIZED SYSTEM FOR INVESTMENT SCHEDULE FROM LA ***", important=True)
                update_master_data_across_reports("Investment Schedule", file_path, month=month, year=year)

    # Handle file uploads for IA
    _emit_log("*** IA FILE UPLOAD SECTION REACHED ***", important=True)
    ia_uploaded_files = {}
    ia_months = {}
    ia_years = {}
    if show_ia:
        categories = [
            "Disbursement with Budget",
            "Information Request from Credit",
            "Net Portfolio",
            "YARD STOCK AS AT"
        ]
        
        for category in categories:
            field_name = category.lower().replace(' ', '_')
            
            # Special handling for YARD STOCK AS AT (uses date instead of month/year)
            if category == "YARD STOCK AS AT":
                date_field = "yard_stock_date"  # Use the actual form field name
                date = request.form.get(date_field, "").strip()
                _emit_log(f"YARD STOCK AS AT - Date field: {date_field}, Date value: {date}", important=True)
                
                # Handle file upload
                if "yard_stock_file" in request.files:  # Use the actual form field name
                    file = request.files["yard_stock_file"]
                    _emit_log(f"YARD STOCK AS AT - File found: {file.filename if file else 'None'}", important=True)
                    if file and file.filename:
                        # Date is REQUIRED when file is uploaded
                        if not date:
                            flash(f"Date is required when uploading {category} file.", "error")
                            return redirect(url_for("master_data_form", **request.form.to_dict()))

                        # Save file to Master Inputs directory
                        category_dir = MASTER_INPUTS_DIR / category
                        category_dir.mkdir(parents=True, exist_ok=True)

                        # Generate filename with date (format: YARD STOCK AS AT 31-07-2025 FINAL.xlsx)
                        date_obj = datetime.strptime(date, "%Y-%m-%d")
                        formatted_date = date_obj.strftime("%d-%m-%Y")
                        filename = f"YARD STOCK AS AT {formatted_date} FINAL.xlsx"

                        file_path = category_dir / filename
                        _emit_log(f"YARD STOCK AS AT - Saving file to: {file_path}", important=True)
                        file.save(str(file_path))
                        ia_uploaded_files[category] = str(file_path)
                        _emit_log(f"Uploaded {category} file: {filename}", important=True)

                        # Use centralized master data update system
                        update_master_data_across_reports(category, file_path, date=date)
            else:
                # Regular handling for other categories (month/year based)
                month_field = f"{field_name}_month"
                year_field = f"{field_name}_year"

                # Get month and year selection
                month = request.form.get(month_field, "").strip()
                year = request.form.get(year_field, "").strip()

                # Handle file upload
                if f"{field_name}_file" in request.files:
                    file = request.files[f"{field_name}_file"]
                    if file and file.filename:
                        # Month and Year are REQUIRED when file is uploaded
                        if not month or not year:
                            flash(f"Month and Year are required when uploading {category} file.", "error")
                            return redirect(url_for("master_data_form", **request.form.to_dict()))

                        # Validate year - must be exactly 4 digits, no symbols or characters
                        if not year.isdigit() or len(year) != 4:
                            flash(f"Invalid year for {category}: {year}. Please enter a valid 4-digit year (e.g., 2025).", "error")
                            return redirect(url_for("master_data_form", **request.form.to_dict()))
                        year_int = int(year)
                        if year_int < 2000 or year_int > 2100:
                            flash(f"Invalid year for {category}: {year}. Please enter a year between 2000 and 2100.", "error")
                            return redirect(url_for("master_data_form", **request.form.to_dict()))

                        # Save month/year for later use
                        ia_months[month_field] = month
                        ia_years[year_field] = year

                        # Save file to Master Inputs directory
                        category_dir = MASTER_INPUTS_DIR / category
                        category_dir.mkdir(parents=True, exist_ok=True)

                        # Generate filename with month and year (exact format as specified)
                        if category == "Disbursement with Budget":
                            full_month = get_full_month_name(month)
                            filename = f"Disbursement with Budget - {full_month} {year}.xlsx"
                        elif category == "Information Request from Credit":
                            full_month = get_full_month_name(month)
                            filename = f"Information Request from Credit-{full_month} {year}.xlsx"
                        elif category == "Net Portfolio":
                            full_month = get_full_month_name(month)
                            filename = f"Net Portfolio-{full_month}-{year}.xlsx"
                        else:
                            filename = file.filename

                        file_path = category_dir / filename
                        file.save(str(file_path))
                        ia_uploaded_files[category] = str(file_path)
                        _emit_log(f"Uploaded {category} file: {filename}", important=True)

                        # Use centralized master data update system
                        update_master_data_across_reports(category, file_path, month=month, year=year)
                    elif month and year:
                        # Month/year provided but no file - validate and save
                        # Validate year - must be exactly 4 digits, no symbols or characters
                        if not year.isdigit() or len(year) != 4:
                            flash(f"Invalid year for {category}: {year}. Please enter a valid 4-digit year (e.g., 2025).", "error")
                            return redirect(url_for("master_data_form", **request.form.to_dict()))
                        year_int = int(year)
                        if year_int < 2000 or year_int > 2100:
                            flash(f"Invalid year for {category}: {year}. Please enter a year between 2000 and 2100.", "error")
                            return redirect(url_for("master_data_form", **request.form.to_dict()))

                        ia_months[month_field] = month
                        ia_years[year_field] = year
    
    # Handle file uploads for C1-C2
    c1c2_uploaded_files = {}
    c1c2_dates = {}
    c1c2_months = {}
    c1c2_years = {}
    if show_c1c2:
        categories = [
            ("Cadre", "date"),
            ("Unutilized Amount", "month_year")
        ]
        
        for category, type_info in categories:
            field_name = category.lower().replace(' ', '_')
            
            if type_info == "date":
                # Handle Cadre with date picker
                date_field = f"{field_name}_date"
                date_value = request.form.get(date_field, "").strip()

                # Handle file upload
                if f"{field_name}_file" in request.files:
                    file = request.files[f"{field_name}_file"]
                    if file and file.filename:
                        # Date is REQUIRED when file is uploaded
                        if not date_value:
                            flash(f"Date is required when uploading {category} file.", "error")
                            return redirect(url_for("master_data_form", **request.form.to_dict()))

                        # Save date for later use
                        c1c2_dates[date_field] = date_value

                        # Save file to Master Inputs directory
                        category_dir = MASTER_INPUTS_DIR / category
                        category_dir.mkdir(parents=True, exist_ok=True)

                        # Generate filename with date (format: Cadre 31 07 2025.xlsx)
                        # Convert date from YYYY-MM-DD to DD MM YYYY format
                        date_obj = datetime.strptime(date_value, "%Y-%m-%d")
                        formatted_date = date_obj.strftime("%d %m %Y")
                        filename = f"Cadre {formatted_date}.xlsx"

                        file_path = category_dir / filename
                        file.save(str(file_path))
                        c1c2_uploaded_files[category] = str(file_path)
                        _emit_log(f"Uploaded {category} file: {filename}", important=True)

                        # Use centralized master data update system
                        update_master_data_across_reports(category, file_path, date=date_value)
                elif date_value:
                    # Date provided but no file - still save the date
                    c1c2_dates[date_field] = date_value
            
            elif type_info == "month_year":
                # Handle Unutilized Amount with month/year
                month_field = f"{field_name}_month"
                year_field = f"{field_name}_year"

                month = request.form.get(month_field, "").strip()
                year = request.form.get(year_field, "").strip()

                # Handle file upload
                if f"{field_name}_file" in request.files:
                    file = request.files[f"{field_name}_file"]
                    if file and file.filename:
                        # Month and Year are REQUIRED when file is uploaded
                        if not month or not year:
                            flash(f"Month and Year are required when uploading {category} file.", "error")
                            return redirect(url_for("master_data_form", **request.form.to_dict()))

                        # Validate year - must be exactly 4 digits, no symbols or characters
                        if not year.isdigit() or len(year) != 4:
                            flash(f"Invalid year for {category}: {year}. Please enter a valid 4-digit year (e.g., 2025).", "error")
                            return redirect(url_for("master_data_form", **request.form.to_dict()))
                        year_int = int(year)
                        if year_int < 2000 or year_int > 2100:
                            flash(f"Invalid year for {category}: {year}. Please enter a year between 2000 and 2100.", "error")
                            return redirect(url_for("master_data_form", **request.form.to_dict()))

                        # Save month/year for later use
                        c1c2_months[month_field] = month
                        c1c2_years[year_field] = year

                        # Save file to Master Inputs directory
                        category_dir = MASTER_INPUTS_DIR / category
                        category_dir.mkdir(parents=True, exist_ok=True)

                        # Generate filename with month and year (format: Unutilized Amount JULY 2025.xlsx)
                        full_month = get_full_month_name(month)
                        filename = f"Unutilized Amount {full_month.upper()} {year}.xlsx"

                        file_path = category_dir / filename
                        file.save(str(file_path))
                        c1c2_uploaded_files[category] = str(file_path)
                        _emit_log(f"Uploaded {category} file: {filename}", important=True)

                        # Use centralized master data update system
                        update_master_data_across_reports(category, file_path, month=month, year=year)
                elif month and year:
                    # Month/year provided but no file - validate and save
                    # Validate year - must be exactly 4 digits, no symbols or characters
                    if not year.isdigit() or len(year) != 4:
                        flash(f"Invalid year for {category}: {year}. Please enter a valid 4-digit year (e.g., 2025).", "error")
                        return redirect(url_for("master_data_form", **request.form.to_dict()))
                    year_int = int(year)
                    if year_int < 2000 or year_int > 2100:
                        flash(f"Invalid year for {category}: {year}. Please enter a year between 2000 and 2100.", "error")
                        return redirect(url_for("master_data_form", **request.form.to_dict()))

                    c1c2_months[month_field] = month
                    c1c2_years[year_field] = year

    # Save to Master_Data.xlsx
    success, msg = save_master_data(sofp_data, ia_data, c1c6_data)
    if not success:
        flash(f"Failed to save master data: {msg}", "error")
        return redirect(url_for("master_data_form"))
    
    # Get report data from session
    report_ids_with_dates = session.get('report_ids_with_dates', {})
    if not report_ids_with_dates:
        flash("No report data found. Please start again.", "error")
        return redirect(url_for("index"))
    
    # Clear session data
    session.pop('report_ids_with_dates', None)
    
    # Note: File copying is now handled by the centralized update_master_data_across_reports() function
    # which is called during the individual file upload processing above
    
    # Validate SOFP file exists for IA automation (IA and C1C2 run together)
    if show_ia or show_c1c2:
        try:
            _emit_log("Validating SOFP file exists for IA automation (IA and C1C2 run together)", important=True)
            is_valid, validation_msg = validate_sofp_file_for_ia(report_ids_with_dates)
            if not is_valid:
                flash(f"Validation failed: {validation_msg}", "error")
                return redirect(url_for("index"))
        except Exception as e:
            _emit_log(f"Failed to validate SOFP file for IA: {e}", important=True)
            flash(f"Validation error: {e}", "error")
            return redirect(url_for("index"))
    
    # Validate SOFP file exists for GA IS automation
    if show_ga_is:
        try:
            _emit_log("Validating SOFP file exists for GA IS automation", important=True)
            is_valid, validation_msg = validate_sofp_file_for_ga_is(report_ids_with_dates)
            if not is_valid:
                flash(f"Validation failed: {validation_msg}", "error")
                return redirect(url_for("index"))
        except Exception as e:
            _emit_log(f"Failed to validate SOFP file for GA IS: {e}", important=True)
            flash(f"Validation error: {e}", "error")
            return redirect(url_for("index"))

    # Validate Prod. wise file exists for C8 automation
    if "NBD-QF-23-C8" in report_ids_with_dates:
        try:
            _emit_log("Validating Prod. wise Class. of Loans file exists for C8 automation", important=True)
            is_valid, validation_msg = validate_prod_wise_file_for_c8(report_ids_with_dates)
            if not is_valid:
                flash(f"Validation failed: {validation_msg}", "error")
                return redirect(url_for("index"))
        except Exception as e:
            _emit_log(f"Failed to validate Prod. wise file for C8: {e}", important=True)
            flash(f"Validation error: {e}", "error")
            return redirect(url_for("index"))

    # Validate Set 7 reports (C2-C6) - must select all together and check required files
    set7_reports = {"NBD-MF-20-C2", "NBD-MF-20-C3", "NBD-MF-20-C4", "NBD-MF-20-C5", "NBD-MF-20-C6"}
    selected_set7_reports = set(report_ids_with_dates.keys()) & set7_reports
    
    if selected_set7_reports:
        # Check if all Set 7 reports are selected
        if len(selected_set7_reports) != len(set7_reports):
            missing_reports = set7_reports - selected_set7_reports
            flash(f"Set 7 reports must be selected together. Missing: {', '.join(missing_reports)}", "error")
            return redirect(url_for("index"))
        
        # Validate required files exist
        try:
            _emit_log("Validating required files for Set 7 automation", important=True)
            is_valid, validation_msg = validate_set7_required_files(report_ids_with_dates)
            if not is_valid:
                flash(f"Validation failed: {validation_msg}", "error")
                return redirect(url_for("index"))
        except Exception as e:
            _emit_log(f"Failed to validate Set 7 required files: {e}", important=True)
            flash(f"Validation error: {e}", "error")
            return redirect(url_for("index"))
    
    # Start automation
    return start_automation_with_data(report_ids_with_dates, sofp_data, ia_data, uploaded_files, months)


@app.route("/mark-completed", methods=["POST"])
def mark_completed():
    """Mark a report as completed by appending date to completed_dates.txt"""
    report_id = request.form.get("report_id", "").strip()
    date_str = request.form.get("date", "").strip()
    
    if not report_id or not date_str:
        flash("Please select a report and date.", "error")
        return redirect(url_for("index"))
    
    if report_id not in REPORT_STRUCTURE:
        flash("Invalid report selected.", "error")
        return redirect(url_for("index"))
    
    try:
        # Convert date from YYYY-MM-DD to DD-MM-YYYY format
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        formatted_date = dt.strftime("%d-%m-%Y")
        
        # Get the report folder name
        folder_name = REPORT_STRUCTURE[report_id]["folder"]
        completed_file = OUTPUTS_DIR / folder_name / "completed_dates.txt"
        
        # Ensure the directory exists
        completed_file.parent.mkdir(parents=True, exist_ok=True)
        
        # Append the date to completed_dates.txt
        with open(completed_file, "a", encoding="utf-8") as f:
            f.write(formatted_date + "\n")
        
        flash(f"Marked {REPORT_STRUCTURE[report_id]['name']} as completed for {formatted_date}.", "success")
        
    except ValueError:
        flash("Invalid date format. Please use the date picker.", "error")
    except Exception as e:
        flash(f"Failed to mark as completed: {e}", "error")
    
    return redirect(url_for("index"))


def start_automation_with_data(report_ids_with_dates: dict, sofp_data: dict, ia_data: dict, uploaded_files: dict = None, months: dict = None):
    """Start automation with master data."""
    # Initialize new run
    _initialize_new_run()
    global _report_checklist, _report_messages, _last_output_dirs, _stop_requested, _download_files, _sofp_download_file
    _report_checklist = []
    _report_messages = {}
    _last_output_dirs = {}
    _download_files = []
    _sofp_download_file = None
    _stop_requested = False
    
    # PROCESS ISOLATION: Wait for master data form to complete
    _emit_log("*** PROCESS ISOLATION: WAITING FOR MASTER DATA TO COMPLETE ***", important=True)
    import time
    time.sleep(3)  # Wait 3 seconds for master data form to complete
    
    # Check for any locked files and wait if necessary
    _emit_log("*** CHECKING FOR FILE LOCKS ***", important=True)
    max_wait_time = 30  # Maximum wait time in seconds
    wait_interval = 1   # Check every 1 second
    waited = 0
    
    while waited < max_wait_time:
        # Check if any output files are locked
        locked_files = []
        for report_id in report_ids_with_dates:
            if report_id in REPORT_STRUCTURE:
                folder_name = REPORT_STRUCTURE[report_id]["folder"]
                if report_id == "NBD-MF-10-GA-11-IS-SET8":
                    output_dir = OUTPUTS_DIR / "NBD-MF-10-GA & NBD-MF-11-IS"
                else:
                    output_dir = OUTPUTS_DIR / folder_name
                
                if output_dir.exists():
                    # Find latest date folder
                    last_date, date_folder = find_latest_completed_date(output_dir)
                    if date_folder and date_folder.exists():
                        # Check for locked files
                        for f in date_folder.iterdir():
                            if f.is_file() and f.suffix.lower() in ['.xlsx', '.xlsm']:
                                try:
                                    # Try to open file in exclusive mode to check if it's locked
                                    with open(f, 'r+b') as test_file:
                                        pass
                                except (PermissionError, OSError):
                                    locked_files.append(str(f))
        
        if locked_files:
            _emit_log(f"Found {len(locked_files)} locked files, waiting... ({waited}s/{max_wait_time}s)", important=True)
            time.sleep(wait_interval)
            waited += wait_interval
        else:
            _emit_log("No locked files found, proceeding with automation", important=True)
            break
    
    if waited >= max_wait_time:
        _emit_log("WARNING: Maximum wait time reached, proceeding anyway", important=True)
    
    # Clean up any existing duplicate files before starting automation
    _emit_log("Cleaning up duplicate files before starting automation", important=True)
    cleanup_duplicate_ia_files()
    cleanup_duplicate_c1c2_files()

    _set_status(running=True, stage="starting", message=f"Starting {len(report_ids_with_dates)} reports")

    # Run reports in background thread
    def run_in_background():
        _set_status(stage="running", message="Running selected reports")
        success, msg = run_selected_reports(report_ids_with_dates)
        _set_status(running=False, stage="done" if success else "error", message=msg)

    thread = threading.Thread(target=run_in_background, daemon=True)
    thread.start()

    flash(f"Started running {len(report_ids_with_dates)} reports", "success")
    return redirect(url_for("index"))


@app.route("/status", methods=["GET"])
def status():
    with _status_lock:
        return jsonify(_status)


@app.route("/logs", methods=["GET"])
def logs():
    with _log_lock:
        content = "\n".join(_log_buffer)
    return app.response_class(content, mimetype="text/plain")


@app.route("/status-feed", methods=["GET"])
def status_feed():
    content = "\n".join(_status_feed)
    return app.response_class(content, mimetype="text/plain")


@app.route("/report-messages/<report_id>", methods=["GET"])
def report_messages(report_id: str):
    buf = _report_messages.get(report_id)
    if not buf:
        return app.response_class("", mimetype="text/plain")
    return app.response_class("\n".join(buf), mimetype="text/plain")


@app.route("/report-checklist", methods=["GET"])
def report_checklist():
    # Update elapsed times for running reports
    for item in _report_checklist:
        if item["status"] == "running":
            _update_report_elapsed_time(item["name"])
    return jsonify(_report_checklist)


@app.route("/stop", methods=["POST"])
def stop():
    global _stop_requested
    _stop_requested = True
    _set_status(stage="stopping", message="Stopping requested")

    with _proc_lock:
        for report_id, proc in _running_processes.items():
            try:
                proc.terminate()
                _emit_log(f"Stopped {report_id}", important=True)
            except Exception:
                pass
        _running_processes.clear()

    return jsonify({"ok": True})


@app.route("/download-latest", methods=["GET"])
def download_latest():
    """Download the prepared zip file with completed report files."""
    global _download_files, _sofp_download_file

    _emit_log(f"Download requested. Current _download_files: {len(_download_files) if _download_files else 0} files", important=True)

    # If no files available (download already completed), reset status and redirect to main page
    if not _download_files:
        _emit_log("No files in _download_files, resetting status and redirecting to main page", important=True)
        _reset_status()  # Reset status to idle to stop frontend polling
        return redirect(url_for("index"))

    if not _download_files[0].exists():
        _emit_log("No valid files in _download_files, attempting to prepare files", important=True)
        # If no prepared zip exists, create one on-demand
        try:
            auto_download_files()
        except Exception as e:
            _emit_log(f"Failed to prepare files for download: {e}", important=True)
            flash("Failed to prepare files for download.", "error")
            return redirect(url_for("index"))

    if not _download_files or not _download_files[0].exists():
        _emit_log("Still no files available after auto_download_files(), redirecting to main page", important=True)
        flash("No files available for download.", "error")
        return redirect(url_for("index"))

    path = _download_files[0]

    # Extract the download filename for display
    if path.suffix.lower() in [".xlsx", ".xlsb"]:
        download_name = path.name
    else:
        timestamp = path.stem.replace("reports_", "")
        download_name = f"reports_{timestamp}.zip"

    # Clear download files to prevent re-download on page refresh
    _download_files = []
    _sofp_download_file = None
    _emit_log(f"Download completed: {download_name}. Cleared download session.", important=True)

    # Reset status to idle to stop frontend polling after download
    _reset_status()
    _emit_log("Status reset to idle after download initiated", important=True)

    # Store download completion info in session for display after redirect
    from flask import session
    session['download_completed'] = True
    session['download_filename'] = download_name

    # If the queued item is a single .xlsx, send as-is; otherwise send the zip
    try:
        if path.suffix.lower() in [".xlsx", ".xlsb"]:
            return send_file(str(path), as_attachment=True, download_name=download_name)
        else:
            return send_file(str(path), as_attachment=True, download_name=download_name)
    except Exception as e:
        _emit_log(f"Failed to send file: {e}", important=True)
        flash(f"Failed to send file: {e}", "error")
        _reset_status()  # Reset status on error as well
        return redirect(url_for("index"))


if __name__ == "__main__":
    # Kill any running Excel instances to prevent COM errors
    kill_excel_instances()

    ensure_directories_exist()

    # Production-safe defaults
    # host = os.environ.get("FLASK_HOST", "0.0.0.0")
    # port = int(os.environ.get("PORT", 5000))

    app.run(host="127.0.0.1", port=5000, debug=True)

    # print(f"Starting Waitress server on http://{host}:{port}")
    # serve(app, host=host, port=port)
