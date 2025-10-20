import os
from pathlib import Path
import openpyxl
import sys
import argparse
from datetime import datetime
import re
import logging

def get_month_year_from_filename(filename):
    # Example: "NBD-MF-20-C1 to C6 July 2025.xlsx"
    parts = filename.split()
    for i, part in enumerate(parts):
        if part in {"Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","January","February","March","April","May","June","July","August","September","October","November","December"}:
            month = part
            year = parts[i+1].replace(".xlsx","")
            return month, year
    return None, None


def find_files_safely(search_dir, pattern, description):
    """Safely find files and provide detailed error messages"""
    print(f"Looking for {description} in: {search_dir}")
    print(f"Search pattern: {pattern}")
    
    if not search_dir.exists():
        print(f"Directory doesn't exist: {search_dir}")
        return None
    
    files = list(search_dir.glob(pattern))
    print(f"Found {len(files)} files matching pattern:")
    
    if not files:
        print(f"No files found matching pattern: {pattern}")
        print("Available files in directory:")
        try:
            all_files = [f.name for f in search_dir.iterdir() if f.is_file() and f.suffix == '.xlsx']
            if all_files:
                for file in sorted(all_files):
                    print(f"  - {file}")
            else:
                print("  No .xlsx files found in directory")
        except Exception as e:
            print(f"  Error listing files: {e}")
        return None
    
    for file in files:
        print(f"  [OK] {file.name}")
    
    return files[0]  # Return the first match


def setup_logging():
    """Setup organized logging for C6 report"""
    # Configure logging
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    logger = logging.getLogger(__name__)
    
    # Create organized log directory: logs/YYYY-MM-DD/frequency/report_name.log
    try:
        frequency = "monthly"
        date_folder = datetime.now().strftime('%Y-%m-%d')
        run_timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        logs_dir = Path(__file__).parent.parent / "logs" / date_folder / frequency
        logs_dir.mkdir(parents=True, exist_ok=True)
        run_log_file = logs_dir / f"NBD_MF_20_C6_{run_timestamp}.log"
    except Exception:
        # Fallback to simple structure
        logs_dir = Path(__file__).parent / "logs"
        logs_dir.mkdir(parents=True, exist_ok=True)
        run_log_file = logs_dir / f"NBD_MF_20_C6_{run_timestamp}.log"
    
    file_handler = logging.FileHandler(run_log_file, encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    logger.addHandler(file_handler)
    
    logger.info(f"📄 Detailed log file: {run_log_file}")
    return logger

def main():
    # Parse command line arguments
    parser = argparse.ArgumentParser(description='NBD MF 20 C6 Report Generator')
    parser.add_argument('--working-dir', type=str, help='Working directory path')
    parser.add_argument('--month', type=str, help='Report month (e.g., Jul)')
    parser.add_argument('--year', type=str, help='Report year (e.g., 2025)')
    
    args = parser.parse_args()
    
    # Setup logging
    logger = setup_logging()
    
    # Determine working directory - new structure: working/NBD_MF_20_C1_C6/<dated-folder>/
    if args.working_dir:
        # Use provided working directory - it should point to the date folder
        latest_folder = Path(args.working_dir)
        c1_c6_working_dir = latest_folder.parent
        working_dir = c1_c6_working_dir.parent
    else:
        # Fallback to default behavior
        base_dir = Path(__file__).resolve().parent.parent
        working_dir = base_dir / "working"
        c1_c6_working_dir = working_dir / "NBD_MF_20_C1_C6"
        
        if not c1_c6_working_dir.exists():
            print(f"[ERROR] C1-C6 working directory does not exist: {c1_c6_working_dir}")
            print("Please ensure the working/NBD_MF_20_C1_C6 directory exists with dated folders.")
            sys.exit(1)
        
        # Find the latest dated folder
        dated_folders = [d for d in c1_c6_working_dir.iterdir() if d.is_dir()]
        if not dated_folders:
            print("[ERROR] No dated folders found in C1-C6 working directory")
            sys.exit(1)
        
        # Use the first dated folder (assuming only one exists as per new structure)
        latest_folder = dated_folders[0]
        print(f"[OK] Using dated folder: {latest_folder.name}")
    
    logger.info(f"Base directory: {Path(__file__).resolve().parent.parent}")
    logger.info(f"Working directory: {working_dir}")
    logger.info(f"C1-C6 working directory: {c1_c6_working_dir}")
    logger.info(f"Latest folder: {latest_folder}")
    print(f"Base directory: {Path(__file__).resolve().parent.parent}")
    print(f"Working directory: {working_dir}")
    print(f"C1-C6 working directory: {c1_c6_working_dir}")
    print(f"Latest folder: {latest_folder}")
    
    if not latest_folder.exists():
        print(f"[ERROR] Latest folder does not exist: {latest_folder}")
        sys.exit(1)
    
    print(f"\nSearching for files in dated folder: {latest_folder}")
    
    # Files are directly in the dated folder (no subfolders in new structure)
    search_folder = latest_folder
    print(f"Search folder: {search_folder}")
    
    # Both files should be in the same dated folder
    afl_subfolder = search_folder  # Use the same folder for both files
    
    # Search for files in the dated folder
    # First try the expected naming patterns
    file_c1_c6 = find_files_safely(search_folder, "NBD-MF-20-C1 to C6*.xlsx", "C1-C6 file")
    file_afl = find_files_safely(afl_subfolder, "NBD-MF-01-SOFP & SOCI AFL Monthly FS*.xlsx", "AFL file")
    
    # If not found, try alternative patterns
    if file_c1_c6 is None:
        print("Trying alternative patterns for C1-C6 file...")
        file_c1_c6 = find_files_safely(search_folder, "*C1*C6*.xlsx", "C1-C6 file (alternative)")
        if file_c1_c6 is None:
            file_c1_c6 = find_files_safely(search_folder, "*NBD*MF*20*.xlsx", "C1-C6 file (alternative 2)")
    
    if file_afl is None:
        print("Trying alternative patterns for AFL file...")
        file_afl = find_files_safely(afl_subfolder, "*AFL*.xlsx", "AFL file (alternative)")
        if file_afl is None:
            file_afl = find_files_safely(afl_subfolder, "*SOFP*.xlsx", "AFL file (alternative 2)")
        if file_afl is None:
            file_afl = find_files_safely(afl_subfolder, "*SOCI*.xlsx", "AFL file (alternative 3)")
    
    if file_c1_c6 is None:
        print(f"\n[ERROR] Could not find C1 to C6 file in {search_folder}")
        print("Available Excel files:")
        try:
            xlsx_files = list(search_folder.glob("*.xlsx"))
            for f in xlsx_files:
                print(f"  - {f.name}")
        except Exception as e:
            print(f"  Error listing files: {e}")
        sys.exit(1)
    
    if file_afl is None:
        print(f"\n[ERROR] Could not find AFL file in {afl_subfolder}")
        print("Available Excel files:")
        try:
            xlsx_files = list(afl_subfolder.glob("*.xlsx"))
            for f in xlsx_files:
                print(f"  - {f.name}")
        except Exception as e:
            print(f"  Error listing files: {e}")
        sys.exit(1)
    
    # Extract month and year from filename
    month, year = get_month_year_from_filename(file_c1_c6.name)
    if not month or not year:
        print(f"[ERROR] Could not parse month/year from filename: {file_c1_c6.name}")
        print("Expected format: 'NBD-MF-20-C1 to C6 [Month] [Year].xlsx'")
        sys.exit(1)
    
    print(f"\n[OK] Parsed month: {month}, year: {year}")
    
    # Save the report directly in the NBD_MF_20_C1_C6 subfolder where source files are located
    out_folder = search_folder
    print(f"Output folder: {out_folder}")
    
    try:
        # Load workbooks
        print(f"\nLoading workbooks...")
        wb_c1_c6 = openpyxl.load_workbook(file_c1_c6, data_only=True)
        wb_afl = openpyxl.load_workbook(file_afl, data_only=True)
        print(f"[OK] Workbooks loaded successfully")
        
        # Check if required sheets exist
        required_sheets = {
            'c1_c6': ["3 Year Summary", "NBD_NBL-MF-20-C6"],
            'afl': ["NBD-MF-02-SOCI"]
        }
        
        print(f"\nAvailable sheets in C1-C6 file: {wb_c1_c6.sheetnames}")
        print(f"Available sheets in AFL file: {wb_afl.sheetnames}")
        
        # Validate required sheets
        for sheet in required_sheets['c1_c6']:
            if sheet not in wb_c1_c6.sheetnames:
                print(f"[ERROR] Required sheet '{sheet}' not found in C1-C6 file")
                print(f"Available sheets: {wb_c1_c6.sheetnames}")
                return
        
        for sheet in required_sheets['afl']:
            if sheet not in wb_afl.sheetnames:
                print(f"[ERROR] Required sheet '{sheet}' not found in AFL file")
                print(f"Available sheets: {wb_afl.sheetnames}")
                return
        
        print(f"[OK] All required sheets found")
        
        # Get worksheets
        src_ws = wb_afl["NBD-MF-02-SOCI"]
        tgt_ws = wb_c1_c6["3 Year Summary"]
        
        print(f"\nProcessing data...")
        
        # Helper function to safely get numeric values
        def get_numeric_value(worksheet, cell_ref, description=""):
            """Safely extract numeric value from a cell, handling text/None values"""
            try:
                value = worksheet[cell_ref].value
                print(f"{description} ({cell_ref}): {value}")
                
                if value is None:
                    return 0
                elif isinstance(value, (int, float)):
                    return value
                elif isinstance(value, str):
                    # Try to extract number from string if possible
                    try:
                        # Remove common text and try to parse number
                        clean_value = value.replace(',', '').strip()
                        return float(clean_value)
                    except ValueError:
                        print(f"  [WARNING] Cell {cell_ref} contains text '{value}' - using 0")
                        return 0
                else:
                    print(f"  [WARNING] Cell {cell_ref} contains unexpected type {type(value)} - using 0")
                    return 0
            except Exception as e:
                print(f"  [ERROR] Error reading {cell_ref}: {e} - using 0")
                return 0
        
        # 1. Interest Income
        interest_income = get_numeric_value(src_ws, "B10", "Interest Income")
        tgt_ws["B5"] = interest_income
        
        # 2. Interest Expenses
        interest_expenses = get_numeric_value(src_ws, "B11", "Interest Expenses")
        tgt_ws["B6"] = interest_expenses
        
        # 3. Non-interest Income
        net_fee_commission = get_numeric_value(src_ws, "B12", "Net Fee Commission")
        net_other_income = get_numeric_value(src_ws, "B13", "Net Other Income")
        non_interest_income = net_fee_commission + net_other_income
        print(f"Non-interest Income: {net_fee_commission} + {net_other_income} = {non_interest_income}")
        tgt_ws["B7"] = non_interest_income
        
        # 4. Realized profit/losses from sale of securities
        realized_profit_loss = get_numeric_value(src_ws, "B14", "Realized profit/loss")
        tgt_ws["B8"] = realized_profit_loss
        
        # 5. Extraordinary /Irregular Item of Income /Expenses
        print(f"\nExtraordinary items:")
        real_estate_gains = get_numeric_value(src_ws, "B15", "Gains/(Losses) on sale of Real estates")
        investment_property_gains = get_numeric_value(src_ws, "B16", "Gains/(Losses) on sale of Investment Properties")
        ppe_gains = get_numeric_value(src_ws, "B17", "Gains/(Losses) on sale of property, plant & Equipment")
        revaluation_gains = get_numeric_value(src_ws, "B18", "Gains/(Losses) on revaluation of Investment properties")
        
        extra_items = real_estate_gains + investment_property_gains + ppe_gains + revaluation_gains
        print(f"Total extraordinary items: {real_estate_gains} + {investment_property_gains} + {ppe_gains} + {revaluation_gains} = {extra_items}")
        tgt_ws["B9"] = extra_items
        
        # Update column headers and item names in C6 sheet
        tgt_c6_ws = wb_c1_c6["NBD_NBL-MF-20-C6"]
        print(f"Updating column headers and item names in C6 sheet...")
        
        # Change 3rd column (C) header to "1st Year"
        tgt_c6_ws["C1"] = "1st Year"
        
        # Change 4th column (D) header to "2nd Year"
        tgt_c6_ws["D1"] = "2nd Year"
        
        # Update item names in column B
        item_names = {
            "B6": "Total Operating Income",
            "B7": "Interest Income from Loans & Advances", 
            "B8": "Interest Expenses on Deposits",
            "B9": "Fee & Commission Income",
            "B10": "Trading & Investment Income",
            "B11": "Other Operating Income"
        }
        
        for cell, new_name in item_names.items():
            tgt_c6_ws[cell] = new_name
            print(f"Updated {cell}: {new_name}")
        
        print(f"[OK] Column headers updated: C1='1st Year', D1='2nd Year'")
        print(f"[OK] Item names updated in column B")
        print(f"[OK] Data processing completed")
        
        # Save updated workbook with original filename (like Ctrl+S)
        wb_c1_c6.save(file_c1_c6)
        print(f"\n[OK] Report saved successfully to: {file_c1_c6}")
        
    except FileNotFoundError as e:
        print(f"[ERROR] File not found: {e}")
        sys.exit(1)
    except PermissionError as e:
        print(f"[ERROR] Permission error (file might be open in Excel): {e}")
        sys.exit(1)
    except Exception as e:
        print(f"[ERROR] An error occurred: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()