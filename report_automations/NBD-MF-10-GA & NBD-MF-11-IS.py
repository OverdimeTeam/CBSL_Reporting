import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Alignment
import os
import shutil
import logging
from datetime import datetime
import multiprocessing
import traceback
import sys
import time
import win32com.client as win32
import re
import warnings

# Fix Unicode encoding for Windows console
if sys.platform.startswith('win'):
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

warnings.filterwarnings('ignore')

def is_file_open(file_path):
    """Check if a file is currently open by another process."""
    try:
        with open(file_path, 'a'):
            return False
    except PermissionError:
        return True
    except Exception:
        return False

def wait_for_file_access(file_path, max_attempts=10, delay=3):
    """Wait for file to become accessible."""
    for attempt in range(max_attempts):
        if not is_file_open(file_path):
            return True
        print(f"File {file_path} is open, waiting... (attempt {attempt + 1}/{max_attempts})")
        time.sleep(delay)
    return False

def validate_required_files():
    """Validate that all required files exist in the current directory, ignoring date parts."""
    print("="*60)
    print("VALIDATING REQUIRED FILES")
    print("="*60)
    
    # Required file base names (ignoring date parts)
    required_files = [
        "NBD-MF-10-GA & NBD-MF-11-IS",
        "NBD-MF-01-SOFP & SOCI AFL Monthly FS", 
        "Investment Schedule",
        "FD Base as at",
        "Borrowing report",
        "MATCAP",
        "CBSL Provision Comparison"
    ]
    
    # Get all files in current directory
    current_files = os.listdir('.')
    
    found_files = {}
    missing_files = []
    
    print(f"Scanning directory for {len(required_files)} required files...")
    print(f"Found {len(current_files)} files in directory")
    print()
    
    for required_base in required_files:
        found = False
        matching_file = None
        duplicate_files = []
        
        for file in current_files:
            # Skip certain file types that shouldn't be considered duplicates
            skip_extensions = ['.py', '.log', '.txt', '.bak', '.tmp']
            skip_prefixes = ['~$']  # Excel temporary files
            
            if any(file.startswith(prefix) for prefix in skip_prefixes):
                continue
            if any(file.lower().endswith(ext) for ext in skip_extensions):
                continue
            
            # Remove file extension for comparison
            file_base = os.path.splitext(file)[0]
            
            # Check if the required base name is contained in the file name
            if required_base.lower() in file_base.lower():
                if not found:
                    found = True
                    matching_file = file
                else:
                    # Found a duplicate
                    duplicate_files.append(file)
        
        if found:
            if duplicate_files:
                # Multiple files found with same base name
                all_matching = [matching_file] + duplicate_files
                print(f"⚠️  {required_base:<35} -> MULTIPLE FILES FOUND:")
                print(f"    Using: {matching_file}")
                for dup_file in duplicate_files:
                    print(f"    Duplicate: {dup_file}")
                print(f"    ERROR: Please remove duplicate files or rename them!")
                missing_files.append(required_base)  # Treat as missing due to ambiguity
            else:
                found_files[required_base] = matching_file
                print(f"✅ {required_base:<35} -> {matching_file}")
        else:
            missing_files.append(required_base)
            print(f"❌ {required_base:<35} -> NOT FOUND")
    
    print()
    print("="*60)
    
    if missing_files:
        print(f"❌ VALIDATION FAILED: {len(missing_files)} files missing or have duplicates!")
        print("Issues found:")
        for missing in missing_files:
            print(f"   - {missing}")
        print()
        print("Please ensure:")
        print("1. All required files are present in the directory")
        print("2. No duplicate files with the same base name exist")
        print("3. File names can have different dates/months/years, but the base name must match")
        print("4. Remove or rename duplicate files to avoid conflicts")
        return False, found_files
    else:
        print(f"✅ VALIDATION SUCCESSFUL: All {len(required_files)} required files found!")
        print()
        return True, found_files

class PreciseExcelDataTransfer:
    """Precise Excel data transfer - only values to specific cells, preserve everything else."""
    
    def __init__(self):
        # File names - will be dynamically set based on actual files found
        self.main_file = None
        self.sofp_source = None
        self.investment_source = None
        self.fd_base_file = None
        self.borrowing_source = None
        
        # Initialize file names by finding actual files
        self._initialize_file_names()
        self.log_messages = []
        self.excel_app = None
        
        # Setup detailed logging to file
        self.setup_detailed_logging()
    
    def _initialize_file_names(self):
        """Initialize file names by finding actual files in the directory."""
        current_files = os.listdir('.')
        
        # File mapping: base name -> actual file found
        file_mappings = {
            "NBD-MF-10-GA & NBD-MF-11-IS": "main_file",
            "NBD-MF-01-SOFP & SOCI AFL Monthly FS": "sofp_source", 
            "Investment Schedule": "investment_source",
            "FD Base as at": "fd_base_file",
            "Borrowing report": "borrowing_source"
        }
        
        for base_name, attribute_name in file_mappings.items():
            found_file = None
            duplicate_files = []
            
            for file in current_files:
                # Skip certain file types that shouldn't be considered duplicates
                skip_extensions = ['.py', '.log', '.txt', '.bak', '.tmp']
                skip_prefixes = ['~$']  # Excel temporary files
                
                if any(file.startswith(prefix) for prefix in skip_prefixes):
                    continue
                if any(file.lower().endswith(ext) for ext in skip_extensions):
                    continue
                
                file_base = os.path.splitext(file)[0]
                if base_name.lower() in file_base.lower():
                    if not found_file:
                        found_file = file
                    else:
                        duplicate_files.append(file)
            
            if found_file:
                if duplicate_files:
                    # Multiple files found with same base name
                    print(f"⚠️  DUPLICATE FILES FOUND for {base_name}:")
                    print(f"    Primary: {found_file}")
                    for dup_file in duplicate_files:
                        print(f"    Duplicate: {dup_file}")
                    print(f"    ERROR: Using {found_file} but please remove duplicates!")
                    setattr(self, attribute_name, found_file)  # Use first found file
                else:
                    setattr(self, attribute_name, found_file)
                    print(f"✅ Found {base_name}: {found_file}")
            else:
                print(f"❌ Missing {base_name}")
                setattr(self, attribute_name, None)
    
    def setup_detailed_logging(self):
        """Setup detailed logging to file with timestamp-based filename."""
        try:
            # Write log file in the same directory as this script
            script_dir = os.path.dirname(os.path.abspath(__file__))
            log_dir = script_dir
            
            # Create timestamped log file with requested naming: "NBD-MF-10-GA & NBD-MF-11-IS_LOG_Time_Date"
            now = datetime.now()
            time_part = now.strftime("%H%M%S")
            date_part = now.strftime("%Y%m%d")
            log_filename = os.path.join(log_dir, f"NBD-MF-10-GA & NBD-MF-11-IS_LOG_{time_part}_{date_part}.log")
            
            # Configure logging
            logging.basicConfig(
                level=logging.INFO,
                format='%(asctime)s - %(levelname)s - %(message)s',
                handlers=[
                    logging.FileHandler(log_filename, encoding='utf-8'),
                    logging.StreamHandler()  # Also print to console
                ]
            )
            
            self.detailed_logger = logging.getLogger(__name__)
            self.detailed_logger.info(f"=== DETAILED EXCEL TRANSFER LOG STARTED ===")
            self.detailed_logger.info(f"Log file: {os.path.abspath(log_filename)}")
            self.detailed_logger.info(f"Working directory: {os.getcwd()}")
            
        except Exception as e:
            print(f"Warning: Could not setup detailed logging: {e}")
            self.detailed_logger = None
    
    def log_detailed(self, message, level="INFO"):
        """Log detailed message to both file and console with professional formatting."""
        if self.detailed_logger:
            # Remove emojis for professional log file output
            professional_message = self._remove_emojis(message)
            if level == "ERROR":
                self.detailed_logger.error(professional_message)
            elif level == "WARNING":
                self.detailed_logger.warning(professional_message)
            else:
                self.detailed_logger.info(professional_message)
    
    def _remove_emojis(self, text):
        """Remove emojis from text for professional logging."""
        import re
        # Remove common emojis used in the script
        emoji_pattern = re.compile(
            "["
            "\U0001F300-\U0001F5FF"  # symbols & pictographs
            "\U0001F680-\U0001F6FF"  # transport & map symbols
            "\U0001F1E0-\U0001F1FF"  # flags (iOS)
            "\U00002702-\U000027B0"  # dingbats
            "\U000024C2-\U0001F251"  # enclosed characters
            "\U0001F600-\U0001F64F"  # emoticons
            "\U0001F680-\U0001F6FF"  # transport & map
            "\U0001F1E0-\U0001F1FF"  # flags (iOS)
            "\U00002600-\U000026FF"  # miscellaneous symbols
            "\U00002700-\U000027BF"  # dingbats
            "]+", flags=re.UNICODE)
        return emoji_pattern.sub('', text).strip()
        
    def _parse_numeric(self, value):
        """Parse a numeric value that may be in accounting format like '(21.00)', with commas, or as a number."""
        if value is None:
            return None
        # If it's already a number, return as float
        if isinstance(value, (int, float)):
            try:
                return float(value)
            except Exception:
                return None
        # Convert to string and normalize
        try:
            s = str(value).strip()
            if s == "" or s == "-":
                return None
            # Remove commas
            s_clean = s.replace(",", "")
            # Handle accounting negatives like (21.00)
            if s_clean.startswith("(") and s_clean.endswith(")"):
                s_clean = "-" + s_clean[1:-1]
            return float(s_clean)
        except Exception:
            return None
        
    def log(self, message):
        """Log message with timestamp and user-friendly formatting."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_msg = f"[{timestamp}] {message}"
        try:
            print(log_msg)
        except UnicodeEncodeError:
            print(log_msg.encode('ascii', 'ignore').decode('ascii'))
        self.log_messages.append(log_msg)
    
    def start_excel(self):
        """Start Excel application with robust error handling."""
        try:
            try:
                self.excel_app = win32.GetActiveObject("Excel.Application")
                self.log("Connected to existing Excel instance")
            except:
                self.excel_app = win32.Dispatch("Excel.Application")
                self.log("Created new Excel instance")
            
            # Make Excel invisible to prevent windows from showing
            self.excel_app.Visible = False
            self.log("Excel set to invisible mode")
            
            try:
                self.excel_app.DisplayAlerts = False
                self.log("DisplayAlerts disabled")
            except Exception as e:
                self.log(f"Warning: Could not set DisplayAlerts: {e}")
            
            try:
                self.excel_app.ScreenUpdating = False  
                self.log("ScreenUpdating disabled")
            except Exception as e:
                self.log(f"Warning: Could not set ScreenUpdating: {e}")
            
            self.log("Skipping calculation mode setting to avoid COM errors")
            self.log("✅ Excel started successfully - ready for data transfer")
            return True
            
        except Exception as e:
            self.log(f"❌ Failed to start Excel: {e}")
            self.log("💡 Solutions to try:")
            self.log("   1. Close all Excel applications completely")
            self.log("   2. End any Excel processes in Task Manager")
            self.log("   3. Restart your computer if the issue persists")
            return False

    # ============================================================
    # File detection helpers (non-breaking additions)
    # These methods detect input files in the current folder using
    # required keywords and support both .xlsx and .xlsb extensions.
    # They do not change processing logic; they only set file paths
    # before the original flow runs.
    # ============================================================
    def _detect_file_by_keywords(self, required_phrase, search_dir='.'):
        """
        Find a file in current working directory whose filename contains the
        required_phrase (case-insensitive), ignoring date parts, and whose
        extension is .xlsx or .xlsb. If multiple matches exist, pick the most
        recently modified. Returns absolute path or None if not found.
        """
        # Each statement below is line-commented to ensure clarity and traceability.
        try:
            # Build an informational log line indicating which phrase and directory are being searched
            self.log(f"[INFO] Detecting file for: {required_phrase} | in: {os.path.abspath(search_dir)}")
            # Initialize an empty list to accumulate (modified_time, absolute_path) tuples for candidate files
            candidates = []
            # Iterate every entry present in the specified directory
            for name in os.listdir(search_dir):
                # Normalize the filename to lowercase for case-insensitive matching
                lower = name.lower()
                # Check conditions: contains required keyword, has Excel extension, and is not a temporary file
                excel_extensions = ('.xlsx', '.xlsb', '.xlsm')
                if required_phrase.lower() in lower and lower.endswith(excel_extensions) and not lower.startswith('~'):
                    # Construct the absolute path for the current filename
                    full = os.path.abspath(os.path.join(search_dir, name))
                    try:
                        # Attempt to retrieve the last-modified timestamp to aid in selecting the newest file
                        mtime = os.path.getmtime(full)
                    except Exception:
                        # On any failure to read metadata, default to zero so such entries sort to the end
                        mtime = 0
                    # Record the candidate with its modification time for later sorting
                    candidates.append((mtime, full))
            # If no candidates matched, emit a warning and return None to signal no detection
            if not candidates:
                self.log(f"[WARN] No file found for phrase: {required_phrase}")
                return None
            # Sort candidates by modification time descending so the newest file is first
            candidates.sort(reverse=True)
            # Select the path portion of the newest candidate
            chosen = candidates[0][1]
            # Log the absolute path of the selected file for transparency
            self.log(f"[INFO] Selected file: {chosen}")
            # Return the absolute path to the caller
            return chosen
        except Exception as e:
            # Log the error with context and print the traceback for deep diagnostics
            self.log(f"[ERROR] Detection failed for phrase '{required_phrase}': {e}")
            traceback.print_exc()
            # Return None to indicate detection did not succeed
            return None

    def detect_input_files(self):
        """
        Detect and set all required file paths based on keywords, supporting
        .xlsx, .xlsb, and .xlsm formats. Keeps defaults if detection finds nothing.
        """
        # Every operational step is documented inline to preserve and explain behavior.
        try:
            # Announce the start of keyword-based detection with Excel file support
            self.log("[INFO] Starting keyword-based input file detection (.xlsx/.xlsb/.xlsm)")
            # Capture the current working directory as a starting reference point
            cwd = os.getcwd()
            # Initialize project_root to the current directory for upward traversal
            project_root = cwd
            # Attempt to locate the 'CBSL_Reporting-main' anchor folder by walking up to five levels
            for _ in range(5):
                # Construct a candidate path at the current level
                candidate = os.path.join(project_root, 'CBSL_Reporting-main')
                # If the candidate exists as a directory, stop searching upward
                if os.path.isdir(candidate):
                    break
                # Otherwise, move one level up and continue the search
                project_root = os.path.dirname(project_root)
            # Resolve the 'working/monthly' path if the anchor folder was found
            base_working = os.path.join(candidate, 'working', 'monthly') if os.path.isdir(candidate) else None
            # If the monthly base path is missing, log a warning and fall back to current directory search
            if not base_working or not os.path.isdir(base_working):
                self.log(f"[WARN] Could not locate monthly folder from cwd: {cwd}")
                search_dir = '.'
            else:
                # Enumerate subdirectories under the monthly base to determine the latest one by modification time
                monthly_dirs = [os.path.join(base_working, d) for d in os.listdir(base_working) if os.path.isdir(os.path.join(base_working, d))]
                # If no monthly subdirectories exist, warn and fall back to current directory
                if not monthly_dirs:
                    self.log(f"[WARN] No subfolders under: {base_working}")
                    search_dir = '.'
                else:
                    # Sort monthly directories by last modified time (newest first)
                    monthly_dirs.sort(key=lambda p: os.path.getmtime(p), reverse=True)
                    # Pick the most recently modified monthly folder
                    chosen_month = monthly_dirs[0]
                    # Log which monthly folder was selected
                    self.log(f"[INFO] Selected monthly folder: {chosen_month}")
                    # Within the chosen monthly folder, list its immediate subdirectories (expecting exactly one)
                    inner_dirs = [os.path.join(chosen_month, d) for d in os.listdir(chosen_month) if os.path.isdir(os.path.join(chosen_month, d))]
                    # If no inner folder exists, search directly within the chosen monthly folder
                    if not inner_dirs:
                        self.log(f"[WARN] No inner subfolder found under: {chosen_month}")
                        search_dir = chosen_month
                    else:
                        # Select the first (and reportedly only) inner subfolder
                        chosen_inner = inner_dirs[0]
                        # Log the chosen inner subfolder for auditability
                        self.log(f"[INFO] Selected inner folder: {chosen_inner}")
                        # Prefer a specific target subfolder named 'NBD-MF-10-GA & NBD-MF-11-IS' inside the inner folder
                        target_dir = os.path.join(chosen_inner, 'NBD-MF-10-GA & NBD-MF-11-IS')
                        # If that target subfolder exists, use it for searching; otherwise, fall back to the inner folder
                        if os.path.isdir(target_dir):
                            search_dir = target_dir
                        else:
                            self.log(f"[WARN] Target folder not found: {target_dir}. Falling back to inner folder")
                            search_dir = chosen_inner

            # Define the mapping from instance attributes to keyword phrases for detection (date-agnostic)
            mapping = {
                'main_file': "NBD-MF-10-GA & NBD-MF-11-IS",
                'sofp_source': "NBD-MF-01-SOFP & SOCI AFL Monthly FS",
                'investment_source': "Investment Schedule",
                'fd_base_file': "FD Base as at",
                'borrowing_source': "Borrowing report",
            }
            # Iterate over each attribute/phrase pair and attempt detection within the resolved search directory
            for attr, phrase in mapping.items():
                # Invoke the helper to find the newest matching file by keyword and extension
                detected = self._detect_file_by_keywords(phrase, search_dir=search_dir)
                if detected:
                    # Assign the absolute path to the corresponding instance attribute for downstream use
                    setattr(self, attr, detected)
                    # Log the resolved path for this attribute
                    self.log(f"[INFO] {attr} -> {detected}")
                else:
                    # If detection failed, retain the original default value and log a warning for visibility
                    self.log(f"[WARN] {attr} not found by detection; keeping default: {getattr(self, attr)}")
            # Indicate that the detection phase has completed
            self.log("[INFO] File detection completed")
        except Exception as e:
            # Log any unexpected exceptions and include a traceback for detailed debugging
            self.log(f"[ERROR] File detection encountered an error: {e}")
            traceback.print_exc()
    
    def refresh_borrowings_pivot_tables(self):
        """
        NEW FEATURE: Refresh all pivot tables in Borrowings sheet
        Finds and refreshes all pivot tables without needing to know their names
        """
        self.log("Refreshing pivot tables in Borrowings sheet...")
        
        try:
            # Log the file and sheet being accessed prior to opening
            self.log(f"[INFO] Opening destination workbook for pivot refresh: {self.main_file} | sheet: Borrowings")
            dest_wb = self.excel_app.Workbooks.Open(os.path.abspath(self.main_file))
            dest_ws = dest_wb.Worksheets("Borrowings")
            
            # Get all pivot tables in the worksheet
            pivot_tables = dest_ws.PivotTables()
            pivot_count = pivot_tables.Count
            
            self.log(f"Borrowings: Found {pivot_count} pivot table(s)")
            
            if pivot_count > 0:
                refreshed_count = 0
                for i in range(1, pivot_count + 1):
                    try:
                        # Corrected: Use Item method to access pivot table
                        pivot_table = pivot_tables.Item(i)
                        pivot_name = pivot_table.Name
                        self.log(f"Borrowings: Refreshing pivot table '{pivot_name}'")
                        pivot_table.RefreshTable()
                        refreshed_count += 1
                        self.log(f"Borrowings: Successfully refreshed '{pivot_name}'")
                    except Exception as e:
                        self.log(f"Borrowings: Error refreshing pivot table {i}: {e}")
                        continue
                
                self.log(f"Borrowings: Refreshed {refreshed_count}/{pivot_count} pivot tables")
            else:
                self.log("Borrowings: No pivot tables found in the sheet")
            
            # Final verification: log C16 formula and value to confirm formula set correctly
            try:
                c16_formula_final = str(dest_ws.Cells(16, 3).Formula)
                c16_value_final = dest_ws.Cells(16, 3).Value
                self.log_detailed(f"VERIFICATION: C16 FORMULA: {c16_formula_final}")
                self.log_detailed(f"VERIFICATION: C16 VALUE: {c16_value_final}")
            except Exception as e:
                self.log_detailed(f"WARNING: Failed to read back C16 formula/value: {e}", "WARNING")
            
            dest_wb.Save()
            dest_wb.Close(SaveChanges=False)
            
            return True
            
        except Exception as e:
            self.log(f"Borrowings pivot table refresh error: {e}")
            traceback.print_exc()
            return False

    def copy_borrowing_data(self):
        """
        NEW FEATURE: Copy borrowing data from Borrowing report 31-07-2025.xlsx to main file
        Copy from row 755+ in source to row 3+ in destination:
        - C2 (Institution2) -> A2 (Source of Facility)
        - E2 (Type) -> B2 (Type of Rate)  
        - F2 (Amount) -> C2 (Amount)
        - G2 (Rate) -> D2 (Rate)
        - I2 (Maturity Date) -> E2 (Maturity Date)
        Clear existing formulas before pasting values
        THEN calculate Residual Period and Bucket columns
        """
        self.log("Starting Borrowing data transfer...")

        source_wb = None
        dest_wb = None
        try:
            source_wb = self.excel_app.Workbooks.Open(os.path.abspath(self.borrowing_source))
            source_ws = source_wb.Worksheets("Borrowing")
            dest_wb = self.excel_app.Workbooks.Open(os.path.abspath(self.main_file))
            dest_ws = dest_wb.Worksheets("Borrowings")
        finally:
            # Do not close here; close after data transfer is complete
            pass
        
        # Column mapping: source -> destination
        column_map = {
            3: 1,   # C (Institution2) -> A (Source of Facility)
            5: 2,   # E (Type) -> B (Type of Rate)
            6: 3,   # F (Amount) -> C (Amount)
            7: 4,   # G (Rate) -> D (Rate)
            9: 5    # I (Maturity Date) -> E (Maturity Date)
        }
        
        # Clear existing data in destination (starting from row 3)
        cleared_count = 0
        for row in range(3, 100):  # Clear up to row 100
            for col in [1, 2, 3, 4, 5, 6, 7]:  # A, B, C, D, E, F, G columns
                try:
                    cell = dest_ws.Cells(row, col)
                    if not str(cell.Formula).startswith('='):
                        cell.ClearContents()
                        cleared_count += 1
                except:
                    continue
        
        self.log(f"Borrowing: Cleared {cleared_count} old values in destination")
        
        # Copy borrowing data starting from source row 755
        copied_count = 0
        rows_processed = 0
        dest_row_start = 3  # Start pasting at row 3 in destination
        
        # Find last row with data in source (scan multiple columns to be sure)
        last_row_c = source_ws.Cells(source_ws.Rows.Count, 3).End(win32.constants.xlUp).Row  # Column C
        last_row_f = source_ws.Cells(source_ws.Rows.Count, 6).End(win32.constants.xlUp).Row  # Column F
        last_row = max(last_row_c, last_row_f, 1767)  # Use 1767 as minimum or higher if data exists
        
        self.log(f"Borrowing: Last row in column C: {last_row_c}, column F: {last_row_f}")
        self.log(f"Borrowing: Using last row: {last_row}")
        self.log(f"Borrowing: Scanning source from row 755 to {last_row}")
        
        # Store destination rows that will need formulas
        destination_rows_with_data = []
        
        for src_row in range(755, last_row + 1):
            # Check if row has data (check multiple columns for better detection)
            institution_value = source_ws.Cells(src_row, 3).Value  # Column C
            amount_value = source_ws.Cells(src_row, 6).Value      # Column F
            
            # Row is valid if it has data in Institution2 OR Amount column
            if ((institution_value is not None and str(institution_value).strip() != "") or 
                (amount_value is not None and str(amount_value).strip() != "")):
                
                dest_row = dest_row_start + rows_processed
                row_copied = 0
                
                # Copy each column according to mapping
                for src_col, dest_col in column_map.items():
                    try:
                        source_value = source_ws.Cells(src_row, src_col).Value
                        if source_value is not None and str(source_value).strip() != "":
                            dest_cell = dest_ws.Cells(dest_row, dest_col)
                            dest_cell.ClearContents()  # Clear any existing formula
                            dest_cell.Value = source_value  # Set only the value
                            row_copied += 1
                            copied_count += 1
                    except Exception:
                        continue
                
                if row_copied > 0:
                    destination_rows_with_data.append(dest_row)
                    rows_processed += 1
                    if rows_processed <= 10 or rows_processed % 100 == 0:  # Log first 10 and every 100th
                        self.log(f"Borrowing: Copied row {src_row} -> row {dest_row} ({row_copied} cells)")
                
                # Remove the 50 rows limit to process all data
                if rows_processed >= 2000:  # Safety limit increased
                    self.log("Borrowing: Reached 2000 rows limit - stopping")
                    break
        
        self.log(f"Borrowing: Processed {rows_processed} rows, copied {copied_count} values")
        
        # NEW: Calculate Residual Period and Bucket for each row with data
        self.log("Borrowing: Calculating Residual Period and Bucket columns...")
        formulas_added = 0
        
        for dest_row in destination_rows_with_data:
            try:
                # Column F: Residual Period = Maturity Date (E) - F1
                residual_formula = f"=E{dest_row}-$F$1"
                dest_ws.Cells(dest_row, 6).Formula = residual_formula  # Column F
                
                # Column G: Bucket = VLOOKUP based on Residual Period
                bucket_formula = f"=VLOOKUP(F{dest_row},T!$A$2:$C$9,3,TRUE)"
                dest_ws.Cells(dest_row, 7).Formula = bucket_formula  # Column G
                
                formulas_added += 2
                
                if formulas_added <= 20:  # Log first 10 rows
                    self.log(f"Borrowing: Row {dest_row} - F: {residual_formula}, G: {bucket_formula}")
                    
            except Exception as e:
                self.log(f"Borrowing: Error adding formulas to row {dest_row}: {e}")
                continue
        
        self.log(f"Borrowing: Added {formulas_added} formulas ({len(destination_rows_with_data)} rows)")
        
        # Verify F1 exists and has a date value
        try:
            f1_value = dest_ws.Cells(1, 6).Value  # F1
            f1_formula = str(dest_ws.Cells(1, 6).Formula)
            self.log(f"Borrowing: F1 value={f1_value}, formula={f1_formula}")
            
            if f1_value is None:
                self.log("WARNING: F1 is empty - Residual Period calculations may not work correctly")
        except Exception as e:
            self.log(f"Borrowing: Error checking F1: {e}")
        
        dest_wb.Save()
        source_wb.Close(SaveChanges=False)
        dest_wb.Close(SaveChanges=False)
        
        return True

    # The following except block was incorrectly indented and placed outside of a try block.
    # It should be inside the function where the try block starts.
    # Here, we remove the misplaced except block and keep only the function definition.

    def verify_borrowing_amounts(self):
        """
        NEW FEATURE: Verify sum of Amount (F column) in source equals Borrowings C1 in destination
        """
        self.log("Verifying borrowing amounts...")
        
        try:
            self.log(f"[INFO] Opening borrowing source for verification: {self.borrowing_source} | sheet: Borrowing")
            source_wb = self.excel_app.Workbooks.Open(os.path.abspath(self.borrowing_source))
            source_ws = source_wb.Worksheets("Borrowing")
            self.log(f"[INFO] Opening main destination for verification: {self.main_file} | sheet: Borrowings")
            dest_wb = self.excel_app.Workbooks.Open(os.path.abspath(self.main_file))
            dest_ws = dest_wb.Worksheets("Borrowings")
            
            # Calculate sum of Amount column (F) from source starting row 755
            amount_sum = 0
            count = 0
            
            # Find last row with data (scan multiple columns)
            last_row_c = source_ws.Cells(source_ws.Rows.Count, 3).End(win32.constants.xlUp).Row  # Column C
            last_row_f = source_ws.Cells(source_ws.Rows.Count, 6).End(win32.constants.xlUp).Row  # Column F
            last_row = max(last_row_c, last_row_f, 1767)  # Use 1767 as minimum or higher if data exists
            
            self.log(f"Borrowing Verification: Scanning from row 755 to {last_row}")
            
            for src_row in range(755, last_row + 1):
                # Check if row has data in Institution2 OR Amount column
                institution_value = source_ws.Cells(src_row, 3).Value  # Column C
                amount_value = source_ws.Cells(src_row, 6).Value      # Column F
                
                # Process row if it has data in either column
                if ((institution_value is not None and str(institution_value).strip() != "") or 
                    (amount_value is not None and str(amount_value).strip() != "")):
                    
                    if amount_value is not None:
                        try:
                            amount_numeric = float(amount_value)
                            amount_sum += amount_numeric
                            count += 1
                        except (ValueError, TypeError):
                            continue
            
            # Get C1 value from destination
            c1_value = dest_ws.Cells(1, 3).Value  # C1 in Borrowings sheet
            
            self.log(f"Borrowing Verification: Source sum F column (Amount): {amount_sum} ({count} records)")
            self.log(f"Borrowing Verification: Destination C1 value: {c1_value}")
            
            # Compare values (allow small floating point differences)
            if c1_value is not None:
                try:
                    c1_numeric = float(c1_value)
                    difference = abs(amount_sum - c1_numeric)
                    matches = difference < 0.01  # Allow 0.01 difference for floating point
                    
                    self.log(f"Borrowing Verification: Difference: {difference}")
                    self.log(f"Borrowing Verification: Match: {matches}")
                    
                    source_wb.Close(SaveChanges=False)
                    dest_wb.Close(SaveChanges=False)
                    
                    return matches
                    
                except (ValueError, TypeError):
                    self.log("Borrowing Verification: C1 value is not numeric")
                    source_wb.Close(SaveChanges=False)
                    dest_wb.Close(SaveChanges=False)
                    return False
            else:
                self.log("Borrowing Verification: C1 is empty")
                source_wb.Close(SaveChanges=False)
                dest_wb.Close(SaveChanges=False)
                return False
            
        except Exception as e:
            self.log(f"Borrowing verification error: {e}")
            traceback.print_exc()
            return False
    
    def close_excel(self):
        """Close Excel application safely."""
        try:
            if self.excel_app:
                try:
                    self.excel_app.DisplayAlerts = True
                except Exception:
                    pass
                
                try:
                    self.excel_app.ScreenUpdating = True
                except Exception:
                    pass
                
                self.excel_app.Quit()
                self.excel_app = None
                self.log("Excel closed successfully")
        except Exception as e:
            self.log(f"Error closing Excel: {e}")
            traceback.print_exc()
    
    def check_files_exist(self):
        """Check all required files exist."""
        # UPDATED: Added borrowing file to the required files list
        files = [self.main_file, self.sofp_source, self.investment_source, self.fd_base_file, self.borrowing_source]
        for file in files:
            if not os.path.exists(file):
                self.log(f"ERROR: File not found: {file}")
                return False
            size_mb = os.path.getsize(file) / (1024 * 1024)
            self.log(f"Found: {file} ({size_mb:.1f} MB)")
        return True
    
    def copy_sofp_values_only(self):
        """Copy ONLY VALUES from SOFP columns D and E - preserve all formulas in destination."""
        self.log("Starting SOFP data transfer (values only)...")
        self.log_detailed("=== SOFP DATA TRANSFER STARTED ===")
        
        try:
            self.log(f"[INFO] Opening SOFP source: {self.sofp_source} | sheet: NBD-MF-01-SOFP")
            self.log_detailed(f"SOURCE FILE: {self.sofp_source}")
            self.log_detailed(f"SOURCE SHEET: NBD-MF-01-SOFP")
            source_wb = self.excel_app.Workbooks.Open(os.path.abspath(self.sofp_source))
            source_ws = source_wb.Worksheets("NBD-MF-01-SOFP")
            
            self.log(f"[INFO] Opening main destination: {self.main_file} | sheet: NBD-MF-01-SOFP")
            self.log_detailed(f"DESTINATION FILE: {self.main_file}")
            self.log_detailed(f"DESTINATION SHEET: NBD-MF-01-SOFP")
            dest_wb = self.excel_app.Workbooks.Open(os.path.abspath(self.main_file))
            dest_ws = dest_wb.Worksheets("NBD-MF-01-SOFP")
            
            last_row = source_ws.Cells(source_ws.Rows.Count, 4).End(win32.constants.xlUp).Row
            last_row = min(last_row, 1000)
            
            self.log(f"SOFP: Copying values from {last_row} rows")
            self.log_detailed(f"PROCESSING ROWS: 1 to {last_row}")
            self.log_detailed(f"COPY COLUMNS: D (column 4) and E (column 5)")
            self.log_detailed(f"COPY STRATEGY: Values only, preserve existing formulas in destination")
            
            copied_count = 0
            for row in range(1, last_row + 1):
                try:
                    # Column D processing
                    source_d_value = source_ws.Cells(row, 4).Value
                    dest_d_cell = dest_ws.Cells(row, 4)
                    dest_d_formula = str(dest_d_cell.Formula)
                    
                    if source_d_value is not None and not dest_d_formula.startswith('='):
                        dest_d_cell.Value = source_d_value
                        copied_count += 1
                        self.log_detailed(f"COPIED: Row {row}, Column D: '{source_d_value}' (Source: D{row} -> Destination: D{row})")
                    elif source_d_value is not None:
                        self.log_detailed(f"SKIPPED: Row {row}, Column D: Destination has formula '{dest_d_formula}'")
                    
                    # Column E processing
                    source_e_value = source_ws.Cells(row, 5).Value
                    dest_e_cell = dest_ws.Cells(row, 5)
                    dest_e_formula = str(dest_e_cell.Formula)
                    
                    if source_e_value is not None and not dest_e_formula.startswith('='):
                        dest_e_cell.Value = source_e_value
                        copied_count += 1
                        self.log_detailed(f"COPIED: Row {row}, Column E: '{source_e_value}' (Source: E{row} -> Destination: E{row})")
                    elif source_e_value is not None:
                        self.log_detailed(f"SKIPPED: Row {row}, Column E: Destination has formula '{dest_e_formula}'")
                        
                except Exception as e:
                    self.log_detailed(f"ERROR: Row {row} processing failed: {e}", "ERROR")
                    continue
            
            self.log(f"SOFP: Copied {copied_count} values (preserved all formulas)")
            self.log_detailed(f"TOTAL VALUES COPIED: {copied_count}")
            self.log_detailed(f"FORMULAS PRESERVED: All existing formulas in destination were kept intact")
            
            dest_wb.Save()
            self.log_detailed(f"SAVED: Destination file {self.main_file}")
            source_wb.Close(SaveChanges=False)
            dest_wb.Close(SaveChanges=False)
            self.log_detailed("=== SOFP DATA TRANSFER COMPLETED ===")
            
            return True
            
        except Exception as e:
            self.log(f"SOFP transfer error: {e}")
            self.log_detailed(f"CRITICAL ERROR in SOFP transfer: {e}", "ERROR")
            traceback.print_exc()
            return False
    
    def copy_system_tb_values_only(self):
        """Copy ONLY VALUES from System TB - preserve all formulas in destination and force text in column A."""
        self.log("Starting System TB data transfer (values only)...")
        self.log_detailed("=== SYSTEM TB DATA TRANSFER STARTED ===")
        
        try:
            self.log(f"[INFO] Opening SOFP source: {self.sofp_source} | sheet: System TB")
            self.log_detailed(f"SOURCE FILE: {self.sofp_source}")
            self.log_detailed(f"SOURCE SHEET: System TB")
            source_wb = self.excel_app.Workbooks.Open(os.path.abspath(self.sofp_source))
            source_ws = source_wb.Worksheets("System TB")
            
            self.log(f"[INFO] Opening main destination: {self.main_file} | sheet: System TB")
            self.log_detailed(f"DESTINATION FILE: {self.main_file}")
            self.log_detailed(f"DESTINATION SHEET: System TB")
            dest_wb = self.excel_app.Workbooks.Open(os.path.abspath(self.main_file))
            dest_ws = dest_wb.Worksheets("System TB")
            
            # Clear old values (preserve formulas)
            self.log_detailed("CLEARING PHASE: Removing old values while preserving formulas")
            cleared_count = 0
            for row in range(2, 500):
                for col in [1, 2, 3]:
                    try:
                        cell = dest_ws.Cells(row, col)
                        cell_formula = str(cell.Formula)
                        if not cell_formula.startswith('='):
                            cell.Value = None
                            cleared_count += 1
                    except:
                        continue
            self.log(f"System TB: Cleared {cleared_count} old values (kept formulas)")
            self.log_detailed(f"CLEARED VALUES: {cleared_count} cells (rows 2-499, columns A-C)")
            
            # Copy new values
            last_row = source_ws.Cells(source_ws.Rows.Count, 1).End(win32.constants.xlUp).Row
            last_row = min(last_row, 500)
            
            self.log_detailed(f"PROCESSING ROWS: Source rows 3 to {last_row} -> Destination rows 2 to {last_row-1}")
            self.log_detailed(f"COPY COLUMNS: A (column 1), B (column 2), C (column 3)")
            self.log_detailed(f"SPECIAL FORMATTING: Column A will be formatted as text to preserve leading zeros")
            
            copied_count = 0
            for src_row in range(3, last_row + 1):
                dest_row = src_row - 1
                for col in [1, 2, 3]:
                    try:
                        source_value = source_ws.Cells(src_row, col).Value
                        if source_value is not None and str(source_value).strip() != "":
                            dest_cell = dest_ws.Cells(dest_row, col)
                            dest_formula = str(dest_cell.Formula)
                            if not dest_formula.startswith('='):
                                # Force column A to text to preserve leading zeros
                                if col == 1:
                                    dest_cell.NumberFormat = "@"
                                    dest_cell.Value = str(source_value)
                                    self.log_detailed(f"COPIED (TEXT): Row {src_row}->{dest_row}, Column A: '{source_value}' (formatted as text)")
                                else:
                                    dest_cell.Value = source_value
                                    col_letter = self.get_column_letter(col)
                                    self.log_detailed(f"COPIED: Row {src_row}->{dest_row}, Column {col_letter}: '{source_value}'")
                                copied_count += 1
                            else:
                                col_letter = self.get_column_letter(col)
                                self.log_detailed(f"SKIPPED: Row {src_row}->{dest_row}, Column {col_letter}: Destination has formula '{dest_formula}'")
                    except Exception as e:
                        self.log_detailed(f"ERROR: Row {src_row}->{dest_row}, Column {col} processing failed: {e}", "ERROR")
                        continue
            
            self.log(f"System TB: Copied {copied_count} values (preserved all formulas, column A as text)")
            self.log_detailed(f"TOTAL VALUES COPIED: {copied_count}")
            self.log_detailed(f"FORMULAS PRESERVED: All existing formulas in destination were kept intact")
            self.log_detailed(f"TEXT FORMATTING: Column A values formatted as text to preserve leading zeros")
            
            dest_wb.Save()
            self.log_detailed(f"SAVED: Destination file {self.main_file}")
            source_wb.Close(SaveChanges=False)
            dest_wb.Close(SaveChanges=False)
            self.log_detailed("=== SYSTEM TB DATA TRANSFER COMPLETED ===")
            
            return True
            
        except Exception as e:
            self.log(f"System TB transfer error: {e}")
            self.log_detailed(f"CRITICAL ERROR in System TB transfer: {e}", "ERROR")
            traceback.print_exc()
            return False
    
    def verify_nbdmf10_formulas(self):
        """Verify C6 = N6 and C7 = N7 in NBD-MF-10-GA sheet."""
        self.log("Verifying NBD-MF-10-GA sheet formulas...")
        
        try:
            self.log(f"[INFO] Opening main destination for NBD-MF-10-GA verification: {self.main_file} | sheet: NBD-MF-10-GA")
            dest_wb = self.excel_app.Workbooks.Open(os.path.abspath(self.main_file))
            ws = dest_wb.Worksheets("NBD-MF-10-GA")
            
            c6_value = ws.Cells(6, 3).Value
            n6_value = ws.Cells(6, 14).Value
            c7_value = ws.Cells(7, 3).Value
            n7_value = ws.Cells(7, 14).Value
            
            self.log(f"NBD-MF-10-GA: C6={c6_value}, N6={n6_value} | Match: {c6_value == n6_value}")
            self.log(f"NBD-MF-10-GA: C7={c7_value}, N7={n7_value} | Match: {c7_value == n7_value}")
            
            verification_passed = (c6_value == n6_value) and (c7_value == n7_value)
            
            dest_wb.Close(SaveChanges=False)
            
            if verification_passed:
                self.log("NBD-MF-10-GA: Verification PASSED")
            else:
                self.log("NBD-MF-10-GA: Verification FAILED")
            
            return verification_passed
            
        except Exception as e:
            self.log(f"NBD-MF-10-GA verification error: {e}")
            traceback.print_exc()
            return False
    
    def clean_investment_sheet_errors(self):
        """Clean #VALUE! and other errors from Investment sheet before data transfer."""
        self.log("Cleaning errors from Investment sheet...")
        
        try:
            self.log(f"[INFO] Opening main destination for Investment cleanup: {self.main_file} | sheet: Investments")
            dest_wb = self.excel_app.Workbooks.Open(os.path.abspath(self.main_file))
            dest_ws = dest_wb.Worksheets("Investments")
            
            cleaned_count = 0
            error_patterns = ["#VALUE!", "#N/A", "#REF!", "#DIV/0!", "#NAME?", "#NULL!"]
            
            for row in range(3, 51):
                for col in range(1, 27):
                    try:
                        cell = dest_ws.Cells(row, col)
                        cell_value = cell.Value
                        cell_formula = str(cell.Formula)
                        
                        has_error = False
                        if cell_value is not None:
                            cell_str = str(cell_value).strip().upper()
                            for error_pattern in error_patterns:
                                if error_pattern in cell_str:
                                    has_error = True
                                    break
                        
                        if cell_formula != "None":
                            for error_pattern in error_patterns:
                                if error_pattern in cell_formula:
                                    has_error = True
                                    break
                        
                        if has_error:
                            if not cell_formula.startswith('=') or len(cell_formula) < 20:
                                cell.Value = None
                                cleaned_count += 1
                        
                    except Exception:
                        continue
            
            self.log(f"Investment sheet: Cleaned {cleaned_count} error cells")
            
            dest_wb.Save()
            dest_wb.Close(SaveChanges=False)
            
            return True
            
        except Exception as e:
            self.log(f"Investment sheet error cleanup failed: {e}")
            traceback.print_exc()
            return False
    
    def copy_investment_values_only(self):
        """Copy ONLY VALUES from Investment sheet with precise column mapping."""
        self.log("Starting Investment data transfer (values only)...")
        self.log_detailed("=== INVESTMENT DATA TRANSFER STARTED ===")
        
        try:
            self.log(f"[INFO] Opening Investment source: {self.investment_source} | sheet: Investment")
            self.log_detailed(f"SOURCE FILE: {self.investment_source}")
            self.log_detailed(f"SOURCE SHEET: Investment")
            source_wb = self.excel_app.Workbooks.Open(os.path.abspath(self.investment_source))
            source_ws = source_wb.Worksheets("Investment")
            
            self.log(f"[INFO] Opening main destination: {self.main_file} | sheet: Investments")
            self.log_detailed(f"DESTINATION FILE: {self.main_file}")
            self.log_detailed(f"DESTINATION SHEET: Investments")
            dest_wb = self.excel_app.Workbooks.Open(os.path.abspath(self.main_file))
            dest_ws = dest_wb.Worksheets("Investments")
            
            # Verify headers in rows 1 and 2
            row1_a = str(dest_ws.Cells(1, 1).Value or "").lower()
            row2_a = str(dest_ws.Cells(2, 1).Value or "").lower()
            if "treasury" in row2_a and any(bucket in row1_a for bucket in ["less than", "7-30", "1-3", "3-6", "over"]):
                self.log("Investment: Headers confirmed in rows 1 and 2")
                self.log_detailed(f"HEADER VERIFICATION: PASSED - Row 1: '{row1_a}', Row 2: '{row2_a}'")
            else:
                self.log("Investment: Warning - Expected headers in rows 1 and 2")
                self.log_detailed(f"HEADER VERIFICATION: WARNING - Row 1: '{row1_a}', Row 2: '{row2_a}'")
            
            # Clear existing data values in destination (preserve headers in rows 1 and 2, and formulas)
            target_columns = [1, 2, 3, 5, 6, 7, 8, 9, 10, 11, 13]  # A, B, C, E, F, G, H, I, J, K, M
            self.log_detailed(f"CLEARING PHASE: Removing old values from rows 3-49, columns: {[self.get_column_letter(c) for c in target_columns]}")
            cleared_count = 0
            
            for row in range(3, 50):
                for col in target_columns:
                    try:
                        cell = dest_ws.Cells(row, col)
                        cell_formula = str(cell.Formula)
                        if not cell_formula.startswith('='):
                            cell.Value = None
                            cleared_count += 1
                    except:
                        continue
            
            self.log(f"Investment: Cleared {cleared_count} old values (kept headers and formulas)")
            self.log_detailed(f"CLEARED VALUES: {cleared_count} cells (rows 3-49, target columns)")
            
            # Column mapping: source -> destination
            column_map = {
                1: 1,   # A (Treasury Bonds) -> A
                2: 2,   # B (ISIN) -> B
                3: 3,   # C (Cost Rs.(000)) -> C
                5: 5,   # E (Rate) -> E
                6: 6,   # F (Date of Investment) -> F
                7: 7,   # G (Date of Maturity) -> G
                8: 8,   # H (Coupon) -> H
                9: 9,   # I (Face Value) -> I
                10: 10, # J (Value As per IFRSs Rs. (000)) -> J
                13: 11  # M (M2M Gain) -> K (M2M Gain/(Loss))
            }
            
            self.log_detailed("COLUMN MAPPING:")
            for src_col, dest_col in column_map.items():
                src_letter = self.get_column_letter(src_col)
                dest_letter = self.get_column_letter(dest_col)
                self.log_detailed(f"  Source {src_letter} -> Destination {dest_letter}")
            
            copied_count = 0
            rows_processed = 0
            
            # Copy values row by row from source row 12+ to destination row 4+
            last_row = source_ws.Cells(source_ws.Rows.Count, 1).End(win32.constants.xlUp).Row
            last_row = min(last_row, 100)
            
            self.log_detailed(f"PROCESSING ROWS: Source rows 13 to {last_row} -> Destination rows 3 to {3 + (last_row - 13)}")
            self.log_detailed(f"COPY STRATEGY: Values only, preserve existing formulas in destination")
            self.log_detailed(f"SPECIAL PROCESSING: M2M Gain (column M) values will be divided by 1000")
            
            for src_row in range(13, last_row + 1):
                dest_row = 3 + (src_row - 13)
                first_cell = source_ws.Cells(src_row, 1).Value
                if first_cell is None or str(first_cell).strip() == "":
                    if rows_processed > 0:
                        self.log_detailed(f"STOPPING: Empty first cell at source row {src_row} after processing {rows_processed} rows")
                        break
                    continue
                
                row_values_copied = 0
                for src_col, dest_col in column_map.items():
                    try:
                        source_value = source_ws.Cells(src_row, src_col).Value
                        if source_value is not None and str(source_value).strip() != "":
                            original_value = source_value
                            if src_col == 13 and isinstance(source_value, (int, float)):
                                source_value = source_value / 1000
                                self.log_detailed(f"ADJUSTED: Row {src_row}->{dest_row}, Column M: {original_value} -> {source_value} (divided by 1000)")
                            
                            dest_cell = dest_ws.Cells(dest_row, dest_col)
                            dest_formula = str(dest_cell.Formula)
                            if not dest_formula.startswith('='):
                                dest_cell.Value = source_value
                                row_values_copied += 1
                                copied_count += 1
                                src_letter = self.get_column_letter(src_col)
                                dest_letter = self.get_column_letter(dest_col)
                                self.log_detailed(f"COPIED: Row {src_row}->{dest_row}, {src_letter}->{dest_letter}: '{source_value}'")
                            else:
                                src_letter = self.get_column_letter(src_col)
                                dest_letter = self.get_column_letter(dest_col)
                                self.log_detailed(f"SKIPPED: Row {src_row}->{dest_row}, {src_letter}->{dest_letter}: Destination has formula '{dest_formula}'")
                    except Exception as e:
                        self.log_detailed(f"ERROR: Row {src_row}->{dest_row}, Column {src_col}->{dest_col} processing failed: {e}", "ERROR")
                        continue
                
                if row_values_copied > 0:
                    rows_processed += 1
                    if rows_processed <= 5 or rows_processed % 10 == 0:  # Log first 5 and every 10th
                        self.log_detailed(f"PROGRESS: Processed {rows_processed} rows, {row_values_copied} values in current row")
            
            self.log(f"Investment: Processed {rows_processed} rows, copied {copied_count} values")
            self.log("Investment: Headers in rows 1 and 2, data from row 4+ (all formulas preserved)")
            self.log_detailed(f"TOTAL ROWS PROCESSED: {rows_processed}")
            self.log_detailed(f"TOTAL VALUES COPIED: {copied_count}")
            self.log_detailed(f"FORMULAS PRESERVED: All existing formulas in destination were kept intact")
            self.log_detailed(f"HEADERS PRESERVED: Rows 1 and 2 kept intact")
            
            dest_wb.Save()
            self.log_detailed(f"SAVED: Destination file {self.main_file}")
            source_wb.Close(SaveChanges=False)
            dest_wb.Close(SaveChanges=False)
            self.log_detailed("=== INVESTMENT DATA TRANSFER COMPLETED ===")
            
            return True
            
        except Exception as e:
            self.log(f"Investment transfer error: {e}")
            self.log_detailed(f"CRITICAL ERROR in Investment transfer: {e}", "ERROR")
            traceback.print_exc()
            return False
    
    def copy_investment_specific_cells(self):
        """
        NEW FEATURE: Copy specific cells from Investment Schedule to main file
        - Copy J63 from Investment sheet to L52 in Investments sheet
        - Verify L54 equals J64 in source Investment sheet and X54 in Investments sheet
        - Copy I88 to L65, I89 to L66
        - Verify L68 equals I91
        - Copy I99 to J74
        """
        self.log("Starting Investment specific cells transfer...")
        
        try:
            self.log(f"[INFO] Opening Investment source for specific cells: {self.investment_source} | sheet: Investment")
            source_wb = self.excel_app.Workbooks.Open(os.path.abspath(self.investment_source))
            source_ws = source_wb.Worksheets("Investment")
            self.log(f"[INFO] Opening main destination for specific cells: {self.main_file} | sheet: Investments")
            dest_wb = self.excel_app.Workbooks.Open(os.path.abspath(self.main_file))
            dest_ws = dest_wb.Worksheets("Investments")
            
            # Step 1: Copy J63 to L52
            j63_value = source_ws.Cells(63, 10).Value  # J63
            if j63_value is not None:
                # Clear existing value in L52 without affecting formulas/formatting
                l52_cell = dest_ws.Cells(52, 12)  # L52
                if not str(l52_cell.Formula).startswith('='):
                    l52_cell.Value = j63_value
                    self.log(f"Investment: Copied J63 ({j63_value}) to L52")
                else:
                    self.log("Investment: L52 has formula - skipped")
            
            # Compare only before decimal point values
            def compare_before_decimal(val1, val2):
                if val1 is None or val2 is None:
                    return val1 == val2
                try:
                    int_val1 = int(float(val1)) if val1 is not None else None
                    int_val2 = int(float(val2)) if val2 is not None else None
                    return int_val1 == int_val2
                except (ValueError, TypeError):
                    return str(val1) == str(val2)
            
            # Step 2: Verify L54 equals J64 in source and X54 in destination (ignore decimal points for comparison)
            l54_value = dest_ws.Cells(54, 12).Value  # L54
            x54_value = dest_ws.Cells(54, 24).Value  # X54 (from Investments sheet)
            j64_value = source_ws.Cells(64, 10).Value  # J64 (from Investment sheet)
            
            l54_matches_j64 = compare_before_decimal(l54_value, j64_value)
            l54_matches_x54 = compare_before_decimal(l54_value, x54_value)
            
            self.log(f"Investment: L54={l54_value}, X54={x54_value}, J64={j64_value}")
            self.log(f"Investment: L54=J64 (before decimal): {l54_matches_j64}, L54=X54 (before decimal): {l54_matches_x54}")
            
            # Step 3: Copy I88 to L65, I89 to L66
            i88_value = source_ws.Cells(88, 9).Value  # I88
            i89_value = source_ws.Cells(89, 9).Value  # I89
            
            if i88_value is not None:
                l65_cell = dest_ws.Cells(65, 12)  # L65
                if not str(l65_cell.Formula).startswith('='):
                    l65_cell.Value = i88_value
                    self.log(f"Investment: Copied I88 ({i88_value}) to L65")
            
            if i89_value is not None:
                l66_cell = dest_ws.Cells(66, 12)  # L66
                if not str(l66_cell.Formula).startswith('='):
                    l66_cell.Value = i89_value
                    self.log(f"Investment: Copied I89 ({i89_value}) to L66")
            
            # Step 4: Verify L68 equals I91 (ignore decimal points for comparison)
            l68_value = dest_ws.Cells(68, 12).Value  # L68
            i91_value = source_ws.Cells(91, 9).Value  # I91
            l68_matches_i91 = compare_before_decimal(l68_value, i91_value)
            
            self.log(f"Investment: L68={l68_value}, I91={i91_value}")
            self.log(f"Investment: L68=I91 (before decimal): {l68_matches_i91}")
            
            # Step 5: Copy I99 to J74
            i99_value = source_ws.Cells(99, 9).Value  # I99
            if i99_value is not None:
                j74_cell = dest_ws.Cells(74, 10)  # J74
                if not str(j74_cell.Formula).startswith('='):
                    j74_cell.Value = i99_value
                    self.log(f"Investment: Copied I99 ({i99_value}) to J74")
            
            dest_wb.Save()
            source_wb.Close(SaveChanges=False)
            dest_wb.Close(SaveChanges=False)
            
            self.log("Investment: Specific cells transfer completed")
            return True
            
        except Exception as e:
            self.log(f"Investment specific cells transfer error: {e}")
            traceback.print_exc()
            return False
    
    def set_other_assets_manual_values(self):
        """
        NEW FEATURE: Set manual values in Other Assets sheet
        Set specified values in row 131, columns F to K
        """
        self.log("Setting manual values in Other Assets sheet...")
        
        try:
            self.log(f"[INFO] Opening main destination for Other Assets manual values: {self.main_file} | sheet: Other Assets")
            dest_wb = self.excel_app.Workbooks.Open(os.path.abspath(self.main_file))
            dest_ws = dest_wb.Worksheets("Other Assets")
            
            # Define manual values for row 131
            manual_values = {
                6: 111159175,    # F131
                7: 63515193,     # G131
                8: 66403345,     # H131
                9: -28547383,    # I131
                10: -4470936,    # J131
                11: -35180520    # K131
            }
            
            set_count = 0
            for col, value in manual_values.items():
                try:
                    cell = dest_ws.Cells(131, col)
                    cell.ClearContents()  # Clear any existing content/formula
                    cell.Value = value
                    col_letter = self.get_column_letter(col)
                    self.log(f"Other Assets: Set {col_letter}131 to {value}")
                    set_count += 1
                except Exception as e:
                    col_letter = self.get_column_letter(col)
                    self.log(f"Other Assets: Error setting {col_letter}131: {e}")
                    continue
            
            self.log(f"Other Assets: Set {set_count}/6 manual values in row 131")
            
            dest_wb.Save()
            dest_wb.Close(SaveChanges=False)
            
            return True
            
        except Exception as e:
            self.log(f"Other Assets manual values setting error: {e}")
            traceback.print_exc()
            return False
    
    def copy_fd_base_portfolio_data(self):
        """
        NEW FEATURE: Copy from FD Base Portfolio sheet to NBD-MF-10-GA sheet
        Copy Y2→C16, Z2→D16, AA2→E16, AB2→F16, AC2→G16, AD2→H16, AE2→I16, AF2→J16
        Clear existing formulas for ALL cells before pasting values
        """
        self.log("Starting FD Base Portfolio data transfer...")
        self.log_detailed("=== FD BASE PORTFOLIO DATA TRANSFER STARTED ===")
        
        try:
            self.log(f"[INFO] Opening FD Base source: {self.fd_base_file} | sheet: Portfolio")
            self.log_detailed(f"SOURCE FILE: {self.fd_base_file}")
            self.log_detailed(f"SOURCE SHEET: Portfolio")
            source_wb = self.excel_app.Workbooks.Open(os.path.abspath(self.fd_base_file))
            source_ws = source_wb.Worksheets("Portfolio")
            
            self.log(f"[INFO] Opening main destination for FD Base: {self.main_file} | sheet: NBD-MF-10-GA")
            self.log_detailed(f"DESTINATION FILE: {self.main_file}")
            self.log_detailed(f"DESTINATION SHEET: NBD-MF-10-GA")
            dest_wb = self.excel_app.Workbooks.Open(os.path.abspath(self.main_file))
            dest_ws = dest_wb.Worksheets("NBD-MF-10-GA")
            
            # Column mapping: source -> destination (Row 2 to Row 16)
            column_map = {
                25: 3,  # Y2 -> C16
                26: 4,  # Z2 -> D16
                27: 5,  # AA2 -> E16
                28: 6,  # AB2 -> F16
                29: 7,  # AC2 -> G16
                30: 8,  # AD2 -> H16
                31: 9,  # AE2 -> I16
                32: 10  # AF2 -> J16
            }
            
            self.log_detailed("COLUMN MAPPING (Source Row 2 -> Destination Row 16):")
            for src_col, dest_col in column_map.items():
                src_letter = self.get_column_letter(src_col)
                dest_letter = self.get_column_letter(dest_col)
                self.log_detailed(f"  Source {src_letter}2 -> Destination {dest_letter}16")
            
            self.log_detailed("COPY STRATEGY: Clear ALL existing formulas and content, then paste values only")
            
            copied_count = 0
            for src_col, dest_col in column_map.items():
                try:
                    source_value = source_ws.Cells(2, src_col).Value
                    if source_value is not None:
                        dest_cell = dest_ws.Cells(16, dest_col)
                        old_formula = str(dest_cell.Formula)
                        old_value = dest_cell.Value
                        
                        # Clear formula completely for ALL cells before pasting values
                        dest_cell.ClearContents()  # Clear formula and content
                        dest_cell.Value = source_value  # Set only the value
                        
                        col_letter = self.get_column_letter(src_col)
                        dest_letter = self.get_column_letter(dest_col)
                        self.log(f"FD Base: Cleared {dest_letter}16 formula and copied {col_letter}2 ({source_value}) to {dest_letter}16")
                        self.log_detailed(f"COPIED: {col_letter}2 -> {dest_letter}16: '{source_value}'")
                        self.log_detailed(f"  Old formula: '{old_formula}'")
                        self.log_detailed(f"  Old value: '{old_value}'")
                        self.log_detailed(f"  New value: '{source_value}'")
                        
                        copied_count += 1
                    else:
                        col_letter = self.get_column_letter(src_col)
                        dest_letter = self.get_column_letter(dest_col)
                        self.log_detailed(f"SKIPPED: {col_letter}2 -> {dest_letter}16: Source value is None")
                except Exception as e:
                    col_letter = self.get_column_letter(src_col)
                    dest_letter = self.get_column_letter(dest_col)
                    self.log(f"FD Base: Error copying column {src_col}: {e}")
                    self.log_detailed(f"ERROR: {col_letter}2 -> {dest_letter}16 processing failed: {e}", "ERROR")
                    continue
            
            self.log(f"FD Base: Processed {copied_count} cells from Portfolio sheet (ALL formulas cleared before pasting)")
            self.log_detailed(f"TOTAL CELLS PROCESSED: {copied_count}")
            self.log_detailed(f"FORMULAS CLEARED: ALL existing formulas in destination row 16 were cleared")
            self.log_detailed(f"VALUES PASTED: Only values were pasted, no formulas")
            
            dest_wb.Save()
            self.log_detailed(f"SAVED: Destination file {self.main_file}")
            source_wb.Close(SaveChanges=False)
            dest_wb.Close(SaveChanges=False)
            self.log_detailed("=== FD BASE PORTFOLIO DATA TRANSFER COMPLETED ===")
            
            return True
            
        except Exception as e:
            self.log(f"FD Base Portfolio transfer error: {e}")
            self.log_detailed(f"CRITICAL ERROR in FD Base Portfolio transfer: {e}", "ERROR")
            traceback.print_exc()
            return False
    
    def get_column_letter(self, col_num):
        """Convert column number to letter (1=A, 26=Z, 27=AA, etc.)"""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(65 + col_num % 26) + result
            col_num //= 26
        return result
    
    def add_o16_to_c16_if_present(self):
        """
        UPDATED FEATURE: Clear C16, set base formula, copy remaining values, then adjust C16 based on O16
        CORRECTED PROCESS:
        1. Clear C16 completely (formula and value)
        2. Set C16 to base formula with Y2 value: =Y2_value
        3. Copy remaining values from Z2:AF2 to D16:J16 (values only)
        4. Check O16 value and adjust C16 formula accordingly:
           - O16 negative: C16 = Y2_value + |O16_value|
           - O16 positive: C16 = Y2_value - O16_value
           - O16 empty: C16 = Y2_value (no change)
        """
        self.log("Processing C16 formula and copying remaining values...")
        self.log_detailed("=== C16 FORMULA AND REMAINING VALUES PROCESSING STARTED ===")
        
        try:
            self.log(f"[INFO] Opening main destination for C16 processing: {self.main_file} | sheet: NBD-MF-10-GA")
            self.log_detailed(f"DESTINATION FILE: {self.main_file}")
            self.log_detailed(f"DESTINATION SHEET: NBD-MF-10-GA")
            dest_wb = self.excel_app.Workbooks.Open(os.path.abspath(self.main_file))
            dest_ws = dest_wb.Worksheets("NBD-MF-10-GA")
            
            # Get Y2 value (base value from FD Base Portfolio)
            y2_value = dest_ws.Cells(2, 25).Value  # Y2
            o16_value = dest_ws.Cells(16, 15).Value  # O16
            c16_cell = dest_ws.Cells(16, 3)         # C16 cell
            c16_current_value = c16_cell.Value      # C16 current value
            c16_current_formula = str(c16_cell.Formula)  # C16 current formula
            
            self.log_detailed(f"Y2 VALUE (Base): {y2_value}")
            self.log_detailed(f"O16 VALUE (Before): {o16_value}")
            self.log_detailed(f"C16 CURRENT VALUE: {c16_current_value}")
            self.log_detailed(f"C16 CURRENT FORMULA: {c16_current_formula}")
            
            # Determine base numeric: prefer Y2; if Y2 missing, use current C16 value
            base_numeric = None
            base_source = None
            try:
                if y2_value is not None:
                    base_numeric = float(y2_value)
                    base_source = "Y2"
                else:
                    # Fallback to C16 current value
                    if c16_current_value is not None:
                        base_numeric = float(c16_current_value)
                        base_source = "C16"
            except (ValueError, TypeError):
                base_numeric = None

            if base_numeric is not None:
                try:
                    self.log_detailed(f"BASE NUMERIC VALUE ({base_source}): {base_numeric}")
                    
                    # Step 1: Clear C16 completely
                    c16_cell.ClearContents()
                    self.log_detailed("STEP 1: C16 CLEARED - Removed formula and value")
                    
                    # Step 2: Set C16 to base formula with Y2 value
                    base_formula = f"={base_numeric}"
                    c16_cell.Formula = base_formula
                    c16_new_value = c16_cell.Value
                    self.log(f"NBD-MF-10-GA: Set C16 base formula to '{base_formula}'")
                    self.log_detailed(f"STEP 2: C16 BASE FORMULA SET: {base_formula}")
                    self.log_detailed(f"C16 NEW VALUE: {c16_new_value}")
                    
                    # Step 3: Copy remaining values from Z2:AF2 to D16:J16
                    # Column mapping: source -> destination (Row 2 to Row 16)
                    remaining_column_map = {
                        26: 4,  # Z2 -> D16
                        27: 5,  # AA2 -> E16
                        28: 6,  # AB2 -> F16
                        29: 7,  # AC2 -> G16
                        30: 8,  # AD2 -> H16
                        31: 9,  # AE2 -> I16
                        32: 10  # AF2 -> J16
                    }
                    
                    self.log_detailed("STEP 3: COPYING REMAINING VALUES (Z2:AF2 -> D16:J16)")
                    self.log_detailed("COLUMN MAPPING (Source Row 2 -> Destination Row 16):")
                    for src_col, dest_col in remaining_column_map.items():
                        src_letter = self.get_column_letter(src_col)
                        dest_letter = self.get_column_letter(dest_col)
                        self.log_detailed(f"  Source {src_letter}2 -> Destination {dest_letter}16")
                    
                    copied_count = 0
                    for src_col, dest_col in remaining_column_map.items():
                        try:
                            source_value = dest_ws.Cells(2, src_col).Value
                            if source_value is not None:
                                dest_cell = dest_ws.Cells(16, dest_col)
                                old_value = dest_cell.Value
                                old_formula = str(dest_cell.Formula)
                                
                                # Clear any existing formula and set value only
                                dest_cell.ClearContents()
                                dest_cell.Value = source_value
                                
                                src_letter = self.get_column_letter(src_col)
                                dest_letter = self.get_column_letter(dest_col)
                                self.log_detailed(f"COPIED: {src_letter}2 -> {dest_letter}16: '{source_value}'")
                                self.log_detailed(f"  Old value: '{old_value}'")
                                self.log_detailed(f"  Old formula: '{old_formula}'")
                                self.log_detailed(f"  New value: '{source_value}'")
                                
                                copied_count += 1
                            else:
                                src_letter = self.get_column_letter(src_col)
                                dest_letter = self.get_column_letter(dest_col)
                                self.log_detailed(f"SKIPPED: {src_letter}2 -> {dest_letter}16: Source value is None")
                        except Exception as e:
                            src_letter = self.get_column_letter(src_col)
                            dest_letter = self.get_column_letter(dest_col)
                            self.log_detailed(f"ERROR: {src_letter}2 -> {dest_letter}16 processing failed: {e}", "ERROR")
                            continue
                    
                    self.log_detailed(f"STEP 3 COMPLETED: Copied {copied_count} remaining values")

                    # From this point forward, use the current C16 value as the base
                    # (do not use Y2 anymore for adjustments)
                    try:
                        current_c16_numeric = float(c16_cell.Value) if c16_cell.Value is not None else None
                        self.log_detailed(f"ADJUSTMENT BASE (from current C16): {current_c16_numeric}")
                    except (ValueError, TypeError):
                        current_c16_numeric = None
                    
                    # Step 4: Check O16 value and adjust C16 formula accordingly
                    self.log_detailed("STEP 4: CHECKING O16 VALUE AND ADJUSTING C16 FORMULA")
                    
                    if o16_value is not None and str(o16_value).strip() != "" and str(o16_value).strip() != "-":
                        try:
                            # Robust parse for values like (21.00), 21.00, "21,000.50", etc.
                            o16_numeric = self._parse_numeric(o16_value)
                            if o16_numeric is None:
                                raise ValueError("O16 not numeric after parsing")
                            self.log_detailed(f"O16 NUMERIC VALUE: {o16_numeric}")
                            
                            # Choose base for adjustment: prefer current C16 numeric, else fallback to base_numeric
                            adj_base = current_c16_numeric if current_c16_numeric is not None else base_numeric

                            if o16_numeric < 0:
                                # O16 is negative: Add absolute value of O16 to Y2
                                abs_o16 = abs(o16_numeric)
                                final_formula = f"={adj_base}+{abs_o16}"
                                self.log_detailed(f"O16 IS NEGATIVE: Adding absolute value {abs_o16} to base {adj_base}")
                                self.log_detailed(f"FINAL FORMULA: {final_formula}")
                            else:
                                # O16 is positive: Subtract O16 value from Y2
                                final_formula = f"={adj_base}-{o16_numeric}"
                                self.log_detailed(f"O16 IS POSITIVE: Subtracting {o16_numeric} from base {adj_base}")
                                self.log_detailed(f"FINAL FORMULA: {final_formula}")
                            
                            # Set the final formula in C16
                            c16_cell.Formula = final_formula
                            c16_final_value = c16_cell.Value
                            self.log(f"NBD-MF-10-GA: Updated C16 formula to '{final_formula}'")
                            self.log_detailed(f"C16 FINAL VALUE: {c16_final_value}")
                            
                        except (ValueError, TypeError):
                            # O16 is not numeric, keep base formula
                            self.log_detailed(f"O16 IS NOT NUMERIC: Keeping base formula '{base_formula}'")
                    else:
                        # O16 is empty or "-": Keep base formula
                        self.log_detailed(f"O16 IS EMPTY: Keeping base formula '{base_formula}'")
                    
                    self.log_detailed("PROCESSING SUMMARY:")
                    self.log_detailed(f"  C16: Cleared and set to base formula '{base_formula}'")
                    self.log_detailed(f"  D16-J16: Copied {copied_count} values from Z2-AF2")
                    self.log_detailed(f"  C16: Final formula adjusted based on O16 value")
                    self.log_detailed(f"  O16: Display handled automatically by Excel sheet")
                        
                except (ValueError, TypeError) as e:
                    self.log(f"NBD-MF-10-GA: Base value ({base_source}) is not numeric - skipped: {e}")
                    self.log_detailed(f"ERROR: Base value from {base_source} is not numeric: {e}", "ERROR")
            else:
                self.log("NBD-MF-10-GA: Neither Y2 nor C16 contains a numeric base - cannot process C16")
                self.log_detailed("ERROR: Missing numeric base from Y2 and C16", "ERROR")
            
            dest_wb.Save()
            self.log_detailed(f"SAVED: Destination file {self.main_file}")
            dest_wb.Close(SaveChanges=False)
            self.log_detailed("=== C16 FORMULA AND REMAINING VALUES PROCESSING COMPLETED ===")
            
            return True
            
        except Exception as e:
            self.log(f"C16 processing error: {e}")
            self.log_detailed(f"CRITICAL ERROR in C16 processing: {e}", "ERROR")
            traceback.print_exc()
            return False
    
    def copy_borrowings_j_to_m(self):
        """
        NEW FEATURE: Copy values from J3:J9 to M3:M9 and M13 to M14 in Borrowings sheet of main file
        - Copy J3 to M3, J4 to M4, J5 to M5, J6 to M6, J7 to M7, J8 to M8, J9 to M9
        - Copy M13 to M14
        - Clear any existing formulas in destination cells before pasting values
        """
        self.log("Starting Borrowings J3:J9 to M3:M9 and M13 to M14 data transfer...")
        
        try:
            self.log(f"[INFO] Opening main destination for Borrowings J→M copy: {self.main_file} | sheet: Borrowings")
            dest_wb = self.excel_app.Workbooks.Open(os.path.abspath(self.main_file))
            dest_ws = dest_wb.Worksheets("Borrowings")
            
            copied_count = 0
            # Copy J3:J9 to M3:M9
            for row in range(3, 10):  # Rows 3 to 9
                try:
                    source_value = dest_ws.Cells(row, 10).Value  # Column J (10)
                    if source_value is not None and str(source_value).strip() != "":
                        dest_cell = dest_ws.Cells(row, 13)  # Column M (13)
                        dest_cell.ClearContents()  # Clear any existing formula
                        dest_cell.Value = source_value  # Set only the value
                        copied_count += 1
                        self.log(f"Borrowings: Copied J{row} ({source_value}) to M{row}")
                    else:
                        self.log(f"Borrowings: Skipped J{row} (empty or blank)")
                except Exception as e:
                    self.log(f"Borrowings: Error copying J{row} to M{row}: {e}")
                    continue
            
            # Copy M13 to M14
            try:
                source_value = dest_ws.Cells(13, 13).Value  # M13
                if source_value is not None and str(source_value).strip() != "":
                    dest_cell = dest_ws.Cells(14, 13)  # M14
                    dest_cell.ClearContents()  # Clear any existing formula
                    dest_cell.Value = source_value  # Set only the value
                    copied_count += 1
                    self.log(f"Borrowings: Copied M13 ({source_value}) to M14")
                else:
                    self.log(f"Borrowings: Skipped M13 (empty or blank)")
            except Exception as e:
                self.log(f"Borrowings: Error copying M13 to M14: {e}")
            
            self.log(f"Borrowings: Copied {copied_count}/8 cells (J3:J9 to M3:M9 and M13 to M14)")
            
            dest_wb.Save()
            dest_wb.Close(SaveChanges=False)
            
            return True
            
        except Exception as e:
            self.log(f"Borrowings J3:J9 to M3:M9 and M13 to M14 transfer error: {e}")
            traceback.print_exc()
            return False
    
    def fix_c17_formulas(self):
        """
        NEW FEATURE: Detect and fix C17 formulas in NBD-MF-10-GA sheet
        - If formula is like ='NBD-MF-01-SOFP'!D93+Borrowings!N3+1, correct to ='NBD-MF-01-SOFP'!D93+Borrowings!N3-1
        - For constants > +1 (e.g., +5), apply same correction (e.g., to -5) and log warning
        """
        self.log("Detecting and fixing C17 formulas in NBD-MF-10-GA sheet...")
        
        try:
            dest_wb = self.excel_app.Workbooks.Open(os.path.abspath(self.main_file))
            dest_ws = dest_wb.Worksheets("NBD-MF-10-GA")
            
            c17_cell = dest_ws.Cells(17, 3)  # C17
            c17_formula = str(c17_cell.Formula)
            
            self.log(f"NBD-MF-10-GA: C17 formula: {c17_formula}")
            
            if c17_formula.startswith('='):
                # Look for +number or -number at the end of the formula
                match = re.search(r'([+-])(\d+)$', c17_formula)
                if match:
                    sign = match.group(1)
                    number = int(match.group(2))
                    
                    if sign == '+' and number >= 1:
                        new_number = -number
                        new_formula = re.sub(r'[+-]\d+$', f'{new_number}', c17_formula)
                        c17_cell.Formula = new_formula
                        self.log(f"NBD-MF-10-GA: Changed C17 formula from '{c17_formula}' to '{new_formula}'")
                        
                        if number > 1:
                            self.log(f"WARNING: Serious problem detected - C17 formula contained +{number} (corrected to {new_number})")
                        
                        dest_wb.Save()
                        dest_wb.Close(SaveChanges=False)
                        return True
                    elif sign == '-' and number >= 1:
                        self.log(f"NBD-MF-10-GA: C17 formula already has -{number}, no correction needed")
                        dest_wb.Close(SaveChanges=False)
                        return True
                    else:
                        self.log("NBD-MF-10-GA: C17 formula does not end with a valid +number or -number")
                        dest_wb.Close(SaveChanges=False)
                        return False
                else:
                    self.log("NBD-MF-10-GA: C17 formula does not match expected pattern (e.g., +1 or -1 at end)")
                    dest_wb.Close(SaveChanges=False)
                    return False
            else:
                self.log("NBD-MF-10-GA: C17 does not contain a formula")
                dest_wb.Close(SaveChanges=False)
                return False
                
        except Exception as e:
            self.log(f"NBD-MF-10-GA: C17 formula fix error: {e}")
            dest_wb.Close(SaveChanges=False) if 'dest_wb' in locals() else None
            traceback.print_exc()
            return False
    
    def run_precise_transfer(self):
        """Run precise data transfer - only specified cells, preserve everything else."""
        self.log("="*60)
        self.log("🚀 PRECISE Excel Data Transfer - Values Only")
        self.log("="*60)
        self.log_detailed("="*80)
        self.log_detailed("PRECISE EXCEL DATA TRANSFER - DETAILED LOGGING STARTED")
        self.log_detailed("="*80)
        self.log_detailed(f"START TIME: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self.log_detailed(f"WORKING DIRECTORY: {os.getcwd()}")
        
        try:
            # New pre-step: detect files by keywords supporting Excel formats before existence check
            self.log("🔍 STEP 0: Detecting input files by keywords...")
            self.log_detailed("STEP 0: FILE DETECTION - Starting keyword-based file detection")
            self.detect_input_files()
            self.log("✅ STEP 0: File detection completed")
            self.log_detailed("STEP 0: FILE DETECTION - Completed")
            
            self.log("📋 STEP 1: Verifying all required files exist...")
            self.log_detailed("STEP 1: FILE EXISTENCE CHECK - Verifying all required files exist")
            if not self.check_files_exist():
                self.log("❌ STEP 1: File verification failed - missing required files")
                self.log_detailed("STEP 1: FILE EXISTENCE CHECK - FAILED - Missing required files", "ERROR")
                return False
            self.log("✅ STEP 1: All required files found")
            self.log_detailed("STEP 1: FILE EXISTENCE CHECK - PASSED - All files found")
            
            temp_files = [f for f in os.listdir('.') if f.startswith('~') and f.endswith(('.xlsx', '.xlsb', '.xlsm'))]
            if temp_files:
                self.log("❌ ERROR: Excel files are open. Please close all Excel applications first.")
                self.log_detailed(f"STEP 2: TEMP FILE CHECK - FAILED - Found temp files: {temp_files}", "ERROR")
                return False
            self.log("✅ STEP 2: No Excel files are currently open")
            self.log_detailed("STEP 2: TEMP FILE CHECK - PASSED - No temp files found")
            
            self.log("⚙️ STEP 3: Starting Excel application...")
            self.log_detailed("STEP 3: EXCEL APPLICATION - Starting Excel application")
            if not self.start_excel():
                self.log("❌ STEP 3: Failed to start Excel application")
                self.log_detailed("STEP 3: EXCEL APPLICATION - FAILED - Could not start Excel", "ERROR")
                return False
            self.log("✅ STEP 3: Excel application started successfully")
            self.log_detailed("STEP 3: EXCEL APPLICATION - SUCCESS - Excel started")
            
            try:
                success_count = 0
                
                self.log("\n--- Step 1: SOFP Data Transfer (Values Only) ---")
                self.log_detailed("STEP 4: SOFP DATA TRANSFER - Starting SOFP values transfer")
                if self.copy_sofp_values_only():
                    success_count += 1
                    self.log("SUCCESS: SOFP values transferred")
                    self.log_detailed("STEP 4: SOFP DATA TRANSFER - SUCCESS - Values transferred")
                else:
                    self.log("FAILED: SOFP transfer failed")
                    self.log_detailed("STEP 4: SOFP DATA TRANSFER - FAILED - Transfer failed", "ERROR")
                
                self.log("\n--- Step 2: System TB Data Transfer (Values Only) ---")
                self.log_detailed("STEP 5: SYSTEM TB DATA TRANSFER - Starting System TB values transfer")
                if self.copy_system_tb_values_only():
                    success_count += 1
                    self.log("SUCCESS: System TB values transferred")
                    self.log_detailed("STEP 5: SYSTEM TB DATA TRANSFER - SUCCESS - Values transferred")
                else:
                    self.log("FAILED: System TB transfer failed")
                    self.log_detailed("STEP 5: SYSTEM TB DATA TRANSFER - FAILED - Transfer failed", "ERROR")
                
                self.log("\n--- Step 3: NBD-MF-10-GA Verification ---")
                self.log_detailed("STEP 6: NBD-MF-10-GA VERIFICATION - Verifying formulas C6=N6 and C7=N7")
                if self.verify_nbdmf10_formulas():
                    self.log("SUCCESS: NBD-MF-10-GA verification passed")
                    self.log_detailed("STEP 6: NBD-MF-10-GA VERIFICATION - SUCCESS - Formulas verified")
                else:
                    self.log("WARNING: NBD-MF-10-GA verification failed")
                    self.log_detailed("STEP 6: NBD-MF-10-GA VERIFICATION - WARNING - Formula verification failed", "WARNING")
                
                self.log("\n--- Step 4: Investment Data Transfer (Values Only) ---")
                self.log_detailed("STEP 7: INVESTMENT ERROR CLEANUP - Cleaning #VALUE! and other errors")
                if self.clean_investment_sheet_errors():
                    self.log("SUCCESS: Investment sheet errors cleaned")
                    self.log_detailed("STEP 7: INVESTMENT ERROR CLEANUP - SUCCESS - Errors cleaned")
                    self.log_detailed("STEP 8: INVESTMENT DATA TRANSFER - Starting Investment values transfer")
                    if self.copy_investment_values_only():
                        success_count += 1
                        self.log("SUCCESS: Investment values transferred")
                        self.log_detailed("STEP 8: INVESTMENT DATA TRANSFER - SUCCESS - Values transferred")
                    else:
                        self.log("FAILED: Investment transfer failed")
                        self.log_detailed("STEP 8: INVESTMENT DATA TRANSFER - FAILED - Transfer failed", "ERROR")
                else:
                    self.log("FAILED: Investment error cleanup failed")
                    self.log_detailed("STEP 7: INVESTMENT ERROR CLEANUP - FAILED - Error cleanup failed", "ERROR")
                
                self.log("\n--- Step 5: Investment Specific Cells Transfer ---")
                if self.copy_investment_specific_cells():
                    success_count += 1
                    self.log("SUCCESS: Investment specific cells transferred")
                else:
                    self.log("FAILED: Investment specific cells transfer failed")
                
                self.log("\n--- Step 6: Other Assets Manual Values Setting ---")
                if self.set_other_assets_manual_values():
                    success_count += 1
                    self.log("SUCCESS: Other Assets manual values set (F131=111,159,175, G131=63,515,193, H131=66,403,345, I131=-28,547,383, J131=-4,470,936, K131=-35,180,520)")
                else:
                    self.log("FAILED: Other Assets manual values setting failed")
                
                self.log("\n--- Step 7: Deposits Process - FD Base Portfolio Data Transfer ---")
                if self.copy_fd_base_portfolio_data():
                    success_count += 1
                    self.log("SUCCESS: FD Base Portfolio data transferred")
                else:
                    self.log("FAILED: FD Base Portfolio transfer failed")
                
                self.log("\n--- Step 8: Fund Back Difference - Add O16 to C16 if Present ---")
                if self.add_o16_to_c16_if_present():
                    success_count += 1
                    self.log("SUCCESS: O16 to C16 addition completed")
                else:
                    self.log("FAILED: O16 to C16 addition failed")
                
                self.log("\n--- Step 9: Borrowing Data Transfer ---")
                if self.copy_borrowing_data():
                    success_count += 1
                    self.log("SUCCESS: Borrowing data transferred")
                else:
                    self.log("FAILED: Borrowing data transfer failed")
                
                self.log("\n--- Step 10: Borrowing Amount Verification ---")
                if self.verify_borrowing_amounts():
                    self.log("SUCCESS: Borrowing amounts verification passed")
                else:
                    self.log("WARNING: Borrowing amounts verification failed")
                
                self.log("\n--- Step 11: Refresh Borrowings Pivot Tables ---")
                if self.refresh_borrowings_pivot_tables():
                    success_count += 1
                    self.log("SUCCESS: Borrowings pivot tables refreshed")
                else:
                    self.log("FAILED: Borrowings pivot tables refresh failed")
                
                self.log("\n--- Step 12: C17 Formula Fix in NBD-MF-10-GA ---")
                if self.fix_c17_formulas():
                    success_count += 1
                    self.log("SUCCESS: C17 formulas fixed")
                else:
                    self.log("FAILED: C17 formula fix failed")
                
                self.log("\n--- Step 13: Borrowings J3:J9 to M3:M9 Transfer ---")
                if self.copy_borrowings_j_to_m():
                    success_count += 1
                    self.log("SUCCESS: Borrowings J3:J9 to M3:M9 transferred")
                else:
                    self.log("FAILED: Borrowings J3:J9 to M3:M9 transfer failed")
                
                self.log("="*60)
                self.log(f"PRECISE TRANSFER COMPLETED: {success_count}/11 data transfers successful")
                self.log("ALL OTHER SHEETS AND FORMULAS PRESERVED EXACTLY")
                self.log("="*60)
                
                # Final detailed logging summary
                self.log_detailed("="*80)
                self.log_detailed("FINAL TRANSFER SUMMARY")
                self.log_detailed("="*80)
                self.log_detailed(f"TOTAL SUCCESSFUL TRANSFERS: {success_count}/11")
                self.log_detailed(f"TRANSFER SUCCESS RATE: {(success_count/11)*100:.1f}%")
                self.log_detailed(f"END TIME: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                self.log_detailed("PRESERVED: All other sheets and formulas kept intact")
                self.log_detailed("STRATEGY: Values only - no formulas copied from source")
                self.log_detailed("="*80)
                self.log_detailed("PRECISE EXCEL DATA TRANSFER - DETAILED LOGGING COMPLETED")
                self.log_detailed("="*80)
                
                return success_count >= 9
                
            finally:
                self.log_detailed("CLEANUP: Closing Excel application")
                self.close_excel()
                self.log_detailed("CLEANUP: Excel application closed")
                
        except Exception as e:
            self.log(f"Critical error: {e}")
            self.log_detailed(f"CRITICAL ERROR in main process: {e}", "ERROR")
            traceback.print_exc()
            self.log_detailed("CLEANUP: Closing Excel application due to error")
            self.close_excel()
            return False

def main():
    """Main execution function."""
    # Step 1: Validate all required files exist
    print("Starting NBD-MF-10-GA & NBD-MF-11-IS Processing Pipeline")
    print("="*60)
    
    validation_success, found_files = validate_required_files()
    
    if not validation_success:
        print("❌ Cannot proceed - required files are missing!")
        print("Please ensure all required files are present and try again.")
        return False
    
    print("✅ All required files found. Proceeding with processing...")
    print()
    
    # Step 2: Initialize processor and run transfer
    processor = PreciseExcelDataTransfer()
    
    # Verify all required files were found
    if not all([processor.main_file, processor.sofp_source, processor.investment_source, 
                processor.fd_base_file, processor.borrowing_source]):
        print("❌ Some required files are missing!")
        print("Missing files:")
        if not processor.main_file:
            print("   - NBD-MF-10-GA & NBD-MF-11-IS")
        if not processor.sofp_source:
            print("   - NBD-MF-01-SOFP & SOCI AFL Monthly FS")
        if not processor.investment_source:
            print("   - Investment Schedule")
        if not processor.fd_base_file:
            print("   - FD Base as at")
        if not processor.borrowing_source:
            print("   - Borrowing report")
        return False
    
    try:
        success = processor.run_precise_transfer()
        
        if success:
            print("\n" + "="*70)
            print("🎉 SUCCESS! Precise data transfer completed successfully!")
            print("="*70)
            print("✅ Completed Operations:")
            print("   • Only specified cells updated with values")
            print("   • All formulas in destination files preserved")
            print("   • All other sheets completely untouched")
            print("   • File structure and formatting maintained")
            print("   • Investment specific cells transferred")
            print("   • Other Assets manual values set")
            print("   • Deposits Process - FD Base Portfolio data transferred")
            print("   • Fund Back Difference - O16 to C16 addition performed")
            print("   • Borrowing data transferred (.xlsx/.xlsb/.xlsm supported)")
            print("   • Borrowing amounts verified")
            print("   • Borrowings pivot tables refreshed")
            print("   • C17 formulas fixed in NBD-MF-10-GA")
            print("   • Borrowings J3:J9 to M3:M9 transferred")
            print("="*70)
        else:
            print("\n" + "="*70)
            print("❌ FAILED! Precise data transfer did not complete successfully.")
            print("📋 Check log messages above for details.")
            print("="*70)
            
    except Exception as e:
        print(f"💥 Unhandled exception: {e}")
        traceback.print_exc()

if __name__ == "__main__":
    main()

# Configure logging
def setup_logging():
    """Set up logging to file and console with detailed formatting."""
    # Use the same log file as the main process
    script_dir = os.path.dirname(os.path.abspath(__file__))
    now = datetime.now()
    time_part = now.strftime("%H%M%S")
    date_part = now.strftime("%Y%m%d")
    log_filename = os.path.join(script_dir, f"NBD-MF-10-GA & NBD-MF-11-IS_LOG_{time_part}_{date_part}.log")
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(message)s',
        handlers=[
            logging.FileHandler(log_filename, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )
    logger = logging.getLogger()
    logger.info("=== MATCAP PROCESSING LOGGING INITIALIZED ===")
    return logger

def load_cbsl_mapping(cbsl_file_path, logger):
    """Load CBSL mapping as two dictionaries for faster lookup."""
    logger.info(f"Loading CBSL mapping from: {cbsl_file_path}")
    start_time_local = datetime.now()
    
    try:
        # Use openpyxl directly for faster loading without pandas
        wb = openpyxl.load_workbook(cbsl_file_path, read_only=True, data_only=True)
        ws = wb['Portfolio']
        
        # Create dictionaries for fast lookup
        contract_to_npp = {}
        contract_to_product = {}
        
        # Process rows (starting from row 3 as per your note)
        for row in ws.iter_rows(min_row=3, min_col=1, max_col=37, values_only=True):  # A to AK
            contract_no = row[0]  # Column A
            npp_value = row[36]   # Column AK (index 36 as A=0, B=1, ..., AK=36)
            
            if contract_no is None:
                continue
                
            contract_no_str = str(contract_no).strip().upper()
            
            # Store in dictionaries
            contract_to_npp[contract_no_str] = str(npp_value).strip() if npp_value is not None else "N/A"
            
            # Extract product from contract number (positions 2-3)
            if len(contract_no_str) >= 4:
                product = contract_no_str[2:4]
                contract_to_product[contract_no_str] = product
        
        logger.info(f"Loaded CBSL mapping with {len(contract_to_npp)} contracts in {datetime.now() - start_time_local}")
        logger.info(f"Sample CBSL entries: {list(contract_to_npp.items())[:5]}")
        
        wb.close()
        return contract_to_npp, contract_to_product
        
    except Exception as e:
        logger.error(f"Failed to load CBSL mapping: {str(e)}")
        raise

def process_matcap_file_optimized(args):
    current_file_path, output_file_path, cbsl_file_path = args
    # Use the main logger instead of creating a separate one
    process_logger = logging.getLogger()
    
    process_logger.info(f"STARTING MATCAP PROCESSING: {os.path.basename(current_file_path)}")
    overall_start_time = datetime.now()
    
    try:
        # Step 1: Load CBSL mapping as dictionaries for faster lookup
        process_logger.info("STEP 1: Loading CBSL mapping data...")
        contract_to_npp, contract_to_product = load_cbsl_mapping(cbsl_file_path, process_logger)
        
        # Debug: Check if CBSL dict has data
        process_logger.info(f"CBSL NP/P dictionary loaded: {len(contract_to_npp)} contracts")
        process_logger.info(f"CBSL Product dictionary loaded: {len(contract_to_product)} products")
        if len(contract_to_npp) == 0:
            process_logger.warning("CBSL dictionary is empty! This will cause all NP/P values to be N/A")
        
        # Step 2: Create copy of current month file
        process_logger.info("STEP 2: Creating working copy of MATCAP file...")
        if os.path.exists(output_file_path):
            process_logger.warning(f"Output file {output_file_path} already exists, removing it")
            os.remove(output_file_path)
        shutil.copy2(current_file_path, output_file_path)
        process_logger.info(f"Created copy: {output_file_path}")
        
        # Step 3: Direct Excel processing without pandas
        process_logger.info("STEP 3: Processing Excel data structure...")
        step_start_time = datetime.now()
        
        wb = openpyxl.load_workbook(output_file_path)
        ws = wb['MATCAP-01-08-2025']
        
        # Get total rows and monthly column count before modifications
        max_row = ws.max_row
        max_col = ws.max_column
        process_logger.info(f"Original dimensions: {max_row} rows, {max_col} columns")
        
        # Step 4: Delete original Product column (B) and insert 11 new columns
        process_logger.info("STEP 4: Restructuring columns (delete Product, add 11 new columns)...")
        ws.delete_cols(2)  # Delete column B (Product)
        ws.insert_cols(2, amount=11)  # Insert 11 columns
        
        # Step 5: Set new headers (B1-L1)
        inserted_cols = [
            "NP/P",                # B
            "Product",             # C
            "Type",                # D
            '7-30 Days (c)',       # E
            '1-3 Months (d)',      # F
            '3-6 Months (e)',      # G
            '6-12 Months (f)',     # H
            '1-3 Years (g)',       # I
            '3-5 Years (h)',       # J
            'Over 5 Years (i)',    # K
            'Total'                # L
        ]
        
        for col_idx, header in enumerate(inserted_cols, start=2):
            ws.cell(row=1, column=col_idx).value = header
        
        # Fix the first column header to match your sample
        ws.cell(row=1, column=1).value = "CON_NO"
        
        process_logger.info(f"Headers set in {datetime.now() - step_start_time}")
        
        # Step 6: Process data row by row for memory efficiency
        process_logger.info("STEP 6: Processing data rows and applying CBSL mapping...")
        step_start_time = datetime.now()
        
        rows_to_delete = []
        processed_rows = 0
        matched_contracts = 0
        
        # Monthly columns now start at column M (13)
        monthly_start_col = 13
        monthly_end_col = ws.max_column
        
        for row in range(2, max_row + 1):
            contract_no = ws.cell(row=row, column=1).value
            
            if contract_no is None:
                rows_to_delete.append(row)
                continue
            
            contract_no_str = str(contract_no).strip().upper()
            
            # Calculate monthly sum for this row (from monthly columns)
            monthly_sum = 0
            for col in range(monthly_start_col, monthly_end_col + 1):
                cell_value = ws.cell(row=row, column=col).value
                if isinstance(cell_value, (int, float)) and cell_value is not None:
                    monthly_sum += cell_value
            
            # Get Product from contract number (positions 2-3)
            product = contract_no_str[2:4] if len(contract_no_str) >= 4 else ''
            
            # Get NP/P from dictionary
            np_p_value = contract_to_npp.get(contract_no_str, "N/A")
            
            # Determine Type
            type_value = "Leases" if product.upper() in ['LE', 'UV'] else "Loans"
            
            if np_p_value != "N/A":
                matched_contracts += 1
            
            # Debug specific contract if needed
            if row <= 10:  # Log first 10 rows for debugging
                process_logger.info(f"Row {row}: Contract='{contract_no_str}', NP/P='{np_p_value}', MonthlySum={monthly_sum}")
            
            # Mark for deletion if monthly sum is 0 and NP/P is N/A
            if monthly_sum == 0 and np_p_value == "N/A":
                rows_to_delete.append(row)
                continue
            
            # Set the processed values
            ws.cell(row=row, column=2).value = np_p_value  # NP/P
            ws.cell(row=row, column=3).value = product     # Product
            ws.cell(row=row, column=4).value = type_value  # Type
            
            processed_rows += 1
        
        process_logger.info(f"Processed {processed_rows} rows, {matched_contracts} contracts matched CBSL, marked {len(rows_to_delete)} for deletion in {datetime.now() - step_start_time}")
        
        # Step 7: Delete rows with zero sum and N/A (from bottom to top to maintain row numbers)
        process_logger.info("STEP 7: Cleaning up empty rows...")
        step_start_time = datetime.now()
        
        for row in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row)
        
        process_logger.info(f"Deleted {len(rows_to_delete)} rows in {datetime.now() - step_start_time}")
        
        # Step 8: Add formulas for bucket calculations
        process_logger.info("STEP 8: Adding bucket calculation formulas...")
        step_start_time = datetime.now()
        
        current_max_row = ws.max_row
        
        # Formulas for columns E-L (monthly data starting at M)
        formulas = {
            5:  '=IFERROR(IF(ISNUMBER(N{row}), N{row}, 0), 0)',   # Column E
            6:  '=IFERROR(SUM(O{row}:P{row}), 0)',                # Column F
            7:  '=IFERROR(SUM(Q{row}:S{row}), 0)',                # Column G
            8:  '=IFERROR(SUM(T{row}:Y{row}), 0)',                # Column H
            9:  '=IFERROR(SUM(Z{row}:AW{row}), 0)',               # Column I
            10: '=IFERROR(SUM(AX{row}:BU{row}), 0)',              # Column J
            11: '=IFERROR(SUM(BV{row}:CZ{row}), 0)',              # Column K
            12: '=IFERROR(E{row}+F{row}+G{row}+H{row}+I{row}+J{row}+K{row}, 0)'  # Column L total
        }
        
        for col, formula_template in formulas.items():
            for row_num in range(2, current_max_row + 1):
                formula = formula_template.format(row=row_num)
                ws.cell(row=row_num, column=col).value = formula
        
        process_logger.info(f"Added bucket formulas in {datetime.now() - step_start_time}")
        
        # Step 9: Add sum formulas for all columns at the bottom
        process_logger.info("STEP 9: Adding total sum formulas...")
        step_start_time = datetime.now()
        
        # Add blank row and sum row
        blank_row = current_max_row + 1
        sum_row = blank_row + 1
        
        # Add "Total" label
        ws.cell(row=sum_row, column=1).value = "Total"
        
        # Sum formulas for ALL monthly columns (M onwards)
        monthly_start_col = 13  # M
        current_max_col = ws.max_column
        
        for col_idx in range(monthly_start_col, current_max_col + 1):
            col_letter = get_column_letter(col_idx)
            sum_formula = f'=SUM({col_letter}2:{col_letter}{current_max_row})'
            ws.cell(row=sum_row, column=col_idx).value = sum_formula
            ws.cell(row=sum_row, column=col_idx).number_format = '#,##0.00;-#,##0.00;"-"'
        
        process_logger.info(f"Added sum formulas for {current_max_col - monthly_start_col + 1} monthly columns in {datetime.now() - step_start_time}")
        
        # Step 10: Apply formatting to match your sample output
        process_logger.info("STEP 10: Applying final formatting...")
        step_start_time = datetime.now()
        
        # Set column widths - wider for bucket columns to prevent ########## display
        for col_idx in range(1, ws.max_column + 1):
            if col_idx >= 5 and col_idx <= 12:  # Bucket columns E-L need more width
                ws.column_dimensions[get_column_letter(col_idx)].width = 15
            else:
                ws.column_dimensions[get_column_letter(col_idx)].width = 12
        
        # Format bucket columns (E-L) to show whole numbers and dashes for zero
        for col_idx in range(5, 13):
            for row_idx in range(2, current_max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = '#,##0;-#,##0;"-"'
        
        # Format monthly columns (M onwards) to show decimals
        for col_idx in range(monthly_start_col, current_max_col + 1):
            for row_idx in range(2, current_max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = '#,##0.00;-#,##0.00;""'
        
        process_logger.info(f"Applied formatting in {datetime.now() - step_start_time}")
        
        # Save the workbook
        wb.save(output_file_path)
        
        # Final verification
        final_rows = ws.max_row - 2  # Exclude header and sum row
        final_cols = ws.max_column
        process_logger.info(f"Final dimensions: {final_rows} data rows, {final_cols} columns")
        process_logger.info(f"Column structure: CON_NO | NP/P | Product | Type | 7 buckets | Total | Monthly columns (starting at M)")
        
        process_logger.info(f"MATCAP PROCESSING COMPLETED SUCCESSFULLY in {datetime.now() - overall_start_time}")
        return True
        
    except Exception as e:
        wb.save(output_file_path)
        process_logger.error(f"MATCAP PROCESSING FAILED: {str(e)}")
        process_logger.error(traceback.format_exc())
        return False

def convert_csv_to_excel_with_processing(args):
    """Convert CSV to Excel with MATCAP processing optimized for new structure."""
    current_csv_path, output_file_path, cbsl_file_path = args
    
    # Create a dedicated logger for this process
    process_logger = logging.getLogger(f"process_{os.getpid()}")
    if not process_logger.handlers:
        process_logger.addHandler(logging.StreamHandler())
        process_logger.setLevel(logging.INFO)
    
    process_logger.info(f"Starting CSV processing for {current_csv_path}")
    overall_start_time = datetime.now()
    
    try:
        # Step 1: Load CBSL mapping as dictionaries for faster lookup
        process_logger.info("Step 1: Loading CBSL mapping")
        contract_to_npp, contract_to_product = load_cbsl_mapping(cbsl_file_path, process_logger)
        
        # Debug: Check if CBSL dict has data
        process_logger.info(f"CBSL NP/P dictionary size: {len(contract_to_npp)}")
        process_logger.info(f"CBSL Product dictionary size: {len(contract_to_product)}")
        if len(contract_to_npp) == 0:
            process_logger.warning("CBSL dictionary is empty! This will cause all NP/P values to be N/A")
        
        # Step 2: Load CSV data
        process_logger.info("Step 2: Loading CSV data")
        df = pd.read_csv(current_csv_path)
        process_logger.info(f"Loaded CSV with {len(df)} rows and {len(df.columns)} columns")
        
        # Step 3: Create new Excel workbook with proper structure
        process_logger.info("Step 3: Creating Excel workbook with new structure")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MATCAP-01-08-2025"
        
        # Step 4: Set up headers with the new structure including blank column M
        headers = [
            "CON_NO",           # A (existing)
            "NP/P",             # B (new)
            "Product",          # C (new) 
            "Type",             # D (new)
            "7-30 Days (c)",    # E (new)
            "1-3 Months (d)",   # F (new)
            "3-6 Months (e)",   # G (new)
            "6-12 Months (f)",  # H (new)
            "1-3 Years (g)",    # I (new)
            "3-5 Years (h)",    # J (new)
            "Over 5 Years (i)", # K (new)
            "Total",            # L (new)
            " "                 # M (new) - BLANK COLUMN
        ]
        
        # Add all monthly columns from original CSV (these will start at column N now)
        original_columns = list(df.columns)
        monthly_columns = original_columns[1:]  # Skip CON_NO, take all monthly columns
        all_headers = headers + monthly_columns
        
        # Write headers to row 1
        for col_idx, header in enumerate(all_headers, 1):
            ws.cell(row=1, column=col_idx).value = header
        
        process_logger.info(f"Headers set: {len(all_headers)} columns total (including blank column M)")
        
        # Step 5: Process data row by row
        process_logger.info("Step 5: Processing data rows")
        rows_to_delete = []
        processed_rows = 0
        matched_contracts = 0
        
        # Monthly data now starts at column N (14) due to blank column M
        monthly_start_col = 14
        
        for idx, row in df.iterrows():
            excel_row = idx + 2  # Excel rows start at 2 (row 1 is headers)
            
            contract_no = row.iloc[0]  # First column is CON_NO
            
            if pd.isna(contract_no) or contract_no is None:
                rows_to_delete.append(excel_row)
                continue
            
            contract_no_str = str(contract_no).strip().upper()
            
            # Calculate monthly sum for this row
            monthly_sum = 0
            monthly_data = row.iloc[1:]  # All columns after CON_NO
            for value in monthly_data:
                if pd.notna(value) and isinstance(value, (int, float)):
                    monthly_sum += value
            
            # Get Product from contract number (positions 2-3)
            product = contract_no_str[2:4] if len(contract_no_str) >= 4 else ''
            
            # Get NP/P from dictionary
            np_p_value = contract_to_npp.get(contract_no_str, "N/A")
            
            # Determine Type
            type_value = "Leases" if product.upper() in ['LE', 'UV'] else "Loans"
            
            if np_p_value != "N/A":
                matched_contracts += 1
            
            # Debug specific contracts
            if idx < 10:  # Log first 10 rows for debugging
                process_logger.info(f"Row {excel_row}: Contract='{contract_no_str}', NP/P='{np_p_value}', MonthlySum={monthly_sum}")
            
            # Mark for deletion if monthly sum is 0 and NP/P is N/A
            if monthly_sum == 0 and np_p_value == "N/A":
                rows_to_delete.append(excel_row)
                continue
            
            # Set the processed values in Excel
            ws.cell(row=excel_row, column=1).value = contract_no_str  # CON_NO
            ws.cell(row=excel_row, column=2).value = np_p_value      # NP/P
            ws.cell(row=excel_row, column=3).value = product         # Product
            ws.cell(row=excel_row, column=4).value = type_value      # Type
            
            # Column M (13) is blank - leave it empty
            
            # Copy all monthly data starting from column N (14)
            for i, value in enumerate(monthly_data):
                col_idx = monthly_start_col + i
                if pd.notna(value):
                    ws.cell(row=excel_row, column=col_idx).value = value
            
            processed_rows += 1
        
        process_logger.info(f"Processed {processed_rows} rows, {matched_contracts} contracts matched CBSL, marked {len(rows_to_delete)} for deletion")
        
        # Step 6: Delete rows with zero sum and N/A (from bottom to top)
        process_logger.info("Step 6: Deleting rows with zero sum and N/A")
        for row in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row)
        process_logger.info(f"Deleted {len(rows_to_delete)} rows")
        
        # Step 7: Add bucket formulas (columns E-L)
        process_logger.info("Step 7: Adding bucket formulas")
        current_max_row = ws.max_row
        
        # Bucket formulas based on monthly column positions
        # Monthly data now starts at column N (14) due to blank column M, so:
        # N=14, O=15, P=16, Q=17, etc.
        
        for row_num in range(2, current_max_row + 1):
            # E (7-30 Days): =N{row} (column 14)
            ws.cell(row=row_num, column=5).value = f'=IFERROR(IF(ISNUMBER(N{row_num}), N{row_num}, 0), 0)'
            
            # F (1-3 Months): =SUM(O{row}:P{row}) (columns 15:16)
            ws.cell(row=row_num, column=6).value = f'=IFERROR(SUM(O{row_num}:P{row_num}), 0)'
            
            # G (3-6 Months): =SUM(Q{row}:S{row}) (columns 17:19)
            ws.cell(row=row_num, column=7).value = f'=IFERROR(SUM(Q{row_num}:S{row_num}), 0)'
            
            # H (6-12 Months): =SUM(T{row}:Y{row}) (columns 20:25)
            ws.cell(row=row_num, column=8).value = f'=IFERROR(SUM(T{row_num}:Y{row_num}), 0)'
            
            # I (1-3 Years): =SUM(Z{row}:AW{row}) (columns 26:49)
            ws.cell(row=row_num, column=9).value = f'=IFERROR(SUM(Z{row_num}:AW{row_num}), 0)'
            
            # J (3-5 Years): =SUM(AX{row}:BU{row}) (columns 50:73)
            ws.cell(row=row_num, column=10).value = f'=IFERROR(SUM(AX{row_num}:BU{row_num}), 0)'
            
            # K (Over 5 Years): =SUM(BV{row}:CZ{row}) (columns 74:104)
            ws.cell(row=row_num, column=11).value = f'=IFERROR(SUM(BV{row_num}:CZ{row_num}), 0)'
            
            # L (Total): =SUM(E{row}:K{row}) (sum of all buckets)
            ws.cell(row=row_num, column=12).value = f'=IFERROR(E{row_num}+F{row_num}+G{row_num}+H{row_num}+I{row_num}+J{row_num}+K{row_num}, 0)'
        
        process_logger.info("Added bucket formulas for all data rows")
        
        # Step 8: Add sum formulas for all columns at the bottom
        process_logger.info("Step 8: Adding sum formulas at bottom")
        
        # Add blank row and sum row
        blank_row = current_max_row + 1
        sum_row = blank_row + 1
        
        # Add "Total" label
        ws.cell(row=sum_row, column=1).value = "Total"
        
        # Sum formulas for ALL monthly columns (N onwards) - skip blank column M
        current_max_col = ws.max_column
        
        for col_idx in range(monthly_start_col, current_max_col + 1):
            col_letter = get_column_letter(col_idx)
            sum_formula = f'=SUM({col_letter}2:{col_letter}{current_max_row})'
            ws.cell(row=sum_row, column=col_idx).value = sum_formula
            ws.cell(row=sum_row, column=col_idx).number_format = '#,##0.00;-#,##0.00;"-"'
        
        process_logger.info(f"Added sum formulas for {current_max_col - monthly_start_col + 1} monthly columns")
        
        # Step 9: Apply formatting
        process_logger.info("Step 9: Applying formatting")
        
        # Set column widths
        for col_idx in range(1, ws.max_column + 1):
            if col_idx >= 5 and col_idx <= 12:  # Bucket columns E-L need more width
                ws.column_dimensions[get_column_letter(col_idx)].width = 15
            elif col_idx == 13:  # Blank column M - make it narrow
                ws.column_dimensions[get_column_letter(col_idx)].width = 5
            else:
                ws.column_dimensions[get_column_letter(col_idx)].width = 12
        
        # Format bucket columns (E-L) to show whole numbers and dashes for zero
        for col_idx in range(5, 13):
            for row_idx in range(2, current_max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = '#,##0;-#,##0;"-"'
        
        # Format monthly columns (N onwards) to show decimals
        for col_idx in range(monthly_start_col, current_max_col + 1):
            for row_idx in range(2, current_max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = '#,##0.00;-#,##0.00;""'
        
        # Save the workbook
        wb.save(output_file_path)
        
        # Final verification
        final_rows = ws.max_row - 2  # Exclude header and sum row
        final_cols = ws.max_column
        process_logger.info(f"Final dimensions: {final_rows} data rows, {final_cols} columns")
        process_logger.info(f"Column structure: CON_NO | NP/P | Product | Type | 7 buckets | Total | [BLANK COLUMN M] | Monthly columns (starting at N)")
        
        process_logger.info(f"Processing for {current_csv_path} completed successfully in {datetime.now() - overall_start_time}")
        return True
        
    except Exception as e:
        process_logger.error(f"Processing for {current_csv_path} failed: {str(e)}")
        process_logger.error(traceback.format_exc())
        return False

def batch_process_files(file_configs, max_workers=None):
    """
    Process multiple files in parallel with configurable worker count.
    """
    logger = setup_logging()
    
    # Validate all files exist before starting
    valid_tasks = []
    for current, output, cbsl in file_configs:
        if not os.path.exists(current):
            logger.error(f"Current file not found: {current}")
        elif not os.path.exists(cbsl):
            logger.error(f"CBSL file not found: {cbsl}")
        else:
            valid_tasks.append((current, output, cbsl))
    
    if not valid_tasks:
        logger.error("No valid files to process!")
        return False
    
    # Determine optimal worker count
    if max_workers is None:
        max_workers = min(len(valid_tasks), multiprocessing.cpu_count())
    
    logger.info(f"Starting batch processing with {max_workers} workers for {len(valid_tasks)} files")
    batch_start_time = datetime.now()
    
    # Process files in parallel
    with multiprocessing.Pool(processes=max_workers) as pool:
        results = pool.map(process_matcap_file_optimized, valid_tasks)
    
    # Report results
    success_count = sum(results)
    total_time = datetime.now() - batch_start_time
    
    logger.info(f"Batch processing completed in {total_time}")
    logger.info(f"Success rate: {success_count}/{len(valid_tasks)} files processed successfully")
    
    if success_count < len(valid_tasks):
        failed_files = [task[0] for task, result in zip(valid_tasks, results) if not result]
        logger.warning(f"Failed files: {failed_files}")
    
    return success_count == len(valid_tasks)

def create_pivot_table(df):
    """Create pivot table from MATCAP data using proper pivot calculations."""
    print("Creating pivot table")
    print(f"Original data shape: {df.shape}")
    
    # Debug: Check unique values in NP/P and Type columns
    print(f"Unique NP/P values: {df['NP/P'].unique()}")
    print(f"Unique Type values: {df['Type'].unique()}")
    
    # Filter rows with valid NP/P values and non-null Type
    df_filtered = df[df['NP/P'].isin(['NP', 'P']) & df['Type'].notna()].copy()
    print(f"Filtered data shape: {df_filtered.shape}")
    
    # Debug: Show some sample data
    print("\nSample filtered data:")
    print(df_filtered[['NP/P', 'Type', 'Total']].head(10))
    
    # Define numeric columns
    numeric_cols = ['7-30 Days (c)', '1-3 Months (d)', '3-6 Months (e)', 
                    '6-12 Months (f)', '1-3 Years (g)', '3-5 Years (h)', 
                    'Over 5 Years (i)', 'Total']
    
    # Ensure all numeric columns are properly converted
    for col in numeric_cols:
        df_filtered[col] = pd.to_numeric(df_filtered[col], errors='coerce').fillna(0)
    
    # Create pivot table grouped by NP/P and Type
    pivot_table = df_filtered.groupby(['NP/P', 'Type'])[numeric_cols].sum()
    print("\nPivot table structure:")
    print(pivot_table)
    
    # Calculate totals for NP and P categories
    np_data = df_filtered[df_filtered['NP/P'] == 'NP']
    p_data = df_filtered[df_filtered['NP/P'] == 'P']
    
    np_totals = np_data[numeric_cols].sum()
    p_totals = p_data[numeric_cols].sum()
    grand_totals = df_filtered[numeric_cols].sum()
    
    print(f"\nNP totals: {np_totals['Total']}")
    print(f"P totals: {p_totals['Total']}")
    print(f"Grand totals: {grand_totals['Total']}")
    
    # Helper function to safely get pivot values
    def get_pivot_value(pivot_table, np_p, type_val, column):
        try:
            # Try different variations of Type names
            type_variations = [type_val, type_val.lower(), type_val.upper(), 
                             type_val + 's', type_val.rstrip('s')]
            
            for type_var in type_variations:
                if (np_p, type_var) in pivot_table.index:
                    return float(pivot_table.loc[(np_p, type_var), column])
            
            # If no exact match found, return 0
            return 0.0
        except (KeyError, IndexError):
            return 0.0
    
    # Build result DataFrame
    result_data = []
    
    # NP Section
    result_data.append({
        'Row Labels': 'NP',
        'Sum of 7-30 Days (c)': float(np_totals['7-30 Days (c)']),
        'Sum of 1-3 Months (d)': float(np_totals['1-3 Months (d)']),
        'Sum of 3-6 Months (e)': float(np_totals['3-6 Months (e)']),
        'Sum of 6-12 Months (f)': float(np_totals['6-12 Months (f)']),
        'Sum of 1-3 Years (g)': float(np_totals['1-3 Years (g)']),
        'Sum of 3-5 Years (h)': float(np_totals['3-5 Years (h)']),
        'Sum of Over 5 Years (i)': float(np_totals['Over 5 Years (i)']),
        'Sum of Total': float(np_totals['Total'])
    })
    
    # NP Lease
    result_data.append({
        'Row Labels': 'Lease',
        'Sum of 7-30 Days (c)': get_pivot_value(pivot_table, 'NP', 'Lease', '7-30 Days (c)'),
        'Sum of 1-3 Months (d)': get_pivot_value(pivot_table, 'NP', 'Lease', '1-3 Months (d)'),
        'Sum of 3-6 Months (e)': get_pivot_value(pivot_table, 'NP', 'Lease', '3-6 Months (e)'),
        'Sum of 6-12 Months (f)': get_pivot_value(pivot_table, 'NP', 'Lease', '6-12 Months (f)'),
        'Sum of 1-3 Years (g)': get_pivot_value(pivot_table, 'NP', 'Lease', '1-3 Years (g)'),
        'Sum of 3-5 Years (h)': get_pivot_value(pivot_table, 'NP', 'Lease', '3-5 Years (h)'),
        'Sum of Over 5 Years (i)': get_pivot_value(pivot_table, 'NP', 'Lease', 'Over 5 Years (i)'),
        'Sum of Total': get_pivot_value(pivot_table, 'NP', 'Lease', 'Total')
    })
    
    # NP Loans
    result_data.append({
        'Row Labels': 'Loans',
        'Sum of 7-30 Days (c)': get_pivot_value(pivot_table, 'NP', 'Loan', '7-30 Days (c)'),
        'Sum of 1-3 Months (d)': get_pivot_value(pivot_table, 'NP', 'Loan', '1-3 Months (d)'),
        'Sum of 3-6 Months (e)': get_pivot_value(pivot_table, 'NP', 'Loan', '3-6 Months (e)'),
        'Sum of 6-12 Months (f)': get_pivot_value(pivot_table, 'NP', 'Loan', '6-12 Months (f)'),
        'Sum of 1-3 Years (g)': get_pivot_value(pivot_table, 'NP', 'Loan', '1-3 Years (g)'),
        'Sum of 3-5 Years (h)': get_pivot_value(pivot_table, 'NP', 'Loan', '3-5 Years (h)'),
        'Sum of Over 5 Years (i)': get_pivot_value(pivot_table, 'NP', 'Loan', 'Over 5 Years (i)'),
        'Sum of Total': get_pivot_value(pivot_table, 'NP', 'Loan', 'Total')
    })
    
    # P Section
    result_data.append({
        'Row Labels': 'P',
        'Sum of 7-30 Days (c)': float(p_totals['7-30 Days (c)']),
        'Sum of 1-3 Months (d)': float(p_totals['1-3 Months (d)']),
        'Sum of 3-6 Months (e)': float(p_totals['3-6 Months (e)']),
        'Sum of 6-12 Months (f)': float(p_totals['6-12 Months (f)']),
        'Sum of 1-3 Years (g)': float(p_totals['1-3 Years (g)']),
        'Sum of 3-5 Years (h)': float(p_totals['3-5 Years (h)']),
        'Sum of Over 5 Years (i)': float(p_totals['Over 5 Years (i)']),
        'Sum of Total': float(p_totals['Total'])
    })
    
    # P Lease
    result_data.append({
        'Row Labels': 'Lease',
        'Sum of 7-30 Days (c)': get_pivot_value(pivot_table, 'P', 'Lease', '7-30 Days (c)'),
        'Sum of 1-3 Months (d)': get_pivot_value(pivot_table, 'P', 'Lease', '1-3 Months (d)'),
        'Sum of 3-6 Months (e)': get_pivot_value(pivot_table, 'P', 'Lease', '3-6 Months (e)'),
        'Sum of 6-12 Months (f)': get_pivot_value(pivot_table, 'P', 'Lease', '6-12 Months (f)'),
        'Sum of 1-3 Years (g)': get_pivot_value(pivot_table, 'P', 'Lease', '1-3 Years (g)'),
        'Sum of 3-5 Years (h)': get_pivot_value(pivot_table, 'P', 'Lease', '3-5 Years (h)'),
        'Sum of Over 5 Years (i)': get_pivot_value(pivot_table, 'P', 'Lease', 'Over 5 Years (i)'),
        'Sum of Total': get_pivot_value(pivot_table, 'P', 'Lease', 'Total')
    })
    
    # P Loans
    result_data.append({
        'Row Labels': 'Loans',
        'Sum of 7-30 Days (c)': get_pivot_value(pivot_table, 'P', 'Loan', '7-30 Days (c)'),
        'Sum of 1-3 Months (d)': get_pivot_value(pivot_table, 'P', 'Loan', '1-3 Months (d)'),
        'Sum of 3-6 Months (e)': get_pivot_value(pivot_table, 'P', 'Loan', '3-6 Months (e)'),
        'Sum of 6-12 Months (f)': get_pivot_value(pivot_table, 'P', 'Loan', '6-12 Months (f)'),
        'Sum of 1-3 Years (g)': get_pivot_value(pivot_table, 'P', 'Loan', '1-3 Years (g)'),
        'Sum of 3-5 Years (h)': get_pivot_value(pivot_table, 'P', 'Loan', '3-5 Years (h)'),
        'Sum of Over 5 Years (i)': get_pivot_value(pivot_table, 'P', 'Loan', 'Over 5 Years (i)'),
        'Sum of Total': get_pivot_value(pivot_table, 'P', 'Loan', 'Total')
    })
    
    # Grand Total
    result_data.append({
        'Row Labels': 'Grand Total',
        'Sum of 7-30 Days (c)': float(grand_totals['7-30 Days (c)']),
        'Sum of 1-3 Months (d)': float(grand_totals['1-3 Months (d)']),
        'Sum of 3-6 Months (e)': float(grand_totals['3-6 Months (e)']),
        'Sum of 6-12 Months (f)': float(grand_totals['6-12 Months (f)']),
        'Sum of 1-3 Years (g)': float(grand_totals['1-3 Years (g)']),
        'Sum of 3-5 Years (h)': float(grand_totals['3-5 Years (h)']),
        'Sum of Over 5 Years (i)': float(grand_totals['Over 5 Years (i)']),
        'Sum of Total': float(grand_totals['Total'])
    })
    
    return pd.DataFrame(result_data)

def save_pivot_table_to_sheet1(matcap_file, pivot_df):
    """Save pivot table to Sheet1 in MATCAP file with proper formatting."""
    print("Saving pivot table to Sheet1")
    
    try:
        # Check file access
        if not wait_for_file_access(matcap_file):
            print(f"File {matcap_file} is locked by another process. Please close it and try again.")
            return False
        
        # Create temporary file
        temp_file = "TEMP_PIVOT.xlsx"
        if os.path.exists(temp_file):
            os.remove(temp_file)
        shutil.copy2(matcap_file, temp_file)
        
        # Load workbook
        wb = openpyxl.load_workbook(temp_file)
        
        # Remove existing Sheet1 if present
        if 'Sheet1' in wb.sheetnames:
            wb.remove(wb['Sheet1'])
        
        # Create Sheet1 at the beginning
        ws_pivot = wb.create_sheet('Sheet1', 0)
        
        # Define headers
        headers = ['Row Labels', 'Sum of 7-30 Days (c)', 'Sum of 1-3 Months (d)', 
                   'Sum of 3-6 Months (e)', 'Sum of 6-12 Months (f)', 
                   'Sum of 1-3 Years (g)', 'Sum of 3-5 Years (h)', 
                   'Sum of Over 5 Years (i)', 'Sum of Total']
        
        # Write headers at row 3
        for col_idx, header in enumerate(headers, 1):
            cell = ws_pivot.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
        
        # Write data starting at row 4
        for row_idx, (_, row_data) in enumerate(pivot_df.iterrows(), 4):
            # Row Labels
            label_cell = ws_pivot.cell(row=row_idx, column=1)
            label = str(row_data['Row Labels'])
            label_cell.value = label
            if label in ['NP', 'P', 'Grand Total']:
                label_cell.font = Font(bold=True)
                label_cell.alignment = Alignment(indent=0)
            else:
                label_cell.alignment = Alignment(indent=1)
            
            # Numeric columns
            numeric_columns = ['Sum of 7-30 Days (c)', 'Sum of 1-3 Months (d)', 'Sum of 3-6 Months (e)', 
                              'Sum of 6-12 Months (f)', 'Sum of 1-3 Years (g)', 'Sum of 3-5 Years (h)', 
                              'Sum of Over 5 Years (i)', 'Sum of Total']
            
            for col_idx, col_name in enumerate(numeric_columns, 2):
                cell = ws_pivot.cell(row=row_idx, column=col_idx)
                value = row_data[col_name]

                # Always write numeric values; replace missing with 0
                if pd.isna(value):
                    value_to_write = 0
                else:
                    value_to_write = int(value) if value == int(value) else float(value)

                cell.value = value_to_write
                # Show dash for zero values, thousands separator, no decimals
                cell.number_format = '#,##0;-#,##0;-'
                # Bold totals
                if label in ['NP', 'P', 'Grand Total'] and col_name == 'Sum of Total':
                    cell.font = Font(bold=True)
        
        # Adjust column widths
        for col_idx in range(1, len(headers) + 1):
            ws_pivot.column_dimensions[get_column_letter(col_idx)].width = 18
        
        # Save to temp and move
        wb.save(temp_file)
        if not wait_for_file_access(matcap_file):
            print(f"File {matcap_file} is still locked. Cannot replace it.")
            os.remove(temp_file)
            return False
        
        os.remove(matcap_file)
        shutil.move(temp_file, matcap_file)
        print("Pivot table saved successfully to Sheet1")
        return True
    
    except Exception as e:
        print(f"Error saving pivot table: {str(e)}")
        if os.path.exists(temp_file):
            os.remove(temp_file)
        return False

def load_matcap_for_pivot(matcap_file):
    """Load updated MATCAP sheet and compute bucket sums from monthly columns.

    This avoids relying on Excel formulas (which are not evaluated by openpyxl)
    by summing the raw monthly columns directly.
    """
    print("Loading MATCAP data for pivot (computing buckets from monthly columns)")
    wb = openpyxl.load_workbook(matcap_file, data_only=True)
    ws = wb['MATCAP-01-08-2025']

    max_row = ws.max_row
    max_col = ws.max_column

    # Column indices (1-based): A=1, B=2, C=3, D=4, ..., M=13
    col_idx_con_no = 1
    col_idx_np_p = 2
    col_idx_type = 4

    # Fixed bucket columns per design
    idx_7_30 = column_index_from_string('N')
    range_1_3m = (column_index_from_string('O'), column_index_from_string('P'))
    range_3_6m = (column_index_from_string('Q'), column_index_from_string('S'))
    range_6_12m = (column_index_from_string('T'), column_index_from_string('Y'))
    range_1_3y = (column_index_from_string('Z'), column_index_from_string('AW'))
    range_3_5y = (column_index_from_string('AX'), column_index_from_string('BU'))
    # Clamp Over 5 Years to available columns, target BV:CZ
    over5_start = column_index_from_string('BV')
    over5_end_target = column_index_from_string('CZ')
    range_over_5y = (over5_start, min(over5_end_target, max_col))

    records = []
    for row in range(2, max_row + 1):
        con_no = ws.cell(row=row, column=col_idx_con_no).value
        if con_no is None or (isinstance(con_no, str) and con_no.strip().lower() == 'total'):
            continue

        np_p = ws.cell(row=row, column=col_idx_np_p).value
        typ = ws.cell(row=row, column=col_idx_type).value

        if np_p not in ['NP', 'P'] or typ is None or str(typ).strip() == '':
            continue

        # Helper to safely get numeric from a cell
        def num_at(cidx):
            v = ws.cell(row=row, column=cidx).value
            return float(v) if isinstance(v, (int, float)) else 0.0

        # Sum helpers for ranges
        def sum_range(a, b):
            end = min(b, max_col)
            start = min(a, end)
            return sum(num_at(ci) for ci in range(start, end + 1))

        v_7_30 = num_at(idx_7_30) if idx_7_30 <= max_col else 0.0
        v_1_3m = sum_range(*range_1_3m)
        v_3_6m = sum_range(*range_3_6m)
        v_6_12m = sum_range(*range_6_12m)
        v_1_3y = sum_range(*range_1_3y)
        v_3_5y = sum_range(*range_3_5y)
        v_over_5y = sum_range(*range_over_5y)

        total = v_7_30 + v_1_3m + v_3_6m + v_6_12m + v_1_3y + v_3_5y + v_over_5y

        records.append({
            'NP/P': str(np_p),
            'Type': str(typ).strip(),
            '7-30 Days (c)': v_7_30,
            '1-3 Months (d)': v_1_3m,
            '3-6 Months (e)': v_3_6m,
            '6-12 Months (f)': v_6_12m,
            '1-3 Years (g)': v_1_3y,
            '3-5 Years (h)': v_3_5y,
            'Over 5 Years (i)': v_over_5y,
            'Total': total
        })

    wb.close()
    return pd.DataFrame.from_records(records)

def part2_main():
    """Main function for Part 2 - Pivot Table Creation"""
    # File path
    matcap_file = "UPDATED-MATCAP-01-08-2025.xlsx"
    
    # Check if file exists
    if not os.path.exists(matcap_file):
        print(f"File {matcap_file} not found!")
        return
    
    try:
        # Load data and compute buckets from monthly columns (avoid Excel formulas)
        df = load_matcap_for_pivot(matcap_file)

        if df.empty:
            print("No data rows found for pivot. Check the updated workbook.")
            return

        # Create pivot table with proper calculations
        pivot_df = create_pivot_table(df)

        # Save to Sheet1
        success = save_pivot_table_to_sheet1(matcap_file, pivot_df)

        if success:
            print(f"Successfully created pivot table in {matcap_file}, Sheet1")
        else:
            print("Failed to create pivot table")

    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        print(traceback.format_exc())

def copy_matcap_data(excel):
    """
    Copy data from UPDATED-MATCAP-01-08-2025.xlsx to NBD-MF-10-GA & NBD-MF-11-IS July 2025.xlsx
    before creating CBSL pivot table.
    """
    
    try:
        matcap_file = "UPDATED-MATCAP-01-08-2025.xlsx"
        
        if not os.path.exists(matcap_file):
            print(f"WARNING: {matcap_file} not found. Skipping MATCAP data copy.")
            return True
        
        print(f"Processing MATCAP data from: {matcap_file}")
        
        # Open MATCAP file
        matcap_wb = excel.Workbooks.Open(os.path.abspath(matcap_file))
        matcap_ws = matcap_wb.Sheets("Sheet1")
        
        # Define possible NBD file names
        nbd_files = [
            "NBD-MF-10-GA & NBD-MF-11-IS July 2025.xlsx",
            "NBD-MF-10-GA & NBD-MF-11-IS July 2025.xlsb"
        ]
        
        nbd_file_found = None
        for nbd_file in nbd_files:
            if os.path.exists(nbd_file):
                nbd_file_found = nbd_file
                break
                
        if not nbd_file_found:
            print(f"WARNING: NBD file not found. Tried: {', '.join(nbd_files)}")
            matcap_wb.Close(False)
            return False
        
        print(f"Copying MATCAP data to: {nbd_file_found}")
        
        # Open NBD file
        nbd_wb = excel.Workbooks.Open(os.path.abspath(nbd_file_found))
        nbd_ws = nbd_wb.Sheets("Portfolio")
        
        # Define the copy mappings: (source_range, destination_cell_start)
        copy_mappings = [
            # Source Range, Destination Start Cell
            ("B4:B10", "D4"),  # Sheet1 B4-B10 to D4-D10
            ("C4:C10", "E4"),  # Sheet1 C4-C10 to E4-E10  
            ("D4:D10", "F4"),  # Sheet1 D4-D10 to F4-F10
            ("E4:E10", "G4"),  # Sheet1 E4-E10 to G4-G10
            ("F4:F10", "H4"),  # Sheet1 F4-F10 to H4-H10
            ("G4:G10", "I4"),  # Sheet1 G4-G10 to I4-I10
            ("H4:H10", "J4"),  # Sheet1 H4-H10 to J4-J10
            ("I4:I10", "K4")   # Sheet1 I4-I10 to K4-K10
        ]
        
        # Copy data for each mapping
        for source_range, dest_start in copy_mappings:
            try:
                # Get source values
                source_cells = matcap_ws.Range(source_range)
                source_values = source_cells.Value
                
                # Calculate destination range
                dest_row_start = int(dest_start[1:])  # Extract row number from cell reference
                dest_col = dest_start[0]  # Extract column letter
                dest_range = f"{dest_col}{dest_row_start}:{dest_col}{dest_row_start + 6}"  # 7 rows (4-10)
                
                # Paste values to destination
                nbd_ws.Range(dest_range).Value = source_values
                
                print(f"Copied {source_range} to {dest_range}")
                
            except Exception as e:
                print(f"Error copying {source_range} to {dest_start}: {e}")
                continue
        
        # Save and close NBD file
        # After pasting, convert any "-" to 0 in D4:D10, E4:E10, F4:F10, G4:G10, H4:H10, I4:I10, J4:J10, K4:K10
        converted_count = 0
        for col in ["D", "E", "F", "G", "H", "I", "J", "K"]:
            for row in range(4, 11):
                cell = nbd_ws.Range(f"{col}{row}")
                val = cell.Value
                if isinstance(val, str) and val.strip() == "-":
                    cell.Value = 0
                    converted_count += 1
        if converted_count:
            print(f"Normalized '-' to 0 in {converted_count} cells within D4:K10")

        nbd_wb.Save()
        nbd_wb.Close()
        
        # Close MATCAP file
        matcap_wb.Close(False)
        
        print("SUCCESS: MATCAP data copied successfully.")
        return True
        
    except Exception as e:
        print(f"ERROR in MATCAP copy process: {e}")
        return False

def create_cbsl_pivot_table(excel, cbsl_file_path):
    """
    Creates CBSL Pivot Table directly in Excel using win32com (fastest for large files).
    - Uses Excel COM automation (requires Windows + Excel installed).
    - Creates a pivot table on new Sheet1 while keeping Portfolio sheet intact.
    """

    start_time = time.time()

    try:
        print("Opening workbook...")
        wb = excel.Workbooks.Open(os.path.abspath(cbsl_file_path))

        # Reference Portfolio sheet
        ws_data = wb.Sheets("Portfolio")

        # Detect last used row & col (dynamic range detection)
        last_row = ws_data.Cells(ws_data.Rows.Count, 1).End(win32.constants.xlUp).Row
        last_col = ws_data.Cells(2, ws_data.Columns.Count).End(win32.constants.xlToLeft).Column
        print(f"Detected data range: {last_row} rows x {last_col} columns")

        data_range = ws_data.Range(ws_data.Cells(3, 1), ws_data.Cells(last_row, last_col))

        # Delete existing Sheet1 if present
        for sheet in wb.Sheets:
            if sheet.Name == "Sheet1":
                sheet.Delete()
                break

        ws_pivot = wb.Sheets.Add(Before=wb.Sheets(1))
        ws_pivot.Name = "Sheet1"

        # Create pivot cache
        pivot_cache = wb.PivotCaches().Create(
            SourceType=win32.constants.xlDatabase,
            SourceData=data_range
        )

        # Create pivot table
        pivot_table_name = "CBSL_Pivot"
        pivot_table = pivot_cache.CreatePivotTable(
            TableDestination=ws_pivot.Cells(1, 1),
            TableName=pivot_table_name
        )

        # Add row fields
        pivot_table.PivotFields("CBSL P/NP").Orientation = win32.constants.xlRowField
        pivot_table.PivotFields("ACTUAL LEASE/LOAN").Orientation = win32.constants.xlRowField

        # Add value fields (using ARREARS instead of TOTAL ARREARS as shown in screenshots)
        value_fields = [
            "ARREARS", 
            "MOI ADJUSTMENT", 
            "INTEREST RECEIVABLE",
            "DEFERRED PROMOTION", 
            "IIS", 
            "IMP + PRO.DP"
        ]

        for field in value_fields:
            pf = pivot_table.PivotFields(field)
            pf.Orientation = win32.constants.xlDataField
            pf.Function = win32.constants.xlSum
            pf.NumberFormat = "#,##0"
            pf.Name = f"Sum of {field}"

        # Formatting
        ws_pivot.Columns.AutoFit()
        ws_pivot.Rows.AutoFit()
        ws_pivot.Range("A1").Select()

        # Save & close
        wb.Save()
        wb.Close(SaveChanges=True)

        total_time = time.time() - start_time
        print(f"Pivot table created successfully in {total_time:.2f} seconds!")
        print("Open Sheet1 to view results.")
        return True

    except Exception as e:
        print(f"ERROR: {e}")
        return False

def copy_to_nbd_files(excel, cbsl_file_path):
    """
    Copy values from CBSL file to NBD-MF files and perform additional operations
    """
    
    try:
        # Open CBSL file
        cbsl_wb = excel.Workbooks.Open(os.path.abspath(cbsl_file_path))
        cbsl_ws = cbsl_wb.Sheets("Sheet1")
        
        # Define possible NBD file names (both .xlsx and .xlsb formats)
        nbd_files = [
            "NBD-MF-10-GA & NBD-MF-11-IS July 2025.xlsx",
            "NBD-MF-10-GA & NBD-MF-11-IS July 2025.xlsb"
        ]
        
        # Mapping of source to destination cells
        cell_mapping = [
            # Destination, Source
            ("N14", "B3"), ("O14", "C3"), ("P14", "D3"), ("Q14", "E3"), ("R14", "F3"), ("S14", "G3"),
            ("N15", "B4"), ("O15", "C4"), ("P15", "D4"), ("Q15", "E4"), ("R15", "F4"), ("S15", "G4"),
            ("N16", "B5"), ("O16", "C5"), ("P16", "D5"), ("Q16", "E5"), ("R16", "F5"), ("S16", "G5"),
            ("N17", "B6"), ("O17", "C6"), ("P17", "D6"), ("Q17", "E6"), ("R17", "F6"), ("S17", "G6"),
            ("N18", "B7"), ("O18", "C7"), ("P18", "D7"), ("Q18", "E7"), ("R18", "F7"), ("S18", "G7"),
            ("N19", "B8"), ("O19", "C8"), ("P19", "D8"), ("Q19", "E8"), ("R19", "F8"), ("S19", "G8"),
            ("N20", "B9"), ("O20", "C9"), ("P20", "D9"), ("Q20", "E9"), ("R20", "F9"), ("S20", "G9")
        ]
        
        nbd_file_found = None
        for nbd_file in nbd_files:
            if os.path.exists(nbd_file):
                nbd_file_found = nbd_file
                break
                
        if not nbd_file_found:
            print(f"WARNING: NBD file not found. Tried: {', '.join(nbd_files)}")
            cbsl_wb.Close(False)
            return False
        
        print(f"Processing: {nbd_file_found}")
        
        # Open NBD file
        nbd_wb = excel.Workbooks.Open(os.path.abspath(nbd_file_found))
        nbd_ws = nbd_wb.Sheets("Portfolio")
        
        # Clear destination cells (only values, keep formulas)
        for i in range(14, 21):  # Rows 14 to 20
            for col in ["N", "O", "P", "Q", "R", "S"]:
                cell = nbd_ws.Range(f"{col}{i}")
                # Only clear values, keep formulas intact
                cell.Value = None
        
        # Copy values from CBSL to NBD
        for dest, src in cell_mapping:
            value = cbsl_ws.Range(src).Value
            nbd_ws.Range(dest).Value = value
        
        # Clear values in O4 to O8 (preserve formulas)
        for i in range(4, 9):  # O4 to O8
            cell = nbd_ws.Range(f"O{i}")
            # Only clear value, keep formula intact
            cell.Value = None
        
        # Set formula for O9 if it doesn't have one
        o9_cell = nbd_ws.Range("O9")
        if not o9_cell.Formula or not o9_cell.Formula.startswith("="):
            o9_cell.Formula = "=O6+O7+O8"
        
        # Process additional steps
        sofp_file = "NBD-MF-01-SOFP & SOCI AFL Monthly FS July 2025.xlsx"
        if os.path.exists(sofp_file):
            print(f"Processing: {sofp_file}")
            
            # Open SOFP file
            sofp_wb = excel.Workbooks.Open(os.path.abspath(sofp_file))
            sofp_ws = sofp_wb.Sheets("CBSL Provision")
            
            # Copy C9 to O4 (only value, keep formula if exists)
            c9_value = sofp_ws.Range("C9").Value
            nbd_ws.Range("O4").Value = c9_value
            
            # Copy C6 and C7, then add them and paste to O5 (only value, keep formula if exists)
            c6_value = sofp_ws.Range("C6").Value
            c7_value = sofp_ws.Range("C7").Value
            nbd_ws.Range("O5").Value = c6_value + c7_value
            
            # Calculate O6 = O4 + O5 (only value, keep formula if exists)
            o4_value = nbd_ws.Range("O4").Value
            o5_value = nbd_ws.Range("O5").Value
            if o4_value is not None and o5_value is not None:
                nbd_ws.Range("O6").Value = o4_value + o5_value
            
            # Copy H10 to O7 with brackets (only value, keep formula if exists)
            h10_value = sofp_ws.Range("H10").Value
            if h10_value is not None:
                nbd_ws.Range("O7").Value = f"({h10_value})"
            
            # Copy G10 to O8 with brackets (only value, keep formula if exists)
            g10_value = sofp_ws.Range("G10").Value
            if g10_value is not None:
                nbd_ws.Range("O8").Value = f"({g10_value})"
            
            # Close SOFP file
            sofp_wb.Close(False)
            print(f"Completed: {sofp_file}")
        else:
            print(f"WARNING: {sofp_file} not found. Skipping additional operations.")
        
        # Set formula for P7 = N7 - O7 and handle zero values
        p7_cell = nbd_ws.Range("P7")
        p7_cell.Formula = "=N7-O7"
        
        # Apply custom number format to show 0 instead of "-" when result is zero
        p7_cell.NumberFormat = "#,##0;-#,##0;0"
        
        # Copy values from FD Base file
        fd_base_file = "FD Base as at 31.07.2025.xlsx"
        if os.path.exists(fd_base_file):
            print(f"Processing: {fd_base_file}")
            
            # Open FD Base file
            fd_base_wb = excel.Workbooks.Open(os.path.abspath(fd_base_file))
            fd_base_ws = fd_base_wb.Sheets("FD Loan")
            
            # Mapping of FD Base to NBD cells
            fd_cell_mapping = [
                # Destination (NBD), Source (FD Base)
                ("C33", "U23"), ("D33", "V23"), ("E33", "W23"), 
                ("F33", "X23"), ("G33", "Y23"), ("H33", "Z23"),
                ("I33", "AA23"), ("J33", "AB23"), ("K33", "AC23")
            ]
            
            # Copy values from FD Base to NBD
            for dest, src in fd_cell_mapping:
                value = fd_base_ws.Range(src).Value
                nbd_ws.Range(dest).Value = value
            
            # Close FD Base file
            fd_base_wb.Close(False)
            print(f"Completed: {fd_base_file}")
        else:
            print(f"WARNING: {fd_base_file} not found. Skipping FD Base operations.")
        
        # Save and close NBD file
        nbd_wb.Save()
        nbd_wb.Close()
        
        print(f"Completed: {nbd_file_found}")
        
        # Close CBSL file
        cbsl_wb.Close(False)
        
        return True
        
    except Exception as e:
        print(f"ERROR in copy process: {e}")
        return False

def main():
    """Main pipeline function that runs all parts sequentially"""
    # Ensure we run from the directory containing this script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    os.chdir(script_dir)

    # Step 1: Validate all required files exist
    print("Starting MATCAP Processing Pipeline")
    print("="*60)
    
    validation_success, found_files = validate_required_files()
    
    if not validation_success:
        print("❌ Cannot proceed - required files are missing!")
        print("Please ensure all required files are present and try again.")
        return False
    
    print("✅ All required files found. Proceeding with MATCAP processing...")
    print()

    # File names (updated for CSV input) - use found files if available
    current_matcap = found_files.get("MATCAP", "MATCAP-01-08-2025.csv")
    updated_matcap = "UPDATED-MATCAP-01-08-2025.xlsx"
    cbsl_file = found_files.get("CBSL Provision Comparison", "CBSL Provision Comparison - Jul 2025.xlsx")
    
    print(f"Using files:")
    print(f"  - MATCAP: {current_matcap}")
    print(f"  - CBSL: {cbsl_file}")
    print(f"  - Output: {updated_matcap}")
    print()

    # Validate inputs for part 1
    if not os.path.exists(current_matcap):
        raise FileNotFoundError(f"Input MATCAP CSV file not found: {current_matcap}")
    if not os.path.exists(cbsl_file):
        raise FileNotFoundError(f"CBSL mapping file not found: {cbsl_file}")

    # Run Part 1: process CSV and generate Excel workbook
    print("Running Part 1: converting CSV to Excel with MATCAP processing...")
    success = convert_csv_to_excel_with_processing((current_matcap, updated_matcap, cbsl_file))
    if not success:
        raise RuntimeError("Part 1 failed to process the MATCAP CSV file.")

    if not os.path.exists(updated_matcap):
        raise FileNotFoundError(f"Expected updated workbook not found: {updated_matcap}")

    # Run Part 2: create pivot table in Sheet1 and save
    print("Running Part 2: creating pivot table and saving to Sheet1...")
    part2_main()

    # Run Part 3: CBSL Pivot Table and NBD file operations
    print("Running Part 3: CBSL Pivot Table and NBD file operations...")
    
    # Create Excel application instance for Part 3
    excel_app = win32.Dispatch("Excel.Application")
    excel_app.Visible = False  # Keep Excel hidden for speed
    excel_app.DisplayAlerts = False  # Disable alerts for automation

    try:
        # Step 1: Copy MATCAP data before creating pivot table
        print("Step 1: Copying MATCAP data to NBD file...")
        matcap_success = copy_matcap_data(excel_app)

        if matcap_success:
            print("SUCCESS: MATCAP data copied. Now creating pivot table...")
            
            # Step 2: Create pivot table in CBSL file
            pivot_success = create_cbsl_pivot_table(excel_app, cbsl_file)

            if pivot_success:
                print("SUCCESS: Pivot table created. Now copying data to NBD file...")
                # Step 3: Copy data from CBSL to NBD file
                copy_success = copy_to_nbd_files(excel_app, cbsl_file)
                
                if copy_success:
                    print("SUCCESS: All operations completed successfully!")
                else:
                    print("FAILED: Could not copy data to NBD file.")
            else:
                print("FAILED: Could not create pivot table.")
        else:
            print("FAILED: Could not copy MATCAP data.")

    except Exception as e:
        print(f"ERROR in Part 3 process: {e}")
        import traceback
        print(traceback.format_exc())

    finally:
        # Always quit Excel
        excel_app.Quit()

    # After all processing is completed, remove the UPDATED-MATCAP file if it is no longer needed
    try:
        # Check if the updated MATCAP file exists in the working directory
        if os.path.exists(updated_matcap):
            # Attempt to delete the updated MATCAP file to avoid leaving unnecessary artifacts
            os.remove(updated_matcap)
            # Confirm deletion to the user via console output
            print(f"Clean up: Deleted '{updated_matcap}' as it is no longer needed.")
        else:
            # Inform that there is nothing to delete if the file is absent
            print(f"Clean up: '{updated_matcap}' not found; no deletion necessary.")
    except Exception as e:
        # Log a non-fatal warning if deletion fails for any reason (e.g., file locked)
        print(f"Clean up warning: Could not delete '{updated_matcap}': {e}")

    print("Pipeline completed. All operations finished.")

if __name__ == "__main__":
    main()