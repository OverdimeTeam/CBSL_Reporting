import os
import re
import pandas as pd
import calendar
import sys
import datetime as dt
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
import argparse


# === Setup logging to file ===
class Logger:
    def __init__(self, filename):
        self.terminal = sys.stdout
        self.log = open(filename, 'w', encoding='utf-8')

    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)

    def flush(self):
        self.terminal.flush()
        self.log.flush()

    def close(self):
        self.log.close()

# === Step 1: Build dynamic path ===
# Parse command line arguments
parser = argparse.ArgumentParser(description='NBD QF 23 C8 Report Processing')
parser.add_argument('--working-dir', help='Working directory path (date folder)')
parser.add_argument('date', nargs='?', help='Report date in MM/DD/YYYY format')
args = parser.parse_args()

# Current file directory: C:\CBSL\Script\report_automations\
current_dir = os.path.dirname(os.path.abspath(__file__))

# Go one level up to C:\CBSL\Script
base_dir = os.path.abspath(os.path.join(current_dir, os.pardir))

if args.working_dir:
    # Use provided working directory (should be the date folder)
    target_folder = args.working_dir
    data_dir = target_folder
    print(f"Using provided working directory: {target_folder}")
else:
    # Construct path to C:\CBSL\Script\working\NBD_QF_23_C8
    working_dir = os.path.join(base_dir, "working", "NBD_QF_23_C8")

    # Check if working directory exists
    if not os.path.exists(working_dir):
        raise FileNotFoundError(f"Working directory not found: {working_dir}")

    # Find the single date folder inside 'NBD_QF_23_C8'
    subfolders = [f for f in os.listdir(working_dir) if os.path.isdir(os.path.join(working_dir, f))]

    if not subfolders:
        raise FileNotFoundError(f"No date folder found inside '{working_dir}' directory.")

    # Use the first (and should be only) date folder
    if len(subfolders) > 1:
        print(f"Warning: Multiple date folders found: {subfolders}. Using the first one: {subfolders[0]}")

    target_folder = os.path.join(working_dir, subfolders[0])
    data_dir = target_folder  # Files are directly in the date folder
    print(f"Using auto-detected working directory: {target_folder}")

# Initialize logger
log_file = os.path.join(data_dir, "NBD_QF_23_C8_debug.log")
sys.stdout = Logger(log_file)
print(f"=== Script started at {dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===")
print(f"Log file: {log_file}\n")

# === Step 2: Identify Client Rating and Summary files ===
client_rating_file = None
summary_file = None
# unutilized_amount_file = None
prod_wise_class_file = None 

for f in os.listdir(data_dir):
    if f.startswith("C8-Client Rating") and f.endswith(".xlsx"):
        client_rating_file = os.path.join(data_dir, f)
    elif f.startswith("Summary") and f.endswith(".xlsb"):
        summary_file = os.path.join(data_dir, f)
    # elif f.startswith("Unutilized Amount") and f.endswith(".xlsx"):
    #     unutilized_amount_file = os.path.join(data_dir, f)
    elif f.startswith("Prod. wise Class. of Loans ") and f.endswith(".xlsb"):
        prod_wise_class_file = os.path.join(data_dir, f)

if not client_rating_file:
    raise FileNotFoundError("No file found starting with 'C8-Client Rating' and ending with .xlsx")
if not summary_file:
    raise FileNotFoundError("No file found starting with 'Summary' and ending with .xlsb")
# if not unutilized_amount_file:
#     raise FileNotFoundError("No file found starting with 'unutilized_amount' and ending with .xlsb")
if not prod_wise_class_file:
    raise FileNotFoundError("No file found starting with 'Prod. wise Class. of Loans' and ending with .xlsb")

print(f"Selected Client Rating file: {client_rating_file}")
print(f"Selected Summary file: {summary_file}")
# print(f"Selected Unutilized Amount file: {unutilized_amount_file}")
print(f"Selected Summary file: {prod_wise_class_file}")

# === Step 3: Identify latest dated sheet from Client Rating file ===
xls = pd.ExcelFile(client_rating_file, engine="openpyxl")
sheet_names = xls.sheet_names
print("Available sheets in Client Rating file:", sheet_names)

date_pattern = re.compile(r"Rating\s*-(\d{2}-\d{2}-\d{4})")
sheet_dates = {}

for sheet in sheet_names:
    match = date_pattern.search(sheet)
    if match:
        date_str = match.group(1)
        date = pd.to_datetime(date_str, format="%d-%m-%Y", errors="coerce")
        if pd.notnull(date):
            sheet_dates[sheet] = date

if not sheet_dates:
    raise ValueError("No valid sheets found with date format like 'Rating -30-06-2024'")


# Convert UI input to datetime
if args.date:
    input_date_str = args.date
else:
    raise ValueError("Please provide a date argument in MM/DD/YYYY format")

try:
    report_date = pd.to_datetime(input_date_str, format="%m/%d/%Y")
except ValueError:
    raise ValueError("Invalid date format. Please use MM/DD/YYYY format.")

selected_date = report_date
new_sheet_name = f"Rating -{selected_date.strftime('%d-%m-%Y')}"
print(f"Target new sheet name: {new_sheet_name}")

# --- Build sheet_dates dict from existing sheets ---
date_pattern = re.compile(r"Rating\s*-\s*(\d{2}-\d{2}-\d{4})")
sheet_dates = {}
for sheet in sheet_names:
    match = date_pattern.search(sheet)
    if match:
        try:
            sheet_date = pd.to_datetime(match.group(1), format="%d-%m-%Y")
            sheet_dates[sheet] = sheet_date
        except Exception:
            continue

if not sheet_dates:
    raise ValueError("No valid Rating sheets found.")

# --- Determine latest sheet to duplicate ---
if new_sheet_name in sheet_dates:
    # Sheet with selected date already exists, use the latest sheet before it
    filtered_sheets = {s: d for s, d in sheet_dates.items() if d < selected_date}
    if filtered_sheets:
        latest_sheet = max(filtered_sheets, key=lambda k: filtered_sheets[k])
        latest_date = filtered_sheets[latest_sheet]
        print(f"Latest Rating sheet identified before selected date: {latest_sheet} ({latest_date.strftime('%d-%m-%Y')})")
    else:
        # No sheet before selected date, just pick the latest overall
        latest_sheet = max(sheet_dates, key=lambda k: sheet_dates[k])
        latest_date = sheet_dates[latest_sheet]
        print(f"No previous sheet found. Using latest available sheet: {latest_sheet} ({latest_date.strftime('%d-%m-%Y')})")
else:
    # Selected date sheet does not exist, just take latest sheet
    latest_sheet = max(sheet_dates, key=lambda k: sheet_dates[k])
    latest_date = sheet_dates[latest_sheet]
    print(f"Latest Rating sheet to duplicate: {latest_sheet} ({latest_date.strftime('%d-%m-%Y')})")

# === Step 4: Load DataFrames ===
df_latest_data = pd.read_excel(
    client_rating_file, 
    sheet_name=latest_sheet, 
    skiprows=1,
    engine="openpyxl"
)
print("Client Rating DataFrame loaded successfully.")

df_summary = pd.read_excel(
    summary_file, 
    sheet_name="SUMMARY", 
    skiprows=2, 
    engine="pyxlsb"
)
print("Summary DataFrame loaded successfully.")

# df_unutilized_amount = pd.read_excel(
#     unutilized_amount_file,
#     usecols=[0, 2, 6],  # Adjust columns as needed
#     engine="openpyxl"
# )
# print("Unutilized Amount DataFrame loaded successfully.++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
# print(df_unutilized_amount.columns.tolist())
# print(df_unutilized_amount)

df_prod_wise_class = pd.read_excel(
    prod_wise_class_file, 
    sheet_name="NBD-MF-23-C2", 
    skiprows=4,
    usecols=[1,2,3,4,5,6], # B, C, D, E, F, G
    engine="pyxlsb"
)
print(df_prod_wise_class.columns.tolist())
print(df_prod_wise_class)
print("prod wise class DataFrame loaded successfully.")

# Get all sheet names first
with pd.ExcelFile(client_rating_file, engine="openpyxl") as xls:
    matching_sheets = [s for s in xls.sheet_names if s.startswith("Initial Risk Rating")]

    if not matching_sheets:
        raise ValueError("No worksheet found starting with 'Initial Risk Rating'.")

    target_sheet = matching_sheets[0] 
    print(f"Loading sheet: {target_sheet}")

    df_initial_risk_rating = pd.read_excel(
        xls,
        sheet_name=target_sheet,
        skiprows=1,
        engine="openpyxl"
    )

print("Initial Risk Rating DataFrame loaded successfully.")

df_rating_buckets = pd.read_excel(
    client_rating_file, 
    sheet_name="Rating Buckets", 
    skiprows=1, 
    usecols=[1, 2, 3],  # B, C, D
    header=1,           # Use second row (index 1) as column names
    engine="openpyxl"
)

# Rename columns properly
df_rating_buckets.columns = ["Age Bucket From", "Age Bucket To", "Rating"]

print("Rating Buckets DataFrame loaded successfully.")
print(df_rating_buckets.head())

# Remove rows where CONTRACT NO is missing
df_summary.dropna(subset=["CONTRACT NO"], inplace=True)

print("==============================================================================")
print(f"latest data without new contracts: {df_latest_data}")
print(f"DataFrame shape: {df_latest_data.shape}")
print("==============================================================================")

# === Step 5: Normalize column names ===
df_latest_data.columns = df_latest_data.columns.str.strip().str.replace("\n", " ").str.replace("\r", "", regex=True)
df_summary.columns = df_summary.columns.str.strip().str.replace("\n", " ").str.replace("\r", "", regex=True)

# Identify the contract column dynamically in df_latest_data
possible_contract_cols = [col for col in df_latest_data.columns if "contract" in col.lower() and "no" in col.lower()]
if not possible_contract_cols:
    raise KeyError(f"No 'Contract No' column found in Client Rating file. Columns: {df_latest_data.columns.tolist()}")
contract_col_latest = possible_contract_cols[0]
print(f"Detected contract column in latest data: {contract_col_latest}")

# Identify the contract column in df_summary
if "CONTRACT NO" in df_summary.columns:
    summary_contract_col = "CONTRACT NO"
else:
    possible_summary_cols = [col for col in df_summary.columns if "contract" in col.lower() and "no" in col.lower()]
    if not possible_summary_cols:
        raise KeyError(f"No 'CONTRACT NO' column found in Summary file. Columns: {df_summary.columns.tolist()}")
    summary_contract_col = possible_summary_cols[0]

# Normalize contract columns
df_latest_data[contract_col_latest] = df_latest_data[contract_col_latest].astype(str).str.strip()
df_summary[summary_contract_col] = df_summary[summary_contract_col].astype(str).str.strip()

# 1️⃣ Identify closed contracts (in latest_data but missing in summary)
closed_contracts = df_latest_data[~df_latest_data[contract_col_latest].isin(df_summary[summary_contract_col])]

# === Compute next-quarter title ===

# Determine the base report date
# If you already have 'report_date' from earlier in the script, reuse it
# Otherwise, try to pull from df_latest_data or system date
if "report_date" in locals() and pd.notna(report_date):
    base_date = report_date
else:
    # Fallback: try to find any date column in df_latest_data
    possible_date_cols = [col for col in df_latest_data.columns if "date" in col.lower()]
    if possible_date_cols:
        base_date = pd.to_datetime(df_latest_data[possible_date_cols[0]].dropna().iloc[0], errors='coerce')
    else:
        # Final fallback: use today's date
        base_date = pd.Timestamp.today()
    print(f"[INFO] Base date not provided, using {base_date.strftime('%Y-%m-%d')}")

# Ensure valid Timestamp
if pd.isna(base_date):
    raise ValueError("Could not determine a valid base date for next quarter calculation.")

# Compute next-quarter date
from dateutil.relativedelta import relativedelta
next_quarter_date = base_date + relativedelta(months=3)

# Build merged title
month_index = int(next_quarter_date.month)
month_name = calendar.month_name[month_index]
merged_title = f"{month_name} - {next_quarter_date.year}"
print(f"Next quarter title: {merged_title}")



if not closed_contracts.empty:
    # Create header row
    header_row = pd.DataFrame([[merged_title] + [""] * (len(closed_contracts.columns) - 1)],
                              columns=closed_contracts.columns)
    df_closed_contracts = pd.concat([header_row, closed_contracts], ignore_index=True)
    
    # Remove closed contracts from latest data
    df_latest_data = df_latest_data[df_latest_data[contract_col_latest].isin(df_summary[summary_contract_col])]
    print(f"Removed {len(closed_contracts)} closed contracts from latest data.")
else:
    df_closed_contracts = pd.DataFrame(columns=df_summary.columns)
    print("No closed contracts found.")

# 2️⃣ Identify new contracts (in summary but missing in latest_data)
new_contracts = df_summary[~df_summary[summary_contract_col].isin(df_latest_data[contract_col_latest])]

if not new_contracts.empty:
    # Keep both CONTRACT NO and CLIENT NO for mapping
    new_contracts_to_append = new_contracts[[summary_contract_col, "CLIENT NO"]].copy()
    new_contracts_to_append.rename(columns={summary_contract_col: contract_col_latest}, inplace=True)

    # Add empty columns for all other columns to align with df_latest_data
    for col in df_latest_data.columns:
        if col not in new_contracts_to_append.columns:
            new_contracts_to_append[col] = ""

    # Reorder to match df_latest_data column order
    new_contracts_to_append = new_contracts_to_append[df_latest_data.columns]

    # Track the starting index of new contracts for later reference
    new_contracts_start_idx = len(df_latest_data)

    # Append new contracts to df_latest_data
    df_latest_data = pd.concat([df_latest_data, new_contracts_to_append], ignore_index=True)
    print(f"Appended {len(new_contracts)} new contracts to latest data (rows {new_contracts_start_idx} to {len(df_latest_data)-1}).")

    # === Perform VLOOKUP-style mapping to append data ===

    # 1 Update Client No
    if "CLIENT NO" in df_summary.columns and "Client No" in df_latest_data.columns:
        client_lookup = dict(
            zip(
                df_summary[summary_contract_col].astype(str).str.strip(),
                df_summary["CLIENT NO"]
                .apply(lambda x: str(int(x)) if pd.notna(x) and str(x).replace('.', '', 1).isdigit() else str(x).strip())
            )
        )

        # Apply lookup to fill missing Client No
        df_latest_data["Client No"] = df_latest_data.apply(
            lambda row: row["Client No"] if pd.notna(row["Client No"]) and str(row["Client No"]).strip() != ""
            else client_lookup.get(str(row[contract_col_latest]).strip(), ""),
            axis=1
        )

        print("Client No column updated for newly added contracts using summary data.")
    else:
        print("Warning: CLIENT NO or Client No column not found. Skipping Client No update.")

    # 2 Update Contract Wise DPD and Customer Wise DPD

    if "CLIENT DPD" in df_summary.columns and "Contract Wise DPD" in df_latest_data.columns:
    # Build lookup dictionary (remove .0 if numeric)
        dpd_lookup = dict(
            zip(
                df_summary[summary_contract_col].astype(str).str.strip(),
                df_summary["CLIENT DPD"].apply(
                    lambda x: str(int(x)) if pd.notna(x) and str(x).replace('.', '', 1).isdigit() else str(x).strip()
                )
            )
        )

        # Update Contract Wise DPD
        df_latest_data["Contract Wise DPD"] = df_latest_data.apply(
            lambda row: row["Contract Wise DPD"] if pd.notna(row["Contract Wise DPD"]) and str(row["Contract Wise DPD"]).strip() != ""
            else dpd_lookup.get(str(row[contract_col_latest]).strip(), ""),
            axis=1
        )

        # Update Customer Wise DPD
        df_latest_data["Customer Wise DPD"] = df_latest_data.apply(
            lambda row: row["Customer Wise DPD"] if pd.notna(row["Customer Wise DPD"]) and str(row["Customer Wise DPD"]).strip() != ""
            else dpd_lookup.get(str(row[contract_col_latest]).strip(), ""),
            axis=1
        )

        print("Contract Wise DPD and Customer Wise DPD columns updated for newly added contracts using summary data (no decimals).")
    else:
        print("Warning: Contract Wise DPD or Customer Wise DPD not found. Skipping Contract Wise DPD and Customer Wise DPD update.")


    # 3 Update Yard
    if "YARD CONTRACTS" in df_summary.columns and "Yard" in df_latest_data.columns:
        yard_lookup = dict(zip(df_summary[summary_contract_col], df_summary["YARD CONTRACTS"]))

        df_latest_data["Yard"] = df_latest_data.apply(
            lambda row: row["Yard"] if pd.notna(row["Yard"]) and str(row["Yard"]).strip() != ""
                        else yard_lookup.get(row[contract_col_latest], ""),
            axis=1
        )
        print("Yard column updated for newly added contracts using summary data.")
    else:
        print("Warning: YARD CONTRACTS or Yard column not found. Skipping Yard update.")

    # 4 Update P/NP
    if "P/NP" in df_summary.columns and "P/NP" in df_latest_data.columns:
        pnp_lookup = dict(zip(df_summary[summary_contract_col], df_summary["P/NP"]))

        df_latest_data["P/NP"] = df_latest_data.apply(
            lambda row: row["P/NP"] if pd.notna(row["P/NP"]) and str(row["P/NP"]).strip() != ""
                        else pnp_lookup.get(row[contract_col_latest], ""),
            axis=1
        )
        print("P/NP column updated for newly added contracts using summary data.")
    else:
        print("Warning: P/NP column not found. Skipping P/NP update.")

    # 5 Update Stage
    if "STAGE" in df_summary.columns and "Stage" in df_latest_data.columns:
        # Build lookup dictionary with clean, decimal-free Stage values
        stage_lookup = dict(
            zip(
                df_summary[summary_contract_col].astype(str).str.strip(),
                df_summary["STAGE"].apply(
                    lambda x: str(int(x)) if pd.notna(x) and str(x).replace('.', '', 1).isdigit() else str(x).strip()
                )
            )
        )

        # Fill missing Stage values in df_latest_data (VLOOKUP equivalent)
        df_latest_data["Stage"] = df_latest_data.apply(
            lambda row: row["Stage"] if pd.notna(row["Stage"]) and str(row["Stage"]).strip() != ""
            else stage_lookup.get(str(row[contract_col_latest]).strip(), ""),
            axis=1
        )

        print("Stage column updated for newly added contracts using summary data (no decimals).")
    else:
        print("Warning: STAGE or Stage column not found. Skipping Stage update.")

    
    # === Append Quarter Added only for new contracts ===
    # --- Robust Quarter label calculation from a sheet date and safe assignment ---

    def quarter_label_from_sheet_date(sheet_date):
        """
        Given a sheet end date (pd.Timestamp), return the label for the NEXT quarter
        in format like "Apr - Jun 25".
        Assumes sheet_date is a quarter-end date (month one of 3,6,9,12).
        """
        sheet_date = pd.to_datetime(sheet_date)
        # map quarter end month -> quarter index (1..4 as per your specific mapping)
        # Q1 (Apr-Jun) ends 30-Jun -> end month 6 maps to Q1
        mapping = {6: 1, 9: 2, 12: 3, 3: 4}
        m = sheet_date.month
        if m not in mapping:
            # fallback: compute quarter index by month -> determine which quarter the date represents
            qidx = ((m - 1) // 3) + 1
            # adjust qidx so that it matches your quarter mapping (Apr-Jun = 1 etc.)
            # This fallback should rarely be needed if sheet names are standard quarter-end dates.
        else:
            qidx = mapping[m]

        # next quarter index (1..4)
        next_q = (qidx % 4) + 1

        # map quarter index to start month
        q_start_month = {1: 4, 2: 7, 3: 10, 4: 1}[next_q]
        q_end_month = q_start_month + 2

        # determine year for the next quarter:
        # if next quarter's start month is <= sheet_date.month, it means next quarter is in next calendar year
        if q_start_month <= sheet_date.month:
            q_year = sheet_date.year + 1
        else:
            q_year = sheet_date.year

        month_abbrs = [None] + list(calendar.month_abbr[1:])  # 1-indexed
        start_abbr = month_abbrs[q_start_month]
        end_abbr = month_abbrs[q_end_month]
        year_suffix = str(q_year)[-2:]

        return f"{start_abbr} - {end_abbr} {year_suffix}"

    # --- Use the sheet date of the latest rating sheet you selected earlier ---
    # assume latest_sheet and sheet_dates dict exist and latest_sheet_date = sheet_dates[latest_sheet]
    latest_sheet_date = sheet_dates[latest_sheet]  # this is a pd.Timestamp from your sheet detection
    quarter_label = quarter_label_from_sheet_date(latest_sheet_date)

    # Ensure the column exists
    if "Quarter Added" not in df_latest_data.columns:
        df_latest_data["Quarter Added"] = ""

    # Mark only rows that are currently blank (i.e., truly new rows you just appended)
    mask_new = df_latest_data["Quarter Added"].isna() | (df_latest_data["Quarter Added"].astype(str).str.strip() == "")

    # Assign quarter label only to those blank rows
    df_latest_data.loc[mask_new, "Quarter Added"] = quarter_label

    print(f"Quarter Added updated for {mask_new.sum()} new rows with label: {quarter_label}")


    # Sample rating bucket data - matches 'Rating Buckets'!$B$3:$D$8
    # Excel VLOOKUP with TRUE does approximate match on SORTED first column
    # It finds the largest value less than or equal to lookup value
    rating_data = [
        (-1, 0, "A+"),     # Row 3: DPD from -1 to 0
        (1, 30, "A"),      # Row 4: DPD from 1 to 30
        (31, 60, "B+"),    # Row 5: DPD from 31 to 60
        (61, 90, "B"),     # Row 6: DPD from 61 to 90
        (91, 120, "C"),    # Row 7: DPD from 91 to 120 (D7 = "C")
        (121, 5000, "D")   # Row 8: DPD from 121+
    ]
    df_rating_buckets = pd.DataFrame(rating_data, columns=["Age Bucket From", "Age Bucket To", "Rating"])

    # Clean and prepare df_latest_data columns
    df_latest_data["Yard"] = df_latest_data["Yard"].astype(str).str.strip().str.upper()
    df_latest_data["Contract Wise DPD"] = pd.to_numeric(df_latest_data["Contract Wise DPD"], errors="coerce").fillna(0)

    # Excel formula: =IF(AND(E81318="YARD",G81318<121),'Rating Buckets'!$D$7,VLOOKUP(G81318,'Rating Buckets'!$B$3:$D$8,3,TRUE))
    # E81318 = Yard (column E)
    # G81318 = Contract Wise DPD (column G)
    # $D$7 = "C" (4th row in the rating_data, which is row 7 in Excel with headers)
    # VLOOKUP with TRUE = approximate match (finds largest value <= lookup value)

    def compute_rating(row):
        yard = row["Yard"]
        dpd = row["Contract Wise DPD"]

        # IF(AND(E="YARD", G<121), 'Rating Buckets'!$D$7, ...)
        # If Yard is "YARD" AND DPD < 121, return "C" (from D7)
        if yard == "YARD" and dpd < 121:
            return "C"

        # VLOOKUP(G, 'Rating Buckets'!$B$3:$D$8, 3, TRUE)
        # VLOOKUP with TRUE does approximate match:
        # - Lookup column (B) must be sorted in ascending order
        # - Finds the largest value in column B that is <= dpd
        # - Returns corresponding value from column 3 (D = Rating)

        # Filter buckets where "Age Bucket From" <= dpd and sort by "Age Bucket From" descending
        eligible_buckets = df_rating_buckets[df_rating_buckets["Age Bucket From"] <= dpd]

        if not eligible_buckets.empty:
            # Get the bucket with the largest "Age Bucket From" that's <= dpd
            matched_bucket = eligible_buckets.sort_values("Age Bucket From", ascending=False).iloc[0]
            return matched_bucket["Rating"]

        # If no match (shouldn't happen with -1 as minimum), return empty
        return ""

    # Apply logic to dataframe
    df_latest_data["Rating"] = df_latest_data.apply(compute_rating, axis=1)

    print("Rating column appended successfully.")


    # === Append or update Initial Rating column dynamically ==================================

    # Identify sheets
    sheets = pd.ExcelFile(client_rating_file).sheet_names

    # Dynamically detect quarter sheets
    q1_sheet = next((s for s in sheets if "30-06" in s), None)
    q2_sheet = next((s for s in sheets if "30-09" in s), None)
    q3_sheet = next((s for s in sheets if "30-12" in s or "31-12" in s), None)
    q4_sheet = next((s for s in sheets if "31-03" in s or "30-03" in s), None)
    initial_risk_sheet = next((s for s in sheets if "Initial Risk Rating" in s), None)

    print(f"Found sheets - Initial: {initial_risk_sheet}, Q1: {q1_sheet}, Q2: {q2_sheet}, Q3: {q3_sheet}, Q4: {q4_sheet}")

    # Read helper
    def read_rating_sheet(sheet_name):
        return pd.read_excel(client_rating_file, sheet_name=sheet_name, skiprows=1) if sheet_name else pd.DataFrame()

    # Read sheets
    df_initial_risk_rating = read_rating_sheet(initial_risk_sheet)
    df_Q1 = read_rating_sheet(q1_sheet)
    df_Q2 = read_rating_sheet(q2_sheet)
    df_Q3 = read_rating_sheet(q3_sheet)
    df_Q4 = read_rating_sheet(q4_sheet)

    # Combine all available rating DataFrames
    rating_frames = [df_initial_risk_rating, df_Q1, df_Q2, df_Q3, df_Q4]
    rating_frames = [df for df in rating_frames if not df.empty]

    if rating_frames:
        df_all_ratings = pd.concat(rating_frames, ignore_index=True)

        # Standardize and clean
        df_all_ratings = df_all_ratings.rename(columns=lambda x: x.strip())
        if "Contract No" not in df_all_ratings.columns or "Rating" not in df_all_ratings.columns:
            raise ValueError("Missing required columns ('Contract No', 'Rating') in rating sheets.")

        df_all_ratings = df_all_ratings[["Contract No", "Rating"]].dropna(subset=["Contract No"])
        df_all_ratings["Contract No"] = df_all_ratings["Contract No"].astype(str).str.strip()
        df_all_ratings["Rating"] = df_all_ratings["Rating"].astype(str).str.strip()

        # Remove duplicates, latest wins
        df_all_ratings = df_all_ratings.drop_duplicates(subset="Contract No", keep="last")

        # Mapping dictionary
        contract_to_rating = dict(zip(df_all_ratings["Contract No"], df_all_ratings["Rating"]))

        # Ensure Initial Rating column exists
        if "Initial Rating" not in df_latest_data.columns:
            df_latest_data["Initial Rating"] = ""

        df_latest_data["Contract No"] = df_latest_data["Contract No"].astype(str).str.strip()

        before_blank_count = df_latest_data["Initial Rating"].eq("").sum() + df_latest_data["Initial Rating"].isna().sum()

        # Step 1: Map from concatenated rating DataFrames
        mapped_ratings = df_latest_data["Contract No"].map(contract_to_rating)

        # Step 2: Fill only blanks from mapped ratings
        df_latest_data["Initial Rating"] = df_latest_data["Initial Rating"].mask(
            df_latest_data["Initial Rating"].isna() | (df_latest_data["Initial Rating"].astype(str).str.strip() == ""),
            mapped_ratings
        )

        # Step 3: If still missing, fallback to same-row Rating column
        df_latest_data["Initial Rating"] = df_latest_data["Initial Rating"].mask(
            df_latest_data["Initial Rating"].isna() | (df_latest_data["Initial Rating"].astype(str).str.strip() == ""),
            df_latest_data["Rating"]
        )

        after_blank_count = df_latest_data["Initial Rating"].eq("").sum() + df_latest_data["Initial Rating"].isna().sum()
        filled_count = before_blank_count - after_blank_count

        print(f"Initial Rating column appended or updated successfully. {filled_count} cells filled.")

    else:
        print("No valid rating sheets found to update Initial Rating.")

else:
    print("No new contracts to append.")

# === Create Stage-wise Movement Summary (using Initial Rating) ===================================

# Ensure necessary columns exist
required_cols = {"Stage", "Initial Rating"}
if not required_cols.issubset(df_latest_data.columns):
    raise ValueError(f"Missing required columns: {required_cols - set(df_latest_data.columns)}")

# Convert data types safely
df_latest_data["Stage"] = pd.to_numeric(df_latest_data["Stage"], errors="coerce")
df_latest_data["Initial Rating"] = df_latest_data["Initial Rating"].astype(str).str.strip()

# Define expected categories for clarity and ordering
rating_order = ["A+", "A", "B+", "B", "C", "D"]
stage_order = [1, 2, 3]

# Pivot table to count contracts by Initial Rating & Stage
pivot = (
    df_latest_data
    .pivot_table(index="Initial Rating", columns="Stage", values="Contract No", aggfunc="count", fill_value=0)
    .reindex(index=rating_order, columns=stage_order, fill_value=0)
)

# Add Grand Total column (sum across stages)
pivot["Grand Total"] = pivot.sum(axis=1)

# Add Grand Total row (sum across all ratings)
grand_total = pivot.sum(axis=0).to_frame().T
grand_total.index = ["Grand Total"]

# Combine pivot with totals
df_stage_wise_movement = pd.concat([pivot, grand_total], axis=0).reset_index()

# Rename index column for presentation
df_stage_wise_movement = df_stage_wise_movement.rename(columns={"Initial Rating": "Row Labels"})

# Format numbers with commas for readability
df_stage_wise_movement = df_stage_wise_movement.applymap(
    lambda x: f"{int(x):,}" if isinstance(x, (int, float)) and not pd.isna(x) else x
)

# Final preview
print("Stage-wise movement table created successfully (using Initial Rating).")
print(df_stage_wise_movement)


# === Create Rating Movement Summary (Initial Rating → Rating) ===================================

# Ensure necessary columns exist
required_cols = {"Initial Rating", "Rating"}
if not required_cols.issubset(df_latest_data.columns):
    raise ValueError(f"Missing required columns: {required_cols - set(df_latest_data.columns)}")

# Clean and standardize columns
df_latest_data["Initial Rating"] = df_latest_data["Initial Rating"].astype(str).str.strip()
df_latest_data["Rating"] = df_latest_data["Rating"].astype(str).str.strip()

# Define expected rating categories for ordering
rating_order = ["A+", "A", "B+", "B", "C", "D"]

# Create pivot table to count contract transitions
pivot = (
    df_latest_data
    .pivot_table(
        index="Initial Rating",
        columns="Rating",
        values="Contract No",
        aggfunc="count",
        fill_value=0
    )
    .reindex(index=rating_order, columns=rating_order, fill_value=0)
)

# Add Grand Total column (row sum)
pivot["Grand Total"] = pivot.sum(axis=1)

# Add Grand Total row (column sum)
grand_total = pivot.sum(axis=0).to_frame().T
grand_total.index = ["Grand Total"]

# Combine the pivot table and totals
df_rating_movement = pd.concat([pivot, grand_total], axis=0).reset_index()

# Rename index column for clarity
df_rating_movement = df_rating_movement.rename(columns={"Initial Rating": "Row Labels"})

# Format numeric values with commas for readability
df_rating_movement = df_rating_movement.applymap(
    lambda x: f"{int(x):,}" if isinstance(x, (int, float)) and not pd.isna(x) else x
)

# Final output
print("Rating movement table created successfully (Initial Rating → Rating). ==============================================")
print(df_rating_movement)



# # === Clean and Filter Unutilized Amount Data ================================
# # 1. Standardize column names (strip spaces)
# df_unutilized_amount.columns = df_unutilized_amount.columns.str.strip()

# # 2. Remove the last 2 rows
# df_unutilized_amount = df_unutilized_amount.iloc[:-2]

# # 3. Ensure numeric conversion for "(ICAM) D/C" column
# df_unutilized_amount["(ICAM) D/C"] = pd.to_numeric(df_unutilized_amount["(ICAM) D/C"], errors="coerce")

# # 4. Filter rows where "(ICAM) D/C" is non-zero
# df_icam_nonzero = df_unutilized_amount[df_unutilized_amount["(ICAM) D/C"] != 0]

# # 5. Count unique non-zero contracts (based on "ICAM code")
# no_of_mt_clients = df_icam_nonzero["ICAM code"].nunique()

# print(f"Number of non-zero (ICAM) D/C contracts: {no_of_mt_clients}")

# === Extract Key Metrics from df_prod_wise_class ============================

# Clean column names and key text values
df_prod_wise_class.columns = df_prod_wise_class.columns.str.strip()
df_prod_wise_class["Product Type"] = df_prod_wise_class["Product Type"].astype(str).str.strip()

# Extract specific values using case-insensitive matching
def get_stage_value(product_name, stage_col):
    match = df_prod_wise_class.loc[
        df_prod_wise_class["Product Type"].str.lower() == product_name.lower(), stage_col
    ]
    if not match.empty and pd.notna(match.values[0]):
        return int(round(match.values[0]))  # convert to int (remove decimals)
    else:
        return 0
    

# Assign variables
No_of_MT_Clients = get_stage_value("Margin Trading Loans", "Stage 01")
FD_Against_Debtor = get_stage_value("Loans against Cash/Deposits", "Stage 01")
O17 = get_stage_value("Total Loans", "Stage 01")
P17 = get_stage_value("Total Loans", "Stage 02")
Q17 = get_stage_value("Total Loans", "Stage 03")
N17 = get_stage_value("Total Loans", "Total")

# Print for verification
print("Extracted Values from df_prod_wise_class:")
print(f"No_of_MT_Clients: {No_of_MT_Clients}")
print(f"FD_Against_Debtor: {FD_Against_Debtor}")
print(f"O17 (Total Loans - Stage 01): {O17}")
print(f"P17 (Total Loans - Stage 02): {P17}")
print(f"Q17 (Total Loans - Stage 03): {Q17}")
print(f"N17 (Total Loans - Total): {N17}")


# =================================   save workbook ======================================================================

# === Step 1: Accept the date input from command or UI ===
# Example usage: python script.py 30-09-2025
if args.date:
    input_date_str = args.date
else:
    raise ValueError("Please provide a date argument in MM/DD/YYYY format")

try:
    # Parse MM/DD/YYYY format (slashes)
    report_date = pd.to_datetime(input_date_str, format='%m/%d/%Y', errors='raise')
except ValueError:
    raise ValueError("Invalid date format. Please use MM/DD/YYYY format.")

new_sheet_name = f"Rating -{report_date.strftime('%m-%d-%Y')}"
print(f"Target new sheet name: {new_sheet_name}")

new_sheet_name = f"Rating -{report_date.strftime('%m-%d-%Y')}"
print(f"Target new sheet name: {new_sheet_name}")

# === Load workbook ===
wb = load_workbook(client_rating_file)
sheet_names = wb.sheetnames

# === Check if target sheet already exists ===
if new_sheet_name in sheet_names:
    print(f"Sheet '{new_sheet_name}' already exists. Using existing sheet to update data.")
    ws_new = wb[new_sheet_name]
else:
    # === Identify Rating sheets by date pattern ===
    date_pattern = re.compile(r"Rating\s*-\s*(\d{2}-\d{2}-\d{4})")
    sheet_dates = {}

    for sheet in sheet_names:
        match = date_pattern.search(sheet)
        if match:
            try:
                date = pd.to_datetime(match.group(1), format="%d-%m-%Y")
                sheet_dates[sheet] = date
            except Exception:
                continue

    if not sheet_dates:
        raise ValueError("No valid Rating sheets found with date format 'Rating -dd-mm-yyyy'")

    # === Find oldest and latest sheets ===
    latest_sheet = max(sheet_dates, key=lambda k: sheet_dates[k])
    oldest_sheet = min(sheet_dates, key=lambda k: sheet_dates[k])
    print(f"Latest Rating sheet: {latest_sheet}")
    print(f"Oldest Rating sheet: {oldest_sheet}")

    # === Delete oldest sheet ===
    print(f"Deleting oldest sheet: {oldest_sheet}")
    std = wb[oldest_sheet]
    wb.remove(std)

    # === Duplicate latest sheet and rename ===
    print(f"Duplicating '{latest_sheet}' as '{new_sheet_name}'")
    source_ws = wb[latest_sheet]
    ws_new = wb.copy_worksheet(source_ws)
    ws_new.title = new_sheet_name

# === Step 2: Update date in cell A1 (always run) ===
ws_new["A1"] = report_date.strftime("%d %B %Y")
print(f"[INFO] Written UI date '{report_date.strftime('%d %B %Y')}' in '{ws_new.title}'!A1")

# === Step 3: Clear data from row 3 onwards (columns A:J) (always run) ===
for row in ws_new.iter_rows(min_row=3, max_col=10):
    for cell in row:
        cell.value = None
print(f"[INFO] Cleared old data in '{ws_new.title}' from row 3 onwards")

# === Step 4: Paste df_latest_data starting from row 3 (always run) ===
# FIXED: Write to specific cells instead of using append()
for r_idx, row_data in enumerate(dataframe_to_rows(df_latest_data, index=False, header=False), start=3):
    for c_idx, value in enumerate(row_data, start=1):
        ws_new.cell(row=r_idx, column=c_idx, value=value)
print(f"[INFO] Pasted {len(df_latest_data)} rows of df_latest_data into '{ws_new.title}' starting at A3")

# === Step 5: Update Summary sheet (always run) ===
ws_summary = wb["Summary"]

# Clear old data from row 4 onwards (A:J)
for row in ws_summary.iter_rows(min_row=4, max_col=10):
    for cell in row:
        cell.value = None
print(f"[INFO] Cleared old data in 'Summary' from row 4 onwards")

# Paste new data (without header) from row 4
# FIXED: Write to specific cells instead of using append()
for r_idx, row_data in enumerate(dataframe_to_rows(df_latest_data, index=False, header=False), start=4):
    for c_idx, value in enumerate(row_data, start=1):
        ws_summary.cell(row=r_idx, column=c_idx, value=value)
print(f"[INFO] Pasted {len(df_latest_data)} rows into 'Summary' starting at A4")

# === Step 6: Update Pivot sheet (always run) ===
ws_pivot = wb["Pivot"]

def safe_number(val):
    """Convert numeric-like strings to float or int, otherwise return as-is."""
    if isinstance(val, str):
        val = val.strip()
        if val.replace('.', '', 1).isdigit():
            # If it's a whole number, cast to int
            return int(float(val)) if '.' not in val else float(val)
        return val
    return val

# === Clear and write df_stage_wise_movement (A5:E11) ===
for row in ws_pivot.iter_rows(min_row=5, max_row=11, min_col=1, max_col=5):
    for cell in row:
        cell.value = None

for r_idx, row in enumerate(dataframe_to_rows(df_stage_wise_movement, index=False, header=False), start=5):
    for c_idx, val in enumerate(row[:5], start=1):
        ws_pivot.cell(row=r_idx, column=c_idx, value=safe_number(val))

print(f"[INFO] Updated Pivot sheet A5:E11 with {len(df_stage_wise_movement)} rows")

# === Clear and write df_rating_movement (A18:H24) ===
for row in ws_pivot.iter_rows(min_row=18, max_row=24, min_col=1, max_col=8):
    for cell in row:
        cell.value = None

for r_idx, row in enumerate(dataframe_to_rows(df_rating_movement, index=False, header=False), start=18):
    for c_idx, val in enumerate(row[:8], start=1):
        ws_pivot.cell(row=r_idx, column=c_idx, value=safe_number(val))

print(f"[INFO] Updated Pivot sheet A18:H24 with {len(df_rating_movement)} rows")

# === Step: Append Closed Contracts data (always run) ===
ws_closed_contracts = wb["Closed Contracts"]

# Find the last used row in column A
last_row_closed = ws_closed_contracts.max_row

# If the sheet is empty (only header in A1), start from row 2
start_row = last_row_closed + 1 if last_row_closed >= 1 else 2

# Prepare data to paste (without headers)
closed_contracts_data = df_closed_contracts.values.tolist()

# Paste data starting from next empty row
for r_idx, row in enumerate(closed_contracts_data, start=start_row):
    for c_idx, value in enumerate(row, start=1):
        ws_closed_contracts.cell(row=r_idx, column=c_idx, value=value)

# Log summary
print(f"[INFO] Appended {len(closed_contracts_data)} rows to 'Closed Contracts' starting at row {start_row}.")

# === Save workbook ===
wb.save(client_rating_file)
print(f"Workbook saved successfully with sheet '{new_sheet_name}'")

# ======== Close workbook and Excel instance ========
wb.close()




# === FINAL STEP: Output confirmation ===
# print("Data processing completed.")
# print(f"closed contracts: {df_closed_contracts}")
# print(f"\nClosed Contracts DataFrame: {df_closed_contracts}")
# print(f"summary: {df_summary}")
# print(f"Summary DataFrame shape: {df_summary.shape}")
# print(f"latest data with new contracts: {df_latest_data}")
# print(f"Client Rating DataFrame shape (after filtering): {df_latest_data.shape}")
# print(df_initial_risk_rating)
# print(f"Initial Risk Rating DataFrame shape: {df_initial_risk_rating.shape}")
# print(df_rating_buckets)
# print(f"Rating Buckets DataFrame shape: {df_rating_buckets.shape}")
# print(df_initial_risk_rating.head())
print("Script execution completed successfully.")