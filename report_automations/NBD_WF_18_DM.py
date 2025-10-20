import os
from pathlib import Path
import re
import shutil
import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# ------------------------------------------------------------
# STEP 1 – Find and rename NBD-WF-18-DM Deposit Liability file
# ------------------------------------------------------------
def rename_wf18dm_report(working_dir, report_name, old_week, new_week):
    """Find and rename the NBD-WF-18-DM file in the working date folder"""
    working_path = Path(working_dir)

    # Search for the template file directly in the working folder
    patterns = [
        f"{report_name}*.xlsx",
        f"{report_name}*.xlsb",
        f"{report_name}*.xls",
        "NBD*WF*18*DM*.xlsx",
        "NBD*WF*18*DM*.xlsb",
        "NBD*WF*18*DM*.xls"
    ]

    template_file = None
    for pattern in patterns:
        files = list(working_path.glob(pattern))
        if files:
            template_file = files[0]
            break

    if not template_file:
        print(f"⚠️ No template file found for {report_name}. Creating a basic template...")
        # Create a basic template in the working folder
        template_file = working_path / f"{report_name} Week {old_week}.xlsx"

        # Create a basic workbook with required sheets
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create required sheets
        wb.create_sheet("Summary")
        wb.create_sheet("Rec")
        wb.create_sheet("Detailed")
        wb.create_sheet("Portfolio")

        wb.save(template_file)
        print(f"✅ Created basic template: {template_file.name}")

    # Rename the file in the same directory
    new_filename = re.sub(
        fr"Week\s*{old_week}\b",
        f"Week {new_week}",
        template_file.name
    )
    if new_filename == template_file.name:
        # If no week pattern found, add week number
        stem, suffix = os.path.splitext(template_file.name)
        new_filename = f"{stem} Week {new_week}{suffix}"

    new_path = working_path / new_filename

    # Rename the file
    if template_file != new_path:
        shutil.move(template_file, new_path)
        print(f"✅ Renamed: {template_file.name} → {new_filename}")
    else:
        print(f"✅ File already has correct name: {new_filename}")

    return new_path

# ------------------------------------------------------------
# STEP 2 & 3 – Update Summary & Rec last week values
# ------------------------------------------------------------
def update_summary_and_rec(filepath):
    wb = openpyxl.load_workbook(filepath)
    
    # Summary sheet
    if "Summary" in wb.sheetnames:
        ws_summary = wb["Summary"]
        for col in range(1, ws_summary.max_column + 1):
            ws_summary.cell(row=16, column=col).value = ws_summary.cell(row=11, column=col).value
        print("✅ Copied Summary row 11 → row 16")
    
    # Rec sheet
    if "Rec" in wb.sheetnames:
        ws_rec = wb["Rec"]
        for col in range(1, 5):  # A-D
            ws_rec.cell(row=13, column=col).value = ws_rec.cell(row=10, column=col).value
        print("✅ Copied Rec row 10 (A–D) → row 13")
    
    wb.save(filepath)
    wb.close()
    print(f"💾 Summary & Rec updated in {os.path.basename(filepath)}")

# ------------------------------------------------------------
# STEP 4 – Paste data from CentralBankMovementOfDailyDepositBalances
# ------------------------------------------------------------
def paste_from_centralbank_report(target_file, working_dir):
    """Look for CSV/Excel files directly in the working date folder"""
    working_path = Path(working_dir)

    # Locate CentralBankMovement report (Excel or CSV) directly in the working folder
    cb_files = list(working_path.glob("CentralBankMovementOfDailyDepositBalances*.xls*")) + \
               list(working_path.glob("CentralBankMovementOfDailyDepositBalances*.csv"))
    mod_files = list(working_path.glob("MovementOfDepositDetails*.xls*")) + \
                list(working_path.glob("MovementOfDepositDetails*.csv"))
    
    if not cb_files:
        raise FileNotFoundError("No file starting with 'CentralBankMovementOfDailyDepositBalances' found.")
    if not mod_files:
        raise FileNotFoundError("No file starting with 'MovementOfDepositDetails' found.")
    
    cb_file = cb_files[0]
    mod_file = mod_files[0]

    # Load target workbook
    wb_target = openpyxl.load_workbook(target_file)

    # Handle CSV files
    if cb_file.suffix.lower() == '.csv' and mod_file.suffix.lower() == '.csv':
        # Read CSV data
        def read_csv_rows(path):
            with open(path, newline="", encoding="utf-8-sig") as f:
                return list(csv.reader(f))

        cb_rows = read_csv_rows(cb_file)
        mod_rows = read_csv_rows(mod_file)

        # Summary CSV: copy rows 5-12, cols A-I (0..8) → A4:I11
        if "Summary" in wb_target.sheetnames:
            ws_tgt_summary = wb_target["Summary"]
            for r in range(8):
                for c in range(9):
                    source_val = cb_rows[4 + r][0 + c] if len(cb_rows) > 4 + r and len(cb_rows[4 + r]) > c else None
                    ws_tgt_summary.cell(row=4 + r, column=1 + c).value = source_val
            print("✅ Pasted Summary A5:I12 (CSV) → A4:I11")

        # Detailed CSV: copy rows 5-12, cols A-W (0..22) → A8:W15
        if "Detailed" in wb_target.sheetnames:
            ws_tgt_detailed = wb_target["Detailed"]
            for r in range(8):
                for c in range(23):
                    source_val = mod_rows[4 + r][0 + c] if len(mod_rows) > 4 + r and len(mod_rows[4 + r]) > c else None
                    ws_tgt_detailed.cell(row=8 + r, column=1 + c).value = source_val
            print("✅ Pasted Detailed A5:W12 (CSV) → A8:W15")

    else:
        # Handle Excel files
        wb_source = openpyxl.load_workbook(cb_file, data_only=True)

        # Summary sheet paste A5:I12 → Summary!A4:I11
        for sheet_name in wb_source.sheetnames:
            if "CentralBankMovementOfDailyDepositBalances" in sheet_name:
                ws_src_summary = wb_source[sheet_name]
                break
        else:
            raise ValueError("Source sheet 'CentralBankMovementOfDailyDepositBalances' not found.")

        if "Summary" in wb_target.sheetnames:
            ws_tgt_summary = wb_target["Summary"]
            for r in range(8):
                for c in range(9):
                    ws_tgt_summary.cell(row=4+r, column=1+c).value = ws_src_summary.cell(row=5+r, column=1+c).value
            print("✅ Pasted Summary A5:I12 → A4:I11")

        # Detailed sheet paste A5:W12 → Detailed!A8:W15
        for sheet_name in wb_source.sheetnames:
            if "MovementOfDepositDetails" in sheet_name:
                ws_src_detailed = wb_source[sheet_name]
                break
        else:
            raise ValueError("Source sheet 'MovementOfDepositDetails' not found.")

        if "Detailed" in wb_target.sheetnames:
            ws_tgt_detailed = wb_target["Detailed"]
            for r in range(8):
                for c in range(23):
                    ws_tgt_detailed.cell(row=8+r, column=1+c).value = ws_src_detailed.cell(row=5+r, column=1+c).value
            print("✅ Pasted Detailed A5:W12 → A8:W15")

        wb_source.close()

    wb_target.save(target_file)
    wb_target.close()
    print(f"💾 CentralBankMovement data pasted to {os.path.basename(target_file)}")

# ------------------------------------------------------------
# STEP 5 – Validate Detailed sheet
# ------------------------------------------------------------
def validate_detailed_sheet(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if "Detailed" not in wb.sheetnames:
        raise ValueError("'Detailed' sheet not found.")
    ws = wb["Detailed"]
    exceptions = []

    # Check A8:W15 columns not empty
    for col in range(1, 24):
        col_values = [ws.cell(row=r, column=col).value for r in range(8, 16)]
        if all(v is None or str(v).strip() == "" for v in col_values):
            col_letter = get_column_letter(col)
            exceptions.append(f"Column {col_letter} (A8:W15) is completely empty.")

    # Check O17:O24, W17:W24, E17:E24, J17:J24 = 0
    for cell_range in ["O17:O24", "W17:W24", "E17:E24", "J17:J24"]:
        for row in ws[cell_range]:
            for cell in row:
                val = cell.value
                if val not in (0, 0.0):
                    exceptions.append(f"Cell {cell.coordinate} = {val} (should be 0)")

    if exceptions:
        report_path = Path(filepath).parent / "NBD_WF_18_DM_Detailed_Exceptions.txt"
        with open(report_path, "w", encoding="utf-8") as f:
            f.write("⚠️ Detailed Sheet Exception Report ⚠️\n\n")
            for e in exceptions:
                f.write(f"- {e}\n")
        print(f"❗ {len(exceptions)} exceptions found. Report saved to {report_path}")
    else:
        print("✅ No exceptions in Detailed sheet.")

    wb.close()

# ------------------------------------------------------------
# STEP 6 – Copy dates from Summary → Rec
# ------------------------------------------------------------
def copy_summary_dates_to_rec(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws_summary = wb["Summary"]
    ws_rec = wb["Rec"]

    for i in range(8):  # Summary A4:A11 → Rec A3:A10
        date_val = ws_summary.cell(row=4+i, column=1).value
        ws_rec.cell(row=3+i, column=1).value = date_val
        ws_rec.cell(row=3+i, column=3).value = None
        ws_rec.cell(row=3+i, column=4).value = None

    wb.save(filepath)
    wb.close()
    print("✅ Dates copied to Rec sheet and columns C3:D10 cleared.")

# ------------------------------------------------------------
# STEP 7 – Copy FDBaseNew_UserBranch data with date formatting
# ------------------------------------------------------------
def parse_and_format_date(date_val, output_format="%d/%m/%Y"):
    """Convert various date formats to specified output format
    Args:
        date_val: Date value to convert
        output_format: Output format string (default: "%d/%m/%Y" for DD/MM/YYYY)
    """
    if date_val is None or date_val == "":
        return None
    
    # If already a datetime object
    if isinstance(date_val, datetime):
        return date_val.strftime(output_format)
    
    date_str = str(date_val).strip()
    
    # Skip if it's clearly not a date (too long or contains letters except month names)
    if len(date_str) > 50:
        return date_val
    
    # Try multiple date formats
    date_formats = [
        "%d/%m/%Y",      # 18/09/2025
        "%m/%d/%Y",      # 09/18/2025
        "%Y-%m-%d",      # 2025-09-18
        "%d-%m-%Y",      # 18-09-2025
        "%Y/%m/%d",      # 2025/09/18
        "%d.%m.%Y",      # 18.09.2025
        "%d %m %Y",      # 18 09 2025
        "%d/%m/%y",      # 18/09/25
        "%m/%d/%y",      # 09/18/25
        "%y/%m/%d",      # 25/09/18
    ]
    
    for fmt in date_formats:
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime(output_format)
        except ValueError:
            continue
    
    # Try to handle Excel serial dates (numeric)
    try:
        # Excel stores dates as numbers (days since 1900-01-01)
        excel_date = float(date_str)
        # Valid Excel date range: 1 (1900-01-01) to 2958465 (9999-12-31)
        if 1 <= excel_date <= 2958465:
            # Excel incorrectly considers 1900 a leap year, so adjust
            if excel_date < 60:
                dt = datetime(1899, 12, 31) + timedelta(days=excel_date)
            else:
                dt = datetime(1899, 12, 31) + timedelta(days=excel_date - 1)
            return dt.strftime(output_format)
    except (ValueError, TypeError, OverflowError):
        pass
    
    # Return original if all parsing fails
    return date_val

def copy_fdbase_columns(file_path):
    path = Path(file_path)
    copied_data = []
    
    if path.suffix.lower() == '.csv':
        # CSV: Columns B,C,D,H,J,K,L,M,N,Q -> indices 1,2,3,7,9,10,11,12,13,16 (0-based)
        # Mapping: B=0, C=1, D=2, H=3, J=4, K=5, L=6, M=7, N=8, Q=9 in our extracted data
        csv_indices = [1, 2, 3, 7, 9, 10, 11, 12, 13, 16]
        
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = list(csv.reader(f))
            # Start at row 5 (1-based) -> index 4
            for r in range(4, len(reader)):
                row = reader[r]
                row_data = []
                for col_idx in csv_indices:
                    val = row[col_idx] if col_idx < len(row) else None
                    # Keep raw values - will be processed during paste
                    row_data.append(val)
                
                if any(v not in (None, "") for v in row_data):
                    copied_data.append(row_data)
    else:
        # Excel file - Load with data_only to get values
        wb_values = openpyxl.load_workbook(path, data_only=True)
        ws_values = wb_values.active
        
        # Columns: B,C,D,H,J,K,L,M,N,Q
        columns_to_copy = ["B", "C", "D", "H", "J", "K", "L", "M", "N", "Q"]
        
        max_row = ws_values.max_row
        for row in range(5, max_row + 1):
            row_data = []
            for col in columns_to_copy:
                # Get the value - keep datetime objects and numbers as-is
                val = ws_values[f"{col}{row}"].value
                row_data.append(val)
            
            if any(v not in (None, "") for v in row_data):
                copied_data.append(row_data)
        
        wb_values.close()

    print(f"✅ Copied {len(copied_data)} rows from FDBaseNew_UserBranch.")
    return copied_data

# ------------------------------------------------------------
# STEP 8 & 9 – Paste FDBaseNew_UserBranch → Portfolio
# ------------------------------------------------------------
def convert_date_string_to_excel_serial(date_str):
    """Convert date string to Excel serial number"""
    if date_str is None or date_str == "":
        return None
    
    # If it's already a datetime object
    if isinstance(date_str, datetime):
        excel_epoch = datetime(1899, 12, 30)
        delta = date_str - excel_epoch
        return delta.days
    
    # If it's already a number, return it
    if isinstance(date_str, (int, float)):
        return date_str
    
    # Parse the date string
    date_formats = [
        "%d/%m/%Y",      # 18/09/2025
        "%m/%d/%Y",      # 09/18/2025
        "%Y-%m-%d",      # 2025-09-18
        "%d-%m-%Y",      # 18-09-2025
        "%Y/%m/%d",      # 2025/09/18
        "%d/%m/%y",      # 18/09/25
        "%m/%d/%y",      # 09/18/25
    ]
    
    for fmt in date_formats:
        try:
            dt = datetime.strptime(str(date_str).strip(), fmt)
            # Convert to Excel serial number
            # Excel's date system: 1 = January 1, 1900
            excel_epoch = datetime(1899, 12, 30)
            delta = dt - excel_epoch
            return delta.days
        except ValueError:
            continue
    
    # If parsing fails, return original value
    return date_str

def paste_fdbase_to_portfolio(target_file, copied_data):
    wb = openpyxl.load_workbook(target_file)
    ws = wb["Portfolio"]

    # Clear A5:J[last row]
    max_row = ws.max_row
    for row in range(5, max_row + 1):
        for col in range(1, 11):
            ws.cell(row=row, column=col).value = None

    # Mapping of our extracted columns to Portfolio columns:
    # Extracted: B=0, C=1, D=2, H=3, J=4, K=5, L=6, M=7, N=8, Q=9
    # Portfolio: A=1, B=2, C=3, D=4, E=5, F=6, G=7, H=8, I=9, J=10
    # So: extracted index 2 (D) -> Portfolio column 4 (D)
    #     extracted index 3 (H) -> Portfolio column 5 (E)
    
    # Date columns in extracted data: index 2 (source D) and index 3 (source H)
    # These go to Portfolio columns D (4) and E (5)
    date_column_indices = [2, 3]  # Indices in extracted data
    
    # Paste data starting row 5
    for i, row_data in enumerate(copied_data):
        for j, val in enumerate(row_data):
            cell = ws.cell(row=5+i, column=j+1)
            
            # Convert date strings to Excel serial numbers for date columns
            if j in date_column_indices:
                val = convert_date_string_to_excel_serial(val)
                cell.value = val
                # Apply date format to the cell
                if val is not None and isinstance(val, (int, float)):
                    # Portfolio E (column 5) corresponds to extracted index 3
                    if j == 3:
                        cell.number_format = 'MM/DD/YYYY'
                    else:
                        cell.number_format = 'M/D/YYYY'
            else:
                cell.value = val

    # Apply formulas for K,L,M,N,Q (columns 11,12,13,14,17)
    formula_cols = [11, 12, 13, 14, 17]
    template_row = 4  # assume formulas exist in row 4
    for i in range(len(copied_data)):
        for col in formula_cols:
            formula = ws.cell(row=template_row, column=col).value
            if formula and str(formula).startswith("="):
                # Adjust formula row references
                new_formula = str(formula).replace(str(template_row), str(5+i))
                ws.cell(row=5+i, column=col).value = new_formula

    # Normalize existing Portfolio column E (5) dates to true Date type with MM/DD/YYYY
    last_row = ws.max_row
    for r in range(5, last_row + 1):
        cell_e = ws.cell(row=r, column=5)
        val = cell_e.value
        if isinstance(val, str):
            s = val.strip()
            m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})$", s)
            if m:
                d_str, m_str, y_str = m.groups()
                if len(y_str) == 2:
                    y_str = "20" + y_str
                try:
                    year = int(y_str)
                    month = int(m_str)
                    day = int(d_str)
                    dt = datetime(year, month, day)
                    cell_e.value = dt
                    cell_e.number_format = 'MM/DD/YYYY'
                except ValueError:
                    # Leave value as-is if invalid date
                    pass
        elif isinstance(val, datetime):
            cell_e.number_format = 'MM/DD/YYYY'
        elif isinstance(val, (int, float)):
            cell_e.number_format = 'MM/DD/YYYY'

    wb.save(target_file)
    wb.close()
    print("✅ FDBaseNew_UserBranch data pasted to Portfolio with formulas.")
    print("   - Date columns D and E converted to Excel date format (MM/DD/YYYY)")

# ------------------------------------------------------------
# STEP 10 – Update Portfolio column P from InterestPayableAsAtDateReport
# ------------------------------------------------------------
def update_portfolio_column_p(portfolio_file, interest_file):
    wb_portfolio = openpyxl.load_workbook(portfolio_file)
    ws_portfolio = wb_portfolio["Portfolio"]
    max_row_portfolio = ws_portfolio.max_row

    # Clear column P
    for row in range(5, max_row_portfolio + 1):
        ws_portfolio.cell(row=row, column=16).value = None

    interest_path = Path(interest_file)
    interest_data = []

    if interest_path.suffix.lower() == '.csv':
        # CSV file
        with open(interest_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            next(reader, None)  # skip header
            for row in reader:
                fdno = row[2] if len(row) > 2 else None
                intpayable = row[17] if len(row) > 17 else None
                try:
                    intpayable_val = float(intpayable) if intpayable not in (None, "") else None
                except ValueError:
                    intpayable_val = None
                if fdno not in (None, "") and intpayable_val is not None:
                    interest_data.append((str(fdno), intpayable_val))
    else:
        # Excel file
        wb_interest = openpyxl.load_workbook(interest_path, data_only=True)
        ws_interest = wb_interest.active
        for row in range(2, ws_interest.max_row + 1):
            fdno = ws_interest.cell(row=row, column=3).value
            intpayable = ws_interest.cell(row=row, column=18).value
            if fdno is not None and intpayable is not None:
                interest_data.append((str(fdno), float(intpayable)))
        wb_interest.close()

    # Sort by INTPAYABLE ascending
    interest_data.sort(key=lambda x: x[1])
    lookup_dict = {fdno: intpayable for fdno, intpayable in interest_data}

    # VLOOKUP: fill Portfolio!P
    for row in range(5, max_row_portfolio + 1):
        fdno = ws_portfolio.cell(row=row, column=3).value
        if fdno is not None:
            ws_portfolio.cell(row=row, column=16).value = lookup_dict.get(str(fdno), 0)

    wb_portfolio.save(portfolio_file)
    wb_portfolio.close()
    print("✅ Portfolio column P updated from InterestPayableAsAtDateReport.")

# ------------------------------------------------------------
# STEP 11 – Q3 check for Scienter
# ------------------------------------------------------------
def check_portfolio_q3_for_scienter(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Portfolio"]
    q3_val = ws["Q3"].value
    if q3_val is not None and isinstance(q3_val, (int, float)) and q3_val > 100:
        print(f"⚠️ Portfolio!Q3 = {q3_val} > 100 → Scienter system check required (ToDo).")
    else:
        print("✅ Portfolio!Q3 <= 100 → no Scienter check needed.")
    wb.close()

# ------------------------------------------------------------
# MAIN EXECUTION
# ------------------------------------------------------------
def main():
    # Get relative path to working/NBD-WF-18-DM directory
    script_dir = Path(__file__).parent
    base_dir = script_dir.parent / "working" / "NBD-WF-18-DM"

    # Check if working directory exists
    if not base_dir.exists():
        raise FileNotFoundError(f"Working directory not found: {base_dir}")

    # Find the single date folder inside NBD-WF-18-DM
    date_folders = [f for f in base_dir.iterdir() if f.is_dir()]
    if not date_folders:
        raise FileNotFoundError(f"No date folder found inside '{base_dir}' directory.")

    # Use the first (and should be only) date folder
    if len(date_folders) > 1:
        print(f"Warning: Multiple date folders found: {[f.name for f in date_folders]}. Using the first one: {date_folders[0].name}")

    # Working directory is the date folder
    working_date_folder = date_folders[0]
    print(f"Using working directory: {working_date_folder}")

    # Output will be in the same working folder
    output_dir = working_date_folder
    report_name = "NBD-WF-18-DM"
    old_week = 40
    new_week = 41

    print("=" * 60)
    print("Starting NBD-WF-18-DM Deposit Liability Report Processing")
    print("=" * 60)

    try:
        # Step 1 – Rename template file
        print("\n[STEP 1] Renaming template file...")
        renamed_file = rename_wf18dm_report(working_date_folder, report_name, old_week, new_week)

        # Step 2 & 3 – Summary & Rec update
        print("\n[STEP 2-3] Updating Summary & Rec sheets...")
        update_summary_and_rec(renamed_file)

        # Step 4 – Paste CentralBankMovement data (look in the same date folder)
        print("\n[STEP 4] Pasting CentralBankMovement data...")
        paste_from_centralbank_report(renamed_file, working_date_folder)

        # Step 5 – Validate Detailed sheet
        print("\n[STEP 5] Validating Detailed sheet...")
        validate_detailed_sheet(renamed_file)

        # Step 6 – Copy dates Summary → Rec
        print("\n[STEP 6] Copying dates from Summary to Rec...")
        copy_summary_dates_to_rec(renamed_file)

        # Step 7 – Copy FDBaseNew_UserBranch (look in the same date folder)
        print("\n[STEP 7] Copying FDBaseNew_UserBranch data...")
        fdbase_candidates = list(working_date_folder.glob("FDBaseNew_UserBranch*.xlsx")) + \
                            list(working_date_folder.glob("FDBaseNew_UserBranch*.csv"))
        if not fdbase_candidates:
            raise FileNotFoundError("FDBaseNew_UserBranch file not found in the working folder.")
        copied_data = copy_fdbase_columns(fdbase_candidates[0])

        # Step 8 & 9 – Paste FDBaseNew_UserBranch → Portfolio
        print("\n[STEP 8-9] Pasting data to Portfolio sheet...")
        paste_fdbase_to_portfolio(renamed_file, copied_data)

        # Step 10 – Update Portfolio column P
        print("\n[STEP 10] Updating Portfolio column P...")
        interest_candidates = list(working_date_folder.glob("InterestPayableAsAtDateReport*.xlsx")) + \
                              list(working_date_folder.glob("InterestPayableAsAtDateReport*.csv"))
        if not interest_candidates:
            raise FileNotFoundError("InterestPayableAsAtDateReport file not found in the working folder.")
        update_portfolio_column_p(renamed_file, interest_candidates[0])

        # Step 11 – Q3 check for Scienter
        print("\n[STEP 11] Checking Portfolio Q3 for Scienter...")
        check_portfolio_q3_for_scienter(renamed_file)

        print("\n" + "=" * 60)
        print("✅ PROCESSING COMPLETED SUCCESSFULLY!")
        print(f"📁 Output file: {renamed_file}")
        print("=" * 60)

    except Exception as e:
        print("\n" + "=" * 60)
        print(f"❌ ERROR: {str(e)}")
        print("=" * 60)
        raise

if __name__ == "__main__":
    main()