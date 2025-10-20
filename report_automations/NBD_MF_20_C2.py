import os
from pathlib import Path
import openpyxl
import sys

def get_month_year_from_filename(filename):
    # Example: "NBD-MF-20-C1 to C6 July 2025.xlsx"
    parts = filename.split()
    for i, part in enumerate(parts):
        if part in {"Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","January","February","March","April","May","June","July","August","September","October","November","December"}:
            month = part
            year = parts[i+1].replace(".xlsx","")
            return month, year

def find_files_safely(base_dir, pattern, description):
    """Safely find files and provide detailed error messages"""
    print(f"Looking for {description} in: {base_dir}")
    print(f"Search pattern: {pattern}")
    
    files = list(base_dir.glob(pattern))
    print(f"Found {len(files)} files matching pattern:")
    
    if not files:
        print(f"No files found matching pattern: {pattern}")
        print("Available files in directory:")
        try:
            all_files = [f.name for f in base_dir.iterdir() if f.is_file() and f.suffix == '.xlsx']
            if all_files:
                for file in sorted(all_files):
                    print(f"  - {file}")
            else:
                print("  No .xlsx files found in directory")
        except Exception as e:
            print(f"  Error listing files: {e}")
        return None
    
    for file in files:
        print(f"  {file.name}")
    
    return files[0]  # Return the first match

def find_latest_working_folder(base_dir, report_name):
    """Return the latest date folder inside working/<report_name>."""
    working_report_dir = base_dir / "working" / report_name
    if not working_report_dir.exists():
        print(f"ERROR: Working folder not found: {working_report_dir}")
        return None

    subfolders = [f for f in working_report_dir.iterdir() if f.is_dir()]
    if not subfolders:
        print(f"ERROR: No subfolders found under {working_report_dir}")
        return None

    # Since there is only one date folder in working, just return it
    latest_folder = subfolders[0]
    print(f"Working folder for {report_name}: {latest_folder}")
    return latest_folder


def find_file(base_dir, pattern, description):
    """Safely find a single file using glob and provide detailed errors."""
    print(f"\nLooking for {description} in: {base_dir}")
    files = list(base_dir.glob(pattern))
    if not files:
        print(f"No files found for {description} with pattern: {pattern}")
        all_files = [f.name for f in base_dir.iterdir() if f.is_file()]
        if all_files:
            print("Available files in this folder:")
            for f in sorted(all_files):
                print(f"  - {f}")
        else:
            print("No files found in this folder.")
        return None
    print(f"Found {description}: {files[0].name}")
    return files[0]


def prepare_output_folder(base_dir, report_name, working_folder):
    """Create output folder matching working folder name under outputs/<report_name>/"""
    output_report_dir = base_dir / "outputs" / report_name
    output_folder = output_report_dir / working_folder.name
    output_folder.mkdir(parents=True, exist_ok=True)
    print(f"Output folder created: {output_folder}")
    return output_folder

def main():

    base_dir = Path(__file__).resolve().parent.parent
    report_name = "NBD_MF_20_C1_C6"  # Example report

    # Get the latest working folder (single date folder)
    working_folder = find_latest_working_folder(base_dir, report_name)
    if not working_folder:
        return

    # Find files inside the working folder
    file_c1_c6 = find_file(working_folder, "**/NBD-MF-20-C1 to C6*.xlsx", "C1-C6 file")
    file_afl = find_file(working_folder, "**/NBD-MF-01-SOFP & SOCI AFL Monthly FS*.xlsx", "AFL file")

    if not file_c1_c6 or not file_afl:
        print("Required files not found. Cannot continue.")
        return

    # Prepare output folder
    output_folder = prepare_output_folder(base_dir, report_name, working_folder)

    print("\n--- Summary ---")
    print(f"C1-C6 File: {file_c1_c6}")
    print(f"AFL File: {file_afl}")
    print(f"Output folder: {output_folder}")
    
    try:
        # Load workbooks
        print(f"\nLoading workbooks...")
        wb_c1_c6 = openpyxl.load_workbook(file_c1_c6, data_only=True)
        wb_afl = openpyxl.load_workbook(file_afl, data_only=True)
        print(f"Workbooks loaded successfully")
        
        # Check if required sheets exist
        print(f"\nAvailable sheets in C1-C6 file: {wb_c1_c6.sheetnames}")
        print(f"Available sheets in AFL file: {wb_afl.sheetnames}")
        
        # Check for C2 sheet in C1-C6 file
        if "NBD_NBL-MF-20-C2" not in wb_c1_c6.sheetnames:
            print(f"Required sheet 'NBD_NBL-MF-20-C2' not found in C1-C6 file")
            print(f"Available sheets: {wb_c1_c6.sheetnames}")
            return
        
        # Check for SOCI sheet in AFL file
        if "NBD-MF-02-SOCI" not in wb_afl.sheetnames:
            print(f"Required sheet 'NBD-MF-02-SOCI' not found in AFL file")
            print(f"Available sheets: {wb_afl.sheetnames}")
            return
        
        print(f"All required sheets found")
        
        # Get worksheets
        src_ws = wb_afl["NBD-MF-02-SOCI"]
        tgt_ws = wb_c1_c6["NBD_NBL-MF-20-C2"]
        
        print(f"\nProcessing data...")
        
        # Helper function to safely get numeric values
        def get_numeric_value(worksheet, cell_ref, description=""):
            """Safely extract numeric value from a cell, handling text/None values"""
            try:
                value = worksheet[cell_ref].value
                print(f"{description} ({cell_ref}): {value}")
                
                if value is None:
                    return 0
                elif isinstance(value, (int, float)):
                    return value
                elif isinstance(value, str):
                    # Try to extract number from string if possible
                    try:
                        # Remove common text and try to parse number
                        clean_value = value.replace(',', '').strip()
                        return float(clean_value)
                    except ValueError:
                        print(f"  Warning: Cell {cell_ref} contains text '{value}' - using 0")
                        return 0
                else:
                    print(f"  Warning: Cell {cell_ref} contains unexpected type {type(value)} - using 0")
                    return 0
            except Exception as e:
                print(f"  Error reading {cell_ref}: {e} - using 0")
                return 0
        
        # Find "Total Comprehensive Income for the Year" row in AFL file
        print(f"\nSearching for 'Total Comprehensive Income for the Year' in SOCI sheet...")
        
        comprehensive_income_value = None
        comprehensive_income_row = None
        
        # First, find the "Total Comprehensive Income for the Year" row
        for row in range(1, src_ws.max_row + 1):
            for col in range(1, src_ws.max_column + 1):
                cell_value = src_ws.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str) and "Total Comprehensive Income for the Year" in cell_value:
                    print(f"Found 'Total Comprehensive Income for the Year' at row {row}, column {col}")
                    comprehensive_income_row = row
                    break
            if comprehensive_income_row:
                break
        
        if comprehensive_income_row is None:
            print("Could not find 'Total Comprehensive Income for the Year' row in AFL file")
            print("Available cell values in the SOCI sheet:")
            for row in range(1, min(20, src_ws.max_row + 1)):  # Show first 20 rows
                for col in range(1, min(10, src_ws.max_column + 1)):  # Show first 10 columns
                    cell_value = src_ws.cell(row=row, column=col).value
                    if cell_value:
                        print(f"  Row {row}, Col {col}: {cell_value}")
            return
        
        # Find the "Rs.'000" column and get the value
        print(f"Looking for 'Rs.'000' column value...")
        
        # Look for Rs.'000 column header
        rs_column = None
        for col in range(1, src_ws.max_column + 1):
            cell_value = src_ws.cell(row=1, column=col).value
            if cell_value and isinstance(cell_value, str) and "Rs.'000" in cell_value:
                print(f"Found 'Rs.'000' column at column {col}")
                rs_column = col
                break
        
        if rs_column is None:
            print("Could not find 'Rs.'000' column in AFL file")
            return
        
        # Get the comprehensive income value
        comprehensive_income_value = get_numeric_value(src_ws, f"{openpyxl.utils.get_column_letter(rs_column)}{comprehensive_income_row}", "Total Comprehensive Income for the Year")
        
        # Only update if the value is negative (loss)
        current_year_profit_loss = comprehensive_income_value
        current_year_updated = False
        
        if comprehensive_income_value < 0:
            print(f"Comprehensive income is negative ({comprehensive_income_value}), will update Current year's profit(losses)")
            
            # Find "Current year's profit(losses)" row in C2 sheet
            print(f"\nSearching for 'Current year's profit(losses)' in C2 sheet...")
            
            current_year_row = None
            for row in range(1, tgt_ws.max_row + 1):
                for col in range(1, tgt_ws.max_column + 1):
                    cell_value = tgt_ws.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str) and "Current year's profit(losses)" in cell_value:
                        print(f"Found 'Current year's profit(losses)' at row {row}, column {col}")
                        current_year_row = row
                        break
                if current_year_row:
                    break
            
            if current_year_row is None:
                print("Could not find 'Current year's profit(losses)' row in C2 sheet")
                print("Available cell values in C2 sheet:")
                for row in range(1, min(20, tgt_ws.max_row + 1)):  # Show first 20 rows
                    for col in range(1, min(10, tgt_ws.max_column + 1)):  # Show first 10 columns
                        cell_value = tgt_ws.cell(row=row, column=col).value
                        if cell_value:
                            print(f"  Row {row}, Col {col}: {cell_value}")
                return
            
            # Update the current year's profit(losses) value
            # Assuming the value goes in column C (3rd column)
            target_cell = f"C{current_year_row}"
            print(f"Updating cell {target_cell} with value: {current_year_profit_loss}")
            tgt_ws[target_cell] = current_year_profit_loss
            current_year_updated = True
        else:
            print(f"Comprehensive income is positive ({comprehensive_income_value}), NOT updating Current year's profit(losses)")
        
        # Now calculate Adjustments to Tier I capital values
        print(f"\nCalculating Adjustments to Tier I capital values...")
        
        # Find the required values in C2 sheet
        stated_capital = 0
        statutory_reserve_fund = 0
        retained_earnings = 0
        
        # Search for these values in the C2 sheet
        for row in range(1, tgt_ws.max_row + 1):
            for col in range(1, tgt_ws.max_column + 1):
                cell_value = tgt_ws.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    if "Stated Capital" in cell_value:
                        # Get value from column C of the same row
                        stated_capital = get_numeric_value(tgt_ws, f"C{row}", "Stated Capital")
                    elif "Statutory Reserve Fund" in cell_value:
                        # Get value from column C of the same row
                        statutory_reserve_fund = get_numeric_value(tgt_ws, f"C{row}", "Statutory Reserve Fund")
                    elif "Retained Earnings" in cell_value:
                        # Get value from column C of the same row
                        retained_earnings = get_numeric_value(tgt_ws, f"C{row}", "Retained Earnings")
        
        # Calculate Adjustments to Tier I capital
        adjustments_to_tier1 = stated_capital + statutory_reserve_fund + retained_earnings
        print(f"Adjustments to Tier I capital calculation:")
        print(f"  Stated Capital: {stated_capital}")
        print(f"  Statutory Reserve Fund: {statutory_reserve_fund}")
        print(f"  Retained Earnings: {retained_earnings}")
        print(f"  Total Adjustments to Tier I capital: {adjustments_to_tier1}")
        
        # Find "Adjustments to Tier I capital" row and update it
        print(f"\nSearching for 'Adjustments to Tier I capital' in C2 sheet...")
        
        adjustments_row = None
        for row in range(1, tgt_ws.max_row + 1):
            for col in range(1, tgt_ws.max_column + 1):
                cell_value = tgt_ws.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str) and "Adjustments to Tier I capital" in cell_value:
                    print(f"Found 'Adjustments to Tier I capital' at row {row}, column {col}")
                    adjustments_row = row
                    break
            if adjustments_row:
                break
        
        if adjustments_row is None:
            print("Could not find 'Adjustments to Tier I capital' row in C2 sheet")
            return
        
        # Update the adjustments value
        adjustments_cell = f"C{adjustments_row}"
        print(f"Updating cell {adjustments_cell} with value: {adjustments_to_tier1}")
        tgt_ws[adjustments_cell] = adjustments_to_tier1
        
        # Hardcoded value for "Instruments qualified as Tier 2 capital" (dummy data from work hub integration)
        print(f"\nSetting hardcoded value for 'Instruments qualified as Tier 2 capital'...")
        tier2_hardcoded_value = 5000000  # Hardcoded dummy value
        print(f"DUMMY DATA: Using hardcoded value {tier2_hardcoded_value} for 'Instruments qualified as Tier 2 capital'")
        print(f"NOTE: This value comes from work hub integration as master data (dummy purpose)")
        
        # Find "Instruments qualified as Tier 2 capital" row in C2 sheet
        print(f"\nSearching for 'Instruments qualified as Tier 2 capital' in C2 sheet...")
        
        tier2_row = None
        for row in range(1, tgt_ws.max_row + 1):
            for col in range(1, tgt_ws.max_column + 1):
                cell_value = tgt_ws.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str) and "Instruments qualified as Tier 2 capital" in cell_value:
                    print(f"Found 'Instruments qualified as Tier 2 capital' at row {row}, column {col}")
                    tier2_row = row
                    break
            if tier2_row:
                break
        
        if tier2_row is None:
            print("Could not find 'Instruments qualified as Tier 2 capital' row in C2 sheet")
            print("Available cell values in C2 sheet:")
            for row in range(1, min(20, tgt_ws.max_row + 1)):  # Show first 20 rows
                for col in range(1, min(10, tgt_ws.max_column + 1)):  # Show first 10 columns
                    cell_value = tgt_ws.cell(row=row, column=col).value
                    if cell_value:
                        print(f"  Row {row}, Col {col}: {cell_value}")
        else:
            # Update the Tier 2 capital value
            tier2_cell = f"C{tier2_row}"
            print(f"Updating cell {tier2_cell} with hardcoded value: {tier2_hardcoded_value}")
            tgt_ws[tier2_cell] = tier2_hardcoded_value
        
        print(f"Data processing completed")
        print(f"Updated C2 sheet:")
        if current_year_updated:
            print(f"  - Current year's profit(losses): {current_year_profit_loss} (UPDATED)")
        else:
            print(f"  - Current year's profit(losses): NOT UPDATED (comprehensive income was positive)")
        print(f"  - Adjustments to Tier I capital: {adjustments_to_tier1}")
        if tier2_row:
            print(f"  - Instruments qualified as Tier 2 capital: {tier2_hardcoded_value} (HARDCODED DUMMY DATA)")
        else:
            print(f"  - Instruments qualified as Tier 2 capital: NOT FOUND IN SHEET")
        
        # Save updated workbook back to the same original file (in-place like Ctrl+S)
        wb_c1_c6.save(file_c1_c6)
        print(f"\nReport saved successfully to original location: {file_c1_c6}")
        
    except FileNotFoundError as e:
        print(f"File not found: {e}")
    except PermissionError as e:
        print(f"Permission error (file might be open in Excel): {e}")
    except Exception as e:
        print(f"An error occurred: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description='NBD-MF-20-C2 Report Automation')
    parser.add_argument('--working-dir', type=str, help='Working directory (not used, for compatibility)')
    parser.add_argument('--month', type=str, help='Report month (e.g., Jul)')
    parser.add_argument('--year', type=str, help='Report year (e.g., 2025)')
    args = parser.parse_args()

    main()
