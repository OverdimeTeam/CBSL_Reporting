import os
from pathlib import Path
import sys
import openpyxl
from openpyxl import load_workbook
import pandas as pd
import win32com.client as win32
import numpy as np
import shutil
file_car = r"C:\Users\Overdime Tech\Desktop\C1-C6\C3\CBSL_Reporting\outputs\monthly\09-19-2025(1)\CAR Working July 2025.xlsb"
def get_month_year_from_filename(filename):
    parts = filename.split()
    for i, part in enumerate(parts):
        if part in {
            "Jan", "Feb", "Mar", "Apr", "May", "Jun",
            "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
            "January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December",
        }:
            month = part
            year = parts[i + 1].replace(".xlsx", "").replace(".xlsb", "")
            return month, year
    return None, None

def find_first_matching(search_dirs, pattern):
    for search_dir in search_dirs:
        if search_dir.exists():
            files = list(search_dir.glob(pattern))
            if files:
                return files[0]
    return None

def normalize_columns(df):
    """Apply safe normalization strategy to column headers"""
    df = df.copy()
    if isinstance(df.columns, pd.MultiIndex):
        new_cols = []
        for tup in df.columns:
            # Join non-empty parts
            parts = [str(p).strip() for p in tup if str(p).strip() and str(p).lower() != "nan"]
            new_cols.append(" ".join(parts) if parts else "")
        df.columns = new_cols
    else:
        df.columns = [str(c).strip() for c in df.columns]
    
    # [OK] Safe Normalization Strategy
    # 1. Trim spaces (already done above)
    # 2. Unify case to lowercase for consistent matching
    # 3. Replace special characters with underscores for safe identifiers
    normalized_cols = []
    for col in df.columns:
        # Convert to lowercase
        normalized = col.lower()
        # Replace spaces and special characters with underscores
        import re
        normalized = re.sub(r'[^\w]', '_', normalized)
        # Remove multiple consecutive underscores
        normalized = re.sub(r'_+', '_', normalized)
        # Remove leading/trailing underscores
        normalized = normalized.strip('_')
        # Ensure we don't have empty column names
        if not normalized:
            normalized = f"col_{len(normalized_cols)}"
        normalized_cols.append(normalized)
    
    df.columns = normalized_cols
    return df

def pick_column(df, candidates):
    """Pick column from DataFrame using normalized candidate names"""
    # Since columns are now normalized, we need to normalize candidates too
    import re
    
    def normalize_candidate(candidate):
        """Normalize candidate name to match column normalization"""
        normalized = str(candidate).lower()
        # Replace spaces and special characters with underscores
        normalized = re.sub(r'[^\w]', '_', normalized)
        # Remove multiple consecutive underscores
        normalized = re.sub(r'_+', '_', normalized)
        # Remove leading/trailing underscores
        normalized = normalized.strip('_')
        return normalized
    
    # Create mapping of normalized column names to original column names
    cols = {normalize_candidate(c): c for c in df.columns}
    
    # Try exact matches first with normalized candidates
    for cand in candidates:
        normalized_cand = normalize_candidate(cand)
        if normalized_cand in cols:
            return cols[normalized_cand]
    
    # Fallback: try partial matches
    for cand in candidates:
        normalized_cand = normalize_candidate(cand)
        for normalized_col, original_col in cols.items():
            if normalized_cand in normalized_col or normalized_col in normalized_cand:
                return original_col
    
    return None

def safe_number(x):
    """Convert value to float, handling various formats and edge cases"""
    if pd.isna(x) or x is None:
        return 0.0
    if isinstance(x, (int, float)):
        if np.isnan(x) or np.isinf(x):
            return 0.0
        return float(x)
    try:
        # Handle string representations
        str_val = str(x).replace(",", "").strip()
        if str_val == "" or str_val.lower() in ["nan", "none", "null", "-"]:
            return 0.0
        return float(str_val)
    except (ValueError, TypeError):
        print(f"Warning: Could not convert '{x}' to number, using 0.0")
        return 0.0

def validate_data_quality(df, col_name):
    """Validate and report data quality issues"""
    print(f"\n=== Data Quality Report for {col_name} ===")
    print(f"Total rows: {len(df)}")
    print(f"Non-null values: {df[col_name].notna().sum()}")
    print(f"Null values: {df[col_name].isna().sum()}")
    print(f"Zero values: {(df[col_name] == 0).sum()}")
    print(f"Negative values: {(df[col_name] < 0).sum()}")
    print(f"Data type: {df[col_name].dtype}")
    print(f"Sample values: {df[col_name].head().tolist()}")
    
    if df[col_name].dtype in ['object', 'string']:
        print("Warning: Column contains non-numeric data!")
        unique_types = df[col_name].apply(type).value_counts()
        print(f"Value types found: {unique_types}")

def normalize_contract_number(contract_no):
    """Normalize contract numbers for better matching"""
    if pd.isna(contract_no):
        return ""
    
    # Convert to string and clean
    contract_str = str(contract_no).strip()
    
    # Remove common prefixes/suffixes that might cause mismatches
    contract_str = contract_str.replace("Contract-", "").replace("CONTRACT-", "")
    contract_str = contract_str.replace("Cont-", "").replace("CONT-", "")
    contract_str = contract_str.replace("C-", "").replace("c-", "")
    
    # Remove leading zeros (but keep if it's all zeros)
    if contract_str and not contract_str.replace("0", ""):
        return contract_str  # Keep if all zeros
    else:
        contract_str = contract_str.lstrip("0") or "0"  # Remove leading zeros
    
    # Remove any non-alphanumeric characters except hyphens and underscores
    import re
    contract_str = re.sub(r'[^\w\-]', '', contract_str)
    
    return contract_str.upper()

def detailed_contract_comparison(df_port_contracts, df_cbsl_contracts):
    """Perform detailed analysis of contract number mismatches"""
    print("\n=== DETAILED CONTRACT COMPARISON ANALYSIS ===")
    
    # Normalize both sets
    port_normalized = {normalize_contract_number(c): c for c in df_port_contracts}
    cbsl_normalized = {normalize_contract_number(c): c for c in df_cbsl_contracts}
    
    print(f"Portfolio contracts (original): {len(df_port_contracts)}")
    print(f"Portfolio contracts (normalized unique): {len(port_normalized)}")
    print(f"CBSL contracts (original): {len(df_cbsl_contracts)}")
    print(f"CBSL contracts (normalized unique): {len(cbsl_normalized)}")
    
    # Find matches after normalization
    port_norm_set = set(port_normalized.keys())
    cbsl_norm_set = set(cbsl_normalized.keys())

    exact_matches = port_norm_set.intersection(cbsl_norm_set)
    port_only = port_norm_set - cbsl_norm_set
    cbsl_only = cbsl_norm_set - port_norm_set
    
    print(f"\nAfter normalization:")
    print(f"Exact matches: {len(exact_matches)}")
    print(f"Portfolio-only: {len(port_only)}")
    print(f"CBSL-only: {len(cbsl_only)}")
    
    # Show sample mismatches
    if port_only:
        print(f"\nSample Portfolio-only contracts (normalized -> original):")
        for i, norm_contract in enumerate(list(port_only)[:10]):
            orig_contract = port_normalized[norm_contract]
            print(f"  {i+1}. '{norm_contract}' <- '{orig_contract}'")
    
    if cbsl_only:
        print(f"\nSample CBSL-only contracts (normalized -> original):")
        for i, norm_contract in enumerate(list(cbsl_only)[:10]):
            orig_contract = cbsl_normalized[norm_contract]
            print(f"  {i+1}. '{norm_contract}' <- '{orig_contract}'")
    
    # Fuzzy matching for close matches
    print(f"\n=== FUZZY MATCHING ANALYSIS ===")
    from difflib import SequenceMatcher
    
    def similar(a, b, threshold=0.8):
        return SequenceMatcher(None, a, b).ratio() >= threshold
    
    potential_matches = []
    for port_contract in list(port_only)[:50]:  # Limit to first 50 for performance
        for cbsl_contract in cbsl_only:
            similarity = SequenceMatcher(None, port_contract, cbsl_contract).ratio()
            if similarity >= 0.8:
                potential_matches.append((
                    port_contract, 
                    cbsl_contract, 
                    similarity,
                    port_normalized[port_contract],
                    cbsl_normalized[cbsl_contract]
                ))
    
    if potential_matches:
        print(f"Found {len(potential_matches)} potential fuzzy matches (>80% similarity):")
        for port_norm, cbsl_norm, sim, port_orig, cbsl_orig in potential_matches[:10]:
            print(f"  Portfolio: '{port_orig}' ({port_norm}) <-> CBSL: '{cbsl_orig}' ({cbsl_norm}) [Similarity: {sim:.2f}]")
    else:
        print("No high-similarity fuzzy matches found.")
    
    return {
        'exact_matches': exact_matches,
        'port_only': port_only,
        'cbsl_only': cbsl_only,
        'potential_fuzzy_matches': potential_matches,
        'port_normalized_map': port_normalized,
        'cbsl_normalized_map': cbsl_normalized
    }

def enhanced_pnp_lookup(df_port, df_cbsl, cbsl_contract_col, cbsl_pnp_col):
    """Enhanced P/NP lookup with detailed comparison and normalization"""
    print(f"\n=== ENHANCED P/NP LOOKUP PROCESS ===")
    
    # Create CBSL lookup dataframe
    df_cbsl_lookup = df_cbsl[[cbsl_contract_col, cbsl_pnp_col]].copy()
    df_cbsl_lookup.columns = ["contract_no", "p_np_based_on_cbsl_provision"]
    # Ensure Contract No is treated as text
    df_cbsl_lookup["contract_no"] = df_cbsl_lookup["contract_no"].astype(str).str.strip()
    
    # Clean CBSL data
    print("Cleaning CBSL data...")
    initial_cbsl_count = len(df_cbsl_lookup)
    df_cbsl_lookup = df_cbsl_lookup.dropna(subset=["contract_no"])
    df_cbsl_lookup = df_cbsl_lookup[df_cbsl_lookup["contract_no"].astype(str).str.strip() != ""]
    df_cbsl_lookup = df_cbsl_lookup[~df_cbsl_lookup["contract_no"].astype(str).str.lower().isin(["nan", "none", "null"])]
    
    print(f"CBSL records after cleaning: {len(df_cbsl_lookup)} (removed {initial_cbsl_count - len(df_cbsl_lookup)} invalid)")
    
    # Clean Portfolio data
    print("Cleaning Portfolio data...")
    df_port_clean = df_port.copy()
    df_port_clean = df_port_clean.dropna(subset=["contract_no"])
    df_port_clean = df_port_clean[df_port_clean["contract_no"].astype(str).str.strip() != ""]
    # Ensure Contract No is treated as text
    df_port_clean["contract_no"] = df_port_clean["contract_no"].astype(str).str.strip()
    
    print(f"Portfolio records after cleaning: {len(df_port_clean)}")
    
    # Show data samples before matching
    print(f"\n=== DATA SAMPLES BEFORE MATCHING ===")
    print("Sample Portfolio contract numbers:")
    port_sample = df_port_clean["contract_no"].head(10).astype(str).tolist()
    for i, contract in enumerate(port_sample, 1):
        print(f"  {i:2d}. '{contract}'")
    
    print("\nSample CBSL contract numbers:")
    cbsl_sample = df_cbsl_lookup["contract_no"].head(10).astype(str).tolist()
    for i, contract in enumerate(cbsl_sample, 1):
        print(f"  {i:2d}. '{contract}'")
    
    # Perform detailed comparison
    comparison_results = detailed_contract_comparison(
        df_port_clean["contract_no"].unique(),
        df_cbsl_lookup["contract_no"].unique()
    )
    
    # Standard lookup first
    print(f"\n=== PERFORMING STANDARD LOOKUP ===")
    print(f"Portfolio columns before merge: {list(df_port_clean.columns)}")
    print(f"CBSL lookup columns: {list(df_cbsl_lookup.columns)}")
    df_port_result = df_port_clean.merge(df_cbsl_lookup, on="contract_no", how="left", suffixes=("_orig", "_cbsl"))
    print(f"Portfolio columns after merge: {list(df_port_result.columns)}")
    
    # Use the CBSL version of the column (which has the actual values)
    pnp_col = "p_np_based_on_cbsl_provision_cbsl"
    if pnp_col in df_port_result.columns:
        standard_matches = df_port_result[pnp_col].notna().sum()
    else:
        # Fallback to original if CBSL version doesn't exist
        pnp_col = "p_np_based_on_cbsl_provision_orig"
        standard_matches = df_port_result[pnp_col].notna().sum()
    print(f"Standard lookup matches: {standard_matches} out of {len(df_port_result)} ({standard_matches/len(df_port_result)*100:.1f}%)")
    
    # Enhanced lookup with normalization
    print(f"\n=== PERFORMING NORMALIZED LOOKUP ===")
    
    # Add normalized contract columns
    df_port_norm = df_port_clean.copy()
    df_cbsl_norm = df_cbsl_lookup.copy()
    
    df_port_norm["contract_normalized"] = df_port_norm["contract_no"].apply(normalize_contract_number)
    df_cbsl_norm["contract_normalized"] = df_cbsl_norm["contract_no"].apply(normalize_contract_number)
    
    # Remove any empty normalized contracts
    df_port_norm = df_port_norm[df_port_norm["contract_normalized"] != ""]
    df_cbsl_norm = df_cbsl_norm[df_cbsl_norm["contract_normalized"] != ""]
    
    # Perform normalized lookup
    df_cbsl_norm_lookup = df_cbsl_norm[["contract_normalized", "p_np_based_on_cbsl_provision"]].drop_duplicates(subset=["contract_normalized"], keep="first")

    # Merge exact matches (by original Contract No)
    df_port_enhanced = df_port_norm.merge(df_cbsl_lookup, on="contract_no", how="left", suffixes=("_orig", "_cbsl"))
    # Merge normalized matches (by normalized contract)
    df_port_enhanced = df_port_enhanced.merge(
        df_cbsl_norm_lookup.rename(columns={"p_np_based_on_cbsl_provision": "p_np_based_on_cbsl_provision_normalized"}),
        on="contract_normalized",
        how="left"
    )

    # Combine standard and normalized results
    # Use the CBSL version from the first merge
    cbsl_pnp_col = "p_np_based_on_cbsl_provision_cbsl"
    if cbsl_pnp_col in df_port_enhanced.columns:
        df_port_enhanced["p_np_based_on_cbsl_provision_final"] = df_port_enhanced[cbsl_pnp_col].fillna(df_port_enhanced["p_np_based_on_cbsl_provision_normalized"])
    else:
        # Fallback to original
        orig_pnp_col = "p_np_based_on_cbsl_provision_orig"
        df_port_enhanced["p_np_based_on_cbsl_provision_final"] = df_port_enhanced[orig_pnp_col].fillna(df_port_enhanced["p_np_based_on_cbsl_provision_normalized"])
    
    enhanced_matches = df_port_enhanced["p_np_based_on_cbsl_provision_final"].notna().sum()
    print(f"Enhanced lookup matches: {enhanced_matches} out of {len(df_port_enhanced)} ({enhanced_matches/len(df_port_enhanced)*100:.1f}%)")
    print(f"Additional matches from normalization: {enhanced_matches - standard_matches}")
    
    # Final result preparation
    df_result = df_port.copy()  # Start with original portfolio
    
    # Create lookup dictionary from enhanced results
    # Use the CBSL version if available, otherwise use the final combined version
    final_pnp_col = "p_np_based_on_cbsl_provision_final"
    if final_pnp_col in df_port_enhanced.columns:
        lookup_dict = df_port_enhanced.set_index("contract_no")[final_pnp_col].to_dict()
    else:
        # Fallback to CBSL version
        lookup_dict = df_port_enhanced.set_index("contract_no")[pnp_col].to_dict()
    df_result["p_np_based_on_cbsl_provision"] = df_result["contract_no"].map(lookup_dict)
    
    # Show P/NP distribution
    print(f"\n=== P/NP VALUE DISTRIBUTION ===")
    pnp_counts = df_result["p_np_based_on_cbsl_provision"].value_counts(dropna=False)
    for value, count in pnp_counts.items():
        percentage = (count / len(df_result)) * 100
        print(f"  {str(value):20s}: {count:5d} records ({percentage:5.1f}%)")
    
    # Show sample results
    print(f"\n=== SAMPLE MATCHING RESULTS ===")
    matched_sample = df_result[df_result["p_np_based_on_cbsl_provision"].notna()].head(5)
    if len(matched_sample) > 0:
        print("Sample MATCHED records:")
        for idx, row in matched_sample.iterrows():
            print(f"  Contract: {row['contract_no']:15s} -> P/NP: {row['p_np_based_on_cbsl_provision']}")
    
    unmatched_sample = df_result[df_result["p_np_based_on_cbsl_provision"].isna()].head(5)
    if len(unmatched_sample) > 0:
        print("\nSample UNMATCHED records:")
        for idx, row in unmatched_sample.iterrows():
            print(f"  Contract: {row['contract_no']:15s} -> No P/NP found")
    
    # Export detailed matching report (optimized/vectorized)
    print(f"\n=== GENERATING DETAILED MATCHING REPORT ===")
    # Precompute normalized contract for result set
    df_result = df_result.copy()
    df_result["normalized_contract"] = df_result["contract_no"].apply(normalize_contract_number)
    
    # Build sets for O(1) membership checks
    cbsl_contract_set = set(df_cbsl_lookup["contract_no"].astype(str).str.strip())
    cbsl_norm_set = set(df_cbsl_norm["contract_normalized"]) if "contract_normalized" in df_cbsl_norm.columns else set()
    
    # Boolean masks
    exact_mask = df_result["contract_no"].isin(cbsl_contract_set)
    norm_mask = df_result["normalized_contract"].isin(cbsl_norm_set)
    pnp_notna = df_result["p_np_based_on_cbsl_provision"].notna()
    
    # Default match type
    df_result["match_type"] = "No Match"
    
    # Assign match types in priority order
    df_result.loc[pnp_notna & exact_mask, "match_type"] = "Exact Match"
    df_result.loc[pnp_notna & ~exact_mask & norm_mask, "match_type"] = "Normalized Match"
    df_result.loc[pnp_notna & ~exact_mask & ~norm_mask, "match_type"] = "Unknown Match"
    
    # Construct report dataframe directly
    df_matching_report = df_result[[
        "contract_no",
        "normalized_contract",
        "p_np_based_on_cbsl_provision",
        "match_type",
        "client_code",
        "product",
        "gross_with_dp",
    ]].rename(columns={"p_np_based_on_cbsl_provision": "p_np_value"})
    
    # Summary statistics
    print(f"Matching Summary:")
    match_type_counts = df_matching_report["match_type"].value_counts()
    for match_type, count in match_type_counts.items():
        percentage = (count / len(df_matching_report)) * 100
        print(f"  {match_type:18s}: {count:5d} records ({percentage:5.1f}%)")
    
    return df_result, df_matching_report, comparison_results

def enhanced_mortgage_lookup(df_port, df_cbsl_pm, pm_contract_col, pm_value_col):
    """Enhanced mortgage lookup with detailed comparison and normalization"""
    print(f"\n=== ENHANCED MORTGAGE LOOKUP PROCESS ===")
    
    # Create CBSL PropertyMortgage lookup dataframe
    df_pm_lookup = df_cbsl_pm[[pm_contract_col, pm_value_col]].copy()
    df_pm_lookup.columns = ["contract_no", "mortgage_value"]
    # Ensure Contract No is treated as text
    df_pm_lookup["contract_no"] = df_pm_lookup["contract_no"].astype(str).str.strip()
    
    # Clean CBSL PropertyMortgage data
    print("Cleaning CBSL PropertyMortgage data...")
    initial_pm_count = len(df_pm_lookup)
    df_pm_lookup = df_pm_lookup.dropna(subset=["contract_no"])
    df_pm_lookup = df_pm_lookup[df_pm_lookup["contract_no"].astype(str).str.strip() != ""]
    df_pm_lookup = df_pm_lookup[~df_pm_lookup["contract_no"].astype(str).str.lower().isin(["nan", "none", "null"])]
    
    print(f"CBSL PropertyMortgage records after cleaning: {len(df_pm_lookup)} (removed {initial_pm_count - len(df_pm_lookup)} invalid)")
    
    # Clean Portfolio data
    print("Cleaning Portfolio data...")
    df_port_clean = df_port.copy()
    df_port_clean = df_port_clean.dropna(subset=["contract_no"])
    df_port_clean = df_port_clean[df_port_clean["contract_no"].astype(str).str.strip() != ""]
    # Ensure Contract No is treated as text
    df_port_clean["contract_no"] = df_port_clean["contract_no"].astype(str).str.strip()
    
    print(f"Portfolio records after cleaning: {len(df_port_clean)}")
    
    # Show data samples before matching
    print(f"\n=== MORTGAGE DATA SAMPLES BEFORE MATCHING ===")
    print("Sample Portfolio contract numbers:")
    port_sample = df_port_clean["contract_no"].head(10).astype(str).tolist()
    for i, contract in enumerate(port_sample, 1):
        print(f"  {i:2d}. '{contract}'")
    
    print("\nSample CBSL PropertyMortgage contract numbers:")
    pm_sample = df_pm_lookup["contract_no"].head(10).astype(str).tolist()
    for i, contract in enumerate(pm_sample, 1):
        print(f"  {i:2d}. '{contract}'")
    
    print("\nSample CBSL PropertyMortgage values:")
    pm_values_sample = df_pm_lookup["mortgage_value"].head(10).tolist()
    for i, value in enumerate(pm_values_sample, 1):
        print(f"  {i:2d}. {value}")
    
    # Perform detailed comparison for mortgage data
    comparison_results = detailed_contract_comparison(
        df_port_clean["contract_no"].unique(),
        df_pm_lookup["contract_no"].unique()
    )
    
    # Standard lookup first
    print(f"\n=== PERFORMING STANDARD MORTGAGE LOOKUP ===")
    df_port_result = df_port_clean.merge(df_pm_lookup, on="contract_no", how="left")
    
    standard_matches = df_port_result["mortgage_value"].notna().sum()
    print(f"Standard mortgage lookup matches: {standard_matches} out of {len(df_port_result)} ({standard_matches/len(df_port_result)*100:.1f}%)")
    
    # Enhanced lookup with normalization
    print(f"\n=== PERFORMING NORMALIZED MORTGAGE LOOKUP ===")
    
    # Add normalized contract columns
    df_port_norm = df_port_clean.copy()
    df_pm_norm = df_pm_lookup.copy()
    
    df_port_norm["contract_normalized"] = df_port_norm["contract_no"].apply(normalize_contract_number)
    df_pm_norm["contract_normalized"] = df_pm_norm["contract_no"].apply(normalize_contract_number)
    
    # Remove any empty normalized contracts
    df_port_norm = df_port_norm[df_port_norm["contract_normalized"] != ""]
    df_pm_norm = df_pm_norm[df_pm_norm["contract_normalized"] != ""]
    
    # Perform normalized lookup
    df_pm_norm_lookup = df_pm_norm[["contract_normalized", "mortgage_value"]].drop_duplicates(subset=["contract_normalized"], keep="first")

    # Merge exact matches (by original Contract No)
    df_port_enhanced = df_port_norm.merge(df_pm_lookup, on="contract_no", how="left")
    # Merge normalized matches (by normalized contract)
    df_port_enhanced = df_port_enhanced.merge(
        df_pm_norm_lookup.rename(columns={"mortgage_value": "mortgage_value_normalized"}),
        on="contract_normalized",
        how="left"
    )

    # Combine standard and normalized results
    df_port_enhanced["mortgage_value_final"] = df_port_enhanced["mortgage_value"].fillna(df_port_enhanced["mortgage_value_normalized"])
    
    enhanced_matches = df_port_enhanced["mortgage_value_final"].notna().sum()
    print(f"Enhanced mortgage lookup matches: {enhanced_matches} out of {len(df_port_enhanced)} ({enhanced_matches/len(df_port_enhanced)*100:.1f}%)")
    print(f"Additional matches from normalization: {enhanced_matches - standard_matches}")
    
    # Final result preparation
    df_result = df_port.copy()  # Start with original portfolio
    
    # --- Contract Comparison Logic ---
    # Compare contract numbers between Portfolio and CBSL PropertyMortgage
    print(f"\n=== CONTRACT COMPARISON BETWEEN CAR AND CBSL PROPERTYMORTGAGE ===")
    
    # Get unique contract numbers from both datasets
    portfolio_contracts = set(df_port_clean["contract_no"].astype(str).str.strip())
    cbsl_pm_contracts = set(df_pm_lookup["contract_no"].astype(str).str.strip())
    
    print(f"Portfolio contracts: {len(portfolio_contracts)}")
    print(f"CBSL PropertyMortgage contracts: {len(cbsl_pm_contracts)}")
    
    # Find exact matches
    exact_matches = portfolio_contracts.intersection(cbsl_pm_contracts)
    print(f"Exact contract matches: {len(exact_matches)}")
    
    # Create lookup dictionary for exact matches
    lookup_dict = {}
    for contract in exact_matches:
        lookup_dict[contract] = contract  # Map to the same contract number
    
    # Apply mapping to the final result dataframe
    df_result["mortgage"] = df_result["contract_no"].astype(str).str.strip().map(lookup_dict).fillna("#N/A")
    
    # Log results
    total_records = len(df_result)
    matched_records = (df_result["mortgage"] != "#N/A").sum()
    unmatched_records = (df_result["mortgage"] == "#N/A").sum()
    
    print(f"Contract matching results:")
    print(f"  Total portfolio records: {total_records}")
    print(f"  Matched contracts: {matched_records} ({(matched_records/total_records)*100:.1f}%)")
    print(f"  Unmatched contracts: {unmatched_records} ({(unmatched_records/total_records)*100:.1f}%)")
    
    # Show sample matches
    if matched_records > 0:
        sample_matches = df_result[df_result["mortgage"] != "#N/A"].head(5)
        print(f"\nSample matched contracts:")
        for idx, row in sample_matches.iterrows():
            print(f"  Portfolio: {row['contract_no']:15s} -> CBSL: {row['mortgage']}")
    
    # Show sample unmatched
    if unmatched_records > 0:
        sample_unmatched = df_result[df_result["mortgage"] == "#N/A"].head(5)
        print(f"\nSample unmatched contracts:")
        for idx, row in sample_unmatched.iterrows():
            print(f"  Portfolio: {row['contract_no']:15s} -> No match in CBSL PropertyMortgage")

    
    # Show mortgage value distribution
    print(f"\n=== MORTGAGE VALUE DISTRIBUTION ===")
    mortgage_counts = df_result["mortgage"].value_counts(dropna=False)
    for value, count in mortgage_counts.items():
        percentage = (count / len(df_result)) * 100
        print(f"  {str(value)[:30]:30s}: {count:5d} records ({percentage:5.1f}%)")
    
    # Show sample results
    print(f"\n=== SAMPLE MORTGAGE MATCHING RESULTS ===")
    matched_sample = df_result[df_result["mortgage"] != "#N/A"].head(5)
    if len(matched_sample) > 0:
        print("Sample MATCHED mortgage records:")
        for idx, row in matched_sample.iterrows():
            print(f"  Contract: {row['contract_no']:15s} -> Mortgage: {row['mortgage']}")
    
    unmatched_sample = df_result[df_result["mortgage"] == "#N/A"].head(5)
    if len(unmatched_sample) > 0:
        print("\nSample UNMATCHED mortgage records:")
        for idx, row in unmatched_sample.iterrows():
            print(f"  Contract: {row['contract_no']:15s} -> No Mortgage found")
    
    return df_result

def extract_cbsl_provision_values(file_sofp):
    """Extract FD Loan, and Difference values from CBSL Provision sheet"""
    try:
        print(f"\n=== READING CBSL PROVISION VALUES FROM SOFP FILE ===")
        print(f"Reading from: {file_sofp}")
        
        # Read the CBSL Provision sheet
        df_cbsl_prov = pd.read_excel(file_sofp, sheet_name="CBSL Provision", header=None, engine="openpyxl")
        print(f"CBSL Provision sheet shape: {df_cbsl_prov.shape}")
        
        # Extract values based on the structure we observed
        values = {}
        df_cbsl_prov = df_cbsl_prov[
            ~df_cbsl_prov.iloc[:, 1].astype(str).str.contains('FDL|Margin Trading', case=False, na=False)
        ].copy()
        # Look for FD Loan (should be around row 7, column 2)
        for i in range(len(df_cbsl_prov)):
            if pd.notna(df_cbsl_prov.iloc[i, 1]) and "FD Loan" in str(df_cbsl_prov.iloc[i, 1]):
                fd_loan_value = df_cbsl_prov.iloc[i, 2]
                if pd.notna(fd_loan_value):
                    values["FD Loan"] = float(fd_loan_value)
                    print(f"[OK] Found FD Loan: {values['FD Loan']:,.2f}")
                break
        
        # Look for Margin Trading (should be around row 8, column 2)
        # for i in range(len(df_cbsl_prov)):
        #     if pd.notna(df_cbsl_prov.iloc[i, 1]) and "Margin Trading" in str(df_cbsl_prov.iloc[i, 1]):
        #         margin_trading_value = df_cbsl_prov.iloc[i, 2]
        #         if pd.notna(margin_trading_value):
        #             values["Margin Trading"] = float(margin_trading_value)
        #             print(f"[OK] Found Margin Trading: {values['Margin Trading']:,.2f}")
        #         break
        
        # Look for Difference (should be around row 13, column 2)
        for i in range(len(df_cbsl_prov)):
            if pd.notna(df_cbsl_prov.iloc[i, 1]) and "Difference" in str(df_cbsl_prov.iloc[i, 1]):
                difference_value = df_cbsl_prov.iloc[i, 2]
                if pd.notna(difference_value):
                    values["Difference"] = float(difference_value)
                    print(f"[OK] Found Difference: {values['Difference']:,.2f}")
                break
        
        # Look for Net total from "As per Form 1" table (should be around row 33, column 4)
        for i in range(len(df_cbsl_prov)):
            if pd.notna(df_cbsl_prov.iloc[i, 1]) and "Total" in str(df_cbsl_prov.iloc[i, 1]):
                # Check if this is the "As per Form 1" Total row by looking at the context
                if i > 25 and i < 40:  # "As per Form 1" table is around rows 28-33
                    net_value = df_cbsl_prov.iloc[i, 4]  # Net column is column 4
                    if pd.notna(net_value):
                        values["Net_Total"] = float(net_value)
                        print(f"[OK] Found Net Total from As per Form 1: {values['Net_Total']:,.2f}")
                        break
        
        # Filter out FDL and Margin Trading from the Product column

        print(f"\n=== EXTRACTED CBSL PROVISION VALUES ===")
        for key, value in values.items():
            print(f"  {key:15s}: {value:15,.2f}")

        print(f"\n=== FILTERED DATAFRAME (After removing FDL and Margin Trading) ===")
        print(f"Original rows: {len(df_cbsl_prov)}, Filtered rows: {len(df_cbsl_prov)}")
                

        
        return values
        
    except Exception as e:
        print(f"[ERROR] Error reading CBSL Provision values from SOFP file: {e}")
        return {}

def load_risk_weight_categories(file_car):
    """Load risk weight categories from Types sheet in CAR Excel file"""
    try:
        print(f"\n=== LOADING RISK WEIGHT CATEGORIES FROM TYPES SHEET ===")
        
        # Read the Types sheet
        df_types = pd.read_excel(file_car, sheet_name="Types", engine="pyxlsb", header=None)
        print(f"Types sheet shape: {df_types.shape}")
        
        # Create mapping dictionary
        categories = {}
        
        # Column B (index 1) has codes, Column C (index 2) has category names
        for i in range(len(df_types)):
            if pd.notna(df_types.iloc[i, 1]) and pd.notna(df_types.iloc[i, 2]):
                code = str(df_types.iloc[i, 1]).strip()
                category = str(df_types.iloc[i, 2]).strip()
                categories[category] = code
                print(f"[OK] Category: {category} -> Code: {code}")
        
        print(f"\n=== LOADED {len(categories)} RISK WEIGHT CATEGORIES ===")
        return categories
        
    except Exception as e:
        print(f"[ERROR] Error loading risk weight categories: {e}")
        return {}

def categorize_risk_weight(df_port, categories):
    """Categorize contracts based on risk weight criteria"""
    try:
        print(f"\n=== CATEGORIZING RISK WEIGHT FOR {len(df_port)} CONTRACTS ===")
        
        # Initialize the new columns
        df_port["final_category_for_risk_weight"] = ""
        df_port["final_category_for_risk_weight_code"] = ""
        
        # First, implement logic for "Adjustment", "FD Loans", and "Others" in P category
        for idx, row in df_port.iterrows():
            product = str(row["product"]).strip()
            p_np = str(row["p_np_based_on_cbsl_provision"]).strip() if pd.notna(row["p_np_based_on_cbsl_provision"]) else ""
            
            # Logic for P category
            if p_np == "P":
                if product in ["FDL", "FD Loan"]:
                    # FD Loans -> P Others
                    category = "P Others"
                    code = categories.get(category, "")
                    df_port.at[idx, "final_category_for_risk_weight"] = category
                    df_port.at[idx, "final_category_for_risk_weight_code"] = code
                
                # elif product in ["Adjustment", "Margin Trading", "Difference"]:
                #     # Adjustment, Margin Trading, Difference -> P Others
                #     category = "P Others"
                #     code = categories.get(category, "")
                #     df_port.at[idx, "final_category_for_risk_weight"] = category
                #     df_port.at[idx, "final_category_for_risk_weight_code"] = code
                
                elif product in ["LE", "UV"]:
                    # Check if this is AFL Green Loans (exclude from P Leases)
                    contract_no = str(row["contract_no"]).strip()
                    equipment = str(row.get("equipment", "")).strip()
                    if "AFL" in contract_no and "Green" in equipment:
                        # AFL Green Loans -> P Others (excluded from P Leases)
                        category = "P Others"
                        code = categories.get(category, "")
                        df_port.at[idx, "final_category_for_risk_weight"] = category
                        df_port.at[idx, "final_category_for_risk_weight_code"] = code
                    else:
                        # LE and UV products -> P Leases (except AFL Green Loans)
                        category = "P Leases"
                        code = categories.get(category, "")
                        df_port.at[idx, "final_category_for_risk_weight"] = category
                        df_port.at[idx, "final_category_for_risk_weight_code"] = code
                
                else:
                    # Check if this P product is Corporate (P Corporates)
                    corporate_individual = str(row.get("corporate_individual", "")).strip()
                    if corporate_individual == "Corporate":
                        # P products with Corporate -> P Corporates
                        category = "P Corporates"
                        code = categories.get(category, "")
                        df_port.at[idx, "final_category_for_risk_weight"] = category
                        df_port.at[idx, "final_category_for_risk_weight_code"] = code
                    else:
                        # Check if this P product has mortgage value (P Real Estate)
                        mortgage = row.get("mortgage", "")
                        if pd.notna(mortgage) and str(mortgage).strip() != "#N/A" and str(mortgage).strip() != "nan":
                            # P products with mortgage values -> P Real Estate
                            category = "P Real Estate"
                            code = categories.get(category, "")
                            df_port.at[idx, "final_category_for_risk_weight"] = category
                            df_port.at[idx, "final_category_for_risk_weight_code"] = code
                        else:
                            # Default for other P products without mortgage -> P Others
                            df_port.at[idx, "final_category_for_risk_weight"] = "P Others"
                            df_port.at[idx, "final_category_for_risk_weight_code"] = categories.get("P Others", "")
            
            # Logic for NP category
            elif p_np == "NP":
                # Check if NP contract has mortgage value (not #N/A)
                mortgage = row.get("mortgage", "")
                margin_20_percent = str(row.get("margin_20_percent", "")).strip()
                
                if pd.notna(mortgage) and str(mortgage).strip() != "#N/A" and str(mortgage).strip() != "nan":
                    # NP contracts with mortgage values
                    if margin_20_percent == "Below 20%":
                        category = "NP Real Estate Below 20%"
                        code = categories.get(category, "")
                        df_port.at[idx, "final_category_for_risk_weight"] = category
                        df_port.at[idx, "final_category_for_risk_weight_code"] = code
                    elif margin_20_percent == "Above 20%":
                        category = "NP Other Above 20%"
                        code = categories.get(category, "")
                        df_port.at[idx, "final_category_for_risk_weight"] = category
                        df_port.at[idx, "final_category_for_risk_weight_code"] = code
                    else:
                        # NP with mortgage but unknown margin -> leave blank
                        df_port.at[idx, "final_category_for_risk_weight"] = ""
                        df_port.at[idx, "final_category_for_risk_weight_code"] = ""
                else:
                    # NP contracts without mortgage values
                    if margin_20_percent == "Below 20%":
                        category = "NP Other Below 20%"
                        code = categories.get(category, "")
                        df_port.at[idx, "final_category_for_risk_weight"] = category
                        df_port.at[idx, "final_category_for_risk_weight_code"] = code
                    elif margin_20_percent == "Above 20%":
                        category = "NP Other Above 20%"
                        code = categories.get(category, "")
                        df_port.at[idx, "final_category_for_risk_weight"] = category
                        df_port.at[idx, "final_category_for_risk_weight_code"] = code
                    else:
                        # NP without mortgage and unknown margin -> leave blank
                        df_port.at[idx, "final_category_for_risk_weight"] = ""
                        df_port.at[idx, "final_category_for_risk_weight_code"] = ""
            
            # Default for unknown P/NP
            else:
                df_port.at[idx, "final_category_for_risk_weight"] = "P Others"
                df_port.at[idx, "final_category_for_risk_weight_code"] = categories.get("P Others", "")
        
        # Count categorized contracts
        categorized_count = len(df_port[df_port["final_category_for_risk_weight"] != ""])
        print(f"\n=== CATEGORIZATION COMPLETE ===")
        print(f"Categorized contracts: {categorized_count} out of {len(df_port)}")
        
        # Show distribution
        category_dist = df_port["final_category_for_risk_weight"].value_counts()
        print(f"\nCategory distribution:")
        for category, count in category_dist.items():
            print(f"  {category}: {count} contracts")
        
        return df_port
        
    except Exception as e:
        print(f"[ERROR] Error categorizing risk weight: {e}")
        return df_port

def add_cbsl_provision_contracts(df_port, cbsl_values, month, year):
    """Add FD Loan, and Difference as new contract entries to portfolio"""
    if not cbsl_values:
        print("[WARN] No CBSL Provision values to add")
        return df_port
    
    print(f"\n=== ADDING CBSL PROVISION CONTRACTS TO PORTFOLIO ===")
    
    # Create new contract entries
    new_contracts = []
    
    for contract_name, amount in cbsl_values.items():
        # Skip Net_Total as it's only used for H2 cell, not as a contract entry
        if contract_name == "Net_Total":
            continue
            
        if amount != 0:  # Only add non-zero amounts
            # Generate a unique contract number for each CBSL Provision entry
            contract_no = f"CBSL-{contract_name.replace(' ', '')}-{month}-{year}"
            
            new_contract = {
                "contract_no": contract_no,
                "client_code": f"CBSL-{contract_name.replace(' ', '')}",
                "mid": 2,  # Corporate (since these are CBSL provisions)
                "product": contract_name,
                "gross_with_dp": amount,  # Use exact value including negative sign
                "ifrs_provision_with_dp": 0.0,  # No IFRS provision for these
                "iis": 0.0,  # No IIS for these
                "gross_iis_imp": amount,  # Same as gross since no provisions
                "imp_percent": 0.0,  # No impairment
                "margin_20_percent": "Below 20%",  # 0% is below 20%
                "p_np_based_on_cbsl_provision": None,  # Will be filled by CBSL lookup
                "equipment": "",
                "corporate_individual": "Corporate",
                "mortgage": "#N/A",  # No mortgage for these
                "final_category_for_risk_weight": "",
                "a": 0.0,
                "b": 0.0
            }
            
            new_contracts.append(new_contract)
            print(f"[OK] Added contract: {contract_no} - {contract_name}: {amount:,.2f}")
    
    if new_contracts:
        # Convert to DataFrame and append to existing portfolio
        df_new_contracts = pd.DataFrame(new_contracts)
        df_port_updated = pd.concat([df_port, df_new_contracts], ignore_index=True)
        
        print(f"\n=== CBSL PROVISION CONTRACTS ADDED ===")
        print(f"Added {len(new_contracts)} new contract entries")
        print(f"Total portfolio records: {len(df_port)} -> {len(df_port_updated)}")
        
        return df_port_updated
    else:
        print("[WARN] No valid CBSL Provision contracts to add")
        return df_port

def populate_mortgage_from_cbsl(df_port, file_cbsl):
    """Populate Mortgage column from CBSL PropertyMortgage sheet using VLOOKUP-style logic"""
    try:
        print("\n=== READING CBSL PropertyMortgage SHEET FOR MORTGAGE MAPPING ===")
        xls_cbsl_pm = pd.ExcelFile(file_cbsl, engine="pyxlsb")
        print(f"Available sheets in CBSL file: {xls_cbsl_pm.sheet_names}")
        
        # Find PropertyMortgage sheet
        pm_sheet = None
        
        # Prefer exact name
        if "PropertyMortgage" in xls_cbsl_pm.sheet_names:
            pm_sheet = "PropertyMortgage"
            print(f"[OK] Found exact PropertyMortgage sheet")
        else:
            # Try common variants
            for name in xls_cbsl_pm.sheet_names:
                lname = name.lower().replace(" ", "").replace("-", "").replace("_", "")
                if ("property" in lname and "mortgage" in lname) or "propertymortgage" in lname:
                    pm_sheet = name
                    print(f"[OK] Found PropertyMortgage-like sheet: {pm_sheet}")
                    break
        
        if not pm_sheet:
            print("[ERROR] No PropertyMortgage sheet found in CBSL file")
            df_port["mortgage"] = "#N/A"
            return df_port
        
        print(f"Reading CBSL PropertyMortgage sheet: {pm_sheet}")
        
        # Read the PropertyMortgage sheet - try different approaches
        df_pm = None
        
        # Method 1: Try reading with headers from different rows
        for header_row in [0, 1, 2, 3, 4, 5]:
            try:
                print(f"Attempting to read PropertyMortgage with header row {header_row}...")
                df_temp = pd.read_excel(xls_cbsl_pm, sheet_name=pm_sheet, engine="pyxlsb", header=4)
                df_temp = normalize_columns(df_temp)
                
                # Check if we have reasonable columns
                if len(df_temp.columns) > 1 and len(df_temp) > 0:
                    print(f"[OK] Successfully read PropertyMortgage with header row {header_row}")
                    print(f"Columns found: {list(df_temp.columns)[:10]}")  # Show first 10 columns
                    df_pm = df_temp
                    break
                    
            except Exception as e:
                print(f"Failed with header row {header_row}: {e}")
                continue
        
        # Method 2: If no success, try reading raw and finding appropriate data range
        if df_pm is None:
            try:
                print("Trying raw read approach...")
                df_raw = pd.read_excel(xls_cbsl_pm, sheet_name=pm_sheet, engine="pyxlsb", header=None)
                print(f"Raw PropertyMortgage sheet dimensions: {df_raw.shape}")
                
                # Display first few rows to understand structure
                print("First 8 rows of PropertyMortgage sheet:")
                for i in range(min(8, len(df_raw))):
                    print(f"Row {i+1}: {df_raw.iloc[i, :8].tolist()}")  # Show first 8 columns
                
                # Try to find the header row by looking for "CONTRACT" in the first 10 rows
                header_row_found = None
                for i in range(min(10, len(df_raw))):
                    row_values = [str(val).upper() for val in df_raw.iloc[i, :10] if pd.notna(val)]
                    if any("CONTRACT" in val for val in row_values):
                        header_row_found = i
                        print(f"[OK] Found CONTRACT in row {i+1} (0-indexed: {i})")
                        break
                
                if header_row_found is not None:
                    headers = df_raw.iloc[header_row_found, :].values
                    data = df_raw.iloc[header_row_found + 1:, :]
                    
                    # Clean up headers
                    clean_headers = []
                    for h in headers:
                        if pd.notna(h) and str(h).strip():
                            clean_headers.append(str(h).strip())
                        else:
                            clean_headers.append(f"Col_{len(clean_headers)}")
                    
                    # Ensure we don't have more columns than headers
                    n_headers = len(clean_headers)
                    if len(data.columns) > n_headers:
                        data = data.iloc[:, :n_headers]
                    elif len(data.columns) < n_headers:
                        clean_headers = clean_headers[:len(data.columns)]
                    
                    data.columns = clean_headers
                    df_pm = data.reset_index(drop=True)
                    df_pm = normalize_columns(df_pm)
                    
                    print(f"[OK] Successfully extracted PropertyMortgage headers from row {header_row_found + 1}: {clean_headers[:10]}")
                    print(f"PropertyMortgage data shape after processing: {df_pm.shape}")
                else:
                    raise Exception("Could not find CONTRACT header in PropertyMortgage sheet")
                    
            except Exception as e:
                print(f"[ERROR] Failed to read PropertyMortgage with raw approach: {e}")
                df_port["mortgage"] = "#N/A"
                return df_port
        
        if df_pm is None:
            print("[ERROR] Could not read PropertyMortgage sheet successfully")
            df_port["mortgage"] = "#N/A"
            return df_port
        
        # Normalize columns after reading
        df_pm = normalize_columns(df_pm)
        print(f"\n=== PropertyMortgage Sheet Column Analysis & Logging ===")
        print(f"Total columns: {len(df_pm.columns)}")
        print("All PropertyMortgage column names after normalization:")
        for i, col in enumerate(df_pm.columns):
            print(f"  {i+1:2d}. '{col}'")
            col_lower = str(col).lower()
            if "contract" in col_lower:
                print(f"      *** POTENTIAL CONTRACT COLUMN ***")
            if any(term in col_lower for term in ['mortgage', 'property', 'exposure', 'gross']):
                print(f"      *** POTENTIAL MORTGAGE VALUE COLUMN ***")
        print("=" * 50)
        
        # Detect contract column in PropertyMortgage
        pm_contract = None
        pm_candidates = [
            "contract_no", "contrtact_no", "contractno", "contract_number", "facility_no", "facility_number", 
            "agreement_no", "agreement_number"
        ]
        
        # Exclude any YARD-related columns from consideration
        pm_cols_filtered = [c for c in df_pm.columns if "yard" not in str(c).lower()]
        
        # Prefer exact 'contract_no' if present
        if "contract_no" in pm_cols_filtered:
            pm_contract = "contract_no"
            print(f"[OK] Using specific PropertyMortgage contract column: '{pm_contract}' (ignoring YARD columns)")
        
        # First try exact matches among filtered columns
        if not pm_contract:
            for term in pm_candidates:
                if term in pm_cols_filtered:
                    pm_contract = term
                    print(f"[OK] Found EXACT PropertyMortgage contract column match: '{pm_contract}' (ignoring YARD columns)")
                    break
        
        # If no exact match, try partial matches
        if not pm_contract:
            print("No exact contract column match found in PropertyMortgage, searching partial matches (ignoring YARD columns)...")
            for col in pm_cols_filtered:
                if any(pattern in col for pattern in ['contract', 'facility', 'agreement']):
                    pm_contract = col
                    print(f"[OK] Found partial PropertyMortgage contract column match: '{pm_contract}'")
                    break
        
        # Detect mortgage value column with enhanced search based on your Excel formula
        pm_value = None
        pm_value_candidates = [
            # Based on common PropertyMortgage column naming patterns
            "cbsl_gross_exposure", "propertymortgage", "property_mortgage", "mortgage", 
            "gross_exposure", "exposure_amount", "mortgage_amount", "property_value"
        ]
        
        # First try exact matches for mortgage value
        print("Searching for PropertyMortgage value column...")
        for term in pm_value_candidates:
            if term in df_pm.columns:
                pm_value = term
                print(f"[OK] Found EXACT PropertyMortgage value column match: '{pm_value}'")
                break
        
        # If no exact match, try partial matches
        if not pm_value:
            print("No exact PropertyMortgage value column match found, searching partial matches...")
            for col in df_pm.columns:
                if any(pattern in col for pattern in ['mortgage', 'property', 'exposure', 'gross', 'amount', 'value']):
                    pm_value = col
                    print(f"[OK] Found partial PropertyMortgage value column match: '{pm_value}'")
                    break
        
        print(f"\n=== PropertyMortgage Column Detection Results ===")
        print(f"Final PropertyMortgage column selections:")
        print(f"PropertyMortgage Contract column: {pm_contract}")
        print(f"PropertyMortgage Value column: {pm_value}")
        
        # Log PropertyMortgage column detection success/failure
        pm_mapping_status = {
            "Contract": "[OK] FOUND" if pm_contract and pm_contract in df_pm.columns else "[ERROR] NOT FOUND",
            "Value": "[OK] FOUND" if pm_value and pm_value in df_pm.columns else "[ERROR] NOT FOUND"
        }
        
        print(f"\nPropertyMortgage Column Detection Status:")
        for col_name, status in pm_mapping_status.items():
            print(f"  {col_name:10s}: {status}")
        print("=" * 50)
        
        # Show all available columns if no match found
        if not pm_contract or not pm_value:
            print(f"\n[WARN] Missing required PropertyMortgage columns. Available columns:")
            for i, col in enumerate(df_pm.columns, 1):
                print(f"  {i:2d}. '{col}'")
                # Check if this might be a contract or mortgage value column
                col_lower = str(col).lower()
                if any(term in col_lower for term in ['contract', 'facility', 'agreement']):
                    print(f"       ^ Potential CONTRACT column")
                if any(term in col_lower for term in ['mortgage', 'property', 'exposure', 'gross', 'amount', 'value']):
                    print(f"       ^ Potential MORTGAGE VALUE column")
        
        # Proceed with enhanced mortgage lookup if both columns are found
        if pm_contract and pm_value and pm_contract in df_pm.columns and pm_value in df_pm.columns:
            print(f"\n[OK] Both required PropertyMortgage columns found - proceeding with enhanced lookup")
            print(f"   Contract column: '{pm_contract}'")
            print(f"   Value column: '{pm_value}'")
            
            # Use enhanced mortgage lookup function
            df_port = enhanced_mortgage_lookup(df_port, df_pm, pm_contract, pm_value)
            
            # Log mortgage lookup results
            total_records = len(df_port)
            mortgage_found = (df_port["mortgage"] != "#N/A").sum()
            mortgage_not_found = (df_port["mortgage"] == "#N/A").sum()
            
            print(f"\n=== MORTGAGE LOOKUP RESULTS ===")
            print(f"Total records processed: {total_records}")
            print(f"Mortgage values found: {mortgage_found} ({(mortgage_found/total_records)*100:.1f}%)")
            print(f"Mortgage values NOT found: {mortgage_not_found} ({(mortgage_not_found/total_records)*100:.1f}%)")
            print("[OK] Mortgage column populated from CBSL PropertyMortgage sheet using enhanced lookup")
            
        else:
            print("[ERROR] Could not find required PropertyMortgage columns for lookup. Setting Mortgage to #N/A for all records.")
            df_port["mortgage"] = "#N/A"
        
    except Exception as e:
        print(f"[WARN] Failed to populate Mortgage from PropertyMortgage sheet: {e}")
        print("Setting Mortgage to #N/A for all records.")
        df_port["mortgage"] = "#N/A"
    
    return df_port
def update_car_with_prod_contracts(file_prod, file_car, out_file):
    """
    Reads contract numbers from 'Prod. wise Class. of Loans - (month) 2025' (C1 & C2 Working sheet),
    and appends/updates them in 'CAR Working (month) 2025' (Portfolio sheet).
    """
    import pandas as pd

    # Read Prod. wise (C1 & C2 Working)
    df_prod = pd.read_excel(file_prod, sheet_name="C1 & C2 Working", engine="pyxlsb")
    df_prod.columns = [str(c).strip().lower().replace(" ", "_") for c in df_prod.columns]
    
    # Detect contract column in Prod. wise
    contract_col_prod = None
    for c in df_prod.columns:
        if "contract" in c.lower():
            contract_col_prod = c
            break
    if not contract_col_prod:
        raise ValueError("No Contract column found in Prod. wise C1 & C2 Working")
    
    prod_contracts = df_prod[contract_col_prod].dropna().astype(str).str.strip().unique()
    print(f"[INFO] Found {len(prod_contracts)} unique contracts in Prod. wise file")

    # Read CAR Working (Portfolio)
    df_car = pd.read_excel(file_car, sheet_name="Portfolio", engine="pyxlsb", header=3)
    df_car.columns = [str(c).strip().lower().replace(" ", "_") for c in df_car.columns]

    if "contract_no" not in df_car.columns:
        raise ValueError("No Contract No column found in CAR Working Portfolio sheet")

    # Append new contracts
    existing_contracts = set(df_car["contract_no"].astype(str).str.strip())
    new_contracts = [c for c in prod_contracts if c not in existing_contracts]

    df_new = pd.DataFrame({"contract_no": new_contracts})
    df_car_updated = pd.concat([df_car, df_new], ignore_index=True)

    print(f"[OK] Added {len(new_contracts)} new contracts to CAR Working")
    print(f"Total CAR Portfolio contracts: {len(df_car)} → {len(df_car_updated)}")

    # Save updated CAR Working
    df_car_updated.to_excel(out_file, sheet_name="Portfolio", index=False)
    print(f"[SAVED] Updated CAR Working file: {out_file}")

def update_portfolio_with_pivot_data(wb_c1_c6, df_port, out_folder, month, year):
    """Update Portfolio sheet with pivot data without replacing the entire sheet"""
    try:
        print("\n=== UPDATING PORTFOLIO SHEET WITH PIVOT DATA (PRESERVING FORMATTING) ===")
        
        if wb_c1_c6 is None:
            print("[WARN] No workbook provided, skipping Portfolio sheet update")
            return None
            
        # Find Portfolio sheet
        portfolio_sheet = None
        for sheet in wb_c1_c6.sheetnames:
            if "portfolio" in sheet.lower():
                portfolio_sheet = sheet
                break
        
        if not portfolio_sheet:
            print("[WARN] No Portfolio sheet found in workbook")
            return None
            
        ws = wb_c1_c6[portfolio_sheet]
        print(f"Using existing Portfolio sheet: {portfolio_sheet}")
        
        # Create a mapping of contract_no to row data for efficient lookup
        df_port_indexed = df_port.set_index('contract_no')
        
        # Get column headers from the existing sheet to understand the structure
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                headers.append(str(cell_value).strip())
            else:
                headers.append(f"Col_{col}")
        
        print(f"Existing sheet headers: {headers[:10]}...")  # Show first 10 headers
        
        # Find column indices for key fields we want to update
        column_mapping = {}
        key_fields = [
            'p_np_based_on_cbsl_provision',
            'mortgage', 
            'gross_with_dp',
            'ifrs_provision_with_dp',
            'iis',
            'gross_iis_imp',
            'imp_percent',
            'margin_20_percent',
            'final_category_for_risk_weight',
            'final_category_for_risk_weight_code',
            'corporate_individual'
        ]
        
        for field in key_fields:
            for i, header in enumerate(headers):
                if field.lower() in header.lower() or header.lower() in field.lower():
                    column_mapping[field] = i + 1  # Excel is 1-indexed
                    print(f"[OK] Found column for {field}: '{header}' at column {i + 1}")
                    break
        
        # Update data row by row
        updated_count = 0
        for row_num in range(2, ws.max_row + 1):  # Skip header row
            # Get contract number from first column (assuming it's contract_no)
            contract_cell = ws.cell(row=row_num, column=1)
            if not contract_cell.value:
                continue
                
            contract_no = str(contract_cell.value).strip()
            
            # Check if this contract exists in our pivot data
            if contract_no in df_port_indexed.index:
                row_data = df_port_indexed.loc[contract_no]
                
                # Update each field that we found columns for
                for field, col_num in column_mapping.items():
                    if field in row_data.index and pd.notna(row_data[field]):
                        cell = ws.cell(row=row_num, column=col_num)
                        old_value = cell.value
                        new_value = row_data[field]
                        
                        # Only update if value has changed
                        if old_value != new_value:
                            cell.value = new_value
                            updated_count += 1
                            
                            # Debug first few updates
                            if updated_count <= 5:
                                print(f"  Updated {field} for {contract_no}: {old_value} -> {new_value}")
        
        print(f"[OK] Updated {updated_count} cells in Portfolio sheet")
        
        # Save the updated workbook
        out_path = out_folder / f"NBD_MF_20_C3_{month}_{year}_report.xlsx"
        wb_c1_c6.save(out_path)
        print(f"[OK] Updated Portfolio sheet saved to: {out_path}")
        
        return out_path
        
    except Exception as e:
        print(f"[ERROR] Error updating Portfolio sheet: {e}")
        import traceback
        traceback.print_exc()
        return None

def update_c1_c6_with_car_pivot_values(file_c1_c6, file_car, df_port, out_folder):
    """Edit existing C1-C6 file directly and save edited copy into output folder"""
    try:
        print("\n=== UPDATING C1-C6 WITH CAR PIVOT VALUES (IN-PLACE) ===")

        # --- Load existing workbook ---
        wb = load_workbook(file_c1_c6)
        print(f"Available sheets: {wb.sheetnames}")

        # Find Portfolio sheet
        portfolio_sheet = None
        for sheet in wb.sheetnames:
            if any(f'c{i}' in sheet.lower() for i in range(1, 7)) or "portfolio" in sheet.lower():
                portfolio_sheet = sheet
                break
        if not portfolio_sheet:
            portfolio_sheet = wb.sheetnames[0]

        ws = wb[portfolio_sheet]
        print(f"Using sheet: {portfolio_sheet}")

        # --- Build pivot from CAR Working data ---
        pivot_data = df_port.groupby('final_category_for_risk_weight')['gross_iis_imp'].sum().reset_index()
        pivot_data = pivot_data.merge(
            df_port.groupby('final_category_for_risk_weight')['final_category_for_risk_weight_code'].first().reset_index(),
            on='final_category_for_risk_weight'
        )

        category_mapping = {
            str(row['final_category_for_risk_weight_code']): row['gross_iis_imp']
            for _, row in pivot_data.iterrows()
        }

        # --- Update worksheet directly ---
        updated_count = 0
        total_gross_iis_imp = 0

        for row in ws.iter_rows(min_row=2):  # Skip header
            category_code = row[0].value  # First column
            if category_code and str(category_code) in category_mapping:
                new_val = category_mapping[str(category_code)]
                row[2].value = new_val  # Update 3rd column (Gross-IIS-IMP)
                updated_count += 1
                total_gross_iis_imp += new_val

        # Update or add Grand Total row
        found_total = False
        for row in ws.iter_rows():
            if row[1].value and "Grand Total" in str(row[1].value):
                row[2].value = total_gross_iis_imp
                found_total = True
                break
        if not found_total:
            ws.append([None, "Grand Total", total_gross_iis_imp])

        print(f"[OK] Updated {updated_count} rows in {portfolio_sheet}, total {total_gross_iis_imp:,.2f}")

        # --- Save to output folder ---
        out_folder.mkdir(parents=True, exist_ok=True)
        out_path = out_folder / file_c1_c6.name

        # Copy original to output first (to preserve everything), then overwrite
        shutil.copy(file_c1_c6, out_path)
        wb.save(out_path)

        print(f"[OK] Edited file saved to: {out_path}")
        return out_path

    except Exception as e:
        print(f"[ERROR] Error updating C1-C6: {e}")
        return None


def update_pivot_table(file_car, df_port):
    """Update the pivot table in CAR Working file with new categorization data"""
    try:
        print("Creating pivot table summary...")
        
        # Create pivot table data
        pivot_data = df_port.groupby('final_category_for_risk_weight').agg({
            'gross_iis_imp': 'sum',
            'gross_with_dp': 'sum', 
            'ifrs_provision_with_dp': 'sum',
            'contract_no': 'count'
        }).round(2)
        
        # Add risk weight codes
        pivot_data['risk_weight_code'] = pivot_data.index.map(
            lambda x: df_port[df_port['final_category_for_risk_weight'] == x]['final_category_for_risk_weight_code'].iloc[0]
        )
        
        # Reorder columns
        pivot_data = pivot_data[['risk_weight_code', 'gross_iis_imp', 'gross_with_dp', 'ifrs_provision_with_dp', 'contract_no']]
        pivot_data.columns = ['Risk Weight Code', 'Gross-IIS-IMP', 'Gross with DP', 'IFRS Provision', 'Count']
        
        # Sort by risk weight code
        pivot_data = pivot_data.sort_index()
        
        print("Pivot table data:")
        print(pivot_data)
        
        # Save pivot table to Excel
        pivot_out_path = Path(file_car).parent / "Pivot_Table_Summary.xlsx"
        with pd.ExcelWriter(pivot_out_path, engine='openpyxl') as writer:
            pivot_data.to_excel(writer, sheet_name='Pivot Summary', index=True)
            
            # Also create a detailed breakdown
            detailed_data = df_port.groupby(['final_category_for_risk_weight_code', 'final_category_for_risk_weight']).agg({
                'gross_iis_imp': 'sum',
                'contract_no': 'count'
            }).round(2)
            detailed_data.columns = ['Gross-IIS-IMP', 'Count']
            detailed_data.to_excel(writer, sheet_name='Detailed Breakdown', index=True)
        
        print(f"[OK] Pivot table summary saved to: {pivot_out_path}")
        
        # Try to update the actual CAR Working file pivot sheet
        excel_pivot = None
        wb_pivot = None
        try:
            excel_pivot = win32.Dispatch("Excel.Application")
            
            # Set Excel properties with error handling
            try:
                excel_pivot.Visible = False
            except Exception as visible_error:
                print(f"[WARN] Could not set Excel Visible property for pivot: {visible_error}")
                
            try:
                excel_pivot.DisplayAlerts = False
            except Exception as alerts_error:
                print(f"[WARN] Could not set Excel DisplayAlerts property for pivot: {alerts_error}")
            
            wb_pivot = excel_pivot.Workbooks.Open(str(file_car))
            
            # Check if Pivot sheet exists
            try:
                ws_pivot = wb_pivot.Worksheets("Pivot")
                print("Found existing Pivot sheet, updating...")
                
                # Clear existing data
                ws_pivot.Cells.Clear()
                
                # Write headers in the exact format from the image
                ws_pivot.Cells(1, 1).Value = "Row Labels"
                ws_pivot.Cells(1, 2).Value = "Sum of Gross-IIS-IMP"
                
                # Format headers with blue background and bold text
                ws_pivot.Cells(1, 1).Interior.Color = 0xCCE5FF  # Light blue background
                ws_pivot.Cells(1, 2).Interior.Color = 0xCCE5FF  # Light blue background
                ws_pivot.Cells(1, 1).Font.Bold = True
                ws_pivot.Cells(1, 2).Font.Bold = True
                
                # Write data in the exact order from the image (adjusted for actual categories)
                row_order = [
                    "NP Other Above 20%",
                    "NP Other Below 20%", 
                    "P Corporates",
                    "P Leases",
                    "P Others",
                    "P Real Estate"
                ]
                
                row_num = 2
                for category in row_order:
                    if category in pivot_data.index:
                        ws_pivot.Cells(row_num, 1).Value = category
                        ws_pivot.Cells(row_num, 2).Value = pivot_data.loc[category, 'Gross-IIS-IMP']
                        row_num += 1
                
                # Add Grand Total row
                grand_total_row = row_num
                ws_pivot.Cells(grand_total_row, 1).Value = "Grand Total"
                ws_pivot.Cells(grand_total_row, 2).Value = pivot_data['Gross-IIS-IMP'].sum()
                
                # Format Grand Total row with blue background and bold text
                ws_pivot.Cells(grand_total_row, 1).Interior.Color = 0xCCE5FF  # Light blue background
                ws_pivot.Cells(grand_total_row, 2).Interior.Color = 0xCCE5FF  # Light blue background
                ws_pivot.Cells(grand_total_row, 1).Font.Bold = True
                ws_pivot.Cells(grand_total_row, 2).Font.Bold = True
                
                # Add additional total rows
                row_num += 1
                
                # Total IFRS Provision row
                ws_pivot.Cells(row_num, 1).Value = "Total IFRS Provision"
                ws_pivot.Cells(row_num, 2).Value = pivot_data['IFRS Provision'].sum()
                ws_pivot.Cells(row_num, 1).Interior.Color = 0xCCE5FF  # Light blue background
                ws_pivot.Cells(row_num, 2).Interior.Color = 0xCCE5FF  # Light blue background
                ws_pivot.Cells(row_num, 1).Font.Bold = True
                ws_pivot.Cells(row_num, 2).Font.Bold = True
                
                row_num += 1
                
                # Total Gross with DP row
                ws_pivot.Cells(row_num, 1).Value = "Total Gross with DP"
                ws_pivot.Cells(row_num, 2).Value = pivot_data['Gross with DP'].sum()
                ws_pivot.Cells(row_num, 1).Interior.Color = 0xCCE5FF  # Light blue background
                ws_pivot.Cells(row_num, 2).Interior.Color = 0xCCE5FF  # Light blue background
                ws_pivot.Cells(row_num, 1).Font.Bold = True
                ws_pivot.Cells(row_num, 2).Font.Bold = True
                
                row_num += 1
                
                # Total Gross-IIS-IMP row (same as Grand Total but with different label)
                ws_pivot.Cells(row_num, 1).Value = "Total Gross-IIS-IMP"
                ws_pivot.Cells(row_num, 2).Value = pivot_data['Gross-IIS-IMP'].sum()
                ws_pivot.Cells(row_num, 1).Interior.Color = 0xCCE5FF  # Light blue background
                ws_pivot.Cells(row_num, 2).Interior.Color = 0xCCE5FF  # Light blue background
                ws_pivot.Cells(row_num, 1).Font.Bold = True
                ws_pivot.Cells(row_num, 2).Font.Bold = True
                
                # Format numbers with commas (no decimals for whole numbers)
                for i in range(2, row_num + 1):
                    ws_pivot.Cells(i, 2).NumberFormat = "#,##0"
                
                # Auto-fit columns
                ws_pivot.Columns.AutoFit()
                
                wb_pivot.Save()
                print("[OK] Pivot sheet updated successfully in CAR Working file")
                
            except Exception as e:
                print(f"[WARN] Could not update Pivot sheet: {e}")
                print("Pivot table data saved to separate file instead")
            
        except Exception as e:
            print(f"[WARN] Could not update CAR Working file directly: {e}")
            print("Pivot table data saved to separate file instead")
            
        finally:
            # Clean up pivot Excel objects
            try:
                if wb_pivot is not None:
                    wb_pivot.Close(SaveChanges=False)
                    wb_pivot = None
            except Exception as close_error:
                print(f"[WARN] Error closing pivot workbook: {close_error}")
                
            try:
                if excel_pivot is not None:
                    excel_pivot.Quit()
                    excel_pivot = None
            except Exception as quit_error:
                print(f"[WARN] Error quitting pivot Excel: {quit_error}")
                
            # Force cleanup
            try:
                import gc
                gc.collect()
            except:
                pass
            
    except Exception as e:
        print(f"[ERROR] Error updating pivot table: {e}")


def main(wb_c1_c6, file_car, file_prod, file_cbsl, file_sofp=None, out_folder=None, file_c1_c6=None):
    """
    Main function to process C3 report data and return modified workbook
    
    Args:
        wb_c1_c6: openpyxl workbook object for C1-C6 file
        file_car: Path to CAR Working file
        file_prod: Path to Prod. wise Class. of Loans file
        file_cbsl: Path to CBSL Provision Comparison file
        file_sofp: Path to SOFP file (optional)
        out_folder: Path to output folder for saving reports (optional)
        file_c1_c6: Path to C1-C6 file (optional, for updating operations)
    
    Returns:
        Modified openpyxl workbook object
    """
    try:
        print(f"\n=== PROCESSING C3 REPORT ===")
        print(f"CAR Working: {file_car}")
        print(f"Prod wise: {file_prod}")
        print(f"CBSL Provision: {file_cbsl}")
        if file_sofp:
            print(f"SOFP: {file_sofp}")
        
        # Set up output folder if not provided
        if out_folder is None:
            from pathlib import Path
            out_folder = Path("outputs/monthly") / "temp_c3_output"
            out_folder.mkdir(parents=True, exist_ok=True)
            print(f"[INFO] Using default output folder: {out_folder}")
        else:
            print(f"[INFO] Using provided output folder: {out_folder}")
        
        # Extract month and year from CAR Working filename
        month, year = get_month_year_from_filename(file_car.name)
        if not month or not year:
            print(f"[ERROR] Could not parse month/year from filename: {file_car.name}")
            return None
        
        print(f"[OK] Parsed month: {month}, year: {year}")
        
        # Read source data from xlsb files
        print("Reading 'Prod. wise' C1 & C2 Working sheet (xlsb)...")
        
        def read_prod_sheet(path):
            try:
                df = pd.read_excel(path, sheet_name="C1 & C2 Working", engine="pyxlsb", header=[0, 1])
                if isinstance(df.columns, pd.MultiIndex):
                    new_cols = []
                    for top, sub in df.columns:
                        name = str(sub).strip() if str(sub).strip() and str(sub).lower() != "nan" else str(top).strip()
                        new_cols.append(name)
                    df.columns = new_cols
                df = normalize_columns(df)
                return df
            except Exception:
                df = pd.read_excel(path, sheet_name="C1 & C2 Working", engine="pyxlsb")
                df = normalize_columns(df)
                return df
        
        df_prod = read_prod_sheet(file_prod)
        print(f"Loaded {len(df_prod)} rows from production data")
        
        # Log Production sheet columns
        print(f"\n=== Production Sheet Column Analysis & Logging ===")
        print(f"Total columns: {len(df_prod.columns)}")
        print("All Production column names after normalization:")
        for i, col in enumerate(df_prod.columns):
            print(f"  {i+1:2d}. '{col}'")
            if "contract" in str(col).lower():
                print(f"      *** POTENTIAL CONTRACT COLUMN ***")
            if "client" in str(col).lower():
                print(f"      *** POTENTIAL CLIENT COLUMN ***")
            if "product" in str(col).lower():
                print(f"      *** POTENTIAL PRODUCT COLUMN ***")
            if "gross" in str(col).lower() and "dp" in str(col).lower():
                print(f"      *** POTENTIAL GROSS WITH DP COLUMN ***")
            if "irfs" in str(col).lower() or "ifrs" in str(col).lower():
                print(f"      *** POTENTIAL IFRS PROVISION COLUMN ***")
            if "iis" in str(col).lower():
                print(f"      *** POTENTIAL IIS COLUMN ***")
            if "equipment" in str(col).lower():
                print(f"      *** POTENTIAL EQUIPMENT COLUMN ***")
        print("=" * 50)
        
        # Read CAR Working
        print("Reading 'CAR Working' Portfolio sheet...")
        try:
            df_car = pd.read_excel(file_car, sheet_name="Portfolio", engine="pyxlsb",header=3)
            df_car = normalize_columns(df_car)
            
            # Log CAR Working sheet columns
            print(f"\n=== CAR Working Sheet Column Analysis & Logging ===")
            print(f"Total columns: {len(df_car.columns)}")
            print("All CAR Working column names after normalization:")
            for i, col in enumerate(df_car.columns):
                print(f"  {i+1:2d}. '{col}'")
                if "contract" in str(col).lower():
                    print(f"      *** POTENTIAL CONTRACT COLUMN ***")
                if "p_np" in str(col).lower() or "cbsl" in str(col).lower():
                    print(f"      *** POTENTIAL P/NP COLUMN ***")
                if "mortgage" in str(col).lower():
                    print(f"      *** POTENTIAL MORTGAGE COLUMN ***")
            print("=" * 50)
            
        except Exception as e:
            print(f"[WARN] Could not read Portfolio sheet from CAR Working ({e})")
            df_car = pd.DataFrame()
        
        # Read CBSL data - ACCESS 3RD LEVEL COLUMNS DIRECTLY
        print("Reading 'CBSL Provision Comparison' - Portfolio sheet...")
        df_cbsl = None
        try:
            xls_cbsl = pd.ExcelFile(file_cbsl, engine="pyxlsb")
            print(f"Available sheets in CBSL file: {xls_cbsl.sheet_names}")
            
            # Find the appropriate sheet
            candidate_names = [
                "Portfolio", "CBSL Portfolio", "Portfolio Sheet", "Working", "CBSL Working", 
                "CBSL", "Sheet1", "Sheet", "C1 & C2 Working"
            ]
            
            chosen_sheet = None
            
            # First, try to find Portfolio sheet specifically
            if "Portfolio" in xls_cbsl.sheet_names:
                chosen_sheet = "Portfolio"
                print(f"[OK] Found Portfolio sheet in CBSL file")
            else:
                # Look for Portfolio-like sheet names
                for sheet_name in xls_cbsl.sheet_names:
                    if "portfolio" in sheet_name.lower():
                        chosen_sheet = sheet_name
                        print(f"[OK] Found Portfolio-like sheet: {chosen_sheet}")
                        break
            
            # If no Portfolio sheet found, use candidate names
            if chosen_sheet is None:
                for name in candidate_names:
                    if name in xls_cbsl.sheet_names:
                        chosen_sheet = name
                        print(f"[WARN] No Portfolio sheet found, using: {chosen_sheet}")
                        break
            
            # Use first sheet as fallback
            if chosen_sheet is None:
                chosen_sheet = xls_cbsl.sheet_names[0]
                print(f"[WARN] Using first sheet as fallback: {chosen_sheet}")
            
            print(f"Reading CBSL sheet: {chosen_sheet}")
            
            # Read the sheet with row 3 as column names, starting from column C
            try:
                # Read raw data first to understand structure
                df_temp = pd.read_excel(xls_cbsl, sheet_name=chosen_sheet, engine="pyxlsb", header=None)
                print(f"Raw CBSL sheet dimensions: {df_temp.shape}")
                
                # Display first few rows to understand structure
                print("First 6 rows of CBSL sheet:")
                for i in range(min(6, len(df_temp))):
                    print(f"Row {i+1} (Excel row {i+1}): {df_temp.iloc[i, :10].tolist()}")  # Show first 10 columns
                
                # Use row 3 (index 2) as headers, starting from column C (index 2) - Excel row 3
                if len(df_temp) > 2:
                    headers = df_temp.iloc[2, :].values  # Row 3 (0-indexed as 2), from column C onwards
                    data = df_temp.iloc[3:, :]  # Data from row 4 onwards, column C onwards
                    
                    # Clean up headers - use row 3 column names directly
                    clean_headers = []
                    for h in headers:
                        if pd.notna(h) and str(h).strip():
                            clean_headers.append(str(h).strip())
                        else:
                            clean_headers.append(f"Col_{len(clean_headers)}")
                    
                    # Ensure we don't have more columns than headers
                    n_headers = len(clean_headers)
                    if len(data.columns) > n_headers:
                        data = data.iloc[:, :n_headers]
                    elif len(data.columns) < n_headers:
                        clean_headers = clean_headers[:len(data.columns)]
                    
                    data.columns = clean_headers
                    df_cbsl = data.reset_index(drop=True)
                    
                    print(f"[OK] Successfully extracted headers from Excel row 3: {clean_headers[:15]}")
                    print(f"CBSL data shape after processing: {df_cbsl.shape}")
                else:
                    raise Exception("Sheet has insufficient rows for row 3 header extraction")
                
            except Exception as e:
                print(f"[ERROR] Failed to read CBSL with row 3 headers: {e}")
                # Fallback: try simple read
                try:
                    df_cbsl = pd.read_excel(xls_cbsl, sheet_name=chosen_sheet, engine="pyxlsb")
                    df_cbsl = normalize_columns(df_cbsl)
                    print(f"[WARN] Using fallback method - simple read")
                except Exception as e2:
                    print(f"[ERROR] Fallback also failed: {e2}")
                    return
            
            print(f"[OK] Using CBSL sheet: {chosen_sheet}")
            print(f"CBSL sheet shape: {df_cbsl.shape}")
            
            # Normalize columns after reading
            df_cbsl = normalize_columns(df_cbsl)
            
            print(f"\n=== CBSL Column Analysis & Logging ===")
            print(f"Total columns: {len(df_cbsl.columns)}")
            print("All CBSL column names after normalization:")
            for i, col in enumerate(df_cbsl.columns):
                print(f"  {i+1:2d}. '{col}'")
                if "cbsl" in str(col).lower() or "p_np" in str(col).lower():
                    print(f"      *** POTENTIAL P/NP COLUMN ***")
                if "contract" in str(col).lower():
                    print(f"      *** POTENTIAL CONTRACT COLUMN ***")
            print("=" * 50)
        
        except Exception as e:
            print(f"[ERROR] Failed to open CBSL xlsb: {e}")
            return
        
        # Map required columns
        print("\n=== Column Mapping ===")
        print(f"Available columns in Prod sheet: {list(df_prod.columns)}")
    
        col_contract = pick_column(df_prod, ["contract_no", "contractno", "contract_number"]) or "contract_no"
        col_client = pick_column(df_prod, ["client_code", "clientcode", "client_number"]) or "client_code"
        col_product = pick_column(df_prod, ["product"]) or "product"
        col_gross_dp = pick_column(df_prod, ["gross_outstanding_with_dp", "gross_with_dp", "gross_outstanding"]) or "gross_outstanding_with_dp"
        col_ifrs_dp = pick_column(df_prod, ["irfs_provision_imp_prov_dp", "ifrs_provision_imp_prov_dp", "ifrs_provision_with_dp", "ifrs_provision"]) or "irfs_provision_imp_prov_dp"
        col_iis = pick_column(df_prod, ["iis"]) or "iis"
        col_equipment = pick_column(df_prod, ["equipment"]) or "equipment"
        
        print(f"\n=== Production Sheet Column Mapping Results ===")
        print(f"Mapped columns:")
        print(f"  Contract: {col_contract}")
        print(f"  Client: {col_client}")
        print(f"  Product: {col_product}")
        print(f"  Gross DP: {col_gross_dp}")
        print(f"  IFRS DP: {col_ifrs_dp}")
        print(f"  IIS: {col_iis}")
        print(f"  Equipment: {col_equipment}")
        
        # Log column mapping success/failure
        mapping_status = {
            "Contract": "[OK] FOUND" if col_contract in df_prod.columns else "[ERROR] NOT FOUND",
            "Client": "[OK] FOUND" if col_client in df_prod.columns else "[ERROR] NOT FOUND", 
            "Product": "[OK] FOUND" if col_product in df_prod.columns else "[ERROR] NOT FOUND",
            "Gross DP": "[OK] FOUND" if col_gross_dp in df_prod.columns else "[ERROR] NOT FOUND",
            "IFRS DP": "[OK] FOUND" if col_ifrs_dp in df_prod.columns else "[ERROR] NOT FOUND",
            "IIS": "[OK] FOUND" if col_iis in df_prod.columns else "[ERROR] NOT FOUND",
            "Equipment": "[OK] FOUND" if col_equipment in df_prod.columns else "[ERROR] NOT FOUND"
        }
        
        print(f"\nColumn Mapping Status:")
        for col_name, status in mapping_status.items():
            print(f"  {col_name:10s}: {status}")
        print("=" * 50)
        
        missing = [c for c in [col_contract, col_client, col_product, col_gross_dp, col_ifrs_dp, col_iis, col_equipment] 
                   if c not in df_prod.columns]
        if missing:
            print(f"[ERROR] Missing expected columns in Prod sheet: {missing}")
            return
        
        # Validate source data quality BEFORE processing
        print("\n=== Source Data Validation ===")
        for col in [col_gross_dp, col_ifrs_dp, col_iis]:
            validate_data_quality(df_prod, col)
        
        # Create portfolio dataframe with proper data conversion
        print("\n=== Creating Portfolio DataFrame ===")
        df_port = pd.DataFrame({
            "contract_no": df_prod[col_contract],
            "client_code": df_prod[col_client],
            "product": df_prod[col_product],
            "equipment": df_prod[col_equipment],
        })
        
        # Convert numeric columns with validation
        print("Converting numeric columns...")
        df_port["gross_with_dp"] = df_prod[col_gross_dp].apply(safe_number)
        df_port["ifrs_provision_with_dp"] = df_prod[col_ifrs_dp].apply(safe_number)
        df_port["iis"] = df_prod[col_iis].apply(safe_number)
        
        # Validate converted data
        for col in ["gross_with_dp", "ifrs_provision_with_dp", "iis"]:
            validate_data_quality(df_port, col)
        
        # Calculate Mid from Client Code first digit
        def compute_mid(code):
            s = str(code).strip()
            return 1 if s.startswith("1") else (2 if s.startswith("2") else None)
        
        df_port["mid"] = df_port["client_code"].map(compute_mid)
        
        # Calculate Gross-IIS-IMP with detailed logging
        print("\n=== Calculating Gross-IIS-IMP ===")
        def calculate_gross_iis_imp(row):
            try:
                gross = row["gross_with_dp"]
                iis = row["iis"]
                ifrs = row["ifrs_provision_with_dp"]
                
                # Validate inputs
                if pd.isna(gross) or pd.isna(iis) or pd.isna(ifrs):
                    if row.name < 5:  # Debug first few rows
                        print(f"Row {row.name}: NaN detected - Gross: {gross}, IIS: {iis}, IFRS: {ifrs}")
                    return 0.0
                
                result = gross - iis - ifrs
                
                # Debug first few rows
                if row.name < 5:
                    print(f"Row {row.name}: {gross} - {iis} - {ifrs} = {result}")
                
                return result
                
            except Exception as e:
                print(f"Error calculating Gross-IIS-IMP for row {row.name}: {e}")
                print(f"  Values - Gross: {row['gross_with_dp']}, IIS: {row['iis']}, IFRS: {row['ifrs_provision_with_dp']}")
                return 0.0
        
        df_port["gross_iis_imp"] = df_port.apply(calculate_gross_iis_imp, axis=1)
        
        # Add missing columns that will be needed later
        df_port["p_np_based_on_cbsl_provision"] = None
        df_port["equipment"] = df_port["equipment"].fillna("")
        df_port["corporate_individual"] = ""
        df_port["mortgage"] = "#N/A"
        df_port["final_category_for_risk_weight"] = ""
        df_port["a"] = 0.0
        df_port["b"] = 0.0
        
        # Extract CBSL Provision values from SOFP file and add as new contracts
        if file_sofp:
            cbsl_provision_values = extract_cbsl_provision_values(file_sofp)
            if cbsl_provision_values:
                df_port = add_cbsl_provision_contracts(df_port, cbsl_provision_values, month, year)
            else:
                print("[WARN] No CBSL Provision values extracted from SOFP file")
        else:
            print("[WARN] SOFP file not found - skipping CBSL Provision contract addition")
        
        # Load risk weight categories from Types sheet
        categories = load_risk_weight_categories(file_car)
        
        # Round to 2 decimal places
        df_port["gross_iis_imp"] = df_port["gross_iis_imp"].round(2)
        
        # Final validation of calculated column
        validate_data_quality(df_port, "gross_iis_imp")
        
        # IMP%: Corrected calculation - IMP = (IIS + IFRS Provision with DP) / Gross with DP
        def compute_imp_pct(row):
            gross = row["gross_with_dp"]
            iis = row["iis"]
            ifrs = row["ifrs_provision_with_dp"]
            
            # Handle edge cases
            if gross == 0:
                return 0.0
            
            # Calculate as (IIS + IFRS) / Gross
            imp_value = (iis + ifrs) / gross
            return imp_value
        
        df_port["imp_percent"] = df_port.apply(compute_imp_pct, axis=1)
        # Round IMP% to 4 decimal places for percentage precision
        df_port["imp_percent"] = df_port["imp_percent"].round(4)
        df_port["margin_20_percent"] = df_port["imp_percent"].map(lambda x: "Above 20%" if x >= 0.2 else "Below 20%")
    
        # Corporate/Individual classification based on Client Code first digit
        def classify_corporate_individual(code):
            s = str(code).strip()
            if s.startswith("1"):
                return "Individual"
            if s.startswith("2"):
                return "Corporate"
            return ""
        
        df_port["corporate_individual"] = df_port["client_code"].map(classify_corporate_individual)
        
        # Enhanced CBSL P/NP lookup
        print(f"\n=== CBSL P/NP Column Detection ===")
        print(f"CBSL sheet columns: {list(df_cbsl.columns)}")
        
        # Look for contract column with enhanced search
        cbsl_contract = None
        contract_search_terms = [
            "contract_no", "contractno", "contract_number", "facility_no", "facility_number", 
            "agreement_no", "agreement_number"
        ]
        
        # Exclude any YARD-related columns from consideration
        cbsl_cols_filtered = [c for c in df_cbsl.columns if "yard" not in str(c).lower()]
        
        # Prefer exact 'contract_no' if present
        if "contract_no" in cbsl_cols_filtered:
            cbsl_contract = "contract_no"
            print(f"[OK] Using specific CBSL contract column: '{cbsl_contract}' (ignoring YARD columns)")
        
        # First try exact matches among filtered columns
        if not cbsl_contract:
            for term in contract_search_terms:
                if term in cbsl_cols_filtered:
                    cbsl_contract = term
                    print(f"[OK] Found EXACT contract column match: '{cbsl_contract}' (ignoring YARD columns)")
                    break
        
        # If no exact match, try partial matches
        if not cbsl_contract:
            print("No exact contract column match found, searching partial matches (ignoring YARD columns)...")
            for col in cbsl_cols_filtered:
                if any(pattern in col for pattern in ['contract', 'facility', 'agreement']):
                    cbsl_contract = col
                    print(f"[OK] Found partial contract column match: '{cbsl_contract}'")
                    break
        
        # Look for CBSL P/NP column with enhanced search
        cbsl_pnp = None
        pnp_search_terms = [
            "cbsl_p_np", "cbsl_pnp", "p_np", "pnp", "classification", "status", "category"
        ]
        
        # First try exact matches for P/NP
        print("Searching for P/NP column...")
        for term in pnp_search_terms:
            if term in df_cbsl.columns:
                cbsl_pnp = term
                print(f"[OK] Found EXACT P/NP column match: '{cbsl_pnp}'")
                break
        
        # If no exact match, try partial matches
        if not cbsl_pnp:
            print("No exact P/NP column match found, searching partial matches...")
            for col in df_cbsl.columns:
                if any(pattern in col for pattern in ['p_np', 'pnp', 'cbsl', 'classification', 'status']):
                    cbsl_pnp = col
                    print(f"[OK] Found partial P/NP column match: '{cbsl_pnp}'")
                    break
        
        print(f"\n=== CBSL Column Detection Results ===")
        print(f"Final column selections:")
        print(f"CBSL Contract column: {cbsl_contract}")
        print(f"CBSL P/NP column: {cbsl_pnp}")
        
        # Log CBSL column detection success/failure
        cbsl_mapping_status = {
            "Contract": "[OK] FOUND" if cbsl_contract and cbsl_contract in df_cbsl.columns else "[ERROR] NOT FOUND",
            "P/NP": "[OK] FOUND" if cbsl_pnp and cbsl_pnp in df_cbsl.columns else "[ERROR] NOT FOUND"
        }
        
        print(f"\nCBSL Column Detection Status:")
        for col_name, status in cbsl_mapping_status.items():
            print(f"  {col_name:10s}: {status}")
        print("=" * 50)
        
        # Show all available columns if no match found
        if not cbsl_contract or not cbsl_pnp:
            print(f"\n[WARN] Missing required columns. Available CBSL columns:")
            for i, col in enumerate(df_cbsl.columns, 1):
                print(f"  {i:2d}. '{col}'")
                # Check if this might be a contract or P/NP column
                col_lower = str(col).lower()
                if any(term in col_lower for term in ['contract', 'facility', 'agreement']):
                    print(f"       ^ Potential CONTRACT column")
                if any(term in col_lower for term in ['p/np', 'pnp', 'cbsl', 'classification', 'status']):
                    print(f"       ^ Potential P/NP column")
        
        # Proceed with enhanced lookup if both columns are found
        if cbsl_contract and cbsl_pnp and cbsl_contract in df_cbsl.columns and cbsl_pnp in df_cbsl.columns:
            # Use enhanced P/NP lookup function
            df_port, df_matching_report, comparison_results = enhanced_pnp_lookup(
                df_port, df_cbsl, cbsl_contract, cbsl_pnp
            )
            
            # Save detailed matching report
            matching_report_path = out_folder / f"P_NP_Matching_Report_{month}_{year}.xlsx"
            print(f"\n=== SAVING DETAILED MATCHING REPORT ===")
            
            with pd.ExcelWriter(matching_report_path, engine='openpyxl') as writer:
                # Main matching report
                df_matching_report.to_excel(writer, sheet_name='Matching_Report', index=False)
                
                # Summary statistics
                summary_data = []
                match_type_counts = df_matching_report["match_type"].value_counts()
                total_records = len(df_matching_report)
                
                for match_type, count in match_type_counts.items():
                    percentage = (count / total_records) * 100
                    summary_data.append({
                        'Match Type': match_type,
                        'Count': count,
                        'Percentage': f"{percentage:.1f}%"
                    })
                
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='Summary', index=False)
                
                # P/NP value distribution
                pnp_counts = df_matching_report["p_np_value"].value_counts(dropna=False)
                pnp_distribution_data = []
                for value, count in pnp_counts.items():
                    percentage = (count / total_records) * 100
                    pnp_distribution_data.append({
                        'P/NP Value': str(value),
                        'Count': count,
                        'Percentage': f"{percentage:.1f}%"
                    })
                
                df_pnp_dist = pd.DataFrame(pnp_distribution_data)
                df_pnp_dist.to_excel(writer, sheet_name='PNP_Distribution', index=False)
                
                # Unmatched records for further investigation
                df_unmatched = df_matching_report[df_matching_report["match_type"] == "No Match"]
                if len(df_unmatched) > 0:
                    df_unmatched.to_excel(writer, sheet_name='Unmatched_Records', index=False)
                
                # Auto-fit columns for all sheets
                workbook = writer.book
                
                # Auto-fit all sheets
                for sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        # Check header length
                        if column[0].row == 1:
                            header_length = len(str(column[0].value)) if column[0].value else 0
                            max_length = max(max_length, header_length)
                        
                        # Check data length (sample first 500 rows for performance)
                        sample_size = min(500, ws.max_row)
                        for i, cell in enumerate(column):
                            if i >= sample_size:
                                break
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
            
            print(f"[OK] Detailed matching report saved to: {matching_report_path}")
            
        else:
            print("[ERROR] Could not find required CBSL columns for lookup. Setting P/NP to None for all records.")
            df_port["p_np_based_on_cbsl_provision"] = None
            
            # Create a simple report showing the issue
            issue_report = pd.DataFrame({
                'Issue': ['Missing CBSL Contract Column', 'Missing CBSL P/NP Column'],
                'Found': [cbsl_contract is not None, cbsl_pnp is not None],
                'Column_Name': [cbsl_contract or 'NOT FOUND', cbsl_pnp or 'NOT FOUND']
            })
            
            issue_report_path = out_folder / f"CBSL_Column_Issues_{month}_{year}.xlsx"
            issue_report.to_excel(issue_report_path, index=False)
            print(f"[WARN] Issue report saved to: {issue_report_path}")
        
        # Populate Mortgage from CBSL PropertyMortgage sheet using enhanced VLOOKUP-style logic
        df_port = populate_mortgage_from_cbsl(df_port, file_cbsl)
        
        # Categorize contracts based on risk weight criteria (after P/NP lookup and mortgage lookup)
        df_port = categorize_risk_weight(df_port, categories)
        
        # Also populate Mortgage in CAR Working file if it exists
        if len(df_car) > 0:
            print(f"\n=== POPULATING MORTGAGE IN CAR WORKING FILE ===")
            print(f"CAR Working file has {len(df_car)} records")
            
            # Check if CAR Working already has a mortgage column
            car_mortgage_col = pick_column(df_car, ["mortgage", "propertymortgage", "property_mortgage"])
            if car_mortgage_col:
                print(f"CAR Working already has mortgage column: '{car_mortgage_col}'")
                print("Updating existing mortgage column with CBSL PropertyMortgage data...")
            else:
                print("CAR Working does not have mortgage column, will create new one...")
            
            df_car = populate_mortgage_from_cbsl(df_car, file_cbsl)
            
            # Log final results for CAR Working
            car_total_records = len(df_car)
            car_mortgage_found = (df_car["mortgage"] != "#N/A").sum()
            car_mortgage_not_found = (df_car["mortgage"] == "#N/A").sum()
            
            print(f"\n=== CAR WORKING MORTGAGE LOOKUP RESULTS ===")
            print(f"Total CAR Working records: {car_total_records}")
            print(f"Mortgage values found: {car_mortgage_found} ({(car_mortgage_found/car_total_records)*100:.1f}%)")
            print(f"Mortgage values NOT found: {car_mortgage_not_found} ({(car_mortgage_not_found/car_total_records)*100:.1f}%)")
            print("[OK] Mortgage column populated in CAR Working file")
        
        # Final data summary
        print("\n=== FINAL DATA SUMMARY ===")
        print(f"Total portfolio records: {len(df_port)}")
        print("Column summary:")
        for col in df_port.columns:
            if df_port[col].dtype in ['int64', 'float64']:
                non_zero = (df_port[col] != 0).sum()
                print(f"  {col}: min={df_port[col].min():.2f}, max={df_port[col].max():.2f}, mean={df_port[col].mean():.2f}, non-zero={non_zero}")
            else:
                non_null = df_port[col].notna().sum()
                unique_vals = df_port[col].nunique()
                print(f"  {col}: non-null={non_null}, unique_values={unique_vals}")
        
        # Show final P/NP statistics
        if "p_np_based_on_cbsl_provision" in df_port.columns:
            pnp_final_stats = df_port["p_np_based_on_cbsl_provision"].value_counts(dropna=False)
            print(f"\nFinal P/NP Distribution:")
            total_records = len(df_port)
            for value, count in pnp_final_stats.items():
                percentage = (count / total_records) * 100
                print(f"  {str(value):15s}: {count:5d} records ({percentage:5.1f}%)")
        
        # Show final Mortgage statistics
        if "mortgage" in df_port.columns:
            mortgage_final_stats = df_port["mortgage"].value_counts(dropna=False)
            print(f"\nFinal Mortgage Distribution:")
            total_records = len(df_port)
            # Show top 10 values and count of #N/A
            na_count = (df_port["mortgage"] == "#N/A").sum()
            non_na_count = total_records - na_count
            print(f"  {'#N/A (No Match)':30s}: {na_count:5d} records ({(na_count/total_records)*100:5.1f}%)")
            print(f"  {'Found Values':30s}: {non_na_count:5d} records ({(non_na_count/total_records)*100:5.1f}%)")
            
            # Show sample of actual mortgage values found
            mortgage_values = df_port[df_port["mortgage"] != "#N/A"]["mortgage"]
            if len(mortgage_values) > 0:
                print(f"  Sample mortgage values found: {mortgage_values.head().tolist()}")
        
        # Reorder columns to match CAR Excel format
        car_column_order = [
            "contract_no",
            "client_code", 
            "mid",
            "product",
            "gross_with_dp",
            "ifrs_provision_with_dp",
            "iis",
            "gross_iis_imp",
            "imp_percent",
            "margin_20_percent",
            "p_np_based_on_cbsl_provision",
            "equipment",
            "corporate_individual",
            "mortgage",
            "final_category_for_risk_weight",
            "final_category_for_risk_weight_code",
            "a",
            "b"
        ]
        
        # Add missing columns with default values
        for col in car_column_order:
            if col not in df_port.columns:
                if col in ["corporate_individual", "final_category_for_risk_weight", "final_category_for_risk_weight_code"]:
                    df_port[col] = ""  # Empty string for text columns
                elif col in ["a", "b"]:
                    df_port[col] = 0.0  # Zero for numeric columns
                else:
                    df_port[col] = None  # None for other missing columns
        
        # Reorder the DataFrame columns
        df_port = df_port[car_column_order]
        
        print(f"[OK] Reordered columns to match CAR Excel format: {list(df_port.columns)}")
        
        # Update Portfolio sheet with pivot data (preserving existing formatting)
        if wb_c1_c6 is not None:
            print("\n=== UPDATING EXISTING PORTFOLIO SHEET (PRESERVING FORMATTING) ===")
            out_path = update_portfolio_with_pivot_data(wb_c1_c6, df_port, out_folder, month, year)
            if out_path:
                print(f"[OK] Portfolio sheet updated successfully: {out_path}")
            else:
                print("[WARN] Portfolio sheet update failed, falling back to new file creation")
                # Fallback: create new file if update fails
                out_path = out_folder / f"NBD_MF_20_C3_{month}_{year}_report.xlsx"
                with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
                    df_port.to_excel(writer, sheet_name="Portfolio", index=False)
                print(f"[OK] New report file created: {out_path}")
        else:
            # Use pandas ExcelWriter approach for new file creation
            print("Using pandas ExcelWriter approach for new file...")
            out_path = out_folder / f"NBD_MF_20_C3_{month}_{year}_report.xlsx"
            
            with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
                df_port.to_excel(writer, sheet_name="Portfolio", index=False)
                
                # Auto-fit columns
                workbook = writer.book
                ws = workbook["Portfolio"]
                print("Auto-fitting column widths...")
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    # Check header length
                    if column[0].row == 1:  # Header row
                        header_length = len(str(column[0].value)) if column[0].value else 0
                        max_length = max(max_length, header_length)
                    
                    # Check data length (sample first 1000 rows for performance)
                    sample_size = min(1000, len(df_port))
                    for i, cell in enumerate(column):
                        if i >= sample_size + 1:  # +1 for header row
                            break
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    # Set column width (with some padding, cap at 50 characters)
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            print(f"\n[OK] New report file created successfully: {out_path}")
        
        # Create detailed mortgage matching report
        if "mortgage" in df_port.columns:
            print(f"\n=== SAVING DETAILED MORTGAGE MATCHING REPORT ===")
            mortgage_report_path = out_folder / f"Mortgage_Matching_Report_{month}_{year}.xlsx"
            
            # Prepare mortgage report data
            df_mortgage_report = df_port[["contract_no", "client_code", "product", "gross_with_dp", "mortgage"]].copy()
            df_mortgage_report["match_status"] = df_mortgage_report["mortgage"].apply(
                lambda x: "No Match" if x == "#N/A" else "Match Found"
            )
            
            with pd.ExcelWriter(mortgage_report_path, engine='openpyxl') as writer:
                # Main mortgage report
                df_mortgage_report.to_excel(writer, sheet_name='Mortgage_Report', index=False)
                
                # Summary statistics
                match_summary = df_mortgage_report["match_status"].value_counts()
                summary_data = []
                total_records = len(df_mortgage_report)
                
                for status, count in match_summary.items():
                    percentage = (count / total_records) * 100
                    summary_data.append({
                        'Match Status': status,
                        'Count': count,
                        'Percentage': f"{percentage:.1f}%"
                    })
                
                df_mortgage_summary = pd.DataFrame(summary_data)
                df_mortgage_summary.to_excel(writer, sheet_name='Summary', index=False)
                
                # Auto-fit columns for both sheets
                workbook = writer.book
                
                # Auto-fit Mortgage_Report sheet
                mortgage_ws = workbook['Mortgage_Report']
                for column in mortgage_ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    # Check header length
                    if column[0].row == 1:
                        header_length = len(str(column[0].value)) if column[0].value else 0
                        max_length = max(max_length, header_length)
                    
                    # Check data length (sample first 500 rows for performance)
                    sample_size = min(500, len(df_mortgage_report))
                    for i, cell in enumerate(column):
                        if i >= sample_size + 1:
                            break
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    mortgage_ws.column_dimensions[column_letter].width = adjusted_width
                
                # Auto-fit Summary sheet
                summary_ws = workbook['Summary']
                for column in summary_ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    summary_ws.column_dimensions[column_letter].width = adjusted_width
                
                # Unmatched records for investigation
                df_mortgage_unmatched = df_mortgage_report[df_mortgage_report["match_status"] == "No Match"]
                if len(df_mortgage_unmatched) > 0:
                    df_mortgage_unmatched.to_excel(writer, sheet_name='Unmatched_Records', index=False)
                
                # Matched records sample
                df_mortgage_matched = df_mortgage_report[df_mortgage_report["match_status"] == "Match Found"]
                if len(df_mortgage_matched) > 0:
                    # Sample of first 100 matched records
                    df_mortgage_matched.head(100).to_excel(writer, sheet_name='Matched_Sample', index=False)
            
            print(f"[OK] Detailed mortgage matching report saved to: {mortgage_report_path}")
        
        # Compare with existing CAR Working file
        print(f"\n=== COMPARING WITH EXISTING CAR WORKING FILE ===")
        if len(df_car) > 0:
            print("Comparing generated portfolio with existing CAR Working Portfolio sheet...")
            
            # Check if CAR Working has P/NP column
            car_pnp_col = pick_column(df_car, ["p_np_based_on_cbsl", "p_np_based_on_cbsl_provision", "cbsl_p_np", "p_np"])
    
            # Fallback: partial-match search if not found
            if not car_pnp_col:
                for col in df_car.columns:
                    if any(pattern in col for pattern in ["p_np", "pnp", "cbsl", "classification", "status"]):
                        car_pnp_col = col
                        print(f"[OK] Fallback found CAR P/NP-like column: '{car_pnp_col}'")
                        break
    
            if car_pnp_col:
                print(f"Found P/NP column in CAR Working: '{car_pnp_col}'")
                
                # Compare P/NP values
                car_contract_col = pick_column(df_car, ["contract_no", "contractno", "contract_number"])
                
                if car_contract_col:
                    df_car_comparison = df_car[[car_contract_col, car_pnp_col]].copy()
                    df_car_comparison.columns = ["contract_no", "car_p_np"]
                    df_car_comparison["contract_no"] = df_car_comparison["contract_no"].astype(str).str.strip()
                    
                    # Merge with our results
                    df_comparison = df_port[["contract_no", "p_np_based_on_cbsl_provision"]].merge(
                        df_car_comparison, on="contract_no", how="outer", suffixes=("_New", "_CAR")
                    )
                    
                    # Identify differences
                    df_comparison["Match"] = df_comparison["p_np_based_on_cbsl_provision"] == df_comparison["car_p_np"]
                    
                    matches = df_comparison["Match"].sum()
                    total_comparable = len(df_comparison.dropna(subset=["p_np_based_on_cbsl_provision", "car_p_np"]))
                    
                    print(f"P/NP Comparison with existing CAR Working:")
                    print(f"  Matching records: {matches}")
                    print(f"  Total comparable: {total_comparable}")
                    if total_comparable > 0:
                        print(f"  Match percentage: {(matches/total_comparable)*100:.1f}%")
                    
                    # Save comparison report
                    comparison_path = out_folder / f"CAR_vs_New_PNP_Comparison_{month}_{year}.xlsx"
                    
                    # Save with auto-fit columns
                    with pd.ExcelWriter(comparison_path, engine='openpyxl') as writer:
                        df_comparison.to_excel(writer, sheet_name='Comparison', index=False)
                        
                        # Auto-fit columns
                        workbook = writer.book
                        ws = workbook['Comparison']
                        for column in ws.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            
                            # Check header length
                            if column[0].row == 1:
                                header_length = len(str(column[0].value)) if column[0].value else 0
                                max_length = max(max_length, header_length)
                            
                            # Check data length (sample first 500 rows for performance)
                            sample_size = min(500, ws.max_row)
                            for i, cell in enumerate(column):
                                if i >= sample_size:
                                    break
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            
                            adjusted_width = min(max_length + 2, 50)
                            ws.column_dimensions[column_letter].width = adjusted_width
                    
                    print(f"[OK] CAR comparison report saved to: {comparison_path}")
                else:
                    print("[ERROR] Could not find Contract No column in CAR Working file")
            else:
                print("[ERROR] Could not find P/NP column in CAR Working file")
            
            # Check if CAR Working has Mortgage column and compare
            car_mortgage_col = pick_column(df_car, ["mortgage", "propertymortgage", "property_mortgage"])
            
            if car_mortgage_col:
                print(f"\nFound Mortgage column in CAR Working: '{car_mortgage_col}'")
                
                car_contract_col = pick_column(df_car, ["contract_no", "contractno", "contract_number"])
                
                if car_contract_col:
                    df_car_mortgage = df_car[[car_contract_col, car_mortgage_col]].copy()
                    df_car_mortgage.columns = ["contract_no", "car_mortgage"]
                    df_car_mortgage["contract_no"] = df_car_mortgage["contract_no"].astype(str).str.strip()
                    
                    # Merge with our results
                    df_mortgage_comparison = df_port[["contract_no", "mortgage"]].merge(
                        df_car_mortgage, on="contract_no", how="outer", suffixes=("_New", "_CAR")
                    )
                    
                    # Identify differences (considering #N/A as no match)
                    df_mortgage_comparison["Match"] = (
                        (df_mortgage_comparison["mortgage"] == df_mortgage_comparison["car_mortgage"]) |
                        ((df_mortgage_comparison["mortgage"] == "#N/A") & (pd.isna(df_mortgage_comparison["car_mortgage"])))
                    )
                    
                    matches = df_mortgage_comparison["Match"].sum()
                    total_comparable = len(df_mortgage_comparison.dropna(subset=["contract_no"]))
                    
                    print(f"Mortgage Comparison with existing CAR Working:")
                    print(f"  Matching records: {matches}")
                    print(f"  Total comparable: {total_comparable}")
                    if total_comparable > 0:
                        print(f"  Match percentage: {(matches/total_comparable)*100:.1f}%")
                    
                    # Save mortgage comparison report
                    mortgage_comparison_path = out_folder / f"CAR_vs_New_Mortgage_Comparison_{month}_{year}.xlsx"
                    
                    # Save with auto-fit columns
                    with pd.ExcelWriter(mortgage_comparison_path, engine='openpyxl') as writer:
                        df_mortgage_comparison.to_excel(writer, sheet_name='Mortgage_Comparison', index=False)
                        
                        # Auto-fit columns
                        workbook = writer.book
                        ws = workbook['Mortgage_Comparison']
                        for column in ws.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            
                            # Check header length
                            if column[0].row == 1:
                                header_length = len(str(column[0].value)) if column[0].value else 0
                                max_length = max(max_length, header_length)
                            
                            # Check data length (sample first 500 rows for performance)
                            sample_size = min(500, ws.max_row)
                            for i, cell in enumerate(column):
                                if i >= sample_size:
                                    break
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            
                            adjusted_width = min(max_length + 2, 50)
                            ws.column_dimensions[column_letter].width = adjusted_width
                    
                    print(f"[OK] CAR mortgage comparison report saved to: {mortgage_comparison_path}")
                else:
                    print("[ERROR] Could not find Contract No column in CAR Working file for mortgage comparison")
            else:
                print("[ERROR] Could not find Mortgage column in CAR Working file")
                
        else:
            print("[WARN] No CAR Working data available for comparison")
        
        # Update CAR Working file
        excel = None
        wb_car = None
        try:
            print(f"\n=== UPDATING CAR WORKING FILE ===")
            print("Updating CAR Working .xlsb Portfolio sheet via Excel automation...")
            
            # Initialize Excel with better error handling
            try:
                excel = win32.Dispatch("Excel.Application")
                if excel is None:
                    print("[ERROR] Failed to create Excel application")
                    return
            except Exception as excel_error:
                print(f"[ERROR] Error creating Excel application: {excel_error}")
                return
                
            # Set Excel properties with error handling
            try:
                excel.Visible = False
            except Exception as visible_error:
                print(f"[WARN] Could not set Excel Visible property: {visible_error}")
                # Continue anyway
                
            try:
                excel.DisplayAlerts = False
            except Exception as alerts_error:
                print(f"[WARN] Could not set Excel DisplayAlerts property: {alerts_error}")
                # Continue anyway
            
            # Open workbook with error handling
            try:
                wb_car = excel.Workbooks.Open(str(file_car))
                if wb_car is None:
                    print("[ERROR] Failed to open CAR Working file")
                    if excel is not None:
                        excel.Quit()
                    return
            except Exception as open_error:
                print(f"[ERROR] Error opening CAR Working file: {open_error}")
                if excel is not None:
                    excel.Quit()
                return
                
            try:
                try:
                    ws_car = wb_car.Worksheets("Portfolio")
                except Exception:
                    ws_car = wb_car.Worksheets.Add()
                    ws_car.Name = "Portfolio"
                
                # Clear existing contents
                try:
                    ws_car.UsedRange.ClearContents()
                except Exception:
                    pass
                
                # Prepare data - CAR Working expects headers to start from row 4
                headers = list(df_port.columns)
                
                # Convert data to proper format for Excel
                df_excel = df_port.copy()
                
                # Convert NaN values to empty strings for Excel compatibility
                df_excel = df_excel.fillna("")
                
                # Convert data to list of lists for Excel
                data_rows = df_excel.values.tolist()
                n_cols = len(headers)
                n_rows = len(data_rows)
                
                print(f"Preparing to write {n_rows} rows and {n_cols} columns to Excel")
                
                # Calculate column letter
                def col_to_letter(n):
                    s = ""
                    while n > 0:
                        n, r = divmod(n - 1, 26)
                        s = chr(65 + r) + s
                    return s
                
                end_col_letter = col_to_letter(n_cols)
                
                # Calculate totals for row 3
                print("Calculating totals for row 3...")
                total_ifrs = df_port["ifrs_provision_with_dp"].sum()
                total_iis = df_port["iis"].sum()
                total_gross_iis_imp = df_port["gross_iis_imp"].sum()
                
                print(f"Totals calculated:")
                print(f"  Total IFRS Provision with DP: {total_ifrs:,.2f}")
                print(f"  Total IIS: {total_iis:,.2f}")
                print(f"  Total Gross-IIS-IMP: {total_gross_iis_imp:,.2f}")
                
                # Extract Net total from SOFP file for H2
                net_total_from_sofp = None
                if file_sofp:
                    cbsl_provision_values = extract_cbsl_provision_values(file_sofp)
                    if cbsl_provision_values and "Net_Total" in cbsl_provision_values:
                        net_total_from_sofp = cbsl_provision_values["Net_Total"]
                        print(f"  Net Total from SOFP (for H2): {net_total_from_sofp:,.2f}")
                    else:
                        print("[WARN] Net Total not found in SOFP file")
                else:
                    print("[WARN] SOFP file not available for Net Total extraction")
                
                # Find column indices for the totals
                ifrs_col_idx = headers.index("ifrs_provision_with_dp") if "ifrs_provision_with_dp" in headers else -1
                iis_col_idx = headers.index("iis") if "iis" in headers else -1
                gross_iis_imp_col_idx = headers.index("gross_iis_imp") if "gross_iis_imp" in headers else -1
                
                # Write totals in row 3
                print("Writing totals in row 3...")
                try:
                    # Write "TOTAL" label in column A, row 3
                    ws_car.Cells(3, 1).Value = "TOTAL"
                    
                    # Write totals in respective columns
                    if ifrs_col_idx >= 0:
                        ifrs_col_letter = col_to_letter(ifrs_col_idx + 1)
                        ws_car.Cells(3, ifrs_col_idx + 1).Value = total_ifrs
                        print(f"[OK] Total IFRS written to {ifrs_col_letter}3: {total_ifrs:,.2f}")
                    
                    if iis_col_idx >= 0:
                        iis_col_letter = col_to_letter(iis_col_idx + 1)
                        ws_car.Cells(3, iis_col_idx + 1).Value = total_iis
                        print(f"[OK] Total IIS written to {iis_col_letter}3: {total_iis:,.2f}")
                    
                    if gross_iis_imp_col_idx >= 0:
                        gross_iis_imp_col_letter = col_to_letter(gross_iis_imp_col_idx + 1)
                        ws_car.Cells(3, gross_iis_imp_col_idx + 1).Value = total_gross_iis_imp
                        print(f"[OK] Total Gross-IIS-IMP written to {gross_iis_imp_col_letter}3: {total_gross_iis_imp:,.2f}")
                    
                    # Write Gross with DP total to E3 (in the gross_with_dp column)
                    gross_dp_col_idx = headers.index("gross_with_dp") if "gross_with_dp" in headers else -1
                    if gross_dp_col_idx >= 0:
                        total_gross_dp = df_port["gross_with_dp"].sum()
                        gross_dp_col_letter = col_to_letter(gross_dp_col_idx + 1)
                        ws_car.Cells(3, gross_dp_col_idx + 1).Value = total_gross_dp
                        print(f"[OK] Total Gross with DP written to {gross_dp_col_letter}3: {total_gross_dp:,.2f}")
                    else:
                        print("[WARN] Could not write Total Gross with DP to E3 - column not found")
                    
                    # Format the totals row
                    if ifrs_col_idx >= 0:
                        ws_car.Cells(3, ifrs_col_idx + 1).NumberFormat = "0.00"
                    if iis_col_idx >= 0:
                        ws_car.Cells(3, iis_col_idx + 1).NumberFormat = "0.00"
                    if gross_iis_imp_col_idx >= 0:
                        ws_car.Cells(3, gross_iis_imp_col_idx + 1).NumberFormat = "0.00"
                    if gross_dp_col_idx >= 0:
                        ws_car.Cells(3, gross_dp_col_idx + 1).NumberFormat = "0.00"
                    
                    # Write Net total to H2 (above the Gross-IIS-IMP total)
                    if net_total_from_sofp is not None and gross_iis_imp_col_idx >= 0:
                        gross_iis_imp_col_letter = col_to_letter(gross_iis_imp_col_idx + 1)
                        ws_car.Cells(2, gross_iis_imp_col_idx + 1).Value = net_total_from_sofp
                        ws_car.Cells(2, gross_iis_imp_col_idx + 1).NumberFormat = "0.00"
                        print(f"[OK] Net Total from SOFP written to {gross_iis_imp_col_letter}2: {net_total_from_sofp:,.2f}")
                    else:
                        print("[WARN] Could not write Net Total to H2 - value not available or column not found")
                    
                        
                except Exception as e:
                    print(f"[ERROR] Error writing totals in row 3: {e}")
                    # Continue with headers even if totals fail
                    pass
                
                # Write headers starting from row 4
                print(f"Writing headers to range A4:{end_col_letter}4")
                try:
                    header_range = ws_car.Range(f"A4:{end_col_letter}4")
                    header_range.Value = headers  # Direct assignment without wrapping in list
                    print("[OK] Headers written successfully")
                except Exception as e:
                    print(f"[ERROR] Error writing headers: {e}")
                    raise
                
                # Write data starting from row 5 (if there's data)
                if data_rows and n_rows > 0:
                    print(f"Writing data to range A5:{end_col_letter}{4 + n_rows}")
                    try:
                        data_range = ws_car.Range(f"A5:{end_col_letter}{4 + n_rows}")
                        data_range.Value = data_rows
                        print("[OK] Data written successfully")
                    except Exception as e:
                        print(f"[ERROR] Error writing data: {e}")
                        raise
                    
                    # Format numeric columns to 2 decimal places
                    numeric_columns = ["gross_with_dp", "ifrs_provision_with_dp", "iis", "gross_iis_imp", "a", "b"]
                    percentage_columns = ["imp_percent"]
                    
                    for i, col_name in enumerate(headers):
                        col_letter = col_to_letter(i + 1)
                        if col_name in numeric_columns:
                            numeric_range = ws_car.Range(f"{col_letter}5:{col_letter}{4 + n_rows}")
                            numeric_range.NumberFormat = "0.00"
                        elif col_name in percentage_columns:
                            pct_range = ws_car.Range(f"{col_letter}5:{col_letter}{4 + n_rows}")
                            pct_range.NumberFormat = "0.0000"  # 4 decimal places for percentage
                
                # Save copy to outputs
                car_out_path = out_folder / file_car.name
                print(f"Saving CAR Working copy to: {car_out_path}")
                wb_car.SaveCopyAs(str(car_out_path))
                print("[OK] CAR Working copy saved with enhanced mortgage lookup data.")
                
                # Update pivot table in CAR Working file
                print("\n=== UPDATING PIVOT TABLE ===")
                update_pivot_table(file_car, df_port)
                
                # Update C1-C6 file with CAR pivot values
                if file_c1_c6 is not None:
                    print("\n=== UPDATING C1-C6 WITH CAR PIVOT VALUES ===")
                    update_c1_c6_with_car_pivot_values(file_c1_c6, file_car, df_port, out_folder)
                else:
                    print("\n[WARN] Skipping C1-C6 update - file_c1_c6 parameter not provided")
                
            finally:
                # Clean up Excel objects properly
                try:
                    if wb_car is not None:
                        print("Closing workbook...")
                        wb_car.Close(SaveChanges=False)
                        wb_car = None
                except Exception as close_error:
                    print(f"[WARN] Error closing workbook: {close_error}")
                    
                try:
                    if excel is not None:
                        print("Quitting Excel application...")
                        excel.Quit()
                        excel = None
                except Exception as quit_error:
                    print(f"[WARN] Error quitting Excel: {quit_error}")
                    
                # Force cleanup
                try:
                    import gc
                    gc.collect()
                except:
                    pass
                
        except Exception as e:
            print(f"[WARN] Could not update/save CAR Working .xlsb: {e}")
            print(f"Error type: {type(e).__name__}")
            
            # Enhanced cleanup in case of error
            try:
                if wb_car is not None:
                    print("Emergency closing workbook...")
                    wb_car.Close(SaveChanges=False)
                    wb_car = None
            except Exception as emergency_close:
                print(f"[WARN] Emergency workbook close failed: {emergency_close}")
                
            try:
                if excel is not None:
                    print("Emergency quitting Excel...")
                    excel.Quit()
                    excel = None
            except Exception as emergency_quit:
                print(f"[WARN] Emergency Excel quit failed: {emergency_quit}")
                
            # Force cleanup
            try:
                import gc
                gc.collect()
            except:
                pass
    
        # Summary of enhancements
        print(f"\n=== ENHANCEMENT SUMMARY ===")
        print("[OK] Enhanced P/NP lookup with normalization and fuzzy matching")
        print("[OK] Enhanced Mortgage lookup from CBSL PropertyMortgage sheet")
        print("[OK] VLOOKUP-style functionality implemented for mortgage data")
        print("[OK] Detailed matching reports generated for both P/NP and Mortgage")
        print("[OK] Comparison reports with existing CAR Working file")
        print("[OK] Updated CAR Working file with enhanced data")
        print("[OK] Comprehensive column logging for all Excel files")
        print("[OK] Column detection and mapping status tracking")
        print(f"[OK] All outputs saved to: {out_folder}")
        
        # Final logging summary
        print(f"\n=== FINAL LOGGING SUMMARY ===")
        print("[PROCESS] Column Analysis Completed for:")
        print("  • Production Sheet (C1 & C2 Working)")
        print("  • CAR Working Sheet (Portfolio)")
        print("  • CBSL Provision Comparison Sheet")
        print("  • CBSL PropertyMortgage Sheet")
        print("[LIST] All column names normalized and logged")
        print("[INFO] Column detection status tracked for each sheet")
        print("[OK] Enhanced error handling and debugging information")
        
        return wb_c1_c6
        
    except Exception as e:
        print(f"[ERROR] Error in C3 report processing: {e}")
        print(f"Error type: {type(e).__name__}")
        import traceback
        traceback.print_exc()
        return wb_c1_c6

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description='NBD-MF-20-C3 Report Automation')
    parser.add_argument('--month', type=str, help='Report month (e.g., Jul)', required=False)
    parser.add_argument('--year', type=str, help='Report year (e.g., 2025)', required=False)
    parser.add_argument('--working-dir', type=str, help='Working directory (for compatibility)', required=False)
    args = parser.parse_args()

    # Determine working directory
    script_dir = Path(__file__).resolve().parent.parent
    working_monthly = script_dir / "working" / "monthly"

    # Find the date folder (should be only one)
    if working_monthly.exists():
        subdirs = [d for d in working_monthly.iterdir() if d.is_dir()]
        if len(subdirs) == 1:
            date_folder = subdirs[0]
            print(f"Found date folder: {date_folder}")
        else:
            print(f"Error: Expected 1 folder in {working_monthly}, found {len(subdirs)}")
            sys.exit(1)
    else:
        print(f"Error: Working directory not found: {working_monthly}")
        sys.exit(1)

    # Look for NBD_MF_20_C1_C6 subfolder
    c1_c6_folder = date_folder / "NBD_MF_20_C1_C6"
    if c1_c6_folder.exists():
        working_dir = c1_c6_folder
        print(f"Found C1-C6 working directory: {working_dir}")
    else:
        working_dir = date_folder
        print(f"NBD_MF_20_C1_C6 folder not found, using date folder: {working_dir}")

    # Find required files in the C1-C6 folder and its Input subfolder
    search_dirs = [working_dir, working_dir / "Input"]

    # Find CAR Working file
    file_car = find_first_matching(search_dirs, "CAR Working*.xlsb")
    if not file_car:
        print("Error: CAR Working file not found")
        sys.exit(1)

    # Find Prod wise file
    file_prod = find_first_matching(search_dirs, "Prod. wise Class. of Loans*.xlsb")
    if not file_prod:
        print("Error: Prod. wise Class. of Loans file not found")
        sys.exit(1)

    # Find CBSL Provision file (try both .xlsx and .xlsb)
    file_cbsl = find_first_matching(search_dirs, "*CBSL Provision*.xlsx")
    if not file_cbsl:
        file_cbsl = find_first_matching(search_dirs, "*CBSL Provision*.xlsb")
    if not file_cbsl:
        print("Error: CBSL Provision Comparison file not found")
        sys.exit(1)

    # Find C1-C6 file (different naming patterns)
    file_c1_c6 = find_first_matching(search_dirs, "NBD-MF-20-C1 to C6*.xlsx")
    if not file_c1_c6:
        file_c1_c6 = find_first_matching(search_dirs, "C1 C2 C3 C4 C5 C6*.xlsx")
    if not file_c1_c6:
        print("Error: C1-C6 file not found")
        sys.exit(1)

    print(f"\n=== NBD-MF-20-C3 Report Automation ===")
    print(f"CAR Working: {file_car}")
    print(f"Prod wise: {file_prod}")
    print(f"CBSL Provision: {file_cbsl}")
    print(f"C1-C6 file: {file_c1_c6}")

    # Kill any open Excel instances to prevent Protected View issues
    print("\nClosing any open Excel instances...")
    try:
        import subprocess
        subprocess.run(
            ['taskkill', '/F', '/IM', 'EXCEL.EXE'],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            timeout=5
        )
        import time
        time.sleep(2)  # Wait for Excel to close
        print("Excel instances closed")
    except Exception as e:
        print(f"Note: Could not close Excel instances: {e}")

    # Load C1-C6 workbook
    print(f"\nLoading C1-C6 workbook...")
    wb_c1_c6 = openpyxl.load_workbook(file_c1_c6)

    # Run main function (output to the C1-C6 working directory)
    out_folder = working_dir
    wb_result = main(
        wb_c1_c6=wb_c1_c6,
        file_car=file_car,
        file_prod=file_prod,
        file_cbsl=file_cbsl,
        out_folder=out_folder,
        file_c1_c6=file_c1_c6
    )

    if wb_result:
        # Save the updated workbook
        output_file = file_c1_c6
        print(f"\nSaving updated workbook to: {output_file}")
        wb_result.save(output_file)
        print("C3 report completed successfully!")

        # Clean up temporary files
        print("\n=== Cleaning up temporary files ===")
        temp_patterns = [
            "CAR_vs_New_Mortgage_Comparison*.xlsx",
            "CAR_vs_New_PNP_Comparison_*.xlsx",
            "Mortgage_Matching_Report_*.xlsx",
            "NBD_MF_20_C3_*.xlsx",
            "P_NP_Matching_Report_*.xlsx"
        ]

        import glob
        removed_count = 0
        for pattern in temp_patterns:
            # Search in working directory
            matches = list(working_dir.glob(pattern))
            for temp_file in matches:
                try:
                    temp_file.unlink()
                    print(f"Removed: {temp_file.name}")
                    removed_count += 1
                except Exception as e:
                    print(f"Warning: Could not remove {temp_file.name}: {e}")

        if removed_count > 0:
            print(f"Cleaned up {removed_count} temporary file(s)")
        else:
            print("No temporary files found to clean up")
    else:
        print("C3 report processing failed")
        sys.exit(1)