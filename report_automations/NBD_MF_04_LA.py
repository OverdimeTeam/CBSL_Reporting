#!/usr/bin/env python3
"""
Monthly Liquidity Report Automation Script
Central Bank Sri Lanka - NBD_MF_04_LA

This script generates the monthly liquidity report by consolidating data from 5 weekly AFL files.

USAGE:
    Called automatically by NBD_MF_15_LA.py when 5 weekly AFL files exist

WORKFLOW:
1. Find all 5 weekly AFL Liquidity files for the current month
2. Extract A3:C25 range from NBD-WF-15-LA sheet from each weekly file
3. Copy data with formatting to respective Week 1-5 sheets (A1:H22 area)
4. For Week 5, use Week 4 data if Week 5 file doesn't exist
5. Fill NBD_WF_04_LA sheet columns C-H with H column values from each Week sheet
6. Save the monthly report with date-based filename

FOLDER STRUCTURE:
- working/weekly/07-06-2025/NBD_MF_04_LA/ - Template location (FIXED)
- working/weekly/{MM-DD-YYYY}/NBD_MF_04_LA/Inputs/ - Contains weekly AFL files (DYNAMIC)
- working/monthly/{MM}/{DD}/{YYYY}/ - Output location for monthly report (month-end date)

COLUMN MAPPING FOR NBD_WF_04_LA SHEET:
- C5:C15 ← Week 1 H3:H13
- C19 ← Week 1 H14
- C22:C27 ← Week 1 H15:H20
- Same pattern for D (Week 2), E (Week 3), F (Week 4), G (Week 5), H (Week 5 duplicate)
"""

import os
import sys
from datetime import datetime
import calendar
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging
from pathlib import Path
import shutil

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('monthly_liquidity_automation.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


class MonthlyLiquidityReport:
    def __init__(self, base_dir=".", report_date=None, afl_files=None):
        """
        Initialize the Monthly Liquidity Report generator

        Args:
            base_dir (str): Base directory path
            report_date (datetime): Report date (to determine month/year)
            afl_files (list): List of tuples (week_number, file_path, date)
        """
        self.base_dir = Path(base_dir)
        self.report_date = report_date
        self.afl_files = afl_files or []

        # Calculate month-end date for the report month
        year = self.report_date.year
        month = self.report_date.month
        last_day = calendar.monthrange(year, month)[1]  # Get last day of month (28-31)
        self.month_end_date = self.report_date.replace(day=last_day)

        # Determine output folder: working/monthly/{MM}/{DD}/{YYYY}
        month_str = f"{self.month_end_date.month:02d}"
        day_str = f"{self.month_end_date.day:02d}"
        year_str = f"{self.month_end_date.year}"
        self.output_folder = self.base_dir / "working" / "monthly" / month_str / day_str / year_str

        # Template location (FIXED folder)
        self.template_folder = self.base_dir / "working" / "weekly" / "07-06-2025" / "NBD_MF_04_LA"

        logger.info(f"Initialized Monthly Liquidity Report for {self.report_date.strftime('%B %Y')}")
        logger.info(f"Month-end date: {self.month_end_date.strftime('%m/%d/%Y')}")
        logger.info(f"Template folder: {self.template_folder}")
        logger.info(f"Output folder: {self.output_folder}")
        logger.info(f"Found {len(self.afl_files)} weekly AFL files")
        for week_num, file_path, file_date in self.afl_files:
            logger.info(f"  Week {week_num}: {file_path.name} ({file_date.strftime('%m/%d/%Y')})")

    def copy_range_with_formatting(self, source_sheet, dest_sheet, source_range, dest_start_cell):
        """
        Copy a range of cells from source to destination, preserving formatting

        Args:
            source_sheet: Source worksheet
            dest_sheet: Destination worksheet
            source_range: String like "A3:C25"
            dest_start_cell: String like "A1"
        """
        try:
            # Parse source range
            from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

            # Get source range cells
            source_cells = source_sheet[source_range]

            # Parse destination start cell
            dest_col_letter, dest_row = coordinate_from_string(dest_start_cell)
            dest_col_idx = column_index_from_string(dest_col_letter)

            # Copy each cell
            for row_idx, row in enumerate(source_cells):
                for col_idx, source_cell in enumerate(row):
                    dest_row_num = dest_row + row_idx
                    dest_col_num = dest_col_idx + col_idx

                    dest_cell = dest_sheet.cell(row=dest_row_num, column=dest_col_num)

                    # Copy value
                    if source_cell.value is not None:
                        dest_cell.value = source_cell.value

                    # Copy formatting
                    if source_cell.has_style:
                        dest_cell.font = source_cell.font.copy()
                        dest_cell.fill = source_cell.fill.copy()
                        dest_cell.border = source_cell.border.copy()
                        dest_cell.alignment = source_cell.alignment.copy()
                        dest_cell.number_format = source_cell.number_format

            logger.info(f"  ✓ Copied range {source_range} with formatting to {dest_start_cell}")
            return True

        except Exception as e:
            logger.error(f"Error copying range with formatting: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False

    def extract_weekly_data(self, week_number, afl_file_path):
        """
        Extract A3:C25 range from NBD-WF-15-LA sheet in weekly AFL file

        Args:
            week_number (int): Week number (1-5)
            afl_file_path (Path): Path to the weekly AFL file

        Returns:
            tuple: (success, worksheet or None)
        """
        wb = None
        try:
            logger.info(f"Extracting data from Week {week_number}: {afl_file_path.name}")

            # Load the weekly AFL file
            wb = openpyxl.load_workbook(afl_file_path, data_only=False)

            # Check if NBD-WF-15-LA sheet exists
            if "NBD-WF-15-LA" not in wb.sheetnames:
                logger.error(f"Sheet 'NBD-WF-15-LA' not found in {afl_file_path.name}")
                logger.info(f"Available sheets: {wb.sheetnames}")
                return False, None

            source_sheet = wb["NBD-WF-15-LA"]
            logger.info(f"  ✓ Found NBD-WF-15-LA sheet in Week {week_number} file")

            return True, (wb, source_sheet)

        except Exception as e:
            logger.error(f"Error extracting data from Week {week_number}: {e}")
            import traceback
            logger.error(traceback.format_exc())
            if wb:
                try:
                    wb.close()
                except:
                    pass
            return False, None

    def find_monthly_template(self):
        """
        Find the NBD-MF-04-LA monthly template file in FIXED folder (07-06-2025)

        Returns:
            Path to template file, or None if not found
        """
        try:
            # Look for template in FIXED folder
            template_files = list(self.template_folder.glob("NBD-MF-04-LA*.xlsx"))

            if not template_files:
                logger.error(f"Monthly template NBD-MF-04-LA not found in {self.template_folder}")
                return None

            template_file = template_files[0]
            logger.info(f"Found monthly template: {template_file.name}")
            return template_file

        except Exception as e:
            logger.error(f"Error finding monthly template: {e}")
            return None

    def fill_nbd_wf_04_la_sheet(self, template_wb):
        """
        Fill NBD_WF_04_LA sheet columns C-H with H column values from Week 1-5 sheets

        Mapping:
        - C5:C15 ← Week 1 H3:H13
        - C19 ← Week 1 H14
        - C22:C27 ← Week 1 H15:H20
        - Same pattern for D (Week 2), E (Week 3), F (Week 4), G (Week 5), H (Week 5 duplicate)

        Args:
            template_wb: The workbook with Week sheets already populated
        """
        try:
            logger.info("\nFilling NBD_WF_04_LA sheet with weekly data...")

            # Check if NBD_WF_04_LA sheet exists
            if "NBD_WF_04_LA" not in template_wb.sheetnames:
                logger.error("Sheet 'NBD_WF_04_LA' not found in template")
                return False

            nbd_sheet = template_wb["NBD_WF_04_LA"]

            # Column mapping: C=Week1, D=Week2, E=Week3, F=Week4, G=Week5, H=Week5
            week_to_column = {
                1: 'C',  # Week 1 → Column C
                2: 'D',  # Week 2 → Column D
                3: 'E',  # Week 3 → Column E
                4: 'F',  # Week 4 → Column F
                5: 'G',  # Week 5 → Column G
            }

            # Process weeks 1-5
            for week_num in range(1, 6):
                week_sheet_name = f"Week {week_num}"

                if week_sheet_name not in template_wb.sheetnames:
                    logger.warning(f"Sheet '{week_sheet_name}' not found, skipping...")
                    continue

                week_sheet = template_wb[week_sheet_name]
                dest_col = week_to_column[week_num]

                logger.info(f"  Processing Week {week_num} → Column {dest_col}...")

                # Copy H3:H13 to C5:C15 (11 cells)
                for i in range(11):
                    source_row = 3 + i  # H3 to H13
                    dest_row = 5 + i    # C5 to C15
                    source_value = week_sheet[f"H{source_row}"].value
                    nbd_sheet[f"{dest_col}{dest_row}"].value = source_value

                # Copy H14 to C19 (1 cell)
                source_value = week_sheet["H14"].value
                nbd_sheet[f"{dest_col}19"].value = source_value

                # Copy H15:H20 to C22:C27 (6 cells)
                for i in range(6):
                    source_row = 15 + i  # H15 to H20
                    dest_row = 22 + i    # C22 to C27
                    source_value = week_sheet[f"H{source_row}"].value
                    nbd_sheet[f"{dest_col}{dest_row}"].value = source_value

                logger.info(f"    ✓ Week {week_num} data copied to column {dest_col}")

            # Duplicate Week 5 (column G) to column H
            logger.info("  Duplicating Week 5 data from column G to column H...")

            # Copy G5:G15 to H5:H15
            for i in range(11):
                row_num = 5 + i
                nbd_sheet[f"H{row_num}"].value = nbd_sheet[f"G{row_num}"].value

            # Copy G19 to H19
            nbd_sheet["H19"].value = nbd_sheet["G19"].value

            # Copy G22:G27 to H22:H27
            for i in range(6):
                row_num = 22 + i
                nbd_sheet[f"H{row_num}"].value = nbd_sheet[f"G{row_num}"].value

            logger.info("    ✓ Column H filled with Week 5 data")
            logger.info("✓ NBD_WF_04_LA sheet filled successfully")

            return True

        except Exception as e:
            logger.error(f"Error filling NBD_WF_04_LA sheet: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False

    def generate_monthly_report(self):
        """
        Generate the monthly liquidity report from 4-5 weekly AFL files
        """
        template_wb = None
        weekly_wbs = []

        try:
            logger.info("\n" + "=" * 60)
            logger.info("GENERATING MONTHLY LIQUIDITY REPORT")
            logger.info("=" * 60)

            # Validate we have at least 5 weekly files
            if len(self.afl_files) < 5:
                logger.error(f"Need at least 5 weekly files, only have {len(self.afl_files)}")
                return False

            # Find monthly template
            template_path = self.find_monthly_template()
            if not template_path:
                logger.error("Monthly template not found - cannot generate report")
                return False

            # Create output filename with month and year
            month_name = self.report_date.strftime("%B")  # e.g., "July"
            year = self.report_date.strftime("%Y")  # e.g., "2025"
            output_filename = f"NBD-MF-04-LA Liquid Assets - {month_name} {year}.xlsx"

            # Create output folder if it doesn't exist
            self.output_folder.mkdir(parents=True, exist_ok=True)
            logger.info(f"Output folder: {self.output_folder}")

            output_path = self.output_folder / output_filename

            # Copy template to output location
            logger.info(f"Creating monthly report: {output_filename}")
            shutil.copy2(template_path, output_path)

            # Open the output file
            template_wb = openpyxl.load_workbook(output_path)
            logger.info(f"Opened template: {template_path.name}")
            logger.info(f"Available sheets: {template_wb.sheetnames}")

            # Process weeks 1-4 (from actual AFL files)
            for week_num, afl_file_path, file_date in self.afl_files:
                if week_num > 4:
                    continue  # Only process weeks 1-4 from files

                logger.info(f"\nProcessing Week {week_num}...")

                # Extract data from weekly file
                success, wb_data = self.extract_weekly_data(week_num, afl_file_path)

                if not success:
                    logger.error(f"Failed to extract Week {week_num} data")
                    continue

                wb, source_sheet = wb_data
                weekly_wbs.append(wb)

                # Find destination sheet in template
                dest_sheet_name = f"Week {week_num}"
                if dest_sheet_name not in template_wb.sheetnames:
                    logger.error(f"Sheet '{dest_sheet_name}' not found in template")
                    continue

                dest_sheet = template_wb[dest_sheet_name]
                logger.info(f"  ✓ Found destination sheet: {dest_sheet_name}")

                # Copy A3:C25 range with formatting to A1 (to map to A1:H22 area)
                self.copy_range_with_formatting(
                    source_sheet=source_sheet,
                    dest_sheet=dest_sheet,
                    source_range="A3:C25",
                    dest_start_cell="A1"
                )

            # Handle Week 5
            logger.info("\nProcessing Week 5...")

            # Check if we have a Week 5 file
            week5_file = None
            for week_num, afl_file_path, file_date in self.afl_files:
                if week_num == 5:
                    week5_file = afl_file_path
                    break

            if week5_file:
                # Use actual Week 5 data
                logger.info(f"Found Week 5 file: {week5_file.name}")
                success, wb_data = self.extract_weekly_data(5, week5_file)

                if success:
                    wb, source_sheet = wb_data
                    weekly_wbs.append(wb)

                    if "Week 5" in template_wb.sheetnames:
                        dest_sheet = template_wb["Week 5"]
                        self.copy_range_with_formatting(
                            source_sheet=source_sheet,
                            dest_sheet=dest_sheet,
                            source_range="A3:C25",
                            dest_start_cell="A1"
                        )
                        logger.info("  ✓ Week 5 populated with actual Week 5 data")
            else:
                # Duplicate Week 4 data to Week 5
                logger.info("No Week 5 file found - duplicating Week 4 data to Week 5...")
                if "Week 4" in template_wb.sheetnames and "Week 5" in template_wb.sheetnames:
                    week4_sheet = template_wb["Week 4"]
                    week5_sheet = template_wb["Week 5"]

                    self.copy_range_with_formatting(
                        source_sheet=week4_sheet,
                        dest_sheet=week5_sheet,
                        source_range="A1:C23",
                        dest_start_cell="A1"
                    )
                    logger.info("  ✓ Week 5 populated with Week 4 data")

            # Fill NBD_WF_04_LA sheet with weekly H column data
            success = self.fill_nbd_wf_04_la_sheet(template_wb)
            if not success:
                logger.warning("Failed to fill NBD_WF_04_LA sheet, but continuing...")

            # Save the monthly report
            template_wb.save(output_path)
            logger.info(f"\n✓ Monthly report saved: {output_path}")

            logger.info("\n" + "=" * 60)
            logger.info("MONTHLY REPORT GENERATION COMPLETED SUCCESSFULLY")
            logger.info(f"Output file: {output_filename}")
            logger.info(f"Location: {output_path}")
            logger.info("=" * 60)

            return True

        except Exception as e:
            logger.error(f"Error generating monthly report: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False

        finally:
            # Close all workbooks
            if template_wb:
                try:
                    template_wb.close()
                except:
                    pass

            for wb in weekly_wbs:
                try:
                    wb.close()
                except:
                    pass


def main():
    """
    Main function - typically called by NBD_MF_15_LA.py
    """
    logger.info("NBD_MF_04_LA Monthly Report Generator")
    logger.info("This script is typically called automatically by NBD_MF_15_LA.py")


if __name__ == "__main__":
    main()
