import os
from pathlib import Path
import openpyxl
import sys
import argparse
import logging
from datetime import datetime

def get_month_year_from_filename(filename):
    # Example: "NBD-MF-20-C1 to C6 July 2025.xlsx"
    parts = filename.split()
    for i, part in enumerate(parts):
        if part in {"Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","January","February","March","April","May","June","July","August","September","October","November","December"}:
            month = part
            year = parts[i+1].replace(".xlsx","")
            return month, year
    return None, None

def find_files_safely(base_dir, pattern, description):
    """Safely find files and provide detailed error messages"""
    print(f"Looking for {description} in: {base_dir}")
    print(f"Search pattern: {pattern}")
    
    files = list(base_dir.glob(pattern))
    print(f"Found {len(files)} files matching pattern:")
    
    if not files:
        print(f"No files found matching pattern: {pattern}")
        print("Available files in directory:")
        try:
            all_files = [f.name for f in base_dir.iterdir() if f.is_file() and f.suffix == '.xlsx']
            if all_files:
                for file in sorted(all_files):
                    print(f"  - {file}")
            else:
                print("  No .xlsx files found in directory")
        except Exception as e:
            print(f"  Error listing files: {e}")
        return None
    
    for file in files:
        print(f"  {file.name}")
    
    return files[0]  # Return the first match

def setup_logging():
    """Setup organized logging for C4 report"""
    # Configure logging
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    logger = logging.getLogger(__name__)
    
    # Create organized log directory: logs/YYYY-MM-DD/frequency/report_name.log
    try:
        frequency = "monthly"
        date_folder = datetime.now().strftime('%Y-%m-%d')
        run_timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        logs_dir = Path(__file__).parent.parent / "logs" / date_folder / frequency
        logs_dir.mkdir(parents=True, exist_ok=True)
        run_log_file = logs_dir / f"NBD_MF_20_C4_{run_timestamp}.log"
    except Exception:
        # Fallback to simple structure
        logs_dir = Path(__file__).parent / "logs"
        logs_dir.mkdir(parents=True, exist_ok=True)
        run_log_file = logs_dir / f"NBD_MF_20_C4_{run_timestamp}.log"
    
    file_handler = logging.FileHandler(run_log_file, encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    logger.addHandler(file_handler)
    
    logger.info(f"📄 Detailed log file: {run_log_file}")
    return logger

def main():
    # Parse command line arguments
    parser = argparse.ArgumentParser(description='NBD MF 20 C4 Report Generator')
    parser.add_argument('--working-dir', type=str, help='Working directory path')
    parser.add_argument('--month', type=str, help='Report month (e.g., Jul)')
    parser.add_argument('--year', type=str, help='Report year (e.g., 2025)')
    
    args = parser.parse_args()
    
    # Setup logging
    logger = setup_logging()
    
    # Determine working directory
    if args.working_dir:
        # Use provided working directory - it should point to the date folder
        monthly_folder = Path(args.working_dir)
        working_dir = monthly_folder.parent.parent
    else:
        # Fallback to default behavior
        base_dir = Path(__file__).resolve().parent.parent
        working_dir = base_dir / "working"
        monthly_working_dir = working_dir / "monthly"
        
        # Find the monthly folder (there should be only one)
        monthly_folders = [f for f in monthly_working_dir.iterdir() if f.is_dir()]
        
        if not monthly_folders:
            print(f"No monthly folders found in: {monthly_working_dir}")
            return
        
        if len(monthly_folders) > 1:
            print(f"Multiple monthly folders found, using the first one: {monthly_folders[0].name}")
        
        monthly_folder = monthly_folders[0]
    
    logger.info(f"Base directory: {Path(__file__).resolve().parent.parent}")
    logger.info(f"Working directory: {working_dir}")
    logger.info(f"Monthly folder: {monthly_folder}")
    print(f"Base directory: {Path(__file__).resolve().parent.parent}")
    print(f"Working directory: {working_dir}")
    print(f"Monthly folder: {monthly_folder}")
    print(f"Current working directory: {Path.cwd()}")
    
    file_c1_c6 = None
    file_unutilized = None
    
    # Look for the specific files in the monthly folder and subdirectories
    print(f"\nSearching for files in: {monthly_folder}")
    
    # The final working folder is NBD_MF_20_C1_C6 inside the date folder
    c1_c6_subfolder = monthly_folder / "NBD_MF_20_C1_C6"
    print(f"Final working folder: {c1_c6_subfolder}")
    
    # Look for C1-C6 file in subdirectories
    c1_c6_pattern = "**/NBD-MF-20-C1 to C6*.xlsx"
    c1_c6_files = list(monthly_folder.glob(c1_c6_pattern))
    
    if c1_c6_files:
        file_c1_c6 = c1_c6_files[0]
        print(f"Found C1-C6 file: {file_c1_c6}")
    else:
        print(f"Could not find C1-C6 file with pattern: {c1_c6_pattern}")
        print("Available files in directory and subdirectories:")
        for f in monthly_folder.rglob("*.xlsx"):
            print(f"  - {f.relative_to(monthly_folder)}")
        return
    
    # Look for Unutilized Credit Limits file in subdirectories
    unutilized_pattern = "**/Unutilized Credit Limits*.xlsx"
    unutilized_files = list(monthly_folder.glob(unutilized_pattern))
    
    if unutilized_files:
        file_unutilized = unutilized_files[0]
        print(f"Found Unutilized Credit Limits file: {file_unutilized}")
    else:
        print(f"Could not find Unutilized Credit Limits file with pattern: {unutilized_pattern}")
        print("Available files in directory and subdirectories:")
        for f in monthly_folder.rglob("*.xlsx"):
            print(f"  - {f.relative_to(monthly_folder)}")
        return
    
    # Extract month and year from filename
    month, year = get_month_year_from_filename(file_c1_c6.name)
    if not month or not year:
        print(f"Could not parse month/year from filename: {file_c1_c6.name}")
        print("Expected format: 'NBD-MF-20-C1 to C6 [Month] [Year].xlsx'")
        return
    
    print(f"\nParsed month: {month}, year: {year}")
    
    try:
        # Load workbooks
        print(f"\nLoading workbooks...")
        wb_c1_c6 = openpyxl.load_workbook(file_c1_c6, data_only=True)
        wb_unutilized = openpyxl.load_workbook(file_unutilized, data_only=True)
        print(f"Workbooks loaded successfully")
        
        # Check if required sheets exist
        print(f"\nAvailable sheets in C1-C6 file: {wb_c1_c6.sheetnames}")
        print(f"Available sheets in Unutilized file: {wb_unutilized.sheetnames}")
        
        # Check for C4 sheet in C1-C6 file
        if "NBD_NBL-MF-20-C4" not in wb_c1_c6.sheetnames:
            print(f"Required sheet 'NBD_NBL-MF-20-C4' not found in C1-C6 file")
            print(f"Available sheets: {wb_c1_c6.sheetnames}")
            return
        
        # Find any sheet that starts with "Working"
        working_sheet_name = None
        for sheet_name in wb_unutilized.sheetnames:
            if sheet_name.startswith("Working"):
                working_sheet_name = sheet_name
                break
        
        if working_sheet_name is None:
            print(f"No sheet starting with 'Working' found in Unutilized file")
            print(f"Available sheets: {wb_unutilized.sheetnames}")
            return
        
        print(f"Found working sheet: {working_sheet_name}")
        
        print(f"All required sheets found")
        
        # Get worksheets
        src_ws = wb_unutilized[working_sheet_name]
        tgt_ws = wb_c1_c6["NBD_NBL-MF-20-C4"]
        
        print(f"\nProcessing data...")
        
        # Helper function to safely get numeric values
        def get_numeric_value(worksheet, cell_ref, description=""):
            """Safely extract numeric value from a cell, handling text/None values"""
            try:
                value = worksheet[cell_ref].value
                print(f"{description} ({cell_ref}): {value}")
                
                if value is None:
                    return 0
                elif isinstance(value, (int, float)):
                    return value
                elif isinstance(value, str):
                    # Try to extract number from string if possible
                    try:
                        # Remove common text and try to parse number
                        clean_value = value.replace(',', '').strip()
                        return float(clean_value)
                    except ValueError:
                        print(f"  Warning: Cell {cell_ref} contains text '{value}' - using 0")
                        return 0
                else:
                    print(f"  Warning: Cell {cell_ref} contains unexpected type {type(value)} - using 0")
                    return 0
            except Exception as e:
                print(f"  Error reading {cell_ref}: {e} - using 0")
                return 0
        
        # Find the "Unutilized Loan Balances" column and get total from last row
        print(f"\nSearching for 'Unutilized Loan Balances' column in {working_sheet_name} sheet...")
        
        total_value = None
        unutilized_col = None
        
        # First, find the "Unutilized Loan Balances" column header
        for row in range(1, min(10, src_ws.max_row + 1)):  # Check first 10 rows for headers
            for col in range(1, src_ws.max_column + 1):
                cell_value = src_ws.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str) and "Unutilized Loan Balances" in cell_value:
                    print(f"Found 'Unutilized Loan Balances' header at row {row}, column {col}")
                    unutilized_col = col
                    break
            if unutilized_col:
                break
        
        if unutilized_col is None:
            print("Could not find 'Unutilized Loan Balances' column header in unutilized file")
            print("Available cell values in the first 10 rows:")
            for row in range(1, min(10, src_ws.max_row + 1)):
                for col in range(1, min(10, src_ws.max_column + 1)):
                    cell_value = src_ws.cell(row=row, column=col).value
                    if cell_value:
                        print(f"  Row {row}, Col {col}: {cell_value}")
            return
        
        # Now find the last row with data in the Unutilized Loan Balances column
        print(f"Looking for total value in the last row of column {unutilized_col}...")
        
        # Find the last row with data in this column
        last_row = 1
        for row in range(src_ws.max_row, 0, -1):  # Start from bottom and go up
            cell_value = src_ws.cell(row=row, column=unutilized_col).value
            if cell_value is not None and cell_value != "":
                last_row = row
                break
        
        print(f"Last row with data in Unutilized Loan Balances column: {last_row}")
        
        # Get the value from the last row of the Unutilized Loan Balances column
        col_letter = openpyxl.utils.get_column_letter(unutilized_col)
        total_value = get_numeric_value(src_ws, f"{col_letter}{last_row}", f"Unutilized Loan Balances Total (Row {last_row})")
        
        if total_value is None or total_value == 0:
            print("Could not find valid total value in the last row of Unutilized Loan Balances column")
            print("Please check the file structure and ensure the data is in the expected format")
            return
        
        # Find the "Commitments" section in C4 sheet
        print(f"\nSearching for 'Commitments' section in C4 sheet...")
        
        commitments_row = None
        for row in range(1, tgt_ws.max_row + 1):
            for col in range(1, tgt_ws.max_column + 1):
                cell_value = tgt_ws.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str) and "Commitments" in cell_value:
                    print(f"Found 'Commitments' at row {row}, column {col}")
                    commitments_row = row
                    break
        
        if commitments_row is None:
            print("Could not find 'Commitments' section in C4 sheet")
            print("Available cell values in C4 sheet:")
            for row in range(1, min(20, tgt_ws.max_row + 1)):  # Show first 20 rows
                for col in range(1, min(10, tgt_ws.max_column + 1)):  # Show first 10 columns
                    cell_value = tgt_ws.cell(row=row, column=col).value
                    if cell_value:
                        print(f"  Row {row}, Col {col}: {cell_value}")
            return
        
        # Find "with an original maturity up to 1 year" row
        print(f"\nSearching for 'with an original maturity up to 1 year' in C4 sheet...")
        
        maturity_row = None
        for row in range(commitments_row, tgt_ws.max_row + 1):
            for col in range(1, tgt_ws.max_column + 1):
                cell_value = tgt_ws.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str) and "With an original maturity, up to 1 year" in cell_value:
                    print(f"Found 'With an original maturity, up to 1 year' at row {row}, column {col}")
                    maturity_row = row
                    break
        
        if maturity_row is None:
            print("Could not find 'with an original maturity up to 1 year' in C4 sheet")
            return
        
        # Find the "Principal amount of Off-Balance Sheet Item" column (column C)
        print(f"\nLooking for 'Principal amount of Off-Balance Sheet Item' column...")
        
        # Column C is the 3rd column
        target_col = 3  # Column C
        target_cell = f"C{maturity_row}"
        
        print(f"Updating cell {target_cell} with value: {total_value}")
        tgt_ws[target_cell] = total_value
        
        print(f"Data processing completed")
        print(f"Updated C4 sheet: {target_cell} = {total_value}")
        
        # Save updated workbook directly to the same location
        wb_c1_c6.save(file_c1_c6)
        print(f"\nReport saved successfully to: {file_c1_c6}")
        
    except FileNotFoundError as e:
        print(f"File not found: {e}")
    except PermissionError as e:
        print(f"Permission error (file might be open in Excel): {e}")
    except Exception as e:
        print(f"An error occurred: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()