import pandas as pd
import numpy as np
import os
import glob
from datetime import datetime, timedelta
import argparse
from pathlib import Path

from FixedLoans import create_fixed_loans_df
from MT import create_marginal_loans_df
from FDLquarter import create_FDL_quarter_df
from disbursement_processor import get_disbursement_df
from mt_report_consolidator import consolidate_mt_reports

# Define constants - paths relative to script root directory
SCRIPT_ROOT = os.path.dirname(os.path.abspath(__file__))

parent_directory = Path(SCRIPT_ROOT).parent

# Allow passing the <date-folder> explicitly; fallback to auto-detect
parser = argparse.ArgumentParser(description='Process summary using new folder structure')
parser.add_argument('--working-dir', type=str, help='Path to working date folder (working/NBD-QF-23-C3-C10-SF-FI/<date-folder>)')
args, unknown = parser.parse_known_args()

if args.working_dir:
    provided_path = Path(args.working_dir)
    if not provided_path.exists() or not provided_path.is_dir():
        raise FileNotFoundError(f"Working directory not found: {provided_path}")
    # If the provided path is the base folder (working/NBD-QF-23-C3-C10-SF-FI), pick its single <date-folder>
    subdirs = [d for d in provided_path.iterdir() if d.is_dir()]
    if subdirs:
        if len(subdirs) > 1:
            print(f"Warning: multiple <date-folder>s found under base directory, using the first: {subdirs[0].name}")
        date_folder = subdirs[0]
    else:
        # Treat the provided path as the <date-folder>
        date_folder = provided_path
else:
    # New structure: working/NBD-QF-23-C3-C10-SF-FI/<date-folder>/ (exactly one folder)
    base_working_dir = parent_directory / "working" / "NBD-QF-23-C3-C10-SF-FI"
    if not base_working_dir.exists():
        raise FileNotFoundError(f"Working directory not found: {base_working_dir}")
    date_folders = [p for p in base_working_dir.iterdir() if p.is_dir()]
    if not date_folders:
        raise FileNotFoundError(f"No <date-folder> found inside: {base_working_dir}")
    if len(date_folders) > 1:
        # Use the first one by default; adjust if you need specific selection logic
        print(f"Warning: multiple <date-folder>s found, using the first: {date_folders[0].name}")
    date_folder = date_folders[0]

# Get current date for file name generation
now = datetime.now()
current_month_date = now.replace(day=1) - timedelta(days=1)  # Last day of previous month

# Folder paths (operate strictly within <date-folder>)
INPUT_FOLDER = str(date_folder / "Input")
OUTPUT_FOLDER = str(date_folder)

def find_file_with_pattern(folder_path, pattern):
    """Find file using glob pattern within folder_path (including subfolders) and return the first match"""
    # Search recursively to allow files placed outside Input or in subfolders of the date folder
    search_pattern = os.path.join(folder_path, "**", pattern)
    matches = glob.glob(search_pattern, recursive=True)
    if matches:
        return matches[0]
    else:
        print(f"Warning: No file found matching pattern: {pattern}")
        return None

def generate_dynamic_file_paths():
    """Generate all file paths dynamically based on current date"""
    # Current month report is for previous month (end date)
    current_month_date = now.replace(day=1) - timedelta(days=1)
    current_month_str = current_month_date.strftime("%Y-%m-%d")  # 2025-09-30
    current_month_name = current_month_date.strftime("%b %Y")  # Sep 2025
    current_month_short = current_month_date.strftime("%b %y")  # Sep 25

    # Next month (for Net Portfolio)
    next_month = now
    next_month_name = next_month.strftime("%b-%Y")  # Oct-2025

    # Previous month
    prev_month = (current_month_date.replace(day=1) - timedelta(days=1))
    prev_month_name = prev_month.strftime("%b %Y")  # Aug 2025
    last_q_monthname = (current_month_date.replace(day=1) - timedelta(days=90)).strftime("%b %Y")  # Jun 2025
    file_paths = {
        'file_path': find_file_with_pattern(INPUT_FOLDER, f"3. Summary_{current_month_str}*.xlsb"),
        'file_path_portfolio': find_file_with_pattern(INPUT_FOLDER, f"1. Portfolio_4_ECL_-{current_month_str}*.xlsb"),
        'file_path_netPortfolio': find_file_with_pattern(INPUT_FOLDER, f"Net Portfolio-{next_month_name}*.xlsx"),
        'cbsl_provision': find_file_with_pattern(INPUT_FOLDER, f"CBSL Provision Comparison - {current_month_name}*.xlsb"),
        'file_path_cbsl_sec': find_file_with_pattern(INPUT_FOLDER, f"1.Sector Report- {current_month_name}*.xlsb"),
        'file_path_district': find_file_with_pattern(INPUT_FOLDER, "JOBHD*.xlsx"),
        'file_path_micro': find_file_with_pattern(INPUT_FOLDER, "Copy of Information Request from Credit - C10C11*.xlsx"),
        'file_path_reschedule': find_file_with_pattern(INPUT_FOLDER, "Reschedule Contract Details - Internal*.xlsx"),
        'file_path_monthlyreport': find_file_with_pattern(INPUT_FOLDER, f"Prod. wise Class. of Loans - {next_month_name}*.xlsb"),
        'file_path_fixedloans': find_file_with_pattern(INPUT_FOLDER, "LoanAsAtDate*.csv") or find_file_with_pattern(INPUT_FOLDER, "LoanAsAtDate*.xlsx"),
        'file_path_MT': find_file_with_pattern(INPUT_FOLDER, f"Unutilized Amount {prev_month_name}*.xlsx"),
        'report_file': find_file_with_pattern(OUTPUT_FOLDER, f"Quarter Classification Loans & Advance-{last_q_monthname}*.xlsx")
    }

    return file_paths

# Generate file paths dynamically
print(f"Looking for files in: {INPUT_FOLDER}")
print(f"Date folder: {date_folder.name}")

file_paths_dict = generate_dynamic_file_paths()

# Extract individual file paths from dictionary
file_path = file_paths_dict['file_path']
file_path_portfolio = file_paths_dict['file_path_portfolio']
file_path_netPortfolio = file_paths_dict['file_path_netPortfolio']
cbsl_provision = file_paths_dict['cbsl_provision']
file_path_cbsl_sec = file_paths_dict['file_path_cbsl_sec']
file_path_district = file_paths_dict['file_path_district']
file_path_micro = file_paths_dict['file_path_micro']
file_path_reschedule = file_paths_dict['file_path_reschedule']
file_path_monthlyreport = file_paths_dict['file_path_monthlyreport']
file_path_fixedloans = file_paths_dict['file_path_fixedloans']
file_path_MT = file_paths_dict['file_path_MT']
report_file = file_paths_dict['report_file']

# Verify all files were found
print("\nFile paths resolved:")
for key, path in file_paths_dict.items():
    status = "✓" if path else "✗"
    print(f"{status} {key}: {path if path else 'NOT FOUND'}")

df = pd.read_excel(file_path, sheet_name='SUMMARY', skiprows=2, engine='pyxlsb')
portfolio_df = pd.read_excel(file_path_portfolio, sheet_name='Portfolio', skiprows=2, engine='pyxlsb')
net_portfolio_df = pd.read_excel(file_path_netPortfolio)
cbsl_provision = pd.read_excel(cbsl_provision, sheet_name='Portfolio', skiprows=2, engine='pyxlsb')



# Find the STAGE column index
stage_col_index = df.columns.get_loc('STAGE')

# Keep only columns up to and including STAGE
df = df.iloc[:, :stage_col_index + 1]

maindf = df.drop(columns=['RELATED-PARTY CONTRACT','STAFF CONTRACT','ACTUAL LEASE/LOAN'])

maindf = maindf.dropna(subset=['CONTRACT NO'])

contract_to_freq = dict(zip(net_portfolio_df["CONTRACT_NO"], net_portfolio_df["CON_RNTFREQ"]))
contract_to_sector = dict(zip(net_portfolio_df["CONTRACT_NO"], net_portfolio_df["SEC_DESC"]))
contract_to_sub_sec= dict(zip(net_portfolio_df["CONTRACT_NO"], net_portfolio_df["SUB_SECTOR"]))

contract_to_prov_cat = dict(zip(cbsl_provision["CONTRACT NO"], cbsl_provision["PROVISION CATEGORY"]))
contract_to_prov_guid = dict(zip(cbsl_provision["CONTRACT NO"], cbsl_provision["FINAL PROVISION (CBSL GUIDELINE)"]))
contract_to_prov_pnp = dict(zip(cbsl_provision["CONTRACT NO"], cbsl_provision["CBSL P/NP"]))

# Map the frequency values to the main dataframe
maindf['Frequency'] = maindf['CONTRACT NO'].map(contract_to_freq)

# Add Repayment Cycle column with Excel formulas
# Excel row 3 corresponds to df index 0 (since we skipped 2 rows)
maindf['Repayment Cycle'] = [f'=IF(OR(AE{i+3}="M",AE{i+3}="D"),"Monthly Basis","Other")' for i in range(len(maindf))]
maindf['CBSL DPD'] = [f'=F{i + 3}' for i in range(len(maindf))]
maindf['Provision Category'] = maindf['CONTRACT NO'].map(contract_to_prov_cat)

# Replace 'NO PROVISION' with 'Performing' in Provision Category column
maindf['Provision Category'] = maindf['Provision Category'].replace('NO PROVISION', 'Performing')
maindf['Provision - CBSL Guideline'] = maindf['CONTRACT NO'].map(contract_to_prov_guid)
maindf['CBSL P/NP'] = maindf['CONTRACT NO'].map(contract_to_prov_pnp)
maindf['Forbone Loans (Yes/No)'] = [f'=IF(ISNUMBER(MATCH(A{i + 3},\'MOI List -31 Jan\'!A:A, 0)), "Yes", "No")' for i in range(len(maindf))]
maindf['EXP+DP-IIS'] = [f'=Y{i + 3}-T{i + 3}' for i in range(len(maindf))]
maindf['Gross-Imp-IIS'] = [f'=E{i+3}-J{i+3}-T{i+3}' for i in range(len(maindf))]
maindf['Exp+DP-IMP-IIS'] = [f'=Y{i+3}-Z{i+3}-T{i+3}' for i in range(len(maindf))]
maindf['Collateral/Security Type'] = [f'=IF(ISNUMBER(MATCH(A{i+3},\'Property Mortgage List\'!A:A,0)),"Immovable Properties",IF(OR(AA{i+3}="UV",AA{i+3}="LE",AA{i+3}="AT"),"Vehicles and Machinery",IF(ISNUMBER(MATCH(R{i+3},Cat_List!$A$2:$A$9,0)),"Vehicles and Machinery",IF(A{i+3}="Margin Trading","Shares and Debt Securities-Listed","Personal and Corporate Guarantees"))))' for i in range(len(maindf))]
maindf['Sector'] = maindf['CONTRACT NO'].map(contract_to_sector)
maindf['Sub-Sector'] = maindf['CONTRACT NO'].map(contract_to_sub_sec)
#sector working calculation
#TODO: Add sector working calculation

df_sector = pd.read_excel(file_path_cbsl_sec, sheet_name='Portfolio', skiprows=1, engine='pyxlsb')
contract_to_cbsl_sec= dict(zip(df_sector["Contract No"], df_sector["CBSL Sector 1 Final"]))
maindf['CBSL Sector'] = maindf['CONTRACT NO'].map(contract_to_cbsl_sec)

df_district = pd.read_excel(file_path_district)
contract_to_district = dict(zip(df_district["CLM_CODE"], df_district["DISTRICT"]))
maindf['CustomerDistrict(ClientMain)'] = maindf['CLIENT NO'].map(contract_to_district)

# Clean up column names
net_portfolio_df.columns = net_portfolio_df.columns.str.strip()

# Then try again
df_unique = net_portfolio_df.drop_duplicates(subset=['CLIENT_CODE'], keep='first')
contract_to_client_name = dict(zip(df_unique["CLIENT_CODE"], df_unique["CLM_NAME"]))
maindf['Customer Name'] = maindf['CLIENT NO'].map(contract_to_client_name)

df_micro_dis = pd.read_excel(file_path_micro, sheet_name=0, skiprows=4)
df_micro_port = pd.read_excel(file_path_micro, sheet_name=1, skiprows=3)

# Assuming your DataFrames are named df1 and df2

# Select relevant columns from the first DataFrame
df1_relevant = df_micro_port[['Client Code', 'Contract No', 'Product', 'Customer Name', 'Contract Amount', 'Micro/Small/Medium', 'Remarks']]

# Select relevant columns from the second DataFrame
df2_relevant = df_micro_dis[['Client Code', 'Contract No', 'Product', 'Customer Name', 'Net Monthly Disbursement (LKR)', 'Micro/Small/Medium']]

# Concatenate them
combined_df = pd.concat([df1_relevant, df2_relevant], ignore_index=True)

contract_to_ct_micro = dict(zip(combined_df["Contract No"], combined_df["Micro/Small/Medium"]))
maindf['Micro/Small/Medium'] = maindf['CONTRACT NO'].map(contract_to_ct_micro)

#TODO: get gender from DF and concat new data sheet and old sheet to be paste as values
maindf['Gender'] = [f'=IF(LEFT(B{i + 3},1)="2",_xlfn.XLOOKUP(B1,BusinessGender!A:A,BusinessGender!C:C,"No Data"),IF(LEFT(AT{i + 3},3)="Mr.","Male",IF(LEFT(AT{i + 3},3)="Mr ","Male",IF(LEFT(AT{i + 3},3)="Rev","Male",IF(LEFT(AT{i + 3},4)="Miss","Female",IF(LEFT(AT{i + 3},3)="Ms.","Female",IF(LEFT(AT{i + 3},4)="Mrs.","Female"))))))' for i in range(len(maindf))]

df_reschedule = pd.read_excel(file_path_reschedule)
contract_to_old_contract = dict(zip(df_reschedule["CLO_NEWCONNO"], df_reschedule["CON_NO"]))
maindf['Old Contract No (Before Reschedule)'] = maindf['CONTRACT NO'].map(contract_to_old_contract)

contract_to_grantAm= dict(zip(net_portfolio_df["CONTRACT_NO"], net_portfolio_df["CONTRACT_AMOUNT"]))
maindf['Grant Amount'] = maindf['CONTRACT NO'].map(contract_to_grantAm)

df_monthly_report = pd.read_excel(file_path_monthlyreport, sheet_name='C1 & C2 Working', skiprows=1, engine='pyxlsb')
contract_to_initialVal= dict(zip(df_monthly_report["Contract No"], df_monthly_report["Initial Valuation"]))
maindf['Initial Valuation'] = maindf['CONTRACT NO'].map(contract_to_initialVal)

maindf['LTV %'] = np.where(maindf['Initial Valuation'].isna() | maindf['Grant Amount'].isna(), np.nan, maindf['Grant Amount'] / maindf['Initial Valuation'])

maindf['WALTV %'] = [f'=IF(AO{i+3}="Vehicles and Machinery",(E{i+3}/$AZ$1)*AZ{i+3},0)' for i in range(len(maindf))]
fixedloan = create_fixed_loans_df(file_path_fixedloans, df_monthly_report, len(maindf))
# Reset column names to match by position
fixedloan.columns = maindf.columns

# Now concat
maindf = pd.concat([maindf, fixedloan], ignore_index=True)
maindf['Top 50 Clients'] = [f'=VLOOKUP(B{i+3},\'Top50\'!A:E,5,0)' for i in range(len(maindf))]


marginal_loan = create_marginal_loans_df(file_path_MT, df_district, combined_df)

df_FDL_quarter = create_FDL_quarter_df(df_monthly_report, file_path_fixedloans)

disbursement_df = get_disbursement_df(maindf, net_portfolio_df)

MT_disbursement = consolidate_mt_reports()
# Display the resulting dataframe
print("Data loaded successfully!")
print(f"Shape: {disbursement_df.shape}")
print("\nFirst few rows:")
print(maindf.head())


output_path = r'disoutput_summary.xlsx'
disbursement_df.to_excel(output_path, index=False)
print(f"\nData written to {output_path}")

# Open the report file and update sheets
from openpyxl import load_workbook

print(f"\nUpdating {report_file}...")

# First convert .xlsb to .xlsx if needed
if report_file.endswith('.xlsb'):
    import xlwings as xw
    print("Converting .xlsb to .xlsx...")
    app = xw.App(visible=False)
    try:
        wb_temp = xw.Book(report_file)
        xlsx_file = report_file.replace('.xlsb', '.xlsx')
        wb_temp.save(xlsx_file)
        wb_temp.close()
        report_file = xlsx_file
    finally:
        app.quit()

# Now use openpyxl
wb = load_workbook(report_file)

# 1. Summary sheet - clear from row 3 onwards and paste maindf without headings
if 'Summary' in wb.sheetnames:
    ws_summary = wb['Summary']
    # Clear data from row 3 onwards
    ws_summary.delete_rows(3, ws_summary.max_row - 2)
    # Paste maindf to row 3 without headings
    maindf_clean = maindf.replace([np.inf, -np.inf], np.nan).fillna('')
    for i, row in enumerate(maindf_clean.values, start=3):
        for j, val in enumerate(row, start=1):
            cell = ws_summary.cell(row=i, column=j)
            if isinstance(val, str) and val.startswith('='):
                cell.value = val  # openpyxl auto-detects formulas
            else:
                cell.value = val.item() if isinstance(val, (np.integer, np.floating)) else val
    print("Updated SUMMARY sheet")

# 2. MT sheet - clear from row 2 onwards and paste marginal_loan without headings
if 'MT' in wb.sheetnames:
    ws_mt = wb['MT']
    # Clear data from row 2 onwards
    ws_mt.delete_rows(2, ws_mt.max_row - 1)
    # Paste marginal_loan to row 2 without headings (starting at column C)
    marginal_clean = marginal_loan.replace([np.inf, -np.inf], np.nan).fillna('')
    for i, row in enumerate(marginal_clean.values, start=2):
        for j, val in enumerate(row, start=3):  # Start at column C (3)
            cell = ws_mt.cell(row=i, column=j)
            if isinstance(val, str) and val.startswith('='):
                cell.value = val
            else:
                cell.value = val.item() if isinstance(val, (np.integer, np.floating)) else val
    print("Updated MT sheet")

# 3. FDL Quarter sheet - clear from row 3 onwards and paste df_FDL_quarter without headings
if 'FDL Quarter' in wb.sheetnames:
    ws_fdl = wb['FDL Quarter']
    # Clear data from row 3 onwards
    ws_fdl.delete_rows(3, ws_fdl.max_row - 2)
    # Paste df_FDL_quarter to row 3 without headings
    fdl_clean = df_FDL_quarter.replace([np.inf, -np.inf], np.nan).fillna('')
    for i, row in enumerate(fdl_clean.values, start=3):
        for j, val in enumerate(row, start=1):
            cell = ws_fdl.cell(row=i, column=j)
            if isinstance(val, str) and val.startswith('='):
                cell.value = val
            else:
                cell.value = val.item() if isinstance(val, (np.integer, np.floating)) else val
    print("Updated FDL Quarter sheet")

# 3. FDL Quarter sheet - clear from row 3 onwards and paste df_FDL_quarter without headings
if 'Disbursement' in wb.sheetnames:
    ws_fdl = wb['Disbursement']
    # Clear data from row 3 onwards
    ws_fdl.delete_rows(5, ws_fdl.max_row - 2)
    # Paste df_FDL_quarter to row 3 without headings
    disbursement_df = disbursement_df.replace([np.inf, -np.inf], np.nan).fillna('')
    for i, row in enumerate(disbursement_df.values, start=5):
        for j, val in enumerate(row, start=1):
            cell = ws_fdl.cell(row=i, column=j)
            if isinstance(val, str) and val.startswith('='):
                cell.value = val
            else:
                cell.value = val.item() if isinstance(val, (np.integer, np.floating)) else val
    print("Updated Disbursement sheet")

# MT-Disbursement
    # Save the workbook - generate output name dynamically # October 25
    output_report = f'Quarter Classification Loans & Advance-July.xlsx'

    wb.save(output_report)

print(f"\nReport file {output_report} updated successfully!")

updated_report = pd.read_excel(output_report, sheet_name='Summary', skiprows=1)
print(f'Updated report shape: {updated_report[['Client No', 'EXP+DP-IIS']]}')
# Extract required columns from first dataframe
updated_report_top50 = updated_report[['Contract No', 'Client No', 'Customer Name',]].copy()
updated_report_top50['EXP+DP-IIS'] = updated_report['EXP+DP'] - updated_report['IIS']

# Assuming updated_MT_report is a dataframe, extract required columns
# If it's a list as shown, you need to read it first from somewhere
# For now, assuming it's already a dataframe:
updated_MT_report_subset = marginal_loan[['ICAM code', '(ICAM) Debtor']].copy()

# Rename columns in MT report to match the first dataframe
updated_MT_report_subset.rename(columns={
    'ICAM code': 'Client No',
    '(ICAM) Debtor': 'EXP+DP-IIS'
}, inplace=True)

# Concatenate the two dataframes (one under another)
combined_df = pd.concat([updated_report_top50, updated_MT_report_subset], ignore_index=True)

# Group by CLIENT NO and sum the EXP+DP-IIS values
pivoted_df = combined_df.groupby('Client No', as_index=False)['EXP+DP-IIS'].sum()

# Sort by EXP+DP-IIS from highest to lowest
final_df = pivoted_df.sort_values(by='EXP+DP-IIS', ascending=False).reset_index(drop=True)

if 'Top50' in wb.sheetnames:
    ws_fdl = wb['Top50']
    
    # Clear data in only columns A and B (from row 1 to max_row)
    max_row = ws_fdl.max_row
    for row in range(1, max_row + 1):
        ws_fdl.cell(row=row, column=1).value = None  # Column A
        ws_fdl.cell(row=row, column=2).value = None  # Column B
    
    # Paste final_df to Column A1 (including headers)
    final_df_clean = final_df.replace([np.inf, -np.inf], np.nan).fillna('')
    
    # Write headers first (starting at row 1)
    for j, col_name in enumerate(final_df_clean.columns, start=1):
        ws_fdl.cell(row=1, column=j, value=col_name)
    
    # Write data (starting at row 2)
    for i, row in enumerate(final_df_clean.values, start=2):
        for j, val in enumerate(row, start=1):
            cell = ws_fdl.cell(row=i, column=j)
            if isinstance(val, str) and val.startswith('='):
                cell.value = val
            else:
                cell.value = val.item() if isinstance(val, (np.integer, np.floating)) else val
    
    print("Updated Disbursement sheet with final_df in columns A and B")

wb.save(output_report)
print("Final DataFrame:")
print(final_df)

#get the final df top 50 rows
final_df_top50 = final_df.head(50).copy()
print(f'Final Top 50 shape: {final_df_top50[["Client No"]]}')
final_df_top50['Top 50 Clients'] = "TOP 50"
output_path = r'top50.xlsx'
final_df_top50.to_excel(output_path, index=False)
contract_toTop50 = dict(zip(final_df_top50["Client No"], final_df_top50['Top 50 Clients']))

temp = pd.DataFrame()
temp['Contract No'] = np.nan
temp['Client No'] = marginal_loan['ICAM code']
temp['Customer Name'] = marginal_loan['Name of the Client']
temp['EXP+DP-IIS'] = marginal_loan['(ICAM) Debtor']

updated_report_top50 = pd.concat([updated_report_top50, temp], ignore_index=True)

updated_report_top50['Top 50 Clients'] = updated_report_top50['Client No'].map(contract_toTop50)
output_path = r'finaltop50.xlsx'
updated_report_top50.to_excel(output_path, index=False)

Filtered_T50 = updated_report_top50[updated_report_top50['Top 50 Clients'] == "TOP 50"]
print(f'Filtered Top 50 shape: {Filtered_T50.head()}')

# merge Filtered T50 with df_district to get ID number (fall back to df_district mapping only)
contract_to_id = dict(zip(df_district["CLM_CODE"], df_district["CLM_IDNO"]))
Filtered_T50['NIC/Company Registration No.'] = Filtered_T50['Client No'].map(contract_to_id)

#merge Filtered T50 with df monthly report to get Product Type
contract_to_prod_type = dict(zip(df_monthly_report["Contract No"], df_monthly_report["Product Type"]))
Filtered_T50['Type of Facility'] = Filtered_T50['Contract No'].map(contract_to_prod_type)
print(f'Filtered Top 50 shape: {Filtered_T50.head()}')

datasett50 = Filtered_T50[['Contract No', 'Client No', 'Customer Name', 'NIC/Company Registration No.', 'Type of Facility']]
# Write to ws_fdl C6 Working Working Sheet staring from A1
if 'C6 Working' in wb.sheetnames: 
    ws_c6 = wb['C6 Working']
    
    # Clear contents from columns A to E, from row 1 onwards
    for row in range(1, ws_c6.max_row + 1):
        for col in range(1, 6):  # Columns A to E
            ws_c6.cell(row=row, column=col).value = None
    
    # Paste Filtered_T50 to columns A to E with headings
    filtered_clean = datasett50.replace([np.inf, -np.inf], np.nan).fillna('')
    
    # Write the headings in row 1 (column names)
    for j, col_name in enumerate(filtered_clean.columns[:5], start=1):
        ws_c6.cell(row=1, column=j).value = col_name
    
    # Write the data from row 2 onwards
    for i, row in enumerate(filtered_clean.values, start=2):  # Start from row 2 to leave room for headings
        for j, val in enumerate(row[:5], start=1):  # Only write columns A to E
            cell = ws_c6.cell(row=i, column=j)
            if isinstance(val, str) and val.startswith('='): 
                cell.value = val  # openpyxl auto-detects formulas
            else:
                cell.value = val.item() if isinstance(val, (np.integer, np.floating)) else val
    
    print("Updated C6 Working sheet with Filtered_T50")

    # Check each row: if column E is empty but column B has data, write "Margin Trading Loans" to column E
    rows_updated = 0
    for row in range(2, ws_c6.max_row + 1):  # Start from row 2 to skip header
        col_b_value = ws_c6.cell(row=row, column=2).value  # Column B
        col_e_value = ws_c6.cell(row=row, column=5).value  # Column E

        # If column B has data but column E is empty, write "Margin Trading Loans" to column E
        if col_b_value is not None and col_b_value != '':
            if col_e_value is None or col_e_value == '':
                ws_c6.cell(row=row, column=5).value = "Margin Trading Loans"
                rows_updated += 1
        else:
            # Stop when column B has no data
            break

    print(f"Added 'Margin Trading Loans' to column E for {rows_updated} rows")

# Copy columns from C6 Working to NBD-QF-23-C6 sorted worksheet with values only
if 'NBD-QF-23-C6 sorted' in wb.sheetnames:
    ws_c6_sorted = wb['NBD-QF-23-C6 sorted']

    # Save the workbook first before reloading with data_only
    wb.save(output_report)

    # Load the workbook again with data_only=True to get calculated values from formulas
    wb_data_only = load_workbook(output_report, data_only=True)
    ws_c6_data_only = wb_data_only['C6 Working']

    # Find the end of data in C6 Working sheet (column B as reference)
    last_data_row = 1
    for row in range(2, ws_c6_data_only.max_row + 1):
        if ws_c6_data_only.cell(row=row, column=2).value is not None and ws_c6_data_only.cell(row=row, column=2).value != '':
            last_data_row = row
        else:
            break

    # Clear existing data in NBD-QF-23-C6 sorted from A6 to C(end of data) without deleting rows
    if ws_c6_sorted.max_row >= 6:
        for row in range(6, ws_c6_sorted.max_row + 1):
            for col in range(1, 4):  # Columns A to C
                ws_c6_sorted.cell(row=row, column=col).value = None

    # Column mapping: C6 Working -> NBD-QF-23-C6 sorted
    # A -> A, C -> B, D -> C
    column_mapping = {
        1: 1,   # A to A
        3: 2,   # C to B
        4: 3,   # D to C
    }

    # Copy data from row 2 to last_data_row in C6 Working to row 6 onwards in NBD-QF-23-C6 sorted (values only)
    for i, source_row in enumerate(range(2, last_data_row + 1), start=6):
        for source_col, dest_col in column_mapping.items():
            value = ws_c6_data_only.cell(row=source_row, column=source_col).value
            # Convert numpy types to Python native types
            if isinstance(value, (np.integer, np.floating)):
                value = value.item()
            ws_c6_sorted.cell(row=i, column=dest_col).value = value

    wb_data_only.close()
    print(f"Copied data from C6 Working to NBD-QF-23-C6 sorted ({last_data_row - 1} rows)")
else:
    print("Warning: 'NBD-QF-23-C6 sorted' sheet not found in workbook")

wb.save(output_report)


