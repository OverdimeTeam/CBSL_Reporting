import os
from pathlib import Path
import openpyxl
import sys

def get_month_year_from_filename(filename):
    # Example: "NBD-MF-20-C1 to C6 July 2025.xlsx"
    parts = filename.split()
    for i, part in enumerate(parts):
        if part in {"Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","January","February","March","April","May","June","July","August","September","October","November","December"}:
            month = part
            year = parts[i+1].replace(".xlsx","")
            return month, year
    return None, None

def find_files_safely(base_dir, pattern, description):
    """Safely find files and provide detailed error messages"""
    print(f"Looking for {description} in: {base_dir}")
    print(f"Search pattern: {pattern}")
    
    files = list(base_dir.glob(pattern))
    print(f"Found {len(files)} files matching pattern:")
    
    if not files:
        print(f"❌ No files found matching pattern: {pattern}")
        print("Available files in directory:")
        try:
            all_files = [f.name for f in base_dir.iterdir() if f.is_file() and f.suffix == '.xlsx']
            if all_files:
                for file in sorted(all_files):
                    print(f"  - {file}")
            else:
                print("  No .xlsx files found in directory")
        except Exception as e:
            print(f"  Error listing files: {e}")
        return None
    
    for file in files:
        print(f"  ✅ {file.name}")
    
    return files[0]  # Return the first match

def main():
    # Get the base directory (parent of report_automations)
    base_dir = Path(__file__).resolve().parent.parent
    
    # Based on your file structure, input files are in working/monthly/NBD-MF-20-C1 to C6 month 2025/
    working_dir = base_dir / "working"
    monthly_working_dir = working_dir / "monthly"
    outputs_monthly_dir = base_dir / "outputs" / "monthly"
    
    print(f"Base directory: {base_dir}")
    print(f"Working directory: {working_dir}")
    print(f"Monthly working directory: {monthly_working_dir}")
    print(f"Current working directory: {Path.cwd()}")
    
    # Create outputs monthly directory if it doesn't exist
    outputs_monthly_dir.mkdir(parents=True, exist_ok=True)
    
    # Search in the working/monthly directory first
    search_dirs = [
        monthly_working_dir,
        working_dir,
        base_dir
    ]
    
    file_c1_c6 = None
    file_afl = None
    
    # Search for C1-C6 file in multiple locations
    for search_dir in search_dirs:
        if search_dir.exists():
            print(f"\nSearching in: {search_dir}")
            
            # Look for the specific file you have
            c1_c6_pattern = "**/NBD-MF-20-C1 to C6*.xlsx"
            files = list(search_dir.glob(c1_c6_pattern))
            
            if files:
                file_c1_c6 = files[0]
                print(f"✅ Found C1-C6 file: {file_c1_c6}")
                break
    
    if file_c1_c6 is None:
        print("\n❌ Could not find C1 to C6 file. Looking for any similar files...")
        for search_dir in search_dirs:
            if search_dir.exists():
                xlsx_files = list(search_dir.rglob("*.xlsx"))
                if xlsx_files:
                    print(f"Excel files in {search_dir}:")
                    for f in xlsx_files:
                        print(f"  - {f.relative_to(base_dir)}")
        return
    
    # Search for AFL file
    for search_dir in search_dirs:
        if search_dir.exists():
            afl_pattern = "**/NBD-MF-01-SOFP & SOCI AFL Monthly FS*.xlsx"
            files = list(search_dir.glob(afl_pattern))
            
            if files:
                file_afl = files[0]
                print(f"✅ Found AFL file: {file_afl}")
                break
    
    if file_afl is None:
        print("\n❌ Could not find AFL file in any search directory")
        print("Please ensure the AFL file exists in one of these locations:")
        for search_dir in search_dirs:
            print(f"  - {search_dir}")
        return
    
    # Extract month and year from filename
    month, year = get_month_year_from_filename(file_c1_c6.name)
    if not month or not year:
        print(f"❌ Could not parse month/year from filename: {file_c1_c6.name}")
        print("Expected format: 'NBD-MF-20-C1 to C6 [Month] [Year].xlsx'")
        return
    
    print(f"\n✅ Parsed month: {month}, year: {year}")
    
    # Create output folder in outputs/monthly
    out_folder = outputs_monthly_dir / f"{year}_{month}"
    out_folder.mkdir(parents=True, exist_ok=True)
    print(f"Output folder: {out_folder}")
    
    try:
        # Load workbooks
        print(f"\nLoading workbooks...")
        wb_c1_c6 = openpyxl.load_workbook(file_c1_c6, data_only=True)
        wb_afl = openpyxl.load_workbook(file_afl, data_only=True)
        print(f"✅ Workbooks loaded successfully")
        
        # Check if required sheets exist
        required_sheets = {
            'c1_c6': ["3 Year Summary", "NBD_NBL-MF-20-C6"],
            'afl': ["NBD-MF-02-SOCI"]
        }
        
        print(f"\nAvailable sheets in C1-C6 file: {wb_c1_c6.sheetnames}")
        print(f"Available sheets in AFL file: {wb_afl.sheetnames}")
        
        # Validate required sheets
        for sheet in required_sheets['c1_c6']:
            if sheet not in wb_c1_c6.sheetnames:
                print(f"❌ Required sheet '{sheet}' not found in C1-C6 file")
                print(f"Available sheets: {wb_c1_c6.sheetnames}")
                return
        
        for sheet in required_sheets['afl']:
            if sheet not in wb_afl.sheetnames:
                print(f"❌ Required sheet '{sheet}' not found in AFL file")
                print(f"Available sheets: {wb_afl.sheetnames}")
                return
        
        print(f"✅ All required sheets found")
        
        # Get worksheets
        src_ws = wb_afl["NBD-MF-02-SOCI"]
        tgt_ws = wb_c1_c6["3 Year Summary"]
        
        print(f"\nProcessing data...")
        
        # Helper function to safely get numeric values
        def get_numeric_value(worksheet, cell_ref, description=""):
            """Safely extract numeric value from a cell, handling text/None values"""
            try:
                value = worksheet[cell_ref].value
                print(f"{description} ({cell_ref}): {value}")
                
                if value is None:
                    return 0
                elif isinstance(value, (int, float)):
                    return value
                elif isinstance(value, str):
                    # Try to extract number from string if possible
                    try:
                        # Remove common text and try to parse number
                        clean_value = value.replace(',', '').strip()
                        return float(clean_value)
                    except ValueError:
                        print(f"  ⚠️ Warning: Cell {cell_ref} contains text '{value}' - using 0")
                        return 0
                else:
                    print(f"  ⚠️ Warning: Cell {cell_ref} contains unexpected type {type(value)} - using 0")
                    return 0
            except Exception as e:
                print(f"  ❌ Error reading {cell_ref}: {e} - using 0")
                return 0
        
        # 1. Interest Income
        interest_income = get_numeric_value(src_ws, "B10", "Interest Income")
        tgt_ws["B5"] = interest_income
        
        # 2. Interest Expenses
        interest_expenses = get_numeric_value(src_ws, "B11", "Interest Expenses")
        tgt_ws["B6"] = interest_expenses
        
        # 3. Non-interest Income
        net_fee_commission = get_numeric_value(src_ws, "B12", "Net Fee Commission")
        net_other_income = get_numeric_value(src_ws, "B13", "Net Other Income")
        non_interest_income = net_fee_commission + net_other_income
        print(f"Non-interest Income: {net_fee_commission} + {net_other_income} = {non_interest_income}")
        tgt_ws["B7"] = non_interest_income
        
        # 4. Realized profit/losses from sale of securities
        realized_profit_loss = get_numeric_value(src_ws, "B14", "Realized profit/loss")
        tgt_ws["B8"] = realized_profit_loss
        
        # 5. Extraordinary /Irregular Item of Income /Expenses
        print(f"\nExtraordinary items:")
        real_estate_gains = get_numeric_value(src_ws, "B15", "Gains/(Losses) on sale of Real estates")
        investment_property_gains = get_numeric_value(src_ws, "B16", "Gains/(Losses) on sale of Investment Properties")
        ppe_gains = get_numeric_value(src_ws, "B17", "Gains/(Losses) on sale of property, plant & Equipment")
        revaluation_gains = get_numeric_value(src_ws, "B18", "Gains/(Losses) on revaluation of Investment properties")
        
        extra_items = real_estate_gains + investment_property_gains + ppe_gains + revaluation_gains
        print(f"Total extraordinary items: {real_estate_gains} + {investment_property_gains} + {ppe_gains} + {revaluation_gains} = {extra_items}")
        tgt_ws["B9"] = extra_items
        

        
        # Update column headers and item names in C6 sheet
        tgt_c6_ws = wb_c1_c6["NBD_NBL-MF-20-C6"]
        print(f"Updating column headers and item names in C6 sheet...")
        
        # Change 3rd column (C) header to "1st Year"
        tgt_c6_ws["C1"] = "1st Year"
        
        # Change 4th column (D) header to "2nd Year"
        tgt_c6_ws["D1"] = "2nd Year"
        
        # Update item names in column B
        item_names = {
            "B6": "Total Operating Income",
            "B7": "Interest Income from Loans & Advances", 
            "B8": "Interest Expenses on Deposits",
            "B9": "Fee & Commission Income",
            "B10": "Trading & Investment Income",
            "B11": "Other Operating Income"
        }
        
        for cell, new_name in item_names.items():
            tgt_c6_ws[cell] = new_name
            print(f"Updated {cell}: {new_name}")
        
        print(f"✅ Column headers updated: C1='1st Year', D1='2nd Year'")
        print(f"✅ Item names updated in column B")
        print(f"✅ Data processing completed (Interest Income excluded from C6 sheet)")
        
        # Save updated workbook
        out_path = out_folder / f"NBD_MF_20_C1_C6_{month}_{year}_report.xlsx"
        wb_c1_c6.save(out_path)
        print(f"\n✅ Report saved successfully to: {out_path}")
        
    except FileNotFoundError as e:
        print(f"❌ File not found: {e}")
    except PermissionError as e:
        print(f"❌ Permission error (file might be open in Excel): {e}")
    except Exception as e:
        print(f"❌ An error occurred: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()