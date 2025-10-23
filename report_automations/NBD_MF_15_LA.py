#!/usr/bin/env python3
"""
AFL Liquidity Report Automation Script - Date-Driven Workflow
Central Bank Sri Lanka - NBD_MF_15_LA

This script automates the complete AFL Liquidity report generation workflow:

USAGE:
python NBD_MF_15_LA.py 10/07/2025
python NBD_MF_15_LA.py 07/10/2025 --no-bots

INTEGRATED BOT EXECUTION:
0a. Run TB from ERP bot (Selenium) - downloads TB-COA file for specified date (retries indefinitely until downloaded)
0b. Run M2M from SASIA bot (Selenium) - downloads M2M file to Input folder

FILE MANAGEMENT LOGIC:
- FIXED Master Data Folder: working/weekly/07-06-2025/NBD_MF_04_LA/Input
  * ALL source files are ALWAYS read from here (TB-COA, Loan Schedule, Bank Balances, etc.)
  * Bot downloads go here

- Starting Template: working/weekly/07-06-2025/NBD_MF_04_LA/Inputs/AFL Liquidity - Week_27.xlsx
  * Used only on first run if no other AFL files exist

- Latest AFL Detection:
  * Searches ALL date folders for most recent AFL Liquidity file (by modification time)
  * Uses latest AFL as working file for updates

- Versioned Output Folders: working/weekly/{MM-DD-YYYY}/NBD_MF_04_LA/Inputs
  * Creates new versioned folder if date folder exists: 07-13-2025, 07-13-2025(2), 07-13-2025(3), etc.
  * Saves updated AFL file to the new versioned folder

EXAMPLE WORKFLOW:
Run 1 (07-06-2025): Use Week_27.xlsx → Save to 07-06-2025/NBD_MF_04_LA/Inputs
Run 2 (07-06-2025): Use AFL from 07-06-2025 → Save to 07-06-2025(2)/NBD_MF_04_LA/Inputs
Run 3 (07-13-2025): Use latest AFL from 07-06-2025(2) → Save to 07-13-2025/NBD_MF_04_LA/Inputs
Run 4 (07-20-2025): Use latest AFL from 07-13-2025 → Save to 07-20-2025/NBD_MF_04_LA/Inputs

DATE UPDATES IN FILE:
- Creates: AFL Liquidity - Week_XX.xlsx (week number calculated from date)
- Updates all date cells inside the file:
  * Treasury Input C1: Oct-07
  * Liquid Assets D1: Oct-07
  * Borrowings D1: Oct-07
  * NBD-WF-15-LA C3: Oct 07, 2025
  * ALCO C1, D1, E1: 07-Oct-25

DATA PROCESSING (ALL FILES REQUIRED):
1. Run TB bot (retries until TB-COA downloaded to FIXED Input folder)
2. Find latest AFL file across all folders (or use starting template)
3. Copy TB-COA data to working file's TB sheet (REQUIRED)
4. Extract M2M values and update Treasury Input C23, C24, C25 (OPTIONAL)
5. Extract Loan Summary values from Loan Schedule file (REQUIRED)
6. Update working file's Treasury Input sheet with Loan Summary values (REQUIRED)
7. Extract bank balances from Daily Bank Balances file (for specified date) (REQUIRED)
8. Update working file's Treasury Input sheet with bank balances (REQUIRED)
9. Extract Deposit Liability data and update working file (REQUIRED)

VALIDATION & OUTPUT:
10. Validate Liquid Assets sheet (C93 and D93 must be 0)
11. Create new AFL Liquidity - Week_XX.xlsx file with updated dates (only if validation passes)
12. Generate exceptions report if validation fails
13. Trigger monthly report generation if 5+ weekly files exist

MONTHLY REPORT GENERATION (AUTO-TRIGGERED WHEN 5 AFL FILES EXIST):
1. Find 5 most recent AFL Liquidity files across all folders
2. Copy NBD-WF-15-LA sheet A3:C25 from each AFL file to NBD-MF-04-LA Week 1-5 sheets
3. Fill NBD-MF-04-LA sheet columns C-H with H column data from Week 1-5 sheets:
   - Column C ← Week 1 H column
   - Column D ← Week 2 H column
   - Column E ← Week 3 H column
   - Column F ← Week 4 H column
   - Column G ← Week 5 H column
   - Column H ← Week 5 H column (duplicate)
4. Save updated NBD-MF-04-LA file to template location
5. Copy completed monthly report to last week's folder (e.g., 07-31-2025/NBD_MF_04_LA/)

COMMAND-LINE OPTIONS:
date              : Report date in MM/DD/YYYY or DD/MM/YYYY format (default: yesterday)
--no-bots         : Skip Selenium bots and use existing files
"""

import os
import sys
from datetime import datetime, timedelta
import openpyxl
from openpyxl.utils import get_column_letter
import glob
import logging
from pathlib import Path
import shutil
import subprocess
import time
from dotenv import load_dotenv

# Setup logging with UTF-8 encoding to avoid Windows console errors
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('afl_liquidity_automation.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

class AFLLiquidityAutomation:
    def __init__(self, base_dir=".", input_date=None, run_bots=True):
        """
        Initialize the AFL Liquidity automation script

        Args:
            base_dir (str): Base directory path
            input_date (datetime): Date for report generation (defaults to yesterday)
            run_bots (bool): Whether to run Selenium bots at startup (default: True)
        """
        self.base_dir = Path(base_dir)

        # Set date - if no date provided, use yesterday's date
        if input_date is None:
            self.report_date = datetime.now() - timedelta(days=1)
        else:
            self.report_date = input_date

        self.run_bots = run_bots

        # Determine week number for the current month
        self.week_of_month = self.get_week_of_month(self.report_date)

        # Calculate ISO week number for the year (for Deposit Liability)
        self.week_of_year = self.report_date.isocalendar()[1]

        # Create date-based folder name for outputs
        self.date_folder_name = self.report_date.strftime("%m-%d-%Y")

        # Extract month and year names for file patterns
        self.month_name = self.report_date.strftime("%B")  # e.g., "August"
        self.year = self.report_date.strftime("%Y")  # e.g., "2025"
        self.mmdd = self.report_date.strftime("%m%d")  # e.g., "0809"

        # WORKING folders - NEW STRUCTURE (no weekly folders)
        # When run from app.py, working directory is already set to the date folder
        # Use current working directory as the base
        self.working_folder = Path.cwd()

        # Input subfolder (singular) for bot downloads and master data files
        self.input_folder = self.working_folder / "Input"

        # Inputs subfolder (plural) for AFL files
        self.output_inputs_folder = self.working_folder / "Inputs"

        # Bots folder
        self.bots_folder = self.base_dir / "bots"

        # Create folder structure
        self.setup_folder_structure()

        logger.info(f"Initialized AFL Liquidity Automation for date: {self.report_date.strftime('%Y-%m-%d')}")
        logger.info(f"Week of month: {self.week_of_month}")
        logger.info(f"FIXED Input folder (master data - read from): {self.input_folder}")
        logger.info(f"DYNAMIC Output folder (AFL files - write to): {self.output_inputs_folder}")
        logger.info(f"Bots folder: {self.bots_folder}")

        # Verify folders exist
        if not self.input_folder.exists():
            logger.error(f"Input folder does not exist: {self.input_folder}")
        else:
            logger.info("✓ Input folder found successfully")
            files = list(self.input_folder.glob("*"))
            logger.info(f"Files in Input folder:")
            for f in files:
                logger.info(f"  - {f.name}")

        if not self.output_inputs_folder.exists():
            logger.warning(f"Output Inputs folder does not exist yet: {self.output_inputs_folder}")
        else:
            logger.info("✓ Output Inputs folder found")
            files = list(self.output_inputs_folder.glob("*"))
            logger.info(f"Files in Output Inputs folder:")
            for f in files:
                logger.info(f"  - {f.name}")

        # File paths will be determined dynamically
        self.afl_liquidity_file = None
        self.working_file = None  # This will be the latest test file or AFL Liquidity file
        self.loan_schedule_file = None
        self.tb_coa_file = None
        self.daily_bank_balance_file = None
        self.deposit_liability_file = None
        self.m2m_file = None  # M2M file for Treasury Input C23, C24, C25

    def get_week_of_month(self, date):
        """
        Calculate which week of the month this date falls into (1-5)
        """
        first_day = date.replace(day=1)
        # Calculate the week number within the month
        day_of_month = date.day
        week_number = ((day_of_month - 1) // 7) + 1
        return min(week_number, 5)  # Cap at 5 weeks

    def get_versioned_folder_path(self, base_folder_path):
        """
        Get a versioned folder path if the folder already exists
        Returns: Path object for the folder (either original or versioned)

        Examples:
        - 07-06-2025 -> 07-06-2025 (if doesn't exist)
        - 07-06-2025 -> 07-06-2025(2) (if exists)
        - 07-06-2025 -> 07-06-2025(3) (if (2) also exists)
        """
        try:
            original_path = Path(base_folder_path)

            # If folder doesn't exist, use it as-is
            if not original_path.exists():
                return original_path

            # Folder exists, find next available version number
            version = 2
            while True:
                versioned_path = Path(f"{base_folder_path}({version})")
                if not versioned_path.exists():
                    logger.info(f"Folder {original_path.name} exists, creating versioned folder: {versioned_path.name}")
                    return versioned_path
                version += 1

        except Exception as e:
            logger.error(f"Error getting versioned folder path: {e}")
            return Path(base_folder_path)

    def setup_folder_structure(self):
        """
        Create folder structure for weekly reports (NEW STRUCTURE - no weekly folders)
        Working folder is already set to working/NBD_MF_15_LA/{DD-MM-YYYY} by app.py
        Structure: working/NBD_MF_15_LA/{DD-MM-YYYY}/
                   ├── Input/  (bot downloads and master data)
                   └── Inputs/ (AFL files)
        """
        try:
            logger.info(f"Setting up folder structure (NEW - simplified)...")

            # Ensure working folder exists
            self.working_folder.mkdir(parents=True, exist_ok=True)
            logger.info(f"✓ Working folder: {self.working_folder}")

            # Ensure Input subfolder exists (singular - for bot downloads and master data)
            self.input_folder.mkdir(parents=True, exist_ok=True)
            logger.info(f"✓ Input folder: {self.input_folder}")

            # Ensure Inputs subfolder exists (plural - for AFL files)
            self.output_inputs_folder.mkdir(parents=True, exist_ok=True)
            logger.info(f"✓ Inputs folder (AFL files): {self.output_inputs_folder}")

            logger.info("Folder structure setup completed successfully")

        except Exception as e:
            logger.error(f"Error setting up folder structure: {e}")
            import traceback
            logger.error(traceback.format_exc())

    def get_5_most_recent_afl_files(self):
        """
        Find the 5 most recent AFL Liquidity files from outputs folder (completed weeks)
        Returns: List of file paths sorted by modification time (oldest to newest)
        """
        try:
            logger.info("Searching for 5 most recent AFL Liquidity files in outputs...")

            all_afl_files = []

            # NEW STRUCTURE: Search in outputs folder for completed date folders
            outputs_dir = self.base_dir / "outputs" / "NBD_MF_15_LA"

            if not outputs_dir.exists():
                logger.warning(f"Outputs directory not found: {outputs_dir}")
                return []

            # Search through all date folders in outputs directory
            for date_folder in outputs_dir.glob("*"):
                if not date_folder.is_dir():
                    continue

                # Find AFL Liquidity files directly in date folder (new structure)
                afl_files = list(date_folder.glob("AFL Liquidity - Week_*.xlsx"))
                for afl_file in afl_files:
                    all_afl_files.append(afl_file)

            if len(all_afl_files) < 5:
                logger.warning(f"Only found {len(all_afl_files)} AFL files in outputs. Need 5 for monthly report.")
                return []

            # Sort by modification time (oldest to newest) and take the 5 most recent
            all_afl_files.sort(key=lambda f: f.stat().st_mtime)
            five_most_recent = all_afl_files[-5:]  # Get last 5 (most recent)

            logger.info(f"✓ Found 5 most recent AFL files (in chronological order):")
            for idx, afl_file in enumerate(five_most_recent, 1):
                mod_time = datetime.fromtimestamp(afl_file.stat().st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                logger.info(f"  {idx}. {afl_file.parent.name}/{afl_file.name} (modified: {mod_time})")

            return five_most_recent

        except Exception as e:
            logger.error(f"Error finding 5 most recent AFL files: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return []

    def get_monthly_afl_files(self):
        """
        Find all AFL Liquidity files for the current month from outputs folder
        Returns: List of tuples (week_number, file_path, date)
        """
        try:
            current_month = self.report_date.month
            current_year = self.report_date.year

            afl_files = []

            # NEW STRUCTURE: Search in outputs folder for completed date folders
            outputs_dir = self.base_dir / "outputs" / "NBD_MF_15_LA"

            if not outputs_dir.exists():
                logger.warning(f"Outputs directory not found: {outputs_dir}")
                return []

            # Search through all date folders in outputs directory
            for date_folder in outputs_dir.glob("*"):
                if not date_folder.is_dir():
                    continue

                try:
                    # Parse folder name (DD-MM-YYYY or DD-MM-YYYY(n))
                    folder_name = date_folder.name
                    # Remove version suffix if present
                    if "(" in folder_name:
                        folder_name = folder_name.split("(")[0]

                    folder_date = datetime.strptime(folder_name, "%d-%m-%Y")

                    # Check if it's in the same month and year
                    if folder_date.month == current_month and folder_date.year == current_year:
                        # Calculate week of month for this folder
                        week_num = self.get_week_of_month(folder_date)

                        # Look for AFL file directly in date folder (new structure)
                        afl_pattern = list(date_folder.glob("AFL Liquidity - Week_*.xlsx"))
                        if afl_pattern:
                            afl_files.append((week_num, afl_pattern[0], folder_date))
                            logger.info(f"Found Week {week_num} AFL file: {afl_pattern[0].name}")
                except ValueError:
                    # Skip folders that don't match date format
                    continue

            # Sort by week number
            afl_files.sort(key=lambda x: x[0])

            return afl_files

        except Exception as e:
            logger.error(f"Error finding monthly AFL files: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return []

    def copy_cells_with_formatting(self, source_sheet, target_sheet, source_range, target_start_cell):
        """
        Copy cells from source to target with all formatting, formulas, and styles

        Args:
            source_sheet: Source worksheet object
            target_sheet: Target worksheet object
            source_range: String like "A3:C25"
            target_start_cell: String like "A1"
        """
        try:
            from openpyxl.utils import range_boundaries
            from copy import copy

            # Parse source range
            min_col, min_row, max_col, max_row = range_boundaries(source_range)

            # Parse target start cell
            target_min_col, target_min_row, _, _ = range_boundaries(target_start_cell)

            # Calculate offsets
            row_offset = target_min_row - min_row
            col_offset = target_min_col - min_col

            logger.info(f"  Copying {source_range} → starting at {target_start_cell}")

            # Copy cells with all properties
            for row_idx in range(min_row, max_row + 1):
                for col_idx in range(min_col, max_col + 1):
                    source_cell = source_sheet.cell(row=row_idx, column=col_idx)
                    target_cell = target_sheet.cell(row=row_idx + row_offset, column=col_idx + col_offset)

                    # Copy value
                    target_cell.value = source_cell.value

                    # Copy formatting
                    if source_cell.has_style:
                        target_cell.font = copy(source_cell.font)
                        target_cell.border = copy(source_cell.border)
                        target_cell.fill = copy(source_cell.fill)
                        target_cell.number_format = copy(source_cell.number_format)
                        target_cell.protection = copy(source_cell.protection)
                        target_cell.alignment = copy(source_cell.alignment)

            # Copy row heights
            for row_idx in range(min_row, max_row + 1):
                if source_sheet.row_dimensions[row_idx].height:
                    target_sheet.row_dimensions[row_idx + row_offset].height = source_sheet.row_dimensions[row_idx].height

            # Copy column widths
            for col_idx in range(min_col, max_col + 1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                target_col_letter = openpyxl.utils.get_column_letter(col_idx + col_offset)
                if source_sheet.column_dimensions[col_letter].width:
                    target_sheet.column_dimensions[target_col_letter].width = source_sheet.column_dimensions[col_letter].width

            logger.info(f"  ✓ Copied {(max_row - min_row + 1)} rows x {(max_col - min_col + 1)} columns")
            return True

        except Exception as e:
            logger.error(f"Error copying cells with formatting: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False

    def fill_nbd_mf_04_la_sheet(self, monthly_wb):
        """
        Fill NBD-MF-04-LA sheet columns C-H with H column values from Week 1-5 sheets

        Mapping:
        - C5:C15 ← Week 1 H3:H13
        - C19 ← Week 1 H14
        - C22:C27 ← Week 1 H15:H20
        - Same pattern for D (Week 2), E (Week 3), F (Week 4), G (Week 5), H (Week 5 duplicate)

        Args:
            monthly_wb: The workbook with Week sheets already populated
        """
        try:
            logger.info("\n--- Filling NBD-MF-04-LA Sheet ---")

            # Check if NBD-MF-04-LA sheet exists
            if "NBD-MF-04-LA" not in monthly_wb.sheetnames:
                logger.error("Sheet 'NBD-MF-04-LA' not found in monthly report")
                logger.info(f"Available sheets: {monthly_wb.sheetnames}")
                return False

            nbd_sheet = monthly_wb["NBD-MF-04-LA"]

            # Column mapping: C=Week1, D=Week2, E=Week3, F=Week4, G=Week5, H=Week5
            week_to_column = {
                1: 'C',  # Week 1 → Column C
                2: 'D',  # Week 2 → Column D
                3: 'E',  # Week 3 → Column E
                4: 'F',  # Week 4 → Column F
                5: 'G',  # Week 5 → Column G
            }

            # Process weeks 1-5
            for week_num in range(1, 6):
                week_sheet_name = f"Week {week_num}"

                if week_sheet_name not in monthly_wb.sheetnames:
                    logger.warning(f"Sheet '{week_sheet_name}' not found, skipping...")
                    continue

                week_sheet = monthly_wb[week_sheet_name]
                dest_col = week_to_column[week_num]

                logger.info(f"  Processing Week {week_num} → Column {dest_col}...")

                # Copy H3:H13 to C5:C15 (11 cells)
                for i in range(11):
                    source_row = 3 + i  # H3 to H13
                    dest_row = 5 + i    # C5 to C15
                    source_value = week_sheet[f"H{source_row}"].value
                    nbd_sheet[f"{dest_col}{dest_row}"].value = source_value

                # Copy H14 to C19 (1 cell)
                source_value = week_sheet["H14"].value
                nbd_sheet[f"{dest_col}19"].value = source_value

                # Copy H15:H20 to C22:C27 (6 cells)
                for i in range(6):
                    source_row = 15 + i  # H15 to H20
                    dest_row = 22 + i    # C22 to C27
                    source_value = week_sheet[f"H{source_row}"].value
                    nbd_sheet[f"{dest_col}{dest_row}"].value = source_value

                logger.info(f"    ✓ Week {week_num} data copied to column {dest_col}")

            # Duplicate Week 5 (column G) to column H
            logger.info("  Duplicating Week 5 data from column G to column H...")

            # Copy G5:G15 to H5:H15
            for i in range(11):
                row_num = 5 + i
                nbd_sheet[f"H{row_num}"].value = nbd_sheet[f"G{row_num}"].value

            # Copy G19 to H19
            nbd_sheet["H19"].value = nbd_sheet["G19"].value

            # Copy G22:G27 to H22:H27
            for i in range(6):
                row_num = 22 + i
                nbd_sheet[f"H{row_num}"].value = nbd_sheet[f"G{row_num}"].value

            logger.info("    ✓ Column H filled with Week 5 data")
            logger.info("✓ NBD-MF-04-LA sheet filled successfully")

            return True

        except Exception as e:
            logger.error(f"Error filling NBD-MF-04-LA sheet: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False

    def generate_monthly_report_from_5_afl_files(self):
        """
        Generate monthly report by copying data from 5 most recent AFL files
        to NBD-MF-04-LA Liquid Assets file

        Process:
        1. Find 5 most recent AFL files
        2. For each AFL file, copy NBD-WF-15-LA sheet A3:C25
        3. Paste into NBD-MF-04-LA file Week 1, Week 2, Week 3, Week 4, Week 5 sheets
        4. Fill NBD-MF-04-LA sheet with data from Week sheets
        5. Save and copy to last week's folder

        Returns: True if successful, False otherwise
        """
        monthly_wb = None
        try:
            logger.info("\n" + "=" * 60)
            logger.info("MONTHLY REPORT GENERATION - THE CELL PASTING GAME")
            logger.info("=" * 60)

            # Step 1: Get 5 most recent AFL files
            five_afl_files = self.get_5_most_recent_afl_files()

            if len(five_afl_files) < 5:
                logger.warning("Cannot generate monthly report - need 5 AFL files")
                return False

            # Determine the last week's folder (5th AFL file's folder)
            last_afl_file = five_afl_files[-1]  # Most recent (5th file)
            last_week_folder = last_afl_file.parent.parent  # Go up to NBD_MF_04_LA folder

            logger.info(f"Last week folder: {last_week_folder.parent.name}/NBD_MF_04_LA")

            # Step 2: Path to monthly report template (NEW STRUCTURE - in current working folder)
            monthly_report_path = self.working_folder / "NBD-MF-04-LA Liquid Assets - Jul 25.xlsm"

            if not monthly_report_path.exists():
                logger.error(f"Monthly report template not found: {monthly_report_path}")
                return False

            logger.info(f"Opening monthly report template: {monthly_report_path.name}")

            # Step 3: Open the monthly report file
            monthly_wb = openpyxl.load_workbook(monthly_report_path, keep_vba=True)

            week_sheets = ["Week 1", "Week 2", "Week 3", "Week 4", "Week 5"]

            # Step 4: Process each of the 5 AFL files
            for idx, afl_file_path in enumerate(five_afl_files, 1):
                week_sheet_name = week_sheets[idx - 1]

                logger.info(f"\n--- Processing AFL File #{idx} ---")
                logger.info(f"Source: {afl_file_path.parent.parent.parent.name}/{afl_file_path.name}")
                logger.info(f"Target: {week_sheet_name} sheet")

                # Open AFL file
                afl_wb = openpyxl.load_workbook(afl_file_path, data_only=True)

                # Check if NBD-WF-15-LA sheet exists
                if "NBD-WF-15-LA" not in afl_wb.sheetnames:
                    logger.error(f"Sheet 'NBD-WF-15-LA' not found in {afl_file_path.name}")
                    logger.error(f"Available sheets: {afl_wb.sheetnames}")
                    afl_wb.close()
                    continue

                # Check if target week sheet exists
                if week_sheet_name not in monthly_wb.sheetnames:
                    logger.error(f"Sheet '{week_sheet_name}' not found in monthly report")
                    logger.error(f"Available sheets: {monthly_wb.sheetnames}")
                    afl_wb.close()
                    continue

                # Get source and target sheets
                source_sheet = afl_wb["NBD-WF-15-LA"]
                target_sheet = monthly_wb[week_sheet_name]

                # Copy A3:C25 from source to target (starting at A1)
                success = self.copy_cells_with_formatting(
                    source_sheet=source_sheet,
                    target_sheet=target_sheet,
                    source_range="A3:C25",
                    target_start_cell="A1"
                )

                afl_wb.close()

                if success:
                    logger.info(f"✓ Successfully copied data to {week_sheet_name}")
                else:
                    logger.error(f"✗ Failed to copy data to {week_sheet_name}")

            # Step 5: Fill NBD-MF-04-LA sheet with weekly data
            logger.info("\n" + "=" * 60)
            fill_success = self.fill_nbd_mf_04_la_sheet(monthly_wb)
            if not fill_success:
                logger.warning("Failed to fill NBD-MF-04-LA sheet, but continuing...")

            # Step 6: Save the monthly report to template location
            logger.info(f"\nSaving monthly report to template location...")
            monthly_wb.save(monthly_report_path)
            logger.info(f"✓ Saved: {monthly_report_path}")

            # Step 7: Copy the completed monthly report to last week's folder
            logger.info(f"\nCopying completed monthly report to last week's folder...")
            destination_path = last_week_folder / monthly_report_path.name

            # Close workbook before copying
            monthly_wb.close()
            monthly_wb = None

            # Copy file
            shutil.copy2(monthly_report_path, destination_path)
            logger.info(f"✓ Copied to: {last_week_folder.parent.name}/NBD_MF_04_LA/{destination_path.name}")

            logger.info("\n" + "=" * 60)
            logger.info("✓ MONTHLY REPORT GENERATED SUCCESSFULLY")
            logger.info(f"✓ Template updated: {monthly_report_path}")
            logger.info(f"✓ Copy saved to: {destination_path}")
            logger.info("=" * 60)

            return True

        except Exception as e:
            logger.error(f"Error generating monthly report: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False
        finally:
            if monthly_wb is not None:
                try:
                    monthly_wb.close()
                except:
                    pass

    def check_and_run_monthly_report(self):
        """
        Check if 5 AFL files exist and trigger monthly report generation
        Returns: True if monthly report was generated, False otherwise
        """
        try:
            logger.info("\n" + "=" * 60)
            logger.info("Checking for monthly report generation...")
            logger.info("=" * 60)

            # Check if 5 or more AFL files exist
            five_afl_files = self.get_5_most_recent_afl_files()

            if len(five_afl_files) < 5:
                logger.info(f"Only {len(five_afl_files)} AFL file(s) found. Need 5 for monthly report.")
                logger.info("Skipping monthly report generation.")
                return False

            logger.info("✓ Found 5 or more AFL files - triggering monthly report generation")

            # Generate the monthly report
            return self.generate_monthly_report_from_5_afl_files()

        except Exception as e:
            logger.error(f"Error checking monthly report: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False

    def run_monthly_report_script(self, afl_files):
        """
        Run the NBD_MF_04_LA.py script for monthly report generation

        Args:
            afl_files: List of tuples (week_number, file_path, date)
        """
        try:
            logger.info("\n" + "=" * 60)
            logger.info("Running Monthly Report Script (NBD_MF_04_LA.py)...")
            logger.info("=" * 60)

            monthly_script = self.base_dir / "report_automations" / "NBD_MF_04_LA.py"

            if not monthly_script.exists():
                logger.error(f"Monthly report script not found: {monthly_script}")
                return False

            # Import and run the monthly report script
            sys.path.insert(0, str(self.base_dir / "report_automations"))

            try:
                from NBD_MF_04_LA import MonthlyLiquidityReport

                # Create monthly report instance
                report = MonthlyLiquidityReport(
                    base_dir=str(self.base_dir),
                    report_date=self.report_date,
                    afl_files=afl_files
                )

                # Run the report generation
                success = report.generate_monthly_report()

                if success:
                    logger.info("✓ Monthly report generated successfully")
                    return True
                else:
                    logger.error("✗ Monthly report generation failed")
                    return False

            except ImportError as e:
                logger.error(f"Failed to import monthly report script: {e}")
                logger.info("Monthly report script may not exist yet - skipping")
                return False

        except Exception as e:
            logger.error(f"Error running monthly report script: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False
        finally:
            # Clean up sys.path
            if str(self.base_dir / "report_automations") in sys.path:
                sys.path.remove(str(self.base_dir / "report_automations"))

    
    def find_loan_schedule_file(self):
        """
        Find the Loan Schedule Excel file in Input folder (no 's')
        Pattern: Loan Schedule {Month} {Year}.xlsx (e.g., Loan Schedule August 2025.xlsx)
        """
        try:
            # Try specific pattern first: Loan Schedule {Month} {Year}.xlsx
            specific_pattern = f"Loan Schedule {self.month_name} {self.year}.xlsx"
            files = list(self.input_folder.glob(specific_pattern))

            # Fallback to generic pattern if specific not found
            if not files:
                logger.info(f"Specific pattern not found: {specific_pattern}, trying generic pattern...")
                generic_pattern = "Loan Schedule*.xlsx"
                files = list(self.input_folder.glob(generic_pattern))

            if not files:
                logger.error(f"No Loan Schedule file found in {self.input_folder}")
                logger.info(f"Expected: Loan Schedule {self.month_name} {self.year}.xlsx or Loan Schedule*.xlsx")
                return None

            self.loan_schedule_file = files[0]
            logger.info(f"Found Loan Schedule file: {files[0].name}")
            return files[0]

        except Exception as e:
            logger.error(f"Error finding Loan Schedule file: {e}")
            return None
    
    def find_tb_coa_file(self):
        """
        Find the TB-COA Excel file downloaded by the bot in Input folder (no 's')
        Looks for pattern like: "*Report*.xlsx" or "TB-COA*.xlsx"
        """
        try:
            # Try multiple patterns to find the TB-COA report
            patterns = [
                "*Report*.xlsx",
                "TB-COA*.xlsx",
                "*TB_COA*.xlsx"
            ]
            
            files = []
            for pattern in patterns:
                files = list(self.input_folder.glob(pattern))
                if files:
                    logger.info(f"Found TB-COA file(s) with pattern: {pattern}")
                    break
            
            if not files:
                logger.error(f"No TB-COA file found in {self.input_folder}")
                logger.info("Searched patterns: *Report*.xlsx, TB-COA*.xlsx, *TB_COA*.xlsx")
                return None
            
            # Use the first file found (or most recent if multiple)
            self.tb_coa_file = files[0]
            logger.info(f"Found TB-COA file: {files[0].name}")
            return files[0]
            
        except Exception as e:
            logger.error(f"Error finding TB-COA file: {e}")
            return None
    
    def find_daily_bank_balance_file(self):
        """
        Find the Daily Bank Balances Excel file in Input folder (no 's')
        Pattern: Daily Bank Balances - {Month} {Year}.xlsx (e.g., Daily Bank Balances - August 2025.xlsx)
        """
        try:
            # Try specific pattern first: Daily Bank Balances - {Month} {Year}.xlsx
            specific_pattern = f"Daily Bank Balances - {self.month_name} {self.year}.xlsx"
            files = list(self.input_folder.glob(specific_pattern))

            # Fallback to generic pattern if specific not found
            if not files:
                logger.info(f"Specific pattern not found: {specific_pattern}, trying generic pattern...")
                generic_pattern = "Daily Bank Balance*.xlsx"
                files = list(self.input_folder.glob(generic_pattern))

            if not files:
                logger.error(f"No Daily Bank Balances file found in {self.input_folder}")
                logger.info(f"Expected: Daily Bank Balances - {self.month_name} {self.year}.xlsx or Daily Bank Balance*.xlsx")
                return None

            self.daily_bank_balance_file = files[0]
            logger.info(f"Found Daily Bank Balances file: {files[0].name}")
            return files[0]
            
        except Exception as e:
            logger.error(f"Error finding Daily Bank Balances file: {e}")
            return None
    
    def find_deposit_liability_file(self):
        """
        Find the Deposit Liability Excel file in Input folder (no 's')
        Pattern: Deposit Liability Week {WeekNumber}.xlsx (e.g., Deposit Liability Week 37.xlsx)
        """
        try:
            # Try specific pattern first: Deposit Liability Week {WeekNumber}.xlsx
            specific_pattern = f"Deposit Liability Week {self.week_of_year}.xlsx"
            files = list(self.input_folder.glob(specific_pattern))

            # Fallback to generic pattern if specific not found
            if not files:
                logger.info(f"Specific pattern not found: {specific_pattern}, trying generic pattern...")
                generic_pattern = "Deposit Liability Week*.xlsx"
                files = list(self.input_folder.glob(generic_pattern))

            if not files:
                logger.error(f"No Deposit Liability file found in {self.input_folder}")
                logger.info(f"Expected: Deposit Liability Week {self.week_of_year}.xlsx or Deposit Liability Week*.xlsx")
                return None

            self.deposit_liability_file = files[0]
            logger.info(f"Found Deposit Liability file: {files[0].name}")
            return files[0]

        except Exception as e:
            logger.error(f"Error finding Deposit Liability file: {e}")
            return None

    def find_m2m_file(self):
        """
        Find the M2M Excel file in Input folder (no 's')
        Pattern: M2M {MMDD}.xlsx (e.g., M2M 0809.xlsx for August 09, 2025)
        """
        try:
            # Try specific pattern first: M2M {MMDD}.xlsx
            specific_pattern = f"M2M {self.mmdd}.xlsx"
            files = list(self.input_folder.glob(specific_pattern))

            # Fallback to generic pattern if specific not found
            if not files:
                logger.info(f"Specific pattern not found: {specific_pattern}, trying generic pattern...")
                generic_pattern = "M2M*.xlsx"
                files = list(self.input_folder.glob(generic_pattern))

            if not files:
                logger.error(f"No M2M file found in {self.input_folder}")
                logger.info(f"Expected: M2M {self.mmdd}.xlsx or M2M*.xlsx")
                return None

            self.m2m_file = files[0]
            logger.info(f"Found M2M file: {files[0].name}")
            return files[0]

        except Exception as e:
            logger.error(f"Error finding M2M file: {e}")
            return None

    def extract_m2m_values(self):
        """
        Extract values from M2M file (Sheet1) and return them

        Extracts:
        - AB46: For Treasury Input C25
        - L46: For Treasury Input C24
        - I46 + K46: For Treasury Input C23 (sum)

        Returns:
            tuple: (c23_value, c24_value, c25_value) or (None, None, None) on error
        """
        wb = None
        try:
            if not self.m2m_file:
                logger.error("M2M file not found")
                return None, None, None

            # Load the workbook
            logger.info(f"Opening M2M file: {self.m2m_file}")
            wb = openpyxl.load_workbook(self.m2m_file, data_only=True)

            # Check if Sheet1 exists
            if "Sheet1" not in wb.sheetnames:
                logger.error(f"Sheet 'Sheet1' not found in M2M file. Available sheets: {wb.sheetnames}")
                return None, None, None

            sheet = wb["Sheet1"]

            # Extract values from specified cells
            ab46_value = sheet["AB46"].value  # For C25
            l46_value = sheet["L46"].value    # For C24
            i46_value = sheet["I46"].value    # Part 1 for C23
            k46_value = sheet["K46"].value    # Part 2 for C23

            logger.info(f"Extracted M2M values:")
            logger.info(f"  AB46: {ab46_value} (for Treasury Input C25)")
            logger.info(f"  L46: {l46_value} (for Treasury Input C24)")
            logger.info(f"  I46: {i46_value} (for Treasury Input C23 - part 1)")
            logger.info(f"  K46: {k46_value} (for Treasury Input C23 - part 2)")

            # Calculate C23 value (I46 + K46)
            c23_value = None
            if i46_value is not None and k46_value is not None:
                try:
                    c23_value = float(i46_value) + float(k46_value)
                    logger.info(f"  Calculated C23 value: {i46_value} + {k46_value} = {c23_value}")
                except (TypeError, ValueError) as e:
                    logger.error(f"Could not calculate sum of I46 and K46: {e}")
            else:
                logger.warning("I46 or K46 is None, cannot calculate sum for C23")

            return c23_value, l46_value, ab46_value

        except Exception as e:
            logger.error(f"Error extracting M2M values: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None, None, None
        finally:
            if wb is not None:
                try:
                    wb.close()
                except:
                    pass

    def update_treasury_input_with_m2m(self, c23_value, c24_value, c25_value):
        """
        Update AFL Liquidity Treasury Input sheet with M2M values

        Args:
            c23_value: Value for C23 (I46 + K46 sum)
            c24_value: Value for C24 (L46)
            c25_value: Value for C25 (AB46)
        """
        wb = None
        try:
            if not self.afl_liquidity_file:
                logger.error("AFL Liquidity file not found")
                return False

            # Load the workbook
            logger.info(f"Opening AFL Liquidity file: {self.afl_liquidity_file}")
            wb = openpyxl.load_workbook(self.afl_liquidity_file)

            # Check if the required sheet exists
            if "Treasury Input" not in wb.sheetnames:
                logger.error(f"Sheet 'Treasury Input' not found. Available sheets: {wb.sheetnames}")
                return False

            sheet = wb["Treasury Input"]

            # Update cells with M2M values
            if c23_value is not None:
                sheet["C23"].value = c23_value
                logger.info(f"✓ Updated Treasury Input C23 with M2M value: {c23_value}")
            else:
                logger.warning("C23 value is None, skipping C23 update")

            if c24_value is not None:
                sheet["C24"].value = c24_value
                logger.info(f"✓ Updated Treasury Input C24 with M2M value: {c24_value}")
            else:
                logger.warning("C24 value is None, skipping C24 update")

            if c25_value is not None:
                sheet["C25"].value = c25_value
                logger.info(f"✓ Updated Treasury Input C25 with M2M value: {c25_value}")
            else:
                logger.warning("C25 value is None, skipping C25 update")

            # Save the workbook
            wb.save(self.afl_liquidity_file)
            wb.close()
            wb = None

            logger.info("✓ Treasury Input sheet updated successfully with M2M values")
            return True

        except Exception as e:
            logger.error(f"Error updating Treasury Input with M2M values: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False
        finally:
            if wb is not None:
                try:
                    wb.close()
                except:
                    pass
    
    def copy_tb_coa_data_with_excel_com(self):
        """
        Copy TB-COA data using Excel COM automation (Windows only)
        This bypasses openpyxl's issues with corrupted files
        """
        excel = None
        tb_wb = None
        afl_wb = None

        try:
            import win32com.client

            logger.info("Using Excel COM automation to handle corrupted file...")

            # Create Excel application
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            # Open TB-COA file
            logger.info(f"Opening TB-COA file with Excel: {self.tb_coa_file}")
            tb_wb = excel.Workbooks.Open(str(self.tb_coa_file.absolute()))

            # Find TB-COA sheet
            sheet_name = None
            for sheet in tb_wb.Sheets:
                if 'TB' in sheet.Name.upper() and 'COA' in sheet.Name.upper():
                    sheet_name = sheet.Name
                    break

            if not sheet_name:
                sheet_name = tb_wb.Sheets(1).Name
                logger.info(f"TB-COA sheet not found by name, using first sheet: {sheet_name}")

            tb_sheet = tb_wb.Sheets(sheet_name)
            logger.info(f"Reading sheet: {sheet_name}")

            # Get used range
            used_range = tb_sheet.UsedRange
            data = used_range.Value

            if data is None:
                logger.error("No data found in TB-COA sheet")
                return False

            # Convert to list if single row
            if not isinstance(data, tuple):
                data = [[data]]
            elif not isinstance(data[0], tuple):
                data = [list(data)]
            else:
                data = [list(row) if row else [] for row in data]

            num_rows = len(data)
            num_cols = len(data[0]) if data else 0
            logger.info(f"Read {num_rows} rows x {num_cols} columns from TB-COA")

            # Close TB workbook
            tb_wb.Close(False)
            tb_wb = None
            excel.Quit()
            excel = None

            # Now write to AFL Liquidity using openpyxl
            logger.info(f"Opening AFL Liquidity file: {self.afl_liquidity_file}")
            afl_wb = openpyxl.load_workbook(self.afl_liquidity_file)

            if "TB" not in afl_wb.sheetnames:
                logger.error(f"Sheet 'TB' not found in AFL Liquidity file. Available sheets: {afl_wb.sheetnames}")
                return False

            tb_dest_sheet = afl_wb["TB"]
            logger.info("TB sheet found in AFL Liquidity file")

            # Unmerge all cells first
            logger.info("Unmerging all cells in TB sheet...")
            merged_ranges = list(tb_dest_sheet.merged_cells.ranges)
            for merged_range in merged_ranges:
                tb_dest_sheet.unmerge_cells(str(merged_range))

            # Clear existing data
            logger.info("Clearing existing data in TB sheet...")
            for row in tb_dest_sheet.iter_rows():
                for cell in row:
                    cell.value = None

            # Write data
            logger.info("Writing data to TB sheet...")
            for row_idx, row_data in enumerate(data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    tb_dest_sheet.cell(row=row_idx, column=col_idx, value=value)

            logger.info(f"Wrote {num_rows} rows x {num_cols} columns to TB sheet")

            # Save
            afl_wb.save(self.afl_liquidity_file)
            afl_wb.close()
            afl_wb = None

            logger.info("TB-COA data copied successfully using Excel COM")
            return True

        except ImportError:
            logger.error("pywin32 not installed. Install with: pip install pywin32")
            return False
        except Exception as e:
            logger.error(f"Error copying TB-COA data with Excel COM: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False
        finally:
            # Ensure proper cleanup
            try:
                if afl_wb is not None:
                    afl_wb.close()
            except:
                pass

            try:
                if tb_wb is not None:
                    tb_wb.Close(False)
            except:
                pass

            try:
                if excel is not None:
                    excel.Quit()
            except:
                pass
    
    def copy_tb_coa_data(self):
        """
        Copy all data from TB-COA sheet to AFL Liquidity TB sheet
        """
        try:
            if not self.tb_coa_file:
                logger.error("TB-COA file not found")
                return False
            
            if not self.afl_liquidity_file:
                logger.error("AFL Liquidity file not found")
                return False
            
            # Try Excel COM automation (works on Windows)
            logger.info("Attempting to use Excel COM automation...")
            return self.copy_tb_coa_data_with_excel_com()
            
        except Exception as e:
            logger.error(f"Error copying TB-COA data: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False
    
    def extract_loan_summary_values(self):
        """
        Extract values from Loan Schedule file (in Input folder)
        Sheet: Loan Summary, Cells: I29 and J29
        """
        wb = None
        try:
            if not self.loan_schedule_file:
                logger.error("Loan Schedule file not found")
                return None, None

            # Load the workbook
            logger.info(f"Opening Loan Schedule file: {self.loan_schedule_file}")
            wb = openpyxl.load_workbook(self.loan_schedule_file, data_only=True)

            # Check if the required sheet exists
            if "Loan Summary" not in wb.sheetnames:
                logger.error(f"Sheet 'Loan Summary' not found. Available sheets: {wb.sheetnames}")
                return None, None

            sheet = wb["Loan Summary"]

            # Get values from I29 and J29
            i29_value = sheet["I29"].value
            j29_value = sheet["J29"].value

            logger.info(f"Extracted values - I29: {i29_value}, J29: {j29_value}")

            return i29_value, j29_value

        except Exception as e:
            logger.error(f"Error extracting Loan Summary values: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None, None
        finally:
            if wb is not None:
                try:
                    wb.close()
                except:
                    pass
    
    def update_deposit_liability_value(self):
        """
        Extract value from Deposit Liability file (NBD-WF-18-DM sheet, D14 cell)
        and update AFL Liquidity file (NBD-WF-15-LA sheet, C7 cell)
        """
        dep_wb = None
        afl_wb = None
        try:
            if not self.deposit_liability_file:
                logger.error("Deposit Liability file not found")
                return False

            if not self.afl_liquidity_file:
                logger.error("AFL Liquidity file not found")
                return False

            # Load Deposit Liability file
            logger.info(f"Opening Deposit Liability file: {self.deposit_liability_file}")
            dep_wb = openpyxl.load_workbook(self.deposit_liability_file, data_only=True)

            # Check if NBD-WF-18-DM sheet exists
            if "NBD-WF-18-DM" not in dep_wb.sheetnames:
                logger.error(f"Sheet 'NBD-WF-18-DM' not found in Deposit Liability file. Available sheets: {dep_wb.sheetnames}")
                return False

            dep_sheet = dep_wb["NBD-WF-18-DM"]

            # Extract value from D14
            d14_value = dep_sheet["D14"].value
            logger.info(f"Extracted value from Deposit Liability D14: {d14_value}")

            # Close deposit workbook
            dep_wb.close()
            dep_wb = None

            # Load AFL Liquidity file
            logger.info(f"Opening AFL Liquidity file: {self.afl_liquidity_file}")
            afl_wb = openpyxl.load_workbook(self.afl_liquidity_file)

            # Check if NBD-WF-15-LA sheet exists
            if "NBD-WF-15-LA" not in afl_wb.sheetnames:
                logger.error(f"Sheet 'NBD-WF-15-LA' not found in AFL Liquidity file. Available sheets: {afl_wb.sheetnames}")
                return False

            afl_sheet = afl_wb["NBD-WF-15-LA"]

            # Update C7 cell
            afl_sheet["C7"].value = d14_value
            logger.info(f"Updated AFL Liquidity NBD-WF-15-LA C7 with value: {d14_value}")

            # Save the workbook
            afl_wb.save(self.afl_liquidity_file)
            afl_wb.close()
            afl_wb = None

            logger.info("Deposit Liability value updated successfully in AFL Liquidity file")
            return True

        except Exception as e:
            logger.error(f"Error updating deposit liability value: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False
        finally:
            if dep_wb is not None:
                try:
                    dep_wb.close()
                except:
                    pass
            if afl_wb is not None:
                try:
                    afl_wb.close()
                except:
                    pass
    
    def update_treasury_input_with_bank_balances(self, target_date="07/25/2025"):
        """
        Update AFL Liquidity Treasury Input sheet with bank balances from Daily Bank Balances file

        Args:
            target_date (str): Date to look up in MM/DD/YYYY format
        """
        afl_wb = None
        bank_wb = None
        try:
            if not self.daily_bank_balance_file:
                logger.error("Daily Bank Balances file not found")
                return False

            if not self.afl_liquidity_file:
                logger.error("AFL Liquidity file not found")
                return False

            # Parse the target date to be flexible with different formats
            try:
                target_dt = datetime.strptime(target_date, '%m/%d/%Y')
            except:
                try:
                    target_dt = datetime.strptime(target_date, '%d/%m/%Y')
                except:
                    logger.error(f"Invalid date format: {target_date}. Use MM/DD/YYYY")
                    return False

            logger.info(f"Looking for date: {target_dt.strftime('%m/%d/%Y')}")

            # First, read GL codes from AFL Liquidity Treasury Input sheet
            logger.info(f"Opening AFL Liquidity file to read GL codes: {self.afl_liquidity_file}")
            afl_wb = openpyxl.load_workbook(self.afl_liquidity_file)

            if "Treasury Input" not in afl_wb.sheetnames:
                logger.error(f"Sheet 'Treasury Input' not found. Available sheets: {afl_wb.sheetnames}")
                return False

            treasury_sheet = afl_wb["Treasury Input"]

            # Extract GL codes from AFL file (A30:A73 and A77:A93)
            afl_gl_codes = []

            # Read from A30 to A73
            for row_idx in range(30, 74):
                gl_code = treasury_sheet.cell(row=row_idx, column=1).value  # Column A
                if gl_code:
                    afl_gl_codes.append((row_idx, str(gl_code).strip()))

            # Read from A77 to A93
            for row_idx in range(77, 94):
                gl_code = treasury_sheet.cell(row=row_idx, column=1).value  # Column A
                if gl_code:
                    afl_gl_codes.append((row_idx, str(gl_code).strip()))

            logger.info(f"Found {len(afl_gl_codes)} GL codes in AFL Treasury Input sheet")
            afl_wb.close()
            afl_wb = None

            # Load Daily Bank Balances file
            logger.info(f"Opening Daily Bank Balances file: {self.daily_bank_balance_file}")
            bank_wb = openpyxl.load_workbook(self.daily_bank_balance_file, data_only=True)

            # Get the first sheet
            bank_sheet = bank_wb.active
            logger.info(f"Reading from sheet: {bank_sheet.title}")

            # Find the column for the target date (row 2 contains dates, starting from column D)
            target_col = None
            logger.info("Searching for target date in Daily Bank Balance file...")

            for col_idx in range(4, bank_sheet.max_column + 1):  # Start from column D (index 4)
                cell_value = bank_sheet.cell(row=2, column=col_idx).value

                if cell_value:
                    date_match = False

                    # Try to match the date in different formats
                    if isinstance(cell_value, datetime):
                        # If it's a datetime object, compare directly
                        if cell_value.date() == target_dt.date():
                            date_match = True
                            date_str = cell_value.strftime('%m/%d/%Y')
                    else:
                        # If it's a string, try to parse it
                        cell_str = str(cell_value).strip()
                        try:
                            cell_dt = datetime.strptime(cell_str, '%m/%d/%Y')
                            if cell_dt.date() == target_dt.date():
                                date_match = True
                                date_str = cell_str
                        except:
                            try:
                                cell_dt = datetime.strptime(cell_str, '%d/%m/%Y')
                                if cell_dt.date() == target_dt.date():
                                    date_match = True
                                    date_str = cell_str
                            except:
                                date_str = cell_str

                    if date_match:
                        target_col = col_idx
                        logger.info(f"✓ FOUND target date {target_date} in column {get_column_letter(col_idx)} (index {col_idx})")
                        break
                    else:
                        logger.debug(f"Column {get_column_letter(col_idx)}: {date_str if 'date_str' in locals() else cell_value}")

            if not target_col:
                logger.error(f"Could not find date {target_date} in Daily Bank Balances file (searched from column D onwards in row 2)")
                return False

            # Build lookup dictionary: GL code -> balance value
            # GL codes are in column B (column 2), starting from row 3
            gl_lookup = {}
            for row_idx in range(3, bank_sheet.max_row + 1):
                gl_code_cell = bank_sheet.cell(row=row_idx, column=2)  # Column B
                balance_cell = bank_sheet.cell(row=row_idx, column=target_col)

                if gl_code_cell.value:
                    gl_code_str = str(gl_code_cell.value).strip()
                    balance_value = balance_cell.value

                    # Handle None values
                    if balance_value is None:
                        balance_value = "-"

                    gl_lookup[gl_code_str] = balance_value
                    logger.debug(f"DBB Row {row_idx}: GL {gl_code_str} = {balance_value}")

            logger.info(f"Built lookup table with {len(gl_lookup)} GL codes from Daily Bank Balance")
            bank_wb.close()
            bank_wb = None

            # Now update AFL Liquidity file with the matched values
            logger.info(f"Reopening AFL Liquidity file to update values: {self.afl_liquidity_file}")
            afl_wb = openpyxl.load_workbook(self.afl_liquidity_file)
            treasury_sheet = afl_wb["Treasury Input"]

            updated_count = 0
            not_found_count = 0

            for row_idx, gl_code in afl_gl_codes:
                if gl_code in gl_lookup:
                    balance = gl_lookup[gl_code]
                    treasury_sheet.cell(row=row_idx, column=3).value = balance  # Column C
                    updated_count += 1
                    logger.info(f"✓ AFL Row {row_idx}: GL {gl_code} = {balance}")
                else:
                    treasury_sheet.cell(row=row_idx, column=3).value = "-"
                    not_found_count += 1
                    logger.warning(f"✗ AFL Row {row_idx}: GL {gl_code} NOT FOUND in Daily Bank Balance, set to '-'")

            logger.info(f"Summary: Updated {updated_count} GL codes, {not_found_count} not found")

            # Save the workbook
            afl_wb.save(self.afl_liquidity_file)
            afl_wb.close()
            afl_wb = None

            logger.info("Treasury Input sheet updated successfully with bank balances")
            return True

        except Exception as e:
            logger.error(f"Error updating Treasury Input with bank balances: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False
        finally:
            if afl_wb is not None:
                try:
                    afl_wb.close()
                except:
                    pass
            if bank_wb is not None:
                try:
                    bank_wb.close()
                except:
                    pass
    
    def update_treasury_input_sheet(self, i29_value, j29_value):
        """
        Update the AFL Liquidity file (in Inputs folder) with Loan Summary values
        Sheet: Treasury Input, Cells: C8, E8, C9, E9
        """
        wb = None
        try:
            if not self.afl_liquidity_file:
                logger.error("AFL Liquidity file not found")
                return False

            # Load the workbook
            logger.info(f"Opening AFL Liquidity file: {self.afl_liquidity_file}")
            wb = openpyxl.load_workbook(self.afl_liquidity_file)

            # Check if the required sheet exists
            if "Treasury Input" not in wb.sheetnames:
                logger.error(f"Sheet 'Treasury Input' not found. Available sheets: {wb.sheetnames}")
                return False

            sheet = wb["Treasury Input"]

            # Update cells with I29 and J29 values
            if i29_value is not None:
                sheet["C8"].value = i29_value
                sheet["E8"].value = i29_value
                logger.info(f"Updated C8 and E8 with I29 value: {i29_value}")
            else:
                logger.warning("I29 value is None, skipping C8 and E8 update")

            if j29_value is not None:
                sheet["C9"].value = j29_value
                sheet["E9"].value = j29_value
                logger.info(f"Updated C9 and E9 with J29 value: {j29_value}")
            else:
                logger.warning("J29 value is None, skipping C9 and E9 update")

            # Save the workbook
            wb.save(self.afl_liquidity_file)
            wb.close()
            wb = None

            logger.info("Treasury Input sheet updated successfully with Loan Summary values")
            return True

        except Exception as e:
            logger.error(f"Error updating Treasury Input sheet: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False
        finally:
            if wb is not None:
                try:
                    wb.close()
                except:
                    pass
    
    def get_next_test_version(self):
        """
        Find the highest test version number and return the next one
        Returns: next version number (int)
        """
        try:
            # Find all AFL Test files
            test_files = list(self.output_inputs_folder.glob("AFL Test *.xlsx"))

            if not test_files:
                # No test files found, start with version 1
                return 1

            # Extract version numbers
            import re
            versions = []
            for file in test_files:
                match = re.search(r'AFL Test (\d+)\.xlsx', file.name)
                if match:
                    versions.append(int(match.group(1)))

            if versions:
                return max(versions) + 1
            else:
                return 1

        except Exception as e:
            logger.error(f"Error finding next test version: {e}")
            return 1

    def find_latest_afl_file_across_all_folders(self):
        """
        Find the latest AFL Liquidity file from outputs folder (completed weeks)
        Returns: Path to latest AFL file, or None if not found
        """
        try:
            logger.info("Searching for latest AFL Liquidity file in outputs...")

            all_afl_files = []

            # NEW STRUCTURE: Search in outputs folder for completed date folders
            outputs_dir = self.base_dir / "outputs" / "NBD_MF_15_LA"

            if not outputs_dir.exists():
                logger.warning(f"Outputs directory not found: {outputs_dir}")
                return None

            # Search through all date folders in outputs directory
            for date_folder in outputs_dir.glob("*"):
                if not date_folder.is_dir():
                    continue

                # Find AFL Liquidity files directly in date folder (new structure)
                afl_files = list(date_folder.glob("AFL Liquidity - Week_*.xlsx"))
                for afl_file in afl_files:
                    all_afl_files.append(afl_file)
                    logger.info(f"  Found: {date_folder.name}/{afl_file.name}")

            if not all_afl_files:
                logger.info("No AFL Liquidity files found in outputs folder")
                return None

            # Get the most recent file by modification time
            latest_file = max(all_afl_files, key=lambda f: f.stat().st_mtime)
            logger.info(f"✓ Latest AFL file: {latest_file.parent.name}/{latest_file.name}")
            logger.info(f"  Modified: {datetime.fromtimestamp(latest_file.stat().st_mtime).strftime('%Y-%m-%d %H:%M:%S')}")
            return latest_file

        except Exception as e:
            logger.error(f"Error finding latest AFL file: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None

    def create_afl_test_copy(self):
        """
        Create a new versioned AFL file from the currently updated working file.
        The working file (which was just updated) becomes the source for the new AFL file.
        Saves to the versioned output Inputs folder (e.g., 07-13-2025/NBD_MF_04_LA/Inputs or 07-13-2025(2)/NBD_MF_04_LA/Inputs)
        and updates date cells inside the file.
        """
        wb = None
        try:
            # Calculate week number from report date
            week_number = self.report_date.isocalendar()[1]

            # Create filename with week number
            copy_filename = f"AFL Liquidity - Week_{week_number}.xlsx"
            copy_path = self.output_inputs_folder / copy_filename

            # Use the current working file (which has been updated) as source
            source_file = self.working_file

            if not source_file:
                logger.error("Working file not set")
                return False

            logger.info(f"Creating {copy_filename} from {source_file.name}...")

            # Create the copy in working Inputs folder
            shutil.copy2(source_file, copy_path)
            logger.info(f"✓ Created: {copy_path.name} in {self.output_inputs_folder}")

            # Now update date cells inside the new file
            logger.info("Updating date cells inside the new file...")
            wb = openpyxl.load_workbook(copy_path)

            # Update Treasury Input sheet - C1 (e.g., "Oct-07")
            if "Treasury Input" in wb.sheetnames:
                treasury_sheet = wb["Treasury Input"]
                treasury_sheet["C1"].value = self.report_date.strftime("%b-%d")
                logger.info(f"Updated Treasury Input C1: {self.report_date.strftime('%b-%d')}")

            # Update Liquid Assets sheet - D1 (e.g., "Oct-07")
            if "Liquid Assets" in wb.sheetnames:
                liquid_sheet = wb["Liquid Assets"]
                liquid_sheet["D1"].value = self.report_date.strftime("%b-%d")
                logger.info(f"Updated Liquid Assets D1: {self.report_date.strftime('%b-%d')}")

            # Update Borrowings sheet - D1 (e.g., "Oct-07")
            if "Borrowings" in wb.sheetnames:
                borrowings_sheet = wb["Borrowings"]
                borrowings_sheet["D1"].value = self.report_date.strftime("%b-%d")
                logger.info(f"Updated Borrowings D1: {self.report_date.strftime('%b-%d')}")

            # Update NBD-WF-15-LA sheet - C3 (e.g., "Oct 07, 2025")
            if "NBD-WF-15-LA" in wb.sheetnames:
                nbd_sheet = wb["NBD-WF-15-LA"]
                nbd_sheet["C3"].value = self.report_date.strftime("%b %d, %Y")
                logger.info(f"Updated NBD-WF-15-LA C3: {self.report_date.strftime('%b %d, %Y')}")

            # Update ALCO sheet - C1, D1, E1 (e.g., "07-Oct-25")
            if "ALCO" in wb.sheetnames:
                alco_sheet = wb["ALCO"]
                alco_date = self.report_date.strftime("%d-%b-%y")

                # Unmerge cells first if they are merged
                cells_to_update = ["C1", "D1", "E1"]
                for cell_ref in cells_to_update:
                    # Check if cell is part of a merged range
                    cell = alco_sheet[cell_ref]
                    if isinstance(cell, openpyxl.cell.cell.MergedCell):
                        # Find and unmerge the range
                        for merged_range in list(alco_sheet.merged_cells.ranges):
                            if cell_ref in merged_range:
                                alco_sheet.unmerge_cells(str(merged_range))
                                logger.info(f"Unmerged cell range: {merged_range}")
                                break

                # Now update the values
                alco_sheet["C1"].value = alco_date
                alco_sheet["D1"].value = alco_date
                alco_sheet["E1"].value = alco_date
                logger.info(f"Updated ALCO C1, D1, E1: {alco_date}")

            # Save the updated workbook
            wb.save(copy_path)
            wb.close()
            wb = None

            # Update working_file to point to the new test file
            self.working_file = copy_path
            logger.info(f"Working file updated to: {copy_path.name}")
            logger.info("✓ All date cells updated successfully")

            return True

        except Exception as e:
            logger.error(f"Error creating AFL Test copy: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False
        finally:
            if wb is not None:
                try:
                    wb.close()
                except:
                    pass

    def run_tb_bot(self):
        """
        Run the TB from ERP bot to download TB-COA file
        Retries indefinitely until the file is successfully downloaded
        Returns: True if successful, False otherwise
        """
        try:
            logger.info("\n" + "=" * 60)
            logger.info("Running TB from ERP Bot...")
            logger.info("=" * 60)

            # Clean up old TB files before downloading (pattern: YYYY_MM_DD_HH_MM_Report.xlsx)
            logger.info(f"Cleaning up old TB files in {self.input_folder}...")
            try:
                old_tb_files = list(self.input_folder.glob("*_*_*_*_*_Report.xlsx"))
                for old_file in old_tb_files:
                    try:
                        old_file.unlink()
                        logger.info(f"✓ Removed old TB file: {old_file.name}")
                    except Exception as e:
                        logger.warning(f"Failed to remove {old_file.name}: {e}")

                if not old_tb_files:
                    logger.info("No old TB files found to clean up")
            except Exception as e:
                logger.warning(f"Failed to clean up old TB files: {e}")

            bot_script = self.bots_folder / "tb_from_erp_bot.py"

            if not bot_script.exists():
                logger.error(f"TB bot script not found: {bot_script}")
                return False

            # Load environment variables
            env_path = self.base_dir / '.env'
            load_dotenv(dotenv_path=env_path)

            # Import and run the bot
            sys.path.insert(0, str(self.bots_folder))

            try:
                from tb_from_erp_bot import ERPLoginBot

                # Get encrypted credentials
                ENC_USERNAME = os.getenv('ENC_USERNAME')
                ENC_PASSWORD = os.getenv('ENC_PASSWORD')

                if not ENC_USERNAME or not ENC_PASSWORD:
                    logger.error("ENC_USERNAME and ENC_PASSWORD not found in environment variables")
                    return False

                # Format report date for bot (DD/MM/YYYY)
                report_date = self.report_date.strftime('%d/%m/%Y')

                # Retry loop - keep trying until TB-COA file is downloaded
                retry_count = 0
                while True:
                    retry_count += 1
                    logger.info(f"\nAttempt #{retry_count} to download TB-COA file...")

                    # Create bot instance with Input folder as download location
                    bot = ERPLoginBot(download_folder=str(self.input_folder))

                    # Run bot automation
                    success = bot.run_full_automation(ENC_USERNAME, ENC_PASSWORD, report_date)

                    # Clean up
                    bot.close_browser()

                    if success:
                        logger.info("✓ TB bot completed successfully")

                        # Verify TB-COA file actually exists
                        tb_file = self.find_tb_coa_file()
                        if tb_file:
                            logger.info(f"✓ TB-COA file confirmed downloaded: {tb_file.name}")
                            return True
                        else:
                            logger.warning("TB bot reported success but file not found. Retrying immediately...")
                            continue
                    else:
                        logger.warning(f"TB bot failed on attempt #{retry_count}. Retrying immediately...")
                        continue

            except ImportError as e:
                logger.error(f"Failed to import TB bot: {e}")
                return False

        except Exception as e:
            logger.error(f"Error running TB bot: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False
        finally:
            # Remove bots folder from path
            if str(self.bots_folder) in sys.path:
                sys.path.remove(str(self.bots_folder))

    def run_m2m_bot(self):
        """
        Run the M2M from SASIA bot
        Returns: True if successful, False otherwise
        """
        try:
            logger.info("\n" + "=" * 60)
            logger.info("Running M2M from SASIA Bot...")
            logger.info("=" * 60)

            bot_script = self.bots_folder / "m2m_from_sasia_bot.py"

            if not bot_script.exists():
                logger.error(f"M2M bot script not found: {bot_script}")
                return False

            # Check if the script has actual content
            with open(bot_script, 'r') as f:
                content = f.read().strip()
                if len(content) < 50:  # If it's just a placeholder
                    logger.warning("M2M bot script appears to be a placeholder, skipping...")
                    return True

            # Run the bot script
            result = subprocess.run(
                [sys.executable, str(bot_script)],
                cwd=str(self.bots_folder),
                capture_output=True,
                text=True
            )

            if result.returncode == 0:
                logger.info("✓ M2M bot completed successfully")
                logger.info(result.stdout)
                time.sleep(2)
                return True
            else:
                logger.error(f"M2M bot failed with return code {result.returncode}")
                logger.error(result.stderr)
                return False

        except Exception as e:
            logger.error(f"Error running M2M bot: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False

    def determine_working_file(self):
        """
        Determine which file to work on:
        - Search ALL folders for latest AFL Liquidity file (by modification time)
        - If found, copy it to temp working location
        - If not found, use the starting template from FIXED folder (07-06-2025/NBD_MF_04_LA/Inputs/AFL Liquidity - Week_27.xlsx)
        Always reads master data from FIXED Input folder (07-06-2025/NBD_MF_04_LA/Input)
        """
        try:
            # Search across ALL folders for the latest AFL file
            latest_afl = self.find_latest_afl_file_across_all_folders()

            if latest_afl:
                # Copy latest AFL to temp working location for modifications (NEW STRUCTURE)
                temp_working_file = self.working_folder / f"temp_{latest_afl.name}"
                shutil.copy2(latest_afl, temp_working_file)
                self.working_file = temp_working_file
                logger.info(f"✓ Working file: Using latest AFL from {latest_afl.parent.name}")
                logger.info(f"  Copied {latest_afl.name} to working folder")
            else:
                # First run - look for starting template in current working folder (NEW STRUCTURE)
                # Template should be in working/NBD_MF_15_LA/{DD-MM-YYYY}/ (copied from outputs by app.py)
                starting_template = self.working_folder / "AFL Liquidity - Week_27.xlsx"

                if not starting_template.exists():
                    logger.error(f"Starting template not found: {starting_template}")
                    logger.info(f"Please place 'AFL Liquidity - Week_27.xlsx' in {self.working_folder}")
                    return None

                # Copy template to temp working location
                temp_working_file = self.working_folder / f"temp_{starting_template.name}"
                shutil.copy2(starting_template, temp_working_file)
                self.working_file = temp_working_file
                logger.info(f"✓ Working file: Using starting template (first run)")
                logger.info(f"  Copied {starting_template.name} from {self.working_folder}")

            return self.working_file

        except Exception as e:
            logger.error(f"Error determining working file: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None

    def validate_liquid_assets(self):
        """
        Validate that Liquid Assets sheet cells C93 and D93 are 0
        Returns: (is_valid, exceptions_list)
        """
        wb = None
        try:
            if not self.afl_liquidity_file:
                logger.error("AFL Liquidity file not found for validation")
                return False, []

            logger.info("Validating Liquid Assets sheet...")
            wb = openpyxl.load_workbook(self.afl_liquidity_file, data_only=True)

            if "Liquid Assets" not in wb.sheetnames:
                logger.error(f"Sheet 'Liquid Assets' not found. Available sheets: {wb.sheetnames}")
                return False, []

            sheet = wb["Liquid Assets"]

            # Get values from C93 and D93
            c93_value = sheet["C93"].value
            d93_value = sheet["D93"].value

            logger.info(f"Liquid Assets validation - C93: {c93_value}, D93: {d93_value}")

            exceptions = []

            # Check if C93 is not 0
            if c93_value is not None and c93_value != 0:
                exceptions.append({
                    'file': self.afl_liquidity_file.name,
                    'sheet': 'Liquid Assets',
                    'cell': 'C93',
                    'value': c93_value,
                    'expected': 0
                })
                logger.warning(f"VALIDATION FAILED: C93 = {c93_value} (expected 0)")

            # Check if D93 is not 0
            if d93_value is not None and d93_value != 0:
                exceptions.append({
                    'file': self.afl_liquidity_file.name,
                    'sheet': 'Liquid Assets',
                    'cell': 'D93',
                    'value': d93_value,
                    'expected': 0
                })
                logger.warning(f"VALIDATION FAILED: D93 = {d93_value} (expected 0)")

            if len(exceptions) == 0:
                logger.info("✓ Validation PASSED: Both C93 and D93 are 0")
                return True, []
            else:
                logger.error(f"✗ Validation FAILED: {len(exceptions)} exception(s) found")
                return False, exceptions

        except Exception as e:
            logger.error(f"Error during validation: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False, []
        finally:
            if wb is not None:
                try:
                    wb.close()
                except:
                    pass

    def create_exceptions_report(self, exceptions):
        """
        Create an exceptions report Excel file with the validation failures

        Args:
            exceptions (list): List of exception dictionaries
        """
        try:
            if not exceptions:
                logger.info("No exceptions to report")
                return True

            # Extract week number from AFL Liquidity filename
            afl_filename = self.afl_liquidity_file.name
            week_number = "XX"  # Default if not found

            # Try to extract week number from filename
            import re
            week_match = re.search(r'Week\s*(\d+)', afl_filename, re.IGNORECASE)
            if week_match:
                week_number = week_match.group(1)

            # Create exceptions report filename
            exceptions_filename = f"AFL Liquidity - week {week_number} - Exceptions.xlsx"
            exceptions_path = self.output_inputs_folder / exceptions_filename

            logger.info(f"Creating exceptions report: {exceptions_filename}")

            # Create a new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Exceptions"

            # Add headers
            headers = ['File', 'Sheet', 'Cell', 'Actual Value', 'Expected Value']
            ws.append(headers)

            # Style headers
            from openpyxl.styles import Font, PatternFill, Alignment
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Add exception rows
            for exception in exceptions:
                ws.append([
                    exception['file'],
                    exception['sheet'],
                    exception['cell'],
                    exception['value'],
                    exception['expected']
                ])

            # Adjust column widths
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15

            # Save the workbook
            wb.save(exceptions_path)
            wb.close()

            logger.info(f"✓ Exceptions report created: {exceptions_path}")
            return True

        except Exception as e:
            logger.error(f"Error creating exceptions report: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False
    
    def run_automation(self):
        """
        Main method to run the AFL Liquidity report automation with integrated bot execution
        """
        logger.info("=" * 60)
        logger.info("Starting AFL Liquidity Report Automation")
        logger.info(f"Report Date: {self.report_date.strftime('%Y-%m-%d')}")
        logger.info("=" * 60)

        try:
            # Step 0: Run Selenium bots if enabled
            if self.run_bots:
                # Run TB from ERP bot first - CRITICAL: Will retry until successful
                logger.info("\nStep 0a: Running TB from ERP Bot...")
                logger.info("NOTE: TB download will retry indefinitely until successful")
                tb_bot_success = self.run_tb_bot()

                if not tb_bot_success:
                    logger.error("TB bot failed to download TB-COA file after retries")
                    logger.error("Cannot proceed without TB-COA file. Stopping automation.")
                    return False

                # Run M2M from SASIA bot
                logger.info("\nStep 0b: Running M2M from SASIA Bot...")
                m2m_bot_success = self.run_m2m_bot()

                if not m2m_bot_success:
                    logger.warning("M2M bot failed, but continuing with automation...")
                    # Don't return False here - M2M is optional
            else:
                logger.info("\nBot execution disabled - skipping Selenium automation")

            # Step 1: Find required files
            logger.info("\nStep 1: Finding required Excel files...")
            logger.info("  - Looking for TB-COA file in Input folder...")
            tb_coa_file = self.find_tb_coa_file()

            logger.info("  - Looking for Loan Schedule in Input folder...")
            loan_schedule_file = self.find_loan_schedule_file()

            logger.info("  - Looking for Daily Bank Balances in Input folder...")
            bank_balance_file = self.find_daily_bank_balance_file()

            logger.info("  - Looking for Deposit Liability in Input folder...")
            deposit_liability_file = self.find_deposit_liability_file()

            # Check all required files - ALL must be present
            missing_files = []
            if not tb_coa_file:
                missing_files.append("TB-COA file")
            if not loan_schedule_file:
                missing_files.append("Loan Schedule file")
            if not bank_balance_file:
                missing_files.append("Daily Bank Balances file")
            if not deposit_liability_file:
                missing_files.append("Deposit Liability file")

            if missing_files:
                logger.error("\n" + "=" * 60)
                logger.error("MISSING REQUIRED FILES - CANNOT PROCEED")
                logger.error("=" * 60)
                logger.error(f"The following required files are missing from Input folder:")
                for file in missing_files:
                    logger.error(f"  ✗ {file}")
                logger.error(f"\nInput folder location: {self.input_folder}")
                logger.error("\nPlease ensure all required files are present before running the automation.")
                logger.error("=" * 60)
                return False

            # Step 1.5: Determine which file to work on (latest AFL or template from working folder)
            logger.info("\nStep 1.5: Determining working file...")
            logger.info(f"  - Will look for NBD-MF-04-LA template in working folder: {self.working_folder}")
            working_file = self.determine_working_file()

            if not working_file:
                logger.error("Could not determine working file. Stopping automation.")
                return False

            # Important: Update self.afl_liquidity_file to point to working file for all operations
            # This ensures all updates happen on the correct file
            self.afl_liquidity_file = working_file

            # Step 2: Copy TB-COA data to working file's TB sheet
            logger.info("\nStep 2: Copying TB-COA data to working file's TB sheet...")
            tb_copy_success = self.copy_tb_coa_data()

            if not tb_copy_success:
                logger.error("Failed to copy TB-COA data")
                return False

            # Step 2.5: Find and extract M2M values
            logger.info("\nStep 2.5: Processing M2M file for Treasury Input C23, C24, C25...")
            logger.info("  - Looking for M2M file in Input folder...")
            m2m_file = self.find_m2m_file()

            if m2m_file:
                # Extract M2M values
                logger.info("  - Extracting values from M2M file...")
                c23_value, c24_value, c25_value = self.extract_m2m_values()

                # Update Treasury Input with M2M values
                if c23_value is not None or c24_value is not None or c25_value is not None:
                    logger.info("  - Updating Treasury Input with M2M values...")
                    m2m_update_success = self.update_treasury_input_with_m2m(c23_value, c24_value, c25_value)

                    if not m2m_update_success:
                        logger.warning("Failed to update Treasury Input with M2M values, but continuing...")
                else:
                    logger.warning("No M2M values extracted, skipping M2M update")
            else:
                logger.warning("M2M file not found, skipping M2M update (C23, C24, C25 will not be updated)")

            # Step 3: Extract Loan Summary values from Input folder
            logger.info("\nStep 3: Extracting Loan Summary values from Input folder...")
            i29_value, j29_value = self.extract_loan_summary_values()

            if i29_value is None and j29_value is None:
                logger.error("Failed to extract Loan Summary values")
                return False

            # Step 4: Update working file's Treasury Input sheet with Loan Summary values
            logger.info("\nStep 4: Updating working file's Treasury Input sheet with Loan Summary values...")
            update_success = self.update_treasury_input_sheet(i29_value, j29_value)

            if not update_success:
                logger.error("Failed to update Treasury Input sheet")
                return False

            # Step 5: Update working file's Treasury Input sheet with bank balances
            logger.info("\nStep 5: Updating working file's Treasury Input sheet with bank balances...")
            # Use the report date for bank balances lookup
            target_date_str = self.report_date.strftime('%m/%d/%Y')
            bank_balance_success = self.update_treasury_input_with_bank_balances(target_date=target_date_str)

            if not bank_balance_success:
                logger.error("Failed to update Treasury Input with bank balances")
                return False

            # Step 6: Update Deposit Liability value in working file
            logger.info("\nStep 6: Updating Deposit Liability value in working file...")
            deposit_success = self.update_deposit_liability_value()

            if not deposit_success:
                logger.error("Failed to update Deposit Liability value")
                return False

            # Step 7: Validate working file before creating test copy
            logger.info("\nStep 7: Validating Liquid Assets sheet (C93 and D93 must be 0)...")
            is_valid, exceptions = self.validate_liquid_assets()

            if not is_valid:
                logger.warning("Validation failed - creating exceptions report...")
                self.create_exceptions_report(exceptions)
                logger.info("\n" + "=" * 60)
                logger.warning("COMPLETED WITH EXCEPTIONS: Check the exceptions report")
                logger.warning("NOT creating test file due to validation errors")
                logger.info("=" * 60)
                return True  # Automation completed, but with validation errors

            # Step 8: Create new versioned AFL Test file (only if validation passed)
            logger.info("\nStep 8: Creating new versioned AFL Test file...")
            copy_success = self.create_afl_test_copy()

            if not copy_success:
                logger.error("Failed to create AFL Test copy")
                return False

            logger.info("\n" + "=" * 60)
            logger.info("SUCCESS: AFL Liquidity Report Automation completed successfully")
            logger.info("All validations passed!")
            logger.info(f"New AFL file created: {self.working_file.name}")
            logger.info("=" * 60)

            # Step 9: Check if we should generate monthly report
            logger.info("\nStep 9: Checking for monthly report generation...")
            monthly_report_generated = self.check_and_run_monthly_report()

            if monthly_report_generated:
                logger.info("✓ Monthly report generation completed")

            return True
            
        except Exception as e:
            logger.error(f"Error in automation process: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False

def main():
    """
    Main function to run the AFL Liquidity automation with integrated workflow

    Usage:
        python NBD_MF_15_LA.py 10/07/2025
        python NBD_MF_15_LA.py 07/10/2025 --no-bots

    Command-line arguments:
        date: Report date in MM/DD/YYYY or DD/MM/YYYY format (optional, defaults to yesterday)
        --no-bots: Skip running Selenium bots (use existing files)
    """
    try:
        # Parse command line arguments
        import argparse
        parser = argparse.ArgumentParser(
            description='AFL Liquidity Report Automation - Date-Driven Workflow',
            formatter_class=argparse.RawDescriptionHelpFormatter,
            epilog="""
Examples:
    python NBD_MF_15_LA.py 10/07/2025
    python NBD_MF_15_LA.py 07/10/2025 --no-bots

Workflow:
1. Run TB from ERP bot (Selenium) - downloads TB-COA file for the specified date
2. Run M2M from SASIA bot (Selenium) - downloads M2M file
3. Determine working file (latest AFL Test X or AFL Liquidity)
4. Copy TB-COA data to working file
5. Extract and update Loan Summary values
6. Update bank balances for the specified date
7. Update Deposit Liability values
8. Validate Liquid Assets sheet (C93 and D93 must be 0)
9. Create new AFL Liquidity - Week_XX.xlsx file with updated dates (if validation passes)

Date Updates in File:
- Treasury Input C1: Oct-07
- Liquid Assets D1: Oct-07
- Borrowings D1: Oct-07
- NBD-WF-15-LA C3: Oct 07, 2025
- ALCO C1, D1, E1: 07-Oct-25
            """
        )
        parser.add_argument('date', nargs='?', type=str, help='Report date in MM/DD/YYYY or DD/MM/YYYY format (default: yesterday)')
        parser.add_argument('--no-bots', action='store_true', help='Skip running Selenium bots (use existing files)')
        args = parser.parse_args()

        report_date = None
        if args.date:
            # Try both MM/DD/YYYY and DD/MM/YYYY formats
            try:
                report_date = datetime.strptime(args.date, '%m/%d/%Y')
            except ValueError:
                try:
                    report_date = datetime.strptime(args.date, '%d/%m/%Y')
                except ValueError:
                    logger.error("Invalid date format. Use MM/DD/YYYY or DD/MM/YYYY (e.g., 10/07/2025 or 07/10/2025)")
                    sys.exit(1)

        run_bots = not args.no_bots

        logger.info("=" * 60)
        logger.info("AFL LIQUIDITY AUTOMATION - DATE-DRIVEN WORKFLOW")
        logger.info("=" * 60)
        if report_date:
            logger.info(f"Report Date: {report_date.strftime('%m/%d/%Y (%B %d, %Y)')}")
        else:
            logger.info("Report Date: Yesterday (default)")
        if run_bots:
            logger.info("Mode: Full automation (with Selenium bots)")
        else:
            logger.info("Mode: Skip bots (use existing files)")
        logger.info("=" * 60)

        # Initialize and run automation
        automation = AFLLiquidityAutomation(
            base_dir="C:\\CBSL\\Script",
            input_date=report_date,
            run_bots=run_bots
        )
        success = automation.run_automation()

        if success:
            logger.info("\n✓ Automation completed successfully")
            sys.exit(0)
        else:
            logger.error("\n✗ Automation failed")
            sys.exit(1)

    except KeyboardInterrupt:
        logger.info("\nAutomation interrupted by user")
        sys.exit(1)
    except Exception as e:
        logger.error(f"\nUnexpected error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        sys.exit(1)

if __name__ == "__main__":
    main()