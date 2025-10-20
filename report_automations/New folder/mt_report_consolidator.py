import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re


def unmerge_cells(file_path):
    """
    Unmerge all cells in the Excel file and fill them with the merged cell's value.

    Args:
    - file_path: Path to the Excel file
    """
    wb = load_workbook(file_path)

    for sheet in wb.worksheets:
        # Get all merged cell ranges
        merged_cells = list(sheet.merged_cells.ranges)

        # Unmerge and fill each range
        for merged_range in merged_cells:
            # Get the value from the top-left cell
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_value = sheet.cell(min_row, min_col).value

            # Unmerge the cells
            sheet.unmerge_cells(str(merged_range))

            # Fill all cells in the range with the value
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row, col).value = top_left_value

    # Save the workbook
    wb.save(file_path)
    wb.close()


def consolidate_mt_reports():
    """
    Consolidate three MT reports into a single dataframe with opening balance from Jun 2025 MT report.

    Returns:
    - consolidated_df: DataFrame with MT No., Client Name, Opening Balance (Mar), and Actual Lending from each report
    """

    # Define file paths
    report_files = [
        r"Input\MT\30-04-2025  Client Available Limits Report.xlsx",
        r"Input\MT\31-05-2025  Client Available Limits Report.xlsx",
        r"Input\MT\30-06-2025  Client Available Limits Report.xlsx"
    ]

    # Path to MT - Jun 2025 report for opening balance
    mt_jun_2025_file = r"Input\MT\MT - Jun 2025.xlsx"

    # Column names for the output
    report_dates = [re.search(r'(\d{2}-\d{2}-\d{4})', file).group(1) for file in report_files]

    # Read opening balance from MT - Jun 2025.xlsx (Dis worksheet, Mar column)
    print(f"Reading opening balance from {mt_jun_2025_file}...")
    opening_balance_df = pd.read_excel(mt_jun_2025_file, sheet_name='Dis')

    # Extract MT No., Client Name, and Opening Balance (Mar column)
    opening_balance_df = opening_balance_df[['Unnamed: 0','Unnamed: 1', 'Unnamed: 2', opening_balance_df.columns[-1]]].copy()
    opening_balance_df.columns = ['MT No.','Gender', 'Client Name', 'Opening Balance']

    # Clean up any NaN values in MT No. or Client Name
    opening_balance_df = opening_balance_df.dropna(subset=['MT No.', 'Client Name'])

    print(f"Loaded opening balance for {len(opening_balance_df)} MT entries")

    # Read all three reports
    dfs = []
    for i, file_path in enumerate(report_files):
        if not os.path.exists(file_path):
            print(f"Warning: File not found - {file_path}")
            continue

        # Unmerge cells before reading
        print(f"Unmerging cells in {file_path}...")
        unmerge_cells(file_path)

        # Read Excel file and skip first 7 columns
        df = pd.read_excel(file_path,skiprows=7)

        # Display available columns for debugging
        if i == 0:
            print(f"Available columns in reports: {df.columns.tolist()}")

        # Select relevant columns (MT No., Client Name, Actual Lending)
        # Adjust column names if they differ in your files
        df_filtered = df[['MT No.', 'Client Name', 'Actual Lending(Rs.)']].copy()
        df_filtered.columns = ['MT No.', 'Client Name', f'Actual Lending {report_dates[i]}']

        dfs.append(df_filtered)

    if len(dfs) == 0:
        print("Error: No files were successfully loaded")
        return pd.DataFrame()

    # Merge all dataframes on MT No. and Client Name
    consolidated_df = dfs[0].reset_index(drop=True)

    for i in range(1, len(dfs)):
        consolidated_df = consolidated_df.merge(
            dfs[i],
            on=['MT No.', 'Client Name'],
            how='outer'
        ).reset_index(drop=True)

    # Merge opening balance as the 3rd column (after MT No. and Client Name)
    consolidated_df = consolidated_df.merge(
        opening_balance_df,
        on=['MT No.', 'Client Name'],
        how='left'
    ).reset_index(drop=True)

    # Reorder columns to place Column 2 and Opening Balance after Client Name
    cols = ['MT No.', 'Client Name', 'Gender', 'Opening Balance'] + [col for col in consolidated_df.columns if col not in ['MT No.', 'Client Name', 'Column 2', 'Opening Balance', 'Gender']]
    consolidated_df = consolidated_df[cols]

    # Fill NaN values with 0 for missing data
    lending_columns = [col for col in consolidated_df.columns if 'Actual Lending' in col or 'Opening Balance' in col]
    consolidated_df[lending_columns] = consolidated_df[lending_columns].fillna(0)

    # Sort by MT No. and reset index
    consolidated_df = consolidated_df.sort_values('MT No.').reset_index(drop=True)

    # Identify missing records (records that have NaN in Gender or Opening Balance)
    # Create boolean masks separately and use them on the same DataFrame to avoid index alignment issues
    gender_mask = consolidated_df['Gender'].isna()
    balance_mask = consolidated_df['Opening Balance'].isna()
    missing_records = consolidated_df[gender_mask | balance_mask].copy()

    # Export missing records to separate Excel file in root folder
    if len(missing_records) > 0:
        missing_records_file = "Missing_Records.xlsx"
        missing_records.to_excel(missing_records_file, index=False)
        print(f"\nFound {len(missing_records)} missing records")
        print(f"Missing records exported to: {missing_records_file}")
    else:
        print("\nNo missing records found")

    print(f"\nConsolidated Report Summary:")
    print(f"Total unique MT entries: {len(consolidated_df)}")
    print(f"\nDataFrame Shape: {consolidated_df.shape}")
    print(f"\nFirst few rows:")
    print(consolidated_df.head())

    # Generate filename with last month's name
    from datetime import datetime, timedelta
    import shutil

    # Get last month's name and year
    today = datetime.today()
    last_month_date = today.replace(day=1) - timedelta(days=1)
    last_month_name = last_month_date.strftime('%b %Y')  # e.g., "May 2025"

    # Extract base workbook name
    base_name = os.path.splitext(os.path.basename(mt_jun_2025_file))[0]  # "MT - Jun 2025"
    workbook_prefix = base_name.split(' - ')[0] if ' - ' in base_name else base_name  # "MT"

    # Create new filename with last month
    output_dir = os.path.dirname(mt_jun_2025_file)
    new_filename = f"{workbook_prefix} - {last_month_name}.xlsx"
    mt_output_file = os.path.join(output_dir, new_filename)

    # Copy original file to new filename if it doesn't exist or if we're updating
    if mt_output_file != mt_jun_2025_file:
        shutil.copy2(mt_jun_2025_file, mt_output_file)
        print(f"Created new workbook: {mt_output_file}")

    # Write consolidated_df to the workbook file, Dis sheet, starting from A1
    print(f"\nWriting consolidated data to {mt_output_file}, sheet 'Dis'...")
    with pd.ExcelWriter(mt_output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        consolidated_df.to_excel(writer, sheet_name='Dis', startrow=0, startcol=0, index=False)
    print(f"Data written successfully to 'Dis' sheet starting from cell A1")

    # Create gender summary table
    print(f"\nCreating gender summary table...")
    gender_summary = consolidated_df.groupby('Gender').agg({
        'Client Name': 'count',
        'Opening Balance': 'sum'
    }).reset_index()
    gender_summary.columns = ['Gender', 'Count of Client Name', 'Sum of Opening Balance']
    print(f"Gender summary table:")
    print(gender_summary)

    # Copy values only from Dis sheet A1:C1000 to Disbursement sheet starting at A3
    print(f"\nCopying values from 'Dis' sheet (A1:C1000) to 'Disbursement' sheet (A3)...")
    wb = load_workbook(mt_output_file)

    # Check if sheets exist
    if 'Dis' not in wb.sheetnames:
        print("Warning: 'Dis' sheet not found")
    elif 'Disbursement' not in wb.sheetnames:
        print("Warning: 'Disbursement' sheet not found")
    else:
        dis_sheet = wb['Dis']
        disbursement_sheet = wb['Disbursement']

        # Copy values only from A1:C1000 in Dis sheet to A3:C1002 in Disbursement sheet
        for row in range(1, 1001):  # Rows 1 to 1000 in Dis sheet
            for col in range(1, 4):  # Columns A, B, C (1, 2, 3)
                cell_value = dis_sheet.cell(row=row, column=col).value
                disbursement_sheet.cell(row=row+2, column=col).value = cell_value

        print(f"Values copied successfully from 'Dis' A1:C1000 to 'Disbursement' A3:C1002")

        
        # Find the last row with data in column A of Disbursement sheet
        last_row = 3  # Start from row 3 (where data starts)
        for row in range(3, disbursement_sheet.max_row + 1):
            if disbursement_sheet.cell(row=row, column=1).value is not None:
                last_row = row
            else:
                break

        print(f"\nCopying formulas from D10, E10, F10, G10 to end of data (row {last_row})...")

        # Get formulas from row 10, columns D, E, F, G (4, 5, 6, 7)
        formula_columns = [4, 5, 6, 7]  # D, E, F, G
        source_row = 10

        print(f"Source formulas from row {source_row}:")
        for col in formula_columns:
            cell = disbursement_sheet.cell(row=source_row, column=col)
            if cell.value:
                print(f"  Column {get_column_letter(col)}{source_row}: {cell.value}")

        # Copy formulas to all rows from 11 to last_row
        from openpyxl.formula.translate import Translator

        for row in range(11, last_row):
            for col in formula_columns:
                source_cell = disbursement_sheet.cell(row=source_row, column=col)
                if source_cell.value and isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                    # Calculate row offset
                    row_offset = row - source_row
                    # Translate formula with proper row offset
                    translated_formula = Translator(source_cell.value, origin=f"{get_column_letter(col)}{source_row}").translate_formula(
                        f"{get_column_letter(col)}{row}", row_delta=row_offset, col_delta=0
                    )
                    disbursement_sheet.cell(row=row, column=col).value = translated_formula

        print(f"Formulas copied to rows 11 through {last_row}")

        # Add total row after the last data row
        total_row = last_row + 1
        disbursement_sheet.cell(row=total_row, column=1).value = "Total"

        print(f"\nAdding totals in row {total_row} for columns D, E, F, G...")

        # Add SUM formulas for columns D, E, F, G
        for col in formula_columns:
            col_letter = get_column_letter(col)
            sum_formula = f"=SUM({col_letter}3:{col_letter}{last_row})"
            disbursement_sheet.cell(row=total_row, column=col).value = sum_formula
            print(f"Column {col_letter}: {sum_formula}")

        # Read Disbursement sheet data (skip first 2 rows) and create pivot table
        print(f"\nReading Disbursement sheet data (skipping 2 rows) for pivot table...")
        disbursement_df = pd.read_excel(mt_output_file, sheet_name='Disbursement', skiprows=2)

        # Create pivot table with Gender as rows, Count of Client Name and Sum of Total
        pivot_table = disbursement_df.pivot_table(
            values=['Client Name', 'Total'],
            index='Male/Female/Other',
            aggfunc={'Client Name': 'count', 'Total': 'sum'}
        ).reset_index()

        # Rename columns for clarity
        pivot_table.columns = ['Male/Female/Other', 'Count of Client Name', 'Sum of Total']

        print(f"Pivot table created:")
        print(pivot_table)

        # Write pivot table to Disbursement sheet starting at J9
        print(f"\nWriting pivot table to Disbursement sheet at J9...")
        start_row = 9
        start_col = 10  # Column J

        # Write headers
        for col_idx, header in enumerate(pivot_table.columns):
            disbursement_sheet.cell(row=start_row, column=start_col + col_idx).value = header

        # Write data rows
        for row_idx, row_data in pivot_table.iterrows():
            for col_idx, value in enumerate(row_data):
                disbursement_sheet.cell(row=start_row + 1 + row_idx, column=start_col + col_idx).value = value

        print(f"Pivot table written to J9:L{start_row + len(pivot_table)}")


        wb.save(mt_output_file)
        print(f"\nWorkbook saved with formulas and totals")

    wb.close()

    return consolidated_df


if __name__ == "__main__":
    # Execute the consolidation
    result_df = consolidate_mt_reports()

    # Optionally save to Excel
    output_file = "MT_Consolidated_Report.xlsx"
    result_df.to_excel(output_file, index=False)
    print(f"\nConsolidated report saved to: {output_file}")
