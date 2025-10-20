import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter

def unmerge_row1_and_save(file_path):
    """
    Opens an Excel file, unmerges all merged cells in row 1, 
    and saves it to the same location.
    
    Parameters:
    file_path (str): Path to the Excel file
    """
    output_file_path = './temp.xlsx'
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    
    # Get the active sheet (or specify sheet name if needed)
    ws = wb['Unutilized-MAR 2025']

    # Get all merged cell ranges
    merged_cells = list(ws.merged_cells.ranges)
    
    # Filter for merged cells that include row 1
    row1_merged = [mc for mc in merged_cells if mc.min_row == 1 and mc.max_row >= 1]
    
    # Unmerge cells in row 1
    for merged_cell in row1_merged:
        # Get the value from the merged cell before unmerging
        top_left_cell = ws.cell(merged_cell.min_row, merged_cell.min_col)
        value = top_left_cell.value
        
        # Unmerge the cells
        ws.unmerge_cells(str(merged_cell))
        
        # Optional: Fill all previously merged cells with the same value
        # Comment out the loop below if you don't want to fill all cells
        for row in range(merged_cell.min_row, merged_cell.max_row + 1):
            for col in range(merged_cell.min_col, merged_cell.max_col + 1):
                ws.cell(row, col).value = value
    
    # Save the workbook to the same location
    wb.save(output_file_path)
    wb.close()
    print(f"File saved successfully to: {file_path}")
    print(f"Unmerged {len(row1_merged)} cell range(s) in row 1")


def create_marginal_loans_df(unutilized_file_path, df_district=None, combined_df=None):

    temp = './temp.xlsx'
    unmerge_row1_and_save(unutilized_file_path)
    unutilized_df = pd.read_excel(temp, sheet_name='Unutilized-MAR 2025')
    print(unutilized_df.columns.tolist())
    margin_trading_df = pd.DataFrame()

    margin_trading_df['Name of the Client'] = unutilized_df['Name of the Client']
    margin_trading_df['ICAM code'] = unutilized_df['ICAM code']
    margin_trading_df['Interest Rate NOV'] = unutilized_df['Intrest Rate ']
    margin_trading_df['Market Value of Portfolio'] = unutilized_df['Market Value of Portfolio']
    margin_trading_df['(ICAM) Debtor'] = unutilized_df['  (ICAM) D/C']
    margin_trading_df['50% of Portfolio'] = margin_trading_df['Market Value of Portfolio'] * 0.5
    margin_trading_df['Limit'] = unutilized_df.iloc[:, 8]
    margin_trading_df['Liability'] = unutilized_df['Liability']
    margin_trading_df['Corporate / Individual'] = unutilized_df['Corporate / Individual']
    margin_trading_df['Repayment Cycle'] = 'Other'
    margin_trading_df['Provision Category'] = 'Performing'
    margin_trading_df['Forbone Loans (Yes/No)'] = 'No'
    margin_trading_df['CBSL Sector'] = 'Financial Services'
    
    # Fix for district mapping
    contract_to_district = dict(zip(df_district["CLM_CODE"], df_district["DISTRICT"]))
    margin_trading_df['CustomerDistrict(ClientMain)'] = margin_trading_df['ICAM code'].map(contract_to_district)
    
    # Fix for Micro/Small/Medium mapping - handle duplicates
    # Option 1: Drop duplicates, keeping the first occurrence
    combined_df_unique = combined_df.drop_duplicates(subset=['Client Code'], keep='first')
    contract_to_ct_micro = dict(zip(combined_df_unique["Client Code"], combined_df_unique["Micro/Small/Medium"]))
    margin_trading_df['Micro/Small/Medium'] = margin_trading_df['ICAM code'].map(contract_to_ct_micro)
    
    # Alternative Option 2: If you want to see which codes are duplicated
    # duplicates = combined_df[combined_df.duplicated(subset=['Client Code'], keep=False)]
    # print("Duplicate Client Codes:")
    # print(duplicates[['Client Code', 'Micro/Small/Medium']].sort_values('Client Code'))

    margin_trading_df['Gender'] = [f'=IF(LEFT(D{i + 2},1)="2",_xlfn.XLOOKUP(B1,BusinessGender!A:A,BusinessGender!C:C,"No Data"),IF(LEFT(C{i + 2},3)="Mr.","Male",IF(LEFT(C{i + 2},3)="Mr ","Male",IF(LEFT(C{i + 2},3)="Rev","Male",IF(LEFT(C{i + 2},4)="Miss","Female",IF(LEFT(C{i + 2},3)="Ms.","Female",IF(LEFT(C{i + 2},4)="Mrs.","Female"))))))' for i in range(len(margin_trading_df))]
    margin_trading_df['TOP50'] = [f'=IFNA(VLOOKUP(D{i + 2},\'Top50\'!A:E,5,0),"Normal")' for i in range(len(margin_trading_df))]

    return margin_trading_df