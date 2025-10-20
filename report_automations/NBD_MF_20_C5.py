import os
from pathlib import Path
import openpyxl
import sys

def get_month_year_from_filename(filename):
    # Example: "NBD-MF-20-C1 to C6 July 2025.xlsx"
    parts = filename.split()
    for i, part in enumerate(parts):
        if part in {"Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","January","February","March","April","May","June","July","August","September","October","November","December"}:
            month = part
            year = parts[i+1].replace(".xlsx","")
            return month, year
    return None, None

def find_files_safely(base_dir, pattern, description):
    """Safely find files and provide detailed error messages"""
    print(f"Looking for {description} in: {base_dir}")
    print(f"Search pattern: {pattern}")
    
    files = list(base_dir.glob(pattern))
    print(f"Found {len(files)} files matching pattern:")
    
    if not files:
        print(f"No files found matching pattern: {pattern}")
        print("Available files in directory:")
        try:
            all_files = [f.name for f in base_dir.iterdir() if f.is_file() and f.suffix == '.xlsx']
            if all_files:
                for file in sorted(all_files):
                    print(f"  - {file}")
            else:
                print("  No .xlsx files found in directory")
        except Exception as e:
            print(f"  Error listing files: {e}")
        return None
    
    for file in files:
        print(f"  {file.name}")
    
    return files[0]  # Return the first match

def main():
    # Get the base directory (parent of report_automations)
    base_dir = Path(__file__).resolve().parent.parent
    
    # New folder structure: working/NBD_MF_20_C1_C6/<dated-folder>/
    working_dir = base_dir / "working"
    c1_c6_working_dir = working_dir / "NBD_MF_20_C1_C6"
    
    print(f"Base directory: {base_dir}")
    print(f"Working directory: {working_dir}")
    print(f"C1-C6 working directory: {c1_c6_working_dir}")
    print(f"Current working directory: {Path.cwd()}")
    
    # Search only in the new C1-C6 working directory structure
    search_dirs = [
        c1_c6_working_dir
    ]
    
    file_c1_c6 = None
    
    # Search for C1-C6 file in the new structure only
    if not c1_c6_working_dir.exists():
        print(f"\nError: C1-C6 working directory does not exist: {c1_c6_working_dir}")
        print("Please ensure the working/NBD_MF_20_C1_C6 directory exists with dated folders.")
        return
    
    print(f"\nSearching in C1-C6 working directory: {c1_c6_working_dir}")
    
    # Look for dated folders in C1-C6 working directory
    dated_folders = [d for d in c1_c6_working_dir.iterdir() if d.is_dir()]
    print(f"Found {len(dated_folders)} dated folders: {[d.name for d in dated_folders]}")
    
    if not dated_folders:
        print("No dated folders found in C1-C6 working directory.")
        return
    
    for dated_folder in dated_folders:
        print(f"\nSearching in dated folder: {dated_folder}")
        
        # Look for the specific file in this dated folder
        c1_c6_pattern = "NBD-MF-20-C1 to C6*.xlsx"
        files = list(dated_folder.glob(c1_c6_pattern))
        
        if files:
            file_c1_c6 = files[0]
            print(f"Found C1-C6 file: {file_c1_c6}")
            break
    
    if file_c1_c6 is None:
        print("\nCould not find C1-C6 file in any dated folder.")
        print("Looking for any Excel files in dated folders...")
        for dated_folder in dated_folders:
            xlsx_files = list(dated_folder.glob("*.xlsx"))
            if xlsx_files:
                print(f"Excel files in {dated_folder.name}:")
                for f in xlsx_files:
                    print(f"  - {f.name}")
        return
    
    # Extract month and year from filename
    month, year = get_month_year_from_filename(file_c1_c6.name)
    if not month or not year:
        print(f"Could not parse month/year from filename: {file_c1_c6.name}")
        print("Expected format: 'NBD-MF-20-C1 to C6 [Month] [Year].xlsx'")
        return
    
    print(f"\nParsed month: {month}, year: {year}")
    
    # Find the dated folder that contains our file (for logging purposes)
    working_folder_name = None
    for dated_folder in dated_folders:
        # Check if this folder contains our target files
        c1_c6_files = list(dated_folder.glob("NBD-MF-20-C1 to C6*.xlsx"))
        if c1_c6_files:
            working_folder_name = dated_folder.name
            print(f"Found working folder: {working_folder_name}")
            break
    
    try:
        # Load workbook
        print(f"\nLoading workbook...")
        wb_c1_c6 = openpyxl.load_workbook(file_c1_c6, data_only=True)
        print(f"Workbook loaded successfully")
        
        # Check if required sheets exist
        print(f"\nAvailable sheets in C1-C6 file: {wb_c1_c6.sheetnames}")
        
        # Check for C5 sheet in C1-C6 file
        if "NBD_NBL-MF-20-C5" not in wb_c1_c6.sheetnames:
            print(f"Required sheet 'NBD_NBL-MF-20-C5' not found in C1-C6 file")
            print(f"Available sheets: {wb_c1_c6.sheetnames}")
            return
        
        print(f"Required sheet found")
        
        # Get worksheet
        tgt_ws = wb_c1_c6["NBD_NBL-MF-20-C5"]
        
        print(f"\nProcessing data...")
        
        # Helper function to safely get numeric values
        def get_numeric_value(worksheet, cell_ref, description=""):
            """Safely extract numeric value from a cell, handling text/None values"""
            try:
                value = worksheet[cell_ref].value
                print(f"{description} ({cell_ref}): {value}")
                
                if value is None:
                    return 0
                elif isinstance(value, (int, float)):
                    return value
                elif isinstance(value, str):
                    # Try to extract number from string if possible
                    try:
                        # Remove common text and try to parse number
                        clean_value = value.replace(',', '').strip()
                        return float(clean_value)
                    except ValueError:
                        print(f"  Warning: Cell {cell_ref} contains text '{value}' - using 0")
                        return 0
                else:
                    print(f"  Warning: Cell {cell_ref} contains unexpected type {type(value)} - using 0")
                    return 0
            except Exception as e:
                print(f"  Error reading {cell_ref}: {e} - using 0")
                return 0
        
        # Read "Guarantees" value from Master_Data.xlsx → NBD-MF-20-C1-C6 sheet (last row, column B)
        print(f"\nFetching 'Guarantees' value from Master_Data.xlsx (sheet: NBD-MF-20-C1-C6)...")
        guarantees_value = None
        try:
            # Master_Data.xlsx is at project root (parent of report_automations)
            project_root = Path(__file__).resolve().parents[1]
            master_path = project_root / "Master_Data.xlsx"
            if not master_path.exists():
                print(f"Master_Data.xlsx not found at {master_path}")
            else:
                md_wb = openpyxl.load_workbook(master_path, data_only=True)
                if "NBD-MF-20-C1-C6" in md_wb.sheetnames:
                    md_ws = md_wb["NBD-MF-20-C1-C6"]
                    if md_ws.max_row >= 2:
                        last_row = md_ws.max_row
                        val = md_ws[f"B{last_row}"] .value
                        if val is not None and str(val).strip() != "":
                            try:
                                guarantees_value = float(str(val).replace(",", "").strip())
                            except Exception:
                                print(f"Warning: Could not convert guarantees '{val}' to number; using 0")
                                guarantees_value = 0.0
                        else:
                            print("Warning: Latest guarantees cell is empty; using 0")
                            guarantees_value = 0.0
                    else:
                        print("Warning: No data rows in NBD-MF-20-C1-C6 sheet; using 0")
                        guarantees_value = 0.0
                    md_wb.close()
                else:
                    print("Warning: Sheet 'NBD-MF-20-C1-C6' not found in Master_Data.xlsx; using 0")
                    guarantees_value = 0.0
        except Exception as e:
            print(f"Error reading Master_Data.xlsx: {e}")
            guarantees_value = 0.0
        
        # Find "Guarantees" row in C5 sheet
        print(f"\nSearching for 'Guarantees' in C5 sheet...")
        
        guarantees_row = None
        for row in range(1, tgt_ws.max_row + 1):
            for col in range(1, tgt_ws.max_column + 1):
                cell_value = tgt_ws.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str) and "Guarantees" in cell_value:
                    print(f"Found 'Guarantees' at row {row}, column {col}")
                    guarantees_row = row
                    break
            if guarantees_row:
                break
        
        if guarantees_row is None:
            print("Could not find 'Guarantees' row in C5 sheet")
            print("Available cell values in C5 sheet:")
            for row in range(1, min(20, tgt_ws.max_row + 1)):  # Show first 20 rows
                for col in range(1, min(10, tgt_ws.max_column + 1)):  # Show first 10 columns
                    cell_value = tgt_ws.cell(row=row, column=col).value
                    if cell_value:
                        print(f"  Row {row}, Col {col}: {cell_value}")
            return
        
        # Update the Guarantees value
        # Assuming the value goes in column C (3rd column)
        guarantees_cell = f"C{guarantees_row}"
        print(f"Updating cell {guarantees_cell} with value from Master_Data.xlsx: {guarantees_value}")
        tgt_ws[guarantees_cell] = guarantees_value
        
        print(f"Data processing completed")
        print(f"Updated C5 sheet:")
        print(f"  - Guarantees: {guarantees_value} (from Master_Data.xlsx)")
        
        # Save updated workbook back to the same location (like Ctrl+S)
        wb_c1_c6.save(file_c1_c6)
        print(f"\nReport saved successfully to original location: {file_c1_c6}")
        
    except FileNotFoundError as e:
        print(f"File not found: {e}")
    except PermissionError as e:
        print(f"Permission error (file might be open in Excel): {e}")
    except Exception as e:
        print(f"An error occurred: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description='NBD-MF-20-C5 Report Automation')
    parser.add_argument('--working-dir', type=str, help='Working directory (not used, for compatibility)')
    parser.add_argument('--month', type=str, help='Report month (e.g., Jul)')
    parser.add_argument('--year', type=str, help='Report year (e.g., 2025)')
    args = parser.parse_args()

    main()
