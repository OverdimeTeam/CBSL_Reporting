import os
import logging
from datetime import datetime
from dateutil.relativedelta import relativedelta
from pathlib import Path
from typing import Any, Dict, Optional, List ,Tuple
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Optional
import shutil

# Configure logging
def setup_logging():
    """Setup detailed logging for the automation script."""
    log_dir = Path(__file__).parent / "logs"
    log_dir.mkdir(exist_ok=True)
    
    # Create log filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = log_dir / f"NBD_MF_01_SOFP_SOCI_{timestamp}.log"
    
    # Configure logging format
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(funcName)s:%(lineno)d - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler()  # Also log to console
        ]
    )
    
    logger = logging.getLogger(__name__)
    logger.info(f"Logging initialized. Log file: {log_file}")
    return logger

# Initialize logger
logger = setup_logging()


def _find_single_subdirectory(parent: Path) -> Optional[Path]:
    """
    Return the only subdirectory inside `parent`, or None if not exactly one.
    """
    logger.info(f"Looking for a single subdirectory in: {parent}")

    if not parent.exists() or not parent.is_dir():
        logger.warning(f"Parent directory does not exist or is not a directory: {parent}")
        return None

    subdirs = [p for p in parent.iterdir() if p.is_dir()]
    if len(subdirs) != 1:
        logger.error(f"Expected exactly 1 subdirectory, found {len(subdirs)}")
        return None

    logger.info(f"Found single subdirectory: {subdirs[0]}")
    return subdirs[0]

def _find_soci_workbook(root_dir: Path) -> Optional[Path]:
    """
    Locate the SOCI workbook under the given root_dir.
    Searches recursively for a file containing 'NBD-MF-01-SOFP & SOCI AFL Monthly FS' in its name.
    """
    logger.info(f"Searching for SOCI workbook under: {root_dir}")

    identifying_fragment = "NBD-MF-01-SOFP & SOCI AFL Monthly FS"
    found_files = []

    for root, _, files in os.walk(root_dir):
        for fname in files:
            if fname.lower().endswith(".xlsx") and identifying_fragment.lower() in fname.lower():
                file_path = Path(root) / fname
                found_files.append(file_path)
                logger.info(f"Found matching SOCI file: {file_path}")

    if not found_files:
        logger.error(f"No SOCI workbook found in {root_dir} containing '{identifying_fragment}'")
        return None

    if len(found_files) > 1:
        logger.warning(f"Multiple SOCI workbooks found: {found_files}")
        logger.info(f"Using first match: {found_files[0]}")

    selected_file = found_files[0]
    logger.info(f"Selected SOCI workbook: {selected_file}")
    return selected_file

def backup_column_E_before_step1() -> Optional[pd.DataFrame]:
    """
    STEP 0: Backup Column E data from Linked TB sheet before any modifications.

    Process:
    1. Find the SOCI workbook
    2. Read Linked TB sheet
    3. Extract Column E data (Account Balance)
    4. Store in DataFrame with Account # for reference
    5. Log first 10 rows
    
    Returns:
        DataFrame with Account # and Column E values, or None if failed
    """
    logger.info("="*50)
    logger.info("STEP 0: Backup Column E Data from Linked TB")
    logger.info("="*50)
    
    try:
        # --- Locate project structure ---
        project_root = Path(__file__).resolve().parents[1]
        working_soci_dir = project_root / "working" / "NBD_MF_01_SOFP_SOCI"
        logger.info(f"Project root: {project_root}")
        logger.info(f"Working SOCI directory: {working_soci_dir}")

        # --- Find the single dated folder inside NBD_MF_01_SOFP_SOCI ---
        date_folder = _find_single_subdirectory(working_soci_dir)
        if date_folder is None:
            logger.error(f"No date folder found inside {working_soci_dir}")
            return None

        logger.info(f"Using dated folder: {date_folder}")

        # --- Find the SOCI workbook inside that folder ---
        workbook_path = _find_soci_workbook(date_folder)
        if workbook_path is None or not workbook_path.exists():
            logger.error(f"SOCI workbook not found in {date_folder}")
            return None

        logger.info(f"SOCI workbook found: {workbook_path}")
        wb = load_workbook(workbook_path)

        sheet_name = "Linked TB"
        logger.info(f"Target sheet: {sheet_name}")
        logger.info(f"Reading from workbook: {workbook_path}")

        # Load the workbook with openpyxl to read exact values
        logger.info("Loading workbook with openpyxl")
        wb = load_workbook(workbook_path, data_only=True)
        
        if sheet_name not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found in workbook")
            logger.info(f"Available sheets: {wb.sheetnames}")
            return None

        ws = wb[sheet_name]
        logger.info(f"Successfully opened sheet: {sheet_name}")

        # Read data from columns A (Account #) and E (Balance)
        data = []
        rows_processed = 0
        
        logger.info(f"Reading data from row 2 to {ws.max_row} (skipping header row 1)")
        
        for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
            account_num = ws.cell(row=row_idx, column=1).value  # Column A
            column_e_value = ws.cell(row=row_idx, column=5).value  # Column E
            
            # Only add rows that have an Account # (skip empty rows)
            if account_num is not None:
                data.append({
                    'Row': row_idx,
                    'Account #': account_num,
                    'Column E Value': column_e_value
                })
                rows_processed += 1

        # Create DataFrame
        df_backup = pd.DataFrame(data)
        
        logger.info(f"Successfully backed up {len(df_backup)} rows from Column E")
        logger.info(f"Total rows processed: {rows_processed}")
        
        # Log first 10 rows
        logger.info("\n" + "="*50)
        logger.info("FIRST 10 ROWS OF COLUMN E BACKUP:")
        logger.info("="*50)
        if not df_backup.empty:
            logger.info("\n" + df_backup.head(10).to_string(index=False))
        else:
            logger.warning("DataFrame is empty - no data to display")
        
        # Log summary statistics
        logger.info("\n" + "="*50)
        logger.info("COLUMN E BACKUP SUMMARY:")
        logger.info("="*50)
        logger.info(f"Total rows: {len(df_backup)}")
        logger.info(f"Non-null values in Column E: {df_backup['Column E Value'].notna().sum()}")
        logger.info(f"Null values in Column E: {df_backup['Column E Value'].isna().sum()}")

        if df_backup['Column E Value'].notna().any():
            # Convert to numeric for statistics (ignore non-numeric values)
            numeric_values = pd.to_numeric(df_backup['Column E Value'], errors='coerce')
            logger.info(f"Sum of Column E values: {numeric_values.sum()}")
            logger.info(f"Min value: {numeric_values.min()}")
            logger.info(f"Max value: {numeric_values.max()}")
        
        logger.info("="*50)
        logger.info("STEP 0 COMPLETED SUCCESSFULLY")
        logger.info("="*50)
        
        return df_backup
        
    except Exception as e:
        logger.error(f"Step 0 failed with exception: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None
 
def read_column_e_from_sofp() -> Optional[Dict[str, Any]]:
    """
    STEP 0a: Read Column E data from NBD-MF-01-SOFP sheet (NO PASTE YET).
    
    Returns:
        Dictionary containing workbook_path, e_values, start_row, last_filled_row
    """
    logger.info("=" * 50)
    logger.info("STEP 0a: Read Column E from SOFP (Backup Phase)")
    logger.info("=" * 50)

    try:
        # --- Locate project structure ---
        project_root = Path(__file__).resolve().parents[1]
        working_soci_dir = project_root / "working" / "NBD_MF_01_SOFP_SOCI"
        logger.info(f"Project root: {project_root}")
        logger.info(f"Working SOCI directory: {working_soci_dir}")

        # --- Find the single dated folder inside working/NBD_MF_01_SOFP_SOCI ---
        date_folder = _find_single_subdirectory(working_soci_dir)
        if date_folder is None or not date_folder.exists():
            logger.error(f"Could not find the dated folder under {working_soci_dir}")
            return None

        logger.info(f"Dated folder found: {date_folder}")

        # --- Find the SOCI workbook (includes SOFP sheet) inside that folder ---
        workbook_path = _find_soci_workbook(date_folder)
        if workbook_path is None or not workbook_path.exists():
            logger.error(f"Could not find SOCI workbook in {date_folder}")
            return None

        logger.info(f"SOCI workbook found: {workbook_path}")
        wb = load_workbook(workbook_path)

        sheet_name = "NBD-MF-01-SOFP"
        logger.info(f"Target workbook: {workbook_path.name}")
        logger.info(f"Target sheet: {sheet_name}")

        # Read column E values (starting from row 3)
        logger.info("Reading column E values (values only)")
        wb_read = load_workbook(workbook_path, data_only=True)
        if sheet_name not in wb_read.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found in workbook")
            logger.info(f"Available sheets: {wb_read.sheetnames}")
            wb_read.close()
            return None

        ws_read = wb_read[sheet_name]
        start_row = 3
        col_a, col_e = 1, 5

        # Find last filled row in column A
        last_filled_row = None
        for r in range(start_row, ws_read.max_row + 1):
            val = ws_read.cell(row=r, column=col_a).value
            if val is not None and (not isinstance(val, str) or val.strip()):
                last_filled_row = r

        if not last_filled_row:
            logger.error("No filled rows found in column A.")
            wb_read.close()
            return None

        e_values = [
            ws_read.cell(row=r, column=col_e).value
            for r in range(start_row, last_filled_row + 1)
        ]
        wb_read.close()
        logger.info(f"Read {len(e_values)} values from column E (rows {start_row}-{last_filled_row})")

        logger.info("=" * 50)
        logger.info("STEP 0a (SOFP READ) COMPLETED SUCCESSFULLY")
        logger.info("=" * 50)
        
        return {
            'workbook_path': workbook_path,
            'e_values': e_values,
            'start_row': start_row,
            'last_filled_row': last_filled_row,
            'sheet_name': sheet_name
        }

    except Exception as e:
        logger.error(f"Error in read_column_e_from_sofp: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None


def read_column_e_from_soci() -> Optional[Dict[str, Any]]:
    """
    STEP 0b: Read Column E data from NBD-MF-02-SOCI sheet (NO PASTE YET).
    
    Returns:
        Dictionary containing workbook_path, e_values, start_row
    """
    logger.info("="*50)
    logger.info("STEP 0b: Read Column E from SOCI (Backup Phase)")
    logger.info("="*50)
    
    try:
        # --- Locate project structure ---
        project_root = Path(__file__).resolve().parents[1]
        working_soci_dir = project_root / "working" / "NBD_MF_01_SOFP_SOCI"
        logger.info(f"Project root: {project_root}")
        logger.info(f"Working SOCI directory: {working_soci_dir}")

        # --- Find the single dated folder inside NBD_MF_01_SOFP_SOCI ---
        date_folder = _find_single_subdirectory(working_soci_dir)
        if date_folder is None or not date_folder.exists():
            logger.error(f"Could not find the dated folder under {working_soci_dir}")
            return None

        logger.info(f"Dated folder found: {date_folder}")

        # --- Find the SOCI workbook inside that dated folder ---
        workbook_path = _find_soci_workbook(date_folder)
        if workbook_path is None or not workbook_path.exists():
            logger.error(f"Could not find SOCI workbook in {date_folder}")
            return None

        logger.info(f"SOCI workbook found: {workbook_path}")
        wb = load_workbook(workbook_path)

        sheet_name = "NBD-MF-02-SOCI"
        logger.info(f"Target sheet: {sheet_name}")

        # Read column E values using pandas
        logger.info("Reading column E values using pandas")
        df_e = pd.read_excel(
            workbook_path,
            sheet_name=sheet_name,
            usecols="E",
            header=None,
            skiprows=2,
            engine="openpyxl",
        )
        logger.info(f"Successfully read {len(df_e)} rows from column E")

        # Convert to list (including NaN values)
        e_values: List[object] = df_e.iloc[:, 0].tolist()
        logger.info(f"Converted to list with {len(e_values)} values")

        logger.info("="*50)
        logger.info("STEP 0b (SOCI READ) COMPLETED SUCCESSFULLY")
        logger.info("="*50)
        
        return {
            'workbook_path': workbook_path,
            'e_values': e_values,
            'start_row': 3,
            'sheet_name': sheet_name
        }
        
    except Exception as e:
        logger.error(f"Error in read_column_e_from_soci: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None

def paste_to_sofp(sofp_data: Dict[str, Any]) -> bool:
    """
    STEP 1a: Paste Column E values to Column G in SOFP sheet.
    Also updates dates in columns D and G (row 2 only).
    
    Args:
        sofp_data: Dictionary containing workbook_path, e_values, start_row, etc.
    
    Returns:
        True if successful, False otherwise
    """
    logger.info("=" * 50)
    logger.info("STEP 1a: Paste Column E → Column G - SOFP")
    logger.info("=" * 50)

    try:
        workbook_path = sofp_data['workbook_path']
        e_values = sofp_data['e_values']
        start_row = sofp_data['start_row']
        last_filled_row = sofp_data['last_filled_row']
        sheet_name = sofp_data['sheet_name']

        # Load workbook
        logger.info("Loading workbook")
        wb_write = load_workbook(workbook_path)
        ws_write = wb_write[sheet_name]

        # Update dates in columns D and G (row 2 only)
        logger.info("Updating dates in columns D and G (row 2 only - increasing month by 1)")
        row_idx = 2
        dates_updated = 0
        
        # Update column D (row 2)
        d_cell = ws_write.cell(row=row_idx, column=4)  # Column D
        if d_cell.value:
            try:
                if isinstance(d_cell.value, datetime):
                    current_date = d_cell.value
                    new_date = current_date + relativedelta(months=1)
                    d_cell.value = new_date
                    logger.info(f"Row {row_idx}, Col D: Updated date from {current_date.strftime('%d/%m/%Y')} to {new_date.strftime('%d/%m/%Y')}")
                    dates_updated += 1
                elif isinstance(d_cell.value, str):
                    try:
                        current_date = datetime.strptime(d_cell.value, "%d/%m/%Y")
                        new_date = current_date + relativedelta(months=1)
                        d_cell.value = new_date
                        logger.info(f"Row {row_idx}, Col D: Updated date from {current_date.strftime('%d/%m/%Y')} to {new_date.strftime('%d/%m/%Y')}")
                        dates_updated += 1
                    except ValueError:
                        pass
            except (ValueError, TypeError):
                pass

        # Update column G (row 2)
        g_cell = ws_write.cell(row=row_idx, column=7)  # Column G
        if g_cell.value:
            try:
                if isinstance(g_cell.value, datetime):
                    current_date = g_cell.value
                    new_date = current_date + relativedelta(months=1)
                    g_cell.value = new_date
                    logger.info(f"Row {row_idx}, Col G: Updated date from {current_date.strftime('%d/%m/%Y')} to {new_date.strftime('%d/%m/%Y')}")
                    dates_updated += 1
                elif isinstance(g_cell.value, str):
                    try:
                        current_date = datetime.strptime(g_cell.value, "%d/%m/%Y")
                        new_date = current_date + relativedelta(months=1)
                        g_cell.value = new_date
                        logger.info(f"Row {row_idx}, Col G: Updated date from {current_date.strftime('%d/%m/%Y')} to {new_date.strftime('%d/%m/%Y')}")
                        dates_updated += 1
                    except ValueError:
                        pass
            except (ValueError, TypeError):
                pass

        logger.info(f"Updated {dates_updated} dates in row 2 (columns D and G)")

        # Write values to column G (preserve formulas)
        logger.info("Writing values to column G")
        col_g = 7
        written_count = 0
        skipped_formulas = 0

        for idx, value in enumerate(e_values, start=start_row):
            g_cell = ws_write.cell(row=idx, column=col_g)
            # Skip if it's a formula cell
            if isinstance(g_cell.value, str) and g_cell.value.startswith("="):
                skipped_formulas += 1
                logger.debug(f"Row {idx}: Skipping formula cell")
                continue
            g_cell.value = value
            written_count += 1
            logger.debug(f"Row {idx}: Set G = {value}")

        wb_write.save(workbook_path)
        wb_write.close()
        logger.info(f"Pasted {written_count} values into Column G (rows {start_row}-{last_filled_row})")
        logger.info(f"Skipped {skipped_formulas} formula cells in Column G")

        logger.info("=" * 50)
        logger.info("STEP 1a (SOFP PASTE) COMPLETED SUCCESSFULLY")
        logger.info("=" * 50)
        return True

    except Exception as e:
        logger.error(f"Error in paste_to_sofp: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def paste_to_soci(soci_data: Dict[str, Any]) -> bool:
    """
    STEP 1b: Paste Column E values to Column G in SOCI sheet.
    Also updates dates in columns D and G.
    
    Args:
        soci_data: Dictionary containing workbook_path, e_values, start_row
    
    Returns:
        True if successful, False otherwise
    """
    logger.info("="*50)
    logger.info("STEP 1b: Paste Column E → Column G - SOCI")
    logger.info("="*50)
    
    try:
        workbook_path = soci_data['workbook_path']
        e_values = soci_data['e_values']
        start_row = soci_data['start_row']
        sheet_name = soci_data['sheet_name']

        # Load workbook with openpyxl to write values while preserving formatting
        logger.info("Loading workbook with openpyxl")
        wb = load_workbook(workbook_path)

        if sheet_name not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found in workbook")
            logger.info(f"Available sheets: {wb.sheetnames}")
            return False

        ws = wb[sheet_name]
        logger.info(f"Working with sheet: {sheet_name}")

        # Update dates in columns D and G (increase month by 1)
        logger.info("Updating dates in columns D and G (increasing month by 1)")

        dates_updated = 0
        for row_idx in range(1, ws.max_row + 1):
            # Update column D
            d_cell = ws.cell(row=row_idx, column=4)  # Column D
            if d_cell.value:
                try:
                    if isinstance(d_cell.value, datetime):
                        current_date = d_cell.value
                        new_date = current_date + relativedelta(months=1)
                        d_cell.value = new_date
                        logger.info(f"Row {row_idx}, Col D: Updated date from {current_date.strftime('%d/%m/%Y')} to {new_date.strftime('%d/%m/%Y')}")
                        dates_updated += 1
                    elif isinstance(d_cell.value, str):
                        try:
                            from datetime import datetime as dt
                            current_date = dt.strptime(d_cell.value, "%d/%m/%Y")
                            new_date = current_date + relativedelta(months=1)
                            d_cell.value = new_date
                            logger.info(f"Row {row_idx}, Col D: Updated date from {current_date.strftime('%d/%m/%Y')} to {new_date.strftime('%d/%m/%Y')}")
                            dates_updated += 1
                        except ValueError:
                            pass
                except (ValueError, TypeError):
                    pass

            # Update column G
            g_cell = ws.cell(row=row_idx, column=7)  # Column G
            if g_cell.value:
                try:
                    if isinstance(g_cell.value, datetime):
                        current_date = g_cell.value
                        new_date = current_date + relativedelta(months=1)
                        g_cell.value = new_date
                        logger.info(f"Row {row_idx}, Col G: Updated date from {current_date.strftime('%d/%m/%Y')} to {new_date.strftime('%d/%m/%Y')}")
                        dates_updated += 1
                    elif isinstance(g_cell.value, str):
                        try:
                            from datetime import datetime as dt
                            current_date = dt.strptime(g_cell.value, "%d/%m/%Y")
                            new_date = current_date + relativedelta(months=1)
                            g_cell.value = new_date
                            logger.info(f"Row {row_idx}, Col G: Updated date from {current_date.strftime('%d/%m/%Y')} to {new_date.strftime('%d/%m/%Y')}")
                            dates_updated += 1
                        except ValueError:
                            pass
                except (ValueError, TypeError):
                    pass

        logger.info(f"Updated {dates_updated} dates in columns D and G")

        # Process column G starting at row 3
        end_row = start_row + len(e_values) - 1 if e_values else start_row - 1
        logger.info(f"Processing rows {start_row} to {end_row}")

        # First pass: clear existing non-formula values in column G
        logger.info("First pass: clearing non-formula values in column G")
        cleared_count = 0
        formula_count = 0
        
        for row in range(start_row, end_row + 1):
            g_cell = ws.cell(row=row, column=7)  # column G
            if isinstance(g_cell.value, str) and g_cell.value.startswith("="):
                formula_count += 1
                logger.debug(f"Row {row}: Preserving formula in G: {g_cell.value}")
                continue
            g_cell.value = None
            cleared_count += 1
        
        logger.info(f"Cleared {cleared_count} cells, preserved {formula_count} formulas")

        # Second pass: write values from E into G
        logger.info("Second pass: writing E values to G")
        written_count = 0
        skipped_count = 0
        
        for idx, value in enumerate(e_values, start=0):
            row = start_row + idx
            g_cell = ws.cell(row=row, column=7)  # column G
            
            # Skip if cell has a formula
            if isinstance(g_cell.value, str) and g_cell.value.startswith("="):
                skipped_count += 1
                logger.debug(f"Row {row}: Skipping formula cell")
                continue
            
            # Write value (convert NaN to None)
            new_value = None if pd.isna(value) else value
            g_cell.value = new_value
            written_count += 1
            logger.debug(f"Row {row}: Set G = {new_value} (from E = {value})")
        
        logger.info(f"Written {written_count} values, skipped {skipped_count} formula cells")

        # Save the workbook
        logger.info("Saving workbook")
        wb.save(workbook_path)
        wb.close()
        logger.info(f"Workbook saved successfully: {workbook_path}")

        logger.info("="*50)
        logger.info("STEP 1b (SOCI PASTE) COMPLETED SUCCESSFULLY")
        logger.info("="*50)
        return True
        
    except Exception as e:
        logger.error(f"Error in paste_to_soci: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False


def find_tb_detail_report(dated_folder: Path) -> Optional[Path]:
    """
    Find the TB Detail Report file directly inside the given dated folder.

    Args:
        dated_folder: Path to the dated folder, e.g.,
                      C:/CBSL/Script/working/NBD_MF_01_SOFP_SOCI/30-09-2025

    Returns:
        Path to the TB Detail Report file, or None if not found.
    """
    logger.info(f"Searching for TB Detail Report in: {dated_folder}")

    if not dated_folder.exists() or not dated_folder.is_dir():
        logger.error(f"Folder does not exist: {dated_folder}")
        return None

    keyword = "TB_DetailReport"
    candidate_files = [
        p for p in dated_folder.iterdir()
        if p.is_file() and keyword in p.name and p.suffix.lower() == ".xlsx"
    ]

    if not candidate_files:
        logger.error(f"No TB Detail Report found in {dated_folder}")
        return None

    # Return the latest by modification time
    latest_file = max(candidate_files, key=lambda p: p.stat().st_mtime)
    logger.info(f"TB Detail Report selected: {latest_file}")
    return latest_file


def read_tb_details_with_win32(file_path: Path) -> Optional[pd.DataFrame]:
    """
    Final fallback function to read TB Detail Report using Win32 COM when all other methods fail.

    Args:
        file_path: Path to the TB Detail Excel file

    Returns:
        DataFrame with columns: Account #, Account Name, Net Balance
    """
    logger.info(f"Attempting Win32 COM to read TB Detail Report from: {file_path}")

    try:
        import win32com.client

        # Create Excel application
        excel_app = win32com.client.Dispatch("Excel.Application")
        excel_app.Visible = False
        excel_app.DisplayAlerts = False

        # Open workbook
        wb = excel_app.Workbooks.Open(
        str(file_path),
        ReadOnly=True
        )
        #ws = wb.Sheets(1)  # First sheet
        ws = wb.Sheets("TB - Details")

        # Read data from columns I, J, K starting from row 6
        data = []
        row_num = 6

        while True:
            # Get values from columns I (9), J (10), K (11)
            account_num = ws.Cells(row_num, 9).Value
            account_name = ws.Cells(row_num, 10).Value
            net_balance = ws.Cells(row_num, 11).Value

            # Stop if all values are None (empty row)
            if account_num is None and account_name is None and net_balance is None:
                # Check next few rows to see if there's more data
                found_more_data = False
                for check_row in range(row_num + 1, row_num + 10):
                    check_val = ws.Cells(check_row, 9).Value
                    if check_val is not None:
                        found_more_data = True
                        break

                if not found_more_data:
                    break
                else:
                    # Skip this empty row but continue
                    row_num += 1
                    continue

            data.append({
                'Account #': account_num,
                'Account Name': account_name,
                'Net Balance': net_balance
            })

            row_num += 1

            # Prevent infinite loop - stop at 10000 rows
            if row_num > 10000:
                break

        # Close workbook and Excel
        wb.Close(False)
        excel_app.Quit()

        df = pd.DataFrame(data)
        # Remove completely empty rows
        df = df.dropna(how='all')

        logger.info(f"Successfully read {len(df)} rows from TB Detail Report using Win32 COM")
        return df

    except Exception as e:
        logger.error(f"Win32 COM approach also failed: {e}")
        # Try to clean up Excel if it was opened
        try:
            excel_app.Quit()
        except:
            pass
        return None


def read_tb_details_with_pandas(file_path: Path) -> Optional[pd.DataFrame]:
    """
    Fallback function to read TB Detail Report using pandas when openpyxl fails.

    Args:
        file_path: Path to the TB Detail Excel file

    Returns:
        DataFrame with columns: Account #, Account Name, Net Balance
    """
    logger.info(f"Using pandas to read TB Detail Report from: {file_path}")

    try:
        # First attempt: pandas with default engine (openpyxl)
        try:
             df_raw = pd.read_excel(
                file_path, 
                header=None, 
                skiprows=5, 
                engine='openpyxl',
                engine_kwargs={'data_only': True, 'read_only': True}
            )
        except Exception as e1:
            logger.warning(f"pandas with default engine failed: {e1}. Trying xlrd engine")
            try:
                # Fallback: pandas with xlrd engine for .xls files
                df_raw = pd.read_excel(file_path, header=None, skiprows=5, engine='xlrd')
            except Exception as e2:
                logger.warning(f"xlrd engine also failed: {e2}. Trying calamine engine")
                try:
                    # Final fallback: pandas with calamine engine
                    df_raw = pd.read_excel(file_path, header=None, skiprows=5, engine='calamine')
                except Exception as e3:
                    logger.warning(f"calamine engine also failed: {e3}. Trying Win32 COM approach")
                    # Final attempt: try Win32 COM (Windows only)
                    return read_tb_details_with_win32(file_path)

        # Extract columns I, J, K (indices 8, 9, 10)
        if df_raw.shape[1] > 10:
            df_extracted = df_raw.iloc[:, [8, 9, 10]].copy()
            df_extracted.columns = ['Account #', 'Account Name', 'Net Balance']

            # Remove completely empty rows
            df_extracted = df_extracted.dropna(how='all')

            logger.info(f"Successfully read {len(df_extracted)} rows from TB Detail Report using pandas")
            return df_extracted
        else:
            logger.error(f"Excel file doesn't have enough columns. Found {df_raw.shape[1]} columns, need at least 11")
            return None

    except Exception as e:
        logger.error(f"Failed to read TB Detail Report with pandas: {e}")
        return None


def read_tb_details(file_path: Path) -> Optional[pd.DataFrame]:
    """
    Read TB Detail Report data from columns I, J, K starting from row 6.
    
    Args:
        file_path: Path to the TB Detail Report file
    
    Returns:
        DataFrame with columns: Account #, Account Name, Net Balance
    """
    logger.info(f"Reading TB Detail Report from: {file_path}")
    
    try:
        # Load the workbook with openpyxl to preserve exact formats
        from openpyxl import load_workbook

        # First attempt: normal loading
        try:
            wb = load_workbook(file_path, data_only=False)
            ws = wb.active  # First sheet (TB Details)
        except (TypeError, ValueError) as e:
            if "expected <class 'int'>" in str(e):
                logger.warning(f"Corrupted merged cells detected in {file_path}. Attempting to load with data_only=True")
                try:
                    # Fallback: load with data_only=True to ignore merged cell formatting issues
                    wb = load_workbook(file_path, data_only=True)
                    ws = wb.active
                except Exception as e2:
                    logger.warning(f"openpyxl loading failed: {e2}. Falling back to pandas Excel reader")
                    # Final fallback: use pandas to read the Excel file
                    return read_tb_details_with_pandas(file_path)
            else:
                raise e
        
        logger.info(f"Opened sheet: {ws.title}")
        
        # Read data starting from row 6, columns I, J, K
        data = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=6, min_col=9, max_col=11), start=6):
            # Column I (Account #), J (Account Name), K (Net Balance)
            account_num = row[0].value
            account_name = row[1].value
            net_balance = row[2].value
            
            # Skip completely empty rows
            if account_num is None and account_name is None and net_balance is None:
                continue
                
            data.append({
                'Account #': account_num,
                'Account Name': account_name,
                'Net Balance': net_balance
            })
        
        df = pd.DataFrame(data)
        
        logger.info(f"Successfully read {len(df)} rows from TB Detail Report")
        logger.info("First 10 rows:")
        logger.info("\n" + df.head(10).to_string(index=False))
        
        return df
        
    except Exception as e:
        logger.error(f"Error reading TB Detail Report: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None


def paste_to_system_tb(df: pd.DataFrame, workbook_path: Path, tb_file_path: Path) -> bool:
    """
    Paste TB data into System TB sheet starting from row 3.
    Copies values directly from source to preserve exact formatting.
    Extends VLOOKUP formulas in column D to match data rows.
    
    Args:
        df: DataFrame with Account #, Account Name, Net Balance
        workbook_path: Path to the SOCI workbook
        tb_file_path: Path to the TB Detail Report (source file)
    
    Returns:
        True if successful, False otherwise
    """
    logger.info("="*50)
    logger.info("Pasting data to System TB sheet")
    logger.info("="*50)
    
    try:
        # Load source workbook (TB Detail Report) to copy exact values and formats
        logger.info(f"Loading source TB Detail Report: {tb_file_path}")
        try:
            source_wb = load_workbook(tb_file_path)
            source_ws = source_wb.active
        except (TypeError, ValueError) as e:
            if "expected <class 'int'>" in str(e):
                logger.warning(f"Corrupted merged cells in source file {tb_file_path}. Using data from DataFrame instead")
                # Since we already have the data in DataFrame, we can skip loading the source file
                # and just use the DataFrame data directly
                source_wb = None
                source_ws = None
            else:
                raise e
        
        # Load target workbook
        logger.info(f"Loading target workbook: {workbook_path}")
        target_wb = load_workbook(workbook_path)
        
        sheet_name = "System TB"
        if sheet_name not in target_wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found in workbook")
            logger.info(f"Available sheets: {target_wb.sheetnames}")
            return False
            
        target_ws = target_wb[sheet_name]
        logger.info(f"Opened sheet: {sheet_name}")

        # Clear previous data in columns A, B, C starting from row 3
        max_row_to_clear = target_ws.max_row
        logger.info(f"Clearing columns A, B, C from row 3 to {max_row_to_clear}")
        
        for row in range(3, max_row_to_clear + 1):
            target_ws[f"A{row}"].value = None
            target_ws[f"B{row}"].value = None
            target_ws[f"C{row}"].value = None

        # Copy data directly from source to target (row 6 onwards from source to row 3 onwards in target)
        target_row = 3
        copied_count = 0

        logger.info(f"Copying data to target starting from row {target_row}")

        if source_wb is not None:
            # Use source workbook if available
            logger.info("Using source workbook data with formatting")
            for row_data in source_ws.iter_rows(min_row=6, min_col=9, max_col=11):
                # Column I, J, K from source
                source_account = row_data[0]
                source_name = row_data[1]
                source_balance = row_data[2]

                # Skip completely empty rows
                if source_account.value is None and source_name.value is None and source_balance.value is None:
                    continue

                # Copy value and number format from source to target
                # Column A (Account #)
                target_ws[f"A{target_row}"].value = source_account.value
                target_ws[f"A{target_row}"].number_format = source_account.number_format

                # Column B (Account Name)
                target_ws[f"B{target_row}"].value = source_name.value
                target_ws[f"B{target_row}"].number_format = source_name.number_format

                # Column C (Net Balance)
                target_ws[f"C{target_row}"].value = source_balance.value
                target_ws[f"C{target_row}"].number_format = source_balance.number_format

                target_row += 1
                copied_count += 1
        else:
            # Use DataFrame data if source workbook failed to load
            logger.info("Using DataFrame data (no formatting from source)")
            for _, row in df.iterrows():
                # Skip completely empty rows
                if pd.isna(row['Account #']) and pd.isna(row['Account Name']) and pd.isna(row['Net Balance']):
                    continue

                # Copy values from DataFrame to target (no formatting available)
                # Column A (Account #)
                target_ws[f"A{target_row}"].value = row['Account #']

                # Column B (Account Name)
                target_ws[f"B{target_row}"].value = row['Account Name']

                # Column C (Net Balance)
                target_ws[f"C{target_row}"].value = row['Net Balance']

                target_row += 1
                copied_count += 1
        
        last_data_row = target_row - 1
        logger.info(f"Copied {copied_count} rows to System TB")
        logger.info(f"Data written from row 3 to row {last_data_row}")

        # Now extend VLOOKUP formulas in column D
        logger.info("Extending VLOOKUP formulas in column D")
        
        # Find the first row with a VLOOKUP formula in column D (should be D3)
        formula_found = False
        base_formula = None
        base_row = None
        
        for row in range(3, last_data_row + 1):
            d_cell = target_ws[f"D{row}"]
            if d_cell.value and isinstance(d_cell.value, str) and d_cell.value.startswith("=VLOOKUP"):
                base_formula = d_cell.value
                base_row = row
                formula_found = True
                logger.info(f"Found VLOOKUP formula in D{row}: {base_formula}")
                break
        
        if formula_found and base_formula:
            import re
            
            # Extract the pattern from the formula
            # Example: =VLOOKUP(A2777,'Linked TB'!A:A,1,FALSE)
            # We need to replace A2777 with A{current_row}
            
            # Parse the formula to understand its structure
            logger.info("Extending formula to all data rows")
            
            for row in range(3, last_data_row + 1):
                d_cell = target_ws[f"D{row}"]
                
                # Replace the row reference in the formula
                # Pattern: A{number} should become A{current_row}
                new_formula = re.sub(r'A\d+', f'A{row}', base_formula)
                
                d_cell.value = new_formula
                
                # Copy number format from base cell if it exists
                if base_row:
                    base_d_cell = target_ws[f"D{base_row}"]
                    if base_d_cell.number_format:
                        d_cell.number_format = base_d_cell.number_format
            
            logger.info(f"Extended VLOOKUP formulas in column D from row 3 to row {last_data_row}")
            
        else:
            logger.warning("No VLOOKUP formula found in column D to extend")

        # Save the workbook
        logger.info("Saving workbook")
        target_wb.save(workbook_path)
        logger.info(f"Workbook saved successfully: {workbook_path}")
        
        return True

    except Exception as e:
        logger.error(f"Error in paste_to_system_tb: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False
    
def tb_detail_to_system_tb() -> bool:
    """
    STEP 2: Read TB Detail Report and paste data into System TB sheet.
    
    Process:
    1. Find the latest TB_DetailReport file
    2. Copy columns I, J, K from row 6 onwards directly with their formats
    3. Paste into System TB sheet columns A, B, C from row 3
    4. Do NOT modify column D
    
    Returns:
        True if successful, False otherwise
    """
    logger.info("="*50)
    logger.info("STEP 2: TB Detail Report → System TB")
    logger.info("="*50)
    
    try:
        # --- Locate project structure ---
        project_root = Path(__file__).resolve().parents[1]
        working_soci_dir = project_root / "working" / "NBD_MF_01_SOFP_SOCI"
        logger.info(f"Project root: {project_root}")
        logger.info(f"Working SOCI directory: {working_soci_dir}")

        # --- Find the single dated folder inside NBD_MF_01_SOFP_SOCI ---
        date_folder = _find_single_subdirectory(working_soci_dir)
        if date_folder is None or not date_folder.exists():
            logger.error(f"Could not find the dated folder under {working_soci_dir}")
            return False

        logger.info(f"Dated folder found: {date_folder}")

        # --- Find TB Detail Report in the dated folder ---
        tb_file = find_tb_detail_report(date_folder)
        if tb_file is None:
            logger.error(f"TB Detail Report not found in {date_folder}")
            return False

        logger.info(f"TB Detail Report found: {tb_file}")

        # --- Read TB data ---
        df_tb = read_tb_details(tb_file)
        if df_tb is None or df_tb.empty:
            logger.error("Failed to load TB Detail Report data or data is empty")
            return False

        # --- Find the SOCI workbook in the same dated folder ---
        workbook_path = _find_soci_workbook(date_folder)
        if workbook_path is None or not workbook_path.exists():
            logger.error(f"Could not find SOCI workbook in {date_folder}")
            return False

        logger.info(f"SOCI workbook found: {workbook_path}")
        wb = load_workbook(workbook_path)

        
        # Paste to System TB (pass tb_file path for direct copying)
        success = paste_to_system_tb(df_tb, workbook_path, tb_file)
        if not success:
            logger.error("Failed to paste data to System TB")
            return False
        
        logger.info("="*50)
        logger.info("STEP 2 COMPLETED SUCCESSFULLY")
        logger.info("="*50)
        return True
        
    except Exception as e:
        logger.error(f"Step 2 failed with exception: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False


def check_missing_accounts_in_linked_tb() -> bool:
    """
    STEP 3: Check which accounts from System TB are missing in Linked TB,
    and append them to the end of Linked TB table.

    Process:
    1. Load System TB and Linked TB sheets
    2. Create vlookup check column
    3. Identify accounts in System TB that don't exist in Linked TB AND have Net Balance > 0
    4. Append missing accounts to the end of Linked TB table (preserving formatting)
    5. Log the results

    Returns:
        True if successful, False otherwise
    """
    logger.info("="*50)
    logger.info("STEP 3: Check Missing Accounts in Linked TB and Append")
    logger.info("="*50)

    try:
        # --- Locate project structure ---
        project_root = Path(__file__).resolve().parents[1]
        working_soci_dir = project_root / "working" / "NBD_MF_01_SOFP_SOCI"

        logger.info(f"Project root: {project_root}")
        logger.info(f"Working SOCI directory: {working_soci_dir}")

        # --- Find the single dated folder inside NBD_MF_01_SOFP_SOCI ---
        date_folder = _find_single_subdirectory(working_soci_dir)
        if date_folder is None or not date_folder.exists():
            logger.error(f"Could not find dated folder under {working_soci_dir}")
            return False

        logger.info(f"Dated folder found: {date_folder}")

        # --- Find TB Detail Report inside the dated folder ---
        tb_file = find_tb_detail_report(date_folder)
        if tb_file is None:
            logger.error(f"TB Detail Report not found in {date_folder}")
            return False

        logger.info(f"TB Detail Report found: {tb_file}")

        # --- Read TB Detail data to get Account Names ---
        logger.info("Reading TB Detail Report to get Account Names")
        df_tb_detail = read_tb_details(tb_file)
        if df_tb_detail is None or df_tb_detail.empty:
            logger.error("Failed to load TB Detail Report data")
            return False

        # Create a dictionary mapping Account # -> Account Name from TB Detail
        tb_detail_map = {}
        for _, row in df_tb_detail.iterrows():
            acc_num = row['Account #']
            acc_name = row['Account Name']
            if pd.notna(acc_num):
                try:
                    acc_num_int = int(acc_num)
                    tb_detail_map[acc_num_int] = acc_name
                except (ValueError, TypeError):
                    tb_detail_map[acc_num] = acc_name

        logger.info(f"Created TB Detail map with {len(tb_detail_map)} account names")

        # --- Find the SOCI workbook inside the same dated folder ---
        workbook_path = _find_soci_workbook(date_folder)
        if workbook_path is None or not workbook_path.exists():
            logger.error(f"Could not find SOCI workbook in {date_folder}")
            return False

        logger.info(f"Loading workbook: {workbook_path}")
        wb = load_workbook(workbook_path)

        
        # Load workbook with openpyxl first to check System TB structure
        wb = load_workbook(workbook_path)
        ws_system = wb['System TB']
        
        # Read System TB data directly from openpyxl to get exact column positions
        # Row 2 should have headers (after skiprows=1)
        logger.info("Reading System TB headers from row 2")
        system_tb_headers = []
        for col_idx in range(1, ws_system.max_column + 1):
            header_value = ws_system.cell(row=2, column=col_idx).value
            system_tb_headers.append(header_value)
            logger.info(f"Column {get_column_letter(col_idx)} (index {col_idx}): {header_value}")
        
        # Now load with pandas for processing
        logger.info("Loading System TB sheet with pandas")
        df_sys_TB = pd.read_excel(
            workbook_path,
            sheet_name='System TB',
            skiprows=1,
            engine='openpyxl'
        )
        logger.info(f"Loaded {len(df_sys_TB)} rows from System TB")
        logger.info(f"System TB DataFrame columns: {df_sys_TB.columns.tolist()}")
        
        # Load Linked TB
        logger.info("Loading Linked TB sheet")
        df_TB_Linked = pd.read_excel(
            workbook_path,
            sheet_name='Linked TB',
            engine='openpyxl'
        )
        logger.info(f"Loaded {len(df_TB_Linked)} rows from Linked TB")
        logger.info(f"Linked TB columns: {df_TB_Linked.columns.tolist()}")
        
        # Create a set of Linked TB Account # for fast lookup
        linked_accounts = set(df_TB_Linked['Account #'].dropna())
        logger.info(f"Found {len(linked_accounts)} unique accounts in Linked TB")
        
        # Remove old Vlook if it exists
        df_sys_TB.drop(columns=['Vlook'], errors='ignore', inplace=True)
        
        # Add vlookup column → Null only if missing in Linked TB
        df_sys_TB['Vlook'] = df_sys_TB['Account #'].apply(
            lambda x: x if x in linked_accounts else "Null"
        )
        
        # Filter only the missing rows with Net Balance > 0
        missing_rows = df_sys_TB[
            (df_sys_TB['Vlook'] == "Null") & 
            (df_sys_TB['Net Balance'] > 0)
        ]
        
        # Log results
        logger.info("="*50)
        logger.info("VLOOKUP CHECK RESULTS")
        logger.info("="*50)
        logger.info(f"\nFirst 10 rows of System TB with vlookup check:")
        logger.info("\n" + df_sys_TB.head(10).to_string(index=False))
        
        logger.info("\n" + "="*50)
        logger.info("Missing Account # rows (VLOOK = Null AND Net Balance > 0):")
        logger.info("="*50)
        if not missing_rows.empty:
            logger.info(f"\nFound {len(missing_rows)} missing accounts with positive balance:")
            logger.info("\n" + missing_rows.to_string(index=False))
            logger.warning(f"WARNING: {len(missing_rows)} accounts from System TB are missing in Linked TB and have Net Balance > 0")
            
            # Now append these missing rows to Linked TB
            logger.info("\n" + "="*50)
            logger.info("APPENDING MISSING ROWS TO LINKED TB")
            logger.info("="*50)
            
            # Reload workbook to ensure we have fresh instance
            wb = load_workbook(workbook_path)
            ws_system = wb['System TB']
            ws_linked = wb['Linked TB']
            
            # Find the actual last row with Account # data in column A (Linked TB)
            last_row = None
            for row_idx in range(ws_linked.max_row, 0, -1):  # Start from bottom and go up
                cell_value = ws_linked.cell(row=row_idx, column=1).value  # Column A
                if cell_value is not None and pd.notna(cell_value):
                    # Check if it's not a header
                    if str(cell_value).strip().lower() != 'account #':
                        last_row = row_idx
                        logger.info(f"Found last Account # in column A at row: {last_row} (Value: {cell_value})")
                        break

            if last_row is None:
                logger.error("Could not find any Account # data in Linked TB column A")
                return False

            # Copy formatting from the last row to use as template
            template_row = last_row
            logger.info(f"Using row {template_row} as formatting template")

            # Insert new rows after the last Account # row (one for each missing account)
            num_rows_to_insert = len(missing_rows)
            insert_position = last_row + 1
            logger.info(f"Inserting {num_rows_to_insert} new row(s) at position {insert_position}")
            ws_linked.insert_rows(insert_position, num_rows_to_insert)

            # Start appending from the newly inserted rows
            start_append_row = insert_position
            logger.info(f"Starting to append from row: {start_append_row}")
            
            # Read data directly from System TB sheet for each missing account
            # We'll match by Account # and read from the same row
            appended_count = 0
            
            for idx, row in missing_rows.iterrows():
                current_row = start_append_row + appended_count
                account_num = row['Account #']

                # Convert account_num to int if it's a float (e.g., 2847.0 -> 2847)
                if pd.notna(account_num):
                    try:
                        account_num = int(account_num)
                    except (ValueError, TypeError):
                        pass  # Keep original value if conversion fails

                # Find this account in System TB sheet to get the exact row
                found_row = None
                last_row = ws_system.max_row + 1  # Calculate last row + 1 for new entry

                for sys_row in range(3, ws_system.max_row + 1):  # Start from row 3 (data starts here)
                    sys_account = ws_system.cell(row=sys_row, column=1).value
                    # Convert System TB account to int as well for comparison
                    if pd.notna(sys_account):
                        try:
                            sys_account = int(sys_account)
                        except (ValueError, TypeError):
                            pass  # Keep original value if conversion fails

                    if sys_account == account_num:  # Column A is Account #
                        found_row = sys_row
                        break

                # If not found, insert at the last row + 1
                if found_row is None:
                    found_row = last_row
                    ws_system.cell(row=found_row, column=1).value = account_num
                
                if found_row:
                    # Read values directly from System TB sheet
                    account_num_value = ws_system.cell(row=found_row, column=1).value  # Column A
                    net_balance_value = ws_system.cell(row=found_row, column=3).value  # Column C

                    # Get Account Name from TB Detail map instead of System TB
                    account_name_value = tb_detail_map.get(account_num, "")

                    logger.info(f"Found Account # {account_num_value} in System TB row {found_row}")
                    logger.info(f"  Account Name (from TB Detail): {account_name_value}")
                    logger.info(f"  Net Balance: {net_balance_value}")
                    
                    # Copy formatting and formulas from template row for all columns
                    import re
                    for col_idx in range(1, ws_linked.max_column + 1):
                        template_cell = ws_linked.cell(row=template_row, column=col_idx)
                        target_cell = ws_linked.cell(row=current_row, column=col_idx)

                        # Copy formatting
                        if template_cell.has_style:
                            target_cell.font = template_cell.font.copy()
                            target_cell.border = template_cell.border.copy()
                            target_cell.fill = template_cell.fill.copy()
                            target_cell.number_format = template_cell.number_format
                            target_cell.protection = template_cell.protection.copy()
                            target_cell.alignment = template_cell.alignment.copy()

                        # Check if template cell has a formula and copy it with updated row references
                        if template_cell.value and isinstance(template_cell.value, str) and template_cell.value.startswith("="):
                            formula = template_cell.value
                            # Replace row references in formula (e.g., A123 -> A{current_row})
                            def replace_row_ref(match):
                                col_letter = match.group(1)
                                return f"{col_letter}{current_row}"
                            new_formula = re.sub(r'([A-Z]+)\d+', replace_row_ref, formula)
                            target_cell.value = new_formula

                    # Now set specific values (overwrite any formulas in these columns)
                    ws_linked.cell(row=current_row, column=1).value = account_num_value  # Column A - Account #
                    ws_linked.cell(row=current_row, column=4).value = account_name_value  # Column D - Account Name
                    ws_linked.cell(row=current_row, column=5).value = net_balance_value  # Column E - Net Balance
                    ws_linked.cell(row=current_row, column=6).value = 0  # Column F - Set to 0

                    logger.info(f"Appended row {current_row}: Account # = {account_num_value}, "
                               f"Account Name = '{account_name_value}', Net Balance = {net_balance_value}, Column F = 0")

                    appended_count += 1
                else:
                    logger.warning(f"Could not find Account # {account_num} in System TB sheet")
            
            logger.info(f"Successfully appended {appended_count} rows to Linked TB")

            # Now check for similar account names and fill IFRS codes
            logger.info("\n" + "="*50)
            logger.info("CHECKING FOR SIMILAR ACCOUNT NAMES TO FILL IFRS CODES")
            logger.info("="*50)

            # Get existing account data from Linked TB (before the newly inserted rows)
            existing_accounts = {}
            for existing_row in range(2, start_append_row):  # From row 2 to before inserted rows
                acc_name = ws_linked.cell(row=existing_row, column=4).value  # Column D - Account Name
                ifrs_code = ws_linked.cell(row=existing_row, column=3).value  # Column C - IFRS Code

                if acc_name and pd.notna(acc_name):
                    # Normalize account name for comparison (lowercase, strip whitespace)
                    normalized_name = str(acc_name).strip().lower()
                    existing_accounts[normalized_name] = {
                        'row': existing_row,
                        'original_name': acc_name,
                        'ifrs_code': ifrs_code
                    }

            logger.info(f"Loaded {len(existing_accounts)} existing account names for comparison")

            # Track accounts that found matches and those that didn't
            matched_accounts = []
            unmatched_accounts = []

            # Check each newly inserted row for similar account names
            for insert_idx in range(appended_count):
                current_row = start_append_row + insert_idx
                new_acc_name = ws_linked.cell(row=current_row, column=4).value  # Column D

                if new_acc_name and pd.notna(new_acc_name):
                    normalized_new_name = str(new_acc_name).strip().lower()

                    match_found = False
                    match_info = None
                    match_type = None

                    # Check for exact match first
                    if normalized_new_name in existing_accounts:
                        match_info = existing_accounts[normalized_new_name]
                        match_found = True
                        match_type = "EXACT"
                    else:
                        # Check for partial match (substring matching)
                        # Split the new account name into words
                        new_name_words = normalized_new_name.split()

                        best_match_score = 0
                        best_match_info = None

                        for existing_name, account_info in existing_accounts.items():
                            # Try to find common substring
                            # Check if new name contains existing name or vice versa
                            if existing_name in normalized_new_name or normalized_new_name in existing_name:
                                # Calculate match score based on length of common substring
                                common_length = min(len(existing_name), len(normalized_new_name))
                                if common_length > best_match_score:
                                    best_match_score = common_length
                                    best_match_info = account_info
                            else:
                                # Check for matching word sequences (at least 2 consecutive words)
                                # Find longest common word sequence
                                for i in range(len(new_name_words)):
                                    for j in range(i + 2, len(new_name_words) + 1):  # At least 2 words
                                        new_phrase = ' '.join(new_name_words[i:j])
                                        if new_phrase in existing_name:
                                            phrase_length = len(new_phrase)
                                            if phrase_length > best_match_score:
                                                best_match_score = phrase_length
                                                best_match_info = account_info

                        # Accept partial match if we found a reasonable match (at least 10 characters or 2 words)
                        if best_match_score >= 10:  # Minimum 10 characters for partial match
                            match_info = best_match_info
                            match_found = True
                            match_type = "PARTIAL"

                    if match_found and match_info:
                        ifrs_code_to_copy = match_info['ifrs_code']

                        logger.info(f"{match_type} MATCH found for '{new_acc_name}'")
                        logger.info(f"  Matching account: '{match_info['original_name']}' (row {match_info['row']})")
                        logger.info(f"  Copying IFRS Code: {ifrs_code_to_copy}")

                        # Copy IFRS code to column C
                        ws_linked.cell(row=current_row, column=3).value = ifrs_code_to_copy

                        matched_accounts.append({
                            'Account #': ws_linked.cell(row=current_row, column=1).value,
                            'Account Name': new_acc_name,
                            'IFRS Code': ifrs_code_to_copy,
                            'Matched With': match_info['original_name'],
                            'Match Type': match_type,
                            'Net Balance': ws_linked.cell(row=current_row, column=5).value
                        })
                    else:
                        # No match found
                        logger.warning(f"NO MATCH found for '{new_acc_name}'")

                        # Collect all data for this row for the report
                        row_data = {
                            'Account #': ws_linked.cell(row=current_row, column=1).value,
                            'IFRS Code': ws_linked.cell(row=current_row, column=2).value,
                            'Line Item': ws_linked.cell(row=current_row, column=3).value,
                            'Account Name': new_acc_name,
                            'Net Balance': ws_linked.cell(row=current_row, column=5).value
                        }
                        unmatched_accounts.append(row_data)

            logger.info(f"\nMatched accounts: {len(matched_accounts)}")
            logger.info(f"Unmatched accounts: {len(unmatched_accounts)}")

            # Save the workbook with IFRS codes filled
            logger.info("\nSaving workbook with IFRS codes filled")
            wb.save(workbook_path)
            logger.info(f"Workbook saved successfully: {workbook_path}")

            # Create Excel report for unmatched accounts if any
            if unmatched_accounts:
                logger.info("\n" + "="*50)
                logger.info("CREATING UNMATCHED ACCOUNTS REPORT")
                logger.info("="*50)

                log_dir = Path(__file__).parent / "logs"
                log_dir.mkdir(exist_ok=True)

                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                report_file = log_dir / f"Unmatched_Accounts_{timestamp}.xlsx"

                # Create DataFrame with unmatched accounts
                df_unmatched = pd.DataFrame(unmatched_accounts)

                # Write to Excel
                df_unmatched.to_excel(report_file, index=False, sheet_name='Unmatched Accounts')

                logger.info(f"Unmatched accounts report created: {report_file}")
                logger.info(f"Total unmatched accounts: {len(unmatched_accounts)}")
                logger.info("\nUnmatched accounts details:")
                logger.info("\n" + df_unmatched.to_string(index=False))

                print(f"\nWARNING: {len(unmatched_accounts)} accounts have no IFRS code match")
                print(f"   Report saved to: {report_file}")
            else:
                logger.info("\nAll accounts matched successfully! No unmatched accounts report needed.")
                print(f"\nAll {len(matched_accounts)} new accounts matched and IFRS codes filled")

        else:
            logger.info("No missing accounts with positive balance found. Nothing to append.")
        
        logger.info("="*50)
        logger.info("STEP 3 COMPLETED SUCCESSFULLY")
        logger.info("="*50)
        return True
        
    except Exception as e:
        logger.error(f"Step 3 failed with exception: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def clear_column_f_values_in_linked_tb() -> bool:
    """
    STEP 4: Clear only values (not formulas) from column F in Linked TB sheet.
    
    Process:
    1. Find the SOCI workbook
    2. Open Linked TB sheet
    3. Clear only non-formula values in column F (skip header row 1)
    4. Preserve all formulas
    
    Returns:
        True if successful, False otherwise
    """
    logger.info("="*50)
    logger.info("STEP 4: Clear Column F Values in Linked TB")
    logger.info("="*50)
    
    try:
        # --- Locate project structure ---
        project_root = Path(__file__).resolve().parents[1]
        working_soci_dir = project_root / "working" / "NBD_MF_01_SOFP_SOCI"

        logger.info(f"Project root: {project_root}")
        logger.info(f"Working SOCI directory: {working_soci_dir}")

        # --- Find the single dated folder (only one exists) ---
        dated_folder = _find_single_subdirectory(working_soci_dir)
        if dated_folder is None or not dated_folder.exists():
            logger.error(f"Could not find dated folder under {working_soci_dir}")
            return False

        logger.info(f"Using dated folder: {dated_folder}")

        # --- Find the SOCI workbook inside the dated folder ---
        workbook_path = _find_soci_workbook(dated_folder)
        if workbook_path is None or not workbook_path.exists():
            logger.error(f"Could not find SOCI workbook in {dated_folder}")
            return False

        logger.info(f"Loading workbook: {workbook_path}")
        wb = load_workbook(workbook_path)
        
        sheet_name = "Linked TB"
        if sheet_name not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found in workbook")
            logger.info(f"Available sheets: {wb.sheetnames}")
            return False
        
        ws = wb[sheet_name]
        logger.info(f"Opened sheet: {sheet_name}")
        
        # Clear only non-formula values in column F (skip header row 1)
        cleared_count = 0
        formula_count = 0
        header_count = 0
        
        logger.info(f"Processing column F from row 2 to {ws.max_row} (skipping header row 1)")
        
        for row_idx in range(1, ws.max_row + 1):
            # Skip header row (row 1 only)
            if row_idx == 1:
                header_count += 1
                logger.debug(f"Row {row_idx}: Skipping header row")
                continue
            
            f_cell = ws.cell(row=row_idx, column=6)  # Column F
            
            # Check if cell has a formula
            if f_cell.value and isinstance(f_cell.value, str) and f_cell.value.startswith("="):
                formula_count += 1
                logger.debug(f"Row {row_idx}: Preserving formula in F: {f_cell.value}")
                continue
            
            # Clear the value if it's not a formula
            if f_cell.value is not None:
                f_cell.value = None
                cleared_count += 1
                logger.debug(f"Row {row_idx}: Cleared value in column F")
        
        logger.info(f"Skipped {header_count} header row, cleared {cleared_count} values, preserved {formula_count} formulas in column F")
        
        # Save the workbook
        logger.info("Saving workbook")
        wb.save(workbook_path)
        logger.info(f"Workbook saved successfully: {workbook_path}")
        
        logger.info("="*50)
        logger.info("STEP 4 COMPLETED SUCCESSFULLY")
        logger.info("="*50)
        return True
        
    except Exception as e:
        logger.error(f"Step 4 failed with exception: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def step5_paste_backup_and_update_headers(df_backup: pd.DataFrame) -> bool:
    """
    STEP 5: Paste Column E backup to Column F and update date headers in Linked TB sheet.
    
    Process:
    Part A: Paste backed-up Column E values into Column F
    1. Match Account # from backup DataFrame
    2. Paste Column E Value into Column F (starting from row 2)
    3. Skip formulas in Column F (preserve them)
    4. Only paste values (not formulas)
    
    Part B: Update date headers by adding 1 month
    5. Read date headers from columns E and F (row 1)
    6. Add 1 month to each date
    7. Update the headers with new dates
    
    Args:
        df_backup: DataFrame from backup_column_e_before_step1() with Account # and Column E Value
    
    Returns:
        True if successful, False otherwise
    """
    logger.info("="*50)
    logger.info("STEP 5: Paste Column E Backup to Column F and Update Date Headers")
    logger.info("="*50)
    
    if df_backup is None or df_backup.empty:
        logger.error("Backup DataFrame is None or empty")
        return False
    
    try:
        # --- Locate project structure ---
        project_root = Path(__file__).resolve().parents[1]
        working_soci_dir = project_root / "working" / "NBD_MF_01_SOFP_SOCI"

        logger.info(f"Project root: {project_root}")
        logger.info(f"Working SOCI directory: {working_soci_dir}")

        # --- Find the single dated folder (only one exists) ---
        dated_folder = _find_single_subdirectory(working_soci_dir)
        if dated_folder is None or not dated_folder.exists():
            logger.error(f"Could not find dated folder under {working_soci_dir}")
            return False

        logger.info(f"Using dated folder: {dated_folder}")

        # --- Find the SOCI workbook inside the dated folder ---
        workbook_path = _find_soci_workbook(dated_folder)
        if workbook_path is None or not workbook_path.exists():
            logger.error(f"Could not find SOCI workbook in {dated_folder}")
            return False

        logger.info(f"Loading workbook: {workbook_path}")
        wb = load_workbook(workbook_path)
                
        # Create a dictionary mapping Account # -> Column E Value for fast lookup
        backup_map = {}
        for _, row in df_backup.iterrows():
            account_num = row['Account #']
            column_e_value = row['Column E Value']
            
            if pd.notna(account_num):
                # Convert to int if it's a float (e.g., 2847.0 -> 2847)
                try:
                    account_num = int(account_num)
                except (ValueError, TypeError):
                    pass
                
                backup_map[account_num] = column_e_value
        
        logger.info(f"Created backup map with {len(backup_map)} account values")
        
        # Load workbook with openpyxl to write values
        wb = load_workbook(workbook_path)
        
        sheet_name = "Linked TB"
        if sheet_name not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found in workbook")
            logger.info(f"Available sheets: {wb.sheetnames}")
            return False
        
        ws = wb[sheet_name]
        logger.info(f"Opened sheet: {sheet_name}")
        
        # ========== PART A: Paste Column E backup to Column F ==========
        logger.info("\n" + "="*50)
        logger.info("PART A: Pasting Column E backup values to Column F")
        logger.info("="*50)
        
        pasted_count = 0
        formula_count = 0
        no_match_count = 0
        header_count = 0
        
        logger.info(f"Processing column F from row 2 to {ws.max_row} (skipping header row 1)")
        
        for row_idx in range(1, ws.max_row + 1):
            # Skip header row (row 1 only)
            if row_idx == 1:
                header_count += 1
                logger.debug(f"Row {row_idx}: Skipping header row")
                continue
            
            f_cell = ws.cell(row=row_idx, column=6)  # Column F
            
            # Check if cell has a formula - preserve it
            if f_cell.value and isinstance(f_cell.value, str) and f_cell.value.startswith("="):
                formula_count += 1
                logger.debug(f"Row {row_idx}: Preserving formula in F: {f_cell.value}")
                continue
            
            # Get Account # from column A
            account_num = ws.cell(row=row_idx, column=1).value  # Column A
            
            if pd.notna(account_num):
                # Convert to int if it's a float
                try:
                    account_num = int(account_num)
                except (ValueError, TypeError):
                    pass
                
                # Look up Column E value from backup
                if account_num in backup_map:
                    column_e_value = backup_map[account_num]
                    f_cell.value = column_e_value
                    pasted_count += 1
                    logger.debug(f"Row {row_idx}: Pasted value {column_e_value} for Account # {account_num}")
                else:
                    # No match found in backup
                    no_match_count += 1
                    logger.debug(f"Row {row_idx}: No match found in backup for Account # {account_num}")
            else:
                # No Account # in this row
                logger.debug(f"Row {row_idx}: No Account # found")
        
        logger.info(f"Skipped {header_count} header row, pasted {pasted_count} values, "
                   f"preserved {formula_count} formulas, {no_match_count} accounts had no match in backup")
        
        # Log first 10 pasted values for verification
        logger.info("\n" + "="*50)
        logger.info("FIRST 10 PASTED VALUES IN COLUMN F:")
        logger.info("="*50)
        for row_idx in range(2, min(12, ws.max_row + 1)):  # Rows 2-11
            account_num = ws.cell(row=row_idx, column=1).value
            f_value = ws.cell(row=row_idx, column=6).value
            logger.info(f"Row {row_idx}: Account # = {account_num}, Column F = {f_value}")
        
        # ========== PART B: Update date headers in columns E and F ==========
        logger.info("\n" + "="*50)
        logger.info("PART B: Updating date headers in columns E and F (adding 1 month)")
        logger.info("="*50)
        
        # Import required date handling modules
        from dateutil.relativedelta import relativedelta
        from datetime import datetime as dt
        
        dates_updated = 0
        
        # Update Column E header (row 1, column 5)
        e_cell = ws.cell(row=1, column=5)
        if e_cell.value:
            try:
                if isinstance(e_cell.value, datetime):
                    current_date = e_cell.value
                    new_date = current_date + relativedelta(months=1)
                    e_cell.value = new_date
                    logger.info(f"Column E header: Updated date from {current_date.strftime('%d/%m/%Y')} to {new_date.strftime('%d/%m/%Y')}")
                    dates_updated += 1
                elif isinstance(e_cell.value, str):
                    try:
                        # Try to parse as date string
                        current_date = dt.strptime(e_cell.value, "%d/%m/%Y")
                        new_date = current_date + relativedelta(months=1)
                        e_cell.value = new_date
                        logger.info(f"Column E header: Updated date from {current_date.strftime('%d/%m/%Y')} to {new_date.strftime('%d/%m/%Y')}")
                        dates_updated += 1
                    except ValueError:
                        logger.warning(f"Column E header contains non-date string: {e_cell.value}")
            except (ValueError, TypeError) as e:
                logger.warning(f"Could not update Column E header: {e}")
        else:
            logger.warning("Column E header is empty")
        
        # Update Column F header (row 1, column 6)
        f_cell = ws.cell(row=1, column=6)
        if f_cell.value:
            try:
                if isinstance(f_cell.value, datetime):
                    current_date = f_cell.value
                    new_date = current_date + relativedelta(months=1)
                    f_cell.value = new_date
                    logger.info(f"Column F header: Updated date from {current_date.strftime('%d/%m/%Y')} to {new_date.strftime('%d/%m/%Y')}")
                    dates_updated += 1
                elif isinstance(f_cell.value, str):
                    try:
                        # Try to parse as date string
                        current_date = dt.strptime(f_cell.value, "%d/%m/%Y")
                        new_date = current_date + relativedelta(months=1)
                        f_cell.value = new_date
                        logger.info(f"Column F header: Updated date from {current_date.strftime('%d/%m/%Y')} to {new_date.strftime('%d/%m/%Y')}")
                        dates_updated += 1
                    except ValueError:
                        logger.warning(f"Column F header contains non-date string: {f_cell.value}")
            except (ValueError, TypeError) as e:
                logger.warning(f"Could not update Column F header: {e}")
        else:
            logger.warning("Column F header is empty")
        
        logger.info(f"Total date headers updated: {dates_updated}")
        
        if dates_updated == 0:
            logger.warning("No dates were updated in headers")
        
        # Save the workbook
        logger.info("\nSaving workbook")
        wb.save(workbook_path)
        logger.info(f"Workbook saved successfully: {workbook_path}")
        
        logger.info("="*50)
        logger.info("STEP 5 COMPLETED SUCCESSFULLY")
        logger.info("="*50)
        return True
        
    except Exception as e:
        logger.error(f"Step 5 failed with exception: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False
    
def find_alcl_management_accounts(base_dir: Path) -> Optional[Path]:
    """
    Find the ALCL Management Accounts file by keyword "ALCL Management Accounts".
    
    Args:
        base_dir: Directory to search in (e.g., NBD_MF_01_SOFP_SOCI folder)
    
    Returns:
        Path to the ALCL Management Accounts file, or None if not found
    """
    logger.info(f"Searching for ALCL Management Accounts in: {base_dir}")
    
    keyword = "ALCL Management Accounts"
    candidate_files = []

    for root, _, files in os.walk(base_dir):
        for file in files:
            if keyword in file and file.lower().endswith(".xlsx"):
                full_path = Path(root) / file
                candidate_files.append(full_path)
                logger.info(f"Found candidate: {full_path}")

    if not candidate_files:
        logger.error(f"No file found containing keyword '{keyword}'")
        return None

    # Pick latest by modified time if multiple files found
    latest_file = max(candidate_files, key=lambda p: p.stat().st_mtime)
    logger.info(f"Latest ALCL Management Accounts selected: {latest_file}")
    return latest_file

def read_alcl_pl_data() -> Optional[pd.DataFrame]:
    """
    STEP 6: Read data from ALCL Management Accounts file.
    
    Process:
    1. Find the ALCL Management Accounts file
    2. Open "P&L (P1)" worksheet
    3. Read Column D and Column I values from rows 10 to 97 (Excel index)
    4. Store in DataFrame
    5. Log first 10 rows
    
    Returns:
        DataFrame with Column D and Column I values, or None if failed
    """
    logger.info("="*50)
    logger.info("STEP 6: Read ALCL Management Accounts P&L Data")
    logger.info("="*50)
    
    try:
        # --- Locate project structure ---
        project_root = Path(__file__).resolve().parents[1]
        working_soci_dir = project_root / "working" / "NBD_MF_01_SOFP_SOCI"

        logger.info(f"Project root: {project_root}")
        logger.info(f"Working SOCI directory: {working_soci_dir}")

        # --- Find the single dated folder inside NBD_MF_01_SOFP_SOCI ---
        date_folder = _find_single_subdirectory(working_soci_dir)
        if date_folder is None or not date_folder.exists():
            logger.error(f"Could not find dated folder under {working_soci_dir}")
            return None

        logger.info(f"Dated folder found: {date_folder}")

        # --- Use the dated folder for ALCL Management Accounts ---
        soci_folder = date_folder
        logger.info(f"Looking for ALCL Management Accounts in: {soci_folder}")

        # Find ALCL Management Accounts file
        alcl_file = find_alcl_management_accounts(soci_folder)
        if alcl_file is None:
            logger.error(f"ALCL Management Accounts file not found in {soci_folder}")
            return None


        sheet_name = "P&L (P1)"
        logger.info(f"Target sheet: {sheet_name}")
        logger.info(f"Reading from file: {alcl_file}")

        # Load the workbook with openpyxl to read exact values
        logger.info("Loading workbook with openpyxl")
        wb = load_workbook(alcl_file, data_only=True)
        
        if sheet_name not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found in workbook")
            logger.info(f"Available sheets: {wb.sheetnames}")
            return None

        ws = wb[sheet_name]
        logger.info(f"Successfully opened sheet: {sheet_name}")

        # Read data from column D and column I, rows 10 to 97 (Excel index)
        data = []
        start_row = 10
        end_row = 97
        
        logger.info(f"Reading Column D and Column I from row {start_row} to row {end_row}")
        
        for row_idx in range(start_row, end_row + 1):
            column_d_value = ws.cell(row=row_idx, column=4).value  # Column D
            column_i_value = ws.cell(row=row_idx, column=9).value  # Column I
            
            data.append({
                'Row': row_idx,
                'Column D Value': column_d_value,
                'Column I Value': column_i_value
            })

        # Create DataFrame
        df_alcl_pl = pd.DataFrame(data)
        
        logger.info(f"Successfully read {len(df_alcl_pl)} rows from Column D and Column I")
        
        # Log first 10 rows
        logger.info("\n" + "="*50)
        logger.info("FIRST 10 ROWS OF ALCL P&L DATA (Columns D & I):")
        logger.info("="*50)
        if not df_alcl_pl.empty:
            logger.info("\n" + df_alcl_pl.head(10).to_string(index=False))
        else:
            logger.warning("DataFrame is empty - no data to display")
        
        logger.info("="*50)
        logger.info("STEP 6 COMPLETED SUCCESSFULLY")
        logger.info("="*50)
        
        return df_alcl_pl
        
    except Exception as e:
        logger.error(f"Step 6 failed with exception: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None
    
def paste_alcl_data_to_ma_sheet(df_alcl_pl: pd.DataFrame) -> bool:
    """
    STEP 7: Paste ALCL P&L data to MA sheet.
    
    Process:
    1. Find the SOCI workbook (NBD-MF-01-SOFP & SOCI AFL Monthly FS)
    2. Open MA sheet
    3. Paste Column D values from df_alcl_pl to Column B starting from row 4
    4. Paste Column I values from df_alcl_pl to Column C starting from row 4
    5. Preserve formulas in Columns B and C
    6. Only paste values (not formulas)
    
    Args:
        df_alcl_pl: DataFrame from read_alcl_pl_data() with Column D Value and Column I Value
    
    Returns:
        True if successful, False otherwise
    """
    logger.info("="*50)
    logger.info("STEP 7: Paste ALCL P&L Data to MA Sheet (Columns B & C)")
    logger.info("="*50)
    
    if df_alcl_pl is None or df_alcl_pl.empty:
        logger.error("ALCL P&L DataFrame is None or empty")
        return False
    
    try:
        # --- Locate project structure ---
        project_root = Path(__file__).resolve().parents[1]
        working_soci_dir = project_root / "working" / "NBD_MF_01_SOFP_SOCI"

        logger.info(f"Project root: {project_root}")
        logger.info(f"Working SOCI directory: {working_soci_dir}")

        # --- Find the single dated folder inside NBD_MF_01_SOFP_SOCI ---
        date_folder = _find_single_subdirectory(working_soci_dir)
        if date_folder is None or not date_folder.exists():
            logger.error(f"Could not find dated folder under {working_soci_dir}")
            return False

        logger.info(f"Dated folder found: {date_folder}")

        # --- Find the SOCI workbook inside that folder ---
        workbook_path = _find_soci_workbook(date_folder)
        if workbook_path is None or not workbook_path.exists():
            logger.error(f"Could not find SOCI workbook in {date_folder}")
            return False

        logger.info(f"Loading workbook: {workbook_path}")
        wb = load_workbook(workbook_path)

        
        # Load workbook with openpyxl to write values
        wb = load_workbook(workbook_path)
        
        sheet_name = "MA"
        if sheet_name not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found in workbook")
            logger.info(f"Available sheets: {wb.sheetnames}")
            return False
        
        ws = wb[sheet_name]
        logger.info(f"Opened sheet: {sheet_name}")
        
        # Paste values starting from row 4
        start_row = 4
        pasted_count_b = 0
        pasted_count_c = 0
        formula_count_b = 0
        formula_count_c = 0
        null_count_b = 0
        null_count_c = 0
        
        logger.info(f"Pasting {len(df_alcl_pl)} values to Columns B and C starting from row {start_row}")
        
        for idx, row in df_alcl_pl.iterrows():
            target_row = start_row + idx
            b_cell = ws.cell(row=target_row, column=2)  # Column B
            c_cell = ws.cell(row=target_row, column=3)  # Column C
            
            # Handle Column B (Column D values)
            if b_cell.value and isinstance(b_cell.value, str) and b_cell.value.startswith("="):
                formula_count_b += 1
                logger.debug(f"Row {target_row}: Preserving formula in B: {b_cell.value}")
            else:
                column_d_value = row['Column D Value']
                if pd.isna(column_d_value):
                    b_cell.value = None
                    null_count_b += 1
                    logger.debug(f"Row {target_row}: Pasted None to Column B (null value)")
                else:
                    b_cell.value = column_d_value
                    pasted_count_b += 1
                    logger.debug(f"Row {target_row}: Pasted value {column_d_value} to Column B")
            
            # Handle Column C (Column I values)
            if c_cell.value and isinstance(c_cell.value, str) and c_cell.value.startswith("="):
                formula_count_c += 1
                logger.debug(f"Row {target_row}: Preserving formula in C: {c_cell.value}")
            else:
                column_i_value = row['Column I Value']
                if pd.isna(column_i_value):
                    c_cell.value = None
                    null_count_c += 1
                    logger.debug(f"Row {target_row}: Pasted None to Column C (null value)")
                else:
                    c_cell.value = column_i_value
                    pasted_count_c += 1
                    logger.debug(f"Row {target_row}: Pasted value {column_i_value} to Column C")
        
        end_row = start_row + len(df_alcl_pl) - 1
 
        # Log first 10 pasted values for verification
        logger.info("\n" + "="*50)
        logger.info("FIRST 10 PASTED VALUES IN COLUMNS B & C:")
        logger.info("="*50)
        for row_idx in range(start_row, min(start_row + 10, end_row + 1)):
            b_value = ws.cell(row=row_idx, column=2).value
            c_value = ws.cell(row=row_idx, column=3).value
            original_row = df_alcl_pl.iloc[row_idx - start_row]['Row']
            logger.info(f"Row {row_idx}: Column B = {b_value}, Column C = {c_value} (from ALCL row {original_row})")
        
        # Save the workbook
        logger.info("\nSaving workbook")
        wb.save(workbook_path)
        logger.info(f"Workbook saved successfully: {workbook_path}")
        
        logger.info("="*50)
        logger.info("STEP 7 COMPLETED SUCCESSFULLY")
        logger.info("="*50)
        return True
        
    except Exception as e:
        logger.error(f"Step 7 failed with exception: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def read_alcl_multiple_sheets_data() -> Optional[Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]]:
    """
    STEP 8A: Read data from multiple sheets in ALCL Management Accounts file.
    
    Process:
    1. Find the ALCL Management Accounts file
    2. Read from three different sheets:
       - "Audited Format" sheet: Column F, rows 10-50 → df_assets
       - "Notes 2" sheet: Column L, rows 9-12 → df_note4
       - "Sheet3" sheet: Column H, rows 7-22 → df_note10
    3. Store in separate DataFrames
    4. Log data from each sheet
    
    Returns:
        Tuple of (df_assets, df_note4, df_note10), or None if failed
    """
    logger.info("="*50)
    logger.info("STEP 8A: Read ALCL Multiple Sheets Data")
    logger.info("="*50)
    
    try:
        # --- Locate project structure ---
        project_root = Path(__file__).resolve().parents[1]
        working_soci_dir = project_root / "working" / "NBD_MF_01_SOFP_SOCI"

        logger.info(f"Project root: {project_root}")
        logger.info(f"Working SOCI directory: {working_soci_dir}")

        # --- Find the single dated folder inside NBD_MF_01_SOFP_SOCI ---
        date_folder = _find_single_subdirectory(working_soci_dir)
        if date_folder is None or not date_folder.exists():
            logger.error(f"Could not find dated folder under {working_soci_dir}")
            return None

        logger.info(f"Dated folder found: {date_folder}")

        # --- Find ALCL Management Accounts file inside the dated folder ---
        alcl_file = find_alcl_management_accounts(date_folder)
        if alcl_file is None or not alcl_file.exists():
            logger.error(f"ALCL Management Accounts file not found in {date_folder}")
            return None

        logger.info(f"Reading from ALCL Management Accounts file: {alcl_file}")

        # Load the workbook with openpyxl to read exact values
        logger.info("Loading workbook with openpyxl")
        wb = load_workbook(alcl_file, data_only=True)
        logger.info(f"Available sheets: {wb.sheetnames}")

        # ===== READ df_assets: "Audited Format" sheet, Column F, rows 10-50 =====
        logger.info("\n" + "="*50)
        logger.info("Reading df_assets from 'Audited Format' sheet")
        logger.info("="*50)
        
        sheet_name_assets = "Audited Format"
        if sheet_name_assets not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name_assets}' not found in workbook")
            return None
        
        ws_assets = wb[sheet_name_assets]
        data_assets = []
        
        start_row_assets = 10
        end_row_assets = 50
        column_assets = 6  # Column F
        
        logger.info(f"Reading Column F from row {start_row_assets} to row {end_row_assets}")
        
        for row_idx in range(start_row_assets, end_row_assets + 1):
            value = ws_assets.cell(row=row_idx, column=column_assets).value
            data_assets.append({
                'Source_Row': row_idx,
                'Value': value
            })
        
        df_assets = pd.DataFrame(data_assets)
        logger.info(f"Successfully read {len(df_assets)} rows from 'Audited Format' sheet")
        logger.info(f"First 5 rows:\n{df_assets.head().to_string(index=False)}")

        # ===== READ df_note4: "Notes 2" sheet, Column L, rows 9-12 =====
        logger.info("\n" + "="*50)
        logger.info("Reading df_note4 from 'Notes 2' sheet")
        logger.info("="*50)
        
        sheet_name_note4 = "Notes 2"
        if sheet_name_note4 not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name_note4}' not found in workbook")
            return None
        
        ws_note4 = wb[sheet_name_note4]
        data_note4 = []
        
        start_row_note4 = 9
        end_row_note4 = 12
        column_note4 = 12  # Column L
        
        logger.info(f"Reading Column L from row {start_row_note4} to row {end_row_note4}")
        
        for row_idx in range(start_row_note4, end_row_note4 + 1):
            value = ws_note4.cell(row=row_idx, column=column_note4).value
            data_note4.append({
                'Source_Row': row_idx,
                'Value': value
            })
        
        df_note4 = pd.DataFrame(data_note4)
        logger.info(f"Successfully read {len(df_note4)} rows from 'Notes 2' sheet")
        logger.info(f"All rows:\n{df_note4.to_string(index=False)}")

        # ===== READ df_note10: "Sheet3" sheet, Column H, rows 7-22 =====
        logger.info("\n" + "="*50)
        logger.info("Reading df_note10 from 'Sheet3' sheet")
        logger.info("="*50)
        
        sheet_name_note10 = "Sheet3"
        if sheet_name_note10 not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name_note10}' not found in workbook")
            return None
        
        ws_note10 = wb[sheet_name_note10]
        data_note10 = []
        
        start_row_note10 = 7
        end_row_note10 = 22
        column_note10 = 8  # Column H
        
        logger.info(f"Reading Column H from row {start_row_note10} to row {end_row_note10}")
        
        for row_idx in range(start_row_note10, end_row_note10 + 1):
            value = ws_note10.cell(row=row_idx, column=column_note10).value
            data_note10.append({
                'Source_Row': row_idx,
                'Value': value
            })
        
        df_note10 = pd.DataFrame(data_note10)
        logger.info(f"Successfully read {len(df_note10)} rows from 'Sheet3' sheet")
        logger.info(f"First 5 rows:\n{df_note10.head().to_string(index=False)}")

        # Summary
        logger.info("\n" + "="*50)
        logger.info("DATA READING SUMMARY:")
        logger.info("="*50)
        logger.info(f"df_assets: {len(df_assets)} rows (Audited Format, Column F, rows 10-50)")
        logger.info(f"df_note4: {len(df_note4)} rows (Notes 2, Column L, rows 9-12)")
        logger.info(f"df_note10: {len(df_note10)} rows (Sheet3, Column H, rows 7-22)")
        
        logger.info("="*50)
        logger.info("STEP 8A COMPLETED SUCCESSFULLY")
        logger.info("="*50)
        
        return (df_assets, df_note4, df_note10)
        
    except Exception as e:
        logger.error(f"Step 8A failed with exception: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None


def paste_alcl_multiple_sheets_to_ma(df_assets: pd.DataFrame, df_note4: pd.DataFrame, 
                                      df_note10: pd.DataFrame) -> bool:
    """
    STEP 8B: Paste data from multiple ALCL sheets to MA sheet Column G.
    
    Process:
    1. Find the SOCI workbook (NBD-MF-01-SOFP & SOCI AFL Monthly FS)
    2. Open MA sheet
    3. Paste values to Column G at specific ranges:
       - df_assets → Column G, rows 3-43
       - df_note4 → Column G, rows 48-51
       - df_note10 → Column G, rows 55-70
    4. Preserve formulas in Column G
    5. Only paste values (not formulas)
    
    Args:
        df_assets: DataFrame from "Audited Format" sheet
        df_note4: DataFrame from "Notes 2" sheet
        df_note10: DataFrame from "Sheet3" sheet
    
    Returns:
        True if successful, False otherwise
    """
    logger.info("="*50)
    logger.info("STEP 8B: Paste ALCL Multiple Sheets Data to MA Column G")
    logger.info("="*50)
    
    if df_assets is None or df_assets.empty:
        logger.error("df_assets is None or empty")
        return False
    if df_note4 is None or df_note4.empty:
        logger.error("df_note4 is None or empty")
        return False
    if df_note10 is None or df_note10.empty:
        logger.error("df_note10 is None or empty")
        return False
    
    try:
        # --- Locate project structure ---
        project_root = Path(__file__).resolve().parents[1]
        working_soci_dir = project_root / "working" / "NBD_MF_01_SOFP_SOCI"

        logger.info(f"Project root: {project_root}")
        logger.info(f"Working SOCI directory: {working_soci_dir}")

        # --- Find the single dated folder inside NBD_MF_01_SOFP_SOCI ---
        date_folder = _find_single_subdirectory(working_soci_dir)
        if date_folder is None or not date_folder.exists():
            logger.error(f"Could not find dated folder under {working_soci_dir}")
            return False

        logger.info(f"Dated folder found: {date_folder}")

        # --- Find the SOCI workbook inside the dated folder ---
        workbook_path = _find_soci_workbook(date_folder)
        if workbook_path is None or not workbook_path.exists():
            logger.error(f"Could not find SOCI workbook in {date_folder}")
            return False

        logger.info(f"Loading SOCI workbook: {workbook_path}")
        wb = load_workbook(workbook_path)

        
        # Load workbook with openpyxl to write values
        wb = load_workbook(workbook_path)
        
        sheet_name = "MA"
        if sheet_name not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found in workbook")
            logger.info(f"Available sheets: {wb.sheetnames}")
            return False
        
        ws = wb[sheet_name]
        logger.info(f"Opened sheet: {sheet_name}")
        
        column_g = 7  # Column G
        
        # ===== PASTE df_assets to Column G, rows 3-43 =====
        logger.info("\n" + "="*50)
        logger.info("Pasting df_assets to Column G (rows 3-43)")
        logger.info("="*50)
        
        start_row_assets = 3
        pasted_count_assets = 0
        null_count_assets = 0
        formula_count_assets = 0
        
        for idx, row in df_assets.iterrows():
            target_row = start_row_assets + idx
            g_cell = ws.cell(row=target_row, column=column_g)
            
            # Check if cell has a formula - preserve it
            if g_cell.value and isinstance(g_cell.value, str) and g_cell.value.startswith("="):
                formula_count_assets += 1
                logger.debug(f"Row {target_row}: Preserving formula in G: {g_cell.value}")
                continue
            
            value = row['Value']
            if pd.isna(value):
                g_cell.value = None
                null_count_assets += 1
            else:
                g_cell.value = value
                pasted_count_assets += 1
        
        logger.info(f"df_assets: Pasted {pasted_count_assets} values, {null_count_assets} nulls, preserved {formula_count_assets} formulas")

        # ===== PASTE df_note4 to Column G, rows 48-51 =====
        logger.info("\n" + "="*50)
        logger.info("Pasting df_note4 to Column G (rows 48-51)")
        logger.info("="*50)
        
        start_row_note4 = 48
        pasted_count_note4 = 0
        null_count_note4 = 0
        formula_count_note4 = 0
        
        for idx, row in df_note4.iterrows():
            target_row = start_row_note4 + idx
            g_cell = ws.cell(row=target_row, column=column_g)
            
            # Check if cell has a formula - preserve it
            if g_cell.value and isinstance(g_cell.value, str) and g_cell.value.startswith("="):
                formula_count_note4 += 1
                logger.debug(f"Row {target_row}: Preserving formula in G: {g_cell.value}")
                continue
            
            value = row['Value']
            if pd.isna(value):
                g_cell.value = None
                null_count_note4 += 1
            else:
                g_cell.value = value
                pasted_count_note4 += 1
        
        logger.info(f"df_note4: Pasted {pasted_count_note4} values, {null_count_note4} nulls, preserved {formula_count_note4} formulas")

        # ===== PASTE df_note10 to Column G, rows 55-70 =====
        logger.info("\n" + "="*50)
        logger.info("Pasting df_note10 to Column G (rows 55-70)")
        logger.info("="*50)
        
        start_row_note10 = 55
        pasted_count_note10 = 0
        null_count_note10 = 0
        formula_count_note10 = 0
        
        for idx, row in df_note10.iterrows():
            target_row = start_row_note10 + idx
            g_cell = ws.cell(row=target_row, column=column_g)
            
            # Check if cell has a formula - preserve it
            if g_cell.value and isinstance(g_cell.value, str) and g_cell.value.startswith("="):
                formula_count_note10 += 1
                logger.debug(f"Row {target_row}: Preserving formula in G: {g_cell.value}")
                continue
            
            value = row['Value']
            if pd.isna(value):
                g_cell.value = None
                null_count_note10 += 1
            else:
                g_cell.value = value
                pasted_count_note10 += 1
        
        logger.info(f"df_note10: Pasted {pasted_count_note10} values, {null_count_note10} nulls, preserved {formula_count_note10} formulas")

        # ===== VERIFICATION: Log first few values from each range =====
        logger.info("\n" + "="*50)
        logger.info("VERIFICATION - Column G Values:")
        logger.info("="*50)
        
        logger.info("\ndf_assets range (rows 3-7):")
        for row_idx in range(3, min(8, 44)):
            value = ws.cell(row=row_idx, column=column_g).value
            logger.info(f"  Row {row_idx}: {value}")
        
        logger.info("\ndf_note4 range (rows 48-51):")
        for row_idx in range(48, 52):
            value = ws.cell(row=row_idx, column=column_g).value
            logger.info(f"  Row {row_idx}: {value}")
        
        logger.info("\ndf_note10 range (rows 55-59):")
        for row_idx in range(55, min(60, 71)):
            value = ws.cell(row=row_idx, column=column_g).value
            logger.info(f"  Row {row_idx}: {value}")

        # Save the workbook
        logger.info("\nSaving workbook")
        wb.save(workbook_path)
        logger.info(f"Workbook saved successfully: {workbook_path}")
        
        # Final summary
        logger.info("\n" + "="*50)
        logger.info("PASTE SUMMARY:")
        logger.info("="*50)
        total_pasted = pasted_count_assets + pasted_count_note4 + pasted_count_note10
        total_nulls = null_count_assets + null_count_note4 + null_count_note10
        total_formulas = formula_count_assets + formula_count_note4 + formula_count_note10
        logger.info(f"Total pasted: {total_pasted} values")
        logger.info(f"Total nulls: {total_nulls}")
        logger.info(f"Total formulas preserved: {total_formulas}")
        
        logger.info("="*50)
        logger.info("STEP 8B COMPLETED SUCCESSFULLY")
        logger.info("="*50)
        return True
        
    except Exception as e:
        logger.error(f"Step 8B failed with exception: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def find_loan_schedule_file(directory: Path) -> Optional[Path]:
    """
    Find the Loan Schedule file in the given directory.
    
    Args:
        directory: Directory to search in
        
    Returns:
        Path to Loan Schedule file, or None if not found
    """
    try:
        for file in directory.iterdir():
            if file.is_file() and "Loan Schedule" in file.name and file.suffix == ".xlsx":
                logger.info(f"Found Loan Schedule file: {file.name}")
                return file
        
        logger.warning(f"Loan Schedule file not found in {directory}")
        return None
    except Exception as e:
        logger.error(f"Error searching for Loan Schedule file: {e}")
        return None


def read_loan_schedule_data() -> Optional[pd.DataFrame]:
    """
    STEP 9A: Read data from Loan Schedule file.
    
    Process:
    1. Find the Loan Schedule file
    2. Open "Loan Summary" worksheet
    3. Read specific cells from columns N, I, and J
    4. Store in DataFrame with mapping information
    
    Returns:
        DataFrame with loan schedule data and target mapping, or None if failed
    """
    logger.info("="*50)
    logger.info("STEP 9A: Read Loan Schedule Data")
    logger.info("="*50)
    
    try:
        # --- Locate project structure ---
        project_root = Path(__file__).resolve().parents[1]
        working_soci_dir = project_root / "working" / "NBD_MF_01_SOFP_SOCI"

        logger.info(f"Project root: {project_root}")
        logger.info(f"Working SOCI directory: {working_soci_dir}")

        # --- Find the single dated folder inside NBD_MF_01_SOFP_SOCI ---
        date_folder = _find_single_subdirectory(working_soci_dir)
        if date_folder is None or not date_folder.exists():
            logger.error(f"Could not find dated folder under {working_soci_dir}")
            return None

        logger.info(f"Dated folder found: {date_folder}")

        # --- Find Loan Schedule file inside the dated folder ---
        loan_file = find_loan_schedule_file(date_folder)
        if loan_file is None or not loan_file.exists():
            logger.error(f"Loan Schedule file not found in {date_folder}")
            return None

        logger.info(f"Loan Schedule file found: {loan_file}")

        sheet_name = "Loan Summary"
        logger.info(f"Target sheet: {sheet_name}")
        logger.info(f"Reading from file: {loan_file}")

        # Load the workbook with openpyxl to read exact values
        logger.info("Loading workbook with openpyxl")
        wb = load_workbook(loan_file, data_only=True)
        
        if sheet_name not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found in workbook")
            logger.info(f"Available sheets: {wb.sheetnames}")
            return None

        ws = wb[sheet_name]
        logger.info(f"Successfully opened sheet: {sheet_name}")

        # Define the mapping: (source_row, source_column, target_row, description)
        mappings = [
            # Column N mappings
            (86, 14, 3, "Column N, Row 86 → Breakups E3"),
            (53, 14, 4, "Column N, Row 53 → Breakups E4"),
            (29, 14, 5, "Column N, Row 29 → Breakups E5"),
            (59, 14, 8, "Column N, Row 59 → Breakups E8"),
            # Column I mappings
            (54, 9, 15, "Column I, Row 54 → Breakups E15"),
            (59, 9, 19, "Column I, Row 59 → Breakups E19"),
            # Column J mappings
            (54, 10, 25, "Column J, Row 54 → Breakups E25"),
            (59, 10, 29, "Column J, Row 59 → Breakups E29"),
        ]

        # Read data according to mappings
        data = []
        logger.info(f"\nReading {len(mappings)} values from Loan Summary sheet:")
        logger.info("="*50)
        
        for source_row, source_col, target_row, description in mappings:
            value = ws.cell(row=source_row, column=source_col).value
            
            # Get column letter for display
            if source_col == 14:
                col_letter = 'N'
            elif source_col == 9:
                col_letter = 'I'
            elif source_col == 10:
                col_letter = 'J'
            else:
                col_letter = '?'
            
            data.append({
                'Source_Row': source_row,
                'Source_Column': source_col,
                'Source_Column_Letter': col_letter,
                'Value': value,
                'Target_Row': target_row,
                'Description': description
            })
            
            logger.info(f"  {col_letter}{source_row} = {value} → Breakups E{target_row}")

        # Create DataFrame
        df_loan_schedule = pd.DataFrame(data)
        
        logger.info("\n" + "="*50)
        logger.info("LOAN SCHEDULE DATA SUMMARY:")
        logger.info("="*50)
        logger.info(f"Total mappings: {len(df_loan_schedule)}")
        logger.info(f"Non-null values: {df_loan_schedule['Value'].notna().sum()}")
        logger.info(f"Null values: {df_loan_schedule['Value'].isna().sum()}")
        
        # Display all data
        logger.info("\n" + "="*50)
        logger.info("ALL LOAN SCHEDULE DATA:")
        logger.info("="*50)
        logger.info("\n" + df_loan_schedule[['Source_Column_Letter', 'Source_Row', 'Value', 'Target_Row']].to_string(index=False))
        
        # Calculate sum of numeric values
        numeric_values = pd.to_numeric(df_loan_schedule['Value'], errors='coerce')
        if numeric_values.notna().any():
            logger.info(f"\nSum of numeric values: {numeric_values.sum()}")
            logger.info(f"Min value: {numeric_values.min()}")
            logger.info(f"Max value: {numeric_values.max()}")
        
        logger.info("="*50)
        logger.info("STEP 9A COMPLETED SUCCESSFULLY")
        logger.info("="*50)
        
        return df_loan_schedule
        
    except Exception as e:
        logger.error(f"Step 9A failed with exception: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None


def paste_loan_schedule_to_breakups(df_loan_schedule: pd.DataFrame) -> bool:
    """
    STEP 9B: Paste Loan Schedule data to Breakups sheet Column E.
    
    Process:
    1. Find the SOCI workbook (NBD-MF-01-SOFP & SOCI AFL Monthly FS)
    2. Open Breakups sheet
    3. Paste values to Column E at specific rows based on mapping
    4. Only paste values (not formulas)
    
    Args:
        df_loan_schedule: DataFrame from read_loan_schedule_data() with mapping info
    
    Returns:
        True if successful, False otherwise
    """
    logger.info("="*50)
    logger.info("STEP 9B: Paste Loan Schedule Data to Breakups Sheet Column E")
    logger.info("="*50)
    
    if df_loan_schedule is None or df_loan_schedule.empty:
        logger.error("Loan Schedule DataFrame is None or empty")
        return False
    
    try:
        # --- Locate project structure ---
        project_root = Path(__file__).resolve().parents[1]
        working_soci_dir = project_root / "working" / "NBD_MF_01_SOFP_SOCI"

        logger.info(f"Project root: {project_root}")
        logger.info(f"Working SOCI directory: {working_soci_dir}")

        # --- Find the single dated folder inside NBD_MF_01_SOFP_SOCI ---
        date_folder = _find_single_subdirectory(working_soci_dir)
        if date_folder is None or not date_folder.exists():
            logger.error(f"Could not find dated folder under {working_soci_dir}")
            return False

        logger.info(f"Dated folder found: {date_folder}")

        # --- Find the SOCI workbook inside the dated folder ---
        workbook_path = _find_soci_workbook(date_folder)
        if workbook_path is None or not workbook_path.exists():
            logger.error(f"Could not find SOCI workbook in {date_folder}")
            return False

        logger.info(f"Loading workbook: {workbook_path}")
        wb = load_workbook(workbook_path)

        
        # Load workbook with openpyxl to write values
        wb = load_workbook(workbook_path)
        
        sheet_name = "Breakups"
        if sheet_name not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found in workbook")
            logger.info(f"Available sheets: {wb.sheetnames}")
            return False
        
        ws = wb[sheet_name]
        logger.info(f"Opened sheet: {sheet_name}")
        
        column_e = 5  # Column E
        pasted_count = 0
        null_count = 0
        skipped_count = 0
        
        logger.info(f"\nPasting {len(df_loan_schedule)} values to Column E:")
        logger.info("="*50)
        
        for idx, row in df_loan_schedule.iterrows():
            target_row = row['Target_Row']
            value = row['Value']
            source_col_letter = row['Source_Column_Letter']
            source_row = row['Source_Row']
            
            e_cell = ws.cell(row=target_row, column=column_e)
            
            # Check if cell has a formula - skip it and log warning
            if e_cell.value and isinstance(e_cell.value, str) and e_cell.value.startswith("="):
                logger.warning(f"Row {target_row}: Cell E{target_row} has formula '{e_cell.value}' - SKIPPING")
                skipped_count += 1
                continue
            
            # Paste the value
            if pd.isna(value):
                e_cell.value = None
                null_count += 1
                logger.info(f"  E{target_row} = None (from {source_col_letter}{source_row})")
            else:
                e_cell.value = value
                pasted_count += 1
                logger.info(f"  E{target_row} = {value} (from {source_col_letter}{source_row})")
        
        logger.info("\n" + "="*50)
        logger.info("PASTE SUMMARY:")
        logger.info("="*50)
        logger.info(f"Pasted {pasted_count} values")
        logger.info(f"Null values: {null_count}")
        logger.info(f"Skipped (formulas): {skipped_count}")
        
        # Verification: Log all pasted values
        logger.info("\n" + "="*50)
        logger.info("VERIFICATION - All Pasted Values in Column E:")
        logger.info("="*50)
        for idx, row in df_loan_schedule.iterrows():
            target_row = row['Target_Row']
            current_value = ws.cell(row=target_row, column=column_e).value
            logger.info(f"  E{target_row} = {current_value}")

        # Save the workbook
        logger.info("\nSaving workbook")
        wb.save(workbook_path)
        logger.info(f"Workbook saved successfully: {workbook_path}")
        
        logger.info("="*50)
        logger.info("STEP 9B COMPLETED SUCCESSFULLY")
        logger.info("="*50)
        return True
        
    except Exception as e:
        logger.error(f"Step 9B failed with exception: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def find_supporting_schedules_file(directory: Path) -> Optional[Path]:
    """
    Find the Supporting Schedules file in the given directory.
    
    Args:
        directory: Directory to search in
        
    Returns:
        Path to Supporting Schedules file, or None if not found
    """
    try:
        for file in directory.iterdir():
            if file.is_file() and "Supporting Schedules" in file.name and file.suffix == ".xlsx":
                logger.info(f"Found Supporting Schedules file: {file.name}")
                return file
        
        logger.warning(f"Supporting Schedules file not found in {directory}")
        return None
    except Exception as e:
        logger.error(f"Error searching for Supporting Schedules file: {e}")
        return None


def find_supporting_schedules_file(directory: Path) -> Optional[Path]:
    """
    Find the Supporting Schedules file in the given directory.
    
    Args:
        directory: Directory to search in
        
    Returns:
        Path to Supporting Schedules file, or None if not found
    """
    try:
        for file in directory.iterdir():
            if file.is_file() and "Supporting Schedules" in file.name and file.suffix == ".xlsx":
                logger.info(f"Found Supporting Schedules file: {file.name}")
                return file
        
        logger.warning(f"Supporting Schedules file not found in {directory}")
        return None
    except Exception as e:
        logger.error(f"Error searching for Supporting Schedules file: {e}")
        return None


def read_supporting_schedules_data() -> Optional[pd.DataFrame]:
    """
    STEP 10A: Read data from Supporting Schedules file - New Schedule sheet.
    
    Process:
    1. Find the Supporting Schedules file
    2. Open "New Shcedule" worksheet
    3. Read rows 49-108, columns A-G
    4. Store in DataFrame
    
    Returns:
        DataFrame with supporting schedules data, or None if failed
    """
    logger.info("="*50)
    logger.info("STEP 10A: Read Supporting Schedules Data")
    logger.info("="*50)
    
    try:
        # --- Locate project structure ---
        project_root = Path(__file__).resolve().parents[1]
        working_soci_dir = project_root / "working" / "NBD_MF_01_SOFP_SOCI"

        logger.info(f"Project root: {project_root}")
        logger.info(f"Working SOCI directory: {working_soci_dir}")

        # --- Find the single dated folder inside NBD_MF_01_SOFP_SOCI ---
        date_folder = _find_single_subdirectory(working_soci_dir)
        if date_folder is None or not date_folder.exists():
            logger.error(f"Could not find dated folder under {working_soci_dir}")
            return None

        logger.info(f"Dated folder found: {date_folder}")

        # --- Look for Supporting Schedules file in the dated folder ---
        schedules_file = find_supporting_schedules_file(date_folder)
        if schedules_file is None or not schedules_file.exists():
            logger.error(f"Supporting Schedules file not found in {date_folder}")
            return None

        logger.info(f"Supporting Schedules file found: {schedules_file}")


        sheet_name = "New Shcedule"
        logger.info(f"Target sheet: {sheet_name}")
        logger.info(f"Reading from file: {schedules_file}")

        # Load the workbook with openpyxl to read exact values
        logger.info("Loading workbook with openpyxl")
        wb = load_workbook(schedules_file, data_only=True)
        
        if sheet_name not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found in workbook")
            logger.info(f"Available sheets: {wb.sheetnames}")
            return None

        ws = wb[sheet_name]
        logger.info(f"Successfully opened sheet: {sheet_name}")

        # Read data from rows 49-108, columns A-G
        data = []
        start_row = 49
        end_row = 108
        
        logger.info(f"Reading rows {start_row} to {end_row}, columns A-G")
        
        for row_idx in range(start_row, end_row + 1):
            row_data = {
                'Source_Row': row_idx,
                'Column_A': ws.cell(row=row_idx, column=1).value,  # Column A
                'Column_B': ws.cell(row=row_idx, column=2).value,  # Column B
                'Column_C': ws.cell(row=row_idx, column=3).value,  # Column C
                'Column_D': ws.cell(row=row_idx, column=4).value,  # Column D
                'Column_E': ws.cell(row=row_idx, column=5).value,  # Column E
                'Column_F': ws.cell(row=row_idx, column=6).value,  # Column F
                'Column_G': ws.cell(row=row_idx, column=7).value,  # Column G
            }
            data.append(row_data)

        # Create DataFrame
        df_schedules = pd.DataFrame(data)
        
        logger.info(f"Successfully read {len(df_schedules)} rows from New Shcedule sheet")
        
        # Log first 10 rows as sample
        logger.info("\n" + "="*50)
        logger.info("SAMPLE DATA (First 10 rows):")
        logger.info("="*50)
        logger.info("\n" + df_schedules.head(10).to_string(index=False))
        
        logger.info("="*50)
        logger.info("STEP 10A COMPLETED SUCCESSFULLY")
        logger.info("="*50)
        
        return df_schedules
        
    except Exception as e:
        logger.error(f"Step 10A failed with exception: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None


def paste_supporting_schedules_to_cbsl_provision(df_schedules: pd.DataFrame) -> bool:
    """
    STEP 10B: Paste Supporting Schedules data to CBSL Provision sheet.
    
    Process:
    1. Find the SOCI workbook (NBD-MF-01-SOFP & SOCI AFL Monthly FS)
    2. Open CBSL Provision sheet
    3. Paste values from DataFrame to specific locations based on mapping
    4. Preserve formulas, only paste values
    
    Mapping:
    - Block 1: Rows 50-56, Column B → CBSL C6-C13
    - Block 2: Rows 50-53, Columns C-G → CBSL D6-H9
    - Block 3: Rows 60-65, Columns A-B → CBSL E17-F22
    - Block 4: Rows 88-90, Columns B-D → CBSL C31-E33
    - Block 5: Rows 97-99, Columns B-D → CBSL C39-E41
    - Block 6: Rows 105-107, Columns B-D → CBSL C46-E48
    
    Args:
        df_schedules: DataFrame from read_supporting_schedules_data()
    
    Returns:
        True if successful, False otherwise
    """
    logger.info("="*50)
    logger.info("STEP 10B: Paste Supporting Schedules to CBSL Provision Sheet")
    logger.info("="*50)
    
    if df_schedules is None or df_schedules.empty:
        logger.error("Supporting Schedules DataFrame is None or empty")
        return False
    
    try:
        # --- Locate project structure ---
        project_root = Path(__file__).resolve().parents[1]
        working_soci_dir = project_root / "working" / "NBD_MF_01_SOFP_SOCI"

        logger.info(f"Project root: {project_root}")
        logger.info(f"Working SOCI directory: {working_soci_dir}")

        # --- Find the single dated folder inside NBD_MF_01_SOFP_SOCI ---
        date_folder = _find_single_subdirectory(working_soci_dir)
        if date_folder is None or not date_folder.exists():
            logger.error(f"Could not find dated folder under {working_soci_dir}")
            return False

        logger.info(f"Dated folder found: {date_folder}")

        # --- Find the SOCI workbook inside the dated folder ---
        workbook_path = _find_soci_workbook(date_folder)
        if workbook_path is None or not workbook_path.exists():
            logger.error(f"Could not find SOCI workbook in {date_folder}")
            return False

        logger.info(f"Loading workbook: {workbook_path}")
        wb = load_workbook(workbook_path)

        
        # Load workbook with openpyxl to write values
        wb = load_workbook(workbook_path)
        
        sheet_name = "CBSL Provision"
        if sheet_name not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found in workbook")
            logger.info(f"Available sheets: {wb.sheetnames}")
            return False
        
        ws = wb[sheet_name]
        logger.info(f"Opened sheet: {sheet_name}")
        
        total_pasted = 0
        total_null = 0
        total_formula = 0
        
        # Define mapping blocks: (source_start_row, source_end_row, source_col, target_start_row, target_col_idx, target_col_letter, description)
        mapping_blocks = [
            # Block 1: Column B rows 50-56 → Column C rows 6-13 (BUT row 56 → row 13, so 50-55→6-11, 56→13)
            (50, 55, 'Column_B', 6, 3, 'C', 'Block 1a: B50-55 → C6-11'),
            (56, 56, 'Column_B', 13, 3, 'C', 'Block 1b: B56 → C13'),
            
            # Block 2: Columns C-G rows 50-53 → Columns D-H rows 6-9
            (50, 53, 'Column_C', 6, 4, 'D', 'Block 2a: C50-53 → D6-9'),
            (50, 53, 'Column_D', 6, 5, 'E', 'Block 2b: D50-53 → E6-9'),
            (50, 53, 'Column_E', 6, 6, 'F', 'Block 2c: E50-53 → F6-9'),
            (50, 53, 'Column_F', 6, 7, 'G', 'Block 2d: F50-53 → G6-9'),
            (50, 53, 'Column_G', 6, 8, 'H', 'Block 2e: G50-53 → H6-9'),
            
            # Block 3: Columns A-B rows 60-65 → Columns E-F rows 17-22
            (60, 65, 'Column_A', 17, 5, 'E', 'Block 3a: A60-65 → E17-22'),
            (60, 65, 'Column_B', 17, 6, 'F', 'Block 3b: B60-65 → F17-22'),
            
            # Block 4: Columns B-D rows 88-90 → Columns C-E rows 31-33
            (88, 90, 'Column_B', 31, 3, 'C', 'Block 4a: B88-90 → C31-33'),
            (88, 90, 'Column_C', 31, 4, 'D', 'Block 4b: C88-90 → D31-33'),
            (88, 90, 'Column_D', 31, 5, 'E', 'Block 4c: D88-90 → E31-33'),
            
            # Block 5: Columns B-D rows 97-99 → Columns C-E rows 39-41
            (97, 99, 'Column_B', 39, 3, 'C', 'Block 5a: B97-99 → C39-41'),
            (97, 99, 'Column_C', 39, 4, 'D', 'Block 5b: C97-99 → D39-41'),
            (97, 99, 'Column_D', 39, 5, 'E', 'Block 5c: D97-99 → E39-41'),
            
            # Block 6: Columns B-D rows 105-107 → Columns C-E rows 46-48
            (105, 107, 'Column_B', 46, 3, 'C', 'Block 6a: B105-107 → C46-48'),
            (105, 107, 'Column_C', 46, 4, 'D', 'Block 6b: C105-107 → D46-48'),
            (105, 107, 'Column_D', 46, 5, 'E', 'Block 6c: D105-107 → E46-48'),
        ]
        
        logger.info(f"\nProcessing {len(mapping_blocks)} mapping blocks")
        logger.info("="*50)
        
        # Process each mapping block
        for source_start, source_end, source_col, target_start, target_col_idx, target_col_letter, description in mapping_blocks:
            logger.info(f"\n{description}")
            logger.info("-" * 50)
            
            block_pasted = 0
            block_null = 0
            block_formula = 0
            
            # Calculate offset: source_start in df_schedules is at index (source_start - 49)
            # because df_schedules starts from row 49
            for i, source_row in enumerate(range(source_start, source_end + 1)):
                df_index = source_row - 49  # Convert to DataFrame index
                target_row = target_start + i
                
                # Check if index is valid
                if df_index < 0 or df_index >= len(df_schedules):
                    logger.warning(f"  Source row {source_row} out of DataFrame bounds, skipping")
                    continue
                
                # Get the cell
                cell = ws.cell(row=target_row, column=target_col_idx)
                
                # Check if cell has a formula - preserve it
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    block_formula += 1
                    logger.info(f"  {target_col_letter}{target_row}: Preserving formula")
                    continue
                
                # Get value from DataFrame
                value = df_schedules.iloc[df_index][source_col]
                
                # Paste the value
                if pd.isna(value):
                    cell.value = None
                    block_null += 1
                    logger.info(f"  {target_col_letter}{target_row} = None (from row {source_row})")
                else:
                    cell.value = value
                    block_pasted += 1
                    logger.info(f"  {target_col_letter}{target_row} = {value} (from row {source_row})")
            
            logger.info(f"Block summary: {block_pasted} pasted, {block_null} nulls, {block_formula} formulas")
            total_pasted += block_pasted
            total_null += block_null
            total_formula += block_formula
        
        # Save the workbook
        logger.info("\n" + "="*50)
        logger.info("Saving workbook")
        wb.save(workbook_path)
        logger.info(f"Workbook saved successfully: {workbook_path}")
        
        # Final summary
        logger.info("\n" + "="*50)
        logger.info("OVERALL SUMMARY:")
        logger.info("="*50)
        logger.info(f"Total values pasted: {total_pasted}")
        logger.info(f"Total nulls: {total_null}")
        logger.info(f"Total formulas preserved: {total_formula}")
        logger.info(f"Total operations: {total_pasted + total_null + total_formula}")
        
        logger.info("="*50)
        logger.info("STEP 10B COMPLETED SUCCESSFULLY")
        logger.info("="*50)
        return True
        
    except Exception as e:
        logger.error(f"Step 10B failed with exception: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False
        
    except Exception as e:
        logger.error(f"Step 10B failed with exception: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def read_supporting_schedules_writeoff_data() -> Optional[pd.DataFrame]:
    """
    STEP 11A: Read data from Supporting Schedules file - New Schedule sheet for Write Off.
    
    Process:
    1. Find the Supporting Schedules file
    2. Open "New Shcedule" worksheet
    3. Read rows 4-27, columns index and E
    4. Store in DataFrame
    
    Returns:
        DataFrame with supporting schedules data for write off, or None if failed
    """
    logger.info("="*50)
    logger.info("STEP 11A: Read Supporting Schedules Data for Write Off")
    logger.info("="*50)
    
    try:
        # --- Locate project structure ---
        project_root = Path(__file__).resolve().parents[1]
        working_soci_dir = project_root / "working" / "NBD_MF_01_SOFP_SOCI"

        logger.info(f"Project root: {project_root}")
        logger.info(f"Working SOCI directory: {working_soci_dir}")

        # --- Find the single dated folder inside NBD_MF_01_SOFP_SOCI ---
        date_folder = _find_single_subdirectory(working_soci_dir)
        if date_folder is None or not date_folder.exists():
            logger.error(f"Could not find dated folder under {working_soci_dir}")
            return None

        logger.info(f"Dated folder found: {date_folder}")

        # --- Use the dated folder directly for Supporting Schedules ---
        soci_folder = date_folder
        logger.info(f"Looking for Supporting Schedules in: {soci_folder}")

        
        # Find Supporting Schedules file
        schedules_file = find_supporting_schedules_file(soci_folder)
        if schedules_file is None:
            logger.error("Supporting Schedules file not found")
            return None

        sheet_name = "New Shcedule"
        logger.info(f"Target sheet: {sheet_name}")
        logger.info(f"Reading from file: {schedules_file}")

        # Load the workbook with openpyxl to read exact values
        logger.info("Loading workbook with openpyxl")
        wb = load_workbook(schedules_file, data_only=True)
        
        if sheet_name not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found in workbook")
            logger.info(f"Available sheets: {wb.sheetnames}")
            return None

        ws = wb[sheet_name]
        logger.info(f"Successfully opened sheet: {sheet_name}")

        # Read data from rows 4-27, index column and column E
        data = []
        start_row = 4
        end_row = 27
        
        logger.info(f"Reading rows {start_row} to {end_row}, index column and column E")
        
        for row_idx in range(start_row, end_row + 1):
            row_data = {
                'Source_Row': row_idx,
                'Column_E': ws.cell(row=row_idx, column=5).value,  # Column E
            }
            data.append(row_data)

        # Create DataFrame
        df_writeoff = pd.DataFrame(data)
        
        logger.info(f"Successfully read {len(df_writeoff)} rows from New Shcedule sheet")
      
        logger.info("="*50)
        logger.info("STEP 11A COMPLETED SUCCESSFULLY")
        logger.info("="*50)
        
        return df_writeoff
        
    except Exception as e:
        logger.error(f"Step 11A failed with exception: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None


def paste_writeoff_data_to_sheet(df_writeoff: pd.DataFrame) -> bool:
    """
    STEP 11B: Paste Supporting Schedules data to Write Off sheet.
    
    Process:
    1. Find the SOCI workbook (NBD-MF-01-SOFP & SOCI AFL Monthly FS)
    2. Open Write Off sheet
    3. Paste values from DataFrame column E to specific locations
    4. Preserve formulas, only paste values
    
    Mapping:
    - DataFrame Column E rows 5-13 → Write Off Column F rows 3-11 (9 rows)
    - DataFrame Column E rows 17-20 → Write Off Column F rows 15-18 (4 rows)
    
    Args:
        df_writeoff: DataFrame from read_supporting_schedules_writeoff_data()
    
    Returns:
        True if successful, False otherwise
    """
    logger.info("="*50)
    logger.info("STEP 11B: Paste Supporting Schedules to Write Off Sheet")
    logger.info("="*50)
    
    if df_writeoff is None or df_writeoff.empty:
        logger.error("Write Off DataFrame is None or empty")
        return False
    
    try:
        # --- Locate project structure ---
        project_root = Path(__file__).resolve().parents[1]
        working_soci_dir = project_root / "working" / "NBD_MF_01_SOFP_SOCI"

        logger.info(f"Project root: {project_root}")
        logger.info(f"Working SOCI directory: {working_soci_dir}")

        # --- Find the single dated folder inside NBD_MF_01_SOFP_SOCI ---
        date_folder = _find_single_subdirectory(working_soci_dir)
        if date_folder is None or not date_folder.exists():
            logger.error(f"Could not find dated folder under {working_soci_dir}")
            return False

        logger.info(f"Dated folder found: {date_folder}")

        # --- Find the SOCI workbook inside the dated folder ---
        workbook_path = _find_soci_workbook(date_folder)
        if workbook_path is None or not workbook_path.exists():
            logger.error(f"Could not find SOCI workbook in {date_folder}")
            return False

        logger.info(f"Loading workbook: {workbook_path}")
        wb = load_workbook(workbook_path)

        
        # Load workbook with openpyxl to write values
        wb = load_workbook(workbook_path)
        
        sheet_name = "Write Off"
        if sheet_name not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found in workbook")
            logger.info(f"Available sheets: {wb.sheetnames}")
            return False
        
        ws = wb[sheet_name]
        logger.info(f"Opened sheet: {sheet_name}")
        
        target_col_idx = 6  # Column F
        target_col_letter = 'F'
        
        total_pasted = 0
        total_null = 0
        total_formula = 0
        
        # Define mapping blocks: (source_start_row, source_end_row, target_start_row, description)
        mapping_blocks = [
            (5, 13, 3, 'Block 1: E5-13 → F3-11'),
            (17, 20, 15, 'Block 2: E17-20 → F15-18'),
        ]
        
        logger.info(f"\nProcessing {len(mapping_blocks)} mapping blocks")
        logger.info("="*50)
        
        # Process each mapping block
        for source_start, source_end, target_start, description in mapping_blocks:
            logger.info(f"\n{description}")
            logger.info("-" * 50)
            
            block_pasted = 0
            block_null = 0
            block_formula = 0
            
            # Calculate offset: source_start in df_writeoff is at index (source_start - 4)
            # because df_writeoff starts from row 4
            for i, source_row in enumerate(range(source_start, source_end + 1)):
                df_index = source_row - 4  # Convert to DataFrame index (row 4 = index 0)
                target_row = target_start + i
                
                # Check if index is valid
                if df_index < 0 or df_index >= len(df_writeoff):
                    logger.warning(f"  Source row {source_row} out of DataFrame bounds, skipping")
                    continue
                
                # Get the cell
                cell = ws.cell(row=target_row, column=target_col_idx)
                
                # Check if cell has a formula - preserve it
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    block_formula += 1
                    logger.info(f"  {target_col_letter}{target_row}: Preserving formula")
                    continue
                
                # Get value from DataFrame
                value = df_writeoff.iloc[df_index]['Column_E']
                
                # Paste the value
                if pd.isna(value):
                    cell.value = None
                    block_null += 1
                    logger.info(f"  {target_col_letter}{target_row} = None (from row {source_row})")
                else:
                    cell.value = value
                    block_pasted += 1
                    logger.info(f"  {target_col_letter}{target_row} = {value} (from row {source_row})")
            
            logger.info(f"Block summary: {block_pasted} pasted, {block_null} nulls, {block_formula} formulas")
            total_pasted += block_pasted
            total_null += block_null
            total_formula += block_formula
        
        # Verification: Log pasted values
        logger.info("\n" + "="*50)
        logger.info("VERIFICATION - Pasted Values in Write Off Sheet:")
        logger.info("="*50)
        logger.info("\nBlock 1 (F3-F11):")
        for row_idx in range(3, 12):
            value = ws.cell(row=row_idx, column=target_col_idx).value
            logger.info(f"  F{row_idx} = {value}")
        
        logger.info("\nBlock 2 (F15-F18):")
        for row_idx in range(15, 19):
            value = ws.cell(row=row_idx, column=target_col_idx).value
            logger.info(f"  F{row_idx} = {value}")
        
        # Save the workbook
        logger.info("\n" + "="*50)
        logger.info("Saving workbook")
        wb.save(workbook_path)
        logger.info(f"Workbook saved successfully: {workbook_path}")
        
        # Final summary
        logger.info("\n" + "="*50)
        logger.info("OVERALL SUMMARY:")
        logger.info("="*50)
        logger.info(f"Total values pasted: {total_pasted}")
        logger.info(f"Total nulls: {total_null}")
        logger.info(f"Total formulas preserved: {total_formula}")
        logger.info(f"Total operations: {total_pasted + total_null + total_formula}")
        
        logger.info("="*50)
        logger.info("STEP 11B COMPLETED SUCCESSFULLY")
        logger.info("="*50)
        return True
        
    except Exception as e:
        logger.error(f"Step 11B failed with exception: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False

if __name__ == "__main__":
    logger.info("="*50)
    logger.info("SCRIPT EXECUTION STARTED")
    logger.info("="*50)
    logger.info(f"Current working directory: {os.getcwd()}")
    
    # STEP 0: Backup Column E data
    logger.info("")
    df_column_e_backup = backup_column_E_before_step1()
    
    if df_column_e_backup is None:
        logger.error("Step 0 failed - could not backup Column E data")
        print("Step 0 failed: Could not backup Column E data.")
    else:
        logger.info(f"Step 0 completed successfully")
        print(f"Step 0 completed successfully!")
        print(f"   Backed up {len(df_column_e_backup)} rows from Column E")
    
    # STEP 1.1: Copy E → G sofp
    logger.info("")
    result1a = read_column_e_from_sofp()
    
    if result1a is None:
        logger.error("Step 1a failed - no workbook updated")
        print("Step 1a failed: No workbook updated.")
    else:
        logger.info(f"Step 1a completed successfully")
        print(f"Step 1a completed successfully!")
        print(f"   Updated workbook: {result1a}")
    
    # STEP 1.1: Copy E → G soci
    logger.info("")
    result1b = read_column_e_from_soci()

    if result1b is None:
        logger.error("Step 1b failed - no workbook updated")
        print("Step 1b failed: No workbook updated.")
    else:
        logger.info(f"Step 1b completed successfully")
        print(f"Step 1b completed successfully!")
        print(f"   Updated workbook: {result1b}")

    # STEP 1.2: paste data to sofp
    logger.info("")
    result1c = paste_to_sofp(result1a)

    if result1c is None:
        logger.error("Step 1c failed - no workbook updated")
        print("Step 1c failed: No workbook updated.")
    else:
        logger.info(f"Step 1c completed successfully")
        print(f"Step 1c completed successfully!")
        print(f"   Updated workbook: {result1c}")

    # STEP 1.2: paste data to soci
    logger.info("")
    result1d = paste_to_soci(result1b)

    if result1d is None:
        logger.error("Step 1d failed - no workbook updated")
        print("Step 1d failed: No workbook updated.")
    else:
        logger.info(f"Step 1d completed successfully")
        print(f"Step 1d completed successfully!")
        print(f"   Updated workbook: {result1d}")

    # STEP 2: TB Detail Report → System TB
    logger.info("")
    result2 = tb_detail_to_system_tb()
    
    if not result2:
        logger.error("Step 2 failed")
        print("Step 2 failed: Could not update System TB.")
    else:
        logger.info(f"Step 2 completed successfully")
        print(f"Step 2 completed successfully!")
        print(f"   Updated System TB sheet")

    # STEP 3: Check Missing Accounts in Linked TB
    logger.info("")
    result3 = check_missing_accounts_in_linked_tb()

    if not result3:
        logger.error("Step 3 failed")
        print("Step 3 failed: Could not check missing accounts.")
    else:
        logger.info(f"Step 3 completed successfully")
        print(f"Step 3 completed successfully!")
        print(f"   Missing accounts check completed - see log for details")

    # STEP 4: Clear Column F Values in Linked TB
    logger.info("")
    result4 = clear_column_f_values_in_linked_tb()

    if not result4:
        logger.error("Step 4 failed")
        print("Step 4 failed: Could not clear column F values.")
    else:
        logger.info(f"Step 4 completed successfully")
        print(f"Step 4 completed successfully!")
        print(f"   Cleared column F values in Linked TB (formulas preserved)")

    # STEP 5: Paste Column E Backup to Column F and Update Headers (combined)
    logger.info("")
    result5 = step5_paste_backup_and_update_headers(df_column_e_backup)

    if not result5:
        logger.error("Step 5 failed")
        print("Step 5 failed: Could not paste Column E backup and update headers.")
    else:
        logger.info(f"Step 5 completed successfully")
        print(f"Step 5 completed successfully!")
        print(f"   Part A: Pasted Column E backup values to Column F")
        print(f"   Part B: Updated date headers in columns E and F (added 1 month)")

    # STEP 6: Read ALCL Management Accounts P&L Data (Columns D & I)
    logger.info("")
    df_alcl_pl = read_alcl_pl_data()

    if df_alcl_pl is None:
        logger.error("Step 6 failed")
        print("Step 6 failed: Could not read ALCL Management Accounts P&L data.")
    else:
        logger.info(f"Step 6 completed successfully")
        print(f"Step 6 completed successfully!")
        print(f"   Read {len(df_alcl_pl)} rows from P&L (P1) sheet")

    # STEP 7: Paste ALCL Data to MA Sheet (Column B)
    logger.info("")
    result7 = paste_alcl_data_to_ma_sheet(df_alcl_pl)

    if not result7:
        logger.error("Step 7 failed")
        print("Step 7 failed: Could not paste ALCL data to MA sheet.")
    else:
        logger.info(f"Step 7 completed successfully")
        print(f"Step 7 completed successfully!")
        print(f"   Pasted Column D values to MA Column B (starting row 4)")

    # STEP 8A: Read ALCL Multiple Sheets Data (Assets, Note4, Note10)
    logger.info("")
    result8a = read_alcl_multiple_sheets_data()

    if result8a is None:
        logger.error("Step 8A failed")
        print("Step 8A failed: Could not read ALCL multiple sheets data.")
        df_assets = None
        df_note4 = None
        df_note10 = None
    else:
        df_assets, df_note4, df_note10 = result8a
        logger.info(f"Step 8A completed successfully")
        print(f"Step 8A completed successfully!")
        print(f"   df_assets: {len(df_assets)} rows from 'Audited Format' sheet")
        print(f"   df_note4: {len(df_note4)} rows from 'Notes 2' sheet")
        print(f"   df_note10: {len(df_note10)} rows from 'Sheet3' sheet")

    # STEP 8B: Paste ALCL Multiple Sheets Data to MA Column G
    logger.info("")
    result8b = False
    if result8a is not None:
        result8b = paste_alcl_multiple_sheets_to_ma(df_assets, df_note4, df_note10)

    if not result8b:
        logger.error("Step 8B failed")
        print("Step 8B failed: Could not paste ALCL multiple sheets data to MA sheet.")
    else:
        logger.info(f"Step 8B completed successfully")
        print(f"Step 8B completed successfully!")
        print(f"   Pasted df_assets to MA Column G (rows 3-43)")
        print(f"   Pasted df_note4 to MA Column G (rows 48-51)")
        print(f"   Pasted df_note10 to MA Column G (rows 55-70)")

    # STEP 9A: Read Loan Schedule Data
    logger.info("")
    df_loan_schedule = read_loan_schedule_data()

    if df_loan_schedule is None:
        logger.error("Step 9A failed")
        print("Step 9A failed: Could not read Loan Schedule data.")
    else:
        logger.info(f"Step 9A completed successfully")
        print(f"Step 9A completed successfully!")
        print(f"   Read {len(df_loan_schedule)} values from Loan Summary sheet")
        print(f"   Sources: Column N (4 values), Column I (2 values), Column J (2 values)")

    # STEP 9B: Paste Loan Schedule Data to Breakups Sheet
    logger.info("")
    result9b = False
    if df_loan_schedule is not None:
        result9b = paste_loan_schedule_to_breakups(df_loan_schedule)

    if not result9b:
        logger.error("Step 9B failed")
        print("Step 9B failed: Could not paste Loan Schedule data to Breakups sheet.")
    else:
        logger.info(f"Step 9B completed successfully")
        print(f"Step 9B completed successfully!")
        print(f"   Pasted {len(df_loan_schedule)} values to Breakups Column E")
        print(f"   Target rows: E3, E4, E5, E8, E15, E19, E25, E29")

    # STEP 10A: Read Supporting Schedules Data
    logger.info("")
    df_schedules = read_supporting_schedules_data()

    if df_schedules is None:
        logger.error("Step 10A failed")
        print("Step 10A failed: Could not read Supporting Schedules data.")
    else:
        logger.info(f"Step 10A completed successfully")
        print(f"Step 10A completed successfully!")
        print(f"   Read {len(df_schedules)} rows from New Shcedule sheet (rows 45-54)")
        print(f"   Columns: A, B, C, D, E, F, G")

    # STEP 10B: Paste Supporting Schedules Data to CBSL Provision Sheet
    logger.info("")
    result10b = False
    if df_schedules is not None:
        result10b = paste_supporting_schedules_to_cbsl_provision(df_schedules)

    if not result10b:
        logger.error("Step 10B failed")
        print("Step 10B failed: Could not paste Supporting Schedules data to CBSL Provision sheet.")
    else:
        logger.info(f"Step 10B completed successfully")
        print(f"Step 10B completed successfully!")
        print(f"   Pasted {len(df_schedules)} rows to CBSL Provision sheet")
        print(f"   Mapped: B→C, C→D, D→E, E→F, F→G, G→H (starting row 6)")

    # STEP 11A: Read Supporting Schedules Data for Write Off
    logger.info("")
    df_writeoff = read_supporting_schedules_writeoff_data()

    if df_writeoff is None:
        logger.error("Step 11A failed")
        print("Step 11A failed: Could not read Supporting Schedules data for Write Off.")
    else:
        logger.info(f"Step 11A completed successfully")
        print(f"Step 11A completed successfully!")
        print(f"   Read {len(df_writeoff)} rows from New Shcedule sheet (rows 4-27)")
        print(f"   Column: E")

    # STEP 11B: Paste Write Off Data to Write Off Sheet
    logger.info("")
    result11b = False
    if df_writeoff is not None:
        result11b = paste_writeoff_data_to_sheet(df_writeoff)

    if not result11b:
        logger.error("Step 11B failed")
        print("Step 11B failed: Could not paste Write Off data to Write Off sheet.")
    else:
        logger.info(f"Step 11B completed successfully")
        print(f"Step 11B completed successfully!")
        print(f"   Pasted values to Write Off Column F")
        print(f"   Block 1: E5-13 → F3-11, Block 2: E17-20 → F15-18")

# Save final file only if all results exist
    import argparse

    parser = argparse.ArgumentParser(description='Save final AFL Monthly FS file with dynamic date')
    parser.add_argument('--date', type=str, required=True, help='Report date in MM/DD/YYYY format')
    args = parser.parse_args()

    # Parse the input date
def _update_breakups_from_master(source_file_path: Path, report_date: datetime):
    """
    Update Breakups sheet rows 4, 5, and 8 at the column corresponding to the GUI month/year
    using latest values from Master_Data.xlsx → NBD-MF-01-SOFP-SOCI (B,C,D columns).
    """
    try:
        project_root = Path(__file__).resolve().parents[1]
        master_path = project_root / "Master_Data.xlsx"
        if not master_path.exists():
            print(f"Master_Data.xlsx not found at {master_path}; skipping Breakups update")
            return
        md_wb = load_workbook(master_path, data_only=True)
        if "NBD-MF-01-SOFP-SOCI" not in md_wb.sheetnames:
            print("Sheet 'NBD-MF-01-SOFP-SOCI' not found in Master_Data.xlsx; skipping Breakups update")
            md_wb.close()
            return
        md_ws = md_wb["NBD-MF-01-SOFP-SOCI"]
        last_row = md_ws.max_row
        bank_loans = md_ws[f"B{last_row}"].value
        secur_loans = md_ws[f"C{last_row}"].value
        foreign_funding = md_ws[f"D{last_row}"].value
        md_wb.close()

        def to_number(val):
            try:
                if val is None:
                    return 0
                if isinstance(val, (int, float)):
                    return float(val)
                s = str(val).replace(",", "").strip()
                return float(s) if s else 0.0
            except Exception:
                return 0.0

        bank_loans = to_number(bank_loans)
        secur_loans = to_number(secur_loans)
        foreign_funding = to_number(foreign_funding)

        wb = load_workbook(source_file_path, data_only=False)
        if "Breakups" not in wb.sheetnames:
            print("Breakups sheet not found in SOCI file; skipping Breakups update")
            wb.close()
            return
        ws = wb["Breakups"]

        # Find target column in row 2 whose month/year matches GUI date
        target_col = None
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=2, column=col).value
            dt = None
            if isinstance(val, datetime):
                dt = val
            elif isinstance(val, (int, float)):
                # Excel date serials can appear; openpyxl often converts to datetime, but safe guard
                try:
                    from openpyxl.utils.datetime import from_excel
                    dt = from_excel(val)
                except Exception:
                    dt = None
            else:
                # Try parse string like '7/1/2025'
                try:
                    dt = datetime.strptime(str(val), "%m/%d/%Y")
                except Exception:
                    try:
                        dt = datetime.strptime(str(val), "%d/%m/%Y")
                    except Exception:
                        dt = None
            if dt and dt.month == report_date.month and dt.year == report_date.year:
                target_col = col
                break

        if target_col is None:
            print("Could not find matching month/year column in Breakups header row 2; skipping update")
            wb.close()
            return

        ws.cell(row=4, column=target_col).value = bank_loans
        ws.cell(row=5, column=target_col).value = secur_loans
        ws.cell(row=8, column=target_col).value = foreign_funding

        try:
            wb.save(source_file_path)
            print(f"Breakups updated and saved in-place: {source_file_path}")
        finally:
            wb.close()
    except Exception as e:
        print(f"Failed to update Breakups from Master_Data.xlsx: {e}")

try:
    report_date = datetime.strptime(args.date, "%m/%d/%Y")
    month_name = report_date.strftime("%B")  # e.g. "October"
    year = report_date.year
except Exception as e:
    print(f"Invalid date format: {e}")
    exit(1)

if (result1a is not None and result1b is not None and result1c is not None and result1d is not None and 
    result2 and result3 and result4 and result5 and 
    df_alcl_pl is not None and result7 and result8a is not None and result8b and
    df_loan_schedule is not None and result9b and df_schedules is not None and result10b and
    df_writeoff is not None and result11b):

    logger.info("\n" + "="*50)
    logger.info("SAVING FINAL FILE")
    logger.info("="*50)

    try:
        # Create dynamic filename
        final_filename = f"NBD-MF-01-SOFP & SOCI AFL Monthly FS {month_name} {year}.xlsx"

        # Get source file path (handle dict case if needed)
        if isinstance(result1a, dict):
            source_path = result1a.get('path') or result1a.get('file_path') or list(result1a.values())[0]
        else:
            source_path = result1a

        source_path = Path(source_path)

        # Update Breakups from Master_Data.xlsx using GUI date before rename
        try:
            _update_breakups_from_master(source_path, report_date)
        except Exception as _e:
            print(f"Warning: Breakups update skipped: {_e}")
        final_path = source_path.parent / final_filename  # same folder as source file

        logger.info(f"Saving file as: {final_filename}")
        logger.info(f"Full path: {final_path}")

        # Rename in same folder
        source_path.rename(final_path)

        logger.info(f"File renamed successfully: {final_path}")
        print(f"\nFinal file saved as: {final_filename}")
        print(f"   Location: {final_path}")

        # Note: Per requirements, do not remove or copy other files here. Saving and renaming only.

    except Exception as e:
        logger.error(f"Failed to save final file: {e}")
        print(f"\nWARNING: Could not save final file with custom name: {e}")

    # Final summary (MOVED OUTSIDE - prints always)
    logger.info("")
    logger.info("="*50)
    logger.info("SCRIPT EXECUTION COMPLETED")
    logger.info("="*50)
    total_steps = 19
    successful_steps = sum([
        df_column_e_backup is not None,
        result1a is not None,
        result1b is not None,
        result1c is not None,
        result1d is not None,
        result2,
        result3,
        result4,
        result5,
        df_alcl_pl is not None,
        result7,
        result8a is not None,
        result8b,
        df_loan_schedule is not None,
        result9b,
        df_schedules is not None,
        result10b,
        df_writeoff is not None,
        result11b
    ])
    logger.info(f"Successfully completed {successful_steps}/{total_steps} steps")
    print(f"\n{'='*50}")
    print(f"EXECUTION SUMMARY: {successful_steps}/{total_steps} steps completed successfully")
    print(f"{'='*50}")

    # Detailed step results (MOVED OUTSIDE - prints always)
    print("\nDetailed Results:")
    print(f"  Step 0 (Backup Column E): {'PASSED' if df_column_e_backup is not None else 'FAILED'}")
    print(f"  Step 1a (Copy E sofp): {'PASSED' if result1a is not None else 'FAILED'}")
    print(f"  Step 1b (Copy E soci): {'PASSED' if result1b is not None else 'FAILED'}")
    print(f"  Step 1c (Paste to sofp): {'PASSED' if result1c is not None else 'FAILED'}")
    print(f"  Step 1d (Paste to soci): {'PASSED' if result1d is not None else 'FAILED'}")
    print(f"  Step 2 (TB Detail → System TB): {'PASSED' if result2 else 'FAILED'}")
    print(f"  Step 3 (Check Missing Accounts): {'PASSED' if result3 else 'FAILED'}")
    print(f"  Step 4 (Clear Column F): {'PASSED' if result4 else 'FAILED'}")
    print(f"  Step 5 (Paste Backup & Update Headers): {'PASSED' if result5 else 'FAILED'}")
    print(f"  Step 6 (Read ALCL P&L Data): {'PASSED' if df_alcl_pl is not None else 'FAILED'}")
    print(f"  Step 7 (Paste to MA Column B): {'PASSED' if result7 else 'FAILED'}")
    print(f"  Step 8A (Read ALCL Multiple Sheets): {'PASSED' if result8a is not None else 'FAILED'}")
    print(f"  Step 8B (Paste to MA Column G): {'PASSED' if result8b else 'FAILED'}")
    print(f"  Step 9A (Read Loan Schedule): {'PASSED' if df_loan_schedule is not None else 'FAILED'}")
    print(f"  Step 9B (Paste to Breakups Column E): {'PASSED' if result9b else 'FAILED'}")
    print(f"  Step 10A (Read Supporting Schedules): {'PASSED' if df_schedules is not None else 'FAILED'}")
    print(f"  Step 10B (Paste to CBSL Provision): {'PASSED' if result10b else 'FAILED'}")
    print(f"  Step 11A (Read Write Off Data): {'PASSED' if df_writeoff is not None else 'FAILED'}")
    print(f"  Step 11B (Paste to Write Off Sheet): {'PASSED' if result11b else 'FAILED'}")
    print(f"{'='*50}")